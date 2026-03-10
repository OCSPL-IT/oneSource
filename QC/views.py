import logging, time
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required,permission_required
from django.views.decorators.http import require_POST
import openpyxl
from pprint import pprint
import pandas as pd
from .models import *
from .forms import *
from django.http import HttpResponse, Http404, JsonResponse, HttpResponseBadRequest
from django.db import connections, transaction
from django.db.models import Prefetch, Count, DateField, Q,Value
from django.views.decorators.http import require_GET
from django.utils.timezone import localtime
from django.utils import timezone
from datetime import date, timedelta
from django.core.mail import send_mail
from django.conf import settings
from django.db.models.functions import TruncDay, TruncWeek, TruncMonth,Trim
from django.db.models import Func
import itertools, csv
from datetime import datetime
from django.core.management import call_command
from io import StringIO
# ReportLab Platypus imports
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.platypus import Paragraph, Spacer, Table, TableStyle, Flowable, SimpleDocTemplate
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from reportlab.lib.styles import getSampleStyleSheet
from django.core.mail import EmailMessage
from io import BytesIO
import urllib.parse
import csv
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
import io
from decimal import Decimal
from collections import defaultdict
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from django.db.models.query import QuerySet
from django.db.models import Model
import xlsxwriter
import io, uuid, threading, traceback
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.cache import cache
import datetime
from django.urls import reverse
import logging
from django.utils.dateparse import parse_date
from datetime import date


# Initialize custom logger
logger = logging.getLogger('custom_logger')


# Create your views here.


# --- AR No. helper (IP/FG series) -------------------------------------------
def _next_ar_no(ar_type: str) -> str:
    """
    Return next AR number for the given series (e.g., 'IP', 'FG', or 'SFG'):
    QC/<AR>/<YY-YY>/<#####>
    """
    today = timezone.now()
    fy_start = today.year if today.month >= 4 else today.year - 1
    fy_str = f"{str(fy_start)[-2:]}-{str(fy_start + 1)[-2:]}"
    prefix = f"QC/{ar_type}/{fy_str}/"

    last = ( QCEntry.objects
             .filter(ar_no__startswith=prefix)
             .order_by('-ar_no')
             .first() )
    last_seq = int(last.ar_no.split('/')[-1]) if last and last.ar_no else 0
    return f"{prefix}{last_seq + 1:05d}"



def _get_stages_from_bom():
    """
    Returns a list of dicts, one per unique item_name in LocalBOMDetail:
      [
        {
          "item_name":  "<StageName>",
          "fg_name":    "<FGName>",
          "bom_code":   "<BOMCode>",
          "itm_type":   "<ItemTypeFromERP>",
        },
        ...
      ]
    """
    # 1) Grab unique stage names (just the name column)
    names = (
        LocalBOMDetail.objects
        .values_list('item_name', flat=True)
        .distinct()
        .order_by('item_name')
    )

    stages = []
    for name in names:
        # 2) pick the first matching row for this name
        detail = LocalBOMDetail.objects.filter(item_name=name).first()
        if not detail:
            continue
        stages.append({
            "item_name": detail.item_name,
            "fg_name":   detail.fg_name,
            "bom_code":  detail.bom_code,
            "itm_type":  detail.itm_type,
        })

    return stages


@login_required
def api_get_batch_nos(request):
    product_id = request.GET.get('product_id')
    stage = request.GET.get('stage')
    batch_nos = []

    # Find batches for this product & stage
    if product_id and stage:
        # Find product name by product_id
        product = Product.objects.filter(id=product_id).first()
        if product:
            # BmrIssue must match product and stage (if your ERP/BMR mapping includes stage)
            qs = BmrIssue.objects.filter(
                fg_name=product.name,
                # Optionally filter on stage if BmrIssue has a stage field
                # stage=stage
            ).values_list('op_batch_no', flat=True).distinct().order_by('op_batch_no')
            batch_nos = list(qs)

    return JsonResponse({"batch_nos": batch_nos})


@login_required
def product_list(request):
    """ View QC details with server-side search and pagination """
    if not request.user.has_perm('QC.view_product'):
        logger.warning(f"Unauthorized View attempt by {request.user.username}")
        messages.error(request, "You do not have permission to View Product records.")
        return redirect('indexpage')

    # Get the search query from the URL's GET parameters
    search_query = request.GET.get('q', '')

    # Start with the base queryset
    products_qs = Product.objects.prefetch_related('specs').all()

    # If a search query is provided, filter the queryset
    if search_query:
        products_qs = products_qs.filter(
            Q(name__icontains=search_query) |
            Q(stages__icontains=search_query) |
            Q(code__icontains=search_query) |
            Q(item_type__icontains=search_query) |
            Q(specs__name__icontains=search_query)
        ).distinct()

    # Order the final queryset
    products_qs = products_qs.order_by('id')
    
    # This loop should be performed on the final paginated list for efficiency
    # But since it's already here, we will apply it to the filtered list.
    for p in products_qs:
        p.visible_specs = [s for s in p.specs.all() if s.name and s.name.strip()]

    # --- Pagination Logic (applied to the filtered list) ---
    paginator = Paginator(products_qs, 10)  # Show 5 products per page
    page = request.GET.get('page')

    try:
        paged_products = paginator.page(page)
    except PageNotAnInteger:
        paged_products = paginator.page(1)
    except EmptyPage:
        paged_products = paginator.page(paginator.num_pages)

    return render(request, 'products/product_list.html', {
        'products': paged_products,
        'paginator': paginator,
        'page_obj': paged_products,
        'search_query': search_query,  # Pass the query back to the template
    })


@login_required
def product_export_xlsx(request):
    
    if not request.user.has_perm('QC.view_product'):
        logger.warning("[QC][Products][Export] Unauthorized attempt by user=%s", request.user.username)
        messages.error(request, "You do not have permission to export Product records.")
        return redirect('indexpage')

    q = (request.GET.get('q') or '').strip()

    # Use the SAME filtering you already have in product_list:
    qs = Product.objects.all()
    if q:
        qs = qs.filter(
            Q(name__icontains=q) |
            Q(stages__icontains=q) |
            Q(code__icontains=q) |
            Q(item_type__icontains=q) |
            Q(specs__name__icontains=q)
        ).distinct()

    # Build workbook in memory
    output = BytesIO()
    wb = xlsxwriter.Workbook(output, {'in_memory': True})
    ws = wb.add_worksheet("Products")

    # Formats
    title_fmt = wb.add_format({'bold': True, 'font_size': 14})
    sub_fmt   = wb.add_format({'italic': True, 'font_color': '#666666'})
    head_fmt  = wb.add_format({'bold': True, 'border': 1, 'bg_color': '#EAF2FF'})
    cell_fmt  = wb.add_format({'border': 1})
    
    # Title + date
    ws.write(0, 0, "QC Products Master", title_fmt)

    # Headers (order fixed)
    headers = ["Name", "Stage", "Code", "Item Type"]
    start_row = 3
    for c, h in enumerate(headers):
        ws.write(start_row, c, h, head_fmt)

    # Data rows
    r = start_row + 1
    for p in qs.only('name', 'stages', 'code', 'item_type'):
        ws.write(r, 0, p.name or "", cell_fmt)
        ws.write(r, 1, p.stages or "", cell_fmt)
        ws.write(r, 2, p.code or "", cell_fmt)
        ws.write(r, 3, p.item_type or "", cell_fmt)
        r += 1

    # Column widths
    ws.set_column(0, 0, 42)  # Name
    ws.set_column(1, 1, 26)  # Stage
    ws.set_column(2, 2, 18)  # Code
    ws.set_column(3, 3, 22)  # Item Type

    wb.close()
    output.seek(0)

    filename = f"Products_Master.xlsx"
    resp = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp




@login_required
def product_detail(request, pk):
    """Read-only details for a single Product. Requires `view_product`."""
    app_label = Product._meta.app_label  # avoids hardcoding "QC"
    if not request.user.has_perm(f"{app_label}.view_product"):
        logger.warning(f"Unauthorized View attempt by {request.user.username}")
        messages.error(request, "You do not have permission to View Product records.")
        return redirect("indexpage")

    # Pull the product with its specs in one go
    product = get_object_or_404(
        Product.objects.prefetch_related("specs"),
        pk=pk
    )

    # Only show specs that have a non-empty name
    visible_specs = [s for s in product.specs.all() if s.name and s.name.strip()]

    return render(request, "products/product_detail.html", {
        "product": product,
        "visible_specs": visible_specs,
    })


# ---- Add once in this module ----------------------------------------------
_FINISHED_LABELS = {"finished goods", "finished good", "finished", "fg"}
 
def _ensure_fg_name(stages):
    """
    Ensure every stage dict/object has an fg_name.
    For Finished-Goods rows that don't have fg_name, use item_name.
    Works for dicts and lightweight objects.
    """
    out = []
    for s in stages:
        if isinstance(s, dict):
            d = s.copy()
            if not (d.get("fg_name") or "").strip():
                d["fg_name"] = (d.get("item_name") or "").strip()
            out.append(d)
        else:
            # object with attrs
            if not getattr(s, "fg_name", None):
                try:
                    setattr(s, "fg_name", (getattr(s, "item_name", "") or "").strip())
                except Exception:
                    # if it's an immutable namedtuple-like; just keep as-is
                    pass
            out.append(s)
    return out
# ---------------------------------------------------------------------------


@login_required
def product_create(request):
    """ Add QC details (Permission Required: QC.add_product) """
    if not request.user.has_perm('QC.add_product'):
        logger.warning(f"Unauthorized Add attempt by {request.user.username}")
        messages.error(request, "You do not have permission to Add Product records.")
        return redirect('indexpage')
   
    appearance_options = list(
        AppearanceOption.objects.values_list("name", flat=True).order_by("name")
    )
    stages = _get_stages_from_bom()          # include stages + FG (no filtering)
    stages = _ensure_fg_name(stages)         # make sure fg_name is always present
 
    if request.method == "POST":
        form = ProductForm(request.POST)
        temp_product = Product()
        formset = SpecFormSetCreate(request.POST, instance=temp_product)
 
        if form.is_valid() and formset.is_valid():
            product = form.save()
            formset.instance = product
            formset.save()
            messages.success(request, f"Product '{product.name}' created successfully.")
            logger.info(f"Product '{product.name}' created by {request.user.username}.")
            return redirect("qc:product_list")
        else:
            messages.error(
                request,
                "There was an error creating the product. Please correct the errors below."
            )
            logger.warning(f"Product creation failed by {request.user.username}. Errors: {form.errors}, Formset errors: {formset.errors}")
    else:
        form = ProductForm()
        formset = SpecFormSetCreate(instance=Product())
 
    return render(request, "products/product_form.html", {
        "form":               form,
        "formset":            formset,
        "action":             "Create",
        "appearance_options": appearance_options,
        "stages":             stages,
    })




def _get_finished_good_data():
    """
    Queries LocalBOMDetail and BmrIssue for 'Finished Good' items,
    combines them, and returns a unique, sorted list of dictionaries
    for the front-end search.
    """
    # 1. Get products from LocalBOMDetail
    bom_products = (
        LocalBOMDetail.objects
        .filter(itm_type='Finished Good')
        .values('fg_name', 'item_name', 'bom_code', 'itm_type')
    )

    # 2. Get products from BmrIssue
    bmr_products = (
        BmrIssue.objects
        .filter(item_type='Finished Good')
        .annotate(bom_code=Value(''))  # Add an empty bom_code field for consistency
        .values('fg_name', 'item_name', 'bom_code', 'item_type')
    )
    
    # Use a dictionary to store unique products by fg_name, giving priority to BOM data
    product_map = {}

    for product in bom_products:
        product_map[product['fg_name']] = {
            'fg_name': product['fg_name'],
            'stage': product['item_name'],
            'bom_code': product['bom_code'],
            'item_type': product['itm_type']
        }

    for product in bmr_products:
        if product['fg_name'] not in product_map: # Only add if not already present from BOM
             product_map[product['fg_name']] = {
                'fg_name': product['fg_name'],
                'stage': product['item_name'],
                'bom_code': product['bom_code'], # will be ''
                'item_type': product['item_type']
            }

    # Convert the dictionary back to a list and sort it by name
    combined_list = sorted(product_map.values(), key=lambda p: p['fg_name'])
    return combined_list


@login_required
def product_create_fg(request):
    """
    Renders and processes the form for creating a 'Finished Good' product.
    """
    if not request.user.has_perm('QC.add_product'):
        messages.error(request, "You do not have permission to Add Product records.")
        return redirect('indexpage')

    if request.method == "POST":
        form = FinishedGoodProductForm(request.POST)
        # The formset needs a temporary, unsaved model instance to bind to
        temp_product = Product()
        formset = SpecFormSetCreate(request.POST, instance=temp_product)

        if form.is_valid() and formset.is_valid():
            # Save the main product form first
            product = form.save()
            
            # Now associate the saved product with the formset and save the specs
            formset.instance = product
            formset.save()
            
            messages.success(request, f"Finished Good Product '{product.name}' created successfully.")
            logger.info(f"Finished Good Product '{product.name}' created by {request.user.username}.")
            return redirect("qc:product_list")
        else:
            messages.error(request, "Please correct the errors below.")
            logger.warning(f"Finished Good creation failed by {request.user.username}. Errors: {form.errors}, Formset errors: {formset.errors}")
    else:
        form = FinishedGoodProductForm()
        formset = SpecFormSetCreate(instance=Product())

    # This data is needed for both GET and failed POST requests
    finished_good_data = _get_finished_good_data()

    context = {
        "form": form,
        "formset": formset,
        "action": "Create Finished Good Product",
        "finished_good_data": finished_good_data, # Pass data for JavaScript
    }
    return render(request, "products/fg_product_form.html", context)




@login_required
def product_update(request, pk):
    """
    Update QC details. It now directs to the correct form based on whether
    the product is a 'Finished Good'.
    """
    if not request.user.has_perm('QC.change_product'):
        logger.warning(f"Unauthorized Edit attempt by {request.user.username}")
        messages.error(request, "You do not have permission to Update Product records.")
        return redirect('indexpage')
   
    product = get_object_or_404(Product, pk=pk)
    
    # Check if the product's item_type is 'Finished Good'
    is_finished_good = product.item_type and 'finished good' in product.item_type.lower()

    if is_finished_good:
        # --- LOGIC FOR EDITING A FINISHED GOOD PRODUCT ---
        finished_good_data = _get_finished_good_data()

        if request.method == "POST":
            form = FinishedGoodProductForm(request.POST, instance=product)
            formset = SpecFormSetUpdate(request.POST, instance=product)

            if form.is_valid() and formset.is_valid():
                form.save()
                formset.save()
                messages.success(request, f"Product '{product.name}' updated successfully.")
                logger.info(f"Product '{product.name}' updated by {request.user.username}.")
                return redirect("qc:product_list")
            else:
                messages.error(request, "There was an error updating the product. Please correct the errors below.")
                # Optional: Add more detailed logging here
        else:
            form = FinishedGoodProductForm(instance=product)
            formset = SpecFormSetUpdate(instance=product)

        return render(request, "products/fg_product_form.html", {
            "form": form,
            "formset": formset,
            "action": "Update",
            "product": product, # Pass instance for template context if needed
            "finished_good_data": finished_good_data,
        })

    else:
        # --- ORIGINAL LOGIC FOR EDITING OTHER PRODUCT TYPES ---
        stages = _get_stages_from_bom()
        stages = _ensure_fg_name(stages)

        if request.method == "POST":
            form = ProductForm(request.POST, instance=product)
            formset = SpecFormSetUpdate(request.POST, instance=product)

            if form.is_valid() and formset.is_valid():
                form.save()
                formset.save()
                messages.success(request, f"Product '{product.name}' updated successfully.")
                logger.info(f"Product '{product.name}' updated by {request.user.username}.")
                return redirect("qc:product_list")
            else:
                messages.error(request, "There was an error updating the product. Please correct the errors below.")
        else:
            form = ProductForm(instance=product)
            formset = SpecFormSetUpdate(instance=product)

        return render(request, "products/product_form.html", {
            "form": form,
            "formset": formset,
            "action": "Update",
            "stages": stages,
            # Pass original context variables if needed
            "appearance_options": list(AppearanceOption.objects.values_list("name", flat=True).order_by("name")),
        })


@login_required
def product_delete(request, pk):
    """ Delete QC details (Permission Required: QC.delete_product) """
    if not request.user.has_perm('QC.add_product'):
        logger.warning(f"Unauthorized Delete attempt by {request.user.username}")
        messages.error(request, "You do not have permission to delete Product records.")
        return redirect('indexpage')
   
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        logger.info(f"Product '{product.name}' deleted by {request.user.username}.")
        product.delete()
        messages.success(request, f"Product '{product.name}' deleted.")
        return redirect('qc:product_list')
    return render(request, 'products/product_confirm_delete.html', {'product': product})



@login_required
def home(request):
    """
    A simple home view: renders templates/home.html
    """
    return render(request, 'home.html')



@login_required
def import_appearance_view(request):
    # --- Handle Excel import POST ---
    if request.method == "POST":
        form = ImportAppearanceForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = form.cleaned_data['file']
            try:
                wb = openpyxl.load_workbook(excel_file)
            except Exception as e:
                logger.error(f"Could not open the uploaded appearance file: {e}")
                messages.error(request, f"Could not open the uploaded file: {e}")
                # still show list
            else:
                sheet = wb.active
                imported_count = skipped_count = 0
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    raw_name = row[0]
                    if not raw_name:
                        continue
                    name = str(raw_name).strip()
                    if not name:
                        continue
                    obj, created = AppearanceOption.objects.get_or_create(name=name)
                    if created:
                        imported_count += 1
                    else:
                        skipped_count += 1
                messages.success(
                    request,
                    f"Import complete: {imported_count} new option(s) added, {skipped_count} skipped."
                )
                logger.info(f"Appearance import by {request.user.username}: {imported_count} added, {skipped_count} skipped.")
                return redirect("qc:import_appearance")
    else:
        form = ImportAppearanceForm()

    # --- List + filter logic (right side) ---
    query = request.GET.get('q', '').strip()
    all_options = AppearanceOption.objects.all()
    if query:
        all_options = all_options.filter(name__icontains=query)
    paginator = Paginator(all_options, 12)  # 12 per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
   
    return render(request, "products/import_appearance.html", {
        'form': form,
        'page_obj': page_obj,
        'query': query,
    })


@login_required
def import_specs(request, pk):
    product = get_object_or_404(Product, pk=pk)

    if request.method == 'POST':
        form = SpecUploadForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data['excel_file']
            try:
                df = pd.read_excel(file)
            except Exception as e:
                logger.error(f"Spec Excel import failed for product {product.name}: {e}")
                messages.error(request, f"Could not read Excel: {e}")
                return redirect('qc:spec_upload', pk=pk)

            required = {'Name', 'Type', 'Choices', 'Min Value', 'Max Value'}
            missing = required - set(df.columns)
            if missing:
                logger.warning(f"Spec import missing columns: {missing} (User: {request.user.username})")
                messages.error(request, f"Missing columns: {', '.join(missing)}")
                return redirect('qc:spec_upload', pk=pk)

            created = updated = 0
            for _, row in df.iterrows():
                name      = str(row['Name']).strip()
                spec_type = str(row['Type']).strip().lower()
                choices   = str(row['Choices']).strip() or ''
                minv      = row['Min Value']
                maxv      = row['Max Value']

                defaults = {
                    'spec_type':      spec_type,
                    'allowed_choices': choices,
                    'min_val':        minv if pd.notna(minv) else None,
                    'max_val':        maxv if pd.notna(maxv) else None,
                }
                spec, created_flag = Spec.objects.update_or_create(
                    product=product,
                    name=name,
                    defaults=defaults
                )
                if created_flag:
                    created += 1
                else:
                    updated += 1

            messages.success(
                request,
                f"Import complete: {created} created, {updated} updated."
            )
            logger.info(f"Specs imported for '{product.name}' by {request.user.username}: {created} created, {updated} updated.")
            return redirect('qc:product_update', pk=pk)

    else:
        form = SpecUploadForm()

    return render(request, 'products/spec_upload.html', {
        'product': product,
        'form':     form,
    })







#========================== QC Home Page =====================================#

class TruncFortnight(Func):
    function = 'DATEADD'
    template = """
      DATEADD(
        day,
        -(
          DATEDIFF(day, '1900-01-01', %(expressions)s) %%%% 14
        ),
        %(expressions)s
      )
    """
    output_field = DateField()


@login_required
def dashboard(request):
    logger.info("[QC][Dashboard] user=%s accessed QC Home Page", request.user.username)
    status_counts   = QCEntry.objects.order_by().values('status').annotate(count=Count('pk'))
    decision_counts = QCEntry.objects.order_by().values('decision_status').annotate(count=Count('pk'))

    stats = {
        'total':      QCEntry.objects.count(),
        'draft':      next((c['count'] for c in status_counts   if c['status']=='draft'), 0),
        'pending':    next((c['count'] for c in status_counts   if c['status']=='pending_qc'), 0),
        'completed':  next((c['count'] for c in status_counts   if c['status']=='qc_completed'), 0),
        'cancelled':  next((c['count'] for c in status_counts   if c['status']=='cancelled'), 0),
        'approved':   next((c['count'] for c in decision_counts if c['decision_status']=='approved'), 0),
        'variation':  next((c['count'] for c in decision_counts if c['decision_status']=='approved_under_deviation'), 0),
        'rejected':   next((c['count'] for c in decision_counts if c['decision_status']=='rejected'), 0),
        'fail':       next((c['count'] for c in decision_counts if c['decision_status']=='fail'), 0),
        'nodecision': next((c['count'] for c in decision_counts if c['decision_status'] is None), 0),
    }

    freq = request.GET.get('freq', 'daily')
    if freq == 'weekly':
        trunc = TruncWeek('entry_date')
    elif freq == 'monthly':
        trunc = TruncMonth('entry_date')
    elif freq == 'fortnightly':
        trunc = TruncFortnight('entry_date')
    else:
        trunc = TruncDay('entry_date')

    period_data = (QCEntry.objects.annotate(period=trunc).values('period').annotate(count=Count('pk'))
        .order_by('period'))

    labels     = [row['period'].strftime('%Y-%m-%d') for row in period_data]
    counts     = [row['count'] for row in period_data]
    cumulative = list(itertools.accumulate(counts))

    logger.info(
        "[QC][Dashboard] user=%s freq=%s total=%d draft=%d pending=%d completed=%d "
        "approved=%d variation=%d rejected=%d fail=%d nodecision=%d",
        request.user.username, freq,
        stats["total"], stats["draft"], stats["pending"], stats["completed"],
        stats["approved"], stats["variation"], stats["rejected"],
        stats["fail"], stats["nodecision"], )
    return render(request, 'home.html', {'stats': stats,'freq': freq,'chart_labels': labels,
        'chart_counts': counts,'chart_cumulative': cumulative,})


#========================== QC List Page =====================================#


@login_required
def qc_list(request):
    logger.info("[QC][List] user=%s accessed QC List Page", request.user.username)
    """ View QC details with pagination and advanced filtering (Permission Required: QC.view_qcentry) """
    if not request.user.has_perm('QC.view_qcentry'):
        logger.warning(f"[QC] Unauthorized view attempt by user '{request.user.username}' ")
        messages.error(request, "You do not have permission to View QC records.")
        return redirect('indexpage')

    # --- Get all filter parameters from the request ---
    q_param    = request.GET.get('q', '').strip()
    status     = request.GET.get('status')
    decision   = request.GET.get('decision')
    item_type  = request.GET.get('item_type')  # item_type filter
    start_date = request.GET.get('start_date')
    end_date   = request.GET.get('end_date')

    # Base queryset
    qs = QCEntry.objects.select_related('product')

    # 🔴 IMPORTANT: do NOT show SFG entries with no decision in this list
    qs = qs.exclude(Q(ar_type='SFG') & Q(decision_status__isnull=True))

    # --- Apply filters to the queryset ---
    if q_param:
        logger.info(
            "[QC] User '%s' searched QC list with query: '%s'",
            request.user.username, q_param
        )
        qs = qs.filter(
            Q(product__name__icontains=q_param) |
            Q(batch_no__icontains=q_param) |
            Q(entry_no__icontains=q_param)  # use entry_no instead of id
        )

    if status in ['draft', 'pending_qc', 'qc_completed', 'cancelled']:
        logger.info(
            "[QC] User '%s' filtered QC list by status: '%s'",
            request.user.username, status
        )
        qs = qs.filter(status=status)

    # ---------- DECISION FILTER (handles NULL) ----------
    if decision == 'nodecision':
        logger.info(
            "[QC] User '%s' filtered QC list by decision: NULL (no decision)",
            request.user.username
        )
        qs = qs.filter(decision_status__isnull=True)
    elif decision in ['approved', 'approved_under_deviation', 'rejected', 'fail']:
        logger.info(
            "[QC] User '%s' filtered QC list by decision: '%s'",
            request.user.username, decision
        )
        qs = qs.filter(decision_status=decision)

    # <-- Item Type filter -->
    if item_type:
        logger.info(
            "[QC] User '%s' filtered QC list by item type: '%s'",
            request.user.username, item_type
        )
        qs = qs.filter(product__item_type=item_type)

    # Date filters
    if start_date:
        qs = qs.filter(entry_date__gte=start_date)
    if end_date:
        try:
            # make end_date inclusive
            end_date_dt = datetime.strptime(end_date, '%Y-%m-%d')
            qs = qs.filter(entry_date__lt=end_date_dt + timedelta(days=1))
        except (ValueError, TypeError):
            pass

    ordered_entries = qs.order_by('-entry_date', '-id')

    # --- Pagination Logic ---
    paginator = Paginator(ordered_entries, 20)  # Show 20 entries per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    logger.info(
        "[QC] User '%s' loaded QC list page %s of %s.",
        request.user.username, page_obj.number, paginator.num_pages
    )

    # --- Preserve Filters During Pagination ---
    query_params = request.GET.copy()
    if 'page' in query_params:
        del query_params['page']

    # Distinct item types for dropdown
    item_type_choices = (
        Product.objects
        .exclude(item_type__isnull=True)
        .exclude(item_type__exact='')
        .values_list('item_type', flat=True)
        .distinct()
        .order_by('item_type')
    )

    # 🔔 SFG draft notification count (status = pending_qc & no decision)
    sfg_draft_count = QCEntry.objects.filter(
        ar_type='SFG',
        status='pending_qc',
        decision_status__isnull=True,
    ).count()

    # --- Prepare context for rendering ---
    context = {
        'entries': page_obj,
        'query_params': query_params.urlencode(),
        'q': q_param,
        'status': status,
        'decision': decision,
        'item_type': item_type,
        'item_type_choices': item_type_choices,
        'start_date': start_date,
        'end_date': end_date,
        'sfg_draft_count': sfg_draft_count,
    }
    return render(request, 'qc/qc_list.html', context)

#========================== QC Draft List =====================================#

@login_required 
def qc_draft_list(request): 
    """
    Draft QC entries:

    - Only entries with decision_status IS NULL
    - AND status = 'pending_qc'
    """
    logger.info("[QC][DraftList] user=%s accessed QC Draft List Page", request.user.username)

    if not request.user.has_perm('QC.view_qcentry'):
        logger.warning(
            "[QC] Unauthorized draft view attempt by user '%s'",
            request.user.username,
        )
        messages.error(request, "You do not have permission to view draft QC records.")
        return redirect('indexpage')

    # --- Get all filter parameters from the request ---
    q_param    = request.GET.get('q', '').strip()
    # status filter is no longer needed; always pending_qc
    item_type  = request.GET.get('item_type')       # item_type filter
    start_date = request.GET.get('start_date')
    end_date   = request.GET.get('end_date')

    # Base queryset: ONLY entries with NO decision yet AND status = 'pending_qc'
    qs = (
        QCEntry.objects
        .select_related('product')
        .filter(decision_status__isnull=True, status='pending_qc',ar_type='SFG')
    )

    # --- Apply filters to the queryset ---
    if q_param:
        logger.info(
            "[QC][DraftList] User '%s' searched draft list with query: '%s'",
            request.user.username, q_param
        )
        qs = qs.filter(
            Q(product__name__icontains=q_param) |
            Q(batch_no__icontains=q_param) |
            Q(entry_no__icontains=q_param)
        )

    # Item Type filter (same as qc_list)
    if item_type:
        logger.info(
            "[QC][DraftList] User '%s' filtered draft list by item type: '%s'",
            request.user.username, item_type
        )
        qs = qs.filter(product__item_type=item_type)

    # Date filters on entry_date
    if start_date:
        qs = qs.filter(entry_date__gte=start_date)

    if end_date:
        try:
            end_date_dt = datetime.strptime(end_date, '%Y-%m-%d')
            qs = qs.filter(entry_date__lt=end_date_dt + timedelta(days=1))
        except (ValueError, TypeError):
            # ignore invalid date
            pass

    ordered_entries = qs.order_by('-entry_date', '-id')

    # --- Pagination Logic ---
    paginator = Paginator(ordered_entries, 20)  # 20 per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    logger.info(
        "[QC][DraftList] User '%s' loaded draft list page %s of %s.",
        request.user.username, page_obj.number, paginator.num_pages
    )

    # --- Preserve filters during pagination ---
    query_params = request.GET.copy()
    if 'page' in query_params:
        del query_params['page']

    # Item type choices for filter dropdown
    item_type_choices = (
        Product.objects
        .exclude(item_type__isnull=True)
        .exclude(item_type__exact='')
        .values_list('item_type', flat=True)
        .distinct()
        .order_by('item_type')
    )

    context = {
        'entries': page_obj,
        'query_params': query_params.urlencode(),
        'q': q_param,
        # status is fixed to pending_qc in the backend; we can pass for info if needed
        'status': 'pending_qc',
        'decision': None,          # no decision filter here; all are NULL
        'item_type': item_type,
        'item_type_choices': item_type_choices,
        'start_date': start_date,
        'end_date': end_date,
        'is_draft_view': True,     # flag for template if needed
    }
    return render(request, 'qc/draft_report.html', context)



#========================== QC List Export =====================================#

@login_required
def qc_export_excel(request):
    logger.info("[QC][Export] user=%s Download QC Lists", request.user.username)
    import datetime
    if not request.user.has_perm('QC.view_qcentry'):
        messages.error(request, "You do not have permission to export QC records.")
        return redirect('indexpage')
    # ---- Read filters exactly like qc_list ----
    q_param    = (request.GET.get('q') or '').strip()
    status     = request.GET.get('status')
    decision   = request.GET.get('decision')
    item_type  = request.GET.get('item_type')
    start_date = request.GET.get('start_date')
    end_date   = request.GET.get('end_date')

    # Base queryset (same as qc_list, including SFG exclusion)
    qs = QCEntry.objects.select_related('product', 'created_by', 'qc_completed_by')

    # 🔴 IMPORTANT: do NOT export SFG entries with no decision in this list
    qs = qs.exclude(Q(ar_type='SFG') & Q(decision_status__isnull=True))

    # --- Apply filters (mirror qc_list) ---
    if q_param:
        logger.info(
            "[QC][Export] User '%s' searched QC list with query: '%s'",
            request.user.username, q_param
        )
        qs = qs.filter(
            Q(product__name__icontains=q_param) |
            Q(batch_no__icontains=q_param) |
            Q(entry_no__icontains=q_param)
        )

    if status in ['draft', 'pending_qc', 'qc_completed', 'cancelled']:
        logger.info(
            "[QC][Export] User '%s' filtered QC list by status: '%s'",
            request.user.username, status
        )
        qs = qs.filter(status=status)

    # ---------- DECISION FILTER (handles NULL) ----------
    if decision == 'nodecision':
        # special token from dashboard card -> decision_status IS NULL
        logger.info(
            "[QC][Export] User '%s' filtered QC list by decision: NULL (no decision)",
            request.user.username
        )
        qs = qs.filter(decision_status__isnull=True)
    elif decision in ['approved', 'approved_under_deviation', 'rejected', 'fail']:
        logger.info(
            "[QC][Export] User '%s' filtered QC list by decision: '%s'",
            request.user.username, decision
        )
        qs = qs.filter(decision_status=decision)

    if item_type:
        logger.info(
            "[QC][Export] User '%s' filtered QC list by item type: '%s'",
            request.user.username, item_type
        )
        qs = qs.filter(product__item_type=item_type)

    if start_date:
        qs = qs.filter(entry_date__gte=start_date)

    if end_date:
        try:
            end_dt = datetime.datetime.strptime(end_date, '%Y-%m-%d') + datetime.timedelta(days=1)
            qs = qs.filter(entry_date__lt=end_dt)
        except Exception:
            pass

    qs = qs.order_by('-entry_date', '-id')

    # ---- Build workbook in-memory ----
    output = BytesIO()
    wb = xlsxwriter.Workbook(output, {'in_memory': True})
    ws = wb.add_worksheet("QC Entries")

    # Formats
    title_fmt = wb.add_format({'bold': True, 'font_size': 14})
    ason_fmt  = wb.add_format({'italic': True, 'font_color': '#6b7280'})  # grey-ish
    head_fmt  = wb.add_format({'bold': True, 'border': 1, 'bg_color': '#EAF2FF'})
    cell_fmt  = wb.add_format({'border': 1})
    num_fmt   = wb.add_format({'border': 1})
    dt_fmt    = wb.add_format({'border': 1, 'num_format': 'yyyy-mm-dd hh:mm'})

    # ---- Title row (A1 + B1) ----
    today_str = datetime.date.today().strftime('%d-%b-%Y')
    ws.write(0, 0, "QC Entries", title_fmt)           # A1
    ws.write(0, 1, f"As on {today_str}", ason_fmt)    # B1
    ws.set_row(0, 20)

    # Headers
    headers = [
        "Entry No", "Product", "stage", "batch_no", "block", "decision_status",
        "status", "Item Type", "group", "ar_no", "sample_sent_at", "created_by",
        "sample_received_at", "qc_completed_by", "general_remarks", "Qty",
        "instrument_id", "frequency", "sample_description", "test_parameters",
    ]
    start_row = 2  # blank row between title and headers
    for c, h in enumerate(headers):
        ws.write(start_row, c, h, head_fmt)

    # Helper to write datetimes, stripping tzinfo if present
    def write_dt(r, c, dt):
        if not dt:
            ws.write(r, c, "", cell_fmt)
            return
        try:
            if getattr(dt, "tzinfo", None) is not None and dt.tzinfo.utcoffset(dt) is not None:
                dt = dt.replace(tzinfo=None)
        except Exception:
            dt = dt.replace(tzinfo=None)
        ws.write_datetime(r, c, dt, dt_fmt)

    # Rows
    row = start_row + 1
    for e in qs:
        created_by = (
            getattr(e.created_by, "get_full_name", lambda: "")() or e.created_by.username
        ) if e.created_by else ""
        qc_by = (
            getattr(e.qc_completed_by, "get_full_name", lambda: "")() or e.qc_completed_by.username
        ) if e.qc_completed_by else ""

        values = [
            e.entry_no or "",
            e.product.name if e.product_id else "",
            e.get_stage_display() if hasattr(e, "get_stage_display") else (e.stage or ""),
            e.batch_no or "",
            e.block or "",
            e.get_decision_status_display() if e.decision_status else "",
            e.get_status_display() if hasattr(e, "get_status_display") else (e.status or ""),
            e.ar_type or "",
            e.group or "",
            e.ar_no or "",
            None,             # sample_sent_at (datetime)
            created_by,
            None,             # sample_received_at (datetime)
            qc_by,
            e.general_remarks or "",
            e.qty if e.qty is not None else "",
            e.instrument_id or "",
            e.frequency or "",
            e.sample_description or "",
            e.test_parameters or "",
        ]

        # Write simple cells up to "ar_no"
        for col in range(0, 10):
            ws.write(row, col, values[col], cell_fmt)

        write_dt(row, 10, e.sample_sent_at)     # sample_sent_at
        ws.write(row, 11, values[11], cell_fmt) # created_by
        write_dt(row, 12, e.sample_received_at) # sample_received_at
        ws.write(row, 13, values[13], cell_fmt) # qc_completed_by
        ws.write(row, 14, values[14], cell_fmt) # general_remarks

        if e.qty is not None:                   # Qty numeric
            ws.write_number(row, 15, float(e.qty), num_fmt)
        else:
            ws.write(row, 15, "", cell_fmt)

        ws.write(row, 16, values[16], cell_fmt) # instrument_id
        ws.write(row, 17, values[17], cell_fmt) # frequency
        ws.write(row, 18, values[18], cell_fmt) # sample_description
        ws.write(row, 19, values[19], cell_fmt) # test_parameters

        row += 1

    # Column widths
    widths = {
        0:12, 1:36, 2:16, 3:18, 4:16, 5:22, 6:18, 7:10, 8:16, 9:22,
        10:20, 11:20, 12:20, 13:20, 14:32, 15:10, 16:18, 17:14, 18:30, 19:30
    }
    for c, w in widths.items():
        ws.set_column(c, c, w)

    wb.close()
    output.seek(0)

    filename = f"QC_Entries_{datetime.date.today().isoformat()}.xlsx"
    resp = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp

   

@login_required
def qc_detail(request, pk):
    """ View QC details (Permission Required: QC.view_qcentry) """
    if not request.user.has_perm('QC.view_qcentry'):
        logger.warning(f"Unauthorized View attempt by {request.user.username}")
        messages.error(request, "You do not have permission to View QC records.")
        return redirect('indexpage')
    # Fetch the QCEntry object
    entry = get_object_or_404(QCEntry, pk=pk)

    # 1) fetch any saved SpecEntry rows (for lookup)
    spec_entries = (
        SpecEntry.objects
                 .filter(qc_entry=entry)
                 .select_related('spec')
    )

    # 2) find all groups present in this QC entry
    groups = (
        spec_entries
            .values_list('spec__group', flat=True)
            .distinct()
            .order_by('spec__group')
    )

    # 3) for each group, get the specs and their SpecEntry values
    group_rows = []
    entry_by_spec = { se.spec_id: se for se in spec_entries }
    for group in groups:
        specs = Spec.objects.filter(product=entry.product, group=group).order_by('id')
        rows = [
            { 'spec': spec, 'entry': entry_by_spec.get(spec.id) }
            for spec in specs
        ]
        group_rows.append({
            'group': group,
            'rows': rows,
        })

    logger.info(
        "QC detail viewed by %s for QCEntry %s (groups=%s)",
        request.user.username,
        entry.pk,
        list(groups),
    )

    return render(request, 'qc/qc_detail.html', {
        'entry':          entry,
        'group_rows':     group_rows,  # all group+rows
    })



def get_spec_groups_for_product(request):
    product_id = request.GET.get('product_id')
    if not product_id:
        return JsonResponse({'groups': []})
    try:
        groups = list(
            Spec.objects.filter(product_id=product_id, group__isnull=False)
            .exclude(group__exact='')
            .values_list('group', flat=True)
            .distinct().order_by('group')
        )
        return JsonResponse({'groups': groups})
    except Exception:
        return JsonResponse({'error': 'An error occurred'}, status=500)



def get_test_parameters_for_group(request):
    """
    API endpoint that returns a list of test parameter names
    for a given product and spec group.
    """
    product_id = request.GET.get('product_id')
    group = request.GET.get('group')

    if not product_id or not group:
        return JsonResponse({'parameters': []})

    # Find all spec names for the given product and group.
    # These names should correspond to the names in QCTestParameter.
    parameter_names = list(
        Spec.objects.filter(
            product_id=product_id,
            group=group
        ).values_list('name', flat=True)
    )

    # Now, find all QCTestParameter objects that match these names.
    parameters = list(
        QCTestParameter.objects.filter(name__in=parameter_names)
        .values('id', 'name') # Return ID and name for the dropdown
    )
   
    return JsonResponse({'parameters': parameters})


@login_required
def qc_create(request):
    """ Add QC details (Permission Required: QC.add_qcentry) """
    if not request.user.has_perm('QC.add_qcentry'):
        logger.warning(f"Unauthorized Add attempt by {request.user.username}")
        messages.error(request, "You do not have permission to add QC records.")
        return redirect('indexpage')

    # Helper: one-time ERP sync (non-fatal)
    def _maybe_sync_bmr_issue_from_erp():
        if BmrIssue.objects.exists():
            return
        try:
            with connections['readonly_db'].cursor() as cursor:
                cursor.execute(""" /* your ERP sync SQL... */ """)
                columns = [c[0] for c in cursor.description]
                rows = cursor.fetchall() or []
            to_create = []
            for row in rows:
                data = dict(zip(columns, row))
                to_create.append(BmrIssue(
                    bmr_issue_type = data.get('BMR_Issue_Type', ''),
                    bmr_issue_no   = data.get('BMR_Issue_No', ''),
                    bmr_issue_date = data.get('BMR_Issue_Date'),
                    fg_name        = data.get('FG_Name', ''),
                    op_batch_no    = data.get('OP_Batch_No', ''),
                    product_name   = data.get('Product_Name','') or '',
                    block          = data.get('Block','') or '',
                    line_no        = data.get('Line_No'),
                    item_type      = data.get('Item_Type', ''),
                    item_code      = data.get('Item_Code', ''),
                    item_name      = data.get('Item_Name', ''),
                    item_narration = data.get('Item_Narration','') or '',
                    uom            = data.get('UOM', ''),
                    batch_quantity = data.get('Batch_Quantity'),
                ))
            if to_create:
                BmrIssue.objects.bulk_create(to_create)
                logger.info(
                    "BmrIssue sync: %s rows loaded from ERP by %s",
                    len(to_create), request.user.username
                )
        except Exception as sync_exc:
            logger.warning(
                "Could not sync BmrIssue from ERP: %s", sync_exc, exc_info=True
            )
            messages.warning(
                request,
                "Warning: ERP sync failed; some dropdowns may be incomplete."
            )

    # helper: load blocks & equipments from BOTH production_scheduler.equipment
    # and qc_localequipmentmaster (LocalEquipmentMaster) for selected blocks
    def _load_scheduler_equipment():
        """
        Returns:
            block_list: list of distinct block names
            equipment_list: list of dicts with keys:
                - eqp_code
                - eqp_name
                - block_name
        """

        def normalize_block_name(block_value):
            val = (block_value or "").strip()
            if not val:
                return ""

            key = val.upper().replace("_", " ")
            key = " ".join(key.split())

            mapping = {
                "BLOCK A": "A-Block",
                "A BLOCK": "A-Block",
                "A-BLOCK": "A-Block",

                "BLOCK B": "B-Block",
                "B BLOCK": "B-Block",
                "B-BLOCK": "B-Block",

                "BLOCK C": "C-Block",
                "C BLOCK": "C-Block",
                "C-BLOCK": "C-Block",

                "BLOCK D": "D-Block",
                "D BLOCK": "D-Block",
                "D-BLOCK": "D-Block",

                "BLOCK E": "E-Block",
                "E BLOCK": "E-Block",
                "E-BLOCK": "E-Block",

                "TANK FARM": "Tank Farm",
                "FIRE PUMP HOUSE": "Fire Pump House",
                "ELECTRICAL PCC ROOM": "ELECTRICAL PCC ROOM",
                "DG HOUSE": "DG House",
                "BOILER": "Boiler",
                "ETP": "ETP",
                "MEE": "MEE",
                "ETP/MEE": "ETP/MEE",
                "PILOT PLANT": "Pilot Plant",
                "ETP RO": "ETP RO",
                "HT YEARD": "HT Yeard",
            }

            return mapping.get(key, val)

        allowed_blocks = {
            "A-Block",
            "B-Block",
            "C-Block",
            "D-Block",
            "E-Block",
            "Tank Farm",
            "Fire Pump House",
            "ELECTRICAL PCC ROOM",
            "DG House",
            "Boiler",
            "ETP",
            "MEE",
            "ETP/MEE",
            "Pilot Plant",
            "ETP RO",
            "HT Yeard",
        }

        block_set = set()
        equipments = []
        seen_pairs = set()

        # ---- Source 1: production_scheduler.equipment ----
        try:
            with connections["production_scheduler"].cursor() as cursor:
                cursor.execute(
                    """
                    SELECT block, eq_id
                    FROM equipment
                    WHERE block IS NOT NULL
                    AND LTRIM(RTRIM(block)) <> ''
                    """
                )
                rows = cursor.fetchall() or []

            for block_val, eq_id in rows:
                block_text = normalize_block_name(block_val)
                if not block_text:
                    continue

                if block_text not in allowed_blocks:
                    continue

                key = (str(eq_id), block_text)
                if key in seen_pairs:
                    continue
                seen_pairs.add(key)

                block_set.add(block_text)
                equipments.append({
                    "eqp_code": eq_id,
                    "eqp_name": eq_id,
                    "block_name": block_text,
                })

        except Exception as exc:
            logger.error(
                "Error loading equipment from production_scheduler.equipment: %s",
                exc,
                exc_info=True,
            )
            messages.warning(
                request,
                "Warning: Could not load equipment list from production scheduler."
            )

        # ---- Source 2: LocalEquipmentMaster ----
        try:
            qs = (
                LocalEquipmentMaster.objects
                .exclude(block_name__isnull=True)
                .exclude(block_name__exact='')
                .order_by('block_name', 'eqp_name', 'eqp_code')
            )

            for obj in qs:
                block_text = normalize_block_name(obj.block_name)
                if not block_text:
                    continue

                if block_text not in allowed_blocks:
                    continue

                eq_code = obj.eqp_code
                eq_name = obj.eqp_name or obj.eqp_code

                key = (str(eq_code), block_text)
                if key in seen_pairs:
                    continue
                seen_pairs.add(key)

                block_set.add(block_text)
                equipments.append({
                    "eqp_code": eq_code,
                    "eqp_name": eq_name,
                    "block_name": block_text,
                })

        except Exception as exc:
            logger.error(
                "Error loading equipment from LocalEquipmentMaster: %s",
                exc,
                exc_info=True,
            )
            messages.warning(
                request,
                "Warning: Could not load equipment list from LocalEquipmentMaster."
            )

        block_list = sorted(block_set)
        equipment_list = sorted(
            equipments,
            key=lambda e: (e["block_name"], str(e["eqp_code"]))
        )
        return block_list, equipment_list

    try:
        # 1) Possibly sync from ERP (non-blocking)
        _maybe_sync_bmr_issue_from_erp()

        # 2) Batch dropdown with defensive fallback
        try:
            distinct_batch_qs = (
                BmrIssue.objects
                .filter(bmr_issue_date__gte=date(2025, 1, 25))
                .values('op_batch_no')
                .annotate(latest_issue=Max('bmr_issue_date'))
                .order_by('-latest_issue')
            )
            distinct_batch_nos = [row['op_batch_no'] for row in distinct_batch_qs]
            selected_batch_no = request.GET.get('batch_no_filter') or None
            if selected_batch_no:
                qc_query_results = BmrIssue.objects.filter(
                    op_batch_no=selected_batch_no
                ).order_by('bmr_issue_no', 'line_no')
            else:
                qc_query_results = BmrIssue.objects.all().order_by(
                    'bmr_issue_no', 'line_no'
                )
        except Exception as bmr_exc:
            logger.error(
                "Error querying BmrIssue for batch/options: %s",
                bmr_exc,
                exc_info=True,
            )
            distinct_batch_nos = []
            qc_query_results = BmrIssue.objects.none()
            selected_batch_no = None
            messages.warning(
                request, "Warning: Unable to load batch information."
            )

        # 3) NEW: load blocks & equipments from both sources
        block_list, equipment_list = _load_scheduler_equipment()

        # 4) Stage dropdown (used for Stage → Product sync)
        stage_options = (
            Product.objects
            .exclude(stages__isnull=True)
            .exclude(stages__exact='')
            .values('stages', 'id')
            .distinct()
            .order_by('stages')
        )

        # Product dropdown (for showing product name)
        product_list = Product.objects.values('id', 'name').order_by('name')

        # Sample description options
        sample_description_options = list(
            SampleDescriptionOption.objects.order_by("name")
            .values_list("name", flat=True)
        )

        # 5) Form handling
        if request.method == 'POST':
            form = ProductionQCEntryForm(
                request.POST,
                sample_description_options=sample_description_options,
                stage_choices=stage_options,
                fixed_ar_type="IP",
            )
            logger.debug(
                "QCEntry POST DATA by %s: %s",
                request.user.username,
                dict(request.POST),
            )

            if not form.is_valid():
                logger.error("QCEntry FORM ERRORS: %s", form.errors.as_json())
                try:
                    logger.error(
                        "QCEntry FORM NON_FIELD ERRORS: %s",
                        form.non_field_errors().as_json(),
                    )
                except Exception:
                    logger.error("Error accessing non-field errors for QCEntry form.")
                logger.debug(
                    "QCEntry FORM cleaned_data: %s",
                    getattr(form, 'cleaned_data', {}),
                )
                for field, errs in form.errors.items():
                    messages.error(request, f"{field}: {errs}")
            else:
                try:
                    with transaction.atomic():
                        qc_entry = form.save(commit=False)
                        qc_entry.status = 'pending_qc'
                        qc_entry.created_by = request.user
                        qc_entry.ar_type = 'IP'  # harden server-side

                        # assign AR if missing (first submit or user tampering)
                        if not qc_entry.ar_no:
                            qc_entry.ar_no = _next_ar_no('IP')

                        qc_entry.save()

                        # Notify via email
                        send_mail(
                            subject=(
                                f"[QC] New entry ready for QC: "
                                f"{qc_entry.product.name} / {qc_entry.batch_no}"
                            ),
                            message=(
                                f"A new QCEntry (ID: {qc_entry.pk}) was submitted.\n"
                                f"Product: {qc_entry.product.name}\n"
                                f"Batch No: {qc_entry.batch_no}\n"
                                f"Stage: {qc_entry.get_stage_display()}\n"
                                f"Equipment: {qc_entry.equipment_id or '—'}\n"
                                f"Block: {qc_entry.block or '—'}\n"
                                f"Sample Description: {qc_entry.sample_description or '—'}\n"
                                f"Test Required For: {qc_entry.test_required_for or '—'}\n"
                                f"Test Parameter: {qc_entry.test_parameters or '—'}\n"
                            ),
                            from_email=getattr(
                                settings,
                                'DEFAULT_FROM_EMAIL',
                                'workflow@ocspl.com',
                            ),
                            recipient_list=['shakir.s@ocspl.com',
                            'vinod.nimbalkar@ocspl.com',
                            'qc@ocspl.com',
                            'vikas.m@ocspl.com'],
                            fail_silently=True,
                        )
                    logger.info(
                        "QCEntry #%s created by %s for product '%s', batch %s",
                        qc_entry.pk,
                        request.user.username,
                        qc_entry.product.name,
                        qc_entry.batch_no,
                    )
                    messages.success(
                        request,
                        "QC created. QC team has been notified."
                    )
                    return redirect('qc:qc_list')
                except Exception as save_exc:
                    logger.error(
                        "Exception during QCEntry save: %s",
                        save_exc,
                        exc_info=True,
                    )
                    messages.error(
                        request,
                        f"Error while saving QC Entry: {save_exc}"
                    )
        else:
            now = timezone.localtime()
            form = ProductionQCEntryForm(
                initial={
                    'entry_date': now,
                    'sample_sent_at': now,
                    'ar_type': 'IP',
                },
                sample_description_options=sample_description_options,
                stage_choices=stage_options,
                fixed_ar_type="IP",
            )

        return render(request, 'qc/qc_form_phase1.html', {
            'form':               form,
            'equipment_list':     equipment_list,   # merged list from both sources
            'block_list':         block_list,       # raw DB values (A-Block, Tank Farm, etc.)
            'distinct_batch_nos': distinct_batch_nos,
            'selected_batch_no':  selected_batch_no,
            'stage_options':      stage_options,
            'product_list':       product_list,
            'qc_query_results':   qc_query_results,
        })
    except Exception as e:
        logger.error(
            "Exception in qc_create by %s: %s",
            request.user.username,
            e,
            exc_info=True,
        )
        messages.error(
            request,
            "An unexpected error occurred. Please contact the admin."
        )
        return redirect('qc:qc_list')



@login_required
def fgqc_create(request):
    """Create FGQC header (same UX as In-Process, adds Qty)."""
    if not request.user.has_perm('QC.add_qcentry'):
        messages.error(request, "You do not have permission to add QC records.")
        return redirect('indexpage')

    # one-time ERP sync (copied pattern from qc_create; safe if it no-ops)
    def _maybe_sync_bmr_issue_from_erp():
        if BmrIssue.objects.exists():
            return
        try:
            with connections['readonly_db'].cursor() as cursor:
                cursor.execute(""" /* your ERP sync SQL... */ """)
                cols = [c[0] for c in cursor.description]
                rows = cursor.fetchall() or []
            to_create = []
            for r in rows:
                d = dict(zip(cols, r))
                to_create.append(BmrIssue(
                    bmr_issue_type = d.get('BMR_Issue_Type',''),
                    bmr_issue_no   = d.get('BMR_Issue_No',''),
                    bmr_issue_date = d.get('BMR_Issue_Date'),
                    fg_name        = d.get('FG_Name',''),
                    op_batch_no    = d.get('OP_Batch_No',''),
                    product_name   = d.get('Product_Name','') or '',
                    block          = d.get('Block','') or '',
                    line_no        = d.get('Line_No'),
                    item_type      = d.get('Item_Type',''),
                    item_code      = d.get('Item_Code',''),
                    item_name      = d.get('Item_Name',''),
                    item_narration = d.get('Item_Narration','') or '',
                    uom            = d.get('UOM',''),
                    batch_quantity = d.get('Batch_Quantity'),
                ))
            if to_create:
                BmrIssue.objects.bulk_create(to_create)
        except Exception:
            # non-fatal; page will still work with existing data
            pass

    # Helper: load blocks & equipments from BOTH production_scheduler.equipment
    # and qc_localequipmentmaster (LocalEquipmentMaster) for selected blocks
    def _load_scheduler_equipment():
        """
        Returns:
            block_list: list of distinct block names
            equipment_list: list of dicts with keys:
                - eqp_code
                - eqp_name
                - block_name
        """

        def normalize_block_name(block_value):
            val = (block_value or "").strip()
            if not val:
                return ""

            key = val.upper().replace("_", " ")
            key = " ".join(key.split())

            mapping = {
                "BLOCK A": "A-Block",
                "A BLOCK": "A-Block",
                "A-BLOCK": "A-Block",

                "BLOCK B": "B-Block",
                "B BLOCK": "B-Block",
                "B-BLOCK": "B-Block",

                "BLOCK C": "C-Block",
                "C BLOCK": "C-Block",
                "C-BLOCK": "C-Block",

                "BLOCK D": "D-Block",
                "D BLOCK": "D-Block",
                "D-BLOCK": "D-Block",

                "BLOCK E": "E-Block",
                "E BLOCK": "E-Block",
                "E-BLOCK": "E-Block",

                "TANK FARM": "Tank Farm",
                "FIRE PUMP HOUSE": "Fire Pump House",
                "ELECTRICAL PCC ROOM": "ELECTRICAL PCC ROOM",
                "DG HOUSE": "DG House",
                "BOILER": "Boiler",
                "ETP": "ETP",
                "MEE": "MEE",
                "ETP/MEE": "ETP/MEE",
                "PILOT PLANT": "Pilot Plant",
                "ETP RO": "ETP RO",
                "HT YEARD": "HT Yeard",
            }

            return mapping.get(key, val)

        allowed_blocks = {
            "A-Block",
            "B-Block",
            "C-Block",
            "D-Block",
            "E-Block",
            "Tank Farm",
            "Fire Pump House",
            "ELECTRICAL PCC ROOM",
            "DG House",
            "Boiler",
            "ETP",
            "MEE",
            "ETP/MEE",
            "Pilot Plant",
            "ETP RO",
            "HT Yeard",
        }

        block_set = set()
        equipments = []
        seen_pairs = set()

        # ---- Source 1: production_scheduler.equipment ----
        try:
            with connections["production_scheduler"].cursor() as cursor:
                cursor.execute(
                    """
                    SELECT block, eq_id
                    FROM equipment
                    WHERE block IS NOT NULL
                    AND LTRIM(RTRIM(block)) <> ''
                    """
                )
                rows = cursor.fetchall() or []

            for block_val, eq_id in rows:
                block_text = normalize_block_name(block_val)
                if not block_text:
                    continue

                if block_text not in allowed_blocks:
                    continue

                key = (str(eq_id), block_text)
                if key in seen_pairs:
                    continue
                seen_pairs.add(key)

                block_set.add(block_text)
                equipments.append({
                    "eqp_code": eq_id,
                    "eqp_name": eq_id,
                    "block_name": block_text,
                })

        except Exception as exc:
            logger.error(
                "Error loading equipment from production_scheduler.equipment: %s",
                exc,
                exc_info=True,
            )
            messages.warning(
                request,
                "Warning: Could not load equipment list from production scheduler."
            )

        # ---- Source 2: LocalEquipmentMaster ----
        try:
            qs = (
                LocalEquipmentMaster.objects
                .exclude(block_name__isnull=True)
                .exclude(block_name__exact='')
                .order_by('block_name', 'eqp_name', 'eqp_code')
            )

            for obj in qs:
                block_text = normalize_block_name(obj.block_name)
                if not block_text:
                    continue

                if block_text not in allowed_blocks:
                    continue

                eq_code = obj.eqp_code
                eq_name = obj.eqp_name or obj.eqp_code

                key = (str(eq_code), block_text)
                if key in seen_pairs:
                    continue
                seen_pairs.add(key)

                block_set.add(block_text)
                equipments.append({
                    "eqp_code": eq_code,
                    "eqp_name": eq_name,
                    "block_name": block_text,
                })

        except Exception as exc:
            logger.error(
                "Error loading equipment from LocalEquipmentMaster: %s",
                exc,
                exc_info=True,
            )
            messages.warning(
                request,
                "Warning: Could not load equipment list from LocalEquipmentMaster."
            )

        block_list = sorted(block_set)
        equipment_list = sorted(
            equipments,
            key=lambda e: (e["block_name"], str(e["eqp_code"]))
        )
        return block_list, equipment_list

    try:
        _maybe_sync_bmr_issue_from_erp()

        # dropdown sources (same as qc_create)
        try:
            distinct_batch_nos = (
                BmrIssue.objects
                .filter(bmr_issue_date__gte=date(2025, 1, 25))
                .values_list('op_batch_no', flat=True)
                .distinct().order_by('op_batch_no')
            )
            selected_batch_no = request.GET.get('batch_no_filter') or None
            if selected_batch_no:
                qc_query_results = BmrIssue.objects.filter(
                    op_batch_no=selected_batch_no
                ).order_by('bmr_issue_no', 'line_no')
            else:
                qc_query_results = BmrIssue.objects.all().order_by('bmr_issue_no', 'line_no')
        except Exception:
            distinct_batch_nos = []
            qc_query_results = BmrIssue.objects.none()
            selected_batch_no = None

        # NEW: merged block + equipment logic
        block_list, equipment_list = _load_scheduler_equipment()

        stage_options = (
            Product.objects
            .exclude(stages__isnull=True).exclude(stages__exact='')
            .values('stages', 'id').distinct().order_by('stages')
            .annotate(item_type_clean=Trim('item_type'))
            .filter(item_type_clean__iexact='Finished Good')
        )
        product_list = Product.objects.values('id', 'name').order_by('name')
        sample_description_options = list(
            SampleDescriptionOption.objects.order_by("name")
            .values_list("name", flat=True)
        )

        if request.method == "POST":
            form = FGQCEntryForm(
                request.POST,
                sample_description_options=sample_description_options,
                stage_choices=stage_options,
            )
            if form.is_valid():
                with transaction.atomic():
                    qc_entry = form.save(commit=False)
                    qc_entry.status = 'pending_qc'
                    qc_entry.created_by = request.user
                    qc_entry.ar_type = 'FG'
                    if not qc_entry.ar_no:
                        qc_entry.ar_no = _next_ar_no('FG')
                    qc_entry.save()

                    # ==========================================================
                    # START: Inserted Email Notification Logic
                    # ==========================================================
                    send_mail(
                        subject=f"[FGQC] New entry ready for QC: {qc_entry.product.name} / {qc_entry.batch_no}",
                        message=(
                            f"A new FGQC Entry (ID: {qc_entry.pk}) was submitted.\n"
                            f"Product: {qc_entry.product.name}\n"
                            f"Batch No: {qc_entry.batch_no}\n"
                            f"Stage: {qc_entry.get_stage_display()}\n"
                            f"Quantity: {getattr(qc_entry, 'qty', 'N/A')}\n"
                            f"Sample Description: {qc_entry.sample_description or '—'}\n"
                            f"Test Required For: {qc_entry.test_required_for or '—'}\n"
                            f"Test Parameter: {qc_entry.test_parameters or '—'}\n"
                        ),
                        from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', 'workflow@ocspl.com'),
                        recipient_list=['shakir.s@ocspl.com',
                            'vinod.nimbalkar@ocspl.com',
                            'qc@ocspl.com',
                            'vikas.m@ocspl.com'],
                        fail_silently=True,  # Set to False for debugging email issues
                    )
                    # ==========================================================
                    # END: Inserted Email Notification Logic
                    # ==========================================================

                messages.success(request, "FGQC header submitted. QC team has been notified.")
                return redirect('qc:qc_list')
            else:
                messages.error(request, "Please correct the errors below.")
        else:
            now = timezone.localtime()
            form = FGQCEntryForm(
                initial={
                    'entry_date': now,
                    'sample_sent_at': now,
                    'ar_type': 'FG',
                    'ar_no': _next_ar_no('FG'),
                },
                sample_description_options=sample_description_options,
                stage_choices=stage_options,
            )

        return render(request, "qc/fgqc_form.html", {
            "form":               form,
            "equipment_list":     equipment_list,  # merged
            "block_list":         block_list,      # merged
            "distinct_batch_nos": distinct_batch_nos,
            "selected_batch_no":  selected_batch_no,
            "stage_options":      stage_options,
            "product_list":       product_list,
            "qc_query_results":   qc_query_results,
        })

    except Exception as e:
        messages.error(request, "An unexpected error occurred. Please contact the admin.")
        return redirect('qc:qc_list')





@login_required
def sfgqc_create(request):
    """Create SFGQC header."""
    if not request.user.has_perm('QC.add_qcentry'):
        messages.error(request, "You do not have permission to add QC records.")
        return redirect('indexpage')

    # one-time ERP sync (copied pattern from fgqc_create; safe if it no-ops)
    def _maybe_sync_bmr_issue_from_erp():
        if BmrIssue.objects.exists():
            return
        try:
            with connections['readonly_db'].cursor() as cursor:
                # Replace with your actual ERP sync query if needed
                cursor.execute(""" SELECT * FROM your_erp_bmr_view WHERE 1=0 """)
                cols = [c[0] for c in cursor.description]
                rows = cursor.fetchall() or []
            to_create = []
            for r in rows:
                d = dict(zip(cols, r))
                to_create.append(BmrIssue(
                    bmr_issue_type = d.get('BMR_Issue_Type', ''),
                    bmr_issue_no   = d.get('BMR_Issue_No', ''),
                    bmr_issue_date = d.get('BMR_Issue_Date'),
                    fg_name        = d.get('FG_Name', ''),
                    op_batch_no    = d.get('OP_Batch_No', ''),
                    product_name   = d.get('Product_Name','') or '',
                    block          = d.get('Block','') or '',
                    line_no        = d.get('Line_No'),
                    item_type      = d.get('Item_Type', ''),
                    item_code      = d.get('Item_Code', ''),
                    item_name      = d.get('Item_Name', ''),
                    item_narration = d.get('Item_Narration','') or '',
                    uom            = d.get('UOM', ''),
                    batch_quantity = d.get('Batch_Quantity'),
                ))
            if to_create:
                BmrIssue.objects.bulk_create(to_create)
        except Exception:
            # non-fatal; page will still work with existing data
            pass

    # NEW: helper to load blocks & equipments from both sources
    def _get_block_and_equipment():
        """
        Combined:
        - Source 1: production_scheduler.equipment
        - Source 2: LocalEquipmentMaster
        Returns (block_list, equipment_list) where equipment_list is a list
        of dicts with keys: eqp_code, eqp_name, block_name
        """

        def normalize_block_name(block_value):
            val = (block_value or "").strip()
            if not val:
                return ""

            key = val.upper().replace("_", " ")
            key = " ".join(key.split())  # remove extra spaces

            mapping = {
                "BLOCK A": "A-Block",
                "A BLOCK": "A-Block",
                "A-BLOCK": "A-Block",

                "BLOCK B": "B-Block",
                "B BLOCK": "B-Block",
                "B-BLOCK": "B-Block",

                "BLOCK C": "C-Block",
                "C BLOCK": "C-Block",
                "C-BLOCK": "C-Block",

                "BLOCK D": "D-Block",
                "D BLOCK": "D-Block",
                "D-BLOCK": "D-Block",

                "BLOCK E": "E-Block",
                "E BLOCK": "E-Block",
                "E-BLOCK": "E-Block",

                "TANK FARM": "Tank Farm",
                "FIRE PUMP HOUSE": "Fire Pump House",
                "ELECTRICAL PCC ROOM": "ELECTRICAL PCC ROOM",
                "DG HOUSE": "DG House",
                "BOILER": "Boiler",
                "ETP": "ETP",
                "MEE": "MEE",
                "ETP/MEE": "ETP/MEE",
                "PILOT PLANT": "Pilot Plant",
                "ETP RO": "ETP RO",
                "HT YEARD": "HT Yeard",
            }

            return mapping.get(key, val)

        allowed_blocks = {
            "A-Block",
            "B-Block",
            "C-Block",
            "D-Block",
            "E-Block",
            "Tank Farm",
            "Fire Pump House",
            "ELECTRICAL PCC ROOM",
            "DG House",
            "Boiler",
            "ETP",
            "MEE",
            "ETP/MEE",
            "Pilot Plant",
            "ETP RO",
            "HT Yeard",
        }

        block_set = set()
        equipments = []
        seen_pairs = set()  # to avoid duplicates (eqp_code, block_name)

        # ---- Source 1: production_scheduler.equipment ----
        try:
            with connections["production_scheduler"].cursor() as cursor:
                cursor.execute(
                    """
                    SELECT block, eq_id
                    FROM equipment
                    WHERE block IS NOT NULL
                    AND LTRIM(RTRIM(block)) <> ''
                    """
                )
                rows = cursor.fetchall() or []

            for block_val, eq_id in rows:
                block_text = normalize_block_name(block_val)
                if not block_text:
                    continue

                if block_text not in allowed_blocks:
                    continue

                key = (str(eq_id), block_text)
                if key in seen_pairs:
                    continue
                seen_pairs.add(key)

                block_set.add(block_text)
                equipments.append({
                    "eqp_code": eq_id,
                    "eqp_name": eq_id,
                    "block_name": block_text,
                })

        except Exception as exc:
            logger.error(
                "Error loading equipment from production_scheduler.equipment for SFGQC: %s",
                exc,
                exc_info=True,
            )
            messages.warning(
                request,
                "Warning: could not load equipment from production scheduler."
            )

        # ---- Source 2: LocalEquipmentMaster ----
        try:
            qs = (
                LocalEquipmentMaster.objects
                .exclude(block_name__isnull=True)
                .exclude(block_name__exact='')
                .order_by('block_name', 'eqp_name', 'eqp_code')
            )

            for obj in qs:
                block_text = normalize_block_name(obj.block_name)
                if not block_text:
                    continue

                if block_text not in allowed_blocks:
                    continue

                eq_code = obj.eqp_code
                eq_name = obj.eqp_name or obj.eqp_code

                key = (str(eq_code), block_text)
                if key in seen_pairs:
                    continue
                seen_pairs.add(key)

                block_set.add(block_text)
                equipments.append({
                    "eqp_code": eq_code,
                    "eqp_name": eq_name,
                    "block_name": block_text,
                })

        except Exception as exc:
            logger.error(
                "Error loading LocalEquipmentMaster for SFGQC blocks/equipment: %s",
                exc,
                exc_info=True,
            )
            messages.warning(
                request,
                "Warning: could not load block/equipment list from Local Equipment master."
            )

        # Final sorted lists
        block_list = sorted(block_set)
        equipment_list = sorted(
            equipments,
            key=lambda e: (e["block_name"], str(e["eqp_code"]))
        )
        return block_list, equipment_list
    try:
        _maybe_sync_bmr_issue_from_erp()

        # dropdown sources (same as fgqc_create)
        try:
            distinct_batch_nos = (
                BmrIssue.objects
                .filter(bmr_issue_date__gte=date(2025, 1, 25))
                .values_list('op_batch_no', flat=True)
                .distinct()
                .order_by('op_batch_no')
            )
            selected_batch_no = request.GET.get('batch_no_filter') or None
            if selected_batch_no:
                qc_query_results = (
                    BmrIssue.objects
                    .filter(op_batch_no=selected_batch_no)
                    .order_by('bmr_issue_no', 'line_no')
                )
            else:
                qc_query_results = BmrIssue.objects.all().order_by(
                    'bmr_issue_no', 'line_no'
                )
        except Exception:
            distinct_batch_nos = []
            qc_query_results = BmrIssue.objects.none()
            selected_batch_no = None

        # NEW: get block_list + equipment_list from both sources
        block_list, equipment_list = _get_block_and_equipment()

        stage_options = (
            Product.objects
            .exclude(stages__isnull=True)
            .exclude(stages__exact='')
            .values('stages', 'id')
            .distinct()
            .order_by('stages')
            .annotate(item_type_clean=Trim('item_type'))
            .filter(item_type_clean__iexact='Semi Finished Good')
        )

        product_list = Product.objects.values('id', 'name').order_by('name')
        sample_description_options = list(
            SampleDescriptionOption.objects.order_by("name")
            .values_list("name", flat=True)
        )

        if request.method == "POST":
            form = SFGQCEntryForm(
                request.POST,
                sample_description_options=sample_description_options,
                stage_choices=stage_options,
            )
            if form.is_valid():
                with transaction.atomic():
                    qc_entry = form.save(commit=False)
                    qc_entry.status = 'pending_qc'
                    qc_entry.created_by = request.user

                    qc_entry.ar_type = 'SFG'
                    if not qc_entry.ar_no:
                        qc_entry.ar_no = _next_ar_no('SFG')

                    qc_entry.save()

                    send_mail(
                        subject=(
                            f"[SFGQC] New entry ready for QC: "
                            f"{qc_entry.product.name} / {qc_entry.batch_no}"
                        ),
                        message=(
                            f"A new SFGQC Entry (ID: {qc_entry.pk}) was submitted.\n"
                            f"Product: {qc_entry.product.name}\n"
                            f"Batch No: {qc_entry.batch_no}\n"
                            f"Stage: {qc_entry.stage}\n"
                            f"Quantity: {getattr(qc_entry, 'qty', 'N/A')}\n"
                            f"Sample Description: {qc_entry.sample_description or '—'}\n"
                            f"Test Required For: {qc_entry.test_required_for or '—'}\n"
                            f"Test Parameter: {qc_entry.test_parameters or '—'}\n"
                        ),
                        from_email=getattr(
                            settings,
                            'DEFAULT_FROM_EMAIL',
                            'workflow@ocspl.com',
                        ),
                        recipient_list=['shakir.s@ocspl.com',
                            'vinod.nimbalkar@ocspl.com',
                            'qc@ocspl.com',
                            'vikas.m@ocspl.com'],
                        fail_silently=True,
                    )

                messages.success(
                    request,
                    "SFGQC header submitted. QC team has been notified."
                )
                return redirect('qc:qc_list')
            else:
                messages.error(request, "Please correct the errors below.")
        else:
            now = timezone.localtime()
            form = SFGQCEntryForm(
                initial={
                    'entry_date': now,
                    'sample_sent_at': now,
                    'ar_type': 'SFG',
                    'ar_no': _next_ar_no('SFG'),
                },
                sample_description_options=sample_description_options,
                stage_choices=stage_options,
            )

        context = {
            "form":               form,
            "equipment_list":     equipment_list,   # combined list of dicts
            "block_list":         block_list,       # distinct block names
            "distinct_batch_nos": distinct_batch_nos,
            "selected_batch_no":  selected_batch_no,
            "stage_options":      stage_options,
            "product_list":       product_list,
            "qc_query_results":   qc_query_results,
        }
        return render(request, "qc/sfgqc_form.html", context)

    except Exception as e:
        logger.error(
            "Exception in sfgqc_create by %s: %s",
            request.user.username if request.user.is_authenticated else "UNKNOWN",
            e,
            exc_info=True,
        )
        messages.error(
            request,
            "An unexpected error occurred. Please contact the admin."
        )
        return redirect('qc:qc_list')





@login_required
def qc_update(request, pk):
    """ add QC details (Permission Required: QC.change_qcentry) """
    if not request.user.has_perm('QC.change_qcentry'):
        logger.warning(f"Unauthorized update attempt by {request.user.username}")
        messages.error(request, "You do not have permission to edit QC records.")
        return redirect('indexpage')
    
    try:
        qc_entry = get_object_or_404(QCEntry, pk=pk)
        reopen_to = request.GET.get("reopen")

        # --- Reopen actions ---
        if reopen_to == "prod" and qc_entry.status == "qc_completed":
            qc_entry.status = "draft"
            qc_entry.save()
            messages.success(request, "Entry reopened to Production.")
            return redirect("qc:qc_update", pk=pk)

        if reopen_to == "qc" and qc_entry.status == "qc_completed":
            qc_entry.status = "pending_qc"
            qc_entry.save()
            messages.success(request, "Entry reopened to QC.")
            return redirect("qc:qc_update", pk=pk)

        # --- Phase 1: Production header ---
        if qc_entry.status == "draft":
            equipment_list = LocalEquipmentMaster.objects.order_by("eqp_name")
            distinct_batch_nos = (BmrIssue.objects.filter(bmr_issue_date__gte=date(2025, 1, 25)).values_list("op_batch_no", flat=True).distinct().order_by("op_batch_no"))
            stage_options = (Product.objects.exclude(stages__isnull=True).exclude(stages__exact="").values("stages", "id").distinct().order_by("stages"))
            product_list = Product.objects.values("id", "name").order_by("name")
            sample_description_options = list(SampleDescriptionOption.objects.order_by("name").values_list("name", flat=True))
            
            if request.method == "POST":
                form = ProductionQCEntryForm(request.POST, instance=qc_entry, stage_choices=stage_options, sample_description_options=sample_description_options)
                if form.is_valid():
                    sd = form.cleaned_data.get("sample_description")
                    if sd and not SampleDescriptionOption.objects.filter(name=sd).exists():
                        SampleDescriptionOption.objects.create(name=sd)
                    qc = form.save(commit=False)
                    qc.status = "pending_qc"
                    qc.save()
                    messages.success(request, "Header updated and submitted to QC.")
                    return redirect("qc:qc_list")
                messages.error(request, "Please correct the errors below.")
            else:
                form = ProductionQCEntryForm(instance=qc_entry, initial={"entry_date": timezone.localtime().strftime("%Y-%m-%dT%H:%M")}, stage_choices=stage_options, sample_description_options=sample_description_options)
            
            return render(request, "qc/qc_form_phase1.html", {"form": form, "equipment_list": equipment_list, "distinct_batch_nos": distinct_batch_nos, "stage_options": stage_options, "product_list": product_list})

        # --- Phase 2: QC results (Fully Merged and Corrected) ---
        if qc_entry.status == "pending_qc":
            existing_spec_entries = SpecEntry.objects.filter(qc_entry=qc_entry).select_related("spec")
            group_options = (
                Spec.objects.filter(product=qc_entry.product)
                .values_list("group", flat=True)
                .distinct()
                .order_by("group")
            )
            group_options = [g for g in group_options if g]

            if request.method == "POST":
                # Correctly instantiate the form with group_options
                results_form = QCResultsForm(request.POST, instance=qc_entry, group_options=group_options)

                # --- PHASE 2 "SAVE DETAILS" BUTTON LOGIC ---
                if "save_phase2_header" in request.POST:
                    # Only validate the four header fields
                    fields_to_save = ["ar_type", "ar_no", "sample_received_at", "instrument_id"]
                    valid = True
                    # Set the values manually (from cleaned_data)
                    for field in fields_to_save:
                        val = results_form.data.get(field)
                        if val is not None:
                            setattr(qc_entry, field, val)
                        else:
                            valid = False

                    if valid:
                        qc_entry.save(update_fields=fields_to_save)
                        messages.success(request, "Header details saved.")
                    else:
                        messages.error(request, "Failed to save details. Please check the values.")

                    # Re-render the same page (stay in phase-2)
                    return render(
                        request,
                        "qc/qc_form_phase2.html",
                        {
                            "results_form": QCResultsForm(instance=qc_entry, group_options=group_options),
                            "qc_entry": qc_entry,
                            "group_options": group_options,
                            "existing_spec_entries": existing_spec_entries,
                            "appearance_options": AppearanceOption.objects.order_by("name"),
                        },
                    )

                # --- Full form submission logic for QC completion ---
                if results_form.is_valid():
                    with transaction.atomic():
                        qc = results_form.save(commit=False)
                        SpecEntry.objects.filter(qc_entry=qc).delete()
                        valid_spec_ids = set(Spec.objects.filter(product=qc.product, group=qc.group).values_list("id", flat=True))
                        
                        to_create = []
                        for key, raw in request.POST.items():
                            if not key.startswith("spec_result_"): continue
                            try: sid = int(key.split("_")[-1])
                            except ValueError: continue
                            if sid not in valid_spec_ids: continue
                            
                            spec_obj = Spec.objects.get(pk=sid)
                            val = raw.strip()
                            remark = ""
                            
                            # Handle "Appearance" even if misspelled (e.g., "Apperance") or differently cased.
                            name_lower = (spec_obj.name or "").lower()
                            is_appearance = "appear" in name_lower  # matches appearance/apperance/etc.

                            if is_appearance:
                                if val == "complies":
                                    remark = "Pass"
                                elif val == "does_not_comply":
                                    remark = "Fail"
                                # Persist selected appearance option on the product (if any)
                                if val:
                                    ao, _ = AppearanceOption.objects.get_or_create(name=val)
                                    qc.product.appearance_options.add(ao)

                            elif spec_obj.min_val is not None and spec_obj.max_val is not None:
                                try:
                                    n = float(val)
                                    minv, maxv, acv = float(spec_obj.min_val), float(spec_obj.max_val), spec_obj.acceptance_criteria
                                    if minv <= n <= maxv:
                                        remark = "Pass"
                                    elif acv is not None and (n < minv and n >= float(acv)):
                                        remark = "Approved under deviation"
                                    elif acv is not None and (n > maxv and n <= float(acv)):
                                        remark = "Approved under deviation"
                                    else:
                                        remark = "Fail"
                                except ValueError:
                                    remark = ""
                            
                            to_create.append(SpecEntry(qc_entry=qc, spec=spec_obj, value=val or None, remark=remark))
                        
                        if to_create:
                            SpecEntry.objects.bulk_create(to_create)

                        qc.status = "qc_completed"
                        qc.qc_completed_by = request.user
                        qc.save()
                        logger.info(f"QCEntry #{qc.pk} completed by {request.user.username} (QC results saved)")

                        # ——— PDF + Email Attachment Logic ———
                        try:
                            buffer = BytesIO()
                            styles = getSampleStyleSheet()
                            body_style = styles["BodyText"]
                            doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=40, rightMargin=40, topMargin=40, bottomMargin=40)
                            story = []

                            story.append(Paragraph("QC Entry Details", styles["Title"]))
                            story.append(Spacer(1, 12))
                            
                            hdr_data = [
                                ["Entry No.", f"#{qc.entry_no}"], ["Product", qc.product.name],
                                ["Stage", qc.get_stage_display()], ["Specification Group", qc.group or "—"],
                                ["General Remarks", Paragraph(qc.general_remarks or "—", body_style)],
                                ["Batch No.", qc.batch_no], 
                                ["AR No.", qc.ar_no],
                                ["Equipment", qc.equipment_id or "—"], 
                                ["Block", qc.block or "—"],
                                ["Instrument ID", qc.instrument_id or "—"],
                                ["Entry Date (Prod)", qc.entry_date.strftime("%Y-%m-%d %H:%M")],
                                ["Prepared By (Prod)", qc.created_by.get_full_name() if qc.created_by else "—"],
                                ["Sample Sent At", qc.sample_sent_at.strftime("%Y-%m-%d %H:%M") if qc.sample_sent_at else "—"],
                                ["Sample Description", qc.sample_description or "—"],
                                ["Frequency", qc.frequency or "—"],
                                ["Sample Received At QC", qc.sample_received_at.strftime("%Y-%m-%d %H:%M") if qc.sample_received_at else "—"],
                                ["Released by QC At", qc.release_by_qc_at.strftime("%Y-%m-%d %H:%M") if qc.release_by_qc_at else "—"],
                                ["Completed By (QC)", qc.qc_completed_by.get_full_name() if qc.qc_completed_by else "—"],
                                ["Test Required For", qc.test_required_for or "—"],
                                ["Test Parameter", qc.test_parameters or "—"],
                                ["Final QC Decision", qc.get_decision_status_display()],
                            ]
                            hdr_table = Table(hdr_data, colWidths=[150, 330], hAlign="LEFT")
                            hdr_table.setStyle(TableStyle([
                                ("VALIGN", (0, 0), (-1, -1), "TOP"), ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                                ("BACKGROUND", (0, 0), (0, -1), colors.lightgrey), ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                            ]))
                            story.append(hdr_table)
                            story.append(Spacer(1, 24))

                            # --------- MULTI-GROUP EXPORT STARTS HERE --------
                            groups_qs = (
                                SpecEntry.objects.filter(qc_entry=qc)
                                .values_list("spec__group", flat=True)
                                .distinct()
                                .order_by("spec__group")
                            )
                            groups = [g for g in groups_qs if g] or [""]

                            for group in groups:
                                story.append(Spacer(1, 16))
                                story.append(Paragraph(f"Test Parameter Results — Group: {group or '(No Group)'}", styles["Heading3"]))
                                story.append(Spacer(1, 8))
                                story.append(Paragraph(f"<b>Sample Description Notes:</b> {qc.sample_description_text or '—'}", body_style))
                                story.append(Spacer(1, 12))

                                tbl_data = [[Paragraph(f"<b>{h}</b>", styles["Heading4"]) for h in ["#", "Test Parameter", "Specification", "Acceptance", "Value", "Unit", "Remark"]]]
                                specs = (
                                    SpecEntry.objects
                                    .filter(qc_entry=qc, spec__group=group)
                                    .select_related("spec")
                                    .order_by("spec__id")
                                )
                                for idx, se in enumerate(specs, start=1):
                                    spec = se.spec
                                    spec_display = (
                                        f"{spec.min_val} – {spec.max_val}" if spec.min_val is not None else (spec.allowed_choices or "—")
                                    )
                                    acc_display = str(spec.acceptance_criteria) if spec.acceptance_criteria is not None else "—"
                                    tbl_data.append([
                                        Paragraph(str(idx), body_style), Paragraph(spec.name or "—", body_style),
                                        Paragraph(spec_display, body_style), Paragraph(acc_display, body_style),
                                        Paragraph(se.value or "—", body_style), Paragraph(spec.unit or "—", body_style),
                                        Paragraph(se.remark or "—", body_style),
                                    ])
                                tbl = Table(tbl_data, colWidths=[30, 130, 100, 80, 60, 60, 60], repeatRows=1)
                                tbl.setStyle(TableStyle([
                                    ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey), ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                                    ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                                ]))
                                story.append(tbl)

                            doc.build(story)
                            pdf_data = buffer.getvalue()
                            buffer.close()

                            email_body = (f"QCEntry (ID: {qc.pk}) completed on {timezone.localtime():%Y-%m-%d %H:%M}.\n"
                                          f"Product: {qc.product.name}\nBatch No: {qc.batch_no}\n"
                                          f"Final QC Decision: {qc.get_decision_status_display()}\n"
                                          f"Genaral Remark: {qc.general_remarks or 'None'}")
                            email = EmailMessage(
                                subject=f"[QC Completed] Entry ID {qc.pk}", body=email_body,
                                from_email=settings.DEFAULT_FROM_EMAIL, to=
                                [
                                'ganesh.t@ocspl.com',
                                'productionblock.a@ocspl.com',
                                'productionblock.b@ocspl.com',
                                'productionblock.c@ocspl.com',
                                'productionblock.d@ocspl.com',
                                'production.solapur@ocspl.com',
                                'vikas.m@ocspl.com'
                                ]
                            )
                            email.attach(f"QCEntry_{qc.pk}_details.pdf", pdf_data, "application/pdf")
                            email.send(fail_silently=False)

                        except Exception as e:
                            logger.error(f"Email failed for QCEntry #{qc.pk} after saving: {e}", exc_info=True)
                            messages.error(request, f"QC saved but email failed: {e}")

                    messages.success(request, "QC results saved. Production Team notified.")
                    return redirect("qc:qc_detail", pk=qc.pk)
                
                messages.error(request, "Please correct the errors below.")
            else:
                # Correctly instantiate the form for the initial GET request
                results_form = QCResultsForm(instance=qc_entry, group_options=group_options)

            return render(
                request,
                "qc/qc_form_phase2.html",
                {
                    "results_form": results_form,
                    "qc_entry": qc_entry,
                    "group_options": group_options,
                    "existing_spec_entries": existing_spec_entries,
                    "appearance_options": AppearanceOption.objects.order_by("name"),
                },
            )

        if qc_entry.status == "qc_completed":
            messages.error(request, "This QC entry is already completed. Use “Reopen” to edit.")
            return redirect("qc:qc_detail", pk=pk)

        raise Http404("Invalid status for editing")

    except Exception as e:
        logger.error(f"Exception in qc_update for pk={pk} by {request.user.username}: {e}", exc_info=True)
        messages.error(request, "An unexpected error occurred. Please contact the admin.")
        return redirect("qc:qc_list")



@login_required
def qc_delete(request, pk):
    """ add QC details (Permission Required: QC.delete_qcentry) """
    if not request.user.has_perm('QC.delete_qcentry'):
        logger.warning(f"Unauthorized delete attempt by {request.user.username}")
        messages.error(request, "You do not have permission to Delete QC records.")
        return redirect('indexpage')
   
    qc = get_object_or_404(QCEntry, pk=pk)
    if request.method == 'POST':
        logger.info(f"QCEntry #{qc.pk} deleted by {request.user.username}")
        qc.delete()
        messages.success(request, "QC entry deleted.")
        return redirect('qc:qc_list')
    else:
        logger.info(f"QCEntry #{qc.pk} delete confirmation page viewed by {request.user.username}")
    return render(request, 'qc/qc_confirm_delete.html', {'qc_entry': qc})


@login_required
def report(request):
    logger.info(f"User {request.user.username} accessed QC report page.")
    entries = (
        QCEntry.objects
        .select_related('product')
        .prefetch_related(Prefetch('values', queryset=SpecEntry.objects.select_related('spec'), to_attr='specs'))
        .order_by('-created')
    )
    return render(request, 'qc/report.html', {'entries': entries})


@require_GET
def get_product_details(request):
    product_id = request.GET.get("product")
    stage = request.GET.get("stage")

    logger.info(
        f"AJAX get_product_details called by {request.user.username if request.user.is_authenticated else 'Anonymous'} "
        f"for product_id={product_id}, stage={stage}"
    )

    if not product_id:
        logger.warning("No product_id supplied in get_product_details.")
        return JsonResponse({"header": {}, "specs": []}, status=200)

    try:
        pid = int(product_id)
        product = Product.objects.get(pk=pid)
    except (ValueError, Product.DoesNotExist):
        logger.error(f"Invalid product ID or product does not exist: {product_id}")
        return JsonResponse({"error": "Invalid product or ID."}, status=400)

    # Lookups for header
    detail = None
    if stage:
        detail = LocalBOMDetail.objects.filter(
            item_name=product.name, fg_name=stage
        ).first()

    header_data = {
        "default_fg_name": detail.fg_name if detail else "",
        "default_bom_code": detail.bom_code if detail else "",
        "default_type": detail.type if detail else "",
    }

    # Fetch the specs, including unit, spec_type, and allowed_choices
    specs_qs = Spec.objects.filter(product_id=pid).values(
        "id",
        "name",
        "min_val",
        "max_val",
        "group",
        "unit",
        "spec_type",
        "allowed_choices",
        "acceptance_criteria",
    )

    specs_list = []
    for s in specs_qs:
        specs_list.append(
            {
                "id": s["id"],
                "name": s["name"],
                "min_val": s["min_val"],
                "max_val": s["max_val"],
                "group": s.get("group") or "",
                "unit": s.get("unit") or "",
                "spec_type": s.get("spec_type") or "",
                "allowed_choices": s.get("allowed_choices") or "",
                "acceptance_criteria": s.get("acceptance_criteria"),
            }
        )

    logger.debug(
        f"Product details returned for product_id={product_id}: header={header_data}, specs_count={len(specs_list)}"
    )

    return JsonResponse({"header": header_data, "specs": specs_list}, status=200)



@require_GET
def get_specs(request):
    """
    Returns a list of specs for the given product, including:
      - id, name, min_val, max_val
      - group
      - unit
      - specification (the raw choice-list text, e.g., for Appearance)
    """
    product_id = request.GET.get("product")
    logger.info(
        f"AJAX get_specs called by {request.user.username if request.user.is_authenticated else 'Anonymous'} "
        f"for product_id={product_id}"
    )

    if not product_id:
        logger.warning("No product_id supplied in get_specs.")
        return JsonResponse([], safe=False)

    try:
        pid = int(product_id)
    except ValueError:
        logger.error(f"Invalid product ID (not an integer): {product_id}")
        return JsonResponse({"error": "Invalid product ID."}, status=400)

    # Pull in the extra fields from Spec, including 'choices' for the specification
    specs_qs = Spec.objects.filter(product_id=pid).values(
        "id", "name", "min_val", "max_val", "group", "unit", "choices"
    )

    specs_list = []
    for s in specs_qs:
        specs_list.append(
            {
                "id": s["id"],
                "name": s["name"],
                "min_val": s["min_val"],
                "max_val": s["max_val"],
                "group": s.get("group") or "",
                "unit": s.get("unit") or "",
                # Rename the raw choice-list field to "specification"
                "specification": s.get("choices") or "",
            }
        )

    logger.debug(f"{len(specs_list)} specs returned for product_id={product_id}")
    return JsonResponse(specs_list, safe=False)

@login_required
def item_master(request):
    """
    Displays a paginated list of all items from the LocalItemMaster.
    """
    logger.info(f"User {request.user.username} accessed Item Master.")
    
    # Get all items, ordered by name
    item_list = LocalItemMaster.objects.all().order_by('product_name')
    # --- PAGINATION LOGIC ---
    # Create a Paginator instance with the item list, showing 50 items per page
    paginator = Paginator(item_list, 50) 
    # Get the page number from the URL's query parameters (e.g., ?page=2)
    page_number = request.GET.get('page')
    # Get the Page object for the requested page number.
    # .get_page() is used to safely handle invalid or empty page numbers.
    page_obj = paginator.get_page(page_number)
    # Pass the 'page_obj' to the template instead of the full 'item_list'
    return render(request, 'qc/item_master.html', {'page_obj': page_obj})


@login_required
def equipment_master(request):
    """
    Displays a paginated list of all equipment from the LocalEquipmentMaster.
    """
    logger.info(f"User {request.user.username} accessed Equipment Master.")
    # Get all equipment, ordered by name
    eq_list = LocalEquipmentMaster.objects.all().order_by('eqp_name')    
    # --- PAGINATION LOGIC ---
    # Create a Paginator instance with the equipment list, showing 50 items per page
    paginator = Paginator(eq_list, 50) 
    # Get the page number from the URL's query parameters (e.g., ?page=2)
    page_number = request.GET.get('page')
    # Get the Page object for the requested page number.
    # .get_page() is used to safely handle invalid or empty page numbers.
    page_obj = paginator.get_page(page_number)
    # Pass the 'page_obj' to the template instead of the full 'equipment_list'
    return render(request, 'qc/equipment_master.html', {'page_obj': page_obj})

@login_required
def qc_reopen_for_qc(request, pk):
    """ add QC details (Permission Required: QC.change_qcentry) """
    if not request.user.has_perm('QC.change_qcentry'):
        logger.warning(f"Unauthorized Add attempt by {request.user.username}")
        messages.error(request, "You do not have permission to reopen QC records.")
        return redirect('indexpage')
   
    logger.info(f"User {request.user.username} requested QC reopen for QCEntry {pk}.")
    qc = get_object_or_404(QCEntry, pk=pk)
    if qc.status != 'qc_completed':
        logger.warning(f"User {request.user.username} tried to reopen QCEntry {pk} for QC correction, but status is '{qc.status}'.")
        messages.error(request, "Only a completed entry can be reopened.")
        return redirect('qc:qc_detail', pk=pk)
    SpecEntry.objects.filter(qc_entry=qc).delete()
    qc.status = 'pending_qc'
    qc.qc_completed_by = None
    qc.release_by_qc_at = None
    qc.save()
    logger.info(f"QCEntry {pk} reopened for QC correction by {request.user.username}.")
    messages.info(request, f"QC entry #{qc.pk} reopened for QC correction.")
    return redirect('qc:qc_update', pk=pk)


@login_required
def qc_reopen_for_prod(request, pk):
    """ add QC details (Permission Required: QC.change_qcentry) """
    if not request.user.has_perm('QC.change_qcentry'):
        logger.warning(f"Unauthorized Add attempt by {request.user.username}")
        messages.error(request, "You do not have permission to Reopen QC records.")
        return redirect('indexpage')
   
    logger.info(f"User {request.user.username} requested Production reopen for QCEntry {pk}.")
    qc = get_object_or_404(QCEntry, pk=pk)
    if qc.status not in ('pending_qc','qc_completed'):
        logger.warning(f"User {request.user.username} tried to reopen QCEntry {pk} for production, but status is '{qc.status}'.")
        messages.error(request, "Only an entry in QC or completed can be reopened for Production.")
        return redirect('qc:qc_detail', pk=pk)
    SpecEntry.objects.filter(qc_entry=qc).delete()
    qc.status = 'draft'
    qc.qc_completed_by = None
    qc.release_by_qc_at = None
    qc.save()
    logger.info(f"QCEntry {pk} reopened for production by {request.user.username}.")
    messages.info(request, f"QC entry #{qc.pk} reopened for Production.")
    return redirect('qc:qc_update', pk=pk)


@login_required
def qc_cancel(request, pk):
    """ add QC details (Permission Required: QC.delete_qcentry) """
    if not request.user.has_perm('QC.delete_qcentry'):
        logger.warning(f"Unauthorized Cancel attempt by {request.user.username}")
        messages.error(request, "You do not have permission to Cancel QC records.")
        return redirect('indexpage')
   
    logger.info(f"User {request.user.username} attempted to cancel QCEntry {pk}.")
    qc = get_object_or_404(QCEntry, pk=pk)
    if qc.status == 'qc_completed':
        logger.warning(f"User {request.user.username} tried to cancel already completed QCEntry {pk}.")
        messages.error(request, "Cannot cancel: QC is already completed.")
    elif qc.status != 'cancelled':
        qc.status = 'cancelled'
        qc.save()
        logger.info(f"QCEntry {pk} cancelled by {request.user.username}.")
        messages.success(request, f"QC entry #{pk} cancelled.")
    return redirect('qc:qc_list')



@login_required
def mis_report(request):
    """
    MIS report: one row per QCEntry, with dynamic columns for each SpecEntry.
    Supports filtering, sorting, pagination, and CSV export.
    """
    try:
        # --- Get filter, sort, and pagination parameters from request ---
        sel_start_date = request.GET.get("start_date")
        sel_end_date = request.GET.get("end_date")
        sel_product = request.GET.get("product")
        sel_batch = request.GET.get("batch_no")
        sel_stage = request.GET.get("stage")
        sel_group = request.GET.get("group")
        sel_decision = request.GET.get("decision")
        sel_status = request.GET.get("status")

        sel_sort_by = request.GET.get("sort_by", "date")
        sel_sort_dir = request.GET.get("sort_dir", "desc")
        
        # Get the page number from the URL, e.g., ?page=2
        page_number = request.GET.get("page")

        want_export = request.GET.get("export") == "csv"

        logger.info(
            f"[MIS Report] User={request.user.username} Filters: "
            f"start_date={sel_start_date}, end_date={sel_end_date}, product={sel_product}, "
            f"batch={sel_batch}, stage={sel_stage}, group={sel_group}, decision={sel_decision}, "
            f"status={sel_status}, sort_by={sel_sort_by}, sort_dir={sel_sort_dir}, export={want_export}"
        )

        # --- Create a query string of active filters for template links ---
        active_filters = {
            "start_date": sel_start_date, "end_date": sel_end_date, "product": sel_product,
            "batch_no": sel_batch, "stage": sel_stage, "group": sel_group,
            "decision": sel_decision, "status": sel_status,
        }
        filter_query_string = urllib.parse.urlencode({k: v for k, v in active_filters.items() if v})

        # --- Build and filter the base queryset ---
        qs = QCEntry.objects.all()
        if sel_start_date: qs = qs.filter(entry_date__date__gte=sel_start_date)
        if sel_end_date: qs = qs.filter(entry_date__date__lte=sel_end_date)
        if sel_product: qs = qs.filter(product_id=sel_product)
        if sel_batch: qs = qs.filter(batch_no=sel_batch)
        if sel_stage: qs = qs.filter(stage=sel_stage)
        if sel_group: qs = qs.filter(group=sel_group)
        if sel_decision: qs = qs.filter(decision_status=sel_decision)
        if sel_status: qs = qs.filter(status=sel_status)

        # --- Apply sorting ---
        sort_map = {"date": "entry_date", "product_name": "product__name", "batch_no": "batch_no"}
        sort_field = sort_map.get(sel_sort_by, "entry_date")
        if sel_sort_dir == "desc":
            sort_field = f"-{sort_field}"

        ordered_qs = qs.select_related("product").prefetch_related(
            Prefetch("values", queryset=SpecEntry.objects.select_related("spec"), to_attr="specs")
        ).order_by(sort_field)

        # --- Discover all spec names ---
        spec_names = []
        seen = set()
        for entry in ordered_qs:
            for se in entry.specs:
                n = se.spec.name
                if n not in seen:
                    seen.add(n)
                    spec_names.append(n)

        # --- Assemble row data ---
        rows = []
        for entry in ordered_qs:
            row = {
                "date": entry.entry_date.strftime("%Y-%m-%d"),
                "product_name": entry.product.name if entry.product else "—",
                "stage": entry.get_stage_display(),
                "group": entry.group or "—",
                "batch_no": entry.batch_no or "—",
                "reactor_no": entry.equipment_id or "—",
                "sample_description": entry.sample_description or "—",
                "decision_status": entry.get_decision_status_display(),
                "status": entry.get_status_display(),
            }
            spec_values = {se.spec.name: se.value for se in entry.specs}
            row.update(spec_values)
            rows.append(row)

        # --- Handle CSV export ---
        if want_export:
            filename = f"MIS_Report_{datetime.now():%Y%m%d_%H%M}.csv"
            logger.info(
                f"[MIS Report] User={request.user.username} triggered CSV export as {filename} (records={len(rows)})."
            )
            resp = HttpResponse(content_type="text/csv")
            resp["Content-Disposition"] = f'attachment; filename="{filename}"'
            writer = csv.writer(resp)

            # Write header block
            writer.writerow(["MIS Report"])
            product_name = (
                Product.objects.filter(pk=sel_product).first().name
                if sel_product
                else "All"
            )
            writer.writerow(["Product:", product_name, "Batch:", sel_batch or "All"])
            writer.writerow(["Stage:", sel_stage or "All", "Group:", sel_group or "All"])
            writer.writerow(
                ["Decision:", sel_decision or "All", "Status:", sel_status or "All"]
            )
            writer.writerow(
                ["Date Range:", sel_start_date or "–", "to", sel_end_date or "–"]
            )
            writer.writerow(["Generated at:", datetime.now().strftime("%Y-%m-%d %H:%M")])
            writer.writerow([])

            # Write column headings
            headers = [
                "Date",
                "Product",
                "Stage",
                "Group",
                "Batch No.",
                "Reactor No.",
                "Final Decision",
                "Status",
                "Sample Description",
            ] + spec_names
            writer.writerow(headers)

            # Write data rows
            for r in rows:
                vals = [
                    r.get("date"),
                    r.get("product_name"),
                    r.get("stage"),
                    r.get("group"),
                    r.get("batch_no"),
                    r.get("reactor_no"),
                    r.get("decision_status"),
                    r.get("status"),
                    r.get("sample_description"),
                ] + [r.get(n, "") for n in spec_names]
                writer.writerow(vals)

            logger.info(
                f"[MIS Report] User={request.user.username} CSV export completed successfully."
            )
            return resp

        # --- ADD PAGINATION LOGIC ---
        paginator = Paginator(rows, 50)  # Show 50 rows per page
        page_obj = paginator.get_page(page_number) # Safely get the page object

        # --- Get options for filter dropdowns ---
        all_entries = QCEntry.objects.all()
        products = all_entries.values("product__id", "product__name").distinct().order_by("product__name")
        batches = all_entries.filter(batch_no__isnull=False).values_list("batch_no", flat=True).distinct().order_by("batch_no")
        stages = all_entries.filter(stage__isnull=False).values_list("stage", "stage").distinct().order_by("stage")
        groups = all_entries.filter(group__isnull=False).values_list("group", flat=True).distinct().order_by("group")

        logger.info(f"[MIS Report] User={request.user.username} viewed report page (records={len(rows)}).")
        
        # --- Render HTML page ---
        context = {
            "page_obj": page_obj, # Pass the page object instead of 'rows'
            "spec_names": spec_names,
            "products": products,
            "batches": batches,
            "stages": stages,
            "groups": groups,
            "decision_choices": QCEntry.DECISION_CHOICES,
            "status_choices": QCEntry.STATUS_CHOICES,
            "sel_start_date": sel_start_date, "sel_end_date": sel_end_date, "sel_product": sel_product,
            "sel_batch": sel_batch, "sel_stage": sel_stage, "sel_group": sel_group,
            "sel_decision": sel_decision, "sel_status": sel_status,
            "filter_query_string": filter_query_string,
            "sel_sort_by": sel_sort_by, "sel_sort_dir": sel_sort_dir,
        }
        return render(request, "qc/mis_report.html", context)

    except Exception as e:
        logger.error(f"[MIS Report] User={request.user.username} error: {e}", exc_info=True)
        messages.error(request, "An unexpected error occurred while generating the report.")
        return render(request, "qc/mis_report.html", {})








@login_required
def generate_ar_no(request):
    ar_type = request.GET.get("ar_type")
    if not ar_type:
        return JsonResponse({"error": "Missing AR type"}, status=400)

    # Generate financial year string like '25-26'
    today = timezone.now()
    fy_start = today.year if today.month >= 4 else today.year - 1
    fy_str = f"{str(fy_start)[-2:]}-{str(fy_start + 1)[-2:]}"
    prefix = f"QC/{ar_type}/{fy_str}/"

    last_ar = (
        QCEntry.objects
        .filter(ar_no__startswith=prefix)
        .order_by('-ar_no')
        .first()
    )

    last_seq = int(last_ar.ar_no.split('/')[-1]) if last_ar and last_ar.ar_no else 0
    new_ar_no = f"{prefix}{last_seq + 1:05d}"

    return JsonResponse({"ar_no": new_ar_no})




# ------------------------------------------------------------------

SYNC_CACHE_KEY = "qc_sync_job:{job_id}"

def _cache_set(job_id, data: dict):
    cache.set(SYNC_CACHE_KEY.format(job_id=job_id), data, timeout=60*60)  # 1 hour

def _cache_get(job_id):
    return cache.get(SYNC_CACHE_KEY.format(job_id=job_id))

def _run_sync_job(job_id, user_id):
    buf = io.StringIO()
    started = timezone.now()
    _cache_set(job_id, {
        "status": "running",
        "started_at": started.isoformat(),
        "log": "Starting ERP sync...\n",
        "user_id": user_id,
    })
    try:
        # 1) ERP sync
        call_command("sync_erp", stdout=buf, stderr=buf)
        buf.write("\n---\n")

        # 2) BMR master sync
        call_command("sync_bmr_master", stdout=buf, stderr=buf)

        log_text = buf.getvalue()
        _cache_set(job_id, {
            "status": "done",
            "started_at": started.isoformat(),
            "finished_at": timezone.now().isoformat(),
            "log": log_text,
            "user_id": user_id,
        })
    except Exception as exc:
        tb = traceback.format_exc()
        prev = buf.getvalue()
        _cache_set(job_id, {
            "status": "error",
            "started_at": started.isoformat(),
            "finished_at": timezone.now().isoformat(),
            "log": f"{prev}\nERROR: {exc}\n{tb}",
            "user_id": user_id,
        })

@login_required
def sync_erp_start(request):
    """
    Start the sync in a background thread and return a job_id immediately.
    """
    job_id = uuid.uuid4()
    _cache_set(job_id, {
        "status": "queued",
        "started_at": None,
        "log": "Queued...",
        "user_id": request.user.id,
    })
    t = threading.Thread(target=_run_sync_job, args=(job_id, request.user.id), daemon=True)
    t.start()
    return JsonResponse({"status": "queued", "job_id": str(job_id)})

@login_required
def sync_erp_status(request, job_id):
    """
    Poll current status/log for the job.
    """
    try:
        job_uuid = uuid.UUID(str(job_id))
    except Exception:
        raise Http404("Invalid job id")

    data = _cache_get(job_uuid)
    if not data:
        return JsonResponse({"status": "unknown"}, status=404)
    # (Optional) hide full logs until done to reduce payload
    tail = data["log"][-5000:] if data.get("log") else ""
    return JsonResponse({
        "status": data.get("status"),
        "started_at": data.get("started_at"),
        "finished_at": data.get("finished_at"),
        "log_tail": tail,
    })

# ------------------------------------------------------------------



def _norm_q(v: str) -> str:
    return (v or "").strip()

@login_required
def sample_description_options_view(request):
    """
    Single page:
      - Create (POST add)
      - Update (POST edit)
      - Delete (POST delete)
      - GET list with search + pagination
      - Shows Edit form when ?edit_id=<id>
    """
    query = _norm_q(request.GET.get("q", ""))
    edit_id = request.GET.get("edit_id")
    action = (request.POST.get("action") or "").strip().lower()

    # ----------------------------
    # POST actions: add / edit / delete
    # ----------------------------
    if request.method == "POST":
        # DELETE
        if action == "delete":
            obj_id = request.POST.get("id")
            if not obj_id:
                return HttpResponseBadRequest("Missing id")
            obj = get_object_or_404(SampleDescriptionOption, id=obj_id)
            obj.delete()
            messages.success(request, "Sample description deleted!")
            return redirect("qc:sample_description_options")

        # EDIT
        if action == "edit":
            obj_id = request.POST.get("id")
            obj = get_object_or_404(SampleDescriptionOption, id=obj_id)
            form = SampleDescriptionOptionForm(request.POST, instance=obj)
            if form.is_valid():
                form.save()
                messages.success(request, "Sample description updated!")
                return redirect(f"{request.path}?q={query}" if query else request.path)
            else:
                messages.error(request, "Please correct the error below.")
                edit_id = str(obj.id)  # keep edit mode open

        # ADD (default)
        else:
            form = SampleDescriptionOptionForm(request.POST)
            if form.is_valid():
                form.save()
                messages.success(request, "Sample description added!")
                return redirect("qc:sample_description_options")
            else:
                messages.error(request, "Please correct the error below.")
                # fallthrough to render add form with errors

    # ----------------------------
    # GET: build add form and optional edit form
    # ----------------------------
    add_form = SampleDescriptionOptionForm()

    edit_obj = None
    edit_form = None
    if edit_id:
        try:
            edit_obj = SampleDescriptionOption.objects.only("id", "name").get(id=edit_id)
            edit_form = SampleDescriptionOptionForm(instance=edit_obj)
        except SampleDescriptionOption.DoesNotExist:
            edit_obj = None
            edit_form = None

    # If POST edit failed, keep edit form with errors
    if request.method == "POST" and action == "edit":
        edit_form = locals().get("form", edit_form)
        edit_obj = edit_obj or (edit_form.instance if edit_form else None)

    # If POST add failed, keep add form with errors
    if request.method == "POST" and (action not in ("edit", "delete")):
        add_form = locals().get("form", add_form)

    # ----------------------------
    # Queryset optimized
    # ----------------------------
    qs = SampleDescriptionOption.objects.all().only("id", "name", "created_at")

    if query:
        qs = qs.filter(name__icontains=query)

    qs = qs.order_by("name")

    paginator = Paginator(qs, 25)  # 25 per page
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(request, "qc/sample_description_options.html", {
        "form": add_form,
        "edit_form": edit_form,
        "edit_obj": edit_obj,
        "page_obj": page_obj,
        "query": query,
        "total_count": page_obj.paginator.count,
    })


@require_POST
@login_required
def add_sample_description(request):
    """
    AJAX endpoint for quick add.
    """
    name = _norm_q(request.POST.get("name", ""))

    if not name:
        return JsonResponse({"error": "Description name cannot be empty."}, status=400)

    if SampleDescriptionOption.objects.filter(name__iexact=name).exists():
        return JsonResponse({"error": f'The description "{name}" already exists.'}, status=409)

    try:
        new_option = SampleDescriptionOption.objects.create(name=name)
        return JsonResponse({"success": True, "id": new_option.id, "name": new_option.name}, status=201)
    except Exception as e:
        logger.exception("Error creating SampleDescriptionOption")
        return JsonResponse({"error": "An unexpected error occurred on the server."}, status=500)
   



# ────────────────────────────────────────────────────────────────────────────────
#  COA List
# ────────────────────────────────────────────────────────────────────────────────
@login_required
def coa_list(request):
    # Fetch all COAs, most recent first
    records = (
        COARecord.objects
        .select_related('qc_entry', 'created_by')
        .order_by('-created_at')   # ← corrected: use created_at
    )
    return render(request, 'qc/coa_list.html', {
        'records': records,
    })

# ────────────────────────────────────────────────────────────────────────────────
#  COA
# ────────────────────────────────────────────────────────────────────────────────
@login_required
def generate_coa(request, pk):
    entry = get_object_or_404(QCEntry, pk=pk)

    # Fetch spec entries explicitly (used for PDF and template)
    spec_entries = (
        SpecEntry.objects
                 .filter(qc_entry=entry)
                 .select_related("spec")
                 .order_by("spec__group", "spec__id")
    )

    if request.method == "POST":
        form = COAExtraForm(request.POST)
        if form.is_valid():
            # Extract header/extra fields with sensible fallbacks
            issue_date = form.cleaned_data.get("issue_date") or timezone.localdate()
            customer_name = form.cleaned_data.get("customer_name", "")
            quantity = form.cleaned_data.get("quantity", "")
            mfg_date = form.cleaned_data.get("mfg_date", "")
            retest_date = form.cleaned_data.get("retest_date", "")

            # Approval blocks from POST
            prepared_by = request.POST.get("prepared_by", "")
            prepared_date = request.POST.get("prepared_date", "")
            verified_by = request.POST.get("verified_by", "")
            verified_date = request.POST.get("verified_date", "")
            approved_by = request.POST.get("approved_by", "")
            approved_date = request.POST.get("approved_date", "")

            # Build a cleaned dictionary to persist
            clean_extra = {
                "issue_date": str(issue_date),
                "customer_name": customer_name,
                "quantity": str(quantity),
                "mfg_date": str(mfg_date),
                "retest_date": str(retest_date),
                "prepared_by": prepared_by,
                "prepared_date": prepared_date,
                "verified_by": verified_by,
                "verified_date": verified_date,
                "approved_by": approved_by,
                "approved_date": approved_date,
            }

            # --- Start PDF Generation ---
            buffer = io.BytesIO()
            styles = getSampleStyleSheet()

            # Custom styles
            title_style = ParagraphStyle("TitleCustom", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=18, alignment=1)
            label_style = ParagraphStyle("LabelBold", parent=styles["BodyText"], fontName="Helvetica-Bold", fontSize=9, wordWrap="CJK")
            body_style = ParagraphStyle("BodyWrap", parent=styles["BodyText"], fontName="Helvetica", fontSize=9, leading=12, wordWrap="CJK")
            small = ParagraphStyle("Small", parent=styles["BodyText"], fontSize=8)

            doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=40, rightMargin=40, topMargin=40, bottomMargin=40)
            story: list[Flowable] = []

            story.append(Paragraph("CERTIFICATE OF ANALYSIS", title_style))
            story.append(Paragraph("Certificate Code: F/QC/005/01", small))
            story.append(Spacer(1, 16))

            # Header summary table
            hdr_data = [
                [Paragraph("Issue Date Of Certificate", label_style), Paragraph(issue_date.strftime("%d.%m.%Y"), body_style)],
                [Paragraph("Product Name", label_style), Paragraph(entry.product.name or "—", body_style)],
                [Paragraph("Batch No.", label_style), Paragraph(entry.batch_no or "—", body_style)],
                [Paragraph("Quantity", label_style), Paragraph(str(quantity) if quantity else "—", body_style)],
                [Paragraph("Mfg. Date", label_style), Paragraph(mfg_date.strftime(" %b’%Y") if isinstance(mfg_date, date) else (mfg_date or "—"), body_style)],
                [Paragraph("Retest Date", label_style), Paragraph(retest_date.strftime(" %b’%Y") if isinstance(retest_date, date) else (retest_date or "—"), body_style)],
                [Paragraph("Customer Name", label_style), Paragraph(customer_name or "—", body_style)],
            ]
            hdr_table = Table(hdr_data, colWidths=[160, 320], hAlign="LEFT")
            hdr_table.setStyle(TableStyle([
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f1f5fa")),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.grey),
                ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("PADDING", (0, 0), (-1, -1), 5),
            ]))
            story.append(hdr_table)
            story.append(Spacer(1, 20))

            # Test Parameters grouped by spec.group
            if spec_entries.exists():
                grouped = {}
                for se in spec_entries:
                    grouped.setdefault(se.spec.group or "(No Group)", []).append(se)

                for group_name, ses in grouped.items():
                    story.append(Paragraph(f"Test Parameter Results — Group: {group_name}", styles["h4"]))
                    story.append(Spacer(1, 6))

                    # START: UPDATED PDF TABLE HEADER (Remark removed)
                    header_cells = [
                        Paragraph("<b>#</b>", label_style),
                        Paragraph("<b>Test Parameter</b>", label_style),
                        Paragraph("<b>Specification</b>", label_style),
                        Paragraph("<b>Value</b>", label_style),
                        Paragraph("<b>Unit</b>", label_style),
                    ]
                    # END: UPDATED PDF TABLE HEADER
                    
                    tbl_data = [header_cells]
                    for idx, se in enumerate(ses, start=1):
                        spec = se.spec
                        spec_display = "Qualitative"
                        if spec.spec_type == "numeric":
                            spec_display = f"{spec.min_val or '—'} – {spec.max_val or '—'}"
                        elif spec.spec_type == "choice":
                            spec_display = spec.allowed_choices or "—"

                        # START: UPDATED PDF TABLE ROW (Remark removed)
                        row = [
                            Paragraph(str(idx), body_style),
                            Paragraph(spec.name or "—", body_style),
                            Paragraph(spec_display, body_style),
                            Paragraph(se.value or "—", body_style),
                            Paragraph(spec.unit or "—", body_style),
                        ]
                        # END: UPDATED PDF TABLE ROW
                        tbl_data.append(row)

                    # START: UPDATED PDF COLUMN WIDTHS (5 columns)
                    col_widths = [30, 200, 170, 60, 50]
                    # END: UPDATED PDF COLUMN WIDTHS
                    
                    tbl = Table(tbl_data, colWidths=col_widths, repeatRows=1, hAlign="LEFT")
                    tbl.setStyle(TableStyle([
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f0f4fb")),
                        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("FONTSIZE", (0, 0), (-1, -1), 8),
                        ("PADDING", (0, 0), (-1, -1), 4),
                    ]))
                    story.append(tbl)
                    story.append(Spacer(1, 16))

            else:
                story.append(Paragraph("No test parameter results are available for this QC entry.", body_style))
                story.append(Spacer(1, 16))

            # ... (Rest of the PDF generation logic for footers is unchanged) ...
            # Footer approval blocks
            prepared_block = [
                [Paragraph("Prepared By", label_style), Paragraph("", body_style)],
                [Paragraph("Name", label_style), Paragraph(prepared_by or "—", body_style)],
                [Paragraph("Date", label_style), Paragraph(prepared_date or "—", body_style)],
            ]
            verified_block = [
                [Paragraph("Verified By", label_style), Paragraph("", body_style)],
                [Paragraph("Name", label_style), Paragraph(verified_by or "—", body_style)],
                [Paragraph("Date", label_style), Paragraph(verified_date or "—", body_style)],
            ]
            approved_block = [
                [Paragraph("Approved By", label_style), Paragraph("", body_style)],
                [Paragraph("Name", label_style), Paragraph(approved_by or "—", body_style)],
                [Paragraph("Date", label_style), Paragraph(approved_date or "—", body_style)],
            ]

            def make_person_table(block):
                t = Table(block, colWidths=[100, 140])
                t.setStyle(TableStyle([
                    ("BOX", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("BACKGROUND", (0, 0), (0, 0), colors.lightgrey),
                    ("FONTNAME", (0, 0), (0, 0), "Helvetica-Bold"),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ]))
                return t

            pb = make_person_table(prepared_block)
            vb = make_person_table(verified_block)
            ab = make_person_table(approved_block)

            # Combine three side-by-side
            footer_row = Table([[pb, vb, ab]], colWidths=[180, 180, 180])
            footer_row.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
            story.append(Spacer(1, 24))
            story.append(footer_row)
            story.append(Spacer(1, 24))

            # Signature block and document code
            sig_data = [
                [Paragraph("Signature (Stamp / Sign)", label_style), Paragraph("", body_style)],
                [Paragraph("Document Code", label_style), Paragraph("F/QC/005/01", body_style)],
            ]
            sig_table = Table(sig_data, colWidths=[220, 220])
            sig_table.setStyle(TableStyle([
                ("BOX", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTNAME", (0, 1), (0, 1), "Helvetica-Bold"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ]))
            story.append(sig_table)

            # Build and finalize PDF
            doc.build(story)
            pdf_data = buffer.getvalue()
            buffer.close()

            # Persist the COARecord
            coa = COARecord.objects.create(qc_entry=entry, created_by=request.user, extra_data=clean_extra)
            coa.pdf.save(f"COA_QC{entry.pk}_{coa.pk}.pdf", io.BytesIO(pdf_data))

            messages.success(request, "COA generated successfully.")
            return redirect("qc:coa_list") # Using namespace
        else:
            messages.error(request, "Please correct the errors below on the COA form.")
            
    else:
        form = COAExtraForm()

    return render(request, "qc/coa_generate_form.html", {
        "form": form,
        "entry": entry,
        "spec_entries": spec_entries,
    })




@login_required
def qc_test_parameter_view(request):
    """
    Handles CRUD for QCTestParameter objects.
    """
    if request.method == "POST":
        form = QCTestParameterForm(request.POST) # Use updated form
        if form.is_valid():
            form.save()
            messages.success(request, "QC Test Parameter added successfully!")
            return redirect('qc:qc_test_parameter_list') # Redirect to the correct URL name
        else:
            messages.error(request, "Please correct the error below.")
    else:
        form = QCTestParameterForm()

    query = request.GET.get('q', '')
    all_parameters = QCTestParameter.objects.all() # Use updated model
    if query:
        all_parameters = all_parameters.filter(name__icontains=query)

    paginator = Paginator(all_parameters, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        "form": form,
        "page_obj": page_obj,
        "query": query,
    }
    return render(request, "qc/qc_test_parameter_list.html", context)



import re
import math
import datetime as _dt
from decimal import Decimal
from collections import defaultdict, OrderedDict

from django.contrib.auth.decorators import login_required
from django.db.models import Q, Count, QuerySet, Model
from django.db.models.functions import TruncMonth
from django.shortcuts import render

from .models import QCEntry, SpecEntry

# Robust numeric extractor (handles NMT, ≤, <, units, comma decimals)
NUM_RE = re.compile(r'[-+]?(?:\d+(?:[.,]\d*)?|\.\d+)')

def _to_float_or_none(v):
    if v is None:
        return None
    if isinstance(v, (int, float, Decimal)):
        try:
            f = float(v)
            return f if math.isfinite(f) else None
        except Exception:
            return None
    s = str(v).strip().lower()
    s = (
        s.replace("nmt", " ")
         .replace("not more than", " ")
         .replace("not more then", " ")
         .replace("≤", " ")
         .replace("≥", " ")
         .replace("<", " ")
         .replace(">", " ")
         .replace("~", " ")
    )
    m = NUM_RE.search(s.replace(",", "."))
    if not m:
        return None
    try:
        f = float(m.group(0))
        return f if math.isfinite(f) else None
    except ValueError:
        return None

def _spec_counts(qs):
    q_pass = Q(remark__iexact='Pass') | Q(value__iexact='complies')
    q_dev  = Q(remark__icontains='deviation')
    q_fail = Q(remark__iexact='Fail') | Q(remark__iexact='Rejected') | Q(value__iexact='does_not_comply')
    return {
        "total":     qs.count(),
        "approved":  qs.filter(q_pass).count(),
        "variation": qs.filter(q_dev).count(),
        "rejected":  qs.filter(q_fail).count(),
    }

# --- JSON safety helpers -------------------------------------------------------
def _json_coerce_val(v):
    if v is None or isinstance(v, (int, float, bool, str)):
        return v
    if isinstance(v, Decimal):
        try:
            f = float(v)
            if f == float("inf") or f == float("-inf") or f != f:
                return str(v)
            return f
        except Exception:
            return str(v)
    if isinstance(v, (_dt.date, _dt.datetime)):
        return v.isoformat()
    if isinstance(v, QuerySet):
        try:
            return list(v.values())
        except Exception:
            return [str(x) for x in v]
    if isinstance(v, Model):
        return str(v)
    return str(v)

def _json_coerce_dict(d):
    out = {}
    for k, v in d.items():
        if isinstance(v, (list, tuple)):
            out[k] = [
                _json_coerce_dict(item) if isinstance(item, dict)
                else [_json_coerce_val(sub) for sub in item] if isinstance(item, (list, tuple))
                else _json_coerce_val(item)
                for item in v
            ]
        elif isinstance(v, dict):
            out[k] = _json_coerce_dict(v)
        else:
            out[k] = _json_coerce_val(v)
    return out

# --- Querystring normalization -------------------------------------------------
def _qs_val(request, key, default=None):
    """
    Normalize GET value: treat All/empty/null/undefined as not set.
    """
    v = (request.GET.get(key) or "").strip()   # <-- use .strip(), not .trim()
    if not v:
        return default
    if v.lower() in {"all", "none", "null", "undefined"}:
        return default
    return v

# --- Keep only latest row per (batch_no, spec_name) for trend ------------------
def _deduplicate_trend_data(trend_rows):
    """
    Keep the most recent row for each (batch_no, spec_name).
    We normalize timestamps to UTC so sorting never mixes naive/aware datetimes.
    """
    if not trend_rows:
        return []

    def _to_utc(ts):
        if not ts:
            return None
        if isinstance(ts, _dt.datetime):
            # if naive, assume UTC; if aware, convert to UTC
            if ts.tzinfo is None or ts.tzinfo.utcoffset(ts) is None:
                return ts.replace(tzinfo=_dt.timezone.utc)
            return ts.astimezone(_dt.timezone.utc)
        return None

    min_datetime = _dt.datetime.min.replace(tzinfo=_dt.timezone.utc)

    # Sort newest → oldest using normalized UTC timestamps
    sorted_rows = sorted(
        trend_rows,
        key=lambda r: _to_utc(r.get("release_timestamp")) or min_datetime,
        reverse=True
    )

    latest_by_key = {}
    for row in sorted_rows:
        key = (row.get("batch_no"), row.get("spec_name"))
        if not all(key):
            continue
        if key not in latest_by_key:
            latest_by_key[key] = row

    # Now return oldest → newest for chart X-axis
    final_list = list(latest_by_key.values())
    final_list.sort(key=lambda r: _to_utc(r.get("release_timestamp")) or min_datetime)
    return final_list

# ------------------------------------------------------------------------------    

@login_required
def qc_dashboard(request):
    logger.info("User=%s accessed QC Dashboard", request.user.username)
    # 1) Filters
    item            = _qs_val(request, "item")
    group_name      = _qs_val(request, "group")
    test_name       = _qs_val(request, "test_name")
    ar_no           = _qs_val(request, "ar_no")
    stage           = _qs_val(request, "stage")
    batch_no        = _qs_val(request, "batch_no")
    month           = _qs_val(request, "month")
    criticality     = _qs_val(request, "criticality", "all")
    decision_status = _qs_val(request, "decision_status", "total")
    if decision_status not in {"total","approved","approved_under_deviation","rejected","fail","nodecision"}:
        decision_status = "total"

    logger.info(
        "[QC][Dashboard][Filters] item=%s group=%s test=%s ar=%s stage=%s batch=%s month=%s criticality=%s status=%s",
        item, group_name, test_name, ar_no, stage, batch_no, month, criticality, decision_status
    )
    q_is_critical = Q(is_critical=True) | Q(spec__is_critical=True)
    # 2) Base QCEntry queryset
    entries_base = QCEntry.objects.select_related("product")
    if item:
        entries_base = entries_base.filter(product__name__iexact=item)
    if group_name:
        ids_for_group = (SpecEntry.objects
                         .filter(spec__group__iexact=group_name)
                         .values_list("qc_entry_id", flat=True).distinct())
        entries_base = entries_base.filter(id__in=ids_for_group)
    if test_name:
        ids_for_test = (SpecEntry.objects
                        .filter(spec__name__iexact=test_name)
                        .values_list("qc_entry_id", flat=True).distinct())
        entries_base = entries_base.filter(id__in=ids_for_test)
    if ar_no:
        entries_base = entries_base.filter(ar_no__iexact=ar_no)
    if stage:
        entries_base = entries_base.filter(stage__iexact=stage)
    if batch_no:
        entries_base = entries_base.filter(batch_no__iexact=batch_no)
    if month:
        try:
            y, m = month.split("-")
            entries_base = entries_base.filter(entry_date__year=int(y), entry_date__month=int(m))
        except (ValueError, TypeError):
            pass

    # 3) Criticality gate
    if criticality == "critical":
        entry_ids = (SpecEntry.objects.filter(q_is_critical)
                     .values_list("qc_entry_id", flat=True).distinct())
        entries = entries_base.filter(id__in=entry_ids)
        logger.info("[QC][Dashboard] critical filter applied, entries=%d", entries.count())
    elif criticality == "non_critical":
        noncrit_ids = (SpecEntry.objects.exclude(q_is_critical)
                       .values_list("qc_entry_id", flat=True).distinct())
        entries = entries_base.filter(id__in=noncrit_ids)
        logger.info("[QC][Dashboard] non_critical filter applied, entries=%d", entries.count())
    else:
        entries = entries_base
    # 4) Mapping for dependent dropdowns
    raw_map = {}
    for se in SpecEntry.objects.select_related("qc_entry__product", "spec"):
        ce = se.qc_entry
        itm = (ce.product.name or "").strip()
        d = raw_map.setdefault(itm, {
            "tests": set(), "ars": set(), "stages": set(), "batches": set(),
            "groups": set(), "tests_by_group": defaultdict(set)
        })
        if se.spec and se.spec.name:
            d["tests"].add(se.spec.name.strip())
        if se.spec and getattr(se.spec, "group", None):
            gname = (se.spec.group or "").strip()
            if gname:
                d["groups"].add(gname)
                d["tests_by_group"][gname].add(se.spec.name.strip())
        if ce.ar_no:    d["ars"].add(str(ce.ar_no).strip())
        if ce.stage:    d["stages"].add(str(ce.stage).strip())
        if ce.batch_no: d["batches"].add(str(ce.batch_no).strip())

    for qe in QCEntry.objects.select_related("product"):
        itm = ((qe.product.name if qe.product else "") or "").strip()
        d = raw_map.setdefault(itm, {
            "tests": set(), "ars": set(), "stages": set(), "batches": set(),
            "groups": set(), "tests_by_group": defaultdict(set)
        })
        if qe.ar_no:    d["ars"].add(str(qe.ar_no).strip())
        if qe.stage:    d["stages"].add(str(qe.stage).strip())
        if qe.batch_no: d["batches"].add(str(qe.batch_no).strip())

    mapping = {
        itm: {
            "tests":  sorted(list(v["tests"])),
            "ars":    sorted(list(v["ars"])),
            "stages": sorted(list(v["stages"])),
            "batches":sorted(list(v["batches"])),
            "groups": sorted(list(v["groups"])),
            "tests_by_group": {g: sorted(list(ts)) for g, ts in v["tests_by_group"].items()}
        } for itm, v in raw_map.items()
    }
    all_stages = sorted({s for vals in mapping.values() for s in vals["stages"]})
    stages = mapping.get(item, {}).get("stages", []) if item else all_stages
    all_groups = sorted({g for vals in mapping.values() for g in vals.get("groups", [])})
    groups = mapping.get(item, {}).get("groups", []) if item else all_groups
    pending_q = Q(decision_status__isnull=True) | Q(decision_status="")

    # 5) KPI cards
    summary = {
        "total":     entries.count(),
        "approved":  entries.filter(decision_status="approved").count(),
        "variation": entries.filter(decision_status="approved_under_deviation").count(),
        "rejected":  entries.filter(decision_status="rejected").count(),
        "fail":      entries.filter(decision_status="fail").count(),
        "pending":   entries.filter(pending_q).count(),
        "distinct_batches": (entries.exclude(batch_no__isnull=True).exclude(batch_no="")
                   .values("batch_no").distinct().count()), }
    logger.info(
        "[QC][Dashboard][Summary] total=%d approved=%d variation=%d rejected=%d "
        "fail=%d pending=%d distinct_batches=%d",
        summary["total"], summary["approved"], summary["variation"],
        summary["rejected"], summary["fail"], summary["pending"],
        summary["distinct_batches"],
    )
    spec_rows_for_cards = (SpecEntry.objects
                           .filter(qc_entry__in=entries_base.values_list("id", flat=True))
                           .select_related("spec", "qc_entry"))
    critical_summary     = _spec_counts(spec_rows_for_cards.filter(q_is_critical))
    non_critical_summary = _spec_counts(spec_rows_for_cards.exclude(q_is_critical))
    if criticality in ("critical", "non_critical"):
        spec_subset = (spec_rows_for_cards.filter(q_is_critical)
                       if criticality == "critical"
                       else spec_rows_for_cards.exclude(q_is_critical))
        entry_ids_subset = list(spec_subset.values_list("qc_entry_id", flat=True).distinct())
        entries_subset = QCEntry.objects.filter(id__in=entry_ids_subset)
        summary = {
            "total":     entries_subset.count(),
            "approved":  entries_subset.filter(decision_status="approved").count(),
            "variation": entries_subset.filter(decision_status="approved_under_deviation").count(),
            "rejected":  entries_subset.filter(decision_status="rejected").count(),
            "fail":      entries_subset.filter(decision_status="fail").count(),
            "distinct_batches": (entries_subset.exclude(batch_no__isnull=True).exclude(batch_no="").values("batch_no")
                              .distinct().count()), }
        logger.info(
            "[QC][Dashboard][Summary][%s] total=%d approved=%d variation=%d rejected=%d fail=%d distinct_batches=%d",
            criticality, summary["total"], summary["approved"], summary["variation"], summary["rejected"], summary["fail"], summary["distinct_batches"] )
    # 6) Stage Status
    stage_status_qs = (
        entries.values('product__name', 'stage')
        .annotate(
            approved=Count('id', filter=Q(decision_status='approved')),
            variation=Count('id', filter=Q(decision_status='approved_under_deviation')),
            rejected=Count('id', filter=Q(decision_status='rejected')),
        )
        .order_by('product__name', 'stage')
    )
    stage_status = list(stage_status_qs)

    # 7) Detail entries (never filter a sliced queryset)
    pending_q = Q(decision_status__isnull=True) | Q(decision_status="")  # make sure this is defined ABOVE this block

    detail_base = entries

    if decision_status == "nodecision":
        # only NULL / blank decision_status
        detail_base = entries.filter(pending_q)
    elif decision_status != "total":
        # specific decided status
        detail_base = entries.filter(decision_status=decision_status)

    if item or ar_no:
        detail_ids = list(detail_base.order_by("-entry_date").values_list("id", flat=True))
    else:
        detail_ids = list(detail_base.order_by("-entry_date").values_list("id", flat=True)[:50])

    detail_entries = list(
        QCEntry.objects.select_related("product")
        .filter(id__in=detail_ids)
        .order_by("-entry_date")
    )
    # Build specs map once
    raw_specs_map = defaultdict(list)
    for se in (SpecEntry.objects
               .filter(qc_entry_id__in=detail_ids)
               .select_related("spec", "qc_entry")
               .order_by("qc_entry__entry_date", "spec__name")):
        raw_specs_map[se.qc_entry_id].append(se)
    for e in detail_entries:
        specs = raw_specs_map.get(e.id, [])
        if group_name:
            specs = [s for s in specs if getattr(s.spec, "group", None) and s.spec.group.strip().lower() == group_name.strip().lower()]
        if criticality == "critical":
            specs = [s for s in specs if getattr(s, "is_critical", False) or getattr(s.spec, "is_critical", False)]
        elif criticality == "non_critical":
            specs = [s for s in specs if not (getattr(s, "is_critical", False) or getattr(s.spec, "is_critical", False))]
        e.batch_number = e.batch_no
        e.specs = specs
    # 7b) Critical groups
    crit_groups, noncrit_groups = defaultdict(list), defaultdict(list)
    critical_alerts = {"aud": 0, "fail": 0}

    def classify_tag(remark, value):
        rr = (remark or "").strip().lower()
        vv = (str(value).strip().lower() if value is not None else "")
        if "deviation" in rr: return "aud"
        if rr in ("fail", "rejected") or vv == "does_not_comply": return "fail"
        if rr == "pass" or "approved" in rr or vv == "complies": return "pass"
        return "neutral"
    for e in detail_entries:
        for se in raw_specs_map.get(e.id, []):
            if group_name:
                g = (getattr(se.spec, "group", "") or "").strip().lower()
                if g != group_name.strip().lower():
                    continue
            is_crit = getattr(se, "is_critical", False) or getattr(se.spec, "is_critical", False)
            key = (e.product.name if e.product else "—")
            row = {"e": e, "se": se}
            if is_crit:
                tag = classify_tag(se.remark, se.value)
                if tag == "aud":  critical_alerts["aud"] += 1
                if tag == "fail": critical_alerts["fail"] += 1
                crit_groups[key].append(row)
            else:
                noncrit_groups[key].append(row)
    crit_groups    = OrderedDict(sorted(crit_groups.items(), key=lambda kv: (kv[0] or "").lower()))
    noncrit_groups = OrderedDict(sorted(noncrit_groups.items(), key=lambda kv: (kv[0] or "").lower()))
    has_critical_alert = (critical_alerts["aud"] + critical_alerts["fail"]) > 0
    # 7c) Latest Batch Snapshot
    q_inprocess = (
        Q(stage__icontains="ipqc") | Q(stage__icontains="ip qc") |
        Q(stage__icontains="in process") | Q(stage__icontains="in-process") |
        Q(stage__icontains="inprocess")
    )
    q_fgqc = (
        Q(stage__icontains="fg qc") | Q(stage__icontains="fgqc") |
        Q(stage__icontains="finished goods") | Q(stage__icontains="finished good")
    )
    snap_qs = (QCEntry.objects
               .select_related("product")
               .exclude(batch_no__isnull=True).exclude(batch_no=""))
    if item:
        snap_qs = snap_qs.filter(product__name__iexact=item)

    latest_per_product = {}
    for qe in snap_qs.order_by("product_id", "-entry_date", "-entry_no", "-id"):
        if qe.product_id not in latest_per_product:
            latest_per_product[qe.product_id] = qe
    latest_batch_snapshot = []
    for pid, last_entry in latest_per_product.items():
        last_batch = last_entry.batch_no
        product_name = last_entry.product.name if last_entry.product else "—"

        batch_entries = QCEntry.objects.filter(product_id=pid, batch_no=last_batch)

        prefer_qs = batch_entries.filter(q_inprocess).order_by("-entry_date", "-entry_no", "-id")
        stage_tag, stage_label = ("ip", "In-Process")
        if not prefer_qs.exists():
            prefer_qs = batch_entries.filter(q_fgqc).order_by("-entry_date", "-entry_no", "-id")
            stage_tag, stage_label = ("fg", "FGQC")
        if not prefer_qs.exists():
            prefer_qs = batch_entries.order_by("-entry_date", "-entry_no", "-id")
            stage_tag, stage_label = ("other", last_entry.stage or "Latest")
        se_qs = (SpecEntry.objects
                 .filter(qc_entry__in=prefer_qs.values_list("id", flat=True))
                 .select_related("spec", "qc_entry", "qc_entry__product")
                 .order_by("-qc_entry__entry_date", "-qc_entry__entry_no", "-id"))
        if group_name:
            se_qs = se_qs.filter(spec__group__iexact=group_name)

        se = se_qs.first()
        if not se:
            continue
        remark = (se.remark or "").strip().lower()
        valstr = _json_coerce_val(se.value)
        val_str_for_tag = (str(se.value).strip().lower() if se.value is not None else "")

        if remark == "pass" or val_str_for_tag == "complies":
            tag = "pass"
        elif "deviation" in remark:
            tag = "deviation"
        elif remark in ("fail", "rejected") or val_str_for_tag == "does_not_comply":
            tag = "fail"
        else:
            tag = "neutral"

        latest_batch_snapshot.append({
            "product":     product_name,
            "batch_no":    _json_coerce_val(last_batch),
            "stage_tag":   stage_tag,
            "stage_label": stage_label,
            "group":       _json_coerce_val(getattr(se.spec, "group", "") or ""),
            "test":        se.spec.name if se.spec else "—",
            "observed":    valstr,
            "unit":        _json_coerce_val(getattr(se.spec, "unit", "") or ""),
            "remark":      _json_coerce_val(se.remark),
            "result_tag":  tag,
            "ar_no":       _json_coerce_val(se.qc_entry.ar_no),
            "entry_no":    _json_coerce_val(se.qc_entry.entry_no),
            "entry_date":  se.qc_entry.entry_date.strftime("%d/%m/%Y") if se.qc_entry.entry_date else "",
        })
    # 8) Trend data
    filters_applied = any([
        item, group_name, test_name, ar_no, stage, batch_no,
        decision_status != "total", month, criticality != "all",
    ])
    spec_qs = SpecEntry.objects.select_related("qc_entry__product", "spec", "qc_entry")
    if item:       spec_qs = spec_qs.filter(qc_entry__product__name__iexact=item)
    if group_name: spec_qs = spec_qs.filter(spec__group__iexact=group_name)
    if test_name:  spec_qs = spec_qs.filter(spec__name__iexact=test_name)
    if ar_no:      spec_qs = spec_qs.filter(qc_entry__ar_no__iexact=ar_no)
    if stage:      spec_qs = spec_qs.filter(qc_entry__stage__iexact=stage)
    if batch_no:   spec_qs = spec_qs.filter(qc_entry__batch_no__iexact=batch_no)
    if decision_status != "total":
        if decision_status == "nodecision":
            spec_qs = spec_qs.filter(
                Q(qc_entry__decision_status__isnull=True) |
                Q(qc_entry__decision_status="")
            )
        else:
            spec_qs = spec_qs.filter(qc_entry__decision_status=decision_status)
    if criticality == "critical":
        spec_qs = spec_qs.filter(q_is_critical)
    elif criticality == "non_critical":
        spec_qs = spec_qs.exclude(q_is_critical)
    if month:
        try:
            y, m = month.split("-")
            spec_qs = spec_qs.filter(qc_entry__entry_date__year=int(y), qc_entry__entry_date__month=int(m))
        except (ValueError, TypeError):
            pass
    if filters_applied:
        spec_iter = spec_qs.order_by("qc_entry__release_by_qc_at", "qc_entry__id")
    else:
        spec_iter = list(spec_qs.order_by("-qc_entry__release_by_qc_at", "-id")[:1000])[::-1]

    def _collect_numeric_rows(qs_iter):
        out = []
        for se in qs_iter:
            obs = _to_float_or_none(se.value)
            if obs is None:
                continue
            e = se.qc_entry
            if not e.batch_no:
                continue
            out.append({
                "label":     _json_coerce_val(e.batch_no),
                "batch_no":  _json_coerce_val(e.batch_no),
                "ar_no":     _json_coerce_val(e.ar_no),
                "entry_no":  f"#{_json_coerce_val(e.entry_no)}",
                "min_val":   _to_float_or_none(getattr(se.spec, "min_val", None)),
                "max_val":   _to_float_or_none(getattr(se.spec, "max_val", None)),
                "observed":  obs,
                "release_timestamp": e.release_by_qc_at,
                "date":      e.release_by_qc_at.strftime("%Y-%m-%d") if e.release_by_qc_at else "",
                "spec_name": _json_coerce_val(se.spec.name if se.spec else "Unknown"),
                # --- ADD THIS LINE ---
                "decision_status": e.decision_status,
            })
        return out

    all_trend_data = _collect_numeric_rows(spec_iter)
    trend_data = _deduplicate_trend_data(all_trend_data)

    if not trend_data:
        fb_qs = (SpecEntry.objects.select_related("qc_entry__product", "spec", "qc_entry")
                 .exclude(value__isnull=True))
        if item: fb_qs = fb_qs.filter(qc_entry__product__name__iexact=item)
        if group_name: fb_qs = fb_qs.filter(spec__group__iexact=group_name)
        fallback_rows = _collect_numeric_rows(list(fb_qs.order_by("-qc_entry__release_by_qc_at", "-id")[:2000])[::-1])
        trend_data = _deduplicate_trend_data(fallback_rows)[-200:]
        logger.info("[QC][Dashboard][Trend][Fallback] rows=%d", len(trend_data))

    # 9) Dropdown lists
    def unique_normalized(seq):
        seen, out = set(), []
        for s in seq:
            if not s:
                continue
            c = str(s).strip()
            k = c.lower()
            if k in seen:
                continue
            seen.add(k)
            out.append(c)
        return out

    products = unique_normalized(QCEntry.objects.values_list("product__name", flat=True))
    months_qs = (QCEntry.objects.annotate(m=TruncMonth("entry_date"))
                 .values_list("m", flat=True).distinct().order_by("-m"))
    months = list(months_qs)
    # 10) Context
    context = {
        "mapping": _json_coerce_dict(mapping),
        "trend_data": trend_data,
        "summary": summary,
        "detail_entries": detail_entries,
        "products": products,
        "stages": stages,
        "groups": groups,
        "selected_item": item or "All",
        "selected_group": group_name or "All",
        "selected_test": test_name or "All",
        "selected_ar": ar_no or "All",
        "selected_stage": stage or "All",
        "selected_batch": batch_no or "All",
        "selected_month": month or "All",
        "selected_status": decision_status or "total",
        "criticality_supported": True,
        "selected_criticality": criticality or "all",
        "critical_summary": critical_summary,
        "non_critical_summary": non_critical_summary,
        "months": months,
        "stage_status": stage_status,
        "crit_groups": crit_groups,
        "noncrit_groups": noncrit_groups,
        "latest_batch_snapshot": latest_batch_snapshot,
        "critical_alerts": critical_alerts,
        "has_critical_alert": has_critical_alert,
    }
    logger.info("[QC][Dashboard][Render] user=%s summary_total=%d detail_entries=%d", request.user.username, summary.get("total", 0), len(detail_entries))
    return render(request, "qc/dashboard.html", context)



@login_required
def qc_dashboard_export_excel(request):
    import datetime
    """
    Export QC dashboard data (KPIs, trend data, and detailed spec rows)
    into an Excel workbook with an Observed vs Spec Limits chart and an
    Outcomes by Product stacked bar chart.
    """
    logger.info("User=%s requested QC Dashboard Excel export", request.user.username)

    if not request.user.has_perm("QC.view_qcentry"):
        messages.error(request, "You do not have permission to export QC dashboard data.")
        return redirect("indexpage")

    # ---------- 1) Read filters exactly like qc_dashboard ----------
    item            = _qs_val(request, "item")
    group_name      = _qs_val(request, "group")
    test_name       = _qs_val(request, "test_name")
    ar_no           = _qs_val(request, "ar_no")
    stage           = _qs_val(request, "stage")
    batch_no        = _qs_val(request, "batch_no")
    month           = _qs_val(request, "month")
    criticality     = _qs_val(request, "criticality", "all")
    decision_status = _qs_val(request, "decision_status", "total")
    if decision_status not in {"total", "approved", "approved_under_deviation",
                               "rejected", "fail", "nodecision"}:
        decision_status = "total"

    q_is_critical = Q(is_critical=True) | Q(spec__is_critical=True)

    # ---------- 2) Base QCEntry queryset (same as dashboard) ----------
    entries_base = QCEntry.objects.select_related("product")
    if item:
        entries_base = entries_base.filter(product__name__iexact=item)
    if group_name:
        ids_for_group = (
            SpecEntry.objects.filter(spec__group__iexact=group_name)
            .values_list("qc_entry_id", flat=True)
            .distinct()
        )
        entries_base = entries_base.filter(id__in=ids_for_group)
    if test_name:
        ids_for_test = (
            SpecEntry.objects.filter(spec__name__iexact=test_name)
            .values_list("qc_entry_id", flat=True)
            .distinct()
        )
        entries_base = entries_base.filter(id__in=ids_for_test)
    if ar_no:
        entries_base = entries_base.filter(ar_no__iexact=ar_no)
    if stage:
        entries_base = entries_base.filter(stage__iexact=stage)
    if batch_no:
        entries_base = entries_base.filter(batch_no__iexact=batch_no)
    if month:
        try:
            y, m = month.split("-")
            entries_base = entries_base.filter(
                entry_date__year=int(y),
                entry_date__month=int(m),
            )
        except (ValueError, TypeError):
            pass

    # 3) Criticality gate (same logic as qc_dashboard)
    if criticality == "critical":
        entry_ids = (
            SpecEntry.objects.filter(q_is_critical)
            .values_list("qc_entry_id", flat=True)
            .distinct()
        )
        entries = entries_base.filter(id__in=entry_ids)
    elif criticality == "non_critical":
        noncrit_ids = (
            SpecEntry.objects.exclude(q_is_critical)
            .values_list("qc_entry_id", flat=True)
            .distinct()
        )
        entries = entries_base.filter(id__in=noncrit_ids)
    else:
        entries = entries_base

    pending_q = Q(decision_status__isnull=True) | Q(decision_status="")

    # ---------- 4) KPI summary (same style as dashboard) ----------
    summary = {
        "total":           entries.count(),
        "approved":        entries.filter(decision_status="approved").count(),
        "variation":       entries.filter(decision_status="approved_under_deviation").count(),
        "rejected":        entries.filter(decision_status="rejected").count(),
        "fail":            entries.filter(decision_status="fail").count(),
        "pending":         entries.filter(pending_q).count(),
        "distinct_batches": (
            entries
            .exclude(batch_no__isnull=True)
            .exclude(batch_no="")
            .values("batch_no")
            .distinct()
            .count()
        ),
    }

    # If critical / non_critical, mirror qc_dashboard behaviour
    if criticality in ("critical", "non_critical"):
        spec_rows_for_cards = (
            SpecEntry.objects
            .filter(qc_entry__in=entries_base.values_list("id", flat=True))
        )
        if criticality == "critical":
            spec_subset = spec_rows_for_cards.filter(q_is_critical)
        else:
            spec_subset = spec_rows_for_cards.exclude(q_is_critical)

        entry_ids_subset = list(
            spec_subset.values_list("qc_entry_id", flat=True).distinct()
        )
        entries_subset = QCEntry.objects.filter(id__in=entry_ids_subset)
        summary = {
            "total":     entries_subset.count(),
            "approved":  entries_subset.filter(decision_status="approved").count(),
            "variation": entries_subset.filter(decision_status="approved_under_deviation").count(),
            "rejected":  entries_subset.filter(decision_status="rejected").count(),
            "fail":      entries_subset.filter(decision_status="fail").count(),
            "pending":   entries_subset.filter(pending_q).count(),
            "distinct_batches": (
                entries_subset
                .exclude(batch_no__isnull=True).exclude(batch_no="")
                .values("batch_no")
                .distinct()
                .count()
            ),
        }

    # ---------- 4b) Outcomes by Product data (for stacked bar chart) ----------
    product_outcomes = []
    product_qs = (
        entries.values("product__name")
        .annotate(
            approved=Count("id", filter=Q(decision_status="approved")),
            aud=Count("id", filter=Q(decision_status="approved_under_deviation")),
            rejected=Count("id", filter=Q(decision_status="rejected")),
        )
        .order_by("product__name")
    )
    for row in product_qs:
        name = (row["product__name"] or "").strip()
        if not name:
            continue
        product_outcomes.append({
            "product":  name,
            "approved": row["approved"],
            "aud":      row["aud"],
            "rejected": row["rejected"],
        })

    # ---------- 5) Spec queryset for trend & detail ----------
    filters_applied = any([
        item, group_name, test_name, ar_no, stage, batch_no,
        decision_status != "total", month, criticality != "all",
    ])

    spec_qs = SpecEntry.objects.select_related(
        "qc_entry__product", "spec", "qc_entry"
    )
    if item:
        spec_qs = spec_qs.filter(qc_entry__product__name__iexact=item)
    if group_name:
        spec_qs = spec_qs.filter(spec__group__iexact=group_name)
    if test_name:
        spec_qs = spec_qs.filter(spec__name__iexact=test_name)
    if ar_no:
        spec_qs = spec_qs.filter(qc_entry__ar_no__iexact=ar_no)
    if stage:
        spec_qs = spec_qs.filter(qc_entry__stage__iexact=stage)
    if batch_no:
        spec_qs = spec_qs.filter(qc_entry__batch_no__iexact=batch_no)

    if decision_status != "total":
        if decision_status == "nodecision":
            spec_qs = spec_qs.filter(
                Q(qc_entry__decision_status__isnull=True)
                | Q(qc_entry__decision_status="")
            )
        else:
            spec_qs = spec_qs.filter(qc_entry__decision_status=decision_status)

    if criticality == "critical":
        spec_qs = spec_qs.filter(q_is_critical)
    elif criticality == "non_critical":
        spec_qs = spec_qs.exclude(q_is_critical)

    if month:
        try:
            y, m = month.split("-")
            spec_qs = spec_qs.filter(
                qc_entry__entry_date__year=int(y),
                qc_entry__entry_date__month=int(m),
            )
        except (ValueError, TypeError):
            pass

    # ---------- 6) Trend data (same as qc_dashboard) ----------
    if filters_applied:
        spec_iter = spec_qs.order_by("qc_entry__release_by_qc_at", "qc_entry__id")
    else:
        spec_iter = list(
            spec_qs.order_by("-qc_entry__release_by_qc_at", "-id")[:1000]
        )[::-1]

    def _collect_numeric_rows(qs_iter):
        out = []
        for se in qs_iter:
            obs = _to_float_or_none(se.value)
            if obs is None:
                continue
            e = se.qc_entry
            if not e.batch_no:
                continue
            out.append({
                "label":     _json_coerce_val(e.batch_no),
                "batch_no":  _json_coerce_val(e.batch_no),
                "ar_no":     _json_coerce_val(e.ar_no),
                "entry_no":  f"#{_json_coerce_val(e.entry_no)}",
                "min_val":   _to_float_or_none(getattr(se.spec, "min_val", None)),
                "max_val":   _to_float_or_none(getattr(se.spec, "max_val", None)),
                "observed":  obs,
                "release_timestamp": e.release_by_qc_at,
                "date":      e.release_by_qc_at.strftime("%Y-%m-%d") if e.release_by_qc_at else "",
                "spec_name": _json_coerce_val(se.spec.name if se.spec else "Unknown"),
                "decision_status": e.decision_status,
            })
        return out

    all_trend_rows = _collect_numeric_rows(spec_iter)
    trend_data = _deduplicate_trend_data(all_trend_rows)

    if not trend_data:
        fb_qs = (
            SpecEntry.objects
            .select_related("qc_entry__product", "spec", "qc_entry")
            .exclude(value__isnull=True)
        )
        if item:
            fb_qs = fb_qs.filter(qc_entry__product__name__iexact=item)
        if group_name:
            fb_qs = fb_qs.filter(spec__group__iexact=group_name)

        fallback_rows = _collect_numeric_rows(
            list(fb_qs.order_by("-qc_entry__release_by_qc_at", "-id")[:2000])[::-1]
        )
        trend_data = _deduplicate_trend_data(fallback_rows)[-200:]

    # ---------- 7) Detail rows (flattened SpecEntry view) ----------
    detail_qs = spec_qs.order_by(
        "qc_entry__entry_date", "qc_entry__id", "id"
    )

    # ---------- 8) Build Excel workbook ----------
    output = BytesIO()
    wb = xlsxwriter.Workbook(output, {"in_memory": True})

    # Formats
    title_fmt    = wb.add_format({"bold": True, "font_size": 14})
    subtitle_fmt = wb.add_format({"italic": True, "font_color": "#6b7280"})
    head_fmt     = wb.add_format({"bold": True, "border": 1, "bg_color": "#E5F0FF"})
    cell_fmt     = wb.add_format({"border": 1})
    num_fmt      = wb.add_format({"border": 1, "num_format": "0.000"})
    date_fmt     = wb.add_format({"border": 1, "num_format": "yyyy-mm-dd"})
    dt_fmt       = wb.add_format({"border": 1, "num_format": "yyyy-mm-dd hh:mm"})

    # Formats for conditional formatting
    fail_fmt = wb.add_format({"font_color": "#FF0000"})   # red text
    pass_fmt = wb.add_format({"font_color": "#008000"})   # green text
    aud_fmt  = wb.add_format({"bg_color": "#FFF2CC"})     # yellow background

    today_str = datetime.date.today().strftime("%d-%b-%Y")

    # ---------- Sheet 1: Summary ----------
    ws0 = wb.add_worksheet("Summary")

    ws0.write(0, 0, "QC Analytics Dashboard Export", title_fmt)
    ws0.write(1, 0, f"Generated on: {today_str}", subtitle_fmt)

    # Filters used
    row = 3
    ws0.write(row, 0, "Filters", head_fmt); row += 1
    ws0.write(row, 0, "Item", cell_fmt);            ws0.write(row, 1, item or "All", cell_fmt); row += 1
    ws0.write(row, 0, "Group", cell_fmt);           ws0.write(row, 1, group_name or "All", cell_fmt); row += 1
    ws0.write(row, 0, "Test", cell_fmt);            ws0.write(row, 1, test_name or "All", cell_fmt); row += 1
    ws0.write(row, 0, "AR No.", cell_fmt);          ws0.write(row, 1, ar_no or "All", cell_fmt); row += 1
    ws0.write(row, 0, "Stage", cell_fmt);           ws0.write(row, 1, stage or "All", cell_fmt); row += 1
    ws0.write(row, 0, "Batch", cell_fmt);           ws0.write(row, 1, batch_no or "All", cell_fmt); row += 1
    ws0.write(row, 0, "Month", cell_fmt);           ws0.write(row, 1, month or "All", cell_fmt); row += 1
    ws0.write(row, 0, "Criticality", cell_fmt);     ws0.write(row, 1, criticality or "all", cell_fmt); row += 1
    ws0.write(row, 0, "Decision Status", cell_fmt); ws0.write(row, 1, decision_status or "total", cell_fmt)

    # KPI block
    row = 3
    col = 3
    ws0.write(row, col, "KPI", head_fmt); ws0.write(row, col + 1, "Value", head_fmt); row += 1
    kpi_rows = [
        ("Total ARs", summary["total"]),
        ("Distinct Batches", summary["distinct_batches"]),
        ("Approved", summary["approved"]),
        ("Approved (Deviation)", summary["variation"]),
        ("Rejected", summary["rejected"]),
        ("Fail", summary["fail"]),
        ("Pending / No Status", summary["pending"]),
    ]
    for label, val in kpi_rows:
        ws0.write(row, col, label, cell_fmt)
        ws0.write_number(row, col + 1, val, cell_fmt)
        row += 1

    ws0.set_column(0, 1, 18)
    ws0.set_column(3, 4, 22)

    # ---------- Sheet 2: Trend Data + line chart ----------
    ws1 = wb.add_worksheet("Trend Data")
    headers_trend = [
        "Batch No", "AR No", "Entry No", "Release Date",
        "Spec Name", "Observed", "Spec Min", "Spec Max", "Decision Status",
    ]
    for c, h in enumerate(headers_trend):
        ws1.write(0, c, h, head_fmt)

    r = 1
    for row_data in trend_data:
        batch = row_data.get("batch_no") or row_data.get("label") or ""
        ws1.write(r, 0, batch, cell_fmt)
        ws1.write(r, 1, row_data.get("ar_no") or "", cell_fmt)
        ws1.write(r, 2, row_data.get("entry_no") or "", cell_fmt)

        rel_ts = row_data.get("release_timestamp")
        if rel_ts:
            if getattr(rel_ts, "tzinfo", None) is not None:
                rel_ts = rel_ts.replace(tzinfo=None)
            ws1.write_datetime(r, 3, rel_ts, dt_fmt)
        else:
            ws1.write(r, 3, "", cell_fmt)

        ws1.write(r, 4, row_data.get("spec_name") or "", cell_fmt)
        ws1.write_number(r, 5, float(row_data.get("observed") or 0), num_fmt)

        min_v = row_data.get("min_val")
        max_v = row_data.get("max_val")
        if min_v is not None:
            ws1.write_number(r, 6, float(min_v), num_fmt)
        else:
            ws1.write(r, 6, "", cell_fmt)
        if max_v is not None:
            ws1.write_number(r, 7, float(max_v), num_fmt)
        else:
            ws1.write(r, 7, "", cell_fmt)

        ws1.write(r, 8, row_data.get("decision_status") or "", cell_fmt)
        r += 1

    last_row = max(r - 1, 1)
    ws1.set_column(0, 0, 14)
    ws1.set_column(1, 2, 12)
    ws1.set_column(3, 3, 20)
    ws1.set_column(4, 4, 22)
    ws1.set_column(5, 7, 12)
    ws1.set_column(8, 8, 16)

    # --- NEW: conditional formatting for Decision Status column on Trend Data ---
    if last_row >= 1:
        # column I (index 8)
        ws1.conditional_format(1, 8, last_row, 8, {
            "type": "cell",
            "criteria": "==",
            "value": '"approved"',
            "format": pass_fmt,
        })
        ws1.conditional_format(1, 8, last_row, 8, {
            "type": "cell",
            "criteria": "==",
            "value": '"approved_under_deviation"',
            "format": aud_fmt,
        })
        ws1.conditional_format(1, 8, last_row, 8, {
            "type": "cell",
            "criteria": "==",
            "value": '"fail"',
            "format": fail_fmt,
        })

    # Line chart: Observed vs Spec Min/Max (wider, placed below KPI table)
    if last_row >= 1:
        line_chart = wb.add_chart({"type": "line"})
        sheet_name = "Trend Data"

        line_chart.add_series({
            "name":       "Observed",
            "categories": [sheet_name, 1, 0, last_row, 0],
            "values":     [sheet_name, 1, 5, last_row, 5],
            "line":       {"width": 2},
        })
        line_chart.add_series({
            "name":       "Spec Max",
            "categories": [sheet_name, 1, 0, last_row, 0],
            "values":     [sheet_name, 1, 7, last_row, 7],
            "line":       {"dash_type": "dash"},
        })
        line_chart.add_series({
            "name":       "Spec Min",
            "categories": [sheet_name, 1, 0, last_row, 0],
            "values":     [sheet_name, 1, 6, last_row, 6],
            "line":       {"dash_type": "dash"},
        })

        line_chart.set_title({"name": "Observed vs Spec Limits"})
        line_chart.set_x_axis({"name": "Batch"})
        line_chart.set_y_axis({"name": "Value"})
        line_chart.set_legend({"position": "bottom"})

        # Place line chart BELOW the tables, wide across the sheet
        ws0.insert_chart("A15", line_chart, {"x_scale": 2.0, "y_scale": 1.2})

    # ---------- Sheet 3: Detail Data ----------
    ws2 = wb.add_worksheet("Detail Data")
    headers_detail = [
        "Product", "Stage", "Entry Date", "Entry No", "AR No",
        "Batch", "Decision Status",
        "Spec Group", "Spec Name", "Spec Type",
        "Spec Min", "Spec Max",
        "Observed Value", "Unit", "Remark", "Critical?",
    ]
    for c, h in enumerate(headers_detail):
        ws2.write(0, c, h, head_fmt)

    r = 1
    for se in detail_qs.iterator():
        e = se.qc_entry
        prod_name = e.product.name if e.product else ""
        ws2.write(r, 0, prod_name, cell_fmt)
        ws2.write(r, 1, e.stage or "", cell_fmt)

        if e.entry_date:
            dt = e.entry_date
            if getattr(dt, "tzinfo", None) is not None:
                dt = dt.replace(tzinfo=None)
            ws2.write_datetime(r, 2, dt, date_fmt)
        else:
            ws2.write(r, 2, "", cell_fmt)

        ws2.write(r, 3, e.entry_no or "", cell_fmt)
        ws2.write(r, 4, e.ar_no or "", cell_fmt)
        ws2.write(r, 5, e.batch_no or "", cell_fmt)
        ws2.write(r, 6, e.decision_status or "", cell_fmt)

        sp = se.spec
        ws2.write(r, 7, getattr(sp, "group", "") or "", cell_fmt)
        ws2.write(r, 8, sp.name if sp else "", cell_fmt)
        ws2.write(r, 9, getattr(sp, "spec_type", "") or "", cell_fmt)

        if sp and sp.min_val is not None:
            ws2.write_number(r, 10, float(sp.min_val), num_fmt)
        else:
            ws2.write(r, 10, "", cell_fmt)
        if sp and sp.max_val is not None:
            ws2.write_number(r, 11, float(sp.max_val), num_fmt)
        else:
            ws2.write(r, 11, "", cell_fmt)

        if se.value is not None and isinstance(se.value, (int, float)):
            ws2.write_number(r, 12, float(se.value), num_fmt)
        else:
            ws2.write(r, 12, _json_coerce_val(se.value) or "", cell_fmt)

        ws2.write(r, 13, getattr(sp, "unit", "") or "", cell_fmt)
        ws2.write(r, 14, se.remark or "", cell_fmt)

        is_critical = bool(
            getattr(se, "is_critical", False)
            or getattr(sp, "is_critical", False)
        )
        ws2.write(r, 15, "Yes" if is_critical else "No", cell_fmt)
        r += 1

    # Column widths
    ws2.set_column(0, 0, 28)   # product
    ws2.set_column(1, 1, 16)   # stage
    ws2.set_column(2, 2, 12)   # date
    ws2.set_column(3, 5, 12)   # entry/ar/batch
    ws2.set_column(6, 6, 16)   # decision
    ws2.set_column(7, 9, 18)   # spec group/name/type
    ws2.set_column(10, 12, 12) # spec limits & observed
    ws2.set_column(13, 15, 18) # unit, remark, critical

    # === Conditional formatting for Decision Status & Remark (Detail Data) ===
    last_detail_row = r - 1
    if last_detail_row >= 1:
        # Decision Status column (G -> index 6)
        ws2.conditional_format(1, 6, last_detail_row, 6, {
            "type": "cell",
            "criteria": "==",
            "value": '"approved"',
            "format": pass_fmt,
        })
        ws2.conditional_format(1, 6, last_detail_row, 6, {
            "type": "cell",
            "criteria": "==",
            "value": '"approved_under_deviation"',
            "format": aud_fmt,
        })
        ws2.conditional_format(1, 6, last_detail_row, 6, {
            "type": "cell",
            "criteria": "==",
            "value": '"fail"',
            "format": fail_fmt,
        })

        # Remark column (O -> index 14)
        ws2.conditional_format(1, 14, last_detail_row, 14, {
            "type": "cell",
            "criteria": "==",
            "value": '"Pass"',
            "format": pass_fmt,
        })
        ws2.conditional_format(1, 14, last_detail_row, 14, {
            "type": "cell",
            "criteria": "==",
            "value": '"Fail"',
            "format": fail_fmt,
        })
        ws2.conditional_format(1, 14, last_detail_row, 14, {
            "type": "cell",
            "criteria": "==",
            "value": '"Approved under deviation"',
            "format": aud_fmt,
        })

    # ---------- Sheet 4: Outcomes by Product data + stacked bar chart ----------
    ws3 = wb.add_worksheet("OutcomesByProduct")
    ws3.write(0, 0, "Product", head_fmt)
    ws3.write(0, 1, "Approved", head_fmt)
    ws3.write(0, 2, "AUD", head_fmt)
    ws3.write(0, 3, "Rejected", head_fmt)

    pr = 1
    for d in product_outcomes:
        ws3.write(pr, 0, d["product"], cell_fmt)
        ws3.write_number(pr, 1, d["approved"], cell_fmt)
        ws3.write_number(pr, 2, d["aud"], cell_fmt)
        ws3.write_number(pr, 3, d["rejected"], cell_fmt)
        pr += 1

    ws3.set_column(0, 0, 40)
    ws3.set_column(1, 3, 12)

    last_prod_row = max(pr - 1, 1)
    if product_outcomes:
        sheet_prod = "OutcomesByProduct"
        bar_chart = wb.add_chart({"type": "column", "subtype": "stacked"})

        bar_chart.add_series({
            "name": "Approved",
            "categories": [sheet_prod, 1, 0, last_prod_row, 0],
            "values":     [sheet_prod, 1, 1, last_prod_row, 1],
        })
        bar_chart.add_series({
            "name": "AUD",
            "categories": [sheet_prod, 1, 0, last_prod_row, 0],
            "values":     [sheet_prod, 1, 2, last_prod_row, 2],
        })
        bar_chart.add_series({
            "name": "Rejected",
            "categories": [sheet_prod, 1, 0, last_prod_row, 0],
            "values":     [sheet_prod, 1, 3, last_prod_row, 3],
        })

        bar_chart.set_title({"name": "Outcomes by Product"})
        bar_chart.set_x_axis({
            "name": "Product",
            "num_font": {"rotation": -45},
        })
        bar_chart.set_y_axis({"name": "Count"})
        bar_chart.set_legend({"position": "top"})

        # Place stacked bar chart BELOW the line chart on Summary sheet
        ws0.insert_chart("A32", bar_chart, {"x_scale": 2.0, "y_scale": 1.3})

    # Finish & response
    wb.close()
    output.seek(0)
    filename = f"QC_Dashboard_{datetime.date.today().isoformat()}.xlsx"
    resp = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp




@login_required
def qc_product_report(request):
    logger.info("[QC][Report] user=%s accessed QC Report page", request.user.username)
    if not request.user.has_perm('QC.view_qcentry'):
        logger.warning(f"[QC] Unauthorized report view attempt by user '{request.user.username}'")
        messages.error(request, "You do not have permission to view this report.")
        return redirect('indexpage')

    # --- 1. Get Filters and Products ---
    all_products = Product.objects.all().order_by('name')
    product_id_query = request.GET.get('product_id', '').strip()
    start_date_query = request.GET.get('start_date', '').strip()
    end_date_query   = request.GET.get('end_date', '').strip()
    batch_no_query   = request.GET.get('batch_no', '').strip()

    # --- 2. Initialize Context ---
    context = {
        'all_products': all_products,
        'selected_product_id': None,
        'selected_product_label': "",   # 🔹 for product search box
        'start_date': start_date_query,
        'end_date': end_date_query,
        'batch_no': batch_no_query,
        'header_info': [],
        'report_rows': [],
        'product': None,
        'batch_no_choices': [],   # 🔹 for batch dropdown
    }

    # --- 3. Process if a Product is Selected ---
    if product_id_query:
        try:
            context['selected_product_id'] = int(product_id_query)

            # Base queryset (used for both rows + batch list)
            base_qs = (
                QCEntry.objects
                .filter(product_id=product_id_query, status='qc_completed')
                .select_related('product')
                .prefetch_related('values', 'values__spec')
            )

            # Date filters
            if start_date_query:
                base_qs = base_qs.filter(release_by_qc_at__date__gte=start_date_query)
            if end_date_query:
                base_qs = base_qs.filter(release_by_qc_at__date__lte=end_date_query)

            # 🔹 Build list of existing batch numbers for dropdown
            batch_no_choices = (
                base_qs.exclude(batch_no__isnull=True)
                       .exclude(batch_no__exact='')
                       .values_list('batch_no', flat=True)
                       .distinct()
                       .order_by('batch_no')
            )
            context['batch_no_choices'] = list(batch_no_choices)

            # Final queryset for table (apply batch filter if given)
            qs = base_qs
            if batch_no_query:
                qs = qs.filter(batch_no__icontains=batch_no_query)

            qs = qs.order_by('-release_by_qc_at')
            filtered_entries = list(qs)

            # --- 4. Prepare Data for the Report Table ---
            if filtered_entries:
                product = filtered_entries[0].product
                context['product'] = product

                # label text for product search box
                if getattr(product, "stages", None):
                    context['selected_product_label'] = f"{product.name} - {product.stages}"
                else:
                    context['selected_product_label'] = product.name

                # ================= HEADER CONSOLIDATION =================
                all_product_specs = list(
                    Spec.objects.filter(product=product).order_by('id')
                )

                header_specs = []
                seen_spec_names = set()

                for spec in all_product_specs:
                    if spec.name not in seen_spec_names:
                        header_specs.append(spec)
                        seen_spec_names.add(spec.name)

                header_info = []
                for spec in header_specs:
                    main_header = spec.name
                    sub_header = "NA"

                    if spec.spec_type == Spec.TYPE_NUMERIC:
                        min_val, max_val = spec.min_val, spec.max_val
                        if min_val is not None and max_val is not None:
                            sub_header = f"{min_val} - {max_val}".replace('.000', '')
                        elif max_val is not None:
                            sub_header = f"≤ {max_val}".replace('.000', '')
                    elif spec.spec_type == Spec.TYPE_CHOICE:
                        sub_header = spec.allowed_choices

                    header_info.append({'main': main_header, 'sub': sub_header})

                context['header_info'] = header_info

                # ================= ROW DATA GENERATION =================
                for entry in filtered_entries:
                    values_by_name_map = {
                        val.spec.name: val
                        for val in entry.values.all()
                        if val.spec
                    }

                    date_val = (
                        entry.release_by_qc_at.strftime('%d.%m.%y')
                        if entry.release_by_qc_at else "-"
                    )

                    row_data = [
                        date_val,
                        entry.batch_no,
                        entry.qty,
                    ]

                    for header_spec in header_specs:
                        spec_entry_for_row = values_by_name_map.get(header_spec.name)

                        if spec_entry_for_row:
                            if header_spec.name == 'Appearance':
                                value_to_append = spec_entry_for_row.spec.allowed_choices
                            else:
                                value_to_append = spec_entry_for_row.value
                        else:
                            value_to_append = "-"

                        row_data.append(value_to_append)

                    context['report_rows'].append(row_data)

        except (ValueError, Product.DoesNotExist):
            messages.error(request, "Invalid product selected. Please try again.")
            return redirect('qc:qc_product_report')

    # if product_id set but no rows, still build label from master
    if context['selected_product_id'] and not context['selected_product_label']:
        p = all_products.filter(id=context['selected_product_id']).first()
        if p:
            context['selected_product_label'] = (
                f"{p.name} - {p.stages}" if getattr(p, "stages", None) else p.name
            )
    # --- 6. Render the Page ---
    return render(request, 'qc/qc_product_report.html', context)



@login_required
def download_qc_report_excel(request):
    """
    Fetches the filtered QC report data and generates an Excel file for download,
    including a title.
    """
    # --- 1. Get Filters from the request ---
    product_id_query = request.GET.get('product_id', '').strip()
    start_date_query = request.GET.get('start_date', '').strip()
    end_date_query   = request.GET.get('end_date', '').strip()
    batch_no_query   = request.GET.get('batch_no', '').strip()   # 🔹 NEW

    if not request.user.has_perm('QC.view_qcentry'):
        messages.error(request, "You do not have permission to download this report.")
        return redirect('qc:qc_product_report')

    if not product_id_query:
        messages.error(request, "A product must be selected to download the report.")
        return redirect('qc:qc_product_report')

    try:
        # --- 2. Build base queryset (same logic as qc_product_report) ---
        qs = (
            QCEntry.objects
            .filter(product_id=product_id_query, status='qc_completed')
            .select_related('product')
            .prefetch_related('values', 'values__spec')
        )

        # Date filters
        if start_date_query:
            qs = qs.filter(release_by_qc_at__date__gte=start_date_query)
        if end_date_query:
            qs = qs.filter(release_by_qc_at__date__lte=end_date_query)

        # 🔹 Batch filter (same as view)
        if batch_no_query:
            qs = qs.filter(batch_no__icontains=batch_no_query)

        # Stable ordering
        qs = qs.order_by('-release_by_qc_at', 'id')

        filtered_entries = list(qs)

        if not filtered_entries:
            messages.warning(request, "No data found for the selected criteria to download.")
            return redirect('qc:qc_product_report')

        product = filtered_entries[0].product

        # --- 3. Prepare Headers and Rows (aligned with qc_product_report) ---
        all_product_specs = list(Spec.objects.filter(product=product).order_by('id'))

        # unique by spec.name
        header_specs = []
        seen_spec_names = set()
        for spec in all_product_specs:
            if spec.name not in seen_spec_names:
                header_specs.append(spec)
                seen_spec_names.add(spec.name)

        header_info = []
        for spec in header_specs:
            main_header = spec.name
            sub_header = "NA"

            if spec.spec_type == Spec.TYPE_NUMERIC:
                min_val, max_val = spec.min_val, spec.max_val
                if min_val is not None and max_val is not None:
                    sub_header = f"{min_val} - {max_val}".replace('.000', '')
                elif max_val is not None:
                    sub_header = f"≤ {max_val}".replace('.000', '')
            elif spec.spec_type == Spec.TYPE_CHOICE:
                sub_header = spec.allowed_choices

            header_info.append({'main': main_header, 'sub': sub_header})

        # ---- build rows (same logic as page table) ----
        report_rows = []
        for entry in filtered_entries:
            # map by spec.name
            values_by_name_map = {
                val.spec.name: val
                for val in entry.values.all()
                if val.spec
            }

            date_val = (
                entry.release_by_qc_at.strftime('%d.%m.%Y')
                if entry.release_by_qc_at else "-"
            )

            row_data = [date_val, entry.batch_no, entry.qty]

            for header_spec in header_specs:
                spec_entry_for_row = values_by_name_map.get(header_spec.name)

                if spec_entry_for_row:
                    if header_spec.name == 'Appearance':
                        # show allowed_choices for that row's spec (same as HTML)
                        value_to_append = spec_entry_for_row.spec.allowed_choices
                    else:
                        value_to_append = spec_entry_for_row.value
                else:
                    value_to_append = "-"

                row_data.append(value_to_append)

            report_rows.append(row_data)

        # --- 4. Generate the Excel File ---
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet("QC Report")

        title_format = workbook.add_format({
            'bold': True,
            'font_size': 16,
            'align': 'center',
            'valign': 'v_center',
        })
        header_format = workbook.add_format({
            'bold': True,
            'align': 'center',
            'valign': 'v_center',
            'border': 1,
            'bg_color': '#DDEBF7',
        })
        sub_header_format = workbook.add_format({
            'align': 'center',
            'valign': 'v_center',
            'border': 1,
            'bg_color': '#E2EFDA',
        })
        cell_format = workbook.add_format({
            'border': 1,
            'align': 'center',
            'valign': 'v_center',
        })

        last_col_index = 2 + len(header_info)
        worksheet.merge_range(
            0, 0, 0, last_col_index,
            f'QC Monitoring Report for {product.name}',
            title_format,
        )
        worksheet.set_row(0, 30)

        # multi-level header
        worksheet.merge_range('A3:A4', 'Date', header_format)
        worksheet.merge_range('B3:B4', 'Batch No.', header_format)
        worksheet.write('C3', 'Qty.', header_format)
        worksheet.write('C4', 'in Kg', sub_header_format)

        col_num = 3
        for header in header_info:
            worksheet.write(2, col_num, header['main'], header_format)
            worksheet.write(3, col_num, header.get('sub', 'NA'), sub_header_format)
            worksheet.set_column(col_num, col_num, 15)
            col_num += 1

        # data rows
        for row_num, data_row in enumerate(report_rows, start=4):
            for col_idx, cell_data in enumerate(data_row):
                try:
                    num_data = float(cell_data)
                    worksheet.write_number(row_num, col_idx, num_data, cell_format)
                except (ValueError, TypeError):
                    worksheet.write_string(
                        row_num, col_idx,
                        str(cell_data or '-'),
                        cell_format,
                    )

        worksheet.set_column('A:A', 12)
        worksheet.set_column('B:B', 20)
        worksheet.set_column('C:C', 10)

        workbook.close()
        output.seek(0)

        # --- 5. Create and return the HTTP response ---
        filename = f"QC_Report_{product.code}_{timezone.now().strftime('%Y%m%d')}.xlsx"
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    except Exception as e:
        logger.error(f"Error generating Excel report: {e}")
        messages.error(request, "An error occurred while generating the Excel report.")
        return redirect('qc:qc_product_report')



# ─────────────────────────────────────────────────────────────────────────────
# Daily QA Report – optimized, no-flow-change + PDL List view (additive)
# ─────────────────────────────────────────────────────────────────────────────
import importlib
import logging
import time
from datetime import date as _date, datetime as _datetime, timedelta, datetime, time as _time
import time as systime
from django.db import transaction
from django.db.models import (
    Sum, F, Case, When, DateTimeField, ExpressionWrapper,
)
from .forms import DailyQAReportForm, IncomingFS, PDLFS, PDLHeaderForm, PDLOnlyFS  # PDLOnlyFS used by entries page
from .models import DailyQAReport, IncomingGRNCache, PDLSample  # (IncomingMaterial is not used directly here)


# ─── Prefill services loader (unchanged behavior) ────────────────────────────
try:
    _prefill = importlib.import_module("QC.services_prefill")
except Exception:
    class _NoPrefill:
        def fetch_incoming_rm_pm(self, d): return []
        def fetch_pdl_samples(self, d): return []
        def fetch_other_details(self, d): return {}
    _prefill = _NoPrefill()

fetch_incoming_rm_pm = getattr(_prefill, "fetch_incoming_rm_pm", lambda d: [])
fetch_pdl_samples    = getattr(_prefill, "fetch_pdl_samples",    lambda d: [])
_fetch_other_details = getattr(_prefill, "fetch_other_details",  lambda d: {})

# ─── Date parsers (deduped) ──────────────────────────────────────────────────
def _parse_date_base(qs: str | None) -> _date:
    if not qs:
        return _date.today()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return _datetime.strptime(qs, fmt).date()
        except Exception:
            continue
    return _date.today()

# Backwards-compatible aliases (keep existing call sites working)
_parse_daily_date = _parse_date_base
_parse_date       = _parse_date_base

# ─── Small helpers (deduped) ─────────────────────────────────────────────────
def _as_float(v):
    try:
        return float(v)
    except Exception:
        return 0.0

def _mt_from_item_type(s: str | None) -> str:
    s = (s or "").strip().lower()
    return "Packing Material" if s.startswith("packing") else "Raw Material"


# ─── Ensure GRN cache (unchanged, with debounce) ─────────────────────────────
_last_inc_sync: dict[str, float] = {}  # {"YYYY-MM-DD": unix_ts}

def _ensure_grn_cache_for(day: _date):
    key = str(day)
    if IncomingGRNCache.objects.filter(grn_date=day).exists():
        return
    # ✅ use systime.time() instead of time.time()
    if _last_inc_sync.get(key, 0) > systime.time() - 60:
        return
    _last_inc_sync[key] = systime.time()
    try:
        call_command("sync_incoming_grn", date=key, verbosity=0)
    except Exception as e:
        logger.warning("sync_incoming_grn failed for %s: %s", key, e)

def _incoming_rows_from_cache(day: _date):
    rows = []
    qs = IncomingGRNCache.objects.filter(grn_date=day).order_by("grn_no", "item_name")
    for r in qs:
        rows.append({
            "material_type": _mt_from_item_type(r.item_type),
            "material":      (r.item_name or "").strip(),
            "supplier":      (r.supplier_name or "").strip(),
            "qty_mt":        _as_float(r.qty),
            "status":        "Pass",
            "remarks":       "",
        })
    return rows

# ─── Analytical downtime / complaints / mistakes (deduped) ───────────────────
def _sum_analytical_downtime_hours(day: _date) -> Decimal:
    try:
        from .models import AnalyticalDowntime
    except Exception:
        return Decimal("0")

    fields = {f.name for f in AnalyticalDowntime._meta.get_fields()}
    base = (
        AnalyticalDowntime.objects.filter(date=day)
        if "date" in fields else
        AnalyticalDowntime.objects.filter(start_at__date=day)
    )

    for hours_col in ("duration_hours", "hours"):
        if hours_col in fields:
            total = base.aggregate(total=Sum(hours_col)).get("total") or Decimal("0")
            try:
                return (total if isinstance(total, Decimal) else Decimal(str(total))).quantize(Decimal("0.01"))
            except Exception:
                return Decimal("0")

    # duration from datetimes (open → now)
    try:
        from django.db.models import DurationField
        dur_expr = ExpressionWrapper(Coalesce(F("end_at"), Now()) - F("start_at"), output_field=DurationField())
        total = base.order_by().aggregate(total=Sum(dur_expr))["total"]
        if not total:
            return Decimal("0")
        hrs = Decimal(str(total.total_seconds() / 3600.0)).quantize(Decimal("0.01"))
        return hrs if hrs >= 0 else Decimal("0")
    except Exception:
        try:
            secs = 0
            now_ = timezone.now()
            for r in base.only("start_at", "end_at"):
                sa, ea = getattr(r, "start_at", None), getattr(r, "end_at", None) or now_
                if sa and ea:
                    secs += max(0, (ea - sa).total_seconds())
            return Decimal(str(secs / 3600.0)).quantize(Decimal("0.01"))
        except Exception:
            return Decimal("0")

def _dynamic_model(name_candidates: tuple[str, ...]):
    """Import-first existing model by name; return None if not found."""
    mod_name = __name__.rsplit(".", 1)[0] + ".models"
    for nm in name_candidates:
        try:
            return getattr(importlib.import_module(mod_name), nm)
        except Exception:
            continue
    return None

def _count_customer_complaints(day: _date) -> int:
    Model = _dynamic_model(("CustomerComplaint", "CustomerComplaints", "Complaint", "Complaints"))
    if not Model:
        return 0
    fields = {f.name for f in Model._meta.get_fields()}
    date_field = "date" if "date" in fields else "complaint_date" if "complaint_date" in fields else "reported_on" if "reported_on" in fields else None
    return Model.objects.filter(**{date_field: day}).count() if date_field else 0

def _count_analytical_mistakes(day: _date) -> int:
    Model = _dynamic_model(("AnalyticalMistake", "AnalyticalMistakes", "LabMistake", "LabMistakes"))
    if not Model:
        return 0
    fields = {f.name for f in Model._meta.get_fields()}
    date_field = "date" if "date" in fields else "mistake_date" if "mistake_date" in fields else "reported_on" if "reported_on" in fields else None
    return Model.objects.filter(**{date_field: day}).count() if date_field else 0

def _count_first_aid_injury_qc(day):
    """
    Count 'First-aid Injury' incidents for QC on a given date.
    Uses incident_date if present, else record_date.
    Filters to department='QC' (or physical_location name='QC').
    """
    from django.db.models.functions import Coalesce
    from django.db.models import Q, DateField

    # Import from the EHS app (note: your folder is 'EHS')
    try:
        from EHS.models import Lagging_Indicator
    except ModuleNotFoundError:
        # fallback if the app is lowercased on some machines
        from EHS.models import Lagging_Indicator

    eff_date = Coalesce("incident_date", "record_date", output_field=DateField())

    qs = (
        Lagging_Indicator.objects
        .annotate(eff_date=eff_date)
        .filter(eff_date=day)
        .filter(
            Q(department__iexact="QA/QC")
        )
    )
    return qs.count()

def _safe_header_metrics(day: _date) -> dict:
    """Compute header metrics defensively (strings/ints only)."""
    data = {"analytical_downtime_hrs": "0", "customer_complaints": 0, "analytical_mistakes": 0}
    try:
        v = _sum_analytical_downtime_hours(day)
        data["analytical_downtime_hrs"] = str(v.normalize() if isinstance(v, Decimal) else round(float(v), 2))
    except Exception as e:
        logger.warning("downtime metric failed: %s", e)
    try:
        data["customer_complaints"] = int(_count_customer_complaints(day))
    except Exception as e:
        logger.warning("complaints metric failed: %s", e)
    try:
        data["analytical_mistakes"] = int(_count_analytical_mistakes(day))
    except Exception as e:
        logger.warning("mistakes metric failed: %s", e)
    return data

# ─── Process deviations (additive) ───────────────────────────────────────────
try:
    from QC.models import Deviation  # adjust if model lives in this app
except Exception:
    Deviation = None

def _count_process_deviations(on_date: _date) -> int:
    try:
        return Deviation.objects.filter(date=on_date).count() if Deviation else 0
    except Exception:
        return 0

# ─── FG inspections (deduped; one definition only) ──────────────────────────
def count_finished_goods_done(on_date: _date) -> int:
    """
    'Done' = status == 'qc_completed', date = COALESCE(release_by_qc_at, entry_date).
    Only FG (not SFG).
    """
    if QCEntry is None:
        return 0
    tz = timezone.get_current_timezone()
    start = datetime.combine(on_date, _time.min).replace(tzinfo=tz)
    end   = datetime.combine(on_date, _time.max).replace(tzinfo=tz)
    done_ts = Case(
        When(release_by_qc_at__isnull=False, then=F("release_by_qc_at")),
        default=F("entry_date"),
        output_field=DateTimeField(),
    )
    return (
        QCEntry.objects.annotate(done_at=done_ts)
        .filter(ar_type="FG", status="qc_completed", done_at__range=(start, end))
        .distinct()
        .count()
    )

# ─── Header initial fill (keeps your earlier behavior; safe defaults) ────────
def fetch_other_details(on_date: _date) -> dict:
    try:
        report = DailyQAReport.objects.filter(report_date=on_date).first()
        return {
            "analytical_downtime_hrs":    (getattr(report, "analytical_downtime_hrs", 0) or 0) if report else 0,
            "customer_complaints":        (getattr(report, "customer_complaints", 0) or 0) if report else 0,
            "analytical_mistakes":        (getattr(report, "analytical_mistakes", 0) or 0) if report else 0,
            "finished_goods_inspections": count_finished_goods_done(on_date),
            "process_deviations":         _count_process_deviations(on_date),
        }
    except Exception:
        return {
            "analytical_downtime_hrs":    0,
            "customer_complaints":        0,
            "analytical_mistakes":        0,
            "finished_goods_inspections": count_finished_goods_done(on_date),
            "process_deviations":         _count_process_deviations(on_date),
        }

# ─────────────────────────────────────────────────────────────────────────────
# API: Header metrics + Incoming
# ─────────────────────────────────────────────────────────────────────────────
@login_required
@require_GET
def api_fetch_other_header(request):
    try:
        sel = _datetime.strptime(request.GET.get("date", ""), "%Y-%m-%d").date()
    except Exception:
        return JsonResponse({"ok": False, "error": "Invalid date"}, status=400)

    data = {
        "analytical_downtime_hrs": str(_sum_analytical_downtime_hours(sel)),
        "customer_complaints": _count_customer_complaints(sel),
        "analytical_mistakes": _count_analytical_mistakes(sel),
        "first_aid_injury_incidents": _count_first_aid_injury_qc(sel),  # ← NEW
    }
    return JsonResponse({"ok": True, "data": data})

@login_required
@require_GET
def api_fetch_incoming(request):
    try:
        sel = _datetime.strptime(request.GET.get("date", ""), "%Y-%m-%d").date()
    except Exception:
        return JsonResponse({"ok": False, "error": "Invalid date"}, status=400)
    _ensure_grn_cache_for(sel)
    rows = _incoming_rows_from_cache(sel)
    return JsonResponse({"ok": True, "rows": rows})

@login_required
@require_GET
def api_fetch_pdl_samples(request):  # kept for backward compatibility
    try:
        d = _datetime.strptime(request.GET.get("date",""), "%Y-%m-%d").date()
    except Exception:
        return JsonResponse({"ok": False, "error": "Invalid date"}, status=400)
    rows = list(fetch_pdl_samples(d) or [])
    return JsonResponse({"ok": True, "rows": rows})

# --------------------------------------------------------------------------------------------
# FTR %
# --------------------------------------------------------------------------------------------
# --- FTR by Stage & Product ---------------------------------------------------
from datetime import datetime, time
from django.http import JsonResponse
from django.utils import timezone
from django.db.models import Q

# from products.models import QCEntry, SpecEntry  # paths per your project

def _parse_iso_date(s: str):
    try:
        return datetime.strptime((s or "").strip(), "%Y-%m-%d").date()
    except Exception:
        return None

@login_required
@require_GET
def api_ftr_table(request):
    on = _parse_iso_date(request.GET.get("date"))
    if not on:
        return JsonResponse({"ok": False, "error": "Invalid date"}, status=400)

    tz = timezone.get_current_timezone()
    start = datetime.combine(on, time.min).replace(tzinfo=tz)
    end   = datetime.combine(on, time.max).replace(tzinfo=tz)

    # Pull all completed entries on that day (based on release_by_qc_at if present; else entry_date)
    # We’ll do the “done_on_day” filter in Python to keep it robust across DBs.
    qs = (
        QCEntry.objects
        .select_related("product")
        .prefetch_related("values")
        .filter(status__in=["pending_qc", "qc_completed"])  # include entries that reached QC
        .exclude(batch_no__isnull=True).exclude(batch_no="")
    )

    rows_by_pb = {}  # key: (product_name, stage, batch_no) -> {"fail": bool}
    for e in qs:
        done_at = e.release_by_qc_at or e.entry_date
        if not (start <= (done_at or start) <= end):
            continue

        key = ((e.product.name if e.product_id else "").strip(),
               (e.stage or "").strip(),
               (e.batch_no or "").strip())
        if not key[0] or not key[1] or not key[2]:
            continue

        rec = rows_by_pb.setdefault(key, {"fail": False})
        # decision-based fail
        if (e.decision_status or "").lower() in {"rejected", "fail"}:
            rec["fail"] = True

        # spec-based fail (any remark 'Fail')
        if not rec["fail"]:
            for v in e.values.all():
                if (v.remark or "").strip().lower() == "fail":
                    rec["fail"] = True
                    break

    # Aggregate to (product, stage)
    agg = {}  # (product, stage) -> {"tot": int, "fail": int}
    for (prod, stage, _batch), rec in rows_by_pb.items():
        k = (prod, stage)
        a = agg.setdefault(k, {"tot": 0, "fail": 0})
        a["tot"]  += 1
        a["fail"] += 1 if rec["fail"] else 0

    out_rows = []
    for (prod, stage), a in sorted(agg.items()):
        tot = a["tot"] or 0
        fail = a["fail"] or 0
        ftr = (fail / tot * 100.0) if tot else 0.0
        out_rows.append({
            "product": prod,
            "stage": stage,
            "total_batches": tot,
            "fail_batches": fail,
            "ftr_percent": round(ftr, 2),
        })

    return JsonResponse({"ok": True, "rows": out_rows})

# ─────────────────────────────────────────────────────────────────────────────
# Daily QA Report – create / list / detail (unchanged flow)
# ─────────────────────────────────────────────────────────────────────────────


@login_required
@transaction.atomic
def daily_report_create(request):
    header_date = _parse_daily_date(request.GET.get("date"))
    incoming_prefill_date = _parse_daily_date(
        request.GET.get("incoming_date") or request.GET.get("date")
    )

    if request.method == "POST":
        form = DailyQAReportForm(request.POST)
        try:
            incoming_fs = IncomingFS(request.POST, prefix="inc")
        except Exception as e:
            logger.error("IncomingFS bind failed: %s", e)
            incoming_fs = IncomingFS(prefix="inc")

        if form.is_valid() and incoming_fs.is_valid():
            report = form.save(commit=False)
            report.created_by = request.user
            report.save()
            incoming_fs.instance = report
            incoming_fs.save()
            messages.success(request, "Daily QA Report saved successfully.")
            return redirect("qc:daily_report_detail", pk=report.pk)

    else:
        # ───────────────────────────────────────────────
        # Header defaults + safe derived metrics
        # ───────────────────────────────────────────────
        initial_other = {}
        try:
            raw = fetch_other_details(header_date) or {}
            if isinstance(raw, dict):
                initial_other = raw
        except Exception as e:
            logger.warning("fetch_other_details failed: %s", e)

        derived = _safe_header_metrics(header_date)
        form = DailyQAReportForm(
            initial={
                "report_date": header_date,
                **initial_other,
                "analytical_downtime_hrs": derived.get("analytical_downtime_hrs", 0),
                "customer_complaints": derived.get("customer_complaints", 0),
                "analytical_mistakes": derived.get("analytical_mistakes", 0),
                # if you have this field on the model/form
                "incident_first_aid_injury": derived.get(
                    "incident_first_aid_injury", 0
                ),
            }
        )

        # ───────────────────────────────────────────────
        # Prefill Incoming: D-1 cache → fallback → service
        # ───────────────────────────────────────────────
        incoming_rows = []
        try:
            d1 = incoming_prefill_date - timedelta(days=1)
            try:
                tz = timezone.get_current_timezone()
                start = datetime.combine(d1, time.min).replace(tzinfo=tz)
                end = datetime.combine(d1, time.max).replace(tzinfo=tz)
                _ensure_grn_cache_for(d1)
            except Exception as e:
                logger.warning("ensure_grn_cache_for(%s) failed: %s", d1, e)

            rows_cache = list(_incoming_rows_from_cache(d1)) or []
            incoming_rows = rows_cache or list(
                fetch_incoming_rm_pm(incoming_prefill_date) or []
            )
        except Exception as e:
            logger.error("Prefill incoming failed: %s", e)

        try:
            incoming_fs = IncomingFS(prefix="inc")
            incoming_fs.extra = max(1, len(incoming_rows))
            incoming_fs.initial = incoming_rows or [{}]
        except Exception as e:
            logger.error("IncomingFS init failed: %s", e)
            incoming_fs = IncomingFS(prefix="inc")

    # ───────────────────────────────────────────────
    # Extra tables for this date
    # ───────────────────────────────────────────────
    fg_qc_rows = FGProductQCStatus.objects.filter(date=header_date).order_by("product")
    instrument_rows = InstrumentOccupancy.objects.filter(
        date=header_date
    ).order_by("make", "model")

    # ───────────────────────────────────────────────
    # Render the form
    # ───────────────────────────────────────────────
    return render(
        request,
        "quality/daily_report_form.html",
        {
            "form": form,
            "incoming_fs": incoming_fs,
            "mode": "create",
            "incoming_prefill_date": incoming_prefill_date,
            "fg_qc_rows": fg_qc_rows,
            "instrument_rows": instrument_rows,
        },
    )




@login_required
def daily_report_detail(request, pk: int):
    report = get_object_or_404(
        DailyQAReport.objects.prefetch_related("incoming", "pdl_samples"),
        pk=pk
    )
    return render(request, "quality/daily_report_detail.html", {"report": report})

# ─────────────────────────────────────────────────────────────────────────────
# PDL Entries (existing) – unchanged behavior
# ─────────────────────────────────────────────────────────────────────────────
@login_required
@transaction.atomic
def pdl_entries(request):
    sel_date = _parse_date(request.GET.get("date") or request.POST.get("report_date"))
    report, _ = DailyQAReport.objects.get_or_create(
        report_date=sel_date, defaults={"created_by": request.user}
    )

    if request.method == "POST":
        header_form = PDLHeaderForm(request.POST)
        formset = PDLOnlyFS(request.POST, instance=report, prefix="pdl")
        if header_form.is_valid() and formset.is_valid():
            new_date = header_form.cleaned_data["report_date"]
            if new_date != report.report_date:
                report, _ = DailyQAReport.objects.get_or_create(
                    report_date=new_date, defaults={"created_by": request.user}
                )
                formset.instance = report
            formset.save()
            messages.success(request, f"PDL entries saved for {report.report_date:%d-%m-%Y}.")
            from django.urls import reverse
            return redirect(f"{reverse('qc:pdl_entries')}?date={report.report_date:%Y-%m-%d}")

        return render(request, "quality/pdl_entries.html", {
            "mode": "edit",
            "header_form": header_form,
            "form": header_form,
            "formset": formset,
        })

    header_form = PDLHeaderForm(initial={"report_date": sel_date})
    PDLOnlyFS1 = type("PDLOnlyFS1", (PDLOnlyFS,), {"extra": 1})
    child_model = PDLOnlyFS.model
    formset = PDLOnlyFS1(instance=report, prefix="pdl", queryset=child_model.objects.none())

    return render(request, "quality/pdl_entries.html", {
        "mode": "edit",
        "header_form": header_form,
        "form": header_form,
        "formset": formset,
    })

# ─────────────────────────────────────────────────────────────────────────────
# NEW: PDL List (read-only/report) – additive, does not affect entries page
# ─────────────────────────────────────────────────────────────────────────────
from django.utils.dateparse import parse_date
from django.db.models import Q

@login_required
def pdl_list(request):
    """
    Filters:
      ?date=YYYY-MM-DD  (report date), ?result=..., ?q=...
    Export CSV: add ?download=csv
    """
    date_str = (request.GET.get("date") or "").strip()
    q        = (request.GET.get("q") or "").strip()
    result   = (request.GET.get("result") or "").strip()

    qs = PDLSample.objects.select_related("report").all()

    d = parse_date(date_str)
    if d:
        qs = qs.filter(report__report_date=d)
    if q:
        qs = qs.filter(Q(stage__icontains=q) | Q(sample_name__icontains=q) | Q(remarks__icontains=q))
    if result:
        qs = qs.filter(result=result)
    qs = qs.order_by("report__report_date", "stage", "sample_name", "id")

    if request.GET.get("download") == "csv":
        import csv
        resp = HttpResponse(content_type="text/csv; charset=utf-8")
        resp["Content-Disposition"] = f'attachment; filename="PDL_List_{(date_str or "all")}.csv"'
        w = csv.writer(resp)
        w.writerow(["Date", "Stage", "Sample Name", "Result", "Remarks"])
        for r in qs.values("report__report_date","stage","sample_name","result","remarks"):
            w.writerow([r["report__report_date"] or "", r["stage"] or "", r["sample_name"] or "", r["result"] or "", r["remarks"] or ""])
        return resp

    page_obj = Paginator(qs, 25).get_page(request.GET.get("page"))
    try:
        result_choices = list(dict(PDLSample._meta.get_field("result").choices).keys())
    except Exception:
        result_choices = ["Pass", "Fail", "N/A"]

    return render(request, "quality/pdl_list.html", {
        "page_obj": page_obj,
        "rows": page_obj.object_list,
        "date": date_str,
        "q": q,
        "result": result,
        "result_choices": result_choices,
    })

# -------------------------------------------------------------------------------------
#                   CustomerComplaint
# -------------------------------------------------------------------------------------

# views.py
from django.shortcuts import render, redirect, get_object_or_404
from .forms import CustomerComplaintForm
from .models import CustomerComplaint
from QC.models import AlfaProductMaster  # <-- import master

def _norm_key(s: str) -> str:
    # collapse extra spaces and lowercase (handles “2 ,4,6 …”, etc.)
    s = (s or "").strip()
    return " ".join(s.split()).lower()

def complaint_list(request):
    # base queryset (objects; we’ll attach a display field)
    qs = CustomerComplaint.objects.all().order_by("-complaint_date")

    # build a normalized alfa -> finished map (only active rows)
    alfa_map_raw = AlfaProductMaster.objects.filter(is_active=True)\
        .values_list("alfa_name", "finished_product_name")
    alfa_map = {_norm_key(a): f for a, f in alfa_map_raw}

    # attach a computed finished name for display if DB field is empty/wrong
    for c in qs:
        saved = (c.finished_product_name or "").strip()
        derived = alfa_map.get(_norm_key(c.product_name))
        # prefer saved when present, else fallback to derived from master
        c.finished_display = saved or (derived or "")

    return render(
        request,
        "QC/complaint_list.html",
        {"complaints": qs}
    )

def complaint_create(request):
    if request.method == "POST":
        form = CustomerComplaintForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("qc:complaint_list")
    else:
        form = CustomerComplaintForm()
    ctx = {"form": form, "edit_mode": False}
    ctx.update(_alfa_form_context())
    return render(request, "QC/complaint_form.html", ctx)

def complaint_update(request, pk):
    complaint = get_object_or_404(CustomerComplaint, pk=pk)
    if request.method == "POST":
        form = CustomerComplaintForm(request.POST, instance=complaint)
        if form.is_valid():
            form.save()
            return redirect("qc:complaint_list")
    else:
        form = CustomerComplaintForm(instance=complaint)
    ctx = {"form": form, "edit_mode": True}
    ctx.update(_alfa_form_context())
    return render(request, "QC/complaint_form.html", ctx)

# -----------------------------------------------------------------------------------------------
# Excel Dowload Complaint
# ----------------------------------------------------------------------------------------------

import io
import calendar
import pandas as pd
from django.http import HttpResponse
from django.shortcuts import render
from django.db.models import Count
from .models import CustomerComplaint
from QC.models import AlfaProductMaster  # <-- needed to derive finished product


def _norm_key(s: str) -> str:
    s = (s or "").strip()
    return " ".join(s.split()).lower()


def complaint_report_view(request):
    qs = CustomerComplaint.objects.all().order_by("-complaint_date")

    # -------- Optional filters ----------
    start = request.GET.get("start")
    end = request.GET.get("end")
    status = request.GET.get("status")
    product = request.GET.get("product")

    if start:
        qs = qs.filter(complaint_date__gte=start)
    if end:
        qs = qs.filter(complaint_date__lte=end)
    if status:
        qs = qs.filter(status=status)
    if product:
        qs = qs.filter(product_name__icontains=product)

    # -------- Build alfa->finished map (for display & export backfill) ----------
    alfa_map_raw = AlfaProductMaster.objects.filter(is_active=True)\
        .values_list("alfa_name", "finished_product_name")
    alfa_map = {_norm_key(a): f for a, f in alfa_map_raw}

    # Attach a computed finished product for display
    # Prefer saved finished_product_name; fallback to derived from alfa map
    for c in qs:
        saved = (getattr(c, "finished_product_name", "") or "").strip()
        derived = alfa_map.get(_norm_key(c.product_name))
        c.finished_display = saved or (derived or "")

    # -------- Aggregates for on-page charts ----------
    status_counts = list(qs.values("status").annotate(total=Count("id")))
    type_counts = list(qs.values("complaint_type").annotate(total=Count("id")))

    # -------- Excel download ----------
    if "download" in request.GET:
        # We’ll export both Alfa and Finished columns
        cols = [
            "complaint_date", "complaint_no",
            "product_name", "finished_product_name",  # include finished
            "customer_name",
            "nature_of_complaint", "complaint_type",
            "investigation", "corrective_action", "preventive_action", "status",
        ]

        # Build rows with a backfilled finished_product_name
        rows = []
        for c in qs:
            finished = (getattr(c, "finished_product_name", "") or "").strip()
            if not finished:
                finished = alfa_map.get(_norm_key(c.product_name), "")
            rows.append({
                "complaint_date": c.complaint_date,
                "complaint_no": c.complaint_no,
                "product_name": c.product_name,
                "finished_product_name": finished,
                "customer_name": c.customer_name,
                "nature_of_complaint": c.nature_of_complaint,
                "complaint_type": c.complaint_type,
                "investigation": c.investigation,
                "corrective_action": c.corrective_action,
                "preventive_action": c.preventive_action,
                "status": c.status,
            })

        df = pd.DataFrame(rows, columns=cols)

        # Nice date formatting & safe empties
        if not df.empty:
            df["complaint_date"] = pd.to_datetime(df["complaint_date"], errors="coerce")
        else:
            df = pd.DataFrame(columns=cols)

        # Summaries (unchanged)
        status_summary = (
            df["status"].value_counts().rename_axis("Status").reset_index(name="Count")
            if not df.empty and "status" in df
            else pd.DataFrame(columns=["Status", "Count"])
        )
        monthly = (
            df.assign(year=df["complaint_date"].dt.year,
                      month=df["complaint_date"].dt.month)
              .dropna(subset=["complaint_date"])
              .groupby(["year", "month"]).size()
              .reset_index(name="Count")
              .sort_values(["year", "month"])
            if not df.empty else pd.DataFrame(columns=["year", "month", "Count"])
        )
        if not monthly.empty:
            monthly["Month"] = monthly.apply(
                lambda r: f"{calendar.month_abbr[int(r['month'])]} {int(r['year'])}",
                axis=1
            )
        type_summary = (
            df["complaint_type"].value_counts().rename_axis("Complaint Type").reset_index(name="Count")
            if not df.empty and "complaint_type" in df
            else pd.DataFrame(columns=["Complaint Type", "Count"])
        )

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="xlsxwriter", datetime_format="yyyy-mm-dd") as writer:
            # Raw data
            df.to_excel(writer, index=False, sheet_name="Complaints")
            wb = writer.book
            sh_data = writer.sheets["Complaints"]

            header_fmt = wb.add_format({"bold": True, "bg_color": "#DCE6F1", "border": 1})
            for c, name in enumerate(df.columns):
                sh_data.write(0, c, name, header_fmt)

            if not df.empty:
                last_row = len(df)
                last_col = len(df.columns) - 1
                sh_data.autofilter(0, 0, last_row, last_col)
            sh_data.freeze_panes(1, 0)

            for c, col in enumerate(df.columns):
                series = df[col].astype(str).fillna("")
                max_len = max([len(col)] + series.map(len).tolist()) if not df.empty else len(col)
                sh_data.set_column(c, c, min(max_len + 2, 50))

            # Summary sheet (unchanged from your version)
            sh_sum = wb.add_worksheet("Summary")
            row_ptr = 0
            sh_sum.write(row_ptr, 0, "Complaints by Status", header_fmt); row_ptr += 1
            if not status_summary.empty:
                sh_sum.write_row(row_ptr, 0, status_summary.columns.tolist())
                for i, r in status_summary.iterrows():
                    sh_sum.write_row(row_ptr + 1 + i, 0, r.tolist())
                chart1 = wb.add_chart({"type": "pie"})
                chart1.add_series({
                    "name": "Complaints by Status",
                    "categories": ["Summary", row_ptr + 1, 0, row_ptr + len(status_summary), 0],
                    "values":     ["Summary", row_ptr + 1, 1, row_ptr + len(status_summary), 1],
                })
                chart1.set_title({"name": "Complaints by Status"})
                chart1.set_style(10)
                sh_sum.insert_chart(row_ptr, 3, chart1)
                row_ptr += len(status_summary) + 3
            else:
                sh_sum.write(row_ptr, 0, "No data"); row_ptr += 2

            sh_sum.write(row_ptr, 0, "Complaints per Month", header_fmt); row_ptr += 1
            if not monthly.empty:
                cols_m = ["Month", "Count"]
                monthly_out = monthly[cols_m]
                sh_sum.write_row(row_ptr, 0, cols_m)
                for i, r in monthly_out.iterrows():
                    sh_sum.write_row(row_ptr + 1 + i, 0, r.tolist())
                chart2 = wb.add_chart({"type": "column"})
                chart2.add_series({
                    "name": "Complaints per Month",
                    "categories": ["Summary", row_ptr + 1, 0, row_ptr + len(monthly_out), 0],
                    "values":     ["Summary", row_ptr + 1, 1, row_ptr + len(monthly_out), 1],
                })
                chart2.set_title({"name": "Complaints per Month"})
                chart2.set_legend({"position": "none"})
                sh_sum.insert_chart(row_ptr, 3, chart2)
                row_ptr += len(monthly_out) + 3
            else:
                sh_sum.write(row_ptr, 0, "No monthly data"); row_ptr += 2

            sh_sum.write(row_ptr, 0, "Top Complaint Types", header_fmt); row_ptr += 1
            if not type_summary.empty:
                sh_sum.write_row(row_ptr, 0, type_summary.columns.tolist())
                for i, r in type_summary.iterrows():
                    sh_sum.write_row(row_ptr + 1 + i, 0, r.tolist())
                chart3 = wb.add_chart({"type": "bar"})
                chart3.add_series({
                    "name": "Top Complaint Types",
                    "categories": ["Summary", row_ptr + 1, 0, row_ptr + len(type_summary), 0],
                    "values":     ["Summary", row_ptr + 1, 1, row_ptr + len(type_summary), 1],
                })
                chart3.set_title({"name": "Top Complaint Types"})
                chart3.set_legend({"position": "none"})
                sh_sum.insert_chart(row_ptr, 3, chart3)
                row_ptr += len(type_summary) + 3
            else:
                sh_sum.write(row_ptr, 0, "No complaint type data"); row_ptr += 2

        output.seek(0)
        response = HttpResponse(
            output,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=Customer_Complaints_Report.xlsx"
        return response

    # Normal page render
    return render(request, "qc/complaint_report.html", {
        "complaints": qs,                  # qs now has c.finished_display attached
        "status_counts": status_counts,    # already converted to list
        "type_counts": type_counts,        # already converted to list
    })

def _norm_key(s: str) -> str:
    return " ".join((s or "").split()).lower()

@login_required
def complaint_detail(request, pk):
    """
    Read-only detail view for a single Customer Complaint.
    """
    obj = get_object_or_404(CustomerComplaint, pk=pk)

    # Derive finished name from Alfa master if empty (for older rows)
    alfa_map = {
        _norm_key(a): f
        for a, f in AlfaProductMaster.objects.filter(is_active=True)
        .values_list("alfa_name", "finished_product_name")
    }
    finished_display = (obj.finished_product_name or "").strip() \
        or alfa_map.get(_norm_key(obj.product_name), "")

    ctx = {
        "c": obj,
        "finished_display": finished_display,
    }
    return render(request, "QC/complaint_detail.html", ctx)

# ------------------------------------------------------------------------------------------------------
#                       AnalyticalDowntime
# ------------------------------------------------------------------------------------------------------

from django.contrib.auth.decorators import login_required, permission_required
from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Count
from django.db.models.functions import TruncDate
from django.http import HttpResponse, JsonResponse
import io
import pandas as pd

from .models import AnalyticalDowntime, LocalBOMDetail
from .forms import AnalyticalDowntimeForm


# ─────────────────────────────────────────────
# LIST VIEW
# ─────────────────────────────────────────────
@login_required
def downtime_list(request):
    qs = AnalyticalDowntime.objects.all().order_by("-start_at")

    # Quick filters
    status = request.GET.get("status")
    instrument = request.GET.get("instrument")
    stage = request.GET.get("stage")

    if status:
        qs = qs.filter(status=status)
    if instrument:
        qs = qs.filter(instrument_id__icontains=instrument)
    if stage:
        qs = qs.filter(stage__icontains=stage)

    return render(request, "qc/downtime_list.html", {"rows": qs})


# ─────────────────────────────────────────────
# CREATE
# ─────────────────────────────────────────────
@login_required
def downtime_create(request):
    if request.method == "POST":
        form = AnalyticalDowntimeForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("qc:downtime_list")
    else:
        form = AnalyticalDowntimeForm()

    return render(
        request,
        "qc/downtime_form.html",
        {"form": form, "edit_mode": False, "stage_options": getattr(form, "stage_options", [])},
    )

# ─────────────────────────────────────────────
# UPDATE
# ─────────────────────────────────────────────
@login_required
def downtime_update(request, pk):
    obj = get_object_or_404(AnalyticalDowntime, pk=pk)
    if request.method == "POST":
        form = AnalyticalDowntimeForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            return redirect("qc:downtime_list")
    else:
        form = AnalyticalDowntimeForm(instance=obj)

    return render(
        request,
        "qc/downtime_form.html",
        {"form": form, "edit_mode": True, "stage_options": getattr(form, "stage_options", [])},
    )

# ─────────────────────────────────────────────
# REPORT / DASHBOARD
# ─────────────────────────────────────────────
from django.db.models import Count, Sum, F, ExpressionWrapper, DurationField
from django.db.models.functions import TruncDate, Coalesce, Now

@login_required
def downtime_report(request):
    qs = AnalyticalDowntime.objects.all()

    # ---- Filters ------------------------------------------------------------
    start  = request.GET.get("start")
    end    = request.GET.get("end")
    status = request.GET.get("status")
    inst   = request.GET.get("instrument")
    stage  = request.GET.get("stage")

    if start:
        qs = qs.filter(start_at__date__gte=start)
    if end:
        qs = qs.filter(start_at__date__lte=end)
    if status:
        qs = qs.filter(status=status)
    if inst:
        qs = qs.filter(instrument_id__icontains=inst)
    if stage:
        qs = qs.filter(stage__icontains=stage)

    # Table ordering (separate from aggregates)
    table_qs = qs.order_by("-start_at")

    # Clear order_by before GROUP BY / aggregates (SQL Server requirement)
    base = qs.order_by()

    # ---- Aggregations used by the dashboard --------------------------------
    # status buckets the template renders as totals.open / .progress / .closed
    status_counts_qs = base.values("status").annotate(total=Count("id"))

    # normalize into a dict with all keys present
    status_map = {"open": 0, "progress": 0, "closed": 0}
    for row in status_counts_qs:
        status_map[row["status"]] = row["total"]

    # total minutes across all incidents (open incidents use NOW() as end)
    dur_expr = ExpressionWrapper(
        Coalesce(F("end_at"), Now()) - F("start_at"),
        output_field=DurationField(),
    )
    total_duration = base.aggregate(total=Sum(dur_expr))["total"]
    # Convert to minutes (guard None)
    total_minutes = int((total_duration.total_seconds() // 60) if total_duration else 0)

    # categories list for the pie chart
    by_cat = list(
        base.values("category")
            .annotate(total=Count("id"))
            .order_by("category")
    )

    # daily counts (if you need them elsewhere)
    daily_counts = list(
        base.annotate(day=TruncDate("start_at"))
            .values("day")
            .annotate(total=Count("id"))
            .order_by("day")
    )

    # ---- Excel download -----------------------------------------------------
    if "download" in request.GET:
        cols = [
            "incident_no", "instrument_id", "start_at", "end_at", "ongoing",
            "status", "category", "short_reason", "detail_reason",
            "stage", "product_name", "batch_no", "tests_delayed",
            "retest_due_date", "resolved_by", "remarks", "created_at", "updated_at",
        ]
        df = pd.DataFrame(list(base.values(*cols)))  # use base (filtered, no ORDER BY)
        if df.empty:
            df = pd.DataFrame(columns=cols)
        else:
            # nicer datetime text for Excel
            for c in ["start_at", "end_at", "retest_due_date", "created_at", "updated_at"]:
                if c in df.columns:
                    df[c] = pd.to_datetime(df[c]).dt.strftime("%Y-%m-%d %H:%M:%S").fillna("")

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
            df.to_excel(writer, index=False, sheet_name="Downtime Log")
            wb = writer.book
            ws = writer.sheets["Downtime Log"]

            header_fmt = wb.add_format({"bold": True, "bg_color": "#DCE6F1", "border": 1})
            for col_num, col_name in enumerate(df.columns):
                ws.write(0, col_num, col_name, header_fmt)
                ws.set_column(col_num, col_num, 22)

            # Summary sheet with pie chart (by status)
            summary = (
                df["status"].value_counts().rename_axis("Status").reset_index(name="Count")
                if not df.empty and "status" in df.columns else
                pd.DataFrame({"Status": ["Open", "In Progress", "Closed"], "Count": [0, 0, 0]})
            )
            summary.to_excel(writer, index=False, sheet_name="Summary")
            sum_ws = writer.sheets["Summary"]

            chart = wb.add_chart({"type": "pie"})
            # categories/values: ("sheet", first_row, first_col, last_row, last_col)
            chart.add_series({
                "name": "Downtime by Status",
                "categories": ["Summary", 1, 0, len(summary), 0],
                "values":     ["Summary", 1, 1, len(summary), 1],
            })
            chart.set_title({"name": "Downtime by Status"})
            chart.set_style(10)
            sum_ws.insert_chart("D2", chart)

        output.seek(0)
        resp = HttpResponse(
            output,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = "attachment; filename=Analytical_Downtime_Report.xlsx"
        return resp

    # ---- Render HTML report -------------------------------------------------
    return render(
        request,
        "qc/downtime_report.html",
        {
            "rows": table_qs,
            "totals": {
                "open": status_map["open"],
                "progress": status_map["progress"],
                "closed": status_map["closed"],
                "minutes": total_minutes,
            },
            "by_cat": by_cat,
            "daily_counts": daily_counts,
        },
    )

# ------------------------------------------------------------
# Excel Download
# ------------------------------------------------------------
from django.utils.timezone import now

def _filtered_downtime_qs(request):
    qs = AnalyticalDowntime.objects.all()
    start  = request.GET.get("start")
    end    = request.GET.get("end")
    status = request.GET.get("status")
    inst   = request.GET.get("instrument")
    stage  = request.GET.get("stage")

    if start:
        qs = qs.filter(start_at__date__gte=start)
    if end:
        qs = qs.filter(start_at__date__lte=end)
    if status:
        qs = qs.filter(status=status)
    if inst:
        qs = qs.filter(instrument_id__icontains=inst)
    if stage:
        qs = qs.filter(stage__icontains=stage)

    return qs

@login_required
def downtime_export_xlsx(request):
    """
    Standalone Excel download for Analytical Downtime
    (respects the same GET filters as the HTML report).
    """
    qs = _filtered_downtime_qs(request).order_by()  # SQL Server: clear ORDER BY for aggregates/values

    cols = [
        "incident_no", "instrument_id", "start_at", "end_at", "ongoing",
        "status", "category", "short_reason", "detail_reason",
        "stage", "product_name", "batch_no", "tests_delayed",
        "retest_due_date", "resolved_by", "remarks", "created_at", "updated_at",
    ]
    df = pd.DataFrame(list(qs.values(*cols)))
    if df.empty:
        df = pd.DataFrame(columns=cols)
    else:
        # Nice datetime strings
        for c in ["start_at", "end_at", "retest_due_date", "created_at", "updated_at"]:
            if c in df.columns:
                df[c] = pd.to_datetime(df[c]).dt.strftime("%Y-%m-%d %H:%M:%S").fillna("")

        # Add a computed duration (minutes) column
        if "start_at" in df.columns:
            # Using end_at or now() if blank
            def _dur_minutes(row):
                try:
                    sa = pd.to_datetime(row["start_at"])
                    ea_str = row.get("end_at")
                    ea = pd.to_datetime(ea_str) if ea_str else pd.Timestamp(now())
                    return max(int((ea - sa).total_seconds() // 60), 0)
                except Exception:
                    return 0
            df["duration_minutes"] = df.apply(_dur_minutes, axis=1)

    # Build workbook
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        df.to_excel(writer, index=False, sheet_name="Downtime Log")
        wb = writer.book
        ws = writer.sheets["Downtime Log"]

        header_fmt = wb.add_format({"bold": True, "bg_color": "#DCE6F1", "border": 1})
        for col_num, col_name in enumerate(df.columns):
            ws.write(0, col_num, col_name, header_fmt)
            ws.set_column(col_num, col_num, 22)

        # Summary sheet (by Status)
        if not df.empty and "status" in df.columns:
            summary = df["status"].value_counts().rename_axis("Status").reset_index(name="Count")
        else:
            summary = pd.DataFrame({"Status": ["open", "progress", "closed"], "Count": [0, 0, 0]})

        summary.to_excel(writer, index=False, sheet_name="Summary")
        sws = writer.sheets["Summary"]

        pie = wb.add_chart({"type": "pie"})
        pie.add_series({
            "name": "Downtime by Status",
            "categories": ["Summary", 1, 0, len(summary), 0],
            "values":     ["Summary", 1, 1, len(summary), 1],
        })
        pie.set_title({"name": "Downtime by Status"})
        pie.set_style(10)
        sws.insert_chart("D2", pie)

    output.seek(0)
    resp = HttpResponse(
        output,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = "attachment; filename=Analytical_Downtime_Report.xlsx"
    return resp

# ─────────────────────────────────────────────
# API: Fetch Product Name from Stage (AJAX)
# ─────────────────────────────────────────────
@login_required
def api_get_product_from_stage(request):
    """
    Given a Stage name, return the corresponding FG Product
    using the LocalBOMDetail table (ERP-synced).
    """
    stage = request.GET.get("stage", "").strip()
    if not stage:
        return JsonResponse({"ok": False, "error": "Missing stage"})

    try:
        product = (
            LocalBOMDetail.objects
            .filter(item_name__iexact=stage)
            .values_list("fg_name", flat=True)
            .first()
        )
        if not product:
            return JsonResponse({"ok": False, "error": "No matching product found"})
        return JsonResponse({"ok": True, "product_name": product})
    except Exception as e:
        return JsonResponse({"ok": False, "error": str(e)})

# views.py
from django.http import JsonResponse
from django.views.decorators.http import require_GET
from django.db.models.functions import Upper
from django.db.models import Q

from .models import LocalBOMDetail, BmrIssue


@require_GET
def api_product_and_batches(request):
    """
    GET ?stage=<stage name>
    → { ok, product_name, batches: [batch_no,...] }

    This version returns **all batches** from BmrIssue table,
    without filtering or limiting conditions.
    """
    stage_raw = (request.GET.get("stage") or "").strip()
    if not stage_raw:
        return JsonResponse({"ok": False, "error": "Missing stage"}, status=400)

    stage = " ".join(stage_raw.split())

    # ---- Resolve product (try exact, else fallback) ----
    product = (
        LocalBOMDetail.objects
        .filter(item_name__iexact=stage)
        .exclude(fg_name__isnull=True)
        .exclude(fg_name__exact="")
        .values_list("fg_name", flat=True)
        .first()
    )

    if not product:
        product = (
            LocalBOMDetail.objects
            .filter(item_name__icontains=stage)
            .exclude(fg_name__isnull=True)
            .exclude(fg_name__exact="")
            .values_list("fg_name", flat=True)
            .first()
        ) or ""

    # ---- Fetch ALL batches (no filters, no limits) ----
    batches: list[str] = []
    try:
        qs = (
            BmrIssue.objects
            .exclude(op_batch_no__isnull=True)
            .exclude(op_batch_no__exact="")
            .annotate(op_norm=Upper("op_batch_no"))
            .values_list("op_norm", flat=True)
            .order_by("op_norm")
        )
        batches = [b.strip() for b in qs if b and b.strip()]
    except Exception as e:
        return JsonResponse({"ok": False, "error": f"Error fetching batches: {e}"}, status=500)

    return JsonResponse({
        "ok": True,
        "product_name": product,
        "batches": batches,
        "total": len(batches)
    })

from dateutil.relativedelta import relativedelta
@require_GET
def api_all_batches(request):
    """
    Return distinct batches from BmrIssue, limited to the most recent N months.
    Query params:
      - months: int (default 3)
    Response:
      { ok: true, items: [{batch, stage, date}] }
    """
    try:
        months = int(request.GET.get("months", 3))
    except Exception:
        months = 3

    today = timezone.localdate()
    date_from = today - relativedelta(months=max(1, months))

    qs = (
        BmrIssue.objects
        .filter(bmr_issue_date__gte=date_from)
        .exclude(op_batch_no__isnull=True).exclude(op_batch_no__exact="")
        .annotate(_ub=Upper("op_batch_no"))
        .values("_ub", "bmr_issue_type")
        .annotate(last_date=Max("bmr_issue_date"))
        .order_by("-last_date", "_ub")
    )

    items = [
        {
            "batch": r["_ub"],
            "stage": r.get("bmr_issue_type") or "",
            "date":  (r["last_date"].strftime("%Y-%m-%d") if r["last_date"] else "")
        }
        for r in qs[:2000]  # sensible cap
    ]
    return JsonResponse({"ok": True, "items": items})

class QCInstrumentForm(forms.ModelForm):
    class Meta:
        model = QCInstrument
        fields = ["name", "code", "category", "is_active", "notes"]

@login_required
def instrument_master_list(request):
    qs = QCInstrument.objects.all().order_by("category", "name", "code")
    return render(request, "qc/instrument_master_list.html", {"rows": qs})

@login_required
def instrument_master_create(request):
    form = QCInstrumentForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        form.save()
        return redirect("qc:instrument_master")
    return render(request, "qc/instrument_master_form.html", {"form": form})

# -----------------------------------------------
# Deviation
# -----------------------------------------------
# views.py
import io
import json
import pandas as pd
from django.contrib.auth.decorators import login_required, permission_required
from django.db.models import Q, Count
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render

from .models import Deviation, AlfaProductMaster
from .forms import DeviationForm


def _alfa_form_context():
    """
    Provide Alfa Product master helpers for the form template:
    - alfa_options: list of Alfa names for the <datalist>
    - alfa_map_json: {"ALFA": "FINISHED PRODUCT"} for auto-fill via JS
    """
    alfa_options = list(
        AlfaProductMaster.objects.values_list("alfa_name", flat=True).order_by("alfa_name")
    )
    alfa_map = dict(
        AlfaProductMaster.objects.values_list("alfa_name", "finished_product_name")
    )
    return {"alfa_options": alfa_options, "alfa_map_json": json.dumps(alfa_map)}


def _norm_key(s: str) -> str:
    """collapse extra spaces and lowercase for loose matching"""
    s = (s or "").strip()
    return " ".join(s.split()).lower()


@login_required
def deviation_list(request):
    qs = Deviation.objects.all()

    q = (request.GET.get("q") or "").strip()
    status = (request.GET.get("status") or "").strip()

    if q:
        qs = qs.filter(
            Q(product__icontains=q)
            | Q(batch_no__icontains=q)
            | Q(deviation_no__icontains=q)
        )
    if status:
        qs = qs.filter(status=status)

    qs = qs.order_by("-date", "-created_at")

    # ---- Derive finished product for display if it's missing on old rows ----
    # Build a normalized map from the Alfa master once
    alfa_map = {
        _norm_key(a): f
        for a, f in AlfaProductMaster.objects.filter(is_active=True)
        .values_list("alfa_name", "finished_product_name")
    }

    rows = []
    for r in qs:
        saved = getattr(r, "finished_product", "") or ""
        derived = alfa_map.get(_norm_key(r.product))
        # prefer saved value if present; else fallback to derived
        r.finished_display = saved.strip() or (derived or "")
        rows.append(r)

    return render(
        request,
        "qc/deviation_list.html",
        {
            "rows": rows,
            "q": q,
            "status": status,
        },
    )

@login_required
def deviation_create(request):
    if request.method == "POST":
        form = DeviationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("qc:deviation_list")
    else:
        form = DeviationForm()

    ctx = {"form": form, "edit_mode": False}
    ctx.update(_alfa_form_context())
    return render(request, "qc/deviation_form.html", ctx)


@login_required
def deviation_update(request, pk):
    obj = get_object_or_404(Deviation, pk=pk)

    if request.method == "POST":
        form = DeviationForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            return redirect("qc:deviation_list")
    else:
        form = DeviationForm(instance=obj)

    ctx = {"form": form, "edit_mode": True, "obj": obj}
    ctx.update(_alfa_form_context())
    return render(request, "qc/deviation_form.html", ctx)


@login_required
def deviation_report(request):
    base = Deviation.objects.all()

    start = request.GET.get("start")
    end = request.GET.get("end")
    status = request.GET.get("status")

    if start:
        base = base.filter(date__gte=start)
    if end:
        base = base.filter(date__lte=end)
    if status:
        base = base.filter(status=status)

    table = base.order_by("-date", "-created_at")

    # SQL Server-safe aggregation (clear ordering)
    agg_qs = base.order_by()
    by_status = list(
        agg_qs.values("status").annotate(total=Count("id")).order_by("status")
    )

    return render(
        request,
        "qc/deviation_report.html",
        {"rows": table, "by_status": by_status},
    )


@login_required
def deviation_export_xlsx(request):
    base = Deviation.objects.all()

    start = request.GET.get("start")
    end = request.GET.get("end")
    status = request.GET.get("status")

    if start:
        base = base.filter(date__gte=start)
    if end:
        base = base.filter(date__lte=end)
    if status:
        base = base.filter(status=status)

    cols = [
        "deviation_no",
        "date",
        "product",
        "plant",
        "batch_no",
        "description",
        "root_cause",
        "corrective_action",
        "preventive_action",
        "status",
        "created_at",
        "updated_at",
    ]

    df = pd.DataFrame(list(base.order_by().values(*cols)))
    if not df.empty:
        for c in ["date", "created_at", "updated_at"]:
            if c in df.columns:
                df[c] = pd.to_datetime(df[c]).dt.strftime("%Y-%m-%d %H:%M:%S")
    else:
        df = pd.DataFrame(columns=cols)

    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="xlsxwriter") as w:
        df.to_excel(w, index=False, sheet_name="Deviations")
        wb = w.book
        ws = w.sheets["Deviations"]

        header_fmt = wb.add_format({"bold": True, "bg_color": "#DCE6F1", "border": 1})
        for i, c in enumerate(df.columns):
            ws.write(0, i, c, header_fmt)
            ws.set_column(i, i, 24)

        # Simple summary tab
        if "status" in df.columns and not df.empty:
            summary = df["status"].value_counts().rename_axis("Status").reset_index(name="Count")
        else:
            summary = pd.DataFrame({"Status": ["open", "progress", "closed"], "Count": [0, 0, 0]})

        summary.to_excel(w, index=False, sheet_name="Summary")
        sws = w.sheets["Summary"]
        pie = wb.add_chart({"type": "pie"})
        pie.add_series({
            "name": "Deviations by Status",
            "categories": ["Summary", 1, 0, len(summary), 0],
            "values":     ["Summary", 1, 1, len(summary), 1],
        })
        pie.set_title({"name": "Deviations by Status"})
        pie.set_style(10)
        sws.insert_chart("E2", pie)

    out.seek(0)
    resp = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = "attachment; filename=Deviation_Report.xlsx"
    return resp

@login_required
def deviation_detail(request, pk):
    """
    Read-only view for documentation/customer users to see all deviation details.
    """
    entry = get_object_or_404(Deviation, pk=pk)

    # Enrich with derived finished product for display (if missing)
    alfa_map = dict(
        AlfaProductMaster.objects.filter(is_active=True)
        .values_list("alfa_name", "finished_product_name")
    )
    finished_display = entry.finished_product or alfa_map.get(entry.product, "")

    ctx = {
        "entry": entry,
        "finished_display": finished_display,
    }
    return render(request, "QC/deviation_detail.html", ctx)

@login_required
def api_alfa_finished(request):
    alfa = (request.GET.get("alfa") or "").strip()
    if not alfa:
        return JsonResponse({"ok": False, "error": "Missing alfa"})
    rec = (AlfaProductMaster.objects
           .filter(alfa_name__iexact=alfa, is_active=True)
           .values("finished_product_name")
           .first())
    return JsonResponse({"ok": bool(rec), "finished": (rec or {}).get("finished_product_name", "")})

def _query_batches(qs_base, q):
    # 1) exact finished, 2) exact alfa, 3) contains either
    qs = qs_base.filter(product_name__iexact=q)
    if not qs.exists():
        qs = qs_base.filter(fg_name__iexact=q)
    if not qs.exists():
        qs = qs_base.filter(Q(product_name__icontains=q) | Q(fg_name__icontains=q))
    return qs

@login_required
def api_batches_for_product(request):
    """
    Return batches grouped by (batch, stage) with latest date.
    Windows: last 30d → last 180d → all-time (first hit wins).
    """
    q = (request.GET.get("product") or "").strip()
    if not q:
        return JsonResponse({"ok": False, "items": []})

    today = timezone.localdate()
    windows = [30, 180, None]  # None = no date limit
    items = []

    for days in windows:
        base = BmrIssue.objects.all()
        if days:
            base = base.filter(bmr_issue_date__gte=today - timedelta(days=days))

        qs = _query_batches(base, q)
        if not qs.exists():
            continue

        qs = qs.exclude(op_batch_no__isnull=True).exclude(op_batch_no="")
        grouped = (
            qs.annotate(_ub=Upper("op_batch_no"))
              .values("_ub", "bmr_issue_type")
              .annotate(last_date=Max("bmr_issue_date"))
              .order_by("-last_date", "_ub", "bmr_issue_type")
        )
        items = [
            {
                "batch": g["_ub"],
                "stage": g["bmr_issue_type"] or "Unknown",
                "date":  g["last_date"].strftime("%Y-%m-%d") if g["last_date"] else ""
            }
            for g in grouped[:200]
        ]
        break  # stop at the first window that returns data

    return JsonResponse({"ok": True, "items": items})

# ----------------------------------------------------------------------------
# AnalyticalMistake
# ----------------------------------------------------------------------------
from .forms import AnalyticalMistakeForm

def _alfa_form_context():
    alfa_options = list(
        AlfaProductMaster.objects.values_list("alfa_name", flat=True).order_by("alfa_name")
    )
    alfa_map = dict(
        AlfaProductMaster.objects.values_list("alfa_name", "finished_product_name")
    )
    return {"alfa_options": alfa_options, "alfa_map_json": json.dumps(alfa_map)}

def _norm_key(s: str) -> str:
    return " ".join((s or "").split()).lower()

# ---------- List ----------
@login_required
def analytical_mistake_list(request):
    qs = AnalyticalMistake.objects.all().order_by("-date", "-created_at")

    # derive finished for display if missing (for old rows)
    alfa_map = {
        _norm_key(a): f
        for a, f in AlfaProductMaster.objects.filter(is_active=True)
        .values_list("alfa_name", "finished_product_name")
    }
    for r in qs:
        saved = (r.finished_product or "").strip()
        r.finished_display = saved or (alfa_map.get(_norm_key(r.product)) or "")

    return render(request, "QC/analytical_mistake_list.html", {"rows": qs})

# ---------- Create ----------
@login_required
def analytical_mistake_create(request):
    if request.method == "POST":
        form = AnalyticalMistakeForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("qc:analytical_mistake_list")
    else:
        form = AnalyticalMistakeForm()
    ctx = {"form": form, "edit_mode": False}
    ctx.update(_alfa_form_context())
    return render(request, "QC/analytical_mistake_form.html", ctx)

# ---------- Update ----------
@login_required
def analytical_mistake_update(request, pk):
    obj = get_object_or_404(AnalyticalMistake, pk=pk)
    if request.method == "POST":
        form = AnalyticalMistakeForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            return redirect("qc:analytical_mistake_list")
    else:
        form = AnalyticalMistakeForm(instance=obj)
    ctx = {"form": form, "edit_mode": True, "obj": obj}
    ctx.update(_alfa_form_context())
    return render(request, "QC/analytical_mistake_form.html", ctx)

# ---------- Read-only Detail ----------
@login_required
def analytical_mistake_detail(request, pk):
    r = get_object_or_404(AnalyticalMistake, pk=pk)
    alfa_map = dict(
        AlfaProductMaster.objects.filter(is_active=True)
        .values_list("alfa_name", "finished_product_name")
    )
    finished_display = r.finished_product or alfa_map.get(r.product, "")
    return render(request, "QC/analytical_mistake_detail.html",
                  {"entry": r, "finished_display": finished_display})

# ------------------------------------------------
# Excel Report with Visulization
# ------------------------------------------------
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Count
from django.utils.dateparse import parse_date

from .models import AnalyticalMistake

@login_required
def analytical_mistake_report(request):
    qs = AnalyticalMistake.objects.all()

    # ------- filters (robust) -------
    start   = (request.GET.get("start") or "").strip()
    end     = (request.GET.get("end") or "").strip()
    plant   = (request.GET.get("plant") or "").strip()
    product = (request.GET.get("product") or "").strip()   # Alfa or Finished (one box)

    # Dates (safe parsing; ignore if malformed)
    d = parse_date(start)
    if d:
        qs = qs.filter(date__gte=d)
    d = parse_date(end)
    if d:
        qs = qs.filter(date__lte=d)

    # Plant (exact from dropdown)
    if plant:
        qs = qs.filter(plant=plant)

    # Single product box filters BOTH alfa (product) and finished_product
    if product:
        qs = qs.filter(
            Q(product__icontains=product) |
            Q(finished_product__icontains=product)
        )

    qs = qs.order_by("-date", "-created_at")

    # ------- summaries for the page -------
    by_plant = list(
        qs.values("plant")
          .annotate(total=Count("id"))
          .order_by("plant")
    )

    # Group by Alfa product name; skip blank/null names
    by_product = list(
        qs.exclude(product__isnull=True)
          .exclude(product__exact="")
          .values("product")
          .annotate(total=Count("id"))
          .order_by("product")
    )

    # Monthly counts (YYYY, MM)
    by_month = list(
        qs.values("date__year", "date__month")
          .annotate(total=Count("id"))
          .order_by("date__year", "date__month")
    )

    return render(
        request,
        "QC/analytical_mistake_report.html",
        {
            "rows": qs,
            "by_plant": by_plant,
            "by_product": by_product,   # <-- use this in your template
            "by_month": by_month,
            "start": start, "end": end, "plant": plant, "product": product,
        },
    )


@login_required
def analytical_mistake_export_xlsx(request):
    qs = AnalyticalMistake.objects.all().order_by("-date", "-created_at")

    start   = request.GET.get("start")
    end     = request.GET.get("end")
    plant   = request.GET.get("plant")
    product = request.GET.get("product")
    fprod   = request.GET.get("finished")

    if start:   qs = qs.filter(date__gte=start)
    if end:     qs = qs.filter(date__lte=end)
    if plant:   qs = qs.filter(plant=plant)
    if product: qs = qs.filter(product__icontains=product)
    if fprod:   qs = qs.filter(finished_product__icontains=fprod)

    cols = [
        "am_no", "date", "product", "finished_product", "plant", "batch_no",
        "description", "root_cause", "corrective_action", "preventive_action",
        "created_at", "updated_at",
    ]
    df = pd.DataFrame(list(qs.order_by().values(*cols)))
    if not df.empty:
        for c in ["date", "created_at", "updated_at"]:
            if c in df.columns:
                df[c] = pd.to_datetime(df[c]).dt.strftime("%Y-%m-%d %H:%M:%S")
    else:
        df = pd.DataFrame(columns=cols)

    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="xlsxwriter") as w:
        # Sheet 1: Data
        df.to_excel(w, index=False, sheet_name="Analytical Mistakes")
        wb  = w.book
        ws  = w.sheets["Analytical Mistakes"]
        header = wb.add_format({"bold": True, "bg_color": "#DCE6F1", "border": 1})
        for i, c in enumerate(df.columns):
            ws.write(0, i, c, header)
            ws.set_column(i, i, 24)

        # Sheet 2: Summary
        sh = wb.add_worksheet("Summary")

        # by plant
        if "plant" in df and not df.empty:
            plant_summary = (
                df["plant"].fillna("").value_counts().rename_axis("Plant").reset_index(name="Count")
            )
        else:
            plant_summary = pd.DataFrame(columns=["Plant", "Count"])
        # by month
        if not df.empty:
            df["_date"] = pd.to_datetime(df["date"], errors="coerce")
            monthly = (df.dropna(subset=["_date"])
                         .assign(Month=lambda x: x["_date"].dt.to_period("M").astype(str))
                         .groupby("Month").size().reset_index(name="Count"))
        else:
            monthly = pd.DataFrame(columns=["Month", "Count"])

        # write plant table
        row = 0
        sh.write(row, 0, "By Plant", header); row += 1
        if not plant_summary.empty:
            sh.write_row(row, 0, plant_summary.columns.tolist()); row += 1
            for _, r in plant_summary.iterrows():
                sh.write_row(row, 0, r.tolist()); row += 1
            chart1 = wb.add_chart({"type": "pie"})
            chart1.add_series({
                "name": "By Plant",
                "categories": ["Summary", 2, 0, 1 + len(plant_summary), 0],
                "values":     ["Summary", 2, 1, 1 + len(plant_summary), 1],
            })
            chart1.set_title({"name": "Analytical Mistakes by Plant"})
            chart1.set_style(10)
            sh.insert_chart(1, 3, chart1)
        else:
            sh.write(row, 0, "No data"); row += 2

        # write monthly table
        sh.write(row, 0, "By Month", header); row += 1
        if not monthly.empty:
            sh.write_row(row, 0, monthly.columns.tolist()); row += 1
            for _, r in monthly.iterrows():
                sh.write_row(row, 0, r.tolist()); row += 1
            chart2 = wb.add_chart({"type": "column"})
            chart2.add_series({
                "name": "By Month",
                "categories": ["Summary", row - len(monthly), 0, row - 1, 0],
                "values":     ["Summary", row - len(monthly), 1, row - 1, 1],
            })
            chart2.set_title({"name": "Analytical Mistakes per Month"})
            chart2.set_legend({"position": "none"})
            sh.insert_chart(row - len(monthly) - 1, 3, chart2)
        else:
            sh.write(row, 0, "No monthly data"); row += 2

    out.seek(0)
    resp = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = "attachment; filename=Analytical_Mistake_Report.xlsx"
    return resp





# ----------------------------CALIBRATION SCHEDULE ----------------------------------------------

@login_required
def calibration_list(request):
    """List all calibration rows with basic filters + pagination."""
    # ---- Permission check ----
    if not request.user.has_perm("QC.view_qccalibrationschedule"):
        messages.error(request, "You do not have permission to view calibration schedule.")
        logger.warning("User '%s' tried to view QC Calibration list without permission.", request.user.username,)
        return redirect("indexpage")

    qs = QCCalibrationSchedule.objects.select_related("instrument")
    instrument_id = (request.GET.get("instrument_id") or "").strip()
    year = (request.GET.get("year") or "").strip()
    logger.info("User=%s accessed QC Calibration LIST | instrument_id='%s' year='%s'", request.user.username,instrument_id, year,)
    if instrument_id:
        qs = qs.filter(instrument__instument_id__icontains=instrument_id)
    if year:
        qs = qs.filter(schedule_year__icontains=year)

    qs = qs.order_by("instrument__instument_id", "calibration_date")

    # -------- Pagination --------
    page_number = request.GET.get("page") or 1
    paginator = Paginator(qs, 25)  # 25 rows per page (change if needed)

    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    # Build querystring for pagination links (keep filters, drop page)
    params = request.GET.copy()
    params.pop("page", None)
    base_qs = params.urlencode()  # e.g. "instrument_id=GC&year=2026-2027"

    logger.info("User=%s QC Calibration LIST result | total_rows=%s page=%s",request.user.username,paginator.count,
        page_obj.number,)
    context = {"calibrations": page_obj.object_list, "page_obj": page_obj,
        "paginator": paginator,"base_qs": base_qs,"filters": {"instrument_id": instrument_id, "year": year}, }
    return render(request, "qc/calibration/calibration_list.html", context)


@login_required
def calibration_create(request):
    # ---- Permission check ----
    if not request.user.has_perm("QC.add_qccalibrationschedule"):
        messages.error(request, "You do not have permission to add calibration schedule.")
        logger.warning("User '%s' tried to create QC Calibration without permission.",request.user.username, )
        return redirect("indexpage")

    if request.method == "POST":
        form = QCCalibrationScheduleForm(request.POST)

        if form.is_valid():
            obj = form.save()
            messages.success(request, "Calibration schedule created successfully.")
            logger.info("User=%s created QC Calibration | id=%s instrument=%s date=%s year=%s",
                request.user.username, obj.pk,getattr(obj.instrument, "instument_id", None), obj.calibration_date,
                obj.schedule_year,)
            return redirect("qc:calibration_list")

        logger.error(
            "QC Calibration CREATE form invalid | user=%s | errors=%s",
            request.user.username,
            form.errors.as_json(),
        )
        messages.error(request, "Please correct the errors below.")
    else:
        form = QCCalibrationScheduleForm()
        logger.info("User=%s opened QC Calibration CREATE form", request.user.username)

    return render(
        request,
        "qc/calibration/calibration_form.html",
        {"form": form, "mode": "create"},
    )


@login_required
def calibration_edit(request, pk):
    # ---- Permission check ----
    if not request.user.has_perm("QC.change_qccalibrationschedule"):
        messages.error(request, "You do not have permission to edit calibration schedule.")
        logger.warning("User '%s' tried to edit QC Calibration (pk=%s) without permission.",request.user.username, pk, )
        return redirect("indexpage")
    obj = get_object_or_404(QCCalibrationSchedule, pk=pk)
    if request.method == "POST":
        form = QCCalibrationScheduleForm(request.POST, instance=obj)
        if form.is_valid():
            obj = form.save()
            messages.success(request, "Calibration schedule updated successfully.")
            logger.info("User=%s updated QC Calibration | id=%s instrument=%s date=%s year=%s",
                request.user.username, obj.pk, getattr(obj.instrument, "instument_id", None),
                obj.calibration_date, obj.schedule_year, )
            return redirect("qc:calibration_list")

        logger.error("QC Calibration EDIT form invalid | user=%s | pk=%s | errors=%s",request.user.username,
            pk, form.errors.as_json(),)
        messages.error(request, "Please correct the errors below.")
    else:
        form = QCCalibrationScheduleForm(instance=obj)
        logger.info("User=%s opened QC Calibration EDIT form | pk=%s",request.user.username,pk,)
    return render(request,"qc/calibration/calibration_form.html",{"form": form, "mode": "edit", "obj": obj}, )



@login_required
@require_POST
def calibration_delete(request, pk):
    # ---- Permission check ----
    if not request.user.has_perm("QC.delete_qccalibrationschedule"):
        messages.error(request, "You do not have permission to delete calibration schedule.")
        logger.warning("User '%s' tried to delete QC Calibration (pk=%s) without permission.",request.user.username,pk,)
        return redirect("indexpage")  # or redirect("qc:calibration_list")
    obj = get_object_or_404(QCCalibrationSchedule, pk=pk)
    inst_id = getattr(getattr(obj, "instrument", None), "instument_id", None) or str(pk)
    cal_date = getattr(obj, "calibration_date", None)
    year = getattr(obj, "schedule_year", None)

    try:
        obj.delete()
        messages.success(request, f"Calibration schedule for {inst_id} deleted.")
        logger.info("User=%s deleted QC Calibration | pk=%s instrument_id=%s calibration_date=%s schedule_year=%s",
            request.user.username, pk, inst_id, cal_date,  year, )
    except Exception as e:
        messages.error(request, "An unexpected error occurred while deleting the calibration schedule.")
        logger.error( "Error deleting QC Calibration | user=%s pk=%s instrument_id=%s error=%s",request.user.username,
            pk,inst_id,  str(e),  exc_info=True, )
    return redirect("qc:calibration_list")



@login_required
def calibration_export_excel(request):
    import datetime
    """Download filtered QC Calibration Schedule as an Excel file."""
    # ---- Permission check ----
    if not request.user.has_perm("QC.view_qccalibrationschedule"):
        messages.error(request, "You do not have permission to export calibration schedule.")
        return redirect("indexpage")
    # ---- Apply same filters as list view ----
    qs = QCCalibrationSchedule.objects.select_related("instrument")

    instrument_id = (request.GET.get("instrument_id") or "").strip()
    year = (request.GET.get("year") or "").strip()

    if instrument_id:
        qs = qs.filter(instrument__instument_id__icontains=instrument_id)
    if year:
        qs = qs.filter(schedule_year__icontains=year)

    qs = qs.order_by("instrument__instument_id", "calibration_date")

    # ---- Build Excel in memory ----
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {"in_memory": True})
    worksheet = workbook.add_worksheet("Calibration")

    # Formats
    title_fmt = workbook.add_format(
        {
            "bold": True,
            "font_size": 14,
            "align": "center",
            "valign": "vcenter",
        }
    )
    header_fmt = workbook.add_format(
        {
            "bold": True,
            "bg_color": "#E5E7EB",  # light gray
            "border": 1,
            "align": "center",
            "valign": "vcenter",
        }
    )
    text_fmt = workbook.add_format({"border": 1})
    date_fmt = workbook.add_format({"num_format": "dd-mm-yyyy", "border": 1})

    # Column widths
    worksheet.set_column("A:A", 25)  # Instrument ID
    worksheet.set_column("B:B", 25)  # Instrument Name
    worksheet.set_column("C:C", 14)  # Schedule Year
    worksheet.set_column("D:F", 14)  # Dates
    worksheet.set_column("G:G", 40)  # Remarks

    # ---- Title row (merged) ----
    title = "QC Calibration Schedule"
    if year:
        title = f"{title} – {year}"

    worksheet.merge_range(0, 0, 0, 6, title, title_fmt)

    # ---- Header row ----
    headers = [
        "Instrument ID",
        "Instrument Name",
        "Schedule Year",
        "Calibration Date",
        "Calibration Due Date",
        "Reminder Date",
        "Remarks",
    ]

    header_row = 2
    for col, h in enumerate(headers):
        worksheet.write(header_row, col, h, header_fmt)
    # ---- Data rows ----
    row = header_row + 1
    for obj in qs:
        # A: Instrument ID
        worksheet.write(row, 0, obj.instrument.instument_id or "", text_fmt)
        # B: Instrument Name (from master)
        worksheet.write(row, 1, obj.instrument.name or "", text_fmt)
        # C: Schedule year
        worksheet.write(row, 2, obj.schedule_year or "", text_fmt)
        # Dates as real Excel dates
        for col_idx, d in enumerate(
            [
                obj.calibration_date,
                obj.calibration_due_date,
                obj.reminder_date,
            ],
            start=3,
        ):
            if d:
                dt = datetime.datetime.combine(d, datetime.time())
                worksheet.write_datetime(row, col_idx, dt, date_fmt)
            else:
                worksheet.write(row, col_idx, "", text_fmt)
        # G: Remarks
        worksheet.write(row, 6, obj.remarks or "", text_fmt)
        row += 1
    workbook.close()
    output.seek(0)

    # ---- HTTP response ----
    filename = "qc_calibration_schedule.xlsx"
    response = HttpResponse(
        output.read(),
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


#------------------Instrument related  code  -------------------------------------

@login_required
def instrument_list(request):
    """List QC instruments with simple filters + pagination."""
    if not request.user.has_perm("QC.view_qcinstrument"):
        messages.error(request, "You do not have permission to view instruments.")
        logger.warning(
            "User '%s' tried to view QCInstrument list without permission.",
            request.user.username,
        )
        return redirect("indexpage")

    qs = QCInstrument.objects.all().order_by("category", "instument_id", "code")

    search = (request.GET.get("q") or "").strip()
    status = (request.GET.get("status") or "").strip()  # '', 'active', 'inactive'

    if search:
        qs = qs.filter(
            # instrument id / name / code search
            instument_id__icontains=search
        ) | qs.filter(name__icontains=search) | qs.filter(code__icontains=search)

    if status == "active":
        qs = qs.filter(is_active=True)
    elif status == "inactive":
        qs = qs.filter(is_active=False)

    paginator = Paginator(qs, 25)
    page_number = request.GET.get("page") or 1
    page_obj = paginator.get_page(page_number)

    # keep filters in pagination links
    get_params = request.GET.copy()
    if "page" in get_params:
        get_params.pop("page")
    base_qs = get_params.urlencode()

    context = {
        "instruments": page_obj.object_list,
        "page_obj": page_obj,
        "paginator": paginator,
        "filters": {"q": search, "status": status},
        "base_qs": base_qs,
    }
    return render(request, "qc/instrument/instrument_list.html", context)



@login_required
def instrument_create(request):
    """Create a new QCInstrument."""
    if not request.user.has_perm("QC.add_qcinstrument"):
        messages.error(request, "You do not have permission to add instruments.")
        return redirect("qc:instrument_list")

    if request.method == "POST":
        form = QCInstrumentFormNew(request.POST)
        if form.is_valid():
            inst = form.save()
            messages.success(
                request,
                f"Instrument '{inst.instument_id or inst.name}' created successfully.",
            )
            logger.info(
                "User=%s created QCInstrument id=%s",
                request.user.username,
                inst.pk,
            )
            return redirect("qc:instrument_list")
    else:
        form = QCInstrumentFormNew()

    return render(
        request,
        "qc/instrument/instrument_form.html",
        {"form": form, "mode": "create"},
    )


@login_required
def instrument_edit(request, pk):
    """Edit an existing QCInstrument."""
    if not request.user.has_perm("QC.change_qcinstrument"):
        messages.error(request, "You do not have permission to edit instruments.")
        return redirect("qc:instrument_list")

    instrument = get_object_or_404(QCInstrument, pk=pk)

    if request.method == "POST":
        form = QCInstrumentFormNew(request.POST, instance=instrument)
        if form.is_valid():
            inst = form.save()
            messages.success(
                request,
                f"Instrument '{inst.instument_id or inst.name}' updated successfully.",
            )
            logger.info(
                "User=%s edited QCInstrument id=%s",
                request.user.username,
                inst.pk,
            )
            return redirect("qc:instrument_list")
    else:
        form = QCInstrumentFormNew(instance=instrument)

    return render(
        request,
        "qc/instrument/instrument_form.html",
        {"form": form, "mode": "edit", "instrument": instrument},
    )


@login_required
def instrument_delete(request, pk):
    """Soft delete (or hard delete) an instrument – used via small trash button."""
    if not request.user.has_perm("QC.delete_qcinstrument"):
        messages.error(request, "You do not have permission to delete instruments.")
        return redirect("qc:instrument_list")

    instrument = get_object_or_404(QCInstrument, pk=pk)

    if request.method == "POST":
        label = instrument.instument_id or instrument.name or f"ID {instrument.pk}"
        try:
            instrument.delete()
            messages.success(request, f"Instrument '{label}' deleted successfully.")
            logger.info(
                "User=%s deleted QCInstrument id=%s",
                request.user.username,
                pk,
            )
        except ProtectedError:
            messages.error(
                request,
                "This instrument is used in calibration schedule and cannot be deleted.",
            )
        return redirect("qc:instrument_list")

    # If someone hits GET on delete URL, just bounce back.
    return redirect("qc:instrument_list")






##----------------------------------------------------------------------


@login_required
def fg_qc_status_list(request):
    # ---- Permission check ----
    if not request.user.has_perm("QC.view_fgproductqcstatus"):
        messages.error(request, "You do not have permission to view FG QC Status.")
        logger.warning(
            "User='%s' tried to access FG QC Status LIST without permission.",
            request.user.username,
        )
        return redirect("indexpage")

    # we filter by product NAME now
    selected_product = (request.GET.get("product") or "").strip()

    logger.info(
        "User='%s' opened FG QC Status LIST | method=%s | selected_product='%s'",
        request.user.username,
        request.method,
        selected_product,
    )

    qs = FGProductQCStatus.objects.all().order_by("-date", "product")

    # distinct product names from status table itself
    product_names = (
        FGProductQCStatus.objects.order_by("product")
        .values_list("product", flat=True)
        .distinct()
    )

    if selected_product:
        qs = qs.filter(product=selected_product)

    logger.info(
        "FG QC Status LIST result | user='%s' | rows=%s",
        request.user.username,
        qs.count(),
    )
    context = {
        "statuses": qs,
        "products": product_names,
        "selected_product": selected_product,
    }
    return render(request, "qc/fg_qc_status_list.html", context)


@login_required
def fg_qc_status_create(request):
    """
    Create a new FGProductQCStatus row.
    """
    # ---- Permission check ----
    if not request.user.has_perm("QC.add_fgproductqcstatus"):
        messages.error(request, "You do not have permission to add FG QC Status.")
        logger.warning(
            "User='%s' tried to create FG QC Status without permission.",
            request.user.username,
        )
        return redirect("indexpage")

    logger.info(
        "User='%s' opened FG QC Status CREATE | method=%s",
        request.user.username,
        request.method,
    )

    if request.method == "POST":
        form = FGProductQCStatusForm(request.POST)

        if form.is_valid():
            obj = form.save()
            messages.success(request, "FG Product QC Status created successfully.")

            logger.info(
                "User='%s' created FG QC Status | id=%s | product='%s' | date=%s",
                request.user.username,
                getattr(obj, "id", None),
                getattr(obj, "product", None),
                getattr(obj, "date", None),
            )
            return redirect("qc:fg_qc_status_list")

        logger.warning(
            "FGProductQCStatusForm INVALID | user='%s' | errors=%s",
            request.user.username,
            form.errors.as_json(),
        )
        messages.error(request, "Please correct the errors below.")
    else:
        form = FGProductQCStatusForm()

    return render(
        request,
        "qc/fg_qc_status_form.html",
        {"form": form, "is_edit": False},
    )


@login_required
def fg_qc_status_update(request, pk):
    """
    Edit an existing FGProductQCStatus row.
    """
    # ---- Permission check ----
    if not request.user.has_perm("QC.change_fgproductqcstatus"):
        messages.error(request, "You do not have permission to update FG QC Status.")
        logger.warning(
            "User='%s' tried to update FG QC Status without permission | pk=%s",
            request.user.username,
            pk,
        )
        return redirect("indexpage")

    obj = get_object_or_404(FGProductQCStatus, pk=pk)

    logger.info(
        "User='%s' opened FG QC Status UPDATE | pk=%s | method=%s",
        request.user.username,
        pk,
        request.method,
    )

    if request.method == "POST":
        old_product = getattr(obj, "product", None)
        old_date = getattr(obj, "date", None)

        form = FGProductQCStatusForm(request.POST, instance=obj)
        if form.is_valid():
            updated = form.save()
            messages.success(request, "FG Product QC Status updated successfully.")

            logger.info(
                "User='%s' updated FG QC Status | pk=%s | product: '%s' -> '%s' | date: %s -> %s",
                request.user.username,
                pk,
                old_product,
                getattr(updated, "product", None),
                old_date,
                getattr(updated, "date", None),
            )
            return redirect("qc:fg_qc_status_list")

        logger.warning(
            "FGProductQCStatusForm INVALID (update) | user='%s' | pk=%s | errors=%s",
            request.user.username,
            pk,
            form.errors.as_json(),
        )
        messages.error(request, "Please correct the errors below.")
    else:
        form = FGProductQCStatusForm(instance=obj)

    return render(
        request,
        "qc/fg_qc_status_form.html",
        {"form": form, "is_edit": True, "obj": obj},
    )
    

@require_POST
@login_required
def fg_qc_status_delete(request, pk):
    """
    Delete an FGProductQCStatus row (POST only).
    """
    # ---- Permission check ----
    if not request.user.has_perm("QC.delete_fgproductqcstatus"):
        messages.error(request, "You do not have permission to delete FG QC Status.")
        logger.warning(
            "User='%s' tried to delete FG QC Status without permission | pk=%s",
            request.user.username,
            pk,
        )
        return redirect("indexpage")

    obj = get_object_or_404(FGProductQCStatus, pk=pk)

    product_name = obj.product
    record_date = obj.date

    logger.info(
        "User='%s' deleting FG QC Status | pk=%s | product='%s' | date=%s",
        request.user.username,
        pk,
        product_name,
        record_date,
    )

    obj.delete()

    logger.info(
        "User='%s' deleted FG QC Status | pk=%s",
        request.user.username,
        pk,
    )

    messages.success(
        request,
        f"FG Product QC Status for '{product_name}' on {record_date:%d-%m-%Y} deleted successfully.",
    )
    return redirect("qc:fg_qc_status_list")


@login_required
def fg_qc_status_export_excel(request):
    """
    Download FGProductQCStatus rows as an Excel file.

    Uses same filter as list view: ?product=<product_name>
    """
    selected_product = (request.GET.get("product") or "").strip()

    qs = FGProductQCStatus.objects.all().order_by("date", "product")
    if selected_product:
        qs = qs.filter(product=selected_product)

    # ---------------- Workbook setup ----------------
    output = BytesIO()
    wb = xlsxwriter.Workbook(output, {"in_memory": True})
    ws = wb.add_worksheet("FG QC Status")

    # Formats
    title_fmt = wb.add_format(
        {
            "bold": True,
            "font_size": 14,
            "align": "center",
            "valign": "vcenter",
        }
    )
    header_fmt = wb.add_format(
        {
            "bold": True,
            "font_size": 10,
            "align": "center",
            "valign": "vcenter",
            "bg_color": "#1D4ED8",   # blue-700
            "font_color": "#FFFFFF",
            "border": 1,
        }
    )
    date_fmt = wb.add_format(
        {
            "num_format": "dd-mm-yyyy",
            "border": 1,
        }
    )
    num_fmt = wb.add_format(
        {
            "num_format": "#,##0.000",
            "border": 1,
            "align": "right",
        }
    )
    text_left_fmt = wb.add_format(
        {
            "border": 1,
            "align": "left",
        }
    )
    srno_fmt = wb.add_format(
        {
            "border": 1,
            "align": "center",
        }
    )

    # Column widths
    ws.set_column("A:A", 6)   # Sr No
    ws.set_column("B:B", 12)  # Date
    ws.set_column("C:C", 40)  # Product
    ws.set_column("D:G", 15)  # Qty columns
    ws.set_column("H:H", 40)  # Remark

    # ---------------- Title ----------------
    title = "FG Product QC Status"
    if selected_product:
        title += f" – {selected_product}"

    # merge over 8 columns: A..H (0..7)
    ws.merge_range(0, 0, 0, 7, title, title_fmt)

    # optional subtitle row with export date
    subtitle_fmt = wb.add_format(
        {
            "italic": True,
            "font_size": 9,
            "align": "right",
            "valign": "vcenter",
        }
    )
    ws.merge_range(
        1,
        0,
        1,
        7,
        f"Exported on {date.today().strftime('%d-%m-%Y')}",
        subtitle_fmt,
    )

    # ---------------- Header row ----------------
    header_row = 3  # leave one empty row after subtitle
    headers = [
        "Sr No",
        "Date",
        "Product",
        "Approved Qty",
        "Off Spec Qty",
        "Under Analysis",
        "Total Qty",
        "Remark",
    ]

    for col, label in enumerate(headers):
        ws.write(header_row, col, label, header_fmt)

    # ---------------- Data rows ----------------
    row_idx = header_row + 1
    sr_no = 1
    for obj in qs:
        ws.write(row_idx, 0, sr_no, srno_fmt)
        ws.write(row_idx, 1, obj.date, date_fmt)
        ws.write(row_idx, 2, obj.product or "", text_left_fmt)
        ws.write_number(row_idx, 3, float(obj.approved_qty or 0), num_fmt)
        ws.write_number(row_idx, 4, float(obj.off_spec_qty or 0), num_fmt)
        ws.write_number(row_idx, 5, float(obj.under_analysis or 0), num_fmt)
        ws.write_number(row_idx, 6, float(obj.total_qty or 0), num_fmt)
        ws.write(row_idx, 7, obj.remark or "", text_left_fmt)

        row_idx += 1
        sr_no += 1

    wb.close()
    output.seek(0)

    filename = f"FG_Product_QC_Status_{date.today().strftime('%Y%m%d')}.xlsx"
    resp = HttpResponse(
        output.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp






######################### Instrument Occuoancy  #################################


# Pre-filled rows (not stored until selected)
INSTRUMENT_MASTER = [
    ("GC-1",  "Shimadzu", "GC-2014"),
    ("GC-3",  "Agilent",  "GC-7820"),
    ("GC-4",  "Shimadzu", "GC-2014"),
    ("GC-7",  "Agilent",  "GC-8890"),
    ("GC-8",  "Agilent",  "GC-8890"),
    ("GC-9",  "Agilent",  "GC-8890"),
    ("GC-10", "Agilent",  "GC-8890"),
    ("GC-11", "Agilent",  "GC-8890"),
    ("GC-12", "Agilent",  "GC-8890"),
    ("GC-13", "Agilent",  "GC-8890"),
    ("HPLC-1","Agilent",  "HPLC-1260"),
    ("HPLC-2","Shimadzu", "HPLC-LC-2010"),
    ("HPLC-3","Shimadzu", "HPLC-LC-2080"),
    ("HPLC-4","Agilent",  "HPLC-1260"),
]



def _build_master_rows_for_add():
    """Return dict rows (uniform for template)."""
    rows = []
    for idx, (area, make, model) in enumerate(INSTRUMENT_MASTER):
        rows.append({
            "idx": idx,
            "area": area,
            "make": make,
            "model": model,
            "checked": False,
            "occ": "",
            "rem": "",
        })
    return rows


@login_required
def instrument_occupancy_create(request):
    """
    Show fixed instrument rows, save ONLY selected rows.
    Redirect to LIST after save.
    """
    if not request.user.has_perm("QC.add_instrumentoccupancy"):
        messages.error(request, "You do not have permission to add Instrument Occupancy.")
        return redirect("indexpage")

    date_str = (request.GET.get("date") or "").strip()
    selected_date = parse_date(date_str) if date_str else date.today()
    if not selected_date:
        selected_date = date.today()

    if request.method == "POST":
        selected_date = parse_date(request.POST.get("date") or "") or date.today()

        checked_indexes = set()
        for x in request.POST.getlist("select_row"):
            try:
                checked_indexes.add(int(x))
            except:
                pass

        saved = 0
        for idx in sorted(checked_indexes):
            if idx < 0 or idx >= len(INSTRUMENT_MASTER):
                continue

            area, make, model = INSTRUMENT_MASTER[idx]
            occ_val = (request.POST.get(f"occ_{idx}") or "").strip()
            rem_val = (request.POST.get(f"rem_{idx}") or "").strip()

            if not occ_val:
                messages.error(request, f"Please enter % Occupancy for {area} {make} {model}.")
                return redirect(f"{request.path}?date={selected_date.isoformat()}")

            try:
                occ = float(occ_val)
            except:
                messages.error(request, f"Invalid % Occupancy for {area} {make} {model}.")
                return redirect(f"{request.path}?date={selected_date.isoformat()}")

            if occ < 0 or occ > 100:
                messages.error(request, f"% Occupancy must be 0–100 for {area} {make} {model}.")
                return redirect(f"{request.path}?date={selected_date.isoformat()}")

            InstrumentOccupancy.objects.create(
                date=selected_date,
                area=area,
                make=make,
                model=model,
                occupancy_percent=occ,
                remarks=rem_val,
            )
            saved += 1

        messages.success(request, f"Saved {saved} selected row(s).")
        return redirect('qc:instrument_occupancy_list')

    return render(
        request,
        "qc/instrument/instrument_occupancy_form.html",
        {
            "selected_date": selected_date,
            "rows": _build_master_rows_for_add(),  # ✅ always dict rows
            "is_edit": False,
        },
    )


@login_required
def instrument_occupancy_edit(request):
    """
    Bulk edit for a specific date.
    URL: /qc/instrument-occupancy/edit/?date=YYYY-MM-DD
    """
    if not request.user.has_perm("QC.change_instrumentoccupancy"):
        messages.error(request, "You do not have permission to edit Instrument Occupancy.")
        return redirect("indexpage")

    date_str = (request.GET.get("date") or "").strip()
    selected_date = parse_date(date_str) if date_str else date.today()
    if not selected_date:
        selected_date = date.today()

    existing_qs = InstrumentOccupancy.objects.filter(date=selected_date)
    existing_map = {(o.area, o.make, o.model): o for o in existing_qs}

    if request.method == "POST":
        selected_date = parse_date(request.POST.get("date") or "") or selected_date

        checked_indexes = set()
        for x in request.POST.getlist("select_row"):
            try:
                checked_indexes.add(int(x))
            except:
                pass

        saved = 0
        for idx, (area, make, model) in enumerate(INSTRUMENT_MASTER):
            if idx not in checked_indexes:
                continue

            occ_val = (request.POST.get(f"occ_{idx}") or "").strip()
            rem_val = (request.POST.get(f"rem_{idx}") or "").strip()

            if not occ_val:
                messages.error(request, f"Please enter % Occupancy for {area} {make} {model}.")
                return redirect(f"{request.path}?date={selected_date.isoformat()}")

            try:
                occ = float(occ_val)
            except:
                messages.error(request, f"Invalid % Occupancy for {area} {make} {model}.")
                return redirect(f"{request.path}?date={selected_date.isoformat()}")

            if occ < 0 or occ > 100:
                messages.error(request, f"% Occupancy must be 0–100 for {area} {make} {model}.")
                return redirect(f"{request.path}?date={selected_date.isoformat()}")

            obj = existing_map.get((area, make, model))
            if obj:
                obj.occupancy_percent = occ
                obj.remarks = rem_val
                obj.save(update_fields=["occupancy_percent", "remarks"])
            else:
                InstrumentOccupancy.objects.create(
                    date=selected_date,
                    area=area,
                    make=make,
                    model=model,
                    occupancy_percent=occ,
                    remarks=rem_val,
                )

            saved += 1

        messages.success(request, f"Updated {saved} selected row(s) for {selected_date.strftime('%d-%m-%Y')}.")
        return redirect('qc:instrument_occupancy_list')

    # ✅ build dict rows (same template)
    rows = []
    for idx, (area, make, model) in enumerate(INSTRUMENT_MASTER):
        obj = existing_map.get((area, make, model))
        rows.append({
            "idx": idx,
            "area": area,
            "make": make,
            "model": model,
            "checked": bool(obj),
            "occ": "" if not obj else ("" if obj.occupancy_percent is None else str(obj.occupancy_percent)),
            "rem": "" if not obj else (obj.remarks or ""),
        })

    return render(
        request,
        "qc/instrument/instrument_occupancy_form.html",
        {
            "selected_date": selected_date,
            "rows": rows,
            "is_edit": True,
        },
    )


@require_POST
@login_required
def instrument_occupancy_delete(request, pk: int):
    """
    Delete an InstrumentOccupancy record.
    Only accepts POST for safety.
    """
    # ---- Permission check ----
    if not request.user.has_perm("QC.delete_instrumentoccupancy"):
        messages.error(request, "You do not have permission to delete Instrument Occupancy records.")
        logger.warning(
            "User='%s' tried to DELETE Instrument Occupancy pk=%s without permission.",
            request.user.username,
            pk,
        )
        return redirect("indexpage")

    obj = get_object_or_404(InstrumentOccupancy, pk=pk)
    logger.info(
        "User='%s' attempting Instrument Occupancy DELETE | pk=%s",
        request.user.username,
        pk,
    )

    # (optional) keep some identifiers for audit log
    area = getattr(obj, "area", "")
    make = getattr(obj, "make", "")
    model = getattr(obj, "model", "")

    obj.delete()

    logger.info(
        "User='%s' deleted Instrument Occupancy | pk=%s | area='%s' make='%s' model='%s'",
        request.user.username,
        pk,
        area,
        make,
        model,
    )
    messages.success(request, "Instrument occupancy record deleted.")
    return redirect("qc:instrument_occupancy_list")


@login_required
def instrument_occupancy_list(request):
    """
    List InstrumentOccupancy records with filters: Date, Area, Make.
    """
    if not request.user.has_perm("QC.view_instrumentoccupancy"):
        messages.error(request, "You do not have permission to view Instrument Occupancy records.")
        logger.warning("User='%s' tried to view Instrument Occupancy LIST without permission.", request.user.username)
        return redirect("indexpage")

    selected_date_str = (request.GET.get("date") or "").strip()
    selected_area = (request.GET.get("area") or "").strip()
    selected_make = (request.GET.get("make") or "").strip()

    selected_date_obj = parse_date(selected_date_str) if selected_date_str else None
    if selected_date_str and not selected_date_obj:
        messages.error(request, "Invalid date format. Please select a valid date.")
        selected_date_str = ""
        selected_date_obj = None

    qs = InstrumentOccupancy.objects.all().order_by("date", "id")

    if selected_date_obj:
        qs = qs.filter(date=selected_date_obj)

    if selected_area:
        qs = qs.filter(area=selected_area)

    if selected_make:
        qs = qs.filter(make=selected_make)

    # Distinct filter values
    dates = (
        InstrumentOccupancy.objects.order_by("-date")
        .values_list("date", flat=True)
        .distinct()
    )
    areas = (
        InstrumentOccupancy.objects.order_by("area")
        .values_list("area", flat=True)
        .distinct()
    )
    makes = (
        InstrumentOccupancy.objects.order_by("make")
        .values_list("make", flat=True)
        .distinct()
    )

    context = {
        "records": qs,
        "dates": dates,
        "areas": areas,
        "makes": makes,
        "selected_date": selected_date_str,   # keep string for input value
        "selected_area": selected_area,
        "selected_make": selected_make,
        
    }
    return render(request, "qc/instrument/instrument_occupancy_list.html", context)






@login_required
def instrument_occupancy_export_excel(request):
    """
    Download InstrumentOccupancy rows as an Excel file.
    Uses the same filters as the list view: ?area=&make=
    Excel sorted in DESCENDING order (latest first).
    """

    # ---- Permission check ----
    if not request.user.has_perm("QC.view_instrumentoccupancy"):
        messages.error(request, "You do not have permission to export Instrument Occupancy records.")
        logger.warning(
            "User='%s' tried to EXPORT Instrument Occupancy without permission.",
            request.user.username,
        )
        return redirect("indexpage")

    selected_area = (request.GET.get("area") or "").strip()
    selected_make = (request.GET.get("make") or "").strip()

    logger.info(
        "User='%s' started Instrument Occupancy EXPORT | area='%s' make='%s'",
        request.user.username,
        selected_area,
        selected_make,
    )

    # ?? DESCENDING ORDER (Latest first)
    qs = InstrumentOccupancy.objects.all().order_by("date", "id")

    if selected_area:
        qs = qs.filter(area=selected_area)

    if selected_make:
        qs = qs.filter(make=selected_make)

    total_rows = qs.count()

    # ---------------- Workbook setup ----------------
    output = BytesIO()
    wb = xlsxwriter.Workbook(output, {"in_memory": True})
    ws = wb.add_worksheet("Instrument Occupancy")

    # ---------------- Formats ----------------
    title_fmt = wb.add_format({
        "bold": True,
        "font_size": 14,
        "align": "center",
        "valign": "vcenter",
        "font_color": "#FFFFFF",
        "bg_color": "#1E293B",  # slate-800
    })

    subtitle_fmt = wb.add_format({
        "font_size": 9,
        "align": "left",
        "valign": "vcenter",
        "font_color": "#4B5563",
    })

    header_fmt = wb.add_format({
        "bold": True,
        "font_size": 10,
        "align": "center",
        "valign": "vcenter",
        "bg_color": "#E5E7EB",
        "border": 1,
    })

    text_fmt = wb.add_format({
        "font_size": 10,
        "align": "left",
        "valign": "vcenter",
        "border": 1,
    })

    date_fmt = wb.add_format({
        "font_size": 10,
        "align": "center",
        "valign": "vcenter",
        "border": 1,
        "num_format": "dd/mm/yyyy",
    })

    num_fmt = wb.add_format({
        "font_size": 10,
        "align": "right",
        "valign": "vcenter",
        "border": 1,
        "num_format": "0.00",
    })

    sr_fmt = wb.add_format({
        "font_size": 10,
        "align": "center",
        "valign": "vcenter",
        "border": 1,
    })

    # ---------------- Column widths ----------------
    ws.set_column("A:A", 6)   # Sr No
    ws.set_column("B:B", 12)  # Date
    ws.set_column("C:C", 14)  # Area
    ws.set_column("D:D", 16)  # Make
    ws.set_column("E:E", 16)  # Model
    ws.set_column("F:F", 14)  # % Occupancy
    ws.set_column("G:G", 40)  # Remarks

    row = 0

    # ---------------- Title ----------------
    ws.merge_range(row, 0, row, 6, "INSTRUMENT OCCUPANCY REPORT", title_fmt)
    row += 1

    # ---------------- Subtitle ----------------
    now = timezone.localtime()
    subtitle_parts = [f"Exported on: {now:%d-%m-%Y %H:%M}"]

    if selected_area:
        subtitle_parts.append(f"Area = {selected_area}")
    if selected_make:
        subtitle_parts.append(f"Make = {selected_make}")

    ws.merge_range(row, 0, row, 6, " | ".join(subtitle_parts), subtitle_fmt)
    row += 2

    # ---------------- Header ----------------
    headers = ["Sr No", "Date", "Area", "Make", "Model", "% Occupancy", "Remarks"]

    header_row = row
    for col, h in enumerate(headers):
        ws.write(row, col, h, header_fmt)

    # Freeze header row
    ws.freeze_panes(header_row + 1, 0)

    # Enable filter
    ws.autofilter(header_row, 0, header_row, len(headers) - 1)

    row += 1

    # ---------------- Data ----------------
    for idx, obj in enumerate(qs, start=1):
        ws.write(row, 0, idx, sr_fmt)

        if obj.date:
            dt_val = datetime.combine(obj.date, time())
            ws.write_datetime(row, 1, dt_val, date_fmt)
        else:
            ws.write_blank(row, 1, None, date_fmt)

        ws.write(row, 2, obj.area or "", text_fmt)
        ws.write(row, 3, obj.make or "", text_fmt)
        ws.write(row, 4, obj.model or "", text_fmt)
        ws.write_number(row, 5, float(obj.occupancy_percent or 0), num_fmt)
        ws.write(row, 6, obj.remarks or "", text_fmt)

        row += 1

    wb.close()
    output.seek(0)

    filename = f"Instrument_Occupancy_{now:%Y%m%d}.xlsx"

    logger.info(
        "User='%s' completed Instrument Occupancy EXPORT | filename='%s' | rows=%s",
        request.user.username,
        filename,
        total_rows,
    )

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response





@login_required
def qc_powerbi_dashboard(request):
    """
    Display embedded Power BI dashboard in iframe.
    """
    powerbi_url = "https://app.powerbi.com/view?r=eyJrIjoiYzM1ZjQxNmQtZTU0OC00MTY3LTk2YjctNWM0ZjMzNmM1MWM2IiwidCI6ImMxZDAyMjBkLWRhZjMtNGMyZC05YzE0LWZlZWJiY2EwNGVhZCJ9"

    return render(request, "qc/powerbi_dashboard.html", {
        "powerbi_url": powerbi_url
    })
    
    