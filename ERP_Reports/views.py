# erp_reports/views.py
from math import ceil
from io import BytesIO
from django.http import HttpResponse
from django.shortcuts import render
from django.db import connections
import pandas as pd
from email.utils import formataddr
from datetime import date

def cogs_report(request):
    """
    COGS report with paging, filtering, Excel export, and **Send Mail** (Excel attachment).
    """
    # Local imports so you don't have to touch globals/imports at the top of the file
    from django.core.mail import EmailMessage
    from django.contrib import messages
    from django.conf import settings
    from django.shortcuts import redirect

    # -------- Params --------
    user_groups  = request.user.groups.values_list('name', flat=True)
    is_superuser = request.user.is_superuser

    # Base params
    company_id = int(request.GET.get('company_id', 27))
    year_id    = int(request.GET.get('year_id',    7))
    from_date  = request.GET.get('from_date', '2022-04-01')
    to_date    = request.GET.get('to_date')  # None -> GETDATE() in SQL
    cust_code  = int(request.GET.get('cust_code', 0))
    item_id    = int(request.GET.get('item_id',   0))

    # New filters (free-text)
    txn_name      = (request.GET.get('txn_name') or '').strip()
    customer_name = (request.GET.get('customer_name') or '').strip()
    item_name     = (request.GET.get('item_name') or '').strip()

    # Paging
    page      = max(1, int(request.GET.get('page', 1)))
    page_size = max(1, int(request.GET.get('page_size', 50)))
    offset    = (page - 1) * page_size
    sizes     = [25, 50, 100, 200]

    # Helper for safe literal embedding
    def q(s: str) -> str:
        return s.replace("'", "''") if s else s

    to_date_sql = "GETDATE()" if not to_date else f"'{q(to_date)}'"

    # Build extra WHERE for string filters (applied on outer SELECT)
    extra_where = ""
    if txn_name:
        extra_where += f" AND [Transaction Name] LIKE '%{q(txn_name)}%'"
    if customer_name:
        extra_where += f" AND [Customer Name] LIKE '%{q(customer_name)}%'"
    if item_name:
        extra_where += f" AND [Item Name] LIKE '%{q(item_name)}%'"

    # -------- Base SQL (CTEs) --------
    base_ctes = f"""
SET NOCOUNT ON;

DECLARE
    @CompanyID INT   = {company_id},
    @YearId    INT   = {year_id},
    @FromDate  DATE  = '{q(from_date)}',
    @ToDate    DATE  = {to_date_sql},
    @CustCode  INT   = {cust_code},
    @ItemId    INT   = {item_id};

;WITH RawSales AS (
    SELECT
        dt.sName AS [Transaction Name],
        d.lId    AS SalesInvoiceId,
        d.sDocNo AS [Sales Invoice No],
        CONVERT(VARCHAR, CONVERT(DATE, CONVERT(VARCHAR(8), d.dtDocDate ,112)),106) AS [Sales Invoice Date],
        CUST.sName AS [Customer Name],
        dd.lLnkDocId, dd.lLnkLine, dd.lLine,
        ITP.sName AS [Item Type],
        ITM.sCode AS [Item Code],
        ITM.sName AS [Item Name],
        dd.sValue1 AS [Batch No.],
        UOM.sCode  AS [UOM],
        CASE WHEN d.lTypId NOT IN (990,341,1079)
             THEN CONVERT(DECIMAL(18,2), dd.dQty2)
             ELSE CONVERT(DECIMAL(18,2), dd.dQty2) * -1 END              AS [Sales Quantity],
        CASE WHEN d.lTypId NOT IN (990,341,1079)
             THEN CONVERT(DECIMAL(18,2), dd.dRate)
             ELSE CONVERT(DECIMAL(18,2), dd.dRate) * -1 END              AS [Sales Rate],
        CASE WHEN d.lTypId NOT IN (990,341,1079)
             THEN CONVERT(DECIMAL(18,2), dd.dQty2 * dd.dRate)
             ELSE CONVERT(DECIMAL(18,2), dd.dQty2 * -1 * dd.dRate) END   AS [Sales Value],
        CONVERT(DECIMAL(18,2), -(dd.dStkVal / NULLIF(dd.dQty2,0)))
          - (CONVERT(DECIMAL(18,2), dds.dRate3) - CONVERT(DECIMAL(18,2), dds.dRate)) AS [Material Rate],
        CONVERT(DECIMAL(18,2), -(dd.dStkVal))
          - dd.dQty2 * (CONVERT(DECIMAL(18,2), dds.dRate3) - CONVERT(DECIMAL(18,2), dds.dRate)) AS [Material Value],
        CONVERT(DECIMAL(18,2), dds.dRate3) - CONVERT(DECIMAL(18,2), dds.dRate) AS [Other Rate],
        dd.dQty2 * (CONVERT(DECIMAL(18,2), dds.dRate3) - CONVERT(DECIMAL(18,2), dds.dRate)) AS [Other Value],
        CONVERT(DECIMAL(18,2), -(dd.dStkVal / NULLIF(dd.dQty2,0))) AS [COGS Rate],
        CONVERT(DECIMAL(18,2), -(dd.dStkVal)) AS [COGS Value],
        C.sName AS [Cost Centre],
        u.sRemarks AS [Sale Person Name]
    FROM  TXNTYP  dt
    JOIN  TXNHDR  d   ON d.lTypId = dt.lTypId
                     AND d.lTypId IN (341,499,504,650,654,824,825,826,827,828,829,939,940,990,1079)
    JOIN  TXNDET  dd  ON d.lId = dd.lId AND dd.cFlag = 'I'
    LEFT  JOIN TXNDET dds ON dd.lStkId   = dds.lId AND dd.lStkLine = dds.lLine
    LEFT  JOIN BUSMST CUST ON d.lAccId1 = CUST.lId
    LEFT  JOIN ITMMST ITM  ON dd.lItmId = ITM.lId
    LEFT  JOIN ITMTYP ITP  ON ITP.lTypId = dd.lItmTyp
    LEFT  JOIN UNTMST UOM  ON dd.lUntId  = UOM.lId
    LEFT  JOIN DIMMST C    ON dd.lDimId  = C.lId AND C.cTyp = 'C'
    LEFT  JOIN USRMST u    ON d.lEmpId   = u.lId
    WHERE (CUST.lId  = @CustCode OR @CustCode = 0)
      AND (dd.lItmId = @ItemId   OR @ItemId  = 0)
      AND  d.lCompId IN (27,9,28,25,26)
      AND  d.bDel    = 0
      AND  CONVERT(DATE, CONVERT(VARCHAR(8), d.dtDocDate,112))
           BETWEEN @FromDate AND @ToDate
),
BondAdj AS (
    SELECT
        dd.lId   AS LnkDocId,
        dd.lLine AS LnkLine,
        CONVERT(DECIMAL(18,2), -(dd.dStkVal / NULLIF(dd.dQty2,0))) AS NewCOGSRate,
        CONVERT(DECIMAL(18,2), -(dd.dStkVal)) AS NewCOGSValue
    FROM  TXNDET dd
    WHERE dd.lTypId = 902
),
Final AS (
    SELECT
        RS.[Transaction Name],
        RS.[Sales Invoice No],
        RS.[Sales Invoice Date],
        RS.[Customer Name],
        RS.[Cost Centre],
        RS.[Sale Person Name],
        RS.[Item Type],
        RS.[Item Code],
        RS.[Item Name],
        RS.[Batch No.],
        RS.[UOM],
        RS.[Sales Quantity],
        RS.[Sales Rate],
        RS.[Sales Value],
        RS.[Material Rate],
        RS.[Material Value],
        RS.[Other Rate],
        RS.[Other Value],
        COALESCE(BA.NewCOGSRate , RS.[COGS Rate])  AS [COGS Rate],
        COALESCE(BA.NewCOGSValue, RS.[COGS Value]) AS [COGS Value],
        RS.[Sales Rate] - COALESCE(BA.NewCOGSRate, RS.[COGS Rate]) AS [GrossProfitPerKG],
        RS.[Sales Quantity] * (RS.[Sales Rate] - COALESCE(BA.NewCOGSRate, RS.[COGS Rate])) AS [Value],
        CASE WHEN RS.[Sales Value] <> 0
             THEN (RS.[Sales Quantity] * (RS.[Sales Rate] - COALESCE(BA.NewCOGSRate, RS.[COGS Rate])))
                  / RS.[Sales Value] * 100
             ELSE 0 END AS [Percent]
    FROM RawSales RS
    LEFT JOIN BondAdj BA
           ON BA.LnkDocId = RS.lLnkDocId
          AND BA.LnkLine  = RS.lLnkLine
          AND RS.[Transaction Name] = 'Ex Bond Sales Invoice - Domestic'
)
"""

    # Paginated SELECT
    sql_page = f"""
{base_ctes}
SELECT
    [Transaction Name],
    [Sales Invoice No],
    [Sales Invoice Date],
    [Customer Name],
    [Cost Centre],
    [Item Type],
    [Item Code],
    [Item Name],
    [Batch No.],
    [UOM],
    [Sales Quantity],
    [Sales Rate],
    [Sales Value],
    [Material Rate],
    [Material Value],
    [Other Rate],
    [Other Value],
    [COGS Rate],
    [COGS Value],
    [GrossProfitPerKG],
    [Value],
    [Percent],
    COUNT_BIG(*) OVER() AS total_rows
FROM Final
WHERE [Cost Centre] = 'CHEMICALS' {extra_where}
ORDER BY [Sales Invoice Date], [Sales Invoice No]
OFFSET {offset} ROWS FETCH NEXT {page_size} ROWS ONLY;
"""

    # Export SELECT (no pagination, no total_rows col)
    sql_export = f"""
{base_ctes}
SELECT
    [Transaction Name],
    [Sales Invoice No],
    [Sales Invoice Date],
    [Customer Name],
    [Cost Centre],
    [Item Type],
    [Item Code],
    [Item Name],
    [Batch No.],
    [UOM],
    [Sales Quantity],
    [Sales Rate],
    [Sales Value],
    [Material Rate],
    [Material Value],
    [Other Rate],
    [Other Value],
    [COGS Rate],
    [COGS Value],
    [GrossProfitPerKG],
    [Value],
    [Percent]
FROM Final
WHERE [Cost Centre] = 'CHEMICALS' {extra_where}
ORDER BY [Sales Invoice Date], [Sales Invoice No];
"""

    # -------- Export or Send Mail? --------
    want_export = (request.GET.get('export') or '').lower() in ('xlsx', 'excel')
    want_mail   = bool(request.GET.get('send_mail'))

    if want_export or want_mail:
        # Fetch full dataset for export/mail
        with connections['readonly_db'].cursor() as cursor:
            cursor.execute(sql_export)
            cols = [c[0] for c in cursor.description]
            data = cursor.fetchall()

        df = pd.DataFrame(data, columns=cols)

        out = BytesIO()
        with pd.ExcelWriter(out, engine="openpyxl") as xw:
            df.to_excel(xw, index=False, sheet_name="COGS")
            ws = xw.book["COGS"]
            ws.freeze_panes = ws["A2"]   # freeze header
            # auto-width (inspect first 200 rows for width)
            for col_idx, col_name in enumerate(df.columns, start=1):
                try:
                    sample_vals = (str(v) for v in df[col_name].head(200).values)
                    maxlen = max([len(str(col_name)), *[len(s) for s in sample_vals]]) + 2
                except ValueError:
                    maxlen = len(str(col_name)) + 2
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(maxlen, 40)

        if want_mail:
            # Recipients: ?to=email1,email2 OR fallback to DEFAULT_FROM_EMAIL
            to_param = (request.GET.get("to") or "").strip()
            recipients = [e.strip() for e in to_param.split(",") if e.strip()]
            if not recipients:
                # Default: send to the logged-in user if email exists, else DEFAULT_FROM_EMAIL
                if getattr(request.user, "email", ""):
                    recipients = [request.user.email]
                else:
                    recipients = [getattr(settings, "DEFAULT_FROM_EMAIL", "workflow@ocspl.com")]

            subject = "COGS Report"
            # Add quick summary of applied filters in the email body
            filters_text = []
            if from_date:    filters_text.append(f"From: {from_date}")
            if to_date:      filters_text.append(f"To: {to_date}")
            if txn_name:     filters_text.append(f"Txn: {txn_name}")
            if customer_name:filters_text.append(f"Customer: {customer_name}")
            if item_name:    filters_text.append(f"Item: {item_name}")
            filters_line = " | ".join(filters_text) if filters_text else "All records"

            body = (
                "Dear Team,\n\n"
                "Please find attached the latest COGS Report.\n"
                f"Filters: {filters_line}\n\n"
                "Regards,\nWorkflow"
            )

            email = EmailMessage(
                subject=subject,
                body=body,
                from_email=None,  # uses DEFAULT_FROM_EMAIL
                to=recipients,
            )
            email.attach("COGS_Report.xlsx", out.getvalue(),
                         "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            try:
                email.send(fail_silently=False)
                messages.success(request, f"COGS Report emailed to: {', '.join(recipients)}")
            except Exception as e:
                messages.error(request, f"Email failed: {e}")

            # Redirect back to same page WITHOUT the send_mail param
            qs = request.GET.copy()
            qs.pop("send_mail", None)
            clean_url = f"{request.path}?{qs.urlencode()}" if qs else request.path
            return redirect(clean_url)

        # Otherwise: download Excel
        resp = HttpResponse(
            out.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = 'attachment; filename="COGS_Report.xlsx"'
        return resp

    # -------- Execute (paged) --------
    with connections['readonly_db'].cursor() as cursor:
        cursor.execute(sql_page)
        columns = [c[0] for c in cursor.description]
        rows    = cursor.fetchall()

    # -------- Shape data / totals --------
    if rows:
        total_rows = int(rows[0][-1])
        num_pages  = ceil(total_rows / page_size) if page_size else 1
        start_row  = (page - 1) * page_size + 1
        end_row    = min(page * page_size, total_rows)
        columns = columns[:-1]          # strip total_rows
        rows    = [r[:-1] for r in rows]
    else:
        total_rows = 0
        num_pages  = 0
        start_row  = 0
        end_row    = 0

    df = pd.DataFrame(rows, columns=columns)

    # -------- Prev/next URLs (preserve filters) --------
    qs = request.GET.copy()
    qs["page_size"] = page_size
    prev_url = next_url = None
    if page > 1:
        qs["page"] = page - 1
        prev_url = f"{request.path}?{qs.urlencode()}"
    if page < num_pages:
        qs["page"] = page + 1
        next_url = f"{request.path}?{qs.urlencode()}"

    # -------- Render --------
    return render(request, 'PC Reports/cogs.html', {
        'columns':     list(df.columns),
        'table_data':  df.to_dict(orient='records'),
        'page':        page,
        'page_size':   page_size,
        'num_pages':   num_pages,
        'prev_url':    prev_url,
        'next_url':    next_url,
        'sizes':       sizes,
        'total_rows':  total_rows,
        'start_row':   start_row,
        'end_row':     end_row,

        # Keep these so layout (sidebar) matches the rest
        'user_groups': user_groups,
        'is_superuser': is_superuser,
        'active_menu': 'reports',

        # Echo filters to template
        'filters': {
            'from_date': from_date,
            'to_date': to_date or '',
            'txn_name': txn_name,
            'customer_name': customer_name,
            'item_name': item_name,
        }
    })



############################# Ageing Backend (Snapshot-based) #############################

from datetime import date, datetime, timedelta
from decimal import Decimal
from io import StringIO, BytesIO
import csv
import json
import logging
import time
from django.db import connections, transaction
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect
from django.views.decorators.http import require_GET, require_POST
from django.contrib.auth.decorators import login_required
from django.contrib import messages

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

FIXED_FROM_DATE = "2022-03-01"
COMPANY_ID = 27

# ---- Transaction-type groups ---- #
CONSUMPTION_TXN_TYPES = [
    "Material Issue Voucher - Capex",
    "Material Issue-Consumption Voucher - MOC",
    "Maintenance Material Issue-Consumption",
]
INWARD_TXN_TYPES = [
    "CAPEX GRN",
    "Domestic GRN Admin",
    "Domestic GRN Engineering",
    "Domestic GRN RMPM",
]

EXCLUDE_TXN_TYPES = [
    "",
]

EXCESS_BATCH_SQL = (
    "S.batch_no LIKE '%excess%' "
    "OR S.batch_no IN ('decexcess2023', 'EXCESS STOCK', 'Excessstock', 'Decexcess2023')"
)

def _esc(s: str) -> str:
    return (s or "").replace("'", "''")


def _sql_in_list(values):
    return ",".join(f"'{_esc(v)}'" for v in values)


CONSUMPTION_IN_SQL = _sql_in_list(CONSUMPTION_TXN_TYPES)
INWARD_IN_SQL = _sql_in_list(INWARD_TXN_TYPES)
EXCLUDE_IN_SQL = _sql_in_list(EXCLUDE_TXN_TYPES)

ITEM_TYPE_LABEL_BY_ID = {
    57: "Raw Material",
    60: "Finished Good",
    61: "Semi Finished Good",
    62: "WIP FR",
    66: "Packing Material",
    68: "Engineering Material",
    76: "Work in Progress",
    77: "Key Raw Material",
    81: "Intercut",
    82: "Intercut",
}


def _get_date_range(request):
    today = date.today()
    fy_year = today.year - 1 if today.month < 4 else today.year
    current_fy_start = f"{fy_year}-04-01"

    user_from = request.GET.get("from") or current_fy_start
    
    try:
        base_dt = datetime.strptime(FIXED_FROM_DATE, "%Y-%m-%d").date()
        user_dt = datetime.strptime(user_from, "%Y-%m-%d").date()
        if user_dt < base_dt:
            user_dt = base_dt
        from_date = user_dt.isoformat()
    except Exception:
        from_date = current_fy_start

    to_date = request.GET.get("to") or today.isoformat()
    l_from = int(from_date.replace("-", ""))
    l_to = int(to_date.replace("-", ""))
    to_date_expr = f"CONVERT(date, '{to_date}')"
    
    return from_date, to_date, l_from, l_to, to_date_expr


def _get_filters(request):
    raw_ids = (request.GET.get("item_type_ids") or "").strip()
    item_type_ids = []
    if raw_ids:
        for tok in raw_ids.split(","):
            tok = tok.strip()
            if tok:
                try:
                    item_type_ids.append(int(tok))
                except ValueError:
                    pass

    item_name = (request.GET.get("item_name") or "").strip()

    item_name_mode = (request.GET.get("item_name_mode") or "exact").strip().lower()
    if item_name_mode not in ("contains", "exact"):
        item_name_mode = "exact"

    return {
        "item_name": item_name,
        "item_name_mode": item_name_mode,
        "item_code": (request.GET.get("item_code") or "").strip(),
        "item_type": (request.GET.get("item_type") or "").strip(),
        "item_type_ids": item_type_ids,
        "location": (request.GET.get("location") or "").strip(),
        "batch_no": (request.GET.get("batch_no") or "").strip(),
        "transaction_type": (request.GET.get("transaction_type") or "").strip(),
        "inventory_category": (request.GET.get("inventory_category") or "").strip(),
        "inventory_subcategory": (request.GET.get("inventory_subcategory") or "").strip(),
        "age_bucket": (request.GET.get("age_bucket") or "").strip(),
        "material_group": (request.GET.get("material_group") or "").strip(),
        "location_type": (request.GET.get("location_type") or "").strip(),
        "stock_location": (request.GET.get("stock_location") or "").strip(),
    }


def _dbg_print(enabled: bool, *args):
    if enabled:
        try:
            print(*args)
        except Exception:
            pass


def _build_item_name_where(filters, alias="S", debug=False):
    item_name = (filters.get("item_name") or "").strip()
    if not item_name:
        return ""

    mode = (filters.get("item_name_mode") or "exact").strip().lower()
    item_name_nq = item_name.replace('"', "").replace("'", "").strip()
    col_norm = f"REPLACE(REPLACE({alias}.item_name,'\"',''),'''','')"

    if mode == "exact":
        w = (
            f" AND ("
            f"   LTRIM(RTRIM({alias}.item_name)) = N'{_esc(item_name)}' "
            f"   OR LTRIM(RTRIM({col_norm})) = N'{_esc(item_name_nq)}'"
            f" )"
        )
        return w

    w = (
        f" AND ("
        f"   {alias}.item_name LIKE N'%{_esc(item_name)}%' "
        f"   OR {col_norm} LIKE N'%{_esc(item_name_nq)}%'"
        f" )"
    )
    return w

def _build_item_where_only(filters, alias="S", debug=False):
    extra = []

    if filters.get("item_code"):
        extra.append(f"{alias}.item_code LIKE '%{_esc(filters['item_code'])}%'")

    ids = filters.get("item_type_ids") or []
    if ids:
        names = [ITEM_TYPE_LABEL_BY_ID[i] for i in ids if i in ITEM_TYPE_LABEL_BY_ID]
        if names:
            extra.append(
                f"{alias}.item_type IN ({','.join(chr(39) + _esc(n) + chr(39) for n in names)})"
            )
    elif filters.get("item_type"):
        item_type = filters["item_type"]
        if item_type == "(Blank)":
            extra.append(f"({alias}.item_type IS NULL OR {alias}.item_type = '')")
        else:
            extra.append(f"{alias}.item_type = '{_esc(item_type)}'")

    mg = (filters.get("material_group") or "").strip()
    if mg == "eng":
        extra.append(f"{alias}.item_type = 'Engineering Material'")
    elif mg == "rm_pm":
        extra.append(f"{alias}.item_type <> 'Engineering Material'")

    if filters.get("location"):
        loc = filters["location"]
        if loc == "(Blank)":
            extra.append(f"({alias}.location IS NULL OR {alias}.location = '')")
        else:
            extra.append(f"{alias}.location = '{_esc(loc)}'")

    if filters.get("stock_location"):
        extra.append(f"{alias}.stock_location = '{_esc(filters['stock_location'])}'")

    lt = (filters.get("location_type") or "").strip()
    if lt in ("Store", "CAPEX"):
        extra.append(
            f"CASE WHEN {alias}.location LIKE 'CAP%' OR {alias}.location = 'Solapur E-20 Capex Store' "
            f"THEN 'CAPEX' ELSE 'Store' END = '{_esc(lt)}'"
        )

    if filters.get("batch_no"):
        batch = filters["batch_no"]
        if batch == "(Blank)":
            extra.append(f"({alias}.batch_no IS NULL OR {alias}.batch_no = '')")
        else:
            extra.append(f"{alias}.batch_no = '{_esc(batch)}'")
            
    if filters.get("transaction_type"):
        txn = filters["transaction_type"]
        if txn == "(Blank)":
            extra.append(f"({alias}.transaction_type IS NULL OR {alias}.transaction_type = '')")
        else:
            extra.append(f"{alias}.transaction_type = '{_esc(txn)}'")
            
    if filters.get("inventory_category"):
        inv_cat = filters["inventory_category"]
        if inv_cat == "(Blank)":
            extra.append(f"({alias}.inventory_category IS NULL OR {alias}.inventory_category = '')")
        else:
            extra.append(f"{alias}.inventory_category = '{_esc(inv_cat)}'")
            
    if filters.get("inventory_subcategory"):
        inv_subcat = filters["inventory_subcategory"]
        if inv_subcat == "(Blank)":
            extra.append(f"({alias}.inventory_subcategory IS NULL OR {alias}.inventory_subcategory = '')")
        else:
            extra.append(f"{alias}.inventory_subcategory = '{_esc(inv_subcat)}'")

    base = (" AND " + " AND ".join(extra)) if extra else ""
    base += _build_item_name_where(filters, alias, debug=debug)
    _dbg_print(debug, "[item_where_only]", base)
    return base


def _dynamic_age_expr_for_where(alias="S", to_date=None, use_fmd=False, fmd_alias="F"):
    if to_date:
        if use_fmd:
            base_date_expr = f"COALESCE({fmd_alias}.first_movement_date, {alias}.doc_date)"
            return f"DATEDIFF(DAY, {base_date_expr}, CONVERT(date,'{_esc(to_date)}'))"
        return f"DATEDIFF(DAY, {alias}.doc_date, CONVERT(date,'{_esc(to_date)}'))"
    return f"{alias}.age_days"


def _age_bucket_case_for_where(age_expr, alias="S"):
    return (
        f"CASE "
        f"WHEN {alias}.batch_no LIKE '%excess%' OR {alias}.batch_no IN ('decexcess2023', 'EXCESS STOCK', 'Excessstock','Decexcess2023') THEN '>720' "
        f"WHEN {age_expr} BETWEEN 0 AND 60 THEN '0-60' "
        f"WHEN {age_expr} BETWEEN 61 AND 90 THEN '61-90' "
        f"WHEN {age_expr} BETWEEN 91 AND 160 THEN '91-160' "
        f"WHEN {age_expr} BETWEEN 161 AND 360 THEN '161-360' "
        f"WHEN {age_expr} BETWEEN 361 AND 720 THEN '361-720' "
        f"ELSE '>720' END"
    )


def _build_extra_where(
    filters, alias="S", debug=False, to_date=None, use_fmd=False, fmd_alias="F"
):
    base = _build_item_where_only(filters, alias, debug=debug)
    ab = (filters.get("age_bucket") or "").strip()
    if ab:
        age_expr = _dynamic_age_expr_for_where(
            alias, to_date, use_fmd=use_fmd, fmd_alias=fmd_alias
        )
        age_bucket_expr = _age_bucket_case_for_where(age_expr, alias)
        if ab in ("0-60", "61-90", "91-160", "161-360", "361-720", ">720"):
            out = base + f" AND {age_bucket_expr} = '{_esc(ab)}'"
            _dbg_print(debug, "[extra_where_with_age]", out)
            return out
    return base


def _json_safe(v):
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, (date, datetime)):
        return v.isoformat()
    return v


def _sample_rows(rows, n=5):
    return [[_json_safe(x) for x in list(r)] for r in rows[:n]]


def _exec_sql_snapshot(sql, params=None):
    params = params or []
    final = "SET NOCOUNT ON;\n" + sql
    with connections["default"].cursor() as cur:
        cur.execute(final, params) if params else cur.execute(final)
        last_rows, last_cols = [], []
        while True:
            if cur.description:
                last_cols = [c[0] for c in cur.description]
                last_rows = cur.fetchall()
            if not cur.nextset():
                break
    return last_rows, last_cols


def _exec_sql_snapshot_multi(sql, params=None, debug=False):
    params = params or []
    final = "SET NOCOUNT ON;\n" + sql
    t0 = time.time()
    sets, trace = [], []

    with connections["default"].cursor() as cur:
        cur.execute(final, params) if params else cur.execute(final)
        si = 0
        while True:
            if cur.description:
                cols = [c[0] for c in cur.description]
                rows = cur.fetchall()
                sets.append({"cols": cols, "rows": rows})
                if debug:
                    trace.append(
                        {
                            "set_index": si,
                            "type": "SELECT",
                            "cols": cols,
                            "row_count": len(rows),
                            "sample": _sample_rows(rows, 5),
                        }
                    )
            else:
                if debug:
                    trace.append({"set_index": si, "type": "NON_SELECT", "rowcount": cur.rowcount})
            si += 1
            if not cur.nextset():
                break

    dp = None
    if debug:
        db = {}
        try:
            with connections["default"].cursor() as c2:
                c2.execute("SELECT DB_NAME(), @@SERVERNAME;")
                r = c2.fetchone()
                if r:
                    db = {"db_name": r[0], "server_name": r[1]}
        except Exception:
            pass

        dp = {
            "elapsed_ms": int((time.time() - t0) * 1000),
            "db": db,
            "trace": trace,
            "sql": final,
        }
    return sets, dp


@login_required
def inventory_ageing_report(request):
    from_date, _to, *_ = _get_date_range(request)
    run_flag = request.GET.get("run") == "1"
    sc, sm = "", ""
    try:
        with connections["default"].cursor() as cur:
            cur.execute(
                "SET NOCOUNT ON; "
                "SELECT TOP(1) CONVERT(date,created_at), CONVERT(date,MAX(doc_date)) "
                "FROM dbo.inventory_ageing_snapshot "
                "GROUP BY created_at ORDER BY created_at DESC;"
            )
            r = cur.fetchone()
            if r:
                if r[0]:
                    sc = r[0].isoformat()
                if r[1]:
                    sm = r[1].isoformat()
    except Exception as e:
        logger.exception("Snapshot meta: %s", e)

    return render(
        request,
        "Store/inventory_ageing.html",
        {
            "today_str": date.today().isoformat(),
            "from_str": from_date,
            "show_results": run_flag,
            "error_msg": "",
            "total_rows": 0,
            "snapshot_created_at_str": sc,
            "snapshot_max_doc_date_str": sm,
        },
    )

@login_required
@require_GET
def inventory_ageing_summary_api(request):
    from_date, to_date, _lf, _l_to, _ = _get_date_range(request)
    filters = _get_filters(request)

    dbg = request.GET.get("debug") == "1"

    has_age_filter = bool((filters.get("age_bucket") or "").strip())

    iw = (
        _build_extra_where(filters, "S", debug=dbg, to_date=to_date, use_fmd=True, fmd_alias="F")
        if has_age_filter
        else _build_item_where_only(filters, "S", debug=dbg)
    )
    pw = _build_extra_where(filters, "S", debug=dbg, to_date=to_date, use_fmd=True, fmd_alias="F")

    try:
        fd = datetime.strptime(from_date, "%Y-%m-%d").date()
        oe = (fd - timedelta(days=1)).isoformat()
    except Exception:
        oe = from_date

    LT = (
        "CASE WHEN {a}.location LIKE 'CAP%' OR {a}.location = 'Solapur E-20 Capex Store' "
        "THEN 'CAPEX' ELSE 'Store' END"
    )

    AGG = """
    ConsQty              = SUM(CASE WHEN P.IsCons=1 THEN ISNULL(P.issue,0) ELSE 0 END),
    ConsValue            = SUM(CASE WHEN P.IsCons=1 THEN ISNULL(P.issue,0)*ISNULL(P.UnitRate,0) ELSE 0 END),
    InwardQty            = SUM(CASE WHEN P.IsGRN=1  THEN ISNULL(P.receipt,0) ELSE 0 END),
    InwardValue          = SUM(CASE WHEN P.IsGRN=1  THEN ISNULL(P.receipt,0)*ISNULL(P.UnitRate,0) ELSE 0 END),
    StkTrfIssueQty       = SUM(CASE WHEN P.IsCons=0 AND ISNULL(P.issue,0)<>0 THEN ISNULL(P.issue,0) ELSE 0 END),
    StkTrfIssueValue     = SUM(CASE WHEN P.IsCons=0 AND ISNULL(P.issue,0)<>0 THEN ISNULL(P.issue,0)*ISNULL(P.UnitRate,0) ELSE 0 END),
    StkTrfInwardQty      = SUM(CASE WHEN P.IsGRN=0  AND ISNULL(P.receipt,0)<>0 THEN ISNULL(P.receipt,0) ELSE 0 END),
    StkTrfInwardValue    = SUM(CASE WHEN P.IsGRN=0  AND ISNULL(P.receipt,0)<>0 THEN ISNULL(P.receipt,0)*ISNULL(P.UnitRate,0) ELSE 0 END)
    """

    def grp_select_full(col: str) -> str:
        return f"""
;WITH
P0 AS (
    SELECT Name = ISNULL(NULLIF(P.{col},''),'(Blank)'), {AGG}
    FROM #P P
    GROUP BY ISNULL(NULLIF(P.{col},''),'(Blank)')
),
O0 AS (
    SELECT Name = ISNULL(NULLIF(O.{col},''),'(Blank)'), OpeningValue = SUM(O.OpeningValue)
    FROM #O O
    GROUP BY ISNULL(NULLIF(O.{col},''),'(Blank)')
),
C0 AS (
    SELECT Name = ISNULL(NULLIF(C.{col},''),'(Blank)'), ClosingQty = SUM(C.ClosingQty), ClosingValue = SUM(C.ClosingValue)
    FROM #C C
    GROUP BY ISNULL(NULLIF(C.{col},''),'(Blank)')
)
SELECT
    Name = COALESCE(O0.Name, P0.Name, C0.Name),
    OpeningValue = ISNULL(O0.OpeningValue,0),
    ConsQty = ISNULL(P0.ConsQty,0),
    ConsValue = ISNULL(P0.ConsValue,0),
    InwardQty = ISNULL(P0.InwardQty,0),
    InwardValue = ISNULL(P0.InwardValue,0),
    StkTrfIssueQty = ISNULL(P0.StkTrfIssueQty,0),
    StkTrfIssueValue = ISNULL(P0.StkTrfIssueValue,0),
    StkTrfInwardQty = ISNULL(P0.StkTrfInwardQty,0),
    StkTrfInwardValue = ISNULL(P0.StkTrfInwardValue,0),
    ClosingQty = ISNULL(C0.ClosingQty,0),
    ClosingValue = ISNULL(C0.ClosingValue,0)
FROM O0
FULL OUTER JOIN P0 ON P0.Name = O0.Name
FULL OUTER JOIN C0 ON C0.Name = COALESCE(O0.Name, P0.Name);
"""

    def itemwise_full() -> str:
        return f"""
;WITH
P0 AS (
    SELECT
        ItemName = ISNULL(NULLIF(P.item_name,''),'(Blank)'),
        InwardQty            = SUM(CASE WHEN P.IsGRN=1 THEN ISNULL(P.receipt,0) ELSE 0 END),
        InwardValue          = SUM(CASE WHEN P.IsGRN=1 THEN ISNULL(P.receipt,0)*ISNULL(P.UnitRate,0) ELSE 0 END),
        StkTrfInwardQty      = SUM(CASE WHEN P.IsGRN=0 AND ISNULL(P.receipt,0)<>0 THEN ISNULL(P.receipt,0) ELSE 0 END),
        StkTrfInwardValue    = SUM(CASE WHEN P.IsGRN=0 AND ISNULL(P.receipt,0)<>0 THEN ISNULL(P.receipt,0)*ISNULL(P.UnitRate,0) ELSE 0 END),
        ConsQty              = SUM(CASE WHEN P.IsCons=1 THEN ISNULL(P.issue,0) ELSE 0 END),
        ConsValue            = SUM(CASE WHEN P.IsCons=1 THEN ISNULL(P.issue,0)*ISNULL(P.UnitRate,0) ELSE 0 END),
        StkTrfIssueQty       = SUM(CASE WHEN P.IsCons=0 AND ISNULL(P.issue,0)<>0 THEN ISNULL(P.issue,0) ELSE 0 END),
        StkTrfIssueValue     = SUM(CASE WHEN P.IsCons=0 AND ISNULL(P.issue,0)<>0 THEN ISNULL(P.issue,0)*ISNULL(P.UnitRate,0) ELSE 0 END)
    FROM #P P
    GROUP BY ISNULL(NULLIF(P.item_name,''),'(Blank)')
),
O0 AS (
    SELECT
        ItemName = ISNULL(NULLIF(O.item_name,''),'(Blank)'),
        OpeningQty = SUM(O.OpeningQty),
        OpeningValue = SUM(O.OpeningValue)
    FROM #O O
    GROUP BY ISNULL(NULLIF(O.item_name,''),'(Blank)')
),
C0 AS (
    SELECT
        ItemName = ISNULL(NULLIF(C.item_name,''),'(Blank)'),
        ClosingQty = SUM(C.ClosingQty),
        ClosingValue = SUM(C.ClosingValue)
    FROM #C C
    GROUP BY ISNULL(NULLIF(C.item_name,''),'(Blank)')
)
SELECT
    ItemName = COALESCE(O0.ItemName, P0.ItemName, C0.ItemName),
    OpeningQty = ISNULL(O0.OpeningQty,0),
    OpeningValue = ISNULL(O0.OpeningValue,0),
    InwardQty = ISNULL(P0.InwardQty,0),
    InwardValue = ISNULL(P0.InwardValue,0),
    StkTrfInwardQty = ISNULL(P0.StkTrfInwardQty,0),
    StkTrfInwardValue = ISNULL(P0.StkTrfInwardValue,0),
    ConsQty = ISNULL(P0.ConsQty,0),
    ConsValue = ISNULL(P0.ConsValue,0),
    StkTrfIssueQty = ISNULL(P0.StkTrfIssueQty,0),
    StkTrfIssueValue = ISNULL(P0.StkTrfIssueValue,0),
    ClosingQty = ISNULL(C0.ClosingQty,0),
    ClosingValue = ISNULL(C0.ClosingValue,0)
FROM O0
FULL OUTER JOIN P0 ON P0.ItemName = O0.ItemName
FULL OUTER JOIN C0 ON C0.ItemName = COALESCE(O0.ItemName, P0.ItemName);
"""

    sql = f"""
IF OBJECT_ID('tempdb..#FMD') IS NOT NULL DROP TABLE #FMD;

SELECT
    X.company_id,
    X.item_code,
    X.batch_no,
    X.first_movement_date
INTO #FMD
FROM (
    SELECT
        S.company_id,
        S.item_code,
        S.batch_no,
        first_movement_date = CASE
            WHEN {EXCESS_BATCH_SQL}
                THEN CONVERT(date, '{FIXED_FROM_DATE}')
            ELSE MIN(S.doc_date)
        END
    FROM dbo.inventory_ageing_snapshot S
    WHERE S.company_id={COMPANY_ID}
      AND S.doc_date >= CONVERT(date,'{FIXED_FROM_DATE}')
      AND ISNULL(S.transaction_type,'') NOT IN ({EXCLUDE_IN_SQL})
    GROUP BY S.company_id, S.item_code, S.batch_no
) X;

CREATE CLUSTERED INDEX IX_FMD ON #FMD(company_id, item_code, batch_no);

DECLARE @ToDate date = CONVERT(date,'{_esc(to_date)}');

IF OBJECT_ID('tempdb..#O') IS NOT NULL DROP TABLE #O;

SELECT
    X.item_code, X.item_name, X.item_type, X.location, X.batch_no,
    X.inventory_category, X.inventory_subcategory, X.first_movement_date,
    X.age_days, X.age_bucket, X.LocationType, X.transaction_type,
    OpeningQty   = SUM(ISNULL(X.closing,0)),
    OpeningValue = SUM(ISNULL(X.closing_value,0))
INTO #O
FROM (
    SELECT
        S.item_code,S.item_name,S.item_type,S.location,S.batch_no,
        S.inventory_category,S.inventory_subcategory, S.transaction_type,
        S.closing,S.closing_value,
        F.first_movement_date,
        AgeCalc.age_days,
        AgeCalc.age_bucket,
        LocCalc.LocationType
    FROM dbo.inventory_ageing_snapshot S
    LEFT JOIN #FMD F
           ON F.company_id=S.company_id
          AND F.item_code=S.item_code
          AND F.batch_no=S.batch_no
    CROSS APPLY (
        SELECT
            age_days = DATEDIFF(DAY, F.first_movement_date, @ToDate),
            age_bucket = CASE
                WHEN S.batch_no LIKE '%excess%' OR S.batch_no IN ('decexcess2023', 'EXCESS STOCK', 'Excessstock','Decexcess2023') THEN '>720'
                WHEN DATEDIFF(DAY, F.first_movement_date, @ToDate) BETWEEN 0 AND 60 THEN '0-60'
                WHEN DATEDIFF(DAY, F.first_movement_date, @ToDate) BETWEEN 61 AND 90 THEN '61-90'
                WHEN DATEDIFF(DAY, F.first_movement_date, @ToDate) BETWEEN 91 AND 160 THEN '91-160'
                WHEN DATEDIFF(DAY, F.first_movement_date, @ToDate) BETWEEN 161 AND 360 THEN '161-360'
                WHEN DATEDIFF(DAY, F.first_movement_date, @ToDate) BETWEEN 361 AND 720 THEN '361-720'
                ELSE '>720' END
    ) AgeCalc
    CROSS APPLY (
        SELECT LocationType = {LT.format(a='S')}
    ) LocCalc
    WHERE S.company_id={COMPANY_ID}
      AND S.doc_date BETWEEN CONVERT(date,'{FIXED_FROM_DATE}') AND CONVERT(date,'{oe}') {iw}
      AND ISNULL(S.transaction_type,'') NOT IN ({EXCLUDE_IN_SQL})
) X
GROUP BY
    X.item_code,X.item_name,X.item_type,X.location,X.batch_no,
    X.inventory_category,X.inventory_subcategory, X.transaction_type,
    X.first_movement_date, X.age_days,X.age_bucket, X.LocationType;

CREATE INDEX IX_O_age_bucket ON #O(age_bucket);
CREATE INDEX IX_O_loc ON #O(location);

IF OBJECT_ID('tempdb..#P') IS NOT NULL DROP TABLE #P;

SELECT
    X.doc_no,X.doc_date,X.company_name,X.item_type,X.item_code,X.item_name,
    X.unit,X.location,X.LocationType,X.stock_location,X.batch_no,
    X.mfg_date,X.retest_date,X.opening,X.receipt,X.issue,X.closing,X.closing_value,
    X.inventory_category,X.inventory_subcategory,X.row_in_batch,
    X.total_closing_per_batch,X.total_closing_value_per_batch,
    X.first_movement_date,
    X.age_days, X.age_bucket, X.transaction_type,
    UnitRate = CASE WHEN NULLIF(X.closing,0) IS NULL THEN 0 ELSE ISNULL(X.closing_value,0)/NULLIF(X.closing,0) END,
    IsCons   = CASE WHEN X.transaction_type IN ({CONSUMPTION_IN_SQL}) THEN 1 ELSE 0 END,
    IsGRN    = CASE WHEN X.transaction_type IN ({INWARD_IN_SQL}) THEN 1 ELSE 0 END
INTO #P
FROM (
    SELECT
        S.doc_no,S.doc_date,S.company_name,S.item_type,S.item_code,S.item_name,
        S.unit,S.location,S.stock_location,S.batch_no,
        S.mfg_date,S.retest_date,S.opening,S.receipt,S.issue,S.closing,S.closing_value,
        S.inventory_category,S.inventory_subcategory,S.row_in_batch,
        S.total_closing_per_batch,S.total_closing_value_per_batch,
        S.transaction_type,
        F.first_movement_date,
        AgeCalc.age_days,
        AgeCalc.age_bucket,
        LocCalc.LocationType
    FROM dbo.inventory_ageing_snapshot S
    LEFT JOIN #FMD F
           ON F.company_id=S.company_id
          AND F.item_code=S.item_code
          AND F.batch_no=S.batch_no
    CROSS APPLY (
        SELECT
            age_days = DATEDIFF(DAY, F.first_movement_date, @ToDate),
            age_bucket = CASE
                WHEN S.batch_no LIKE '%excess%' OR S.batch_no IN ('decexcess2023', 'EXCESS STOCK', 'Excessstock','Decexcess2023') THEN '>720'
                WHEN DATEDIFF(DAY, F.first_movement_date, @ToDate) BETWEEN 0 AND 60 THEN '0-60'
                WHEN DATEDIFF(DAY, F.first_movement_date, @ToDate) BETWEEN 61 AND 90 THEN '61-90'
                WHEN DATEDIFF(DAY, F.first_movement_date, @ToDate) BETWEEN 91 AND 160 THEN '91-160'
                WHEN DATEDIFF(DAY, F.first_movement_date, @ToDate) BETWEEN 161 AND 360 THEN '161-360'
                WHEN DATEDIFF(DAY, F.first_movement_date, @ToDate) BETWEEN 361 AND 720 THEN '361-720'
                ELSE '>720' END
    ) AgeCalc
    CROSS APPLY (
        SELECT LocationType = {LT.format(a='S')}
    ) LocCalc
    WHERE S.company_id={COMPANY_ID}
      AND S.doc_date BETWEEN CONVERT(date,'{from_date}') AND CONVERT(date,'{to_date}') {pw}
      AND ISNULL(S.transaction_type,'') NOT IN ({EXCLUDE_IN_SQL})
) X;

CREATE INDEX IX_P_age_bucket ON #P(age_bucket);
CREATE INDEX IX_P_itemname ON #P(item_name);
CREATE INDEX IX_P_location ON #P(location);

IF OBJECT_ID('tempdb..#C') IS NOT NULL DROP TABLE #C;

SELECT
    X.item_code, X.item_name, X.item_type, X.location, X.batch_no,
    X.inventory_category, X.inventory_subcategory, X.first_movement_date,
    X.age_days, X.age_bucket, X.LocationType, X.transaction_type,
    ClosingQty   = SUM(ISNULL(X.closing,0)),
    ClosingValue = SUM(ISNULL(X.closing_value,0))
INTO #C
FROM (
    SELECT
        S.item_code,S.item_name,S.item_type,S.location,S.batch_no,
        S.inventory_category,S.inventory_subcategory, S.transaction_type,
        S.closing,S.closing_value,
        F.first_movement_date,
        AgeCalc.age_days,
        AgeCalc.age_bucket,
        LocCalc.LocationType
    FROM dbo.inventory_ageing_snapshot S
    LEFT JOIN #FMD F
           ON F.company_id=S.company_id
          AND F.item_code=S.item_code
          AND F.batch_no=S.batch_no
    CROSS APPLY (
        SELECT
            age_days = DATEDIFF(DAY, F.first_movement_date, @ToDate),
            age_bucket = CASE
                WHEN S.batch_no LIKE '%excess%' OR S.batch_no IN ('decexcess2023', 'EXCESS STOCK', 'Excessstock','Decexcess2023') THEN '>720'
                WHEN DATEDIFF(DAY, F.first_movement_date, @ToDate) BETWEEN 0 AND 60 THEN '0-60'
                WHEN DATEDIFF(DAY, F.first_movement_date, @ToDate) BETWEEN 61 AND 90 THEN '61-90'
                WHEN DATEDIFF(DAY, F.first_movement_date, @ToDate) BETWEEN 91 AND 160 THEN '91-160'
                WHEN DATEDIFF(DAY, F.first_movement_date, @ToDate) BETWEEN 161 AND 360 THEN '161-360'
                WHEN DATEDIFF(DAY, F.first_movement_date, @ToDate) BETWEEN 361 AND 720 THEN '361-720'
                ELSE '>720' END
    ) AgeCalc
    CROSS APPLY (
        SELECT LocationType = {LT.format(a='S')}
    ) LocCalc
    WHERE S.company_id={COMPANY_ID}
      AND S.doc_date BETWEEN CONVERT(date,'{FIXED_FROM_DATE}') AND CONVERT(date,'{to_date}') {iw}
      AND ISNULL(S.transaction_type,'') NOT IN ({EXCLUDE_IN_SQL})
) X
GROUP BY
    X.item_code,X.item_name,X.item_type,X.location,X.batch_no,
    X.inventory_category,X.inventory_subcategory, X.transaction_type,
    X.first_movement_date, X.age_days,X.age_bucket, X.LocationType;

CREATE INDEX IX_C_age_bucket ON #C(age_bucket);
CREATE INDEX IX_C_loc ON #C(location);

/* SELECT 0: KPI */
SELECT
    OpeningQty=ISNULL((SELECT SUM(OpeningQty) FROM #O),0),
    OpeningValue=ISNULL((SELECT SUM(OpeningValue) FROM #O),0),
    {AGG},
    ClosingQty=ISNULL((SELECT SUM(ClosingQty) FROM #C),0),
    ClosingValue=ISNULL((SELECT SUM(ClosingValue) FROM #C),0),
    TotalRows=(SELECT COUNT(1) FROM #P)
FROM #P P;

/* SELECT 1: Sanity */
SELECT Cnt=COUNT(1) FROM #P;

/* SELECT 2: Age bucket */
;WITH
P0 AS (
    SELECT AgeBucket = P.age_bucket, {AGG}
    FROM #P P
    GROUP BY P.age_bucket
),
O0 AS (
    SELECT AgeBucket = O.age_bucket, OpeningValue = SUM(O.OpeningValue)
    FROM #O O
    GROUP BY O.age_bucket
),
C0 AS (
    SELECT AgeBucket = C.age_bucket, ClosingQty = SUM(C.ClosingQty), ClosingValue = SUM(C.ClosingValue)
    FROM #C C
    GROUP BY C.age_bucket
)
SELECT
    AgeBucket = COALESCE(O0.AgeBucket, P0.AgeBucket, C0.AgeBucket),
    OpeningValue = ISNULL(O0.OpeningValue,0),
    ConsQty = ISNULL(P0.ConsQty,0),
    ConsValue = ISNULL(P0.ConsValue,0),
    InwardQty = ISNULL(P0.InwardQty,0),
    InwardValue = ISNULL(P0.InwardValue,0),
    StkTrfIssueQty = ISNULL(P0.StkTrfIssueQty,0),
    StkTrfIssueValue = ISNULL(P0.StkTrfIssueValue,0),
    StkTrfInwardQty = ISNULL(P0.StkTrfInwardQty,0),
    StkTrfInwardValue = ISNULL(P0.StkTrfInwardValue,0),
    ClosingQty = ISNULL(C0.ClosingQty,0),
    ClosingValue = ISNULL(C0.ClosingValue,0)
FROM O0
FULL OUTER JOIN P0 ON P0.AgeBucket = O0.AgeBucket
FULL OUTER JOIN C0 ON C0.AgeBucket = COALESCE(O0.AgeBucket, P0.AgeBucket);

/* SELECT 3: Location */
{grp_select_full('location')}

/* SELECT 4: Location Type */
{grp_select_full('LocationType')}

/* SELECT 5: Item Type */
{grp_select_full('item_type')}

/* SELECT 6: Inv Category */
{grp_select_full('inventory_category')}

/* SELECT 7: Inv Subcategory */
{grp_select_full('inventory_subcategory')}

/* SELECT 8: Batch */
{grp_select_full('batch_no')}

/* SELECT 9: Item-wise */
{itemwise_full()}

/* SELECT 10: Transaction Type */
{grp_select_full('transaction_type')}

/* SELECT 11: Detail */
SELECT TOP(500)
    DocNo=S.doc_no,
    DocDate=CONVERT(varchar(10),S.doc_date,105),
    ItemType=S.item_type,
    ItemCode=S.item_code,
    ItemName=S.item_name,
    Unit=S.unit,
    Location=S.location,
    LocationType={LT.format(a='S')},
    BatchNo=S.batch_no,
    TransactionType=S.transaction_type,
    Opening=S.opening,
    Receipt=S.receipt,
    Issue=S.issue,
    Closing=S.closing,
    ClosingValue=S.closing_value,
    UnitRate=CASE WHEN NULLIF(S.closing,0) IS NULL THEN 0 ELSE ISNULL(S.closing_value,0)/NULLIF(S.closing,0) END,
    ConsQty=CASE WHEN S.transaction_type IN ({CONSUMPTION_IN_SQL}) THEN ISNULL(S.issue,0) ELSE 0 END,
    ConsValue=CASE WHEN S.transaction_type IN ({CONSUMPTION_IN_SQL}) THEN ISNULL(S.issue,0)*(CASE WHEN NULLIF(S.closing,0) IS NULL THEN 0 ELSE ISNULL(S.closing_value,0)/NULLIF(S.closing,0) END) ELSE 0 END,
    InwardQty=CASE WHEN S.transaction_type IN ({INWARD_IN_SQL}) THEN ISNULL(S.receipt,0) ELSE 0 END,
    InwardValue=CASE WHEN S.transaction_type IN ({INWARD_IN_SQL}) THEN ISNULL(S.receipt,0)*(CASE WHEN NULLIF(S.closing,0) IS NULL THEN 0 ELSE ISNULL(S.closing_value,0)/NULLIF(S.closing,0) END) ELSE 0 END,
    StkTrfIssueQty=CASE WHEN S.transaction_type NOT IN ({CONSUMPTION_IN_SQL}) AND ISNULL(S.issue,0)<>0 THEN ISNULL(S.issue,0) ELSE 0 END,
    StkTrfIssueValue=CASE WHEN S.transaction_type NOT IN ({CONSUMPTION_IN_SQL}) AND ISNULL(S.issue,0)<>0 THEN ISNULL(S.issue,0)*(CASE WHEN NULLIF(S.closing,0) IS NULL THEN 0 ELSE ISNULL(S.closing_value,0)/NULLIF(S.closing,0) END) ELSE 0 END,
    StkTrfInwardQty=CASE WHEN S.transaction_type NOT IN ({INWARD_IN_SQL}) AND ISNULL(S.receipt,0)<>0 THEN ISNULL(S.receipt,0) ELSE 0 END,
    StkTrfInwardValue=CASE WHEN S.transaction_type NOT IN ({INWARD_IN_SQL}) AND ISNULL(S.receipt,0)<>0 THEN ISNULL(S.receipt,0)*(CASE WHEN NULLIF(S.closing,0) IS NULL THEN 0 ELSE ISNULL(S.closing_value,0)/NULLIF(S.closing,0) END) ELSE 0 END,
    AgeDays=DATEDIFF(DAY, F.first_movement_date, @ToDate),
    AgeBucket={_age_bucket_case_for_where("DATEDIFF(DAY, F.first_movement_date, @ToDate)", "S")},
    FirstMovementDate=CONVERT(varchar(10),F.first_movement_date,105)
FROM dbo.inventory_ageing_snapshot S
LEFT JOIN #FMD F
       ON F.company_id=S.company_id
      AND F.item_code=S.item_code
      AND F.batch_no=S.batch_no
WHERE S.company_id={COMPANY_ID}
  AND S.doc_date BETWEEN CONVERT(date,'{FIXED_FROM_DATE}') AND CONVERT(date,'{to_date}') {pw}
  AND ISNULL(S.transaction_type,'') NOT IN ({EXCLUDE_IN_SQL})
ORDER BY S.batch_no,S.mfg_date,S.location,S.doc_date,S.doc_no;

DROP TABLE IF EXISTS #FMD;
DROP TABLE IF EXISTS #O;
DROP TABLE IF EXISTS #P;
DROP TABLE IF EXISTS #C;
"""

    sets, dp = _exec_sql_snapshot_multi(sql, debug=dbg)

    def _g(i):
        if 0 <= i < len(sets):
            return sets[i].get("rows") or [], sets[i].get("cols") or []
        return [], []

    EPS = 0.000001

    def _has_any_value(*vals) -> bool:
        try:
            for v in vals:
                if abs(float(v or 0.0)) > EPS:
                    return True
        except Exception:
            return False
        return False

    kr, _ = _g(0)
    kpi = {}
    if kr:
        r = kr[0]
        kpi = {
            "opening_qty": float(r[0] or 0),
            "opening_value": float(r[1] or 0),
            "consumption_qty": float(r[2] or 0),
            "consumption_value": float(r[3] or 0),
            "inward_qty": float(r[4] or 0),
            "inward_value": float(r[5] or 0),
            "stk_trf_issue_qty": float(r[6] or 0),
            "stk_trf_issue_value": float(r[7] or 0),
            "stk_trf_inward_qty": float(r[8] or 0),
            "stk_trf_inward_value": float(r[9] or 0),
            "closing_qty": float(r[10] or 0),
            "closing_value": float(r[11] or 0),
            "total_rows": int(r[12] or 0),
        }

    def _pg(rows, nk="name"):
        out = []
        for r in rows or []:
            name = r[0] or "(Blank)"
            ov = float(r[1] or 0)
            cq = float(r[2] or 0)
            cv = float(r[3] or 0)
            iq = float(r[4] or 0)
            iv = float(r[5] or 0)
            soq = float(r[6] or 0)
            sov = float(r[7] or 0)
            siq = float(r[8] or 0)
            siv = float(r[9] or 0)
            clq = float(r[10] or 0)
            clv = float(r[11] or 0)

            if not _has_any_value(ov, clv, cv, iv, sov, siv, cq, iq, soq, siq, clq):
                continue

            out.append(
                {
                    nk: name,
                    "opening_value": ov,
                    "consumption_qty": cq,
                    "consumption_value": cv,
                    "inward_qty": iq,
                    "inward_value": iv,
                    "stk_trf_issue_qty": soq,
                    "stk_trf_issue_value": sov,
                    "stk_trf_inward_qty": siq,
                    "stk_trf_inward_value": siv,
                    "closing_qty": clq,
                    "closing_value": clv,
                    "total_value": clv,
                }
            )

        out.sort(key=lambda x: x["closing_value"], reverse=True)
        return out

    ar, _ = _g(2)
    ao = ["0-60", "61-90", "91-160", "161-360", "361-720", ">720"]
    am = {r[0]: r for r in (ar or [])}
    age_summary = []
    for s in ao:
        r = am.get(s)
        if not r:
            continue
        ov = float(r[1] or 0)
        cq = float(r[2] or 0)
        cv = float(r[3] or 0)
        iq = float(r[4] or 0)
        iv = float(r[5] or 0)
        soq = float(r[6] or 0)
        sov = float(r[7] or 0)
        siq = float(r[8] or 0)
        siv = float(r[9] or 0)
        clq = float(r[10] or 0)
        clv = float(r[11] or 0)

        if not _has_any_value(ov, clv, cv, iv, sov, siv, cq, iq, soq, siq, clq):
            continue

        age_summary.append(
            {
                "slug": s,
                "label": s.replace("-", "\u2013"),
                "opening_value": ov,
                "consumption_qty": cq,
                "consumption_value": cv,
                "inward_qty": iq,
                "inward_value": iv,
                "stk_trf_issue_qty": soq,
                "stk_trf_issue_value": sov,
                "stk_trf_inward_qty": siq,
                "stk_trf_inward_value": siv,
                "closing_qty": clq,
                "closing_value": clv,
                "total_value": clv,
            }
        )

    location_summary = _pg(_g(3)[0])
    location_type_summary = _pg(_g(4)[0])
    item_type_summary = _pg(_g(5)[0])
    invcat_summary = _pg(_g(6)[0])
    invsubcat_summary = _pg(_g(7)[0])
    batch_summary = _pg(_g(8)[0], "batch_no")
    txntype_summary = _pg(_g(10)[0], "transaction_type")

    iwr, _ = _g(9)
    itemwise_summary = []
    for r in iwr or []:
        name = r[0] or "(Blank)"
        oq = float(r[1] or 0)
        ov = float(r[2] or 0)
        iq = float(r[3] or 0)
        iv = float(r[4] or 0)
        siq = float(r[5] or 0)
        siv = float(r[6] or 0)
        cq = float(r[7] or 0)
        cv = float(r[8] or 0)
        soq = float(r[9] or 0)
        sov = float(r[10] or 0)
        clq = float(r[11] or 0)
        clv = float(r[12] or 0)

        if not _has_any_value(ov, clv, cv, iv, sov, siv, oq, iq, cq, soq, siq, clq):
            continue

        itemwise_summary.append(
            {
                "item_name": name,
                "opening_qty": oq,
                "opening_value": ov,
                "inward_qty": iq,
                "inward_value": iv,
                "stk_trf_inward_qty": siq,
                "stk_trf_inward_value": siv,
                "consumption_qty": cq,
                "consumption_value": cv,
                "stk_trf_issue_qty": soq,
                "stk_trf_issue_value": sov,
                "closing_qty": clq,
                "closing_value": clv,
                "total_value": clv,
            }
        )
    itemwise_summary.sort(key=lambda x: x["closing_value"], reverse=True)

    dr, dc = _g(11)
    detail_rows = [{c: _json_safe(v) for c, v in zip(dc, row)} for row in (dr or [])]

    # ======================= GRAPH =======================
    try:
        g_from = datetime.strptime(from_date, "%Y-%m-%d").date()
    except Exception:
        g_from = datetime.strptime(FIXED_FROM_DATE, "%Y-%m-%d").date()

    try:
        g_to = datetime.strptime(to_date, "%Y-%m-%d").date()
    except Exception:
        g_to = date.today()

    base_graph_date = datetime.strptime(FIXED_FROM_DATE, "%Y-%m-%d").date()

    graph_where = _build_extra_where(filters, "S", debug=dbg, to_date=to_date, use_fmd=True, fmd_alias="F")

    graph_sql = f"""
DECLARE @CompanyId int = {COMPANY_ID};
DECLARE @BaseDate date = '{base_graph_date.isoformat()}';
DECLARE @GraphStart date = '{g_from.isoformat()}';
DECLARE @GraphEnd   date = '{g_to.isoformat()}';

IF OBJECT_ID('tempdb..#FMD') IS NOT NULL DROP TABLE #FMD;

SELECT
    X.company_id,
    X.item_code,
    X.batch_no,
    X.first_movement_date
INTO #FMD
FROM (
    SELECT
        S.company_id,
        S.item_code,
        S.batch_no,
        first_movement_date = CASE
            WHEN {EXCESS_BATCH_SQL}
                THEN CONVERT(date, '{FIXED_FROM_DATE}')
            ELSE MIN(S.doc_date)
        END
    FROM dbo.inventory_ageing_snapshot S
    WHERE S.company_id = @CompanyId
      AND S.doc_date >= @BaseDate
      AND ISNULL(S.transaction_type,'') NOT IN ({EXCLUDE_IN_SQL})
    GROUP BY S.company_id, S.item_code, S.batch_no
) X;

CREATE CLUSTERED INDEX IX_FMD ON #FMD(company_id, item_code, batch_no);

DECLARE @BaseCumValue decimal(38,6) =
(
    SELECT ISNULL(SUM(ISNULL(S.closing_value,0)),0)
    FROM dbo.inventory_ageing_snapshot S
    LEFT JOIN #FMD F
      ON F.company_id = S.company_id
     AND F.item_code  = S.item_code
     AND F.batch_no   = S.batch_no
    WHERE S.company_id = @CompanyId
      AND S.doc_date >= @BaseDate
      AND S.doc_date <  @GraphStart
      AND ISNULL(S.transaction_type,'') NOT IN ({EXCLUDE_IN_SQL})
      {graph_where}
);

;WITH MonthEnds AS (
    SELECT EOMONTH(@GraphStart) AS MonthEnd
    UNION ALL
    SELECT EOMONTH(DATEADD(MONTH, 1, MonthEnd))
    FROM MonthEnds
    WHERE MonthEnd < EOMONTH(@GraphEnd)
),
MonthlyAgg AS (
    SELECT
        MonthEnd = EOMONTH(S.doc_date),

        MonthlyClosingValue = SUM(ISNULL(S.closing_value,0)),

        MonthlyInwardValue = SUM(
            CASE
                WHEN S.transaction_type IN ({INWARD_IN_SQL})
                THEN ISNULL(S.receipt,0) *
                     (CASE WHEN NULLIF(S.closing,0) IS NULL THEN 0
                           ELSE ISNULL(S.closing_value,0)/NULLIF(S.closing,0) END)
                ELSE 0
            END
        ),

        MonthlyConsValue = SUM(
            CASE
                WHEN S.transaction_type IN ({CONSUMPTION_IN_SQL})
                THEN ISNULL(S.issue,0) *
                     (CASE WHEN NULLIF(S.closing,0) IS NULL THEN 0
                           ELSE ISNULL(S.closing_value,0)/NULLIF(S.closing,0) END)
                ELSE 0
            END
        )
    FROM dbo.inventory_ageing_snapshot S
    LEFT JOIN #FMD F
      ON F.company_id = S.company_id
     AND F.item_code  = S.item_code
     AND F.batch_no   = S.batch_no
    WHERE S.company_id = @CompanyId
      AND S.doc_date >= @GraphStart
      AND S.doc_date <= @GraphEnd
      AND ISNULL(S.transaction_type,'') NOT IN ({EXCLUDE_IN_SQL})
      {graph_where}
    GROUP BY EOMONTH(S.doc_date)
),
Merged AS (
    SELECT
        M.MonthEnd,
        MonthlyClosingValue = ISNULL(A.MonthlyClosingValue,0),
        MonthlyInwardValue  = ISNULL(A.MonthlyInwardValue,0),
        MonthlyConsValue    = ISNULL(A.MonthlyConsValue,0)
    FROM MonthEnds M
    LEFT JOIN MonthlyAgg A ON A.MonthEnd = M.MonthEnd
),
Final AS (
    SELECT
        MonthEnd,
        ClosingValue = @BaseCumValue
            + SUM(MonthlyClosingValue) OVER (ORDER BY MonthEnd ROWS UNBOUNDED PRECEDING),
        InwardValue = MonthlyInwardValue,
        ConsumptionValue = MonthlyConsValue
    FROM Merged
)
SELECT MonthEnd, ClosingValue, InwardValue, ConsumptionValue
FROM Final
ORDER BY MonthEnd
OPTION (MAXRECURSION 400);

DROP TABLE IF EXISTS #FMD;
"""

    grow, _gcols = _exec_sql_snapshot(graph_sql)

    labels, closing_vals, inward_vals, cons_vals = [], [], [], []
    for r in grow or []:
        me = r[0]
        try:
            me_dt = me if isinstance(me, date) else datetime.strptime(str(me)[:10], "%Y-%m-%d").date()
        except Exception:
            me_dt = None

        labels.append(me_dt.strftime("%b %y") if me_dt else str(me))
        closing_vals.append(float(r[1] or 0))
        inward_vals.append(float(r[2] or 0))
        cons_vals.append(float(r[3] or 0))
    # ======================= END GRAPH =======================

    data = {
        "kpi": kpi,
        "age_summary": age_summary,
        "location_summary": location_summary,
        "item_type_summary": item_type_summary,
        "batch_summary": batch_summary,
        "txntype_summary": txntype_summary,
        "itemwise_summary": itemwise_summary,
        "location_type_summary": location_type_summary,
        "invcat_summary": invcat_summary,
        "invsubcat_summary": invsubcat_summary,
        "monthly_chart": {
            "labels": labels,
            "closing_values": closing_vals,
            "inward_values": inward_vals,
            "consumption_values": cons_vals,
            "values": closing_vals,
        },
        "detail_rows": detail_rows,
    }

    if dbg and dp:
        data["debug"] = dp

    return JsonResponse(data)



@login_required
@require_GET
def inventory_ageing_export_csv(request):
    from_date, to_date, *_ = _get_date_range(request)
    filters = _get_filters(request)
    dbg = request.GET.get("debug") == "1"

    pw = _build_extra_where(filters, "S", debug=dbg, to_date=to_date, use_fmd=True, fmd_alias="F")

    sql = f"""
;WITH FMD AS (
    SELECT
        X.company_id,
        X.item_code,
        X.batch_no,
        X.first_movement_date
    FROM (
        SELECT
            S.company_id,
            S.item_code,
            S.batch_no,
            first_movement_date = CASE
                WHEN {EXCESS_BATCH_SQL}
                    THEN CONVERT(date, '{FIXED_FROM_DATE}')
                ELSE MIN(S.doc_date)
            END
        FROM dbo.inventory_ageing_snapshot S
        WHERE S.company_id={COMPANY_ID}
          AND S.doc_date >= CONVERT(date,'{FIXED_FROM_DATE}')
          AND ISNULL(S.transaction_type,'') NOT IN ({EXCLUDE_IN_SQL})
        GROUP BY S.company_id, S.item_code, S.batch_no
    ) X
)
SELECT
S.doc_no AS [Doc No],
CONVERT(VARCHAR(10),S.doc_date,105) AS [Doc Date],
S.company_name AS [Company],
S.item_type AS [Item Type],
S.item_code AS [Item Code],
S.item_name AS [Item Name],
S.unit AS [Unit],
S.location AS [Location],
CASE WHEN S.location LIKE 'CAP%' OR S.location='Solapur E-20 Capex Store' THEN 'CAPEX' ELSE 'Store' END AS [Location Type],
S.stock_location AS [Stock Location],
S.batch_no AS [Batch No],
CONVERT(VARCHAR(10),S.mfg_date,105) AS [Mfg Date],
CONVERT(VARCHAR(10),S.retest_date,105) AS [Retest Date],
S.opening AS [Opening Qty],
S.receipt AS [Receipt Qty],
S.issue AS [Issue Qty],
S.closing AS [Closing Qty],
S.closing_value AS [Closing Value],
S.inventory_category AS [Inventory Category],
S.inventory_subcategory AS [Inventory Subcategory],
DATEDIFF(DAY, F.first_movement_date, CONVERT(date,'{_esc(to_date)}')) AS [Age Days],
CASE
    WHEN S.batch_no LIKE '%excess%' OR S.batch_no IN ('decexcess2023', 'EXCESS STOCK', 'Excessstock', 'Decexcess2023') THEN '>720'
    WHEN DATEDIFF(DAY, F.first_movement_date, CONVERT(date,'{_esc(to_date)}')) BETWEEN 0 AND 60 THEN '0-60'
    WHEN DATEDIFF(DAY, F.first_movement_date, CONVERT(date,'{_esc(to_date)}')) BETWEEN 61 AND 90 THEN '61-90'
    WHEN DATEDIFF(DAY, F.first_movement_date, CONVERT(date,'{_esc(to_date)}')) BETWEEN 91 AND 160 THEN '91-160'
    WHEN DATEDIFF(DAY, F.first_movement_date, CONVERT(date,'{_esc(to_date)}')) BETWEEN 161 AND 360 THEN '161-360'
    WHEN DATEDIFF(DAY, F.first_movement_date, CONVERT(date,'{_esc(to_date)}')) BETWEEN 361 AND 720 THEN '361-720'
    ELSE '>720' END AS [Age Bucket],
CONVERT(VARCHAR(10),F.first_movement_date,105) AS [First Movement Date],
S.transaction_type AS [Transaction Type],
CASE WHEN NULLIF(S.closing,0) IS NULL THEN 0 ELSE ISNULL(S.closing_value,0)/NULLIF(S.closing,0) END AS [Unit Rate],
CASE WHEN S.transaction_type IN ({CONSUMPTION_IN_SQL}) THEN ISNULL(S.issue,0) ELSE 0 END AS [Consumption Qty],
CASE WHEN S.transaction_type IN ({CONSUMPTION_IN_SQL}) THEN ISNULL(S.issue,0) * CASE WHEN NULLIF(S.closing,0) IS NULL THEN 0 ELSE ISNULL(S.closing_value,0)/NULLIF(S.closing,0) END ELSE 0 END AS [Consumption Value],
CASE WHEN S.transaction_type IN ({INWARD_IN_SQL}) THEN ISNULL(S.receipt,0) ELSE 0 END AS [Inward Qty],
CASE WHEN S.transaction_type IN ({INWARD_IN_SQL}) THEN ISNULL(S.receipt,0) * CASE WHEN NULLIF(S.closing,0) IS NULL THEN 0 ELSE ISNULL(S.closing_value,0)/NULLIF(S.closing,0) END ELSE 0 END AS [Inward Value],
CASE WHEN S.transaction_type NOT IN ({CONSUMPTION_IN_SQL}) AND ISNULL(S.issue,0)<>0 THEN ISNULL(S.issue,0) ELSE 0 END AS [Stk Trf Issue Qty],
CASE WHEN S.transaction_type NOT IN ({CONSUMPTION_IN_SQL}) AND ISNULL(S.issue,0)<>0 THEN ISNULL(S.issue,0) * CASE WHEN NULLIF(S.closing,0) IS NULL THEN 0 ELSE ISNULL(S.closing_value,0)/NULLIF(S.closing,0) END ELSE 0 END AS [Stk Trf Issue Value],
CASE WHEN S.transaction_type NOT IN ({INWARD_IN_SQL}) AND ISNULL(S.receipt,0)<>0 THEN ISNULL(S.receipt,0) ELSE 0 END AS [Stk Trf Inward Qty],
CASE WHEN S.transaction_type NOT IN ({INWARD_IN_SQL}) AND ISNULL(S.receipt,0)<>0 THEN ISNULL(S.receipt,0) * CASE WHEN NULLIF(S.closing,0) IS NULL THEN 0 ELSE ISNULL(S.closing_value,0)/NULLIF(S.closing,0) END ELSE 0 END AS [Stk Trf Inward Value]
FROM dbo.inventory_ageing_snapshot S
LEFT JOIN FMD F
       ON F.company_id=S.company_id
      AND F.item_code=S.item_code
      AND F.batch_no=S.batch_no
WHERE S.company_id={COMPANY_ID}
  AND S.doc_date BETWEEN CONVERT(date,'{FIXED_FROM_DATE}') AND CONVERT(date,'{to_date}') {pw}
  AND ISNULL(S.transaction_type,'') NOT IN ({EXCLUDE_IN_SQL})
ORDER BY S.batch_no,S.mfg_date,S.location,S.doc_date,S.doc_no
OPTION(RECOMPILE);
"""

    rows, cols = _exec_sql_snapshot(sql)
    buf = StringIO()
    w = csv.writer(buf)
    w.writerow(cols)
    for row in rows:
        w.writerow([_json_safe(v) if v is not None else "" for v in row])

    resp = HttpResponse(buf.getvalue(), content_type="text/csv")
    resp["Content-Disposition"] = f'attachment; filename="inventory_ageing_detail_{to_date}.csv"'
    return resp


@login_required 
@require_POST
def inventory_ageing_rebuild_snapshot(request):
    try:
        with transaction.atomic():
            with connections["default"].cursor() as cur:
                cur.execute("EXEC dbo.usp_rebuild_inventory_ageing_snapshot")
                
        messages.success(request, f"Snapshot rebuilt successfully up to {date.today().isoformat()}.")
    except Exception as e:
        messages.error(request, f"Database Error: {e}")
        
    return redirect("erp_reports:inventory_ageing_report")


# ========================= XLSX helpers =========================


def _safe_sheet_name(name: str) -> str:
    name = (name or "Sheet").strip()
    name = name.replace("/", "-").replace("\\", "-").replace(":", "-")
    return name[:31] if len(name) > 31 else name


def _auto_fit_columns(ws, min_w=10, max_w=45):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            v = cell.value
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[col_letter].width = max(min_w, min(max_w, max_len + 2))


def _style_table(ws, header_row=1, freeze_row=1):
    header_fill = PatternFill("solid", fgColor="EEF2FF")
    header_font = Font(bold=True, color="0F172A")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    max_row = ws.max_row
    max_col = ws.max_column

    for c in range(1, max_col + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for r in range(header_row + 1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            if c == 1:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=False)

    ws.freeze_panes = ws[f"A{freeze_row+1}"]
    ws.auto_filter.ref = ws.dimensions


def _write_sheet_from_rows(
    wb,
    title: str,
    rows: list[dict],
    header_order: list[str] | None = None,
    currency_keys: set[str] | None = None,
    qty_keys: set[str] | None = None,
):
    ws = wb.create_sheet(_safe_sheet_name(title))

    if not rows:
        ws.append(["No data"])
        _auto_fit_columns(ws)
        return

    headers = header_order if header_order else list(rows[0].keys())
    ws.append(headers)

    for r in rows:
        ws.append([(r.get(h, "") if r.get(h, None) is not None else "") for h in headers])

    currency_keys = currency_keys or set()
    qty_keys = qty_keys or set()

    header_to_col = {h: idx + 1 for idx, h in enumerate(headers)}

    for key in currency_keys:
        if key in header_to_col:
            col = header_to_col[key]
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0.00"

    for key in qty_keys:
        if key in header_to_col:
            col = header_to_col[key]
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0.000"

    _style_table(ws)
    _auto_fit_columns(ws)


@login_required
@require_GET
def inventory_ageing_export_xlsx(request):
    api_resp = inventory_ageing_summary_api(request)
    if api_resp.status_code != 200:
        return HttpResponse("Failed to export (API error).", status=500)

    try:
        payload = json.loads(api_resp.content.decode("utf-8"))
    except Exception:
        return HttpResponse("Failed to export (invalid JSON).", status=500)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ------------------- KPI sheet -------------------
    k = payload.get("kpi") or {}
    kpi_rows = [
        {"Metric": "Opening Qty", "Value": k.get("opening_qty", 0)},
        {"Metric": "Opening Value", "Value": k.get("opening_value", 0)},
        {"Metric": "Inward Qty", "Value": k.get("inward_qty", 0)},
        {"Metric": "Inward Value", "Value": k.get("inward_value", 0)},
        {"Metric": "StkTrf Inward Qty", "Value": k.get("stk_trf_inward_qty", 0)},
        {"Metric": "StkTrf Inward Value", "Value": k.get("stk_trf_inward_value", 0)},
        {"Metric": "Consumption Qty", "Value": k.get("consumption_qty", 0)},
        {"Metric": "Consumption Value", "Value": k.get("consumption_value", 0)},
        {"Metric": "StkTrf Issue Qty", "Value": k.get("stk_trf_issue_qty", 0)},
        {"Metric": "StkTrf Issue Value", "Value": k.get("stk_trf_issue_value", 0)},
        {"Metric": "Closing Qty", "Value": k.get("closing_qty", 0)},
        {"Metric": "Closing Value", "Value": k.get("closing_value", 0)},
        {"Metric": "Total Rows", "Value": k.get("total_rows", 0)},
    ]
    _write_sheet_from_rows(wb, "KPI", kpi_rows, header_order=["Metric", "Value"])

    # ------------------- Summary sheets -------------------
    summary_currency = {
        "opening_value",
        "inward_value",
        "stk_trf_inward_value",
        "consumption_value",
        "stk_trf_issue_value",
        "closing_value",
    }
    summary_headers = [
        "name",
        "opening_value",
        "inward_value",
        "stk_trf_inward_value",
        "consumption_value",
        "stk_trf_issue_value",
        "closing_qty",
        "closing_value",
    ]

    _write_sheet_from_rows(
        wb,
        "Age Summary",
        payload.get("age_summary") or [],
        header_order=[
            "slug",
            "label",
            "opening_value",
            "inward_value",
            "stk_trf_inward_value",
            "consumption_value",
            "stk_trf_issue_value",
            "closing_qty",
            "closing_value",
        ],
        currency_keys=summary_currency,
        qty_keys={"closing_qty"}
    )

    _write_sheet_from_rows(
        wb,
        "Location Summary",
        payload.get("location_summary") or [],
        header_order=summary_headers,
        currency_keys=summary_currency,
        qty_keys={"closing_qty"}
    )

    _write_sheet_from_rows(
        wb,
        "ItemType Summary",
        payload.get("item_type_summary") or [],
        header_order=summary_headers,
        currency_keys=summary_currency,
        qty_keys={"closing_qty"}
    )

    _write_sheet_from_rows(
        wb,
        "Batch Summary",
        payload.get("batch_summary") or [],
        header_order=[
            "batch_no",
            "opening_value",
            "inward_value",
            "stk_trf_inward_value",
            "consumption_value",
            "stk_trf_issue_value",
            "closing_qty",
            "closing_value",
        ],
        currency_keys=summary_currency,
        qty_keys={"closing_qty"}
    )
    
    _write_sheet_from_rows(
        wb,
        "TxnType Summary",
        payload.get("txntype_summary") or [],
        header_order=[
            "transaction_type",
            "opening_value",
            "inward_value",
            "stk_trf_inward_value",
            "consumption_value",
            "stk_trf_issue_value",
            "closing_qty",
            "closing_value",
        ],
        currency_keys=summary_currency,
        qty_keys={"closing_qty"}
    )

    _write_sheet_from_rows(
        wb,
        "InvCat Summary",
        payload.get("invcat_summary") or [],
        header_order=summary_headers,
        currency_keys=summary_currency,
        qty_keys={"closing_qty"}
    )

    _write_sheet_from_rows(
        wb,
        "InvSubCat Summary",
        payload.get("invsubcat_summary") or [],
        header_order=summary_headers,
        currency_keys=summary_currency,
        qty_keys={"closing_qty"}
    )

    # ------------------- Itemwise -------------------
    _write_sheet_from_rows(
        wb,
        "Itemwise",
        payload.get("itemwise_summary") or [],
        header_order=[
            "item_name",
            "opening_qty",
            "opening_value",
            "inward_qty",
            "inward_value",
            "stk_trf_inward_qty",
            "stk_trf_inward_value",
            "consumption_qty",
            "consumption_value",
            "stk_trf_issue_qty",
            "stk_trf_issue_value",
            "closing_qty",
            "closing_value",
        ],
        currency_keys={
            "opening_value",
            "inward_value",
            "stk_trf_inward_value",
            "consumption_value",
            "stk_trf_issue_value",
            "closing_value",
        },
        qty_keys={
            "opening_qty",
            "inward_qty",
            "stk_trf_inward_qty",
            "consumption_qty",
            "stk_trf_issue_qty",
            "closing_qty",
        },
    )

    # ------------------- Detail -------------------
    _write_sheet_from_rows(
        wb,
        "Detail Top 500",
        payload.get("detail_rows") or [],
        header_order=[
            "DocNo",
            "DocDate",
            "ItemType",
            "ItemCode",
            "ItemName",
            "Unit",
            "Location",
            "BatchNo",
            "TransactionType",
            "Opening",
            "Receipt",
            "Issue",
            "Closing",
            "ClosingValue",
            "UnitRate",
            "ConsQty",
            "ConsValue",
            "InwardQty",
            "InwardValue",
            "StkTrfIssueQty",
            "StkTrfIssueValue",
            "StkTrfInwardQty",
            "StkTrfInwardValue",
            "AgeDays",
            "AgeBucket",
            "FirstMovementDate",
        ],
        currency_keys={
            "ClosingValue",
            "UnitRate",
            "ConsValue",
            "InwardValue",
            "StkTrfIssueValue",
            "StkTrfInwardValue",
        },
        qty_keys={
            "Opening",
            "Receipt",
            "Issue",
            "Closing",
            "ConsQty",
            "InwardQty",
            "StkTrfIssueQty",
            "StkTrfInwardQty",
        },
    )

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    from_dt = request.GET.get("from", "")
    to_dt = request.GET.get("to", "")
    mg = request.GET.get("material_group", "ALL").upper()
    filename = f"Inventory_Ageing_{mg}_{from_dt}_to_{to_dt}.xlsx".replace(":", "-")

    resp = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


import io
import csv
import math
from datetime import datetime, date
from decimal import Decimal, InvalidOperation

import pandas as pd
from django.contrib import messages
from django.db import connections, transaction
from django.shortcuts import render, redirect

from .forms import GST2BUploadForm
def _parse_money(s):
    """Convert ' 1,33,770 ' -> Decimal('133770.00'); handles blanks/None."""
    if s is None:
        return None
    s = str(s).strip()
    if not s:
        return None
    # remove non-digit except dot and minus
    s = s.replace(',', '').replace('₹', '').replace(' ', '')
    try:
        return Decimal(s)
    except InvalidOperation:
        # sometimes Excel keeps numeric already parsed
        try:
            return Decimal(str(float(s)))
        except Exception:
            return None

def _parse_date(d):
    """
    Accepts 19/04/2024, 01-04-2024, 2024-04-19, 19-Apr-2024, Apr'24 (month bucket).
    Returns a real date. For month-like strings, returns first day of that month.
    """
    if d is None:
        return None
    s = str(d).strip()
    if not s:
        return None

    # common day-first formats
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d-%b-%Y", "%d-%b-%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass

    # Month formats like Apr'24 or Apr-24 -> month bucket
    try:
        s2 = s.replace("’", "'")
        if "'" in s2 and len(s2.split("'")[-1]) in (2,4):
            # e.g., Apr'24
            m = datetime.strptime(s2, "%b'%y")
            return date(m.year, m.month, 1)
    except Exception:
        pass

    # Excel numeric date
    try:
        # pandas/Excel serial? try pandas
        return pd.to_datetime(s, dayfirst=True, errors="coerce").date()
    except Exception:
        return None

def _month_bucket(d):
    if not d:
        return None
    return date(d.year, d.month, 1)

def _norm_invoice_no(s):
    """Match SQL UDF: strip space, '/', '-', '.', ',' and uppercase."""
    if s is None:
        return None
    t = str(s).upper()
    for ch in (' ', '/', '-', '.', ','):
        t = t.replace(ch, '')
    t = t.strip()
    return t or None


def gst2b_upload_view(request):
    """
    Upload GST-2B CSV/XLSX and insert into dbo.GST2B_ERP_Reco as source='GST2B'.
    - Parses/cleans key columns
    - Computes month_bucket from Invoice Date
    - Optionally replaces existing GST2B rows for that month
    """
    if request.method == "POST":
        form = GST2BUploadForm(request.POST, request.FILES)
        if form.is_valid():
            f = form.cleaned_data["file"]
            replace_month = form.cleaned_data["replace_month"]

            # --- 1) Load file into DataFrame (supports CSV or Excel) ---
            try:
                name = (getattr(f, "name", "") or "").lower()
                if name.endswith(".csv"):
                    df = pd.read_csv(f, dtype=str)
                else:
                    df = pd.read_excel(f, dtype=str)
            except Exception as e:
                messages.error(request, f"Could not read file: {e}")
                return redirect(request.path)

            # --- 2) Normalize column names (strip – preserve case) ---
            df.columns = [str(c).strip() for c in df.columns]

            # Expected headings (case-insensitive match)
            COL_SUPPLIER_GSTIN = next((c for c in df.columns if c.lower().startswith("gstin of supplier")), None)
            COL_TRADE_NAME     = next((c for c in df.columns if "trade" in c.lower() or "legal name" in c.lower()), None)
            COL_INVOICE_NO     = next((c for c in df.columns if c.lower().startswith("invoice number")), None)
            COL_INVOICE_DATE   = next((c for c in df.columns if c.lower().startswith("invoice date")), None)
            COL_INVOICE_VALUE  = next((c for c in df.columns if "invoice value" in c.lower()), None)
            COL_TAXABLE_VALUE  = next((c for c in df.columns if "taxable value" in c.lower()), None)
            COL_IGST           = next((c for c in df.columns if "integrated tax" in c.lower() or c.lower().startswith("igst")), None)
            COL_CGST           = next((c for c in df.columns if "central tax" in c.lower() or c.lower().startswith("cgst")), None)
            COL_SGST           = next((c for c in df.columns if "state/ut tax" in c.lower() or c.lower().startswith("sgst")), None)
            COL_POS            = next((c for c in df.columns if c.lower().startswith("place of supply")), None)
            COL_INVOICE_TYPE   = next((c for c in df.columns if c.lower().startswith("invoice type")), None)
            COL_PERIOD         = next((c for c in df.columns if "gstr-1" in c.lower() or "period" in c.lower()), None)
            COL_IRN            = next((c for c in df.columns if c.lower() == "irn"), None)
            COL_IRN_DATE       = next((c for c in df.columns if "irn date" in c.lower()), None)
            COL_REMARKS        = next((c for c in df.columns if c.lower().strip() == "remarks"), None)

            missing = [("GSTIN of supplier", COL_SUPPLIER_GSTIN),
                       ("Invoice number", COL_INVOICE_NO),
                       ("Invoice Date", COL_INVOICE_DATE),
                       ("Invoice Value(₹)", COL_INVOICE_VALUE)]
            missing = [exp for exp, col in missing if col is None]
            if missing:
                messages.error(request, f"Missing columns: {', '.join(missing)}")
                return redirect(request.path)

            # --- 3) Build rows for insert ---
            rows = []
            touched_months = set()

            for _, r in df.iterrows():
                gst_gstin   = (str(r.get(COL_SUPPLIER_GSTIN)) if COL_SUPPLIER_GSTIN else None)
                gst_name    = (str(r.get(COL_TRADE_NAME)) if COL_TRADE_NAME else None)
                inv_no      = (str(r.get(COL_INVOICE_NO)) if COL_INVOICE_NO else None)
                inv_date    = _parse_date(r.get(COL_INVOICE_DATE)) if COL_INVOICE_DATE else None
                inv_value   = _parse_money(r.get(COL_INVOICE_VALUE)) if COL_INVOICE_VALUE else None
                tax_value   = _parse_money(r.get(COL_TAXABLE_VALUE)) if COL_TAXABLE_VALUE else None
                igst        = _parse_money(r.get(COL_IGST)) if COL_IGST else None
                cgst        = _parse_money(r.get(COL_CGST)) if COL_CGST else None
                sgst        = _parse_money(r.get(COL_SGST)) if COL_SGST else None
                pos         = (str(r.get(COL_POS)) if COL_POS else None)
                inv_type    = (str(r.get(COL_INVOICE_TYPE)) if COL_INVOICE_TYPE else None)
                period      = (str(r.get(COL_PERIOD)) if COL_PERIOD else None)
                irn         = (str(r.get(COL_IRN)) if COL_IRN else None)
                irn_date    = _parse_date(r.get(COL_IRN_DATE)) if COL_IRN_DATE else None
                remarks     = (str(r.get(COL_REMARKS)) if COL_REMARKS else None)

                norm_no = _norm_invoice_no(inv_no)
                mb = _month_bucket(inv_date) or _month_bucket(_parse_date(period))

                if not inv_no or not inv_date or not mb or inv_value is None:
                    # skip incomplete rows
                    continue

                touched_months.add(mb)

                rows.append((
                    mb,                     # month_bucket
                    'GST2B',                # source
                    gst_gstin,              # gst_supplier_gstin
                    gst_name,               # gst_trade_name
                    inv_no,                 # gst_invoice_number
                    r.get(COL_INVOICE_DATE),# gst_invoice_date_raw
                    str(r.get(COL_INVOICE_VALUE)) if COL_INVOICE_VALUE else None,  # gst_invoice_value_raw
                    str(r.get(COL_TAXABLE_VALUE)) if COL_TAXABLE_VALUE else None,  # gst_taxable_value_raw
                    str(r.get(COL_IGST)) if COL_IGST else None,                    # gst_igst_raw
                    str(r.get(COL_CGST)) if COL_CGST else None,                    # gst_cgst_raw
                    str(r.get(COL_SGST)) if COL_SGST else None,                    # gst_sgst_raw
                    pos, inv_type, period,
                    irn, str(r.get(COL_IRN_DATE)) if COL_IRN_DATE else None,
                    remarks,

                    norm_no,                # gst_invoice_no_norm
                    inv_date,               # gst_invoice_date
                    inv_value,              # gst_invoice_value
                    tax_value,              # gst_taxable_value
                    igst, cgst, sgst        # igst/cgst/sgst
                ))

            if not rows:
                messages.warning(request, "No valid rows found in file.")
                return redirect(request.path)

            # --- 4) DB insert ---
            insert_sql = """
                INSERT INTO dbo.GST2B_ERP_Reco (
                    month_bucket, source,
                    gst_supplier_gstin, gst_trade_name, gst_invoice_number,
                    gst_invoice_date_raw, gst_invoice_value_raw, gst_taxable_value_raw,
                    gst_igst_raw, gst_cgst_raw, gst_sgst_raw,
                    gst_pos, gst_invoice_type, gst_gstr_period,
                    gst_irn, gst_irn_date_raw, gst_remarks,
                    gst_invoice_no_norm, gst_invoice_date, gst_invoice_value,
                    gst_taxable_value, gst_igst, gst_cgst, gst_sgst
                )
                VALUES (
                    %s, %s, %s, %s, %s,
                    %s, %s, %s,
                    %s, %s, %s,
                    %s, %s, %s,
                    %s, %s, %s,
                    %s, %s, %s,
                    %s, %s, %s, %s
                )
            """

            months_list = sorted(touched_months)

            try:
                with connections["default"].cursor() as cur:
                    with transaction.atomic(using="default"):
                        if replace_month and months_list:
                            placeholders_in = ",".join(["%s"] * len(months_list))
                            cur.execute(
                                f"""
                                DELETE FROM dbo.GST2B_ERP_Reco
                                WHERE source='GST2B'
                                  AND month_bucket IN ({placeholders_in})
                                """,
                                months_list,
                            )
                        cur.executemany(insert_sql, rows)
            except Exception as e:
                messages.error(request, f"Database error while inserting: {e}")
                return redirect(request.path)

            messages.success(request, f"Uploaded {len(rows)} GST-2B rows across {len(months_list)} month(s).")
            return redirect(request.path)
        else:
            messages.error(request, "Invalid form submission.")
            return redirect(request.path)

    # GET
    form = GST2BUploadForm()
    return render(request, "GST2B/gst2b_upload.html", {"form": form})


from datetime import date
from io import BytesIO
from django.http import HttpResponse
from django.shortcuts import render
from django.db import connections


def bank_fc_ledger_report(request):
    company_id = int(request.GET.get("company_id", 27))

    # ---- Default to current month ----
    today = date.today()
    first_of_month = today.replace(day=1).isoformat()

    req_from = (request.GET.get("from_date") or "").strip()
    req_to   = (request.GET.get("to_date") or "").strip()
    from_date = req_from if req_from else first_of_month
    to_date   = req_to   if req_to   else today.isoformat()

    currency   = (request.GET.get("currency") or "").strip()
    sacc_name  = (request.GET.get("sacc_name") or "").strip()
    company    = (request.GET.get("company") or "").strip()  # "Chemical" | "Solapur" | ""

    # UI only (no server paging)
    page_size = max(1, int(request.GET.get("page_size", 100)))

    # WHERE filters used in FINAL SELECT (aliases exist there)
    extra_where_sql = ""
    extra_params: list = []
    if currency:
        extra_where_sql += " AND cur_txn.sCode = %s"
        extra_params.append(currency)
    if sacc_name:
        extra_where_sql += " AND a.sName LIKE %s"
        extra_params.append(f"%{sacc_name}%")

    # Company → Doc No prefix (C/S)
    if company == "Chemical":
        extra_where_sql += " AND LEFT(f.sDocNo, 1) = 'C'"
    elif company == "Solapur":
        extra_where_sql += " AND LEFT(f.sDocNo, 1) = 'S'"

    # Exclusions (prefix-safe; case/space tolerant)
    ex1 = "INTRA STATE BRANCH TRANSFER - INWARD"
    ex2 = "OC Specialities Private Limited (DR)"

    # IMPORTANT: any literal % in the SQL string must be doubled as %% to avoid pyodbc formatting errors.
    sql = f"""
SET NOCOUNT ON;

DECLARE @CompanyID INT = %s, @FromDate DATE = %s, @ToDate DATE = %s;

;WITH H AS (
    SELECT d.lId, d.dtDocDate, d.sDocNo, d.lCurrId, d.sNarr AS sNarrHdr
    FROM dbo.TXNHDR d
    JOIN dbo.TXNTYP t ON t.lTypId = d.lTypId AND t.lFinTyp < 2
    WHERE d.bDel = 0
      AND d.lClosed <= 0
      AND d.lCompId = @CompanyID
      AND CONVERT(date, CONVERT(varchar(8), d.dtDocDate)) BETWEEN @FromDate AND @ToDate
      AND d.sDocNo IS NOT NULL
      AND (CHARINDEX('BP', d.sDocNo) > 0 OR CHARINDEX('BR', d.sDocNo) > 0)
),
S AS (
    SELECT
        da.lAccId,
        a.sCode AS sAccCode,
        a.sName AS sAccName,
        at.lTypId,
        h.lId,
        h.dtDocDate,
        h.sDocNo,
        da.cFlag,
        da.sNarr,
        h.sNarrHdr,
        h.lCurrId,
        SUM(da.dAmtDr)             AS dAmtDr,
        SUM(da.dAmtCr)             AS dAmtCr,
        SUM(COALESCE(da.dFCDr, 0)) AS dFCDr,
        SUM(COALESCE(da.dFCCr, 0)) AS dFCCr
    FROM H h
    JOIN dbo.TXNACC da ON da.lId = h.lId
    JOIN dbo.ACCMST a  ON a.lId  = da.lAccId
    JOIN dbo.ACCTYP at ON at.lTypId = a.lTypId
    WHERE 1=1
      AND UPPER(LTRIM(RTRIM(a.sName))) NOT LIKE UPPER(%s) + N'%%'
      AND UPPER(LTRIM(RTRIM(a.sName))) NOT LIKE UPPER(%s) + N'%%'
    GROUP BY
        da.lAccId, a.sCode, a.sName, at.lTypId,
        h.lId, h.dtDocDate, h.sDocNo, da.cFlag, da.sNarr, h.sNarrHdr, h.lCurrId
),
F AS (
    SELECT
        s.*,
        cur_txn.sCode AS CurrencyCode,
        cur_txn.sName AS CurrencyName
    FROM S s
    JOIN dbo.CURMST cur_txn ON cur_txn.lId = s.lCurrId
),

-- ***** STRICT header map: ONLY true bank headers (codes BK00/BNKO) *****
HeaderMap AS (
    SELECT
        a.sCode AS sAccCode,
        CASE
            WHEN LEFT(UPPER(a.sName), 4) = 'CITI' THEN 'CITI BANK'
            WHEN LEFT(UPPER(a.sName), 4) = 'HDFC' THEN 'HDFC BANK'
            WHEN LEFT(UPPER(a.sName), 3) = 'SCB'  THEN 'SCB BANK'
            ELSE a.sName
        END AS BankName
    FROM dbo.ACCMST a
    WHERE LEFT(a.sCode, 4) IN ('BK00','BNKO')   -- ONLY header accounts
),

-- Bank detected per document by presence of a header account line
DocBank AS (
    SELECT f.sDocNo, MIN(hm.BankName) AS BankName
    FROM F f
    JOIN HeaderMap hm ON hm.sAccCode = f.sAccCode
    GROUP BY f.sDocNo
)
SELECT
    CONVERT(varchar(10), CONVERT(date, CONVERT(varchar(8), f.dtDocDate)), 105) AS [Date],
    f.sDocNo                                   AS [Doc No],
    f.sAccCode                                 AS [Acc Code],
    f.sAccName                                 AS [Account Name],
    f.CurrencyCode                             AS [Currency],
    f.CurrencyName                             AS [Currency Name],
    CAST(f.dAmtDr AS decimal(18,2))            AS [Dr],      -- Base amount
    CAST(f.dAmtCr AS decimal(18,2))            AS [Cr],      -- Base amount
    CAST(f.dAmtDr - f.dAmtCr AS decimal(18,2)) AS [Net],
    NULLIF(f.dFCDr, 0)                         AS [FC Dr],   -- FC amount (any currency)
    NULLIF(f.dFCCr, 0)                         AS [FC Cr],   -- FC amount (any currency)
    LTRIM(RTRIM(REPLACE(REPLACE(REPLACE(REPLACE(f.sNarr,'G&L',''),'G & L',''),'G&amp;L',''),'G &amp; L',''))) AS [Narration],
    LTRIM(RTRIM(REPLACE(REPLACE(REPLACE(REPLACE(f.sNarrHdr,'G&L',''),'G & L',''),'G&amp;L',''),'G &amp; L',''))) AS [Header Narration],
    CASE WHEN hm.sAccCode IS NOT NULL THEN 'Header' ELSE 'Detail' END AS [View],
    COALESCE(hm.BankName, db.BankName, '')     AS [Bank Name],
    CASE
        WHEN CHARINDEX('BP', f.sDocNo) > 0 THEN 'Bank Payment'
        WHEN CHARINDEX('BR', f.sDocNo) > 0 THEN 'Bank Receipt'
        ELSE '0'
    END                                         AS [Payment Type]
FROM F f
LEFT JOIN HeaderMap hm ON hm.sAccCode = f.sAccCode
LEFT JOIN DocBank   db ON db.sDocNo   = f.sDocNo
JOIN dbo.ACCMST a    ON a.lId = f.lAccId
JOIN dbo.CURMST cur_txn ON cur_txn.lId = f.lCurrId
WHERE 1=1
  {extra_where_sql}
  AND UPPER(LTRIM(RTRIM(a.sName))) NOT LIKE UPPER(%s) + N'%%'
  AND UPPER(LTRIM(RTRIM(a.sName))) NOT LIKE UPPER(%s) + N'%%'
ORDER BY CONVERT(date, CONVERT(varchar(8), f.dtDocDate)) ASC, f.sDocNo ASC, f.lId ASC
OPTION (RECOMPILE);
"""

    # Order must match placeholders in SQL
    params = [
        company_id, from_date, to_date,   # DECLARE params
        ex1, ex2,                         # exclusions during S CTE
    ] + extra_params + [
        ex1, ex2                          # exclusions in FINAL WHERE
    ]

    with connections["readonly_db"].cursor() as cur:
        cur.execute(sql, params)
        cols = [c[0] for c in cur.description]
        rows = cur.fetchall()

    table_data = [dict(zip(cols, r)) for r in rows]

    # ========= Excel download of FULL RAW DATA (no summaries) =========
    if (request.GET.get("download") or "").lower() in ("1", "true", "xlsx", "excel"):
        buff = BytesIO()
        try:
            import xlsxwriter
        except Exception:
            import csv
            resp = HttpResponse(content_type="text/csv")
            resp["Content-Disposition"] = 'attachment; filename="bank_fc_ledger_raw.csv"'
            writer = csv.writer(resp)
            writer.writerow(cols)
            for r in table_data:
                writer.writerow([r.get(c) for c in cols])
            return resp

        wb = xlsxwriter.Workbook(buff, {"in_memory": True})
        ws = wb.add_worksheet("Raw")

        hdr = wb.add_format({"bold": True, "bg_color": "#EEF2FF", "border": 1})
        numfmt = wb.add_format({"num_format": "#,##0.00"})
        txtfmt = wb.add_format({})

        # header row
        for j, c in enumerate(cols):
            ws.write(0, j, c, hdr)
            ws.set_column(j, j, max(10, min(40, len(c) + 2)))

        # data rows
        numeric_cols = {"Dr", "Cr", "Net", "FC Dr", "FC Cr"}
        for i, r in enumerate(table_data, start=1):
            for j, c in enumerate(cols):
                v = r.get(c)
                if isinstance(v, (int, float)) and c in numeric_cols:
                    ws.write_number(i, j, float(v), numfmt)
                else:
                    ws.write(i, j, "" if v is None else v, txtfmt)

        wb.close()
        buff.seek(0)
        resp = HttpResponse(
            buff.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = 'attachment; filename="bank_fc_ledger_raw.xlsx"'
        return resp

    # -------- Render web page --------
    return render(
        request,
        "PC Reports/bank_fc_ledger.html",
        {
            "columns": cols,
            "table_data": table_data,   # full dataset
            "page": 1,
            "page_size": page_size,     # UI only
            "num_pages": 1,
            "prev_url": None,
            "next_url": None,
            "sizes": [25, 50, 100, 200, 500],
            "filters": {
                "company_id": company_id,
                "from_date": from_date,
                "to_date": to_date,
                "currency": currency,
                "sacc_name": sacc_name,
                "company": company,
            },
            "currency_choices": [],
        },
    )

from datetime import date
import json

from django.db import connections
from django.shortcuts import render


def rgp_issue_grn_report(request):
    """
    RGP Issue vs RGP GRN report with:
      - RGP_ISSUE Quantity
      - RGP GRN Quantity
      - Balance Qty (0 on all rows except last GRN row per line)
      - Actual BALance (running balance after each GRN)

    Data comes from readonly_db and full detail is sent to frontend
    (table_data + table_data_json) so you can build a dashboard.
    """

    # Fixed company (no filter in UI)
    company_id = 27

    from_date = (request.GET.get("from_date") or "2022-11-01").strip()
    to_date = (request.GET.get("to_date") or date.today().isoformat()).strip()
    supp_id = int(request.GET.get("supp_id", 0))

    # New filter: Transaction Type of Issue (by name, e.g. "RGP Issue Voucher")
    issue_txn_name = (request.GET.get("issue_txn_name") or "").strip()

    # Only run when user clicks "Run"
    run_flag = request.GET.get("run") == "1"
    if not run_flag:
        return render(
            request,
            "Store/rgp_issue_grn.html",
            {
                "columns": [],
                "table_data": [],
                "table_data_json": "[]",
                "show_results": False,
                "error_msg": "",
                "filters": {
                    "from_date": from_date,
                    "to_date": to_date,
                    "supp_id": supp_id,
                    "issue_txn_name": issue_txn_name,
                },
                "total_rows": 0,
                "issue_txn_choices": [],  # empty for initial load
            },
        )

    def esc(s: str) -> str:
        return s.replace("'", "''") if s else s

    # ---------- SQL with fixed INT dates + parsed Required Date ----------
    sql = f"""
SET NOCOUNT ON;

DECLARE
    @CompanyID    INT           = {company_id},
    @FromDate     DATE          = '{esc(from_date)}',
    @ToDate       DATE          = '{esc(to_date)}',
    @SuppId       INT           = {supp_id},
    @IssueTxnName NVARCHAR(200) = '{esc(issue_txn_name)}';

;WITH Base AS
(
    SELECT
        -- Header fields
        t.sName AS [Transcation Type of issue No],

        RGP_ISSUEHDR.sDocNo AS [RGP_ISSUE NO],

        -- dtDocDate stored as INT (YYYYMMDD) → DATE → DD/MM/YYYY text
        CONVERT(
            VARCHAR(10),
            TRY_CONVERT(
                DATE,
                CONVERT(VARCHAR(8), RGP_ISSUEHDR.dtDocDate),
                112
            ),
            103
        ) AS [RGP_ISSUE Date],

        t2.sName AS [Transcation Type of Grn],
        GRNHDR.sDocNo AS [GRN NO],

        CONVERT(
            VARCHAR(10),
            TRY_CONVERT(
                DATE,
                CONVERT(VARCHAR(8), GRNHDR.dtDocDate),
                112
            ),
            103
        ) AS [GRN Date],

        cfRefNo.sValue      AS [Reference No.],
        cfRefDate.sValue    AS [Reference Date],
        cfTrnRefNo.sValue   AS [Transaction Reference No],
        cfTrnRefDate.sValue AS [Transaction Reference Date],

        TRANS.sName AS [Transporter Name],
        SUPP.sCode  AS [Supplier Code - Billing],
        SUPP.sName  AS [Supplier Name - Billing],

        -- Detail fields
        RGP_ISSUEDET.lLine      AS [RGP Requisition Line No],
        ityp.sName              AS [RGP Requisition Item Type],
        itm.sCode               AS [RGP Requisition Item Code],
        itm.sName               AS [RGP Requisition Item Name],

        RGP_ISSUEDET.sNarr      AS [RGP Requisition Item Narration],
        RGP_ISSUEDET.svalue5    AS [AR No.],
        RGP_ISSUEDET.svalue1    AS [Batch No],

        uom.sCode               AS [RGP Requisition UOM],

        RGP_ISSUEDET.dQty2      AS IssueQty,
        GRNDET.dQty2            AS GrnQty,

        -- Required date RAW + parsed DATE
        cfReqDate.sValue        AS [Required by Date],
        TRY_CONVERT(DATE, cfReqDate.sValue, 106) AS [Required Date],

        RGP_ISSUEDET.sNarr      AS [RGP_ISSUE Item Narration],

        -- Total GRN qty per RGP line
        SUM(ISNULL(GRNDET.dQty2, 0)) OVER (
            PARTITION BY RGP_ISSUEDET.lId, RGP_ISSUEDET.lLine
        ) AS TotalGRNQty,

        -- Running GRN qty (for Actual BALance)
        SUM(ISNULL(GRNDET.dQty2, 0)) OVER (
            PARTITION BY RGP_ISSUEDET.lId, RGP_ISSUEDET.lLine
            ORDER BY
                TRY_CONVERT(
                    DATE,
                    CONVERT(VARCHAR(8), GRNHDR.dtDocDate),
                    112
                ),
                GRNHDR.sDocNo,
                GRNDET.lLine
            ROWS UNBOUNDED PRECEDING
        ) AS RunGrnQty,

        -- Identify last GRN row per RGP line
        ROW_NUMBER() OVER (
            PARTITION BY RGP_ISSUEDET.lId, RGP_ISSUEDET.lLine
            ORDER BY
                TRY_CONVERT(
                    DATE,
                    CONVERT(VARCHAR(8), GRNHDR.dtDocDate),
                    112
                ) DESC,
                GRNHDR.sDocNo DESC,
                GRNDET.lLine DESC
        ) AS rn_desc

    FROM TXNHDR RGP_ISSUEHDR
    INNER JOIN TXNDET AS RGP_ISSUEDET
        ON RGP_ISSUEHDR.lId = RGP_ISSUEDET.lId

    LEFT JOIN BUSMST AS SUPP
        ON RGP_ISSUEHDR.lAccId1 = SUPP.lId

    LEFT JOIN BUSMST AS TRANS
        ON RGP_ISSUEHDR.lAccId5 = TRANS.lId

    -- GRN detail rows
    LEFT JOIN TXNDET AS GRNDET
        ON RGP_ISSUEDET.lId   = GRNDET.lLnkDocId
       AND GRNDET.lLnkLine    = RGP_ISSUEDET.lLine
       AND GRNDET.lTypId      IN (540, 850)

    LEFT JOIN TXNHDR GRNHDR
        ON GRNHDR.lId    = GRNDET.lId
       AND GRNHDR.lTypId IN (540, 850)

    LEFT JOIN TXNTYP t
        ON t.lTypId = RGP_ISSUEHDR.lTypId

    LEFT JOIN TXNTYP t2
        ON t2.lTypId = GRNHDR.lTypId

    -- Master joins instead of scalar subqueries
    LEFT JOIN ITMTYP ityp
        ON ityp.lTypId = RGP_ISSUEDET.lItmtyp

    LEFT JOIN ITMMST itm
        ON itm.lId = RGP_ISSUEDET.lItmId

    LEFT JOIN UNTMST uom
        ON uom.lId = RGP_ISSUEDET.lUntId

    LEFT JOIN TXNCF cfReqDate
        ON cfReqDate.lId      = RGP_ISSUEDET.lId
       AND cfReqDate.lLine    = RGP_ISSUEDET.lLine
       AND cfReqDate.lFieldNo = 1

    LEFT JOIN TXNCF cfRefNo
        ON cfRefNo.lId    = GRNHDR.lId
       AND cfRefNo.lLine  = 0
       AND cfRefNo.sName  = 'Reference No.'

    LEFT JOIN TXNCF cfRefDate
        ON cfRefDate.lId   = GRNHDR.lId
       AND cfRefDate.lLine = 0
       AND cfRefDate.sName = 'Reference Date'

    LEFT JOIN TXNCF cfTrnRefNo
        ON cfTrnRefNo.lId   = GRNHDR.lId
       AND cfTrnRefNo.lLine = 0
       AND cfTrnRefNo.sName = 'Transaction Reference No'

    LEFT JOIN TXNCF cfTrnRefDate
        ON cfTrnRefDate.lId   = GRNHDR.lId
       AND cfTrnRefDate.lLine = 0
       AND cfTrnRefDate.sName = 'Transaction Reference Date'

    WHERE
        RGP_ISSUEHDR.lTypId IN (533, 848)
        AND RGP_ISSUEHDR.lCompId = @CompanyID

        -- INT dtDocDate (YYYYMMDD) → DATE and filter between @FromDate / @ToDate
        AND TRY_CONVERT(
                DATE,
                CONVERT(VARCHAR(8), RGP_ISSUEHDR.dtDocDate),
                112
            ) BETWEEN @FromDate AND @ToDate

        AND (RGP_ISSUEHDR.lAccId1 = @SuppId OR @SuppId = 0)
        AND (@IssueTxnName = '' OR t.sName LIKE '%' + @IssueTxnName + '%')
)

SELECT
    [Transcation Type of issue No],
    [RGP_ISSUE NO],
    [RGP_ISSUE Date],
    [Transcation Type of Grn],
    [GRN NO],
    [GRN Date],
    [Reference No.],
    [Reference Date],
    [Transaction Reference No],
    [Transaction Reference Date],
    [Transporter Name],
    [Supplier Code - Billing],
    [Supplier Name - Billing],
    [RGP Requisition Line No],
    [RGP Requisition Item Type],
    [RGP Requisition Item Code],
    [RGP Requisition Item Name],
    [RGP Requisition Item Narration],
    [AR No.],
    [Batch No],
    [RGP Requisition UOM],

    CONVERT(DECIMAL(18,3), IssueQty) AS [RGP_ISSUE Quantity],
    CONVERT(DECIMAL(18,3), GrnQty)   AS [RGP GRN Quantity],

    -- Balance only in last GRN row per RGP line, 0 on previous rows
    CONVERT(DECIMAL(18,3),
        CASE
            WHEN rn_desc = 1
                 THEN ISNULL(IssueQty, 0) - ISNULL(TotalGRNQty, 0)
            ELSE 0
        END
    ) AS [Balance Qty],

    -- Running balance after each GRN
    CONVERT(DECIMAL(18,3),
        ISNULL(IssueQty, 0) - ISNULL(RunGrnQty, 0)
    ) AS [Actual BALance],

    [Required by Date],
    [Required Date],
    [RGP_ISSUE Item Narration]

FROM Base
ORDER BY
    [RGP_ISSUE NO],
    [RGP Requisition Line No],
    [GRN Date],
    [GRN NO];
"""

    # ---------- Execute on readonly_db ----------
    with connections["readonly_db"].cursor() as cur:
        # Protect against % formatting issues
        safe_sql = sql.replace("%", "%%")
        cur.execute(safe_sql)
        rows = cur.fetchall()
        cols = [c[0] for c in cur.description] if cur.description else []

    table_data = [dict(zip(cols, r)) for r in rows]
    table_data_json = json.dumps(table_data, default=str)

    # Unique RGP_ISSUE NO count for summary card
    unique_rgp_issues = {
        row.get("RGP_ISSUE NO") for row in table_data if row.get("RGP_ISSUE NO")
    }
    total_rows = len(unique_rgp_issues)

    # Dropdown options for "Transcation Type of issue No"
    issue_txn_choices = sorted({
        row.get("Transcation Type of issue No")
        for row in table_data
        if row.get("Transcation Type of issue No")
    })

    return render(
        request,
        "Store/rgp_issue_grn.html",
        {
            "columns": cols,
            "table_data": table_data,
            "table_data_json": table_data_json,
            "show_results": True,
            "error_msg": "",
            "total_rows": total_rows,             # unique RGP_ISSUE NO count
            "issue_txn_choices": issue_txn_choices,
            "filters": {
                "from_date": from_date,
                "to_date": to_date,
                "supp_id": supp_id,
                "issue_txn_name": issue_txn_name,
            },
        },
    )



from io import BytesIO
from datetime import date, timedelta
from urllib.parse import urlencode

from django.db import connections
from django.http import HttpResponse
from django.shortcuts import render
from django.contrib.auth.decorators import login_required


@login_required
def retest_dashboard(request):
    """
    Retest Dashboard (snapshot-based)
    - Summary table (Expired / Next 7 / Next 15 / Next 30)
    - Click a category -> shows detail table below (auto scroll via #detailPanel)
    - Location multi-select dropdown (checkboxes)
    - Non-overlapping buckets:
        Expired: < today
        Next 7:  today .. today+7
        Next 15: today+8 .. today+15
        Next 30: today+16 .. today+30
    """

    user_groups  = list(request.user.groups.values_list('name', flat=True))
    is_superuser = request.user.is_superuser

    # ---------- Filters ----------
    company_id = int(request.GET.get("company_id", 27) or 27)
    item_name  = (request.GET.get("item_name") or "").strip()
    batch_no   = (request.GET.get("batch_no") or "").strip()
    limit      = max(10, int(request.GET.get("limit", 100) or 100))

    # location multi-select: ?location=A&location=B
    selected_locations = [x.strip() for x in request.GET.getlist("location") if x.strip()]

    selected_bucket = (request.GET.get("bucket") or "expired").lower()
    if selected_bucket not in ("expired", "next7", "next15", "next30"):
        selected_bucket = "expired"

    want_export = (request.GET.get("export") or "").lower() in ("xlsx", "excel")
    bucket_req  = (request.GET.get("bucket") or "all").lower()

    # ---------- Date ranges ----------
    today = date.today()
    d7  = today + timedelta(days=7)
    d15 = today + timedelta(days=15)
    d30 = today + timedelta(days=30)

    s15 = today + timedelta(days=8)   # start for 15 bucket (non-overlap)
    s30 = today + timedelta(days=16)  # start for 30 bucket (non-overlap)

    # ---------- SQL helpers ----------
    # NOTE: change `location` here if your snapshot column name is different
    base_cte = """
;WITH G AS
(
    SELECT
        CAST(location AS NVARCHAR(200)) AS location,
        item_name,
        UPPER(LTRIM(RTRIM(batch_no))) AS batch_no,
        SUM(closing) AS current_stock,
        MAX(TRY_CONVERT(DATE, retest_date, 103)) AS expiry_date,
        MIN(TRY_CONVERT(DATE, batch_start_date, 103)) AS batch_start_date,
        MIN(doc_no) AS first_doc_no
    FROM dbo.inventory_ageing_snapshot
    WHERE company_id = %s
      AND item_type IN ('Key Raw Material', 'Raw Material')
    GROUP BY
        CAST(location AS NVARCHAR(200)),
        item_name,
        UPPER(LTRIM(RTRIM(batch_no)))
    HAVING
        SUM(closing) > 0
)
"""

    def _outer_filters_sql_and_params():
        """
        Builds WHERE clause applied on G g (not inside snapshot table)
        """
        where_sql = "WHERE 1=1"
        params = []

        if item_name:
            where_sql += " AND g.item_name LIKE %s"
            params.append(f"%{item_name}%")

        if batch_no:
            where_sql += " AND g.batch_no LIKE %s"
            params.append(f"%{batch_no.strip().upper()}%")

        if selected_locations:
            # IN (%s,%s,...) with dynamic placeholders
            placeholders = ",".join(["%s"] * len(selected_locations))
            where_sql += f" AND g.location IN ({placeholders})"
            params.extend(selected_locations)

        return where_sql, params

    outer_filter_sql, outer_params = _outer_filters_sql_and_params()

    # ---------- Load available locations for dropdown ----------
    sql_locations = f"""
SET NOCOUNT ON;
{base_cte}
SELECT DISTINCT g.location
FROM G g
{outer_filter_sql.replace("g.location IN", "g.location IN")} -- keep same outer filters except IN uses selected ones
ORDER BY g.location;
"""
    # For locations list we should NOT filter by selected_locations (else dropdown shrinks),
    # so rebuild filters without location filter:
    def _filters_without_location():
        where_sql = "WHERE 1=1"
        params = []
        if item_name:
            where_sql += " AND g.item_name LIKE %s"
            params.append(f"%{item_name}%")
        if batch_no:
            where_sql += " AND g.batch_no LIKE %s"
            params.append(f"%{batch_no.strip().upper()}%")
        return where_sql, params

    outer_no_loc_sql, outer_no_loc_params = _filters_without_location()

    sql_locations = f"""
SET NOCOUNT ON;
{base_cte}
SELECT DISTINCT g.location
FROM G g
{outer_no_loc_sql}
ORDER BY g.location;
"""
    with connections["default"].cursor() as cur:
        cur.execute(sql_locations, [company_id] + outer_no_loc_params)
        locations = [row[0] for row in cur.fetchall()]

    # ---------- KPI Summary (non-overlapping) ----------
    sql_kpi = f"""
SET NOCOUNT ON;
{base_cte}
SELECT
    SUM(CASE WHEN g.expiry_date IS NOT NULL AND g.expiry_date < %s THEN 1 ELSE 0 END) AS expired_count,
    SUM(CASE WHEN g.expiry_date IS NOT NULL AND g.expiry_date < %s THEN g.current_stock ELSE 0 END) AS expired_stock,

    SUM(CASE WHEN g.expiry_date IS NOT NULL AND g.expiry_date >= %s AND g.expiry_date <= %s THEN 1 ELSE 0 END) AS next7_count,
    SUM(CASE WHEN g.expiry_date IS NOT NULL AND g.expiry_date >= %s AND g.expiry_date <= %s THEN g.current_stock ELSE 0 END) AS next7_stock,

    SUM(CASE WHEN g.expiry_date IS NOT NULL AND g.expiry_date >= %s AND g.expiry_date <= %s THEN 1 ELSE 0 END) AS next15_count,
    SUM(CASE WHEN g.expiry_date IS NOT NULL AND g.expiry_date >= %s AND g.expiry_date <= %s THEN g.current_stock ELSE 0 END) AS next15_stock,

    SUM(CASE WHEN g.expiry_date IS NOT NULL AND g.expiry_date >= %s AND g.expiry_date <= %s THEN 1 ELSE 0 END) AS next30_count,
    SUM(CASE WHEN g.expiry_date IS NOT NULL AND g.expiry_date >= %s AND g.expiry_date <= %s THEN 1 ELSE 0 END) AS next30_stock
FROM G g
{outer_filter_sql};
"""
    kpi_params = (
        [company_id]
        + [today, today]              # expired
        + [today, d7, today, d7]      # next 7
        + [s15, d15, s15, d15]        # next 15
        + [s30, d30, s30, d30]        # next 30
        + outer_params
    )

    with connections["default"].cursor() as cur:
        cur.execute(sql_kpi, kpi_params)
        rr = cur.fetchone() or (0,) * 8

    kpi = {
        "expired_count": int(rr[0] or 0),
        "expired_stock": float(rr[1] or 0),
        "next7_count": int(rr[2] or 0),
        "next7_stock": float(rr[3] or 0),
        "next15_count": int(rr[4] or 0),
        "next15_stock": float(rr[5] or 0),
        "next30_count": int(rr[6] or 0),
        "next30_stock": float(rr[7] or 0),
    }

    summary_rows = [
        {"key": "expired", "label": "Expired (Previous)", "count": kpi["expired_count"], "stock": kpi["expired_stock"]},
        {"key": "next7",   "label": "Expiring in Next 7 Days", "count": kpi["next7_count"], "stock": kpi["next7_stock"]},
        {"key": "next15",  "label": "Expiring in Next 15 Days", "count": kpi["next15_count"], "stock": kpi["next15_stock"]},
        {"key": "next30",  "label": "Expiring in Next 30 Days", "count": kpi["next30_count"], "stock": kpi["next30_stock"]},
    ]

    # ---------- Detail query ----------
    def run_detail(bucket_key: str, top_n: int):
        if bucket_key == "expired":
            where_sql = "g.expiry_date < %s"
            params = [today]
            order_by = "g.expiry_date DESC, g.item_name, g.batch_no"
        elif bucket_key == "next7":
            where_sql = "g.expiry_date >= %s AND g.expiry_date <= %s"
            params = [today, d7]
            order_by = "g.expiry_date ASC, g.item_name, g.batch_no"
        elif bucket_key == "next15":
            where_sql = "g.expiry_date >= %s AND g.expiry_date <= %s"
            params = [s15, d15]
            order_by = "g.expiry_date ASC, g.item_name, g.batch_no"
        elif bucket_key == "next30":
            where_sql = "g.expiry_date >= %s AND g.expiry_date <= %s"
            params = [s30, d30]
            order_by = "g.expiry_date ASC, g.item_name, g.batch_no"
        else:
            bucket_key = "expired"
            where_sql = "g.expiry_date < %s"
            params = [today]
            order_by = "g.expiry_date DESC, g.item_name, g.batch_no"

        sql = f"""
SET NOCOUNT ON;
{base_cte}
SELECT TOP ({top_n})
    g.location,
    g.item_name,
    g.batch_no,
    g.current_stock,
    g.expiry_date,
    g.batch_start_date,
    g.first_doc_no
FROM G g
{outer_filter_sql}
  AND g.expiry_date IS NOT NULL
  AND ({where_sql})
ORDER BY {order_by};
"""
        with connections["default"].cursor() as cur:
            cur.execute(sql, [company_id] + outer_params + params)
            cols = [c[0] for c in cur.description]
            rows = cur.fetchall()
        return cols, rows

    # ---------- Export ----------
    if want_export:
        import pandas as pd

        valid = {"expired", "next7", "next15", "next30", "all"}
        if bucket_req not in valid:
            bucket_req = "all"

        buckets = ["expired", "next7", "next15", "next30"] if bucket_req == "all" else [bucket_req]

        out = BytesIO()
        with pd.ExcelWriter(out, engine="openpyxl") as xw:
            for b in buckets:
                cols, rows = run_detail(b, top_n=2000000000)
                df = pd.DataFrame(rows, columns=cols)

                sheet = {
                    "expired": "Expired",
                    "next7": "Next 7 Days",
                    "next15": "Next 15 Days",
                    "next30": "Next 30 Days",
                }[b]

                df.to_excel(xw, index=False, sheet_name=sheet)
                ws = xw.book[sheet]
                ws.freeze_panes = ws["A2"]
                for col_idx, col_name in enumerate(df.columns, start=1):
                    try:
                        sample_vals = (str(v) for v in df[col_name].head(200).values)
                        maxlen = max([len(str(col_name)), *[len(s) for s in sample_vals]]) + 2
                    except ValueError:
                        maxlen = len(str(col_name)) + 2
                    ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(maxlen, 45)

        resp = HttpResponse(
            out.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = 'attachment; filename="Retest_Dashboard.xlsx"'
        return resp

    # ---------- Detail rows for selected bucket ----------
    cols_d, rows_d = run_detail(selected_bucket, top_n=limit)
    detail_rows = [dict(zip(cols_d, r)) for r in rows_d]

    # ---------- Build querystring prefix (for clean clickable category links) ----------
    qs = request.GET.copy()
    qs.pop("bucket", None)
    qs.pop("export", None)
    base_qs = qs.urlencode()
    qs_prefix = (base_qs + "&") if base_qs else ""

    return render(
        request,
        "Store/retest_dashboard.html",
        {
            "today": today,
            "d7": d7,
            "d15": d15,
            "d30": d30,

            "summary_rows": summary_rows,
            "selected_bucket": selected_bucket,
            "detail_rows": detail_rows,

            "limit": limit,
            "kpi": kpi,

            "filters": {
                "company_id": company_id,
                "item_name": item_name,
                "batch_no": batch_no,
            },

            "locations": locations,
            "selected_locations": selected_locations,

            "qs_prefix": qs_prefix,

            "user_groups": user_groups,
            "is_superuser": is_superuser,
            "active_menu": "reports",
        },
    )
