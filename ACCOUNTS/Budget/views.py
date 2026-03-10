# ------------------------------
# stdlib
# ------------------------------
import io
import re
from collections import defaultdict
from datetime import date
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from functools import reduce
from operator import or_
from urllib import request
from venv import logger
from django.utils.safestring import mark_safe
from django.db import transaction, models as dj_models
from django.core.exceptions import FieldError

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ------------------------------
# django
# ------------------------------
from django.apps import apps
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.contenttypes.models import ContentType
from django.core.cache import cache
from django.core.exceptions import FieldDoesNotExist
from django.core.exceptions import ImproperlyConfigured
from django.db import IntegrityError, transaction
from django.db.models import Q, Count, Max, Min, Sum
from django.http import JsonResponse, HttpResponse, Http404
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse,NoReverseMatch
from django.utils import timezone
from django.utils.http import url_has_allowed_host_and_scheme
from django.views.decorators.csrf import csrf_protect
from django.views.decorators.http import require_POST, require_http_methods
from .forms import RMCBudgetForm, RMCBudgetLineForm
from django.forms import modelformset_factory

# ------------------------------
# local: forms
# ------------------------------
from .forms import *

# ------------------------------
# local: models
# ------------------------------
from .models import *

# ------------------------------
# local: services (✅ avoid name clashes)
# ------------------------------
from .services.maker_checker import (
    mc_get,
    mc_submit,
    mc_check as mc_check_service,   # ✅ IMPORTANT: do not import mc_check directly
    mc_approve,
    mc_reject,
)

from .services.maker_checker import (
    mc_get,
    mc_approve as mc_approve_service,
    mc_reject as mc_reject_service,
)


import logging
import uuid
from django.db import transaction, IntegrityError, DatabaseError
# ✅ IMPORTANT: use ONE logger consistently
prod_logger = logging.getLogger("accounts_budget.production")

# =============================================================================
# Maker-Checker helpers (keep flow, only lock edit when submitted/checked/approved)
# =============================================================================

def _mc_app_label() -> str:
    return MakerCheckerState._meta.app_label

def _perm(codename: str) -> str:
    return f"{_mc_app_label()}.{codename}"

def _is_checker(user) -> bool:
    return bool(user and user.is_authenticated and (user.is_superuser or user.has_perm(_perm("can_check_budgets"))))

def _is_approver(user) -> bool:
    return bool(user and user.is_authenticated and (user.is_superuser or user.has_perm(_perm("can_approve_budgets"))))


def _has_checked_status() -> bool:
    return "CHECKED" in {c[0] for c in MCStatus.choices}

def _is_locked(mc) -> bool:
    # maker must not edit after submit / check / approve
    if not mc:
        return False
    locked = {MCStatus.SUBMITTED, MCStatus.APPROVED}
    if _has_checked_status():
        locked.add("CHECKED")
    return mc.status in locked

# ----------------------------------------------------------------------------
# RMC Draft
# ----------------------------------------------------------------------------
def _mc_mark_draft(obj, scope: str, user):
    """
    Set/keep MC status as DRAFT without changing existing save/submit flow.
    Safe across schema variations (submitted_at vs submitted_on, etc.).
    """
    if not obj or not getattr(obj, "pk", None):
        return None

    st = mc_get(obj, scope)
    now = timezone.now()

    if st is None:
        # Create if missing
        ct = ContentType.objects.get_for_model(obj.__class__)
        st = MakerCheckerState.objects.create(
            content_type=ct,
            object_id=obj.pk,
            scope=scope,
            status=MCStatus.DRAFT,
            created_by=user if hasattr(MakerCheckerState, "created_by") else None,
        )

    # If already approved, do not downgrade silently
    if str(getattr(st, "status", "")).upper() == "APPROVED":
        return st

    st.status = MCStatus.DRAFT

    # Clear submission/check/approve metadata if present
    for fld in ("submitted_at", "submitted_on", "checked_at", "checked_on", "approved_at", "approved_on"):
        if hasattr(st, fld):
            setattr(st, fld, None)

    for fld in ("submitted_by", "checked_by", "approved_by"):
        if hasattr(st, fld):
            setattr(st, fld, None)

    if hasattr(st, "updated_at"):
        st.updated_at = now

    st.save()
    return st

# =============================================================================
# Model resolver (robust)
# =============================================================================

def _get_model(model_token: str):
    """
    Accepts:
      - 'app_label.ModelName'   (recommended)
      - 'namespace.ModelName'   (legacy; may not match app_label)
      - 'ModelName'             (fallback if unique)
      - 'some.long.path.ModelName' (fallback: last segment is ModelName)
    """
    if not model_token:
        raise LookupError("Empty model token")

    token = str(model_token).strip()
    parts = [p for p in token.split(".") if p]
    model_name = parts[-1]

    # 1) try explicit app_label.ModelName
    if len(parts) >= 2:
        for candidate_app_label in (parts[0], parts[-2]):
            try:
                return apps.get_model(candidate_app_label, model_name)
            except Exception:
                pass

    # 2) fallback: search globally by class name (must be unique)
    matches = []
    for m in apps.get_models():
        if m.__name__.lower() == model_name.lower():
            matches.append(m)

    if len(matches) == 1:
        return matches[0]

    if len(matches) > 1:
        raise ImproperlyConfigured(
            f"Ambiguous model '{model_name}' for token '{model_token}'. "
            f"Matches: {[m._meta.label for m in matches]}"
        )

    raise LookupError(f"Could not resolve model from token '{model_token}'.")

def _has_field(model_cls, name: str) -> bool:
    try:
        model_cls._meta.get_field(name)
        return True
    except Exception:
        return False


def _max_field_if_exists(model_cls, preferred: str, fallback: str = "id") -> str:
    return preferred if _has_field(model_cls, preferred) else fallback


def resolve_model_instance(model_token: str, pk: int):
    Model = _get_model(model_token)
    try:
        return Model.objects.get(pk=pk)
    except Model.DoesNotExist:
        return None


# def _safe_next(request, fallback_url_name: str):
#     nxt = request.POST.get("next") or request.GET.get("next") or reverse(fallback_url_name)
#     if url_has_allowed_host_and_scheme(
#         url=nxt,
#         allowed_hosts={request.get_host()},
#         require_https=request.is_secure(),
#     ):
#         return nxt
#     return reverse(fallback_url_name)


# =============================================================================
# Checker action (SUBMITTED -> CHECKED)
# =============================================================================
def _safe_next(request, fallback):
    nxt = request.POST.get("next") or request.GET.get("next")
    if nxt and url_has_allowed_host_and_scheme(
        url=nxt,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ):
        return nxt

    # fallback can be: "app:urlname" OR "/some/path/"
    if isinstance(fallback, str):
        fb = fallback.strip()
        if fb and (":" in fb) and not fb.startswith(("/", "http://", "https://")):
            try:
                return reverse(fb)
            except NoReverseMatch:
                pass

    return fallback


@require_http_methods(["GET", "POST"])
@login_required
def mc_check_view(request, model: str, pk: int, scope: str):
    """
    Checker step:
      SUBMITTED -> CHECKED
    """
    # Permission name must match your MakerCheckerState permissions
    if not request.user.has_perm("Budget.can_check_budgets"):
        messages.error(request, "You are not authorized to check budgets.")
        return redirect(_safe_next(request, reverse("accounts_budget:budget_approvals_inbox")))

    Model = _get_model(model)
    obj = get_object_or_404(Model, pk=pk)
    mc = mc_get(obj, scope)

    if not mc:
        messages.error(request, "Maker-Checker state not found.")
        return redirect(_safe_next(request, reverse("accounts_budget:budget_approvals_inbox")))

    if mc.status != MCStatus.SUBMITTED:
        messages.warning(request, "Only submitted entries can be checked.")
        return redirect(_safe_next(request, reverse("accounts_budget:budget_approvals_inbox")))

    if request.method == "POST":
        remarks = (request.POST.get("remarks") or "").strip()
        st = mc_check_service(obj, scope, request.user, remarks=remarks)
        if st and st.status == "CHECKED":
            messages.success(request, "Checked successfully. Sent to Approver.")
        else:
            messages.warning(request, f"Not checked. Current status: {getattr(st,'status', '-')}")
        return redirect(_safe_next(request, reverse("accounts_budget:budget_approvals_inbox")))

    return render(request, "accounts/budget/mc_check_confirm.html", {
        "title": "Check",
        "object_label": str(obj),
        "mc": mc,
        "next": _safe_next(request, ""),
    })


@require_http_methods(["GET", "POST"])
@login_required
def mc_approve_view(request, model: str, pk: int, scope: str):
    if not _is_approver(request.user):
        messages.error(request, "You are not authorized to approve.")
        return redirect(_safe_next(request, "accounts_budget:budget_approvals_inbox"))

    obj = resolve_model_instance(model, pk)
    if not obj:
        messages.error(request, "Object not found.")
        return redirect(_safe_next(request, "accounts_budget:budget_approvals_inbox"))

    st = mc_get(obj, scope)
    if st is None:
        messages.error(request, "Cannot approve: object is not saved.")
        return redirect(_safe_next(request, "accounts_budget:budget_approvals_inbox"))

    # enforce flow
    if _has_checked_status():
        if st.status != "CHECKED":
            messages.warning(request, f"Only CHECKED entries can be approved. Current: {st.status}")
            return redirect(_safe_next(request, "accounts_budget:budget_approvals_inbox"))
    else:
        if st.status != MCStatus.SUBMITTED:
            messages.warning(request, f"Only SUBMITTED entries can be approved. Current: {st.status}")
            return redirect(_safe_next(request, "accounts_budget:budget_approvals_inbox"))

    if request.method == "POST":
        remarks = (request.POST.get("remarks") or "").strip()
        st2 = mc_approve_service(obj, scope, request.user, remarks=remarks)

        if st2 and st2.status == MCStatus.APPROVED:
            messages.success(request, "Approved successfully.")
        else:
            messages.warning(request, f"Not approved. Current: {getattr(st2, 'status', '-')}")
        return redirect(_safe_next(request, "accounts_budget:budget_approvals_inbox"))

    return render(request, "accounts/budget/mc_approve_confirm.html", {
        "title": "Approve",
        "object_label": str(obj),
        "mc": st,
        "next": _safe_next(request, "accounts_budget:budget_approvals_inbox"),
    })

@require_http_methods(["GET", "POST"])
@login_required
def mc_reject_view(request, model: str, pk: int, scope: str):
    if not _is_approver(request.user):
        messages.error(request, "You are not authorized to reject.")
        return redirect(_safe_next(request, "accounts_budget:budget_approvals_inbox"))

    obj = resolve_model_instance(model, pk)
    if not obj:
        messages.error(request, "Object not found.")
        return redirect(_safe_next(request, "accounts_budget:budget_approvals_inbox"))

    st = mc_get(obj, scope)
    if st is None:
        messages.error(request, "Cannot reject: object is not saved.")
        return redirect(_safe_next(request, "accounts_budget:budget_approvals_inbox"))

    allowed = {MCStatus.SUBMITTED, MCStatus.APPROVED}
    if _has_checked_status():
        allowed.add("CHECKED")

    if st.status not in allowed:
        messages.warning(request, f"Only SUBMITTED/CHECKED/APPROVED can be rejected. Current: {st.status}")
        return redirect(_safe_next(request, "accounts_budget:budget_approvals_inbox"))

    if request.method == "POST":
        form = MCRejectForm(request.POST)
        if form.is_valid():
            mc_reject_service(obj, scope, request.user, remarks=form.cleaned_data["remarks"])
            messages.success(request, "Rejected. Maker can edit now.")
            return redirect(_safe_next(request, "accounts_budget:budget_approvals_inbox"))
    else:
        form = MCRejectForm()

    return render(request, "accounts/budget/mc_reject.html", {
        "title": "Reject",
        "object_label": str(obj),
        "mc": st,
        "form": form,
        "next": _safe_next(request, "accounts_budget:budget_approvals_inbox"),
    })

@require_http_methods(["GET"])
@login_required
def mc_check(request, model: str, pk: int, scope: str):
    """
    Checker action:
      SUBMITTED -> CHECKED
    Uses services/maker_checker.py
    """
    if not _is_checker(request.user):
        messages.error(request, "You are not authorized to check.")
        return redirect(_safe_next(request, "accounts_budget:budget_approvals_inbox"))

    obj = resolve_model_instance(model, pk)
    if not obj:
        messages.error(request, "Object not found.")
        return redirect(_safe_next(request, "accounts_budget:budget_approvals_inbox"))

    st = mc_get(obj, scope)
    if st is None:
        messages.error(request, "Cannot check: object is not saved.")
        return redirect(_safe_next(request, "accounts_budget:budget_approvals_inbox"))

    if st.status != MCStatus.SUBMITTED:
        messages.warning(request, f"Only SUBMITTED entries can be checked. Current: {st.status}")
        return redirect(_safe_next(request, "accounts_budget:budget_approvals_inbox"))

    if not _has_checked_status():
        messages.warning(request, "CHECKED status not enabled in MCStatus. Skipping check step.")
        return redirect(_safe_next(request, "accounts_budget:budget_approvals_inbox"))

    st2 = mc_check_service(obj, scope, request.user)
    if st2 and st2.status == "CHECKED":
        messages.success(request, "Checked successfully. Sent to Approver.")
    else:
        messages.warning(request, f"Not checked. Current status: {getattr(st2, 'status', '-')}")
    return redirect(_safe_next(request, "accounts_budget:budget_approvals_inbox"))


# =============================================================================
# Approver action (CHECKED -> APPROVED) OR (SUBMITTED -> APPROVED if no CHECKED)
# =============================================================================

@require_http_methods(["GET", "POST"])
@login_required
def mc_approve(request, model: str, pk: int, scope: str):
    """
    Approver action (supports optional 3-step flow):
      If CHECKED exists: CHECKED -> APPROVED
      Else:             SUBMITTED -> APPROVED

    Uses services/maker_checker.py so schema differences are handled safely.
    """
    if not _is_approver(request.user):
        messages.error(request, "You are not authorized to approve.")
        return redirect(_safe_next(request, "accounts_budget:budget_approvals_inbox"))

    obj = resolve_model_instance(model, pk)
    if not obj:
        messages.error(request, "Object not found.")
        return redirect(_safe_next(request, "accounts_budget:budget_approvals_inbox"))

    st = mc_get(obj, scope)
    if st is None:
        messages.error(request, "Cannot approve: object is not saved.")
        return redirect(_safe_next(request, "accounts_budget:budget_approvals_inbox"))

    # ✅ enforce pre-condition (UI friendly message)
    if _has_checked_status():
        if st.status != "CHECKED":
            messages.warning(request, f"Only CHECKED entries can be approved. Current: {st.status}")
            return redirect(_safe_next(request, "accounts_budget:budget_approvals_inbox"))
    else:
        if st.status != MCStatus.SUBMITTED:
            messages.warning(request, f"Only SUBMITTED entries can be approved. Current: {st.status}")
            return redirect(_safe_next(request, "accounts_budget:budget_approvals_inbox"))

    if request.method == "POST":
        remarks = (request.POST.get("remarks") or "").strip()
        st2 = mc_approve_service(obj, scope, request.user, remarks=remarks)

        if st2 and st2.status == MCStatus.APPROVED:
            messages.success(request, "Approved successfully.")
        else:
            messages.warning(request, f"Not approved. Current: {getattr(st2, 'status', '-')}")
        return redirect(_safe_next(request, "accounts_budget:budget_approvals_inbox"))

    # ✅ GET -> confirm page (avoids approving by link click)
    return render(request, "accounts/budget/mc_approve_confirm.html", {
        "title": "Approve",
        "object_label": str(obj),
        "mc": st,
        "next": _safe_next(request, "accounts_budget:budget_approvals_inbox"),
    })


@require_http_methods(["GET", "POST"])
@login_required
def mc_reject(request, model: str, pk: int, scope: str):
    """
    Reject / disapprove (approver):
      SUBMITTED/CHECKED/APPROVED -> REJECTED
    """
    if not _is_approver(request.user):
        messages.error(request, "You are not authorized to reject.")
        return redirect(_safe_next(request, "accounts_budget:budget_approvals_inbox"))

    obj = resolve_model_instance(model, pk)
    if not obj:
        messages.error(request, "Object not found.")
        return redirect(_safe_next(request, "accounts_budget:budget_approvals_inbox"))

    st = mc_get(obj, scope)
    if st is None:
        messages.error(request, "Cannot reject: object is not saved.")
        return redirect(_safe_next(request, "accounts_budget:budget_approvals_inbox"))

    allowed = {MCStatus.SUBMITTED, MCStatus.APPROVED}
    if _has_checked_status():
        allowed.add("CHECKED")

    if st.status not in allowed:
        messages.warning(request, f"Only SUBMITTED/CHECKED/APPROVED can be rejected. Current: {st.status}")
        return redirect(_safe_next(request, "accounts_budget:budget_approvals_inbox"))

    if request.method == "POST":
        form = MCRejectForm(request.POST)
        if form.is_valid():
            mc_reject_service(obj, scope, request.user, remarks=form.cleaned_data["remarks"])
            messages.success(request, "Rejected. Maker can edit now.")
            return redirect(_safe_next(request, "accounts_budget:budget_approvals_inbox"))
    else:
        form = MCRejectForm()

    return render(request, "accounts/budget/mc_reject.html", {
        "title": "Reject",
        "object_label": str(obj),
        "mc": st,
        "form": form,
        "next": _safe_next(request, "accounts_budget:budget_approvals_inbox"),
    })

# =============================================================================
# Inbox (Maker->Checker->Approver) + BOM included
# =============================================================================

@login_required
def budget_approvals_inbox(request):
    is_checker = _is_checker(request.user)
    is_approver = _is_approver(request.user)

    if not (is_checker or is_approver):
        messages.error(request, "You are not authorized to view approvals.")
        return redirect("accounts_budget:budget_home")

    # -----------------------------
    # ✅ Category access (NEW)
    # -----------------------------
    if request.user.is_superuser:
        allowed_cats = set(c for c, _ in BudgetCategory.choices)
    else:
        allowed_cats = set(
            UserBudgetCategoryAccess.objects.filter(
                user=request.user,
                can_view=True
            ).values_list("category", flat=True)
        )

    # If user has no access at all, show empty list (or redirect if you prefer)
    if not allowed_cats:
        return render(request, "accounts/budget/budget_approvals_inbox.html", {
            "items": [],
            "status": (request.GET.get("status") or "").strip().upper(),
            "q": (request.GET.get("q") or "").strip(),
            "status_choices": [MCStatus.SUBMITTED, "CHECKED", MCStatus.APPROVED, MCStatus.REJECTED, MCStatus.DRAFT]
                            if _has_checked_status()
                            else [MCStatus.SUBMITTED, MCStatus.APPROVED, MCStatus.REJECTED, MCStatus.DRAFT],
            "is_checker": is_checker,
            "is_approver": is_approver,
        })

    # -----------------------------
    # Status choices / defaults
    # -----------------------------
    if _has_checked_status():
        default_status = MCStatus.SUBMITTED if is_checker else "CHECKED"
        status_choices = [MCStatus.SUBMITTED, "CHECKED", MCStatus.APPROVED, MCStatus.REJECTED, MCStatus.DRAFT]
    else:
        default_status = MCStatus.SUBMITTED
        status_choices = [MCStatus.SUBMITTED, MCStatus.APPROVED, MCStatus.REJECTED, MCStatus.DRAFT]

    status = (request.GET.get("status") or default_status).strip().upper()
    q = (request.GET.get("q") or "").strip()
    valid_statuses = {str(s).upper() for s in status_choices}

    # -----------------------------
    # ContentTypes
    # -----------------------------
    ct_plan     = ContentType.objects.get_for_model(BudgetPlan)
    ct_sales    = ContentType.objects.get_for_model(SalesBudget)
    ct_rmc      = ContentType.objects.get_for_model(RMCBudget)
    ct_prod_fg  = ContentType.objects.get_for_model(ProductionBudgetFG)
    ct_prod_hdr = ContentType.objects.get_for_model(ProductionBudget)
    ct_captive  = ContentType.objects.get_for_model(CaptiveConsumptionBudget)
    ct_bom      = ContentType.objects.get_for_model(ProductionBOM)

    # -----------------------------
    # ✅ Scope -> category mapping (NEW)
    # -----------------------------
    scope_category_map = {
    "SALES": BudgetCategory.SALES,    
    "RMC": BudgetCategory.RM,        
    "PROD": BudgetCategory.PRODUCTION,
    "BOM": BudgetCategory.PRODUCTION,
    "CAPTIVE": BudgetCategory.PRODUCTION,
    }

    # -----------------------------
    # ✅ Build access filter (NEW)
    # -----------------------------
    access_q = Q()

    # budget category pages: BUDGET:<category>
    access_q |= Q(
        content_type=ct_plan,
        scope__in=[f"BUDGET:{c}" for c in allowed_cats]
    )

    # other scopes (SALES/RMC/PROD/BOM/CAPTIVE) based on mapping
    for scope, cat in scope_category_map.items():
        if cat in allowed_cats:
            if scope == "PROD":
                access_q |= Q(content_type=ct_prod_fg, scope="PROD") | Q(content_type=ct_prod_hdr, scope="PROD")
            elif scope == "SALES":
                access_q |= Q(content_type=ct_sales, scope="SALES")
            elif scope == "RMC":
                access_q |= Q(content_type=ct_rmc, scope="RMC")
            elif scope == "CAPTIVE":
                access_q |= Q(content_type=ct_captive, scope="CAPTIVE")
            elif scope == "BOM":
                access_q |= Q(content_type=ct_bom, scope="BOM")

    # -----------------------------
    # MC base queryset (UPDATED)
    # -----------------------------
    mc_qs = (
        MakerCheckerState.objects
        .filter(access_q)
        .exclude(scope__in=["BUDGET:", "BUDGET"])
        .select_related("submitted_by", "checked_by")
        .order_by("-submitted_at", "-updated_at")
    )

    if status in valid_statuses:
        mc_qs = mc_qs.filter(status=status)

    # -----------------------------
    # Search (keep your existing logic, but also respect access)
    # -----------------------------
    if q:
        plan_ids = list(
            BudgetPlan.objects
            .filter(Q(fy__icontains=q) | Q(company_group__icontains=q))
            .values_list("id", flat=True)
        )

        sales_ids = list(
            SalesBudget.objects
            .filter(Q(plan__fy__icontains=q) | Q(plan__company_group__icontains=q))
            .values_list("id", flat=True)
        )

        rmc_ids = list(
            RMCBudget.objects
            .filter(Q(plan__fy__icontains=q) | Q(plan__company_group__icontains=q))
            .values_list("id", flat=True)
        )

        prod_fg_ids = list(
            ProductionBudgetFG.objects
            .filter(Q(plan__fy__icontains=q) | Q(plan__company_group__icontains=q) | Q(fg_name__icontains=q))
            .values_list("id", flat=True)
        )

        prod_hdr_ids = list(
            ProductionBudget.objects
            .filter(Q(plan__fy__icontains=q) | Q(plan__company_group__icontains=q))
            .values_list("id", flat=True)
        )

        captive_ids = list(
            CaptiveConsumptionBudget.objects
            .filter(Q(plan__fy__icontains=q) | Q(plan__company_group__icontains=q))
            .values_list("id", flat=True)
        )

        captive_ids_from_lines = list(
            CaptiveConsumptionLine.objects
            .filter(Q(item_name__icontains=q))
            .values_list("budget_id", flat=True)
            .distinct()
        )
        captive_ids = list(set(captive_ids) | set(captive_ids_from_lines))

        bom_ids = list(
            ProductionBOM.objects
            .filter(Q(fg_name__icontains=q) | Q(fg_alpha_name__icontains=q) | Q(bom_code__icontains=q))
            .values_list("id", flat=True)
        )

        # Category label match (Safety, QAQC, etc.) BUT only those allowed
        cat_label = dict(BudgetCategory.choices)
        matching_cats = [
            c for c, lbl in cat_label.items()
            if (q.lower() in (lbl or "").lower()) and (c in allowed_cats)
        ]

        search_filter = (
            Q(content_type=ct_plan, object_id__in=plan_ids) |
            Q(content_type=ct_sales, object_id__in=sales_ids) |
            Q(content_type=ct_rmc, object_id__in=rmc_ids) |
            Q(content_type=ct_prod_fg, object_id__in=prod_fg_ids) |
            Q(content_type=ct_prod_hdr, object_id__in=prod_hdr_ids) |
            Q(content_type=ct_captive, object_id__in=captive_ids) |
            Q(content_type=ct_bom, object_id__in=bom_ids)
        )

        if matching_cats:
            search_filter |= Q(
                content_type=ct_plan,
                scope__in=[f"BUDGET:{c}" for c in matching_cats]
            )

        # ✅ IMPORTANT: keep access restriction
        mc_qs = mc_qs.filter(search_filter)

    mc_list = list(mc_qs[:500])

    # -----------------------------
    # bulk load objects (same as your code)
    # -----------------------------
    plan_ids     = [m.object_id for m in mc_list if m.content_type_id == ct_plan.id]
    sales_ids    = [m.object_id for m in mc_list if m.content_type_id == ct_sales.id]
    rmc_ids      = [m.object_id for m in mc_list if m.content_type_id == ct_rmc.id]
    prod_fg_ids  = [m.object_id for m in mc_list if m.content_type_id == ct_prod_fg.id]
    prod_hdr_ids = [m.object_id for m in mc_list if m.content_type_id == ct_prod_hdr.id]
    captive_ids  = [m.object_id for m in mc_list if m.content_type_id == ct_captive.id]
    bom_ids      = [m.object_id for m in mc_list if m.content_type_id == ct_bom.id]

    plans_map     = BudgetPlan.objects.in_bulk(plan_ids)
    sales_map     = SalesBudget.objects.select_related("plan").in_bulk(sales_ids)
    rmc_map       = RMCBudget.objects.select_related("plan").in_bulk(rmc_ids)
    prod_fg_map   = ProductionBudgetFG.objects.select_related("plan").in_bulk(prod_fg_ids)
    prod_hdr_map  = ProductionBudget.objects.select_related("plan").in_bulk(prod_hdr_ids)
    captive_map   = CaptiveConsumptionBudget.objects.select_related("plan").in_bulk(captive_ids)
    bom_map       = ProductionBOM.objects.in_bulk(bom_ids)

    cat_label = dict(BudgetCategory.choices)

    # -----------------------------
    # Aggregates (same as your code)
    # -----------------------------
    agg_map = {}
    budget_pairs = []
    for m in mc_list:
        if m.content_type_id != ct_plan.id:
            continue
        cat = (m.scope.split(":", 1)[1] if ":" in m.scope else "").strip()
        if cat:
            budget_pairs.append((m.object_id, cat))

    if budget_pairs:
        uniq_plan_ids = list({p for p, _ in budget_pairs})
        uniq_cats = list({c for _, c in budget_pairs})
        rows = (
            BudgetLine.objects
            .filter(plan_id__in=uniq_plan_ids, category__in=uniq_cats)
            .order_by()
            .values("plan_id", "category")
            .annotate(row_count=Count("id"), total=Sum("total"), last_updated=Max("updated_at"))
        )
        agg_map = {(r["plan_id"], r["category"]): r for r in rows}

    sales_counts = {}
    if sales_ids:
        last_f = _max_field_if_exists(SalesBudgetLine, "updated_at", "id")
        srows = (
            SalesBudgetLine.objects
            .filter(budget_id__in=sales_ids)
            .order_by()
            .values("budget_id")
            .annotate(c=Count("id"), last_updated=Max(last_f))
        )
        sales_counts = {r["budget_id"]: r for r in srows}

    rmc_counts = {}
    if rmc_ids:
        last_f = _max_field_if_exists(RMCBudgetLine, "updated_at", "id")
        rrows = (
            RMCBudgetLine.objects
            .filter(budget_id__in=rmc_ids)
            .order_by()
            .values("budget_id")
            .annotate(c=Count("id"), last_updated=Max(last_f))
        )
        rmc_counts = {r["budget_id"]: r for r in rrows}

    prod_fg_counts = {}
    if prod_fg_ids:
        last_f = _max_field_if_exists(ProductionBudgetLine, "updated_at", "id")
        prows = (
            ProductionBudgetLine.objects
            .filter(budget_id__in=prod_fg_ids)
            .order_by()
            .values("budget_id")
            .annotate(c=Count("id"), last_updated=Max(last_f))
        )
        prod_fg_counts = {r["budget_id"]: r for r in prows}

    captive_agg = {}
    if captive_ids:
        last_f = _max_field_if_exists(CaptiveConsumptionLine, "updated_at", "id")
        cap_rows = (
            CaptiveConsumptionLine.objects
            .filter(budget_id__in=captive_ids)
            .order_by()
            .values("budget_id")
            .annotate(row_count=Count("id"), total=Sum("amount"), last_updated=Max(last_f))
        )
        captive_agg = {r["budget_id"]: r for r in cap_rows}

    bom_input_counts = {}
    bom_eff_counts = {}
    if bom_ids:
        in_rows = (
            ProductionBOMInputLine.objects
            .filter(bom_id__in=bom_ids)
            .order_by()
            .values("bom_id")
            .annotate(c=Count("id"))
        )
        bom_input_counts = {r["bom_id"]: r["c"] for r in in_rows}

        ef_rows = (
            ProductionBOMEffluentLine.objects
            .filter(bom_id__in=bom_ids)
            .order_by()
            .values("bom_id")
            .annotate(c=Count("id"))
        )
        bom_eff_counts = {r["bom_id"]: r["c"] for r in ef_rows}

    # -----------------------------
    # Build items (same as your code)
    # -----------------------------
    items = []
    for mc in mc_list:
        ModelCls = mc.content_type.model_class()
        model_path = f"{ModelCls._meta.app_label}.{ModelCls.__name__}" if ModelCls else ""

        if mc.content_type_id == ct_plan.id:
            plan = plans_map.get(mc.object_id)
            category = (mc.scope.split(":", 1)[1] if ":" in mc.scope else "").strip()
            if not category:
                continue
            a = agg_map.get((mc.object_id, category), {})
            items.append({
                "mc": mc, "plan": plan, "category": category,
                "category_label": cat_label.get(category, category),
                "row_count": a.get("row_count") or 0,
                "total": a.get("total") or 0,
                "last_updated": a.get("last_updated"),
                "model_path": model_path,
                "obj_pk": mc.object_id,
                "view_kind": "BUDGET",
            })
            continue

        if mc.content_type_id == ct_sales.id:
            obj = sales_map.get(mc.object_id)
            plan = getattr(obj, "plan", None)
            a = sales_counts.get(mc.object_id, {})
            items.append({
                "mc": mc, "plan": plan,
                "category": "", "category_label": "Sales Budget",
                "row_count": a.get("c") or 0,
                "total": 0,
                "last_updated": a.get("last_updated") or getattr(obj, "updated_at", None),
                "model_path": model_path,
                "obj_pk": mc.object_id,
                "view_kind": "SALES",
            })
            continue

        if mc.content_type_id == ct_rmc.id:
            obj = rmc_map.get(mc.object_id)
            plan = getattr(obj, "plan", None)
            a = rmc_counts.get(mc.object_id, {})
            items.append({
                "mc": mc, "plan": plan,
                "category": "", "category_label": "RMC Budget",
                "row_count": a.get("c") or 0,
                "total": 0,
                "last_updated": a.get("last_updated") or getattr(obj, "updated_at", None),
                "model_path": model_path,
                "obj_pk": mc.object_id,
                "view_kind": "RMC",
            })
            continue

        if mc.content_type_id == ct_prod_fg.id:
            obj = prod_fg_map.get(mc.object_id)
            plan = getattr(obj, "plan", None)
            fg = getattr(obj, "fg_name", "") or ""
            a = prod_fg_counts.get(mc.object_id, {})
            items.append({
                "mc": mc, "plan": plan,
                "category": "", "category_label": f"Production Budget - {fg}".strip(" -"),
                "row_count": a.get("c") or 0,
                "total": 0,
                "last_updated": a.get("last_updated") or getattr(obj, "updated_at", None),
                "model_path": model_path,
                "obj_pk": mc.object_id,
                "view_kind": "PROD",
            })
            continue

        if mc.content_type_id == ct_prod_hdr.id:
            obj = prod_hdr_map.get(mc.object_id)
            plan = getattr(obj, "plan", None)
            items.append({
                "mc": mc, "plan": plan,
                "category": "", "category_label": "Production Budget",
                "row_count": 0, "total": 0,
                "last_updated": getattr(obj, "updated_at", None),
                "model_path": model_path,
                "obj_pk": mc.object_id,
                "view_kind": "PROD",
            })
            continue

        if mc.content_type_id == ct_captive.id:
            obj = captive_map.get(mc.object_id)
            plan = getattr(obj, "plan", None)
            a = captive_agg.get(mc.object_id, {})
            items.append({
                "mc": mc, "plan": plan,
                "category": "", "category_label": "Captive Consumption",
                "row_count": a.get("row_count") or 0,
                "total": a.get("total") or 0,
                "last_updated": a.get("last_updated") or getattr(obj, "updated_at", None),
                "model_path": model_path,
                "obj_pk": mc.object_id,
                "view_kind": "CAPTIVE",
            })
            continue

        if mc.content_type_id == ct_bom.id:
            bom = bom_map.get(mc.object_id)
            if not bom:
                continue
            in_c = bom_input_counts.get(bom.id, 0)
            ef_c = bom_eff_counts.get(bom.id, 0)
            items.append({
                "mc": mc,
                "plan": None,
                "category": "",
                "category_label": f"BOM - {bom.fg_name}".strip(),
                "row_count": int(in_c) + int(ef_c),
                "total": 0,
                "last_updated": getattr(bom, "updated_at", None),
                "model_path": model_path,
                "obj_pk": bom.id,
                "view_kind": "BOM",
                "fg_name": bom.fg_name,
                "bom_code": getattr(bom, "bom_code", "") or "",
            })
            continue

    return render(request, "accounts/budget/budget_approvals_inbox.html", {
        "items": items,
        "status": status,
        "q": q,
        "status_choices": status_choices,
        "is_checker": is_checker,
        "is_approver": is_approver,
    })

@login_required
@permission_required("Budget.can_approve_budgets", raise_exception=True)
def mc_reopen_view(request, model, pk, scope):
    """
    Reopen an APPROVED entry for editing again.
    Keeps existing flow: uses ?next=... and redirects to inbox by default.
    """
    next_url = request.GET.get("next") or reverse("accounts_budget:budget_approvals_inbox")

    Model = _get_model(model)  # uses your fixed resolver
    obj = get_object_or_404(Model, pk=pk)

    mc = mc_get(obj, scope)
    if not mc:
        messages.error(request, "Approval record not found.")
        return redirect(next_url)

    if mc.status != "APPROVED":
        messages.warning(request, f"Only APPROVED entries can be reopened. Current status: {mc.status}")
        return redirect(next_url)

    # Optional remarks (you can pass &remarks=... or later make it POST)
    reopen_note = (request.GET.get("remarks") or "").strip()

    with transaction.atomic():
        # Core behavior: make it editable again
        mc.status = "DRAFT"

        # Keep history in remarks (without breaking if remarks is null)
        if hasattr(mc, "remarks"):
            stamp = timezone.now().strftime("%Y-%m-%d %H:%M")
            msg = f"Reopened by {request.user} at {stamp}"
            if reopen_note:
                msg += f" | {reopen_note}"
            mc.remarks = (mc.remarks + "\n" if mc.remarks else "") + msg

        # If your MC has these fields, clear approval metadata safely
        for fld in ("approved_at", "approved_on"):
            if hasattr(mc, fld):
                setattr(mc, fld, None)
        for fld in ("approved_by",):
            if hasattr(mc, fld):
                setattr(mc, fld, None)

        # If you have reopen audit fields, fill them
        if hasattr(mc, "reopened_at"):
            mc.reopened_at = timezone.now()
        if hasattr(mc, "reopened_by"):
            mc.reopened_by = request.user

        mc.save()

    messages.success(request, "Budget reopened and moved to DRAFT. It is editable now.")
    return redirect(next_url)
# =============================================================================
# BUDGET HOME / PLAN CREATE / CATEGORY EDIT
# =============================================================================

ALLOWED_BOM_TYPES = ("Key Raw Material", "Raw Material", "Work in Progress","Semi Finished Good","WIP FR",)


def _current_fy_label(d: date) -> str:
    """
    Indian FY: Apr..Mar.
    Example: if date is 2026-01-08 => FY is 2025-26
    """
    if d.month >= 4:
        start = d.year
        end = d.year + 1
    else:
        start = d.year - 1
        end = d.year
    return f"{start}-{str(end)[-2:]}"


def _months_meta(plan: BudgetPlan):
    """
    Only used for UI labels. Fields remain fixed Apr..Mar.
    """
    fy_start = int(plan.fy.split("-")[0])
    return [
        ("apr", f"Apr-{str(fy_start)[-2:]}"),
        ("may", f"May-{str(fy_start)[-2:]}"),
        ("jun", f"Jun-{str(fy_start)[-2:]}"),
        ("jul", f"Jul-{str(fy_start)[-2:]}"),
        ("aug", f"Aug-{str(fy_start)[-2:]}"),
        ("sep", f"Sep-{str(fy_start)[-2:]}"),
        ("oct", f"Oct-{str(fy_start)[-2:]}"),
        ("nov", f"Nov-{str(fy_start)[-2:]}"),
        ("dec", f"Dec-{str(fy_start)[-2:]}"),
        ("jan", f"Jan-{str(fy_start + 1)[-2:]}"),
        ("feb", f"Feb-{str(fy_start + 1)[-2:]}"),
        ("mar", f"Mar-{str(fy_start + 1)[-2:]}"),
    ]


CATEGORY_CARDS = [
    (BudgetCategory.SAFETY.value,      BudgetCategory.SAFETY.label,      "fa-shield-heart"),
    (BudgetCategory.ENVIRONMENT.value, BudgetCategory.ENVIRONMENT.label, "fa-leaf"),
    (BudgetCategory.QAQC.value,        BudgetCategory.QAQC.label,        "fa-flask"),
    (BudgetCategory.ENGINEERING.value, BudgetCategory.ENGINEERING.label, "fa-screwdriver-wrench"),
    (BudgetCategory.UTILITY.value,     BudgetCategory.UTILITY.label,     "fa-bolt"),
    (BudgetCategory.ADMIN.value,       BudgetCategory.ADMIN.label,       "fa-building"),
    (BudgetCategory.HR.value,          BudgetCategory.HR.label,          "fa-users"),
    (BudgetCategory.RD.value,          BudgetCategory.RD.label,          "fa-flask-vial"),
    (BudgetCategory.LOGISTIC.value,    BudgetCategory.LOGISTIC.label,    "fa-truck"),
    (BudgetCategory.FIN_ACCTS.value,   BudgetCategory.FIN_ACCTS.label,   "fa-coins"),
]

def _get_or_create_current_plan(request):
    fy_default = _current_fy_label(date.today())
    plan, _ = BudgetPlan.objects.get_or_create(
        fy=fy_default,
        company_group="",
        defaults={"created_by": request.user},
    )
    return plan


@login_required
def budget_home(request):
    fy_default = _current_fy_label(date.today())
    plan_id = request.GET.get("plan")

    if plan_id:
        plan = get_object_or_404(BudgetPlan, pk=plan_id)
    else:
        plan, _ = BudgetPlan.objects.get_or_create(
            fy=fy_default,
            company_group="",
            defaults={"created_by": request.user},
        )

    plans = BudgetPlan.objects.all().order_by("-id")

    totals_qs = (
        BudgetLine.objects
        .filter(plan=plan)
        .order_by()
        .values("category")
        .annotate(t=Sum("total"))
    )
    totals = {r["category"]: (r["t"] or Decimal("0.00")) for r in totals_qs}

    card_rows = []
    for cat, label, icon in CATEGORY_CARDS:
        card_rows.append({
            "cat": cat,
            "label": label,
            "icon": icon,
            "total": totals.get(cat, Decimal("0.00")) or Decimal("0.00"),
            # Optional display:
            # "mc": mc_get(plan, f"BUDGET:{cat}"),
        })

    return render(request, "accounts/budget/budget_home.html", {
        "plan": plan,
        "plans": plans,
        "cards": card_rows,
    })


@login_required
def budget_plan_create(request):
    if request.method == "POST":
        form = BudgetPlanForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.created_by = request.user
            obj.save()
            messages.success(request, "Budget Plan created.")
            return redirect("accounts_budget:budget_home")
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = BudgetPlanForm(initial={
            "fy": _current_fy_label(date.today()),
            "is_active": True
        })

    return render(request, "accounts/budget/budget_plan_create.html", {"form": form})

# =============================================================================
# BUDGET CATEGORY EDIT
# ============================================================================
@login_required
def budget_category_edit(request, plan_id: int, category: str):
    plan = get_object_or_404(BudgetPlan, pk=plan_id)

    valid_cats = {c for c, _ in BudgetCategory.choices}
    if category not in valid_cats:
        messages.error(request, "Invalid budget category.")
        return redirect("accounts_budget:budget_home")

    # ✅ Maker-checker scope per category (locks only that page)
    scope = f"BUDGET:{category}"
    mc = mc_get(plan, scope)
    locked = _is_locked(mc)

    # Block edits if locked (but allow viewing same URL)
    if request.method == "POST" and locked:
        messages.warning(request, "This entry is locked. It can be edited only if disapproved by approver.")
        return redirect("accounts_budget:budget_category_edit", plan_id=plan.id, category=category)

    # Only show/save this category's lines
    qs = BudgetLine.objects.filter(plan=plan, category=category).order_by("sr_no", "id")


    if request.method == "POST":
        formset = BudgetLineFormSet(
            request.POST,
            instance=plan,
            queryset=qs,  # ✅ filtered inline queryset
            form_kwargs={"locked": locked},  # keeps existing lock mechanism style
        )
        formset.category_value = category

        if formset.is_valid():
            formset.save()

            # ✅ After successful save -> submit (locks)
            mc_submit(plan, scope, request.user)

            messages.success(request, "Budget saved successfully.")
            return redirect("accounts_budget:budget_category_edit", plan_id=plan.id, category=category)

        messages.error(request, "Please correct the highlighted errors.")
    else:
        formset = BudgetLineFormSet(
            instance=plan,
            queryset=qs,  # ✅ filtered inline queryset
            form_kwargs={"locked": locked},
        )
        formset.category_value = category

    # If locked, ensure DELETE also disabled (defensive; mixin already does it)
    if locked:
        for f in formset.forms:
            if "DELETE" in f.fields:
                f.fields["DELETE"].disabled = True

    months = _months_meta(plan)
    label_map = dict(BudgetCategory.choices)
    title = label_map.get(category, "Budget")

    agg = qs.aggregate(
        apr=Sum("apr"), may=Sum("may"), jun=Sum("jun"), jul=Sum("jul"), aug=Sum("aug"), sep=Sum("sep"),
        oct=Sum("oct"), nov=Sum("nov"), dec=Sum("dec"), jan=Sum("jan"), feb=Sum("feb"), mar=Sum("mar"),
        total=Sum("total"),
    )

    return render(request, "accounts/budget/budget_category_edit.html", {
        "plan": plan,
        "category": category,
        "title": title,
        "months": months,
        "formset": formset,
        "agg": agg,

        # ✅ Approval UI flags
        "mc": mc,
        "is_locked": locked,
        "is_approver": _is_approver(request.user),
        "mc_scope": scope,
        "mc_model": "accounts_budget.BudgetPlan",
        "mc_pk": plan.id,
    })

@login_required
def budget_category_current(request, category: str):
    # ✅ Only allow active categories for sidebar/current route
    active_cats = {c.value for c in BudgetCategory.active()}
    if category not in active_cats:
        messages.error(request, "Invalid budget category.")
        return redirect("accounts_budget:budget_home")

    plan = _get_or_create_current_plan(request)
    return redirect("accounts_budget:budget_category_edit", plan_id=plan.id, category=category)

@login_required
def budget_line_attachment_upload(request, line_id):
    line = get_object_or_404(BudgetLine, pk=line_id)

    # BudgetLine belongs to BudgetPlan via FK: line.plan
    plan = getattr(line, "plan", None) or getattr(line, "budget_plan", None)
    if not plan:
        raise Http404("Plan not found for this line.")

    # ✅ maker-checker scope must match Budget category pages
    category = (getattr(line, "category", "") or "").strip()
    scope = f"BUDGET:{category}" if category else "BUDGET:UNKNOWN"

    mc = mc_get(plan, scope)
    if _is_locked(mc):
        messages.warning(request, "Budget is locked. Attachments cannot be changed.")
        return redirect(request.GET.get("next") or "accounts_budget:budget_home")

    if request.method == "POST":
        form = BudgetLineAttachmentForm(request.POST, request.FILES)
        if form.is_valid():
            att = form.save(commit=False)
            att.line = line
            att.uploaded_by = request.user
            att.save()
            messages.success(request, "Attachment uploaded.")
            return redirect(
                request.POST.get("next")
                or request.GET.get("next")
                or "accounts_budget:budget_home"
            )
    else:
        form = BudgetLineAttachmentForm()

    return render(request, "accounts/budget/budget_line_attachment_upload.html", {
        "line": line,
        "form": form,
        "next": request.GET.get("next") or "",
    })


@require_POST
@login_required
def budget_line_attachment_delete(request, att_id):
    att = get_object_or_404(BudgetLineAttachment, pk=att_id)
    line = att.line

    plan = getattr(line, "plan", None) or getattr(line, "budget_plan", None)
    if not plan:
        raise Http404("Plan not found for this line.")

    category = (getattr(line, "category", "") or "").strip()
    scope = f"BUDGET:{category}" if category else "BUDGET:UNKNOWN"

    mc = mc_get(plan, scope)
    if _is_locked(mc):
        messages.warning(request, "Budget is locked. Attachments cannot be changed.")
        return redirect(request.POST.get("next") or "accounts_budget:budget_home")

    att.delete()
    messages.success(request, "Attachment deleted.")
    return redirect(request.POST.get("next") or "accounts_budget:budget_home")

# =============================================================================
# PRODUCTION BOM (same flow; FIX: prevent saving blank/duplicate lines + hide old blank lines)
# =============================================================================
# Only these types should be selectable in BOM input material dropdown
BOM_SELECT_TYPES = [
    # ✅ RM
    "Raw Material",
    "Key Raw Material",

    # ✅ FG variants (your DB shows "Finished Good")
    "Finished Good",
    "Finished Goods",
    "FG",
    "F.G",
    "F.G.",

    # ✅ SFG variants
    "SFG",
    "Semi Finished Good",
    "Semi Finished Goods",
    "Semi-Finished Good",
    "Semi-Finished Goods",

    # ✅ NEW: Packing Material variants
    "Packing Material",
    "Packing Materials",
    "PM",
    "P.M",
    "P.M.",
]

def _bom_select_type_q(type_field: str = "type"):
    """
    Build a Q() filter for selectable BOM material types using the correct field name.
    Works even if ERP schema uses item_type/material_type/material_category instead of 'type'.
    """
    q = Q()

    # exact matches
    for t in BOM_SELECT_TYPES:
        q |= Q(**{f"{type_field}__iexact": t})

    # tolerance (avoid __iregex if SQL Server backend doesn't support it reliably)
    q |= Q(**{f"{type_field}__icontains": "Finished Good"})
    q |= Q(**{f"{type_field}__icontains": "Semi Finished"})
    q |= Q(**{f"{type_field}__icontains": "SFG"})
    q |= Q(**{f"{type_field}__icontains": "FG"})

    # ✅ Packing Material tolerance
    q |= Q(**{f"{type_field}__icontains": "Packing"})
    q |= Q(**{f"{type_field}__icontains": "Pack"})
    q |= Q(**{f"{type_field}__icontains": "PM"})

    return q
# -----------------------------------------------------------------------------
# FG lookup
# -----------------------------------------------------------------------------
@login_required
def production_bom_fg_lookup_json(request):
    q = (request.GET.get("q") or "").strip()

    qs = (
        ERPBOMRow.objects
        .exclude(fg_name__isnull=True)
        .exclude(fg_name__exact="")
    )

    if q:
        qs = qs.filter(fg_name__icontains=q)

    fg_names = list(
        qs.values_list("fg_name", flat=True)
          .distinct()
          .order_by("fg_name")[:50]
    )

    alpha_map = {}
    try:
        alpha_map = {
            r["fg_name"]: (r.get("fg_alpha_name") or "")
            for r in ProductionBudgetFG.objects.filter(fg_name__in=fg_names)
                                              .values("fg_name", "fg_alpha_name")
        }
    except Exception:
        alpha_map = {}

    results = []
    for name in fg_names:
        alpha = alpha_map.get(name, "")
        results.append({
            "value": name,
            "text": f"{name} ({alpha})" if alpha else name,
            "alpha": alpha,
        })

    return JsonResponse({"results": results})

@login_required
def production_bom_material_lookup_json(request):
    """
    Material search for Add Row dropdown:
    Returns materials filtered to selectable BOM types:
      - Key Raw Material
      - Raw Material
      - Finished Goods
      - SFG
      - ✅ Packing Material (+ synonyms)
    Uses ERPBOMRow as source (distinct by bom_item_code).
    """
    q = (request.GET.get("q") or "").strip()

    # ✅ IMPORTANT: make it robust across ERP schemas (type vs item_type vs material_type etc)
    typ_f  = _mfield(ERPBOMRow, "type", "item_type", "material_type", "material_category")
    code_f = _mfield(ERPBOMRow, "bom_item_code", "item_code", "code")
    name_f = _mfield(ERPBOMRow, "bom_item_name", "item_name", "name")
    unit_f = _mfield(ERPBOMRow, "unit", "uom")

    # keep same endpoint behavior, but return safe error if schema mismatch
    if not (typ_f and code_f and name_f):
        return JsonResponse({
            "results": [],
            "error": "ERPBOMRow missing required fields",
            "missing": {"type": typ_f, "code": code_f, "name": name_f}
        }, status=500)

    qs = (
        ERPBOMRow.objects
        # ✅ filter by selectable types INCLUDING PACKING MATERIAL
        .filter(_bom_select_type_q(typ_f))
        .exclude(**{f"{code_f}__isnull": True}).exclude(**{f"{code_f}__exact": ""})
        .exclude(**{f"{name_f}__isnull": True}).exclude(**{f"{name_f}__exact": ""})
    )

    if q:
        qs = qs.filter(
            Q(**{f"{code_f}__icontains": q}) |
            Q(**{f"{name_f}__icontains": q})
        )

    rows = (
        qs.values(code_f, typ_f)
          .annotate(
              material_name=Min(name_f),
              unit=Min(unit_f) if unit_f else Min(name_f),
          )
          .order_by(code_f)[:50]
    )

    results = []
    for r in rows:
        code = (r.get(code_f) or "").strip()
        name = (r.get("material_name") or "").strip()
        typ  = (r.get(typ_f) or "").strip()
        unit = (r.get("unit") or "").strip()

        results.append({
            "value": code,
            "text": name or code,
            "code": code,
            "name": name,
            "type": typ,
            "unit": unit,
        })

    return JsonResponse({"results": results})

# -----------------------------------------------------------------------------
# BOM list
# -----------------------------------------------------------------------------
@login_required
def production_bom_list(request):
    q = (request.GET.get("q") or "").strip()

    qs = ProductionBOM.objects.all().order_by("fg_name")
    if q:
        qs = qs.filter(Q(fg_name__icontains=q) | Q(bom_code__icontains=q))

    boms = list(qs)

    for b in boms:
        mc = None
        try:
            mc = mc_get(b, "BOM")
        except Exception:
            mc = None

        b.approval_status = (getattr(mc, "status", "") or "").upper() if mc else ""
        b.is_locked = _is_locked(mc) if mc else False

    return render(request, "accounts/budget/production_bom_list.html", {
        "q": q,
        "boms": boms,
    })


# -----------------------------------------------------------------------------
# Helpers (prefill) - unchanged
# -----------------------------------------------------------------------------
def _first_attr(obj, names):
    for n in names:
        if hasattr(obj, n):
            v = getattr(obj, n, None)
            if v not in (None, ""):
                return v
    return None


def _erp_code_name_map(item_codes):
    codes = []
    for c in (item_codes or []):
        c = (c or "").strip()
        if c:
            codes.append(c)
    codes = sorted(set(codes))
    if not codes:
        return {}

    rows = (
        ERPBOMRow.objects
        .filter(bom_item_code__in=codes)
        .exclude(bom_item_name__isnull=True).exclude(bom_item_name__exact="")
        .values("bom_item_code")
        .annotate(material_name=Min("bom_item_name"))
    )

    return {r["bom_item_code"]: (r.get("material_name") or "") for r in rows}


def _material_prefill_payload_from_formset(fs):
    codes = []
    for obj in fs.queryset:
        code = _first_attr(obj, ["bom_item_code", "material_code", "material_name"])
        if code:
            codes.append(str(code).strip())

    name_map = _erp_code_name_map(codes)

    payload = []
    for c in sorted(set(codes)):
        nm = name_map.get(c, "").strip()
        text = f"{c} — {nm}" if nm else c
        payload.append({"value": c, "text": text, "name": nm})
    return payload


# -----------------------------------------------------------------------------
# ✅ FIX: robust blank/duplicate handling (SERVER-SIDE)
# -----------------------------------------------------------------------------
def _has_field(Model, name: str) -> bool:
    try:
        Model._meta.get_field(name)
        return True
    except FieldDoesNotExist:
        return False


def _zeroish(v) -> bool:
    if v is None:
        return True
    if isinstance(v, bool):
        return False
    if isinstance(v, Decimal):
        return v == Decimal("0")
    if isinstance(v, (int, float)):
        return v == 0
    s = str(v).strip()
    return s in ("", "0", "0.0", "0.00", "0.000", "0.0000", "0.00000", "0.000000")


def _blank_q_for_input_model(Model):
    """
    Blank Input row = NO material selected AND norms are zero.
    (category/sr_no should NOT create DB rows)
    """
    code_field = "bom_item_code" if _has_field(Model, "bom_item_code") else None
    name_field = "material_name" if _has_field(Model, "material_name") else None
    bn_field = "budget_norm" if _has_field(Model, "budget_norm") else ("norm" if _has_field(Model, "norm") else None)
    tn_field = "target_norm" if _has_field(Model, "target_norm") else ("mat_qty_mt" if _has_field(Model, "mat_qty_mt") else None)

    q = Q()
    if code_field:
        q &= (Q(**{f"{code_field}__isnull": True}) | Q(**{f"{code_field}__exact": ""}))
    if name_field:
        q &= (Q(**{f"{name_field}__isnull": True}) | Q(**{f"{name_field}__exact": ""}))
    if bn_field:
        q &= (Q(**{f"{bn_field}__isnull": True}) | Q(**{bn_field: Decimal("0")}))
    if tn_field:
        q &= (Q(**{f"{tn_field}__isnull": True}) | Q(**{tn_field: Decimal("0")}))

    return q


def _blank_q_for_eff_model(Model):
    wt_field = "waste_type" if _has_field(Model, "waste_type") else None
    wn_field = "waste_name" if _has_field(Model, "waste_name") else None
    bn_field = "waste_budget_norm" if _has_field(Model, "waste_budget_norm") else ("norm" if _has_field(Model, "norm") else None)
    tn_field = "waste_target_norm" if _has_field(Model, "waste_target_norm") else ("qty" if _has_field(Model, "qty") else None)

    q = Q()
    if wt_field:
        q &= (Q(**{f"{wt_field}__isnull": True}) | Q(**{f"{wt_field}__exact": ""}))
    if wn_field:
        q &= (Q(**{f"{wn_field}__isnull": True}) | Q(**{f"{wn_field}__exact": ""}))
    if bn_field:
        q &= (Q(**{f"{bn_field}__isnull": True}) | Q(**{bn_field: Decimal("0")}))
    if tn_field:
        q &= (Q(**{f"{tn_field}__isnull": True}) | Q(**{tn_field: Decimal("0")}))

    return q


def _safe_save_bom_input_formset(fs, bom):
    """
    ✅ Saves input formset without creating blank rows and without duplicates.
    Rules:
      - if DELETE => delete instance
      - if no code + no name => skip (and delete existing)
      - if code/name duplicates => skip (and delete existing)
      - if truly blank (no code/name and norms zero) => skip (and delete existing)
    """
    seen = set()

    for form in fs.forms:
        cd = getattr(form, "cleaned_data", None) or {}
        if not cd:
            continue

        inst = form.instance

        if cd.get("DELETE"):
            if inst and inst.pk:
                inst.delete()
            continue

        code = (cd.get("bom_item_code") or "").strip()
        name = (cd.get("material_name") or "").strip()
        bn = cd.get("budget_norm", cd.get("norm"))
        tn = cd.get("target_norm", cd.get("mat_qty_mt"))

        # truly blank
        if not code and not name and _zeroish(bn) and _zeroish(tn):
            if inst and inst.pk:
                inst.delete()
            continue

        # category-only / not selected material -> do not save
        if not code and not name:
            if inst and inst.pk:
                inst.delete()
            continue

        key = (code or name).strip().lower()
        if key in seen:
            if inst and inst.pk:
                inst.delete()
            continue
        seen.add(key)

        if (not inst.pk) or form.has_changed():
            obj = form.save(commit=False)
            obj.bom = bom
            obj.save()

    # if any form has m2m (rare for inline), keep safe
    if hasattr(fs, "save_m2m"):
        try:
            fs.save_m2m()
        except Exception:
            pass


def _safe_save_bom_eff_formset(fs, bom):
    """
    Same idea for effluent/waste.
    """
    seen = set()

    for form in fs.forms:
        cd = getattr(form, "cleaned_data", None) or {}
        if not cd:
            continue

        inst = form.instance

        if cd.get("DELETE"):
            if inst and inst.pk:
                inst.delete()
            continue

        wt = (cd.get("waste_type") or "").strip()
        wn = (cd.get("waste_name") or "").strip()
        bn = cd.get("waste_budget_norm", cd.get("norm"))
        tn = cd.get("waste_target_norm", cd.get("qty"))

        if not wt and not wn and _zeroish(bn) and _zeroish(tn):
            if inst and inst.pk:
                inst.delete()
            continue

        if not wt and not wn:
            if inst and inst.pk:
                inst.delete()
            continue

        key = f"{wt}||{wn}".lower()
        if key in seen:
            if inst and inst.pk:
                inst.delete()
            continue
        seen.add(key)

        if (not inst.pk) or form.has_changed():
            obj = form.save(commit=False)
            obj.bom = bom
            obj.save()

    if hasattr(fs, "save_m2m"):
        try:
            fs.save_m2m()
        except Exception:
            pass


# -----------------------------------------------------------------------------
# Inputs for FG (JSON) - unchanged
# -----------------------------------------------------------------------------
def _mfield(Model, *names):
    for n in names:
        try:
            Model._meta.get_field(n)
            return n
        except FieldDoesNotExist:
            pass
    return None

@login_required
def production_bom_inputs_for_fg_json(request):
    fg = (request.GET.get("fg_name") or "").strip()
    if not fg:
        return JsonResponse({"rows": []})

    qty_f  = _mfield(ERPBOMRow, "bom_qty", "qty", "mat_qty_mt", "norm")
    seq_f  = _mfield(ERPBOMRow, "seq_id", "seq", "sr_no", "id")
    unit_f = _mfield(ERPBOMRow, "unit", "uom")
    typ_f  = _mfield(ERPBOMRow, "type", "item_type", "material_type", "material_category")
    code_f = _mfield(ERPBOMRow, "bom_item_code", "item_code", "code")
    name_f = _mfield(ERPBOMRow, "bom_item_name", "item_name", "name")
    fg_f   = _mfield(ERPBOMRow, "fg_name", "fg", "fgitemname", "finished_good")

    missing = [k for k,v in {
        "qty": qty_f, "seq": seq_f, "type": typ_f,
        "code": code_f, "name": name_f, "fg": fg_f
    }.items() if not v]
    if missing:
        return JsonResponse({"rows": [], "error": "ERPBOMRow missing fields", "missing": missing}, status=500)

    try:
        qs = (
            ERPBOMRow.objects
            .filter(**{fg_f: fg})
            .values(code_f, typ_f)
            .annotate(
                material_name=Min(name_f),
                unit=Min(unit_f) if unit_f else Min(name_f),
                budget_norm=Sum(qty_f),
                seq_id=Min(seq_f),
            )
            .order_by("seq_id", code_f)
        )

        out = []
        for r in qs:
            code = (r.get(code_f) or "").strip()
            typ  = (r.get(typ_f) or "").strip()
            nm   = (r.get("material_name") or "").strip()
            out.append({
                "bom_item_code": code,
                "material_category": typ,
                "material_name": nm,
                "unit": (r.get("unit") or "").strip() if unit_f else "",
                "budget_norm": str(r.get("budget_norm") or Decimal("0.000000")),
            })

        return JsonResponse({"rows": out})

    except Exception as e:
        return JsonResponse({"rows": [], "error": str(e)}, status=500)
# -----------------------------------------------------------------------------
# Create
# -----------------------------------------------------------------------------
@login_required
def production_bom_create(request):
    """
    Same flow as your current code.
    ✅ Changes:
      - uses safe-save for formsets (prevents blank/duplicate DB rows)
    """

    if request.method == "POST":
        fg = (request.POST.get("fg_name") or "").strip()
        if fg:
            existing = ProductionBOM.objects.filter(fg_name__iexact=fg).first()
            if existing:
                mc_existing = mc_get(existing, "BOM")
                if _is_locked(mc_existing):
                    messages.warning(
                        request,
                        f"BOM already exists for FG '{fg}' and is approved/locked. Opening in View mode."
                    )
                    return redirect("accounts_budget:production_bom_view", bom_id=existing.id)

                messages.warning(request, f"BOM already exists for FG '{fg}'. Opening existing BOM.")
                return redirect("accounts_budget:production_bom_edit", bom_id=existing.id)

        bom = ProductionBOM(created_by=request.user)
        locked = False

        form = ProductionBOMForm(request.POST, instance=bom, locked=locked)

        input_formset = ProductionBOMInputLineFormSet(
            request.POST,
            instance=bom,
            prefix="input",
            locked=locked,
        )
        eff_formset = ProductionBOMEffluentLineFormSet(
            request.POST,
            instance=bom,
            prefix="eff",
            locked=locked,
        )

        if form.is_valid() and input_formset.is_valid() and eff_formset.is_valid():
            action = (request.POST.get("action") or "save").lower().strip()

            with transaction.atomic():
                bom = form.save(commit=False)
                if not bom.created_by_id:
                    bom.created_by = request.user
                bom.save()

                # ✅ Bind + SAFE SAVE (prevents blank rows)
                input_formset.instance = bom
                eff_formset.instance = bom
                _safe_save_bom_input_formset(input_formset, bom)
                _safe_save_bom_eff_formset(eff_formset, bom)

                if action == "submit":
                    mc_submit(bom, "BOM", request.user)
                    messages.success(request, "BOM submitted successfully.")
                else:
                    messages.success(request, "BOM saved successfully.")

            return redirect("accounts_budget:production_bom_edit", bom_id=bom.id)

        material_prefill = _material_prefill_payload_from_formset(input_formset)

        return render(request, "accounts/budget/production_bom_form.html", {
            "form": form,
            "bom": bom,
            "input_formset": input_formset,
            "eff_formset": eff_formset,
            "is_new": True,
            "is_locked": locked,
            "view_only": False,
            "material_prefill_json": mark_safe(json.dumps(material_prefill)),
        })

    # GET
    bom = ProductionBOM(created_by=request.user)
    locked = False

    form = ProductionBOMForm(instance=bom, locked=locked)
    input_formset = ProductionBOMInputLineFormSet(instance=bom, prefix="input", locked=locked)
    eff_formset = ProductionBOMEffluentLineFormSet(instance=bom, prefix="eff", locked=locked)

    return render(request, "accounts/budget/production_bom_form.html", {
        "form": form,
        "bom": bom,
        "input_formset": input_formset,
        "eff_formset": eff_formset,
        "is_new": True,
        "is_locked": locked,
        "view_only": False,
        "material_prefill_json": mark_safe("[]"),
    })


# -----------------------------------------------------------------------------
# Edit
# -----------------------------------------------------------------------------
@login_required
def production_bom_edit(request, bom_id: int):
    bom = get_object_or_404(ProductionBOM, pk=bom_id)

    mc = mc_get(bom, "BOM") if bom.pk else None
    locked = _is_locked(mc) if mc else False

    if locked:
        messages.info(request, "BOM is approved/locked. Opening in View mode.")
        return redirect("accounts_budget:production_bom_view", bom_id=bom.id)

    # ✅ Hide already-saved blank rows (even before next save)
    InputModel = ProductionBOMInputLineFormSet.model
    EffModel = ProductionBOMEffluentLineFormSet.model
    input_qs = InputModel.objects.filter(bom=bom).exclude(_blank_q_for_input_model(InputModel))
    eff_qs = EffModel.objects.filter(bom=bom).exclude(_blank_q_for_eff_model(EffModel))

    if request.method == "POST":
        form = ProductionBOMForm(request.POST, instance=bom, locked=locked)
        input_formset = ProductionBOMInputLineFormSet(
            request.POST, instance=bom, prefix="input", locked=locked, queryset=input_qs
        )
        eff_formset = ProductionBOMEffluentLineFormSet(
            request.POST, instance=bom, prefix="eff", locked=locked, queryset=eff_qs
        )

        if form.is_valid() and input_formset.is_valid() and eff_formset.is_valid():
            action = (request.POST.get("action") or "save").lower().strip()

            with transaction.atomic():
                form.save()

                # ✅ SAFE SAVE (prevents blanks/duplicates)
                _safe_save_bom_input_formset(input_formset, bom)
                _safe_save_bom_eff_formset(eff_formset, bom)

                if action == "submit":
                    mc_submit(bom, "BOM", request.user)
                    messages.success(request, "BOM submitted successfully.")
                else:
                    messages.success(request, "BOM saved successfully.")

            return redirect("accounts_budget:production_bom_edit", bom_id=bom.id)

    else:
        form = ProductionBOMForm(instance=bom, locked=locked)
        input_formset = ProductionBOMInputLineFormSet(instance=bom, prefix="input", locked=locked, queryset=input_qs)
        eff_formset = ProductionBOMEffluentLineFormSet(instance=bom, prefix="eff", locked=locked, queryset=eff_qs)

    material_prefill = _material_prefill_payload_from_formset(input_formset)

    return render(request, "accounts/budget/production_bom_form.html", {
        "form": form,
        "bom": bom,
        "input_formset": input_formset,
        "eff_formset": eff_formset,
        "is_new": False,
        "is_locked": locked,
        "view_only": False,
        "mc": mc,
        "material_prefill_json": mark_safe(json.dumps(material_prefill)),
    })


# -----------------------------------------------------------------------------
# View (always locked)
# -----------------------------------------------------------------------------
@login_required
def production_bom_view(request, bom_id: int):
    """
    View-only BOM page (always locked).
    ✅ Change: hide saved blank rows via queryset exclude
    """
    bom = get_object_or_404(ProductionBOM, pk=bom_id)

    mc = mc_get(bom, "BOM") if bom.pk else None
    locked = True

    InputModel = ProductionBOMInputLineFormSet.model
    EffModel = ProductionBOMEffluentLineFormSet.model
    input_qs = InputModel.objects.filter(bom=bom).exclude(_blank_q_for_input_model(InputModel))
    eff_qs = EffModel.objects.filter(bom=bom).exclude(_blank_q_for_eff_model(EffModel))

    form = ProductionBOMForm(instance=bom, locked=locked)
    input_formset = ProductionBOMInputLineFormSet(instance=bom, prefix="input", locked=locked, queryset=input_qs)
    eff_formset = ProductionBOMEffluentLineFormSet(instance=bom, prefix="eff", locked=locked, queryset=eff_qs)

    material_prefill = _material_prefill_payload_from_formset(input_formset)

    return render(request, "accounts/budget/production_bom_form.html", {
        "form": form,
        "bom": bom,
        "input_formset": input_formset,
        "eff_formset": eff_formset,
        "is_new": False,
        "is_locked": True,
        "view_only": True,
        "mc": mc,
        "material_prefill_json": mark_safe(json.dumps(material_prefill)),
    })


# -----------------------------------------------------------------------------
# JSON (unchanged)
# -----------------------------------------------------------------------------
@login_required
def production_bom_json(request, bom_id: int):
    bom = get_object_or_404(ProductionBOM, pk=bom_id)
    data = {
        "fg_name": bom.fg_name,
        "fg_alpha_name": bom.fg_alpha_name,
        "lines": [
            {"material_name": ln.material_name, "norm": str(ln.norm or Decimal("0"))}
            for ln in bom.lines.all().order_by("sr_no", "id")
        ]
    }
    return JsonResponse(data)


# -----------------------------------------------------------------------------
# Delete (unchanged)
# -----------------------------------------------------------------------------
@require_POST
@login_required
def production_bom_delete(request, pk: int):
    bom = get_object_or_404(ProductionBOM, pk=pk)

    scope = "BOM"
    try:
        mc = mc_get(bom, scope)
    except Exception:
        mc = None

    locked = False
    try:
        locked = _is_locked(mc) if mc else False
    except Exception:
        locked = False

    is_checker = False
    is_approver = False
    try:
        is_checker = _is_checker(request.user)
        is_approver = _is_approver(request.user)
    except Exception:
        pass

    if locked and not (is_checker or is_approver):
        messages.error(request, "This BOM is locked/submitted. It cannot be deleted.")
        return redirect("accounts_budget:production_bom_edit", bom_id=bom.id)

    try:
        with transaction.atomic():
            if mc:
                mc.delete()
            bom.delete()

        messages.success(request, "BOM deleted successfully.")
        return redirect("accounts_budget:production_bom_list")

    except Exception as e:
        messages.error(request, f"Unable to delete BOM: {e}")
        return redirect("accounts_budget:production_bom_edit", bom_id=bom.id)

# =============================================================================
# PRODUCTION BUDGET
# =============================================================================

ALLOWED_BOM_TYPES = ("Key Raw Material", "Raw Material", "Work in Progress", "Semi Finished Good", "WIP FR","Packing Material","Packing Materials",)


def _current_fy_label(d: date) -> str:
    if d.month >= 4:
        start = d.year
        end = d.year + 1
    else:
        start = d.year - 1
        end = d.year
    return f"{start}-{str(end)[-2:]}"


def _is_locked(mc) -> bool:
    return mc and (mc.status in ("SUBMITTED", "APPROVED"))


def _is_approver(user) -> bool:
    return user.has_perm("Budget.can_approve_budgets")


def _dec(v, default=Decimal("0.000000")):
    if v in (None, ""):
        return default
    try:
        s = str(v).replace(",", "").strip()
        if s == "":
            return default
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return default


def _to_bool(v) -> bool:
    s = ("" if v is None else str(v)).strip().lower()
    return s in ("1", "true", "on", "yes", "y")


def _get_target_norm_post(request, code: str):
    raw_t = request.POST.get(f"target_norm__{code}", "").strip()
    if raw_t == "":
        return None
    return _dec(raw_t, default=None)


def _safe_upsert_target_norm(budget_fg: ProductionBudgetFG, code: str, name: str, unit: str, target_norm):
    """
    Saves target_norm into ProductionBudgetLine.
    Supports either:
      - fg_budget FK (preferred) OR
      - fallback does nothing if schema differs.
    """
    try:
        ProductionBudgetLine.objects.update_or_create(
            fg_budget=budget_fg,
            bom_item_code=code,
            defaults={
                "material_name": name or "",
                "target_norm": target_norm,
                "remarks": "",
                "unit": unit or "",
            }
        )
        return True
    except Exception:
        return False


@login_required
def production_budget_home(request):
    plan = BudgetPlan.objects.order_by("-id").first()
    if not plan:
        plan = BudgetPlan.objects.create(
            fy=_current_fy_label(date.today()),
            company_group="",
            created_by=request.user,
            is_active=True,
        )

    budgets = list(ProductionBudgetFG.objects.filter(plan=plan).order_by("fg_name"))

    for b in budgets:
        b.mc = mc_get(b, "PROD")
        b.is_locked = _is_locked(b.mc)

        try:
            b.total_fg_qty = sum(
                (getattr(b, m) or Decimal("0.000000"))
                for m in ["apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec", "jan", "feb", "mar"]
            )
        except Exception:
            b.total_fg_qty = None

    return render(request, "accounts/budget/production_budget_home.html", {
        "plan": plan,
        "budgets": budgets,
    })


# -----------------------------------------------------------------------------
# CREATE: FG from BOM master (not ERP)  ✅ NO ProductionBudgetLine seeding
# -----------------------------------------------------------------------------
@login_required
def production_budget_create(request, plan_id=None):
    if plan_id:
        plan = get_object_or_404(BudgetPlan, id=plan_id)
    else:
        plan = BudgetPlan.objects.filter(is_active=True).order_by("-updated_at", "-id").first()
        if not plan:
            messages.error(request, "No active Budget Plan found.")
            return redirect("accounts_budget:budget_home")

    if request.method == "POST":
        form = ProductionBudgetCreateForm(request.POST)
        if form.is_valid():
            fg_name = (form.cleaned_data.get("fg_name") or "").strip()
            if not fg_name:
                form.add_error("fg_name", "Please select FG.")
                return render(request, "accounts/budget/production_budget_create.html", {"plan": plan, "form": form})

            bom = ProductionBOM.objects.filter(fg_name__iexact=fg_name, is_active=True).first()
            if not bom:
                form.add_error("fg_name", "No active Production BOM master found for this FG.")
                return render(request, "accounts/budget/production_budget_create.html", {"plan": plan, "form": form})

            if not ProductionBOMInputLine.objects.filter(bom=bom).exists():
                form.add_error("fg_name", "No BOM input lines found in BOM master for this FG.")
                return render(request, "accounts/budget/production_budget_create.html", {"plan": plan, "form": form})

            existing = ProductionBudgetFG.objects.filter(plan=plan, fg_name__iexact=fg_name).first()
            if existing:
                messages.info(request, f"Production budget already exists for: {fg_name}")
                return redirect("accounts_budget:production_budget_edit", budget_id=existing.id)

            with transaction.atomic():
                budget = ProductionBudgetFG.objects.create(
                    plan=plan,
                    fg_name=fg_name,
                    created_by=request.user,
                )

            messages.success(request, f"Production budget created for: {fg_name}")
            return redirect("accounts_budget:production_budget_edit", budget_id=budget.id)

        messages.error(request, "Please correct the errors below.")
    else:
        form = ProductionBudgetCreateForm()

    return render(request, "accounts/budget/production_budget_create.html", {"plan": plan, "form": form})


# -----------------------------------------------------------------------------
# EDIT: show BOM inputs from BOM master + save FG months (manual)
# -----------------------------------------------------------------------------
DEC6 = Decimal("0.000000")

# -----------------------------------------------------------
# Helpers
# -----------------------------------------------------------
MONTHS = ["apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec", "jan", "feb", "mar"]


# -------------------------------------------------------------------
# ✅ Helpers (schema-safe)
# -------------------------------------------------------------------
def _txt(v) -> str:
    return (v or "").strip()


def _concrete_fields(model_cls):
    return {f.name for f in model_cls._meta.get_fields() if getattr(f, "concrete", False)}


def _pick_first(existing: set[str], candidates: list[str]) -> str | None:
    for c in candidates:
        if c in existing:
            return c
    return None


def _parse_dec6(v):
    """
    - returns Decimal(6dp) or None if blank
    - returns "__INVALID__" if invalid numeric
    """
    s = (v or "").strip().replace(",", "")
    if s == "":
        return None
    try:
        return Decimal(s).quantize(DEC6, rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError, TypeError):
        return "__INVALID__"


def _decimal_default_for_field(model_cls, fname):
    """
    If field is DecimalField and NOT NULL => return 0.000000
    else => None
    """
    try:
        f = model_cls._meta.get_field(fname)
        if isinstance(f, dj_models.DecimalField) and not getattr(f, "null", False):
            return DEC6
    except Exception:
        pass
    return None


# -----------------------------------------------------------
# Waste calc helpers (kept + improved to 6dp)
# -----------------------------------------------------------
def _calc_eff_qty_fields(budget: "ProductionBudgetFG", norm):
    """
    Returns dict with:
      qty_apr..qty_mar, waste_qty_year, qty_year (alias)
    """
    try:
        n = (norm or DEC6)
        out = {}
        total = DEC6
        for m in MONTHS:
            fg = getattr(budget, m, None) or DEC6
            q = (fg * n).quantize(DEC6, rounding=ROUND_HALF_UP)
            out[f"qty_{m}"] = q
            total = (total + q).quantize(DEC6, rounding=ROUND_HALF_UP)

        out["waste_qty_year"] = total
        out["qty_year"] = total  # alias for older code

        prod_logger.debug(
            "EFF_CALC budget_id=%s norm=%s waste_qty_year=%s",
            getattr(budget, "id", None),
            str(n),
            str(total),
        )
        return out
    except Exception as e:
        prod_logger.exception(
            "EFF_CALC FAILED budget_id=%s norm=%r err=%s",
            getattr(budget, "id", None),
            norm,
            str(e),
        )
        raise


def _apply_eff_calculation_to_line(eln: "ProductionBudgetFGEffluentLine", budget: "ProductionBudgetFG"):
    """
    Schema-safe: fills qty_apr..qty_mar and waste_qty_year if those fields exist.
    (Model.save() also recalcs, but we keep this for schema-safety / logging.)
    """
    fields = _calc_eff_qty_fields(budget, getattr(eln, "waste_norm", None) or DEC6)

    for m in MONTHS:
        fn = f"qty_{m}"
        if hasattr(eln, fn):
            setattr(eln, fn, fields.get(fn))

    if hasattr(eln, "waste_qty_year"):
        eln.waste_qty_year = fields.get("waste_qty_year")
    elif hasattr(eln, "qty_year"):
        eln.qty_year = fields.get("qty_year")


def _ensure_budget_effluent_lines(budget: "ProductionBudgetFG", bom: "ProductionBOM"):
    """
    ✅ Upsert waste lines from BOM into budget every time:
       - creates missing lines
       - keeps existing lines (does NOT wipe manual edits)
       - does NOT require ignore_conflicts (SQL Server safe)
    """
    try:
        src = ProductionBOMEffluentLine.objects.filter(bom=bom).order_by("sr_no", "id")

        existing_qs = ProductionBudgetFGEffluentLine.objects.filter(budget=budget).values(
            "sr_no", "waste_type", "waste_name"
        )

        existing_keys = set()
        for r in existing_qs:
            sr = int(r.get("sr_no") or 0)
            wt = (r.get("waste_type") or "").strip().upper()
            wn = (r.get("waste_name") or "").strip().upper()
            existing_keys.add((sr, wt, wn))

        created = 0
        skipped_blank = 0
        skipped_dup_src = 0

        seen_src = set()

        with transaction.atomic():
            for ln in src:
                wt = (getattr(ln, "waste_type", "") or "").strip()
                wn = (getattr(ln, "waste_name", "") or "").strip()
                sr = int(getattr(ln, "sr_no", 0) or 0)

                if not wt and not wn:
                    skipped_blank += 1
                    continue

                key_u = (sr, wt.upper(), wn.upper())
                if key_u in seen_src:
                    skipped_dup_src += 1
                    continue
                seen_src.add(key_u)

                if key_u in existing_keys:
                    continue

                wnorm = getattr(ln, "waste_budget_norm", None) or DEC6
                wtarget = getattr(ln, "waste_target_norm", None)

                ProductionBudgetFGEffluentLine.objects.create(
                    budget=budget,
                    sr_no=sr,
                    waste_type=wt,
                    waste_name=wn,
                    waste_norm=wnorm,
                    waste_target_norm=wtarget,
                )
                created += 1

        prod_logger.info(
            "EFF_UPSERT done budget_id=%s created=%s skipped_blank=%s skipped_dup_src=%s",
            getattr(budget, "id", None),
            created,
            skipped_blank,
            skipped_dup_src,
        )

    except Exception as e:
        prod_logger.exception(
            "EFF_UPSERT FAILED budget_id=%s bom_id=%s err=%s",
            getattr(budget, "id", None),
            getattr(bom, "id", None),
            str(e),
        )
        raise


# -------------------------------------------------------------------
# ✅ NEW: Robust POST parsing for Waste tab (index-based OR id-based)
#      + supports clearing values (blank)
# -------------------------------------------------------------------
def _post_has_key(request, key: str) -> bool:
    return key in request.POST


def _get_first_present(request, keys: list[str]) -> tuple[bool, str | None]:
    """
    Returns (present?, value). present=True even if value is blank.
    """
    for k in keys:
        if _post_has_key(request, k):
            return True, request.POST.get(k)
    return False, None


def _parse_effluent_updates(request, budget_eff_qs, req_id: str):
    """
    Builds:
      updates[eid] = (bn_present, bn_val, tn_present, tn_val)

    Supports BOTH naming styles:
      - index-based: eff_id__0=19, eff_norm__0, eff_target_norm__0
      - id-based:    waste_norm__19 / eff_norm__19, waste_target_norm__19 / eff_target_norm__19

    Clearing:
      - blank waste_norm => set to 0.000000
      - blank target_norm => set to NULL
    """
    eff_errors = []
    updates = {}

    eff_count_raw = (request.POST.get("eff_count") or "").strip()
    try:
        eff_count = int(eff_count_raw or 0)
    except Exception:
        eff_count = 0

    prod_logger.info("[%s] tab2 eff_count=%s", req_id, eff_count)

    # index-based (if provided)
    if eff_count > 0:
        for i in range(eff_count):
            eid_raw = request.POST.get(f"eff_id__{i}")
            if not eid_raw:
                continue
            try:
                eid = int(eid_raw)
            except Exception:
                prod_logger.warning("[%s] tab2 bad eff_id__%s=%r", req_id, i, eid_raw)
                continue

            bn_present, raw_bn = _get_first_present(
                request,
                [f"eff_norm__{i}", f"waste_norm__{i}", f"eff_norm__{eid}", f"waste_norm__{eid}"]
            )
            tn_present, raw_tn = _get_first_present(
                request,
                [f"eff_target_norm__{i}", f"waste_target_norm__{i}", f"eff_target_norm__{eid}", f"waste_target_norm__{eid}"]
            )

            bn_val = _parse_dec6(raw_bn) if bn_present else None
            tn_val = _parse_dec6(raw_tn) if tn_present else None

            if bn_val == "__INVALID__":
                eff_errors.append(f"Invalid Waste Norm (row {i+1})")
                prod_logger.warning("[%s] INVALID eff_norm row=%s raw=%r", req_id, i, raw_bn)

            if tn_val == "__INVALID__":
                eff_errors.append(f"Invalid Waste Target Norm (row {i+1})")
                prod_logger.warning("[%s] INVALID eff_target row=%s raw=%r", req_id, i, raw_tn)

            updates[eid] = (bn_present, bn_val, tn_present, tn_val)

    # id-based (always supported, can override)
    for eln in budget_eff_qs:
        eid = eln.id

        bn_present, raw_bn = _get_first_present(
            request, [f"waste_norm__{eid}", f"eff_norm__{eid}"]
        )
        tn_present, raw_tn = _get_first_present(
            request, [f"waste_target_norm__{eid}", f"eff_target_norm__{eid}"]
        )

        if not bn_present and not tn_present:
            continue

        bn_val = _parse_dec6(raw_bn) if bn_present else None
        tn_val = _parse_dec6(raw_tn) if tn_present else None

        if bn_val == "__INVALID__":
            eff_errors.append(f"Invalid Waste Norm for {eln.waste_type}/{eln.waste_name}")
            prod_logger.warning("[%s] INVALID eff_norm id=%s raw=%r", req_id, eid, raw_bn)

        if tn_val == "__INVALID__":
            eff_errors.append(f"Invalid Waste Target Norm for {eln.waste_type}/{eln.waste_name}")
            prod_logger.warning("[%s] INVALID eff_target id=%s raw=%r", req_id, eid, raw_tn)

        updates[eid] = (bn_present, bn_val, tn_present, tn_val)

    return updates, eff_errors


@login_required
def production_budget_edit(request, budget_id: int):
    req_id = uuid.uuid4().hex[:8]

    prod_logger.info(
        "[%s] PROD_EDIT enter method=%s budget_id=%s user=%s",
        req_id, request.method, budget_id, getattr(request.user, "username", None)
    )

    budget = get_object_or_404(ProductionBudgetFG, pk=budget_id)
    plan = budget.plan
    fg = _txt(budget.fg_name)

    dup_ids = list(
        ProductionBudgetFG.objects
        .filter(plan=plan, fg_name__iexact=fg)
        .order_by("-id")
        .values_list("id", flat=True)[:10]
    )
    if len(dup_ids) > 1:
        prod_logger.warning("[%s] DUP_BUDGET plan_id=%s fg=%s ids=%s", req_id, plan.id, fg, dup_ids)

    scope = "PROD"
    mc = mc_get(budget, scope)
    locked = _is_locked(mc)

    prod_logger.info(
        "[%s] context plan_id=%s fy=%s fg=%s locked=%s mc_status=%s",
        req_id,
        getattr(plan, "id", None),
        getattr(plan, "fy", None),
        fg,
        locked,
        getattr(mc, "status", None) if mc else None,
    )

    if request.method == "POST" and locked:
        prod_logger.warning("[%s] POST blocked due to locked budget_id=%s", req_id, budget.id)
        messages.warning(request, "This entry is locked. It can be edited only if disapproved by approver.")
        return redirect("accounts_budget:production_budget_edit", budget_id=budget.id)

    bom = ProductionBOM.objects.filter(fg_name__iexact=fg, is_active=True).first()
    if not bom:
        prod_logger.error("[%s] No active BOM found for fg=%s (budget_id=%s)", req_id, fg, budget.id)
        messages.error(request, "No active Production BOM master found for this FG.")
        return redirect("accounts_budget:production_budget_home")

    bom_inputs = ProductionBOMInputLine.objects.filter(bom=bom).order_by("id")
    prod_logger.info("[%s] bom_id=%s bom_inputs=%s", req_id, bom.id, bom_inputs.count())

    _ensure_budget_effluent_lines(budget, bom)
    budget_eff = ProductionBudgetFGEffluentLine.objects.filter(budget=budget).order_by("sr_no", "id")
    prod_logger.info("[%s] budget_effluent=%s", req_id, budget_eff.count())

    input_fields = _concrete_fields(ProductionBOMInputLine)

    code_f = _pick_first(input_fields, ["material_code", "bom_item_code", "item_code", "code"])
    name_f = _pick_first(input_fields, ["material_name", "bom_item_name", "name"])
    unit_f = _pick_first(input_fields, ["unit", "uom"])
    cat_f  = _pick_first(input_fields, ["material_category", "category", "type"])

    bn_f   = _pick_first(input_fields, ["budget_norm", "norm"])
    tn_f   = _pick_first(input_fields, ["target_norm", "mat_qty_mt"])
    cap_f  = _pick_first(input_fields, ["is_captive"])

    prod_logger.info(
        "[%s] fieldmap code=%s name=%s unit=%s cat=%s bn=%s tn=%s cap=%s",
        req_id, code_f, name_f, unit_f, cat_f, bn_f, tn_f, cap_f
    )

    def _fg_qty_from_instance(obj: ProductionBudgetFG):
        return {m: (getattr(obj, m) or DEC6) for m in MONTHS}

    def _sum_qty_year(fg_qty_dict, norm_dec):
        n = (norm_dec or DEC6)
        total = DEC6
        for m in MONTHS:
            total += (fg_qty_dict[m] * n)
        return total

    fg_qty = _fg_qty_from_instance(budget)
    fg_total = sum((fg_qty[m] for m in MONTHS), DEC6)

    input_rows = []
    seen_codes = set()
    dup_codes = 0
    blank_codes = 0

    for idx, ln in enumerate(bom_inputs, start=1):
        code = _txt(getattr(ln, code_f, "")) if code_f else ""
        name = _txt(getattr(ln, name_f, "")) if name_f else ""
        unit = _txt(getattr(ln, unit_f, "")) if unit_f else ""
        cat  = _txt(getattr(ln, cat_f, "")) if cat_f else ""

        if not code:
            blank_codes += 1
        else:
            if code in seen_codes:
                dup_codes += 1
            seen_codes.add(code)

        budget_norm = getattr(ln, bn_f, None) if bn_f else None
        target_norm = getattr(ln, tn_f, None) if tn_f else None
        is_captive  = bool(getattr(ln, cap_f, False)) if cap_f else False

        norm = (budget_norm if budget_norm is not None else DEC6) or DEC6
        bom_qty_year = _sum_qty_year(fg_qty, norm)

        input_rows.append({
            "sr_no": idx,
            "code": code,
            "name": name,
            "bom_item_code": code,
            "bom_item_name": name,
            "unit": unit,
            "category": cat,
            "budget_norm": budget_norm,
            "target_norm": target_norm,
            "is_captive": is_captive,
            "norm": norm,
            "bom_qty_year": bom_qty_year,
        })

    if blank_codes or dup_codes:
        prod_logger.warning(
            "[%s] input_codes issues blank=%s duplicate=%s (POST keying can break)",
            req_id, blank_codes, dup_codes
        )

    # -------------------------------------------------------------------
    # POST: Save header + norms + effluent (robust parsing + supports clearing)
    # -------------------------------------------------------------------
    if request.method == "POST":
        prod_logger.info(
            "[%s] POST keys=%s files=%s",
            req_id,
            list(request.POST.keys())[:120],
            list(request.FILES.keys()),
        )

        action = (request.POST.get("action") or "").strip().lower()
        prod_logger.info(
            "[%s] POST action=%r submit_btn_present=%s save_btn_present=%s",
            req_id,
            action,
            ("submit" in request.POST) or ("action" in request.POST and request.POST.get("action") == "submit"),
            ("save" in request.POST) or ("action" in request.POST and request.POST.get("action") == "save"),
        )
        prod_logger.info("[%s] action=%s", req_id, action)

        prod_logger.info("[%s] FG_POST raw months=%s", req_id, {m: request.POST.get(m) for m in MONTHS})

        head_form = ProductionBudgetFGForm(request.POST, instance=budget)

        # ---------- Tab-1 validations + pending updates
        norm_errors = []
        pending_updates = []

        for ln in bom_inputs:
            code = _txt(getattr(ln, code_f, "")) if code_f else ""
            if not code:
                continue

            raw_bn = request.POST.get(f"norm__{code}")
            raw_tn = request.POST.get(f"target_norm__{code}")

            bn_val = _parse_dec6(raw_bn)
            tn_val = _parse_dec6(raw_tn)

            if bn_val == "__INVALID__":
                norm_errors.append(f"Invalid Budget Norm for {code}")
                prod_logger.warning("[%s] INVALID norm code=%s raw=%r", req_id, code, raw_bn)

            if tn_val == "__INVALID__":
                norm_errors.append(f"Invalid Target Norm for {code}")
                prod_logger.warning("[%s] INVALID target code=%s raw=%r", req_id, code, raw_tn)

            cap_vals = request.POST.getlist(f"is_captive__{code}")
            cap_raw = (cap_vals[-1] if cap_vals else "0")
            cap_val = (str(cap_raw).strip() == "1")

            pending_updates.append((ln, bn_val, tn_val, cap_val))

        prod_logger.info(
            "[%s] tab1 parsed pending_updates=%s norm_errors=%s",
            req_id, len(pending_updates), len(norm_errors)
        )

        # ---------- Tab-2 parsing
        eff_updates_by_id, eff_errors = _parse_effluent_updates(request, budget_eff, req_id)

        prod_logger.info(
            "[%s] tab2 parsed eff_updates=%s eff_errors=%s",
            req_id, len(eff_updates_by_id), len(eff_errors)
        )

        try:
            head_valid = head_form.is_valid()
            if not head_valid:
                prod_logger.warning("[%s] head_form invalid errors=%s", req_id, head_form.errors)

            if head_valid and not norm_errors and not eff_errors:
                with transaction.atomic():
                    prod_logger.info("[%s] TX start saving budget_id=%s", req_id, budget.id)

                    budget = ProductionBudgetFG.objects.select_for_update().get(pk=budget.id)

                    # ✅ Save header first
                    head_form.instance = budget
                    saved_budget = head_form.save()
                    saved_budget.refresh_from_db()

                    prod_logger.info(
                        "[%s] HEAD saved budget_id=%s total_fg=%s months=%s",
                        req_id,
                        saved_budget.id,
                        getattr(saved_budget, "total_fg_qty", None),
                        {m: getattr(saved_budget, m) for m in MONTHS},
                    )

                    fg_qty = _fg_qty_from_instance(saved_budget)
                    fg_total = sum((fg_qty[m] for m in MONTHS), DEC6)

                    # ✅ Update BOM input-line fields
                    updated_lines = 0
                    for (ln, bn_val, tn_val, cap_val) in pending_updates:
                        update_fields = []

                        if bn_f:
                            if bn_val is None:
                                bn_val = _decimal_default_for_field(ProductionBOMInputLine, bn_f)
                            setattr(ln, bn_f, bn_val)
                            update_fields.append(bn_f)

                        if tn_f:
                            if tn_val is None:
                                tn_val = _decimal_default_for_field(ProductionBOMInputLine, tn_f)
                            setattr(ln, tn_f, tn_val)
                            update_fields.append(tn_f)

                        if cap_f:
                            setattr(ln, cap_f, bool(cap_val))
                            update_fields.append(cap_f)

                        if update_fields:
                            ln.save(update_fields=update_fields)
                            updated_lines += 1

                    prod_logger.info("[%s] Updated BOM input lines=%s", req_id, updated_lines)

                    # ✅ Update Waste/Effluent lines
                    eff_saved = 0
                    eff_qs = (
                        ProductionBudgetFGEffluentLine.objects
                        .select_for_update()
                        .filter(budget=saved_budget)
                        .order_by("sr_no", "id")
                    )

                    for eln in eff_qs:
                        upd = eff_updates_by_id.get(eln.id)
                        if upd:
                            bn_present, bn_val, tn_present, tn_val = upd
                            if bn_val == "__INVALID__" or tn_val == "__INVALID__":
                                continue

                            # ✅ blank => clear to 0
                            if bn_present:
                                eln.waste_norm = (bn_val if bn_val is not None else DEC6)

                            # ✅ blank => clear to NULL
                            if tn_present:
                                eln.waste_target_norm = (tn_val if tn_val is not None else None)

                        _apply_eff_calculation_to_line(eln, saved_budget)
                        eln.save()
                        eff_saved += 1

                        prod_logger.debug(
                            "[%s] EFF saved id=%s norm=%s target=%s waste_qty_year=%s",
                            req_id,
                            eln.id,
                            eln.waste_norm,
                            eln.waste_target_norm,
                            getattr(eln, "waste_qty_year", None),
                        )

                    prod_logger.info("[%s] Updated effluent lines=%s", req_id, eff_saved)

                    # ✅ Always submit after a successful save
                    try:
                        mc_before = mc_get(saved_budget, scope)
                        prod_logger.info(
                            "[%s] MC_BEFORE submit budget_id=%s scope=%s mc_id=%s status=%s",
                            req_id, saved_budget.id, scope,
                            getattr(mc_before, "id", None),
                            getattr(mc_before, "status", None),
                        )

                        if not mc_before or (str(getattr(mc_before, "status", "")).upper() != "APPROVED"):
                            st = mc_submit(saved_budget, scope, request.user)
                            prod_logger.info(
                                "[%s] MC_AFTER submit budget_id=%s scope=%s mc_id=%s status=%s",
                                req_id, saved_budget.id, scope,
                                getattr(st, "id", None),
                                getattr(st, "status", None),
                            )
                        else:
                            prod_logger.info("[%s] MC_SKIP submit (already APPROVED) budget_id=%s", req_id, saved_budget.id)

                    except Exception as e:
                        prod_logger.exception("[%s] MC_SUBMIT failed budget_id=%s err=%s", req_id, saved_budget.id, str(e))

                # ✅ SUCCESS MESSAGE + ✅ REDIRECT TO HOME (THIS IS THE REQUIRED CHANGE)
                if action == "submit":
                    messages.success(request, "Production Budget submitted for approval.")
                else:
                    messages.success(request, "Production Budget saved and submitted for approval.")

                prod_logger.info(
                    "[%s] SAVE_OK redirecting to production_budget_home budget_id=%s action=%s",
                    req_id, budget.id, action
                )
                return redirect("accounts_budget:production_budget_home")

            # ---------- validation errors
            if norm_errors or eff_errors:
                messages.error(request, "Please correct Norm/Target values.")
                head_form.add_error(None, "; ".join((norm_errors + eff_errors)[:3]))
                prod_logger.warning("[%s] Validation errors norm=%s eff=%s", req_id, norm_errors[:5], eff_errors[:5])
            else:
                messages.error(request, "Please correct the FG monthly plan errors.")
                prod_logger.warning("[%s] Head form errors: %s", req_id, head_form.errors)

        except IntegrityError as e:
            prod_logger.exception("[%s] IntegrityError saving Production Budget: %s", req_id, str(e))
            messages.error(request, "Save failed due to duplicate/constraint error. Please check logs.")
        except DatabaseError as e:
            prod_logger.exception("[%s] DatabaseError saving Production Budget: %s", req_id, str(e))
            messages.error(request, "Save failed due to database error. Please check logs.")
        except Exception as e:
            prod_logger.exception("[%s] Unexpected error saving Production Budget: %s", req_id, str(e))
            messages.error(request, "Save failed due to unexpected error. Please check logs.")

    else:
        head_form = ProductionBudgetFGForm(instance=budget)

    if locked:
        for name in head_form.fields:
            head_form.fields[name].disabled = True

    # -------------------------------------------------------------------
    # Computed RM rows
    # -------------------------------------------------------------------
    month_totals = {m: DEC6 for m in MONTHS}
    mat_rows = []

    for r in input_rows:
        per_month = {}
        total = DEC6
        for m in MONTHS:
            qty = (fg_qty[m] * (r["norm"] or DEC6))
            per_month[m] = qty
            total += qty
            month_totals[m] += qty

        mat_rows.append({
            "material": r["name"],
            "code": r["code"],
            "unit": r["unit"],
            "category": r["category"],
            "budget_norm": r["budget_norm"],
            "target_norm": r["target_norm"],
            "is_captive": r.get("is_captive", False),
            "per_month": per_month,
            "total": total,
        })

    grand_total = sum(month_totals.values(), DEC6)

    # -------------------------------------------------------------------
    # Waste & Effluent rows (budget-side)
    # -------------------------------------------------------------------
    eff_rows = []
    eff_month_totals = {m: DEC6 for m in MONTHS}

    for ln in budget_eff:
        wt = _txt(getattr(ln, "waste_type", ""))
        wn = _txt(getattr(ln, "waste_name", ""))

        wnorm = getattr(ln, "waste_norm", None) or DEC6
        wtarget = getattr(ln, "waste_target_norm", None)

        per_month = {}
        total = DEC6
        for m in MONTHS:
            qty = (fg_qty[m] * (wnorm or DEC6))
            per_month[m] = qty
            total += qty
            eff_month_totals[m] += qty

        qty_year_db = getattr(ln, "waste_qty_year", None)
        qty_year = (qty_year_db if qty_year_db is not None else total)

        eff_rows.append({
            "id": ln.id,
            "sr_no": getattr(ln, "sr_no", 0) or 0,
            "waste_type": wt,
            "waste_name": wn,
            "waste_norm": wnorm,
            "waste_budget_norm": wnorm,
            "waste_target_norm": wtarget,
            "target_norm": wtarget,
            "per_month": per_month,
            "total": qty_year,
        })

    eff_grand_total = sum(eff_month_totals.values(), DEC6)

    return render(request, "accounts/budget/production_budget_edit.html", {
        "plan": plan,
        "budget": budget,
        "fg_name": fg,
        "bom": bom,
        "head_form": head_form,

        "months": MONTHS,
        "fg_qty": fg_qty,
        "fg_total": fg_total,

        "input_rows": input_rows,
        "mat_rows": mat_rows,
        "month_totals": month_totals,
        "grand_total": grand_total,

        "eff_rows": eff_rows,
        "eff_month_totals": eff_month_totals,
        "eff_grand_total": eff_grand_total,

        "mc": mc,
        "is_locked": locked,
        "mc_scope": scope,
        "mc_model": "accounts_budget.ProductionBudgetFG",
        "mc_pk": budget.id,
        "req_id": req_id,
        "dup_ids": dup_ids,
    })

@login_required
def production_budget_delete(request, budget_id: int):
    budget = get_object_or_404(ProductionBudgetFG, pk=budget_id)
    fg_name = (budget.fg_name or "").strip()

    # ✅ Maker-checker lock protection (same as edit)
    scope = "PROD"
    mc = mc_get(budget, scope)  # your existing helper
    locked = _is_locked(mc)     # your existing helper

    next_url = request.GET.get("next") or request.POST.get("next") or None
    if not next_url:
        next_url = "/accounts/budgets/production/"  # fallback; replace with your home url if needed

    # Do not allow delete when locked
    if locked:
        messages.warning(request, "This entry is locked (submitted/approved). Disapprove first to unlock, then delete.")
        return redirect("accounts_budget:production_budget_edit", budget_id=budget.id)

    if request.method == "POST":
        try:
            with transaction.atomic():
                # --- Optional: delete related norms for this FG ---
                ProductionNorm.objects.filter(fg_name=fg_name).delete()

                # --- Optional: delete related target norms if you have a separate model ---
                # ProductionTargetNorm.objects.filter(fg_name=fg_name).delete()

                # --- Finally delete the budget header ---
                budget.delete()

            messages.success(request, f"Production Budget deleted for: {fg_name}")
            return redirect(next_url)
        except Exception as e:
            messages.error(request, f"Delete failed: {e}")
            return redirect("accounts_budget:production_budget_edit", budget_id=budget.id)

    # GET confirm page
    return render(request, "accounts/budget/production_budget_delete.html", {
        "budget": budget,
        "fg_name": fg_name,
        "next": next_url,
        "mc": mc,
        "is_locked": locked,
    })

# =============================================================================
# SALES BUDGET
# =============================================================================

def fmt_inr(value):
    """Format a number in Indian comma style: 1,23,45,678.00"""
    if not value:
        return "0.00"
    try:
        value = Decimal(str(value))
        str_val = f"{value:.2f}"
        integer_part, decimal_part = str_val.split(".")
        negative = integer_part.startswith("-")
        if negative:
            integer_part = integer_part[1:]
        if len(integer_part) <= 3:
            formatted = integer_part
        else:
            last3 = integer_part[-3:]
            rest = integer_part[:-3]
            groups = []
            while len(rest) > 2:
                groups.append(rest[-2:])
                rest = rest[:-2]
            if rest:
                groups.append(rest)
            groups.reverse()
            formatted = ",".join(groups) + "," + last3
        return ("-" if negative else "") + formatted + "." + decimal_part
    except Exception:
        return str(value)


def fmt_usd(value):
    """Format USD with Western comma style: 1,234,567.00"""
    if not value:
        return "0.00"
    try:
        value = Decimal(str(value))
        str_val = f"{value:.2f}"
        integer_part, decimal_part = str_val.split(".")
        negative = integer_part.startswith("-")
        if negative:
            integer_part = integer_part[1:]
        groups = []
        while len(integer_part) > 3:
            groups.append(integer_part[-3:])
            integer_part = integer_part[:-3]
        if integer_part:
            groups.append(integer_part)
        groups.reverse()
        formatted = ",".join(groups)
        return ("-" if negative else "") + formatted + "." + decimal_part
    except Exception:
        return str(value)


@login_required
def sales_budget_home(request):
    plan = (
        BudgetPlan.objects.filter(is_active=True)
        .order_by("-updated_at", "-id")
        .first()
    )
    if not plan:
        messages.error(request, "No active Budget Plan found.")
        return redirect("accounts_budget:budget_home")

    sales_budget, _ = SalesBudget.objects.get_or_create(
        plan=plan,
        defaults={"created_by": request.user}
    )

    mc = mc_get(sales_budget, "SALES")
    locked = _is_locked(mc)

    # ── Exchange rate from budget header
    inr_usd = Decimal(sales_budget.inr_usd or 0)

    lines = sales_budget.lines.all().order_by("sale_type", "product_name")
    months = ["apr","may","jun","jul","aug","sep","oct","nov","dec","jan","feb","mar"]

    month_totals = {m: (lines.aggregate(v=Sum(m))["v"] or Decimal("0.000")) for m in months}
    total_qty = lines.aggregate(v=Sum("annual_qty_mt"))["v"] or Decimal("0.000")

    # ── Domestic: native INR (stored in DB)
    domestic_amt_inr = (
        lines.filter(sale_type=SaleType.DOMESTIC)
             .aggregate(v=Sum("amt_inr"))["v"] or Decimal("0.00")
    )

    # ── Export: compute USD in Python (rate_usd is per KG, qty stored in MT)
    export_amt_usd = Decimal("0.00")
    lines_list = list(lines)  # single DB hit, reused below

    for r in lines_list:
        if r.sale_type == SaleType.EXPORT:
            rate   = Decimal(r.rate_usd or 0)       # USD per KG
            qty_mt = Decimal(r.annual_qty_mt or 0)  # MT
            qty_kg = qty_mt * 1000                  # convert MT → KG
            r.amt_usd = (rate * qty_kg).quantize(Decimal("0.01"))
            export_amt_usd += r.amt_usd

            # Convert USD → INR for Cr display using header rate
            if inr_usd > 0:
                r.amt_inr_equiv  = (r.amt_usd * inr_usd).quantize(Decimal("0.01"))
                r.amt_inr_cr_fmt = fmt_inr(
                    (r.amt_inr_equiv / Decimal("10000000")).quantize(Decimal("0.01"))
                )
            else:
                r.amt_inr_equiv  = Decimal("0.00")
                r.amt_inr_cr_fmt = "0.00"

            r.amt_inr_fmt = "—"         # no native INR for export
            r.amt_usd_fmt = fmt_usd(r.amt_usd)  # formatted USD per line

        else:
            r.amt_usd        = Decimal("0.00")
            r.amt_usd_fmt    = "0.00"
            r.amt_inr_equiv  = Decimal("0.00")
            r.amt_inr_fmt    = fmt_inr(r.amt_inr)
            r.amt_inr_cr_fmt = fmt_inr(r.amt_inr_cr)

        # MT → KG for display (kept for weighted avg calc below)
        r.annual_qty_kg = (Decimal(r.annual_qty_mt or 0) * 1000).quantize(Decimal("0.000"))

    # ── Convert export USD → INR equiv and compute combined totals
    export_amt_inr_equiv = (export_amt_usd * inr_usd).quantize(Decimal("0.01")) if inr_usd > 0 else Decimal("0.00")

    total_amt_inr   = domestic_amt_inr + export_amt_inr_equiv
    total_amt_usd   = export_amt_usd.quantize(Decimal("0.01"))
    total_amt_cr    = (total_amt_inr / Decimal("10000000")).quantize(Decimal("0.01"))
    domestic_amt_cr = (domestic_amt_inr / Decimal("10000000")).quantize(Decimal("0.01"))
    export_amt_cr   = (export_amt_inr_equiv / Decimal("10000000")).quantize(Decimal("0.01"))

    # ── Monthwise in MT (no decimals)
    month_rows = [
        {
            "label": m,
            "total": (month_totals.get(m) or Decimal("0.000")).quantize(Decimal("1"))
        }
        for m in months
    ]

    # ── Weighted average rates for footer
    export_qty_kg_total = sum(
        Decimal(r.annual_qty_mt or 0) * 1000
        for r in lines_list if r.sale_type == SaleType.EXPORT
    )
    domestic_qty_kg_total = sum(
        Decimal(r.annual_qty_mt or 0) * 1000
        for r in lines_list if r.sale_type == SaleType.DOMESTIC
    )

    export_rate_usd_wavg = (
        sum(
            Decimal(r.rate_usd or 0) * Decimal(r.annual_qty_mt or 0) * 1000
            for r in lines_list if r.sale_type == SaleType.EXPORT
        ) / export_qty_kg_total
    ).quantize(Decimal("0.000")) if export_qty_kg_total else Decimal("0.000")

    domestic_rate_inr_wavg = (
        sum(
            Decimal(r.rate_inr or 0) * Decimal(r.annual_qty_mt or 0) * 1000
            for r in lines_list if r.sale_type == SaleType.DOMESTIC
        ) / domestic_qty_kg_total
    ).quantize(Decimal("0.000")) if domestic_qty_kg_total else Decimal("0.000")

    return render(request, "accounts/budget/sales_budget_home.html", {
        "plan": plan,
        "sales_budget": sales_budget,
        "lines": lines_list,
        "months": months,
        "month_totals": month_totals,
        "month_rows": month_rows,

        "total_qty": total_qty,

        # ── INR totals (domestic + export converted): pre-formatted
        "total_amt_inr":    fmt_inr(total_amt_inr),
        "total_amt_cr":     fmt_inr(total_amt_cr),
        "domestic_amt_inr": fmt_inr(domestic_amt_inr),
        "domestic_amt_cr":  fmt_inr(domestic_amt_cr),
        "export_amt_cr":    fmt_inr(export_amt_cr),

        # ── USD totals: pre-formatted with Western commas
        "total_amt_usd":  fmt_usd(total_amt_usd),
        "export_amt_usd": fmt_usd(export_amt_usd),

        # ── Weighted average rates for footer
        "export_rate_usd_wavg":   export_rate_usd_wavg,
        "domestic_rate_inr_wavg": domestic_rate_inr_wavg,

        "mc": mc,
        "is_locked": locked,
        "is_approver": _is_approver(request.user),
        "mc_scope": "SALES",
        "mc_model": "accounts_budget.SalesBudget",
        "mc_pk": sales_budget.id,
    })

@login_required
def sales_budget_edit(request):
    plan = (
        BudgetPlan.objects.filter(is_active=True)
        .order_by("-updated_at", "-id")
        .first()
    )
    if not plan:
        messages.error(request, "No active Budget Plan found.")
        return redirect("accounts_budget:budget_home")

    sales_budget, _ = SalesBudget.objects.get_or_create(
        plan=plan,
        defaults={"created_by": request.user}
    )

    scope = "SALES"
    mc = mc_get(sales_budget, scope)
    locked = _is_locked(mc)

    months = ["apr","may","jun","jul","aug","sep","oct","nov","dec","jan","feb","mar"]

    if request.method == "POST" and locked:
        messages.warning(request, "Sales Budget is locked. It can be edited only if disapproved by approver.")
        return redirect("accounts_budget:sales_budget_edit")

    qs = SalesBudgetLine.objects.filter(budget=sales_budget).order_by("sale_type", "product_name")

    # ✅ FG choices from ProductionNorm (+ include existing saved values so old rows validate)
    existing_products = set(qs.values_list("product_name", flat=True))
    fg_choices = sales_fg_choices_from_production_norm(extra_names=existing_products)

    if request.method == "POST":
        form = SalesBudgetForm(request.POST, instance=sales_budget)
        formset = SalesBudgetLineFormSet(
            request.POST,
            queryset=qs,
            form_kwargs={"budget": sales_budget, "fg_choices": fg_choices},  # ✅ inject choices
        )

        if form.is_valid() and formset.is_valid():
            # Duplicate check: (budget, product_name, sale_type)
            seen = set()
            duplicate_found = False
            for f in formset.forms:
                if not getattr(f, "cleaned_data", None):
                    continue
                if f.cleaned_data.get("DELETE"):
                    continue

                pn = (f.cleaned_data.get("product_name") or "").strip().upper()
                st = f.cleaned_data.get("sale_type")
                if not pn or not st:
                    continue

                key = (pn, str(st))
                if key in seen:
                    f.add_error("product_name", "Duplicate Product + Type is not allowed in the same Sales Budget.")
                    f.add_error("sale_type", "Duplicate Product + Type is not allowed in the same Sales Budget.")
                    duplicate_found = True
                else:
                    seen.add(key)

            if duplicate_found:
                messages.error(request, "Unable to save. Remove duplicate Product + Type rows.")
            else:
                try:
                    with transaction.atomic():
                        form.save()
                        sales_budget.refresh_from_db()

                        ex = sales_budget.inr_usd or Decimal("0")

                        objs = formset.save(commit=False)
                        for o in objs:
                            o.budget = sales_budget

                            total = Decimal("0")
                            for m in months:
                                total += (getattr(o, m) or Decimal("0"))
                            o.annual_qty_mt = total

                            if ex and ex > 0:
                                if o.sale_type == SaleType.EXPORT:
                                    if o.rate_usd:
                                        o.rate_inr = (Decimal(o.rate_usd) * Decimal(ex))
                                    else:
                                        o.rate_inr = None
                                else:
                                    if o.rate_inr:
                                        o.rate_usd = (Decimal(o.rate_inr) / Decimal(ex))
                                    else:
                                        o.rate_usd = None

                            o.save()

                        for o in formset.deleted_objects:
                            o.delete()

                        if not mc or (mc.status or "").upper() != "APPROVED":
                            mc_submit(sales_budget, scope, request.user)

                    if request.POST.get("action") == "submit":
                        messages.success(request, "Sales Budget submitted for approval.")
                    else:
                        messages.success(request, "Sales Budget saved.")

                    return redirect("accounts_budget:sales_budget_home")

                except IntegrityError:
                    messages.error(
                        request,
                        "Unable to save due to duplicate Product + Type. Remove duplicates and try again."
                    )
        else:
            messages.error(request, "Unable to save. Please correct the errors.")

        print("SalesBudgetForm errors:", form.errors)
        print("Formset non_form_errors:", formset.non_form_errors())
        print("Formset errors:", formset.errors)

    else:
        form = SalesBudgetForm(instance=sales_budget)
        formset = SalesBudgetLineFormSet(
            queryset=qs,
            form_kwargs={"budget": sales_budget, "fg_choices": fg_choices},  # ✅ inject choices
        )

    can_approve = _is_approver(request.user) and mc and mc.status == "SUBMITTED"

    return render(request, "accounts/budget/sales_budget_edit.html", {
        "plan": plan,
        "sales_budget": sales_budget,
        "form": form,
        "formset": formset,
        "mc": mc,
        "is_locked": locked,
        "can_approve": can_approve,
    })

# =============================================================================
# RMC BUDGET (CLEAN + REQUIRED QTY WORKING + CAPTIVE SPLIT)
# =============================================================================
# maker-checker hooks (keep your existing implementations) ----
# mc_get, _is_locked, _is_approver, mc_submit must exist in your codebase
# ------------------------------------------------------------------

MONTHS = ["apr","may","jun","jul","aug","sep","oct","nov","dec","jan","feb","mar"]

# BOM types allowed to sync into RMC (your list)
RMC_INPUT_TYPES = (
    "Key Raw Material",
    "Raw Material",
    "Work in Progress",
    "Semi Finished Good",
    "WIP FR",
)


FG_POST_KEYS = ["fg_names", "selected_fgs"]

_RM_CODE_TAIL_RE = re.compile(r"\(([^()]+)\)\s*$")

def _dedupe_keep_order(items):
    out, seen = [], set()
    for x in items:
        u = (x or "").strip()
        if not u:
            continue
        key = u.upper()
        if key in seen:
            continue
        seen.add(key)
        out.append(u)
    return out

def _parse_fg_payload(raw):
    """
    Accepts:
      - ['FG1', 'FG2', ...]       -> keep as-is
      - "FG1"                     -> single
      - "FG1, FG2, FG3"           -> comma-delimited string (with separator commas)
      - '["FG1","FG2"]'           -> JSON string

    IMPORTANT:
      Do NOT split valid FG names that contain commas, e.g.
      "1,2 DIMETHYL PROPYL AMINE"
      "2,4 DICHLORO BENZYL CHLORIDE"
    """
    if raw is None:
        return []

    def _looks_like_delimited_multi(s: str) -> bool:
        """
        Split only when comma is clearly acting as a separator between values,
        not when it is part of a chemical/FG name.
        Examples that SHOULD split:
          "FG1, FG2"
          "ABC , XYZ"
        Examples that should NOT split:
          "1,2 DIMETHYL PROPYL AMINE"
          "2,4,6 TRIMETHYL PHENYL ACETYL CHLORIDE"
        """
        if not s or "," not in s:
            return False
        return (", " in s) or (" ," in s)

    # already list/tuple -> keep each item intact
    if isinstance(raw, (list, tuple)):
        tokens = []
        for v in raw:
            if v is None:
                continue
            s = str(v).strip()
            if not s:
                continue

            # JSON array string inside list item
            if s.startswith("[") and s.endswith("]"):
                try:
                    arr = json.loads(s)
                    tokens.extend(_parse_fg_payload(arr))
                    continue
                except Exception:
                    pass

            # IMPORTANT: do not split plain list items by comma
            tokens.append(s)

        return _dedupe_keep_order(tokens)

    s = str(raw).strip()
    if not s:
        return []

    # JSON array string
    if s.startswith("[") and s.endswith("]"):
        try:
            arr = json.loads(s)
            return _parse_fg_payload(arr)
        except Exception:
            pass

    # split only if comma is a real delimiter between values
    if _looks_like_delimited_multi(s):
        return _dedupe_keep_order([t.strip() for t in s.split(",") if t.strip()])

    return [s]

def _selected_fgs_from_request(request, keys=FG_POST_KEYS):
    """
    Reads FG multi-select robustly.
    Works even if frontend posts:
      - multiple same-name inputs (getlist)
      - single comma string
      - JSON string
    """
    for k in (keys or []):
        # Prefer getlist (true multi)
        vals = request.POST.getlist(k)
        parsed = _parse_fg_payload(vals)
        if parsed:
            return parsed

        # Fallback single key (comma/json)
        one = request.POST.get(k)
        parsed = _parse_fg_payload(one)
        if parsed:
            return parsed

    return []

def _extract_rm_code(v: str) -> str:
    s = (v or "").strip()
    if not s:
        return ""
    m = _RM_CODE_TAIL_RE.search(s)
    if m:
        tail = (m.group(1) or "").strip()
        if tail:
            return tail
    return s

def _norm_code(v: str) -> str:
    return (v or "").strip().upper()

def _norm_rm_code(v: str) -> str:
    return _norm_code(_extract_rm_code(v))

def _d(v, default=Decimal("0")) -> Decimal:
    if v in (None, "", "None"):
        return default
    if isinstance(v, Decimal):
        return v
    try:
        s = str(v).strip().replace(",", "")
        if not s or s.upper() == "NONE":
            return default
        return Decimal(s)
    except (InvalidOperation, ValueError, TypeError):
        return default

def _q(v, places="0.000000") -> Decimal:
    qv = Decimal(places)
    x = _d(v, Decimal("0"))
    try:
        return x.quantize(qv, rounding=ROUND_HALF_UP)
    except Exception:
        return Decimal("0").quantize(qv, rounding=ROUND_HALF_UP)

def _norm_pt(v: str) -> str:
    s = (v or "").strip().upper()
    if s in ("IMPORT", "IMP", "IMPORT PURCHASE"):
        return "IMPORT"
    return "LOCAL"

def _fg_choices_for_plan(plan):
    # Required: FG list from ProductionBudgetFG for this plan
    return list(
        ProductionBudgetFG.objects
        .filter(plan=plan)
        .exclude(fg_name__isnull=True).exclude(fg_name__exact="")
        .values_list("fg_name", flat=True)
        .distinct()
        .order_by("fg_name")
    )

def _fg_month_qty(plan, fg_name: str) -> dict:
    out = {m: Decimal("0.000000") for m in MONTHS}
    pfg = ProductionBudgetFG.objects.filter(plan=plan, fg_name__iexact=fg_name).first()
    if not pfg:
        return out
    for m in MONTHS:
        out[m] = _q(getattr(pfg, m, 0) or 0, "0.000000")
    return out

def _fg_month_qty_bulk(plan, fg_names):
    """
    Returns {FG_UPPER: {apr..mar: Decimal}} in one query.
    """
    out = {}
    fg_names = _dedupe_keep_order(fg_names or [])
    if not fg_names:
        return out

    rows = (
        ProductionBudgetFG.objects
        .filter(plan=plan, fg_name__in=fg_names)   # SQL Server collation is usually CI; OK in most setups
        .values("fg_name", *MONTHS)
    )

    for r in rows:
        fg_u = (r.get("fg_name") or "").strip().upper()
        if not fg_u:
            continue
        out[fg_u] = {m: _q(r.get(m, 0) or 0, "0.000000") for m in MONTHS}

    return out

def _key(rm_code: str, is_captive: bool) -> str:
    # unique key for qty_map/meta_map
    return f"CODE::{_norm_rm_code(rm_code)}::{'C' if is_captive else 'N'}"

def _erp_meta_for_fgs(fg_names):
    """
    Fill missing rm_name/unit if BOM input line is missing/blank.
    """
    fg_names = [x.strip() for x in (fg_names or []) if (x or "").strip()]
    meta = {}
    if not fg_names:
        return meta

    qs = (
        ERPBOMRow.objects
        .filter(fg_name__in=fg_names, type__in=RMC_INPUT_TYPES)
        .exclude(bom_item_code__isnull=True).exclude(bom_item_code__exact="")
        .values("bom_item_code", "bom_item_name", "unit")
    )

    for r in qs:
        code = _norm_rm_code(r.get("bom_item_code") or "")
        if not code:
            continue
        meta.setdefault(code, {
            "rm_code": code,
            "rm_name": (r.get("bom_item_name") or "").strip(),
            "unit": (r.get("unit") or "").strip(),
        })

    return meta

def _captive_flags_for_fg_codes(fg_names, codes):
    """
    Determine captive/non-captive from ProductionNorm.is_captive (this is your "clicked Captive" state).
    """
    fg_names_u = { (x or "").strip().upper() for x in (fg_names or []) if (x or "").strip() }
    codes_u = { _norm_rm_code(c) for c in (codes or []) if (c or "").strip() }

    flags = {}  # (FG_UPPER, CODE_UPPER) -> bool

    if not fg_names_u or not codes_u:
        return flags

    qs = (
        ProductionNorm.objects
        .filter(fg_name__in=list(fg_names_u), bom_item_code__in=list(codes_u))
        .values("fg_name", "bom_item_code", "is_captive")
    )
    for r in qs:
        fg = (r.get("fg_name") or "").strip().upper()
        code = _norm_rm_code(r.get("bom_item_code") or "")
        flags[(fg, code)] = bool(r.get("is_captive") or False)

    return flags

def _bom_inputs_for_fg(fg_name: str):
    """
    Read norms + material list from ProductionBOMInputLine.

    IMPORTANT:
      - budget_norm is the source for Required Qty
      - target_norm is optional (manual) and should NOT affect Required Qty
      - is_captive (Captive Tick) must come from ProductionBOMInputLine so RMC sync can split captive automatically
    """

    def _boolish(v) -> bool:
        if isinstance(v, bool):
            return v
        if v is None:
            return False
        # numeric truthy (1/0)
        if isinstance(v, (int, float)):
            return v == 1
        s = str(v).strip().lower()
        return s in ("1", "true", "yes", "y", "on", "checked")

    def _first_attr(obj, names):
        """Return first existing attribute value among names; else None."""
        for n in names:
            if hasattr(obj, n):
                return getattr(obj, n)
        return None

    fg = (fg_name or "").strip()
    if not fg:
        return []

    bom = ProductionBOM.objects.filter(fg_name__iexact=fg, is_active=True).first()
    if not bom:
        return []

    qs = ProductionBOMInputLine.objects.filter(bom=bom).order_by("id")
    rows = []

    # Candidate field names for "Captive Tick" across schema variations
    CAPTIVE_FIELDS = (
        "is_captive",
        "captive_tick",
        "captive",
        "is_captive_tick",
        "captive_flag",
        "captive_enabled",
    )

    for ln in qs:
        # -----------------------------
        # Robust code
        # -----------------------------
        code = _norm_rm_code(
            (getattr(ln, "material_code", "") or "").strip()
            or (getattr(ln, "bom_item_code", "") or "").strip()
            or (getattr(ln, "rm_code", "") or "").strip()
        )

        name = (
            (getattr(ln, "material_name", "") or "").strip()
            or (getattr(ln, "bom_item_name", "") or "").strip()
        )

        unit = (getattr(ln, "unit", "") or "").strip()

        cat = (
            (getattr(ln, "material_category", "") or "").strip()
            or (getattr(ln, "type", "") or "").strip()
        )

        # keep your existing category guard
        if cat and cat not in RMC_INPUT_TYPES:
            continue
        if not code and not name:
            continue

        # -----------------------------
        # ✅ Captive Tick (robust read)
        # -----------------------------
        raw_captive = _first_attr(ln, CAPTIVE_FIELDS)
        is_captive = _boolish(raw_captive)

        # -----------------------------
        # ✅ KEEP BOTH NORMS SEPARATELY
        # -----------------------------
        budget_norm = _q(getattr(ln, "budget_norm", None) or 0, "0.000000")

        raw_target = getattr(ln, "target_norm", None)
        target_norm = None if raw_target in (None, "") else _q(raw_target or 0, "0.000000")

        # legacy/display convenience (NOT for required qty)
        effective_norm = budget_norm

        rows.append({
            "code": code,
            "name": name,
            "unit": unit,
            "category": cat,

            # ✅ critical fields
            "budget_norm": budget_norm,
            "target_norm": target_norm,

            # ✅ captive tick for auto-splitting in RMC sync
            "is_captive": is_captive,

            # optional aliases (harmless, but helps compatibility)
            "captive": is_captive,
            "captive_tick": is_captive,

            # legacy/display convenience
            "norm": effective_norm,
        })

    return rows

def _compute_rmc_qty_from_production(plan, fg_names):
    """
    ✅ Computes RM consumption for selected FGs
    ✅ Splits captive vs normal using:
        1) Production "Captive Tick" coming in _bom_inputs_for_fg() dict (preferred)
        2) Fallback: ProductionNorm.is_captive via _captive_flags_for_fg_codes()

    REQUIRED QTY RULE:
      month_qty_rm = FG_month_qty * ProductionBudgetNorm (BUDGET_NORM)
      required_qty (annual) = SUM(apr..mar)

    Returns:
      qty_map[key] -> {apr..mar, annual, required_qty}
      meta_map[key] -> {rm_code, rm_name, unit, category, is_captive}
    """

    def _boolish(v) -> bool:
        """Robust bool parse for values from DB/forms/json ('True','1','yes', etc.)."""
        if isinstance(v, bool):
            return v
        if v is None:
            return False
        s = str(v).strip().lower()
        return s in ("1", "true", "yes", "y", "on")

    def _has_explicit_flag(v) -> bool:
        """
        Meaning: the source actually carried a value (including False/0).
        Only None / "" are treated as 'not provided'.
        """
        return v is not None and str(v).strip() != ""

    def _pick_flag_from_row(r: dict):
        """
        Prefer keys in this order. Return (found, value).
        found=True means it's explicitly present (even False/0).
        """
        for key in ("is_captive", "captive_tick", "captive"):
            if key in r and _has_explicit_flag(r.get(key)):
                return True, r.get(key)
        return False, None

    # normalize FG names (dedupe while preserving order)
    seen = set()
    fg_clean = []
    for x in (fg_names or []):
        v = (x or "").strip()
        if not v:
            continue
        u = v.upper()
        if u in seen:
            continue
        seen.add(u)
        fg_clean.append(v)

    fg_names = fg_clean
    qty_map = defaultdict(lambda: {m: Decimal("0.000000") for m in MONTHS})
    meta_map = {}

    fg_month_map = _fg_month_qty_bulk(plan, fg_names)


    # ERP fallback meta (name/unit)
    erp_meta = _erp_meta_for_fgs(fg_names)

    # Cache BOM inputs per FG (avoid double DB hits and ensure consistency)
    inputs_by_fg = {}
    fg_codes = defaultdict(set)

    for fg in fg_names:
        fg_u = fg.strip().upper()
        inputs = _bom_inputs_for_fg(fg) or []
        inputs_by_fg[fg_u] = inputs

        for r in inputs:
            code = (r.get("code") or "").strip()
            if code:
                fg_codes[fg_u].add(_norm_rm_code(code))

    all_codes = set()
    for s in fg_codes.values():
        all_codes |= s

    # Fallback captive flags from ProductionNorm (existing behavior)
    # Only compute if we actually have codes to check.
    captive_flags = {}
    if all_codes and fg_codes:
        captive_flags = _captive_flags_for_fg_codes(list(fg_codes.keys()), list(all_codes))

    for fg in fg_names:
        fg_u = fg.strip().upper()
        fg_u = fg.strip().upper()
        fg_month = fg_month_map.get(fg_u) or {m: Decimal("0.000000") for m in MONTHS}
   # dict apr..mar (Decimal)
        inputs = inputs_by_fg.get(fg_u) or []

        for r in inputs:
            code_u = _norm_rm_code(r.get("code") or "")
            if not code_u:
                continue

            # ✅ Captive decision priority:
            # 1) use explicit flag from BOM input dict
            # 2) else fallback to ProductionNorm captive_flags
            found, raw_flag = _pick_flag_from_row(r)
            if found:
                is_captive = _boolish(raw_flag)
            else:
                is_captive = bool(captive_flags.get((fg_u, code_u), False))

            k = _key(code_u, is_captive)

            rm_name = (r.get("name") or "").strip()
            unit = (r.get("unit") or "").strip()
            cat = (r.get("category") or "").strip()

            # ✅ Production Budget Norm (budget_norm) is the source of truth
            prod_norm = r.get("budget_norm", None)
            if prod_norm in (None, ""):
                prod_norm = r.get("norm", None)  # legacy fallback only

            prod_norm = _q(prod_norm or 0, "0.000000")

            # compute monthwise consumption (RM requirement)
            for m in MONTHS:
                qty_map[k][m] += (_d(fg_month.get(m, 0), Decimal("0")) * prod_norm)

            # meta
            if k not in meta_map:
                meta_map[k] = {
                    "rm_code": code_u,
                    "rm_name": rm_name,
                    "unit": unit,
                    "category": cat,
                    "is_captive": is_captive,
                }

            # enrich blanks from ERP (do not overwrite)
            em = erp_meta.get(code_u)
            if em:
                if not meta_map[k]["rm_name"] and em.get("rm_name"):
                    meta_map[k]["rm_name"] = em["rm_name"]
                if not meta_map[k]["unit"] and em.get("unit"):
                    meta_map[k]["unit"] = em["unit"]

    # annual totals + required qty + quantize
    for k, d in qty_map.items():
        for m in MONTHS:
            d[m] = _q(d[m], "0.000000")

        annual = _q(sum((d[m] for m in MONTHS), Decimal("0.000000")), "0.000000")
        d["annual"] = annual

        # ✅ required_qty MUST equal annual sum of (FG month qty * budget norm)
        d["required_qty"] = annual

    return qty_map, meta_map

def _assign_qty_to_line(line: RMCBudgetLine, per_month: dict):
    """
    Assigns month quantities + required_qty/total_qty safely.

    Key safety rules:
    - If per_month is empty/None: do nothing (prevents wiping on manual save paths).
    - Only writes month fields when month keys are present.
    - required_qty/total_qty are written only when we can compute a reliable annual value.
    """
    if not per_month or not hasattr(per_month, "get"):
        return

    # Do we have any month values in this payload?
    month_present = any(
        (m in per_month) and (per_month.get(m) not in (None, ""))
        for m in MONTHS
    )

    # Compute annual safely:
    # 1) use explicit annual if present
    # 2) else derive from months if month data is present
    annual = None
    annual_raw = per_month.get("annual", None)

    if annual_raw not in (None, ""):
        annual = _q(annual_raw, "0.000000")
    elif month_present:
        total = 0
        for m in MONTHS:
            total += _q(per_month.get(m, 0), "0.000000")
        annual = _q(total, "0.000000")

    # Only write month fields when we actually have month data
    # (prevents wiping existing month splits if per_month only contains "annual" or is partial)
    if month_present:
        for m in MONTHS:
            fld = f"qty_{m}"
            if hasattr(line, fld):
                setattr(line, fld, _q(per_month.get(m, 0), "0.000000"))

    # Write required_qty/total_qty only when annual is known
    if annual is not None:
        if hasattr(line, "required_qty"):
            line.required_qty = annual
        if hasattr(line, "total_qty"):
            line.total_qty = annual

def _rmc_sync_update_fields(line: RMCBudgetLine):
    """
    Update-fields list used in final sync pass.
    Must include:
      - month qty fields (qty_apr..qty_mar)
      - computed fields (required_qty, total_qty, budget_rate_inr)
      - core identity/meta fields (rm_code, rm_name, unit, is_captive)
      - rate inputs so budget_rate_inr recompute is persisted correctly when you save(update_fields=...)
    """
    fields = []

    # month fields
    for m in MONTHS:
        fld = f"qty_{m}"
        if hasattr(line, fld):
            fields.append(fld)

    # ✅ rate inputs (important: if you change these, computed budget_rate_inr must reflect)
    for f in (
        "purchase_type",
        "local_rate_inr",
        "import_rate_usd",
        "duty_percent",
        "freight_inr",
        "clearance_inr",
    ):
        if hasattr(line, f):
            fields.append(f)

    # ✅ computed + core fields
    for f in (
        "required_qty",     # stored (you assign via _assign_qty_to_line)
        "total_qty",        # computed in model.save()
        "budget_rate_inr",  # computed in model.save()
        "rm_code",
        "rm_name",
        "unit",
        "is_captive",
    ):
        if hasattr(line, f):
            fields.append(f)

    # timestamp (auto_now)
    if hasattr(line, "updated_at"):
        fields.append("updated_at")

    # de-dup (preserve order)
    out, seen = [], set()
    for f in fields:
        if f not in seen:
            out.append(f)
            seen.add(f)
    return out

# -----------------------------------------------------------------------------
# Formset
# -----------------------------------------------------------------------------
RMCBudgetLineFormSet = modelformset_factory(
    RMCBudgetLine,
    form=RMCBudgetLineForm,
    extra=0,
    can_delete=True,
)

# =============================================================================
# Views
# =============================================================================

@login_required
def rmc_budget_home(request):
    plan = BudgetPlan.objects.filter(is_active=True).order_by("-updated_at", "-id").first()
    if not plan:
        messages.error(request, "No active Budget Plan found.")
        return redirect("accounts_budget:budget_home")

    rmc_budget, _ = RMCBudget.objects.get_or_create(
        plan=plan,
        defaults={"created_by": request.user}
    )
    mc = mc_get(rmc_budget, "RMC")
    locked = _is_locked(mc)

    ex = _d(getattr(rmc_budget, "usd_inr", None), Decimal("0"))

    # ✅ SAFE ordering: "is_captive" might not be a DB field
    try:
        lines_qs = rmc_budget.lines.all().order_by("rm_name", "rm_code", "is_captive")
    except FieldError:
        lines_qs = rmc_budget.lines.all().order_by("rm_name", "rm_code")

    lines = []

    for l in lines_qs:
        # ✅ normalize captive bool (works even if field name differs)
        cap = bool(
            getattr(l, "is_captive", False)
            or getattr(l, "captive", False)
            or getattr(l, "captive_tick", False)
        )
        l.is_captive = cap  # force attribute for templates

        pt = _norm_pt(getattr(l, "purchase_type", ""))
        l.is_import = (pt == "IMPORT")

        if l.is_import:
            usd = _d(getattr(l, "import_rate_usd", None), Decimal("0"))
            duty_pct = _d(getattr(l, "duty_percent", None), Decimal("0"))
            freight = _d(getattr(l, "freight_inr", None), Decimal("0"))
            clearance = _d(getattr(l, "clearance_inr", None), Decimal("0"))

            rate_inr = _q(usd * ex, "0.0000") if (ex > 0 and usd > 0) else _q(0, "0.0000")
            custom_duty = _q(rate_inr * duty_pct / Decimal("100"), "0.0000") if duty_pct else _q(0, "0.0000")
            clearance_total = _q(freight + clearance, "0.0000")
        else:
            local_rate = _d(getattr(l, "local_rate_inr", None), Decimal("0"))
            rate_inr = _q(local_rate, "0.0000")
            custom_duty = None
            clearance_total = None

        l.rate_inr_calc = rate_inr
        l.custom_duty_inr_calc = custom_duty
        l.clearance_total_calc = clearance_total

        # ✅ optional convenience for UI (doesn't break anything)
        l.display_name = f"{getattr(l,'rm_name','')}{' [CAPTIVE]' if l.is_captive else ''}"

        lines.append(l)

    # ✅ If you need consistent sorting even when "is_captive" isn't a DB field:
    # lines.sort(key=lambda x: ((getattr(x,"rm_name","") or "").lower(), getattr(x,"rm_code","") or "", bool(getattr(x,"is_captive", False))))

    return render(request, "accounts/budget/rmc_budget_home.html", {
        "plan": plan,
        "rmc_budget": rmc_budget,
        "lines": lines,
        "mc": mc,
        "is_locked": locked,
        "is_approver": _is_approver(request.user),
        "mc_scope": "RMC",
        "mc_model": "accounts_budget.RMCBudget",
        "mc_pk": rmc_budget.id,
    })

# END-----------------------------------ENDPOINT--------------------------------

@login_required
def rmc_budget_create(request):
    plan = BudgetPlan.objects.filter(is_active=True).order_by("-updated_at", "-id").first()
    if not plan:
        messages.error(request, "No active Budget Plan found.")
        return redirect("accounts_budget:budget_home")

    RMCBudget.objects.get_or_create(plan=plan, defaults={"created_by": request.user})
    return redirect("accounts_budget:rmc_budget_edit")
# ==============================================================================
# RMC EDIT (complete) — robust Raw + Captive save + terminal debug of frontend POST
# ==============================================================================

import os
import json
from datetime import datetime
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction

# ------------------------------------------------------------------------------
# DEBUG TOGGLE (set env var DEBUG_RMC_POST=1 to print request payload in terminal)
# ------------------------------------------------------------------------------
DEBUG_RMC_POST = os.environ.get("DEBUG_RMC_POST", "0").strip().lower() in ("1", "true", "yes", "on")


def _dbg_post(request, *, title="RMC DEBUG", prefix="form", max_rows=80):
    """
    Prints what frontend is sending (POST keys + formset mgmt + per-row values).
    Enable with: DEBUG_RMC_POST=1
    """
    if not DEBUG_RMC_POST:
        return

    try:
        print("\n" + "=" * 140)
        print(f"[{datetime.now().isoformat(timespec='seconds')}] {title}")
        print(f"PATH={request.path}  METHOD={request.method}  USER={getattr(request.user,'username',None)}")
        print("-" * 140)

        # management form keys (must exist for formset binding)
        mgmt = {
            f"{prefix}-TOTAL_FORMS": request.POST.get(f"{prefix}-TOTAL_FORMS"),
            f"{prefix}-INITIAL_FORMS": request.POST.get(f"{prefix}-INITIAL_FORMS"),
            f"{prefix}-MIN_NUM_FORMS": request.POST.get(f"{prefix}-MIN_NUM_FORMS"),
            f"{prefix}-MAX_NUM_FORMS": request.POST.get(f"{prefix}-MAX_NUM_FORMS"),
        }
        print("MGMT:", mgmt)

        keys = sorted(list(request.POST.keys()))
        print("POST_KEYS_COUNT:", len(keys))
        print("POST_KEYS_HEAD:", keys[:60], "..." if len(keys) > 60 else "")

        # show only important fields (keeps logs readable)
        keep_contains = (
            "action",
            "selected", "fg",
            f"{prefix}-",          # formset fields
            "row_key", "bucket_key", "rm_key", "key",
            "rm_code", "is_captive",
            "purchase_type",
            "local_rate", "import_rate", "duty", "freight", "clearance",
        )
        important = {}
        for k in keys:
            kl = k.lower()
            if kl.startswith("csrf"):
                continue
            if any(s in k for s in keep_contains):
                vlist = request.POST.getlist(k)
                important[k] = vlist if len(vlist) > 1 else (vlist[0] if vlist else "")

        print("\nIMPORTANT_POST_FIELDS (filtered):")
        print(json.dumps(important, indent=2, ensure_ascii=False))

        # per-row preview from management count
        try:
            total = int(request.POST.get(f"{prefix}-TOTAL_FORMS") or "0")
        except Exception:
            total = 0

        print(f"\nFORMS_BREAKDOWN prefix='{prefix}' TOTAL_FORMS={total}")
        for i in range(min(total, max_rows)):
            base = f"{prefix}-{i}-"
            row = {}
            for field in (
                "id", "rm_code", "rm_name", "unit",
                "is_captive",
                "purchase_type",
                "local_rate_inr", "import_rate_usd", "duty_percent", "freight_inr", "clearance_inr",
                "row_key", "bucket_key", "rm_key", "key",
                "DELETE",
            ):
                kk = base + field
                if kk in request.POST:
                    row[field] = request.POST.get(kk)
            if row:
                print(f"  [{i}] {row}")

        print("=" * 140 + "\n")
    except Exception as e:
        print("[RMC DEBUG] failed:", repr(e))


# ------------------------------------------------------------------------------
# Helpers for RMC Budget Edit (key parsing + map lookup)
# ------------------------------------------------------------------------------
def _rmc_bucket_suffix(cap: bool) -> str:
    return "C" if bool(cap) else "N"


def _as_bool(v) -> bool:
    if isinstance(v, bool):
        return v
    if v is None:
        return False
    s = str(v).strip().lower()
    return s in ("1", "true", "yes", "y", "on", "checked")


def _rmc_parse_any_code_cap(v: str):
    """
    Accepts any of:
      - "26000122|0" / "26000122|1"
      - "CODE::26000122::C" / "CODE::26000122::N"
      - "26000122::C" / "26000122::N"
      - "26000122"
    Returns: (code_only, cap_or_None)
    """
    s = (v or "").strip()
    if not s:
        return ("", None)

    # "::" formats
    if "::" in s:
        parts = [p.strip() for p in s.split("::") if p is not None]
        # examples:
        #   ["CODE","26000122","C"]
        #   ["26000122","C"]
        cap = None
        tail = (parts[-1] or "").strip().upper() if parts else ""
        if tail in ("C", "N", "1", "0", "TRUE", "FALSE"):
            if tail in ("C", "1", "TRUE"):
                cap = True
            elif tail in ("N", "0", "FALSE"):
                cap = False

        # choose code segment
        code = ""
        if len(parts) >= 2 and parts[0].strip().upper() == "CODE":
            code = parts[1]
        else:
            code = parts[0]

        return (_norm_rm_code(code), cap)

    # "|" format
    if "|" in s:
        code, cap = s.split("|", 1)
        return (_norm_rm_code(code), _as_bool(cap))

    return (_norm_rm_code(s), None)


def _split_bucket_key(v: str):
    """
    Backward-compatible alias of your earlier behavior.
    """
    return _rmc_parse_any_code_cap(v)


def _rmc_key_candidates(code: str, cap: bool):
    """
    Produce likely keys that might exist in qty_map/meta_map across older/newer formats.
    Primary is your current _key(code, cap) => "CODE::<code>::C|N"
    """
    code = _norm_rm_code(code)
    suf = _rmc_bucket_suffix(cap)

    candidates = []

    # ✅ canonical (preferred)
    try:
        candidates.append(_key(code, cap))
    except Exception:
        pass

    # ✅ safe fallbacks
    candidates.extend([
        f"CODE::{code}::{suf}",
        f"{code}::{suf}",
        f"{code}|{1 if cap else 0}",
        code,
        f"CODE::{code}",
    ])

    out, seen = [], set()
    for k in candidates:
        if k and k not in seen:
            out.append(k)
            seen.add(k)
    return out


def _rmc_lookup_maps(qty_map, meta_map, code: str, cap: bool):
    """
    Returns: (qk_used, per_dict_or_None, meta_dict_or_None)
    """
    for qk in _rmc_key_candidates(code, cap):
        per = (qty_map or {}).get(qk)
        mm = (meta_map or {}).get(qk)
        if per is not None or mm is not None:
            return qk, per, mm
    return None, None, None

def _rmc_pick_bucket(qty_map, meta_map, rm_code: str, *, cap_hint=None, default_cap=False):
    """
    Decide which bucket to use for this rm_code.

    Priority:
      1) If cap_hint provided and that bucket exists -> use it
      2) If only one bucket exists -> use it
      3) If both buckets exist -> use default_cap (usually existing DB row cap)
      4) If none exists -> fall back to cap_hint/default_cap
    Returns: (cap_bool, qk_used, per_dict, meta_dict)
    """
    rm_code = _norm_rm_code(rm_code)

    hits = []
    for cap_try in (False, True):
        qk, per, mm = _rmc_lookup_maps(qty_map, meta_map, rm_code, cap_try)
        if per is not None or mm is not None:
            hits.append((cap_try, qk, per, mm))

    if not hits:
        cap = bool(cap_hint) if cap_hint is not None else bool(default_cap)
        return cap, None, None, None

    # cap_hint wins if present and exists
    if cap_hint is not None:
        want = bool(cap_hint)
        for cap_try, qk, per, mm in hits:
            if cap_try == want:
                return cap_try, qk, per, mm

    if len(hits) == 1:
        return hits[0]

    # both exist -> prefer default_cap (existing row)
    want = bool(default_cap)
    for cap_try, qk, per, mm in hits:
        if cap_try == want:
            return cap_try, qk, per, mm

    return hits[0]

def _form_index_from_prefix(pfx: str):
    """
    "form-12" -> 12 ; else None
    """
    try:
        # expected: "<prefix>-<i>"
        if "-" in pfx:
            tail = pfx.split("-")[-1]
            return int(tail)
    except Exception:
        return None
    return None


def _list_value_for_form(request, keyname: str, form_index: int):
    """
    If frontend sends non-prefixed list inputs like:
      <input name="row_key" value="26000122|0">
      <input name="row_key" value="28000004|1">
    then request.POST.getlist("row_key") has values in order.
    This safely returns the i-th item for current form.
    """
    if form_index is None:
        return None
    vals = request.POST.getlist(keyname)
    if not vals:
        return None
    if 0 <= form_index < len(vals):
        return vals[form_index]
    return None


# ------------------------------------------------------------------------------
# Update-fields list used in final sync pass
# ------------------------------------------------------------------------------
def _rmc_sync_update_fields(line: "RMCBudgetLine"):
    """
    Must include:
      - month qty fields (qty_apr..qty_mar)
      - rate inputs (so recompute budget_rate_inr is persisted correctly)
      - computed fields (required_qty, total_qty, budget_rate_inr)
      - identity/meta (rm_code, rm_name, unit, is_captive)
    """
    fields = []

    for m in MONTHS:
        fld = f"qty_{m}"
        if hasattr(line, fld):
            fields.append(fld)

    for f in (
        "purchase_type",
        "local_rate_inr",
        "import_rate_usd",
        "duty_percent",
        "freight_inr",
        "clearance_inr",
    ):
        if hasattr(line, f):
            fields.append(f)

    for f in (
        "required_qty",
        "total_qty",
        "budget_rate_inr",
        "rm_code",
        "rm_name",
        "unit",
        "is_captive",
    ):
        if hasattr(line, f):
            fields.append(f)

    if hasattr(line, "updated_at"):
        fields.append("updated_at")

    out, seen = [], set()
    for f in fields:
        if f not in seen:
            out.append(f)
            seen.add(f)
    return out


SYNC_ALL_ACTIONS = {
    "sync_all_inputs",
    "sync_all_input",
    "sync_all",
    "syncall_inputs",
    "syncall",
}

def _is_truthy(v) -> bool:
    s = ("" if v is None else str(v)).strip().lower()
    return s in ("1", "true", "yes", "y", "on")

def _is_sync_all_request(request, action: str) -> bool:
    a = (action or "").strip().lower()

    # common cases:
    if a in SYNC_ALL_ACTIONS:
        return True

    # button posted as its own name, not via "action"
    if "sync_all_inputs" in request.POST or "sync_all_input" in request.POST:
        return True

    # some UIs use this flag (you already use it in rmc_qty_preview)
    if _is_truthy(request.POST.get("all_inputs")):
        return True

    # fallback: contains "sync" and "all"
    if "sync" in a and "all" in a:
        return True

    return False


def _rmc_apply_sync_maps_only(rmc_budget: RMCBudget, qty_map: dict, meta_map: dict):
    """
    Backend-only sync:
      - updates existing lines' month qty + required_qty
      - creates missing lines for any qty_map key with annual>0
      - does NOT depend on formset POST keys
    """

    # Load all existing lines
    all_lines = list(RMCBudgetLine.objects.filter(budget=rmc_budget))

    # Normalize and index existing by (code, cap)
    existing_by_key = {}
    for ln in all_lines:
        raw = getattr(ln, "rm_code", "") or ""
        code_clean, cap_hint = _rmc_parse_any_code_cap(raw)

        if code_clean and code_clean != raw:
            ln.rm_code = code_clean
        if cap_hint is not None and hasattr(ln, "is_captive"):
            ln.is_captive = bool(cap_hint)

        k = (_norm_rm_code(getattr(ln, "rm_code", "") or ""), bool(getattr(ln, "is_captive", False)))
        if k[0]:
            existing_by_key[k] = ln

    # Upsert based on qty_map
    for qk, per in (qty_map or {}).items():
        annual = _q(per.get("annual", 0), "0.000000")
        if annual <= 0:
            continue

        code, cap_hint = _rmc_parse_any_code_cap(str(qk))
        code = _norm_rm_code(code)
        if not code:
            continue
        cap = bool(cap_hint) if cap_hint is not None else False

        # meta
        _, _, mm = _rmc_lookup_maps(qty_map, meta_map, code, cap)
        if not mm:
            mm = (meta_map or {}).get(qk) or {}

        ln = existing_by_key.get((code, cap))
        is_new = False
        if not ln:
            ln = RMCBudgetLine(
                budget=rmc_budget,
                rm_code=code,
                is_captive=cap,
                purchase_type=RMCPurchaseType.LOCAL,
            )
            is_new = True

        # apply qty + required_qty
        _assign_qty_to_line(ln, per)

        # fill meta if empty
        if mm:
            if hasattr(ln, "rm_name") and not (getattr(ln, "rm_name", "") or "").strip():
                ln.rm_name = (mm.get("rm_name") or "").strip()
            if hasattr(ln, "unit") and not (getattr(ln, "unit", "") or "").strip():
                ln.unit = (mm.get("unit") or "").strip()

        # ensure captive + code are correct
        if hasattr(ln, "is_captive"):
            ln.is_captive = cap
        ln.rm_code = code

        if is_new:
            ln.save()
        else:
            uf = _rmc_sync_update_fields(ln)
            if hasattr(ln, "required_qty") and "required_qty" not in uf:
                uf.append("required_qty")
            ln.save(update_fields=uf)

        existing_by_key[(code, cap)] = ln

def _save_rmc_formset_safely(rmc_budget, formset):
    """
    Prevent duplicate-key crash on unique constraint:
      (budget, rm_code, is_captive)

    Strategy:
      - normalize rm_code
      - resolve is_captive bucket
      - if another row already exists for same unique key, UPDATE it
      - else INSERT normally
      - respect deleted rows
    """
    objs = formset.save(commit=False)

    # delete rows explicitly marked for deletion first
    for obj in formset.deleted_objects:
        obj.delete()

    for o in objs:
        o.budget = rmc_budget

        # normalize unique-key fields
        o.rm_code = (getattr(o, "rm_code", "") or "").strip().upper()

        if hasattr(o, "is_captive"):
            o.is_captive = bool(getattr(o, "is_captive", False))
            is_captive = o.is_captive
        else:
            is_captive = False

        # blank key rows should not be saved
        if not o.rm_code:
            continue

        # find another row with same unique key
        existing = (
            RMCBudgetLine.objects
            .filter(
                budget=rmc_budget,
                rm_code=o.rm_code,
                is_captive=is_captive,
            )
            .exclude(pk=o.pk if getattr(o, "pk", None) else None)
            .first()
        )

        if existing:
            # copy editable fields from incoming row to existing row
            editable_fields = [
                "rm_name",
                "unit",
                "purchase_type",
                "required_qty",
                "local_rate_inr",
                "import_rate_usd",
                "duty_percent",
                "freight_inr",
                "clearance_inr",
                "qty_apr", "qty_may", "qty_jun", "qty_jul",
                "qty_aug", "qty_sep", "qty_oct", "qty_nov",
                "qty_dec", "qty_jan", "qty_feb", "qty_mar",
                "apr", "may", "jun", "jul", "aug", "sep",
                "oct", "nov", "dec", "jan", "feb", "mar",
                "annual_qty", "annual_qty_mt",
            ]

            for fld in editable_fields:
                if hasattr(existing, fld) and hasattr(o, fld):
                    setattr(existing, fld, getattr(o, fld))

            # preserve normalized unique key
            existing.budget = rmc_budget
            existing.rm_code = o.rm_code
            if hasattr(existing, "is_captive"):
                existing.is_captive = is_captive

            existing.save()
        else:
            o.save()

    # if model formset has m2m hooks
    if hasattr(formset, "save_m2m"):
        formset.save_m2m()

# ------------------------------------------------------------------------------
# MAIN VIEW
# ------------------------------------------------------------------------------
@login_required
def rmc_budget_edit(request):
    plan = (
        BudgetPlan.objects.filter(is_active=True)
        .order_by("-updated_at", "-id")
        .first()
    )
    if not plan:
        messages.error(request, "No active Budget Plan found.")
        return redirect("accounts_budget:budget_home")

    rmc_budget, _ = RMCBudget.objects.get_or_create(
        plan=plan,
        defaults={"created_by": request.user},
    )

    mc = mc_get(rmc_budget, "RMC")
    locked = _is_locked(mc)
    if locked:
        messages.warning(
            request,
            "RMC Budget is locked. It can be edited only if disapproved by approver."
        )
        return redirect("accounts_budget:rmc_budget_home")

    fg_choices = _fg_choices_for_plan(plan)

    has_captive = "is_captive" in RMC_LINE_MODEL_FIELDS
    has_required_qty = "required_qty" in RMC_LINE_MODEL_FIELDS

    order_fields = ["rm_name", "rm_code"]
    if has_captive:
        order_fields.append("is_captive")

    qs = RMCBudgetLine.objects.filter(budget=rmc_budget).order_by(*order_fields)

    FS_PREFIX = "form"

    # ---------------------------
    # Helpers
    # ---------------------------
    def _get_selected_fgs_db():
        return (
            rmc_budget.get_selected_fgs()
            if hasattr(rmc_budget, "get_selected_fgs")
            else (rmc_budget.selected_fgs or [])
        )

    def _persist_selected_fgs_if_posted(selected_fgs_post, selected_fgs_final):
        if selected_fgs_post is None:
            return
        if hasattr(rmc_budget, "set_selected_fgs"):
            rmc_budget.set_selected_fgs(selected_fgs_final)
            rmc_budget.save(update_fields=["selected_fgs", "updated_at"])
        else:
            rmc_budget.selected_fgs = selected_fgs_final
            rmc_budget.save(update_fields=["selected_fgs", "updated_at"])

    def _make_formset(post_data=None, *, validate_rates=True):
        kwargs = dict(
            queryset=qs,
            form_kwargs={"budget": rmc_budget, "validate_rates": validate_rates},
            prefix=FS_PREFIX,
        )
        if post_data is not None:
            return RMCBudgetLineFormSet(post_data, **kwargs)
        return RMCBudgetLineFormSet(**kwargs)

    def _norm_fg_list(lst):
        return [(x or "").strip().upper() for x in _dedupe_keep_order(lst or []) if (x or "").strip()]

    def _pick_fg_for_sync(action_value, selected_fgs_now):
        sync_all = _is_sync_all_request(request, action_value)
        if sync_all:
            return fg_choices
        return selected_fgs_now or []

    def _save_formset_rows_safely(_formset):
        """
        Maintain existing flow, but prevent SQL unique-key crash on:
        (budget, rm_code, is_captive)
        """
        objs = _formset.save(commit=False)

        # delete explicit removals first
        for obj in _formset.deleted_objects:
            obj.delete()

        for o in objs:
            o.budget = rmc_budget

            # normalize key fields
            if hasattr(o, "rm_code"):
                o.rm_code = _norm_rm_code(getattr(o, "rm_code", "") or "")

            if hasattr(o, "purchase_type"):
                o.purchase_type = _norm_pt(getattr(o, "purchase_type", ""))

            cap_val = False
            if has_captive and hasattr(o, "is_captive"):
                cap_val = bool(getattr(o, "is_captive", False))
                o.is_captive = cap_val

            code = _norm_rm_code(getattr(o, "rm_code", "") or "")
            if not code:
                continue

            existing = (
                RMCBudgetLine.objects
                .filter(
                    budget=rmc_budget,
                    rm_code=code,
                    **({"is_captive": cap_val} if has_captive else {})
                )
                .exclude(pk=o.pk if getattr(o, "pk", None) else None)
                .first()
            )

            if existing:
                # merge instead of inserting duplicate key
                editable_fields = [
                    "rm_name",
                    "unit",
                    "purchase_type",
                    "local_rate_inr",
                    "import_rate_usd",
                    "duty_percent",
                    "freight_inr",
                    "clearance_inr",
                    "required_qty",
                    "qty_apr", "qty_may", "qty_jun", "qty_jul",
                    "qty_aug", "qty_sep", "qty_oct", "qty_nov",
                    "qty_dec", "qty_jan", "qty_feb", "qty_mar",
                    "apr", "may", "jun", "jul", "aug", "sep",
                    "oct", "nov", "dec", "jan", "feb", "mar",
                    "annual", "annual_qty", "annual_qty_mt",
                ]
                for fld in editable_fields:
                    if hasattr(existing, fld) and hasattr(o, fld):
                        setattr(existing, fld, getattr(o, fld))

                existing.budget = rmc_budget
                existing.rm_code = code
                if has_captive and hasattr(existing, "is_captive"):
                    existing.is_captive = cap_val
                if hasattr(existing, "purchase_type"):
                    existing.purchase_type = _norm_pt(getattr(existing, "purchase_type", ""))

                existing.save()
            else:
                o.rm_code = code
                o.save()

        if hasattr(_formset, "save_m2m"):
            _formset.save_m2m()

    # ---------------------------
    # GET
    # ---------------------------
    form = RMCBudgetForm(instance=rmc_budget)
    formset = _make_formset(validate_rates=True)

    selected_fgs_db = _get_selected_fgs_db()
    selected_fgs = selected_fgs_db

    if request.method == "POST":
        _dbg_post(request, title="RMC EDIT POST", prefix=FS_PREFIX)

        action = (request.POST.get("action") or "save").lower().strip()
        strict_rates = (action == "submit")

        # FG posted?
        fg_posted = any(k in request.POST for k in ("fg_names", "selected_fgs"))
        selected_fgs_post = _selected_fgs_from_request(request) if fg_posted else None

        if selected_fgs_post is not None:
            selected_fgs = selected_fgs_post

        # FG changed?
        fg_changed = False
        if selected_fgs_post is not None:
            fg_changed = (_norm_fg_list(selected_fgs_post) != _norm_fg_list(selected_fgs_db))

        # Explicit sync?
        sync_requested = _is_sync_all_request(request, action) or _is_truthy(request.POST.get("all_inputs"))

        # ✅ only sync when FG actually changed OR user asked to sync
        DO_SYNC = bool(fg_changed or sync_requested)

        form = RMCBudgetForm(request.POST, instance=rmc_budget)
        formset = _make_formset(request.POST, validate_rates=strict_rates)

        ok_form = form.is_valid()
        ok_fs = formset.is_valid()

        if DEBUG_RMC_POST:
            print("[RMC DEBUG] action=", action, "strict=", strict_rates)
            print("[RMC DEBUG] fg_posted=", fg_posted, "fg_changed=", fg_changed, "sync_requested=", sync_requested, "DO_SYNC=", DO_SYNC)

        if ok_form and ok_fs:
            try:
                with transaction.atomic():
                    form.save()

                    # persist FG selection only if posted
                    _persist_selected_fgs_if_posted(selected_fgs_post, selected_fgs)

                    # -----------------------------
                    # ✅ 1) ALWAYS save user's edits
                    #     duplicate-safe by unique key
                    # -----------------------------
                    _save_formset_rows_safely(formset)

                    # -----------------------------
                    # ✅ 2) ONLY when DO_SYNC = True:
                    #    - compute maps
                    #    - update required qty
                    #    - add missing lines
                    #    - delete not-required lines
                    # -----------------------------
                    if DO_SYNC:
                        fg_for_sync = _pick_fg_for_sync(action, selected_fgs)

                        qty_map, meta_map = _compute_rmc_qty_from_production(plan, fg_for_sync)

                        # backend-only sync; should not overwrite user rates
                        _rmc_apply_sync_maps_only(rmc_budget, qty_map, meta_map)

                        # if FG selection posted (changed), keep only required materials
                        if selected_fgs_post is not None:
                            required_set = set()
                            for qk, per in (qty_map or {}).items():
                                annual = _q(per.get("annual", 0), "0.000000")
                                if annual <= 0:
                                    continue
                                code, cap_hint = _rmc_parse_any_code_cap(str(qk))
                                code = _norm_rm_code(code)
                                if not code:
                                    continue
                                cap = bool(cap_hint) if cap_hint is not None else False
                                if not has_captive:
                                    cap = False
                                required_set.add((code, cap))

                            only = ["id", "rm_code"]
                            if has_captive:
                                only.append("is_captive")

                            delete_ids = []
                            for ln in RMCBudgetLine.objects.filter(budget=rmc_budget).only(*only):
                                code = _norm_rm_code(getattr(ln, "rm_code", "") or "")
                                cap = bool(getattr(ln, "is_captive", False)) if has_captive else False
                                if (code, cap) not in required_set:
                                    delete_ids.append(ln.id)

                            if delete_ids:
                                RMCBudgetLine.objects.filter(budget=rmc_budget, id__in=delete_ids).delete()

                    if action == "submit":
                        mc_submit(rmc_budget, "RMC", request.user)
                        messages.success(request, "RMC Budget saved and submitted for approval.")
                        return redirect("accounts_budget:rmc_budget_home")

                messages.success(request, "RMC Budget saved.")
                return redirect("accounts_budget:rmc_budget_edit")

            except IntegrityError:
                messages.error(
                    request,
                    "Duplicate RM Code + Type found. "
                    "Only one row is allowed for the same RM Code in the same bucket (RAW/CAPTIVE)."
                )
                # rebuild queryset after failed save so page reloads current DB state
                qs = RMCBudgetLine.objects.filter(budget=rmc_budget).order_by(*order_fields)
                formset = _make_formset(validate_rates=strict_rates)

        else:
            messages.error(request, "Please correct the errors.")

    auto_sync_on_load = (request.GET.get("autosync") or "").strip().lower() in ("1", "true", "yes", "on")

    return render(request, "accounts/budget/rmc_budget_edit.html", {
        "plan": plan,
        "rmc_budget": rmc_budget,
        "form": form,
        "formset": formset,
        "selected_fgs": selected_fgs,
        "fg_choices": fg_choices,
        "mc": mc,
        "is_locked": locked,
        "is_approver": _is_approver(request.user),
        "mc_scope": "RMC",
        "mc_model": "accounts_budget.RMCBudget",
        "mc_pk": rmc_budget.id,
        "auto_sync_on_load": auto_sync_on_load,
    })

@require_POST
@csrf_protect
@login_required
def rmc_qty_preview(request):
    plan = BudgetPlan.objects.filter(is_active=True).order_by("-updated_at", "-id").first()
    if not plan:
        return JsonResponse({"ok": False, "error": "No active Budget Plan found."}, status=400)

    all_inputs = (request.POST.get("all_inputs") or "").strip().lower() in ("1", "true", "yes", "on")

    if all_inputs:
        fg_names = _fg_choices_for_plan(plan)
    else:
        
        fg_names = _selected_fgs_from_request(request)


    qty_map, meta_map = _compute_rmc_qty_from_production(plan, fg_names)

    def _as_bool(v) -> bool:
        if isinstance(v, bool):
            return v
        s = ("" if v is None else str(v)).strip().lower()
        return s in ("1", "true", "yes", "y", "on")

    def _split_key(qk):
        """
        Accepts:
          - "RM001|1"
          - "RM001" (old)
          - ("RM001", True/1)
        Returns: (rm_code, is_captive_bool)
        """
        if isinstance(qk, (tuple, list)) and len(qk) >= 2:
            return (str(qk[0] or "").strip().upper(), _as_bool(qk[1]))

        s = (str(qk or "")).strip()
        if "|" in s:
            code, cap = s.split("|", 1)
            return (code.strip().upper(), _as_bool(cap))
        return (s.strip().upper(), False)

    def _canon_key(rm_code: str, is_captive: bool) -> str:
        return f"{(rm_code or '').strip().upper()}|{1 if is_captive else 0}"

    items = []

    for qk, per in (qty_map or {}).items():
        # annual qty
        annual = _q(per.get("annual", 0), "0.000000")
        if annual <= 0:
            continue

        # 1) try meta lookup by current key
        mm = (meta_map or {}).get(qk) or {}

        # 2) normalize rm_code + captive from key/meta
        rm_code_from_key, cap_from_key = _split_key(qk)

        rm_code = (mm.get("rm_code") or rm_code_from_key or "").strip().upper()
        is_captive = _as_bool(mm.get("is_captive")) if "is_captive" in mm else cap_from_key

        key = _canon_key(rm_code, is_captive)

        # 3) if meta_map is keyed by canonical key, pick it up (keeps names/units accurate)
        mm2 = (meta_map or {}).get(key)
        if mm2:
            mm = mm2

        rm_name = (mm.get("rm_name") or "").strip()
        unit = (mm.get("unit") or "").strip()

        items.append({
            "key": key,  # ✅ stable bucket-aware key
            "rm_code": rm_code,
            "rm_name": rm_name,
            "unit": unit,
            "is_captive": 1 if is_captive else 0,  # ✅ JS-friendly
            "months": {m: str(_q(per.get(m, 0), "0.000000")) for m in MONTHS},
            "annual": str(annual),
            "required_qty": str(annual),
        })

    # Keep your sort order; captive grouped last by default (False < True)
    items.sort(
        key=lambda x: (
            (x.get("rm_name") or "").lower(),
            (x.get("rm_code") or ""),
            bool(x.get("is_captive") or 0),
        )
    )

    return JsonResponse({"ok": True, "items": items, "count": len(items)})

# =============================================================================
# RMC Excel report (UNCHANGED LOGIC - kept as-is)
# =============================================================================

@login_required
def rmc_budget_report_excel(request):
    plan = (
        BudgetPlan.objects.filter(is_active=True)
        .order_by("-updated_at", "-id")
        .first()
    )
    if not plan:
        messages.error(request, "No active Budget Plan found.")
        return redirect("accounts_budget:budget_home")

    rmc_budget = RMCBudget.objects.filter(plan=plan).first()
    if not rmc_budget:
        messages.error(request, "RMC Budget not found. Please create it first.")
        return redirect("accounts_budget:rmc_budget_home")

    # ✅ ALL lines from DB (kept)
    lines = list(
        rmc_budget.lines.all().order_by(
            "rm_name",
            "rm_code",
            "is_captive" if hasattr(RMCBudgetLine, "is_captive") else "rm_code",
        )
    )
    ex = _d(getattr(rmc_budget, "usd_inr", None), Decimal("0"))

    # --------------------------------------------------------------------------
    # Helpers (local)
    # --------------------------------------------------------------------------
    def _bool(v) -> bool:
        if isinstance(v, bool):
            return v
        if v is None:
            return False
        s = str(v).strip().lower()
        return s in ("1", "true", "yes", "y", "on", "checked")

    def _canon_code(v: str) -> str:
        return _norm_rm_code(v or "")

    def _budget_rate_inr_for_line(l: "RMCBudgetLine") -> Decimal:
        """
        Final budget ₹/Kg:
          - LOCAL: local_rate_inr
          - IMPORT: (usd*ex) + duty + freight + clearance
        """
        p = _norm_pt(getattr(l, "purchase_type", ""))
        local_inr = _d(getattr(l, "local_rate_inr", None), Decimal("0"))
        usd = _d(getattr(l, "import_rate_usd", None), Decimal("0"))
        duty_pct = _d(getattr(l, "duty_percent", None), Decimal("0"))
        freight = _d(getattr(l, "freight_inr", None), Decimal("0"))
        clearance = _d(getattr(l, "clearance_inr", None), Decimal("0"))

        if p == "IMPORT":
            rate_inr = (usd * ex) if (ex and ex > 0 and usd > 0) else Decimal("0")
            custom_duty = (rate_inr * duty_pct / Decimal("100")) if rate_inr > 0 else Decimal("0")
            return _q(rate_inr + custom_duty + freight + clearance, "0.0000")

        return _q(local_inr, "0.0000")

    # --------------------------------------------------------------------------
    # Price maps (RMCBudgetLine → final budget rate ₹/kg)
    # --------------------------------------------------------------------------
    price_by_code_cap = {}  # (code, cap_bool/None) -> Decimal rate
    price_by_name_cap = {}  # (NAME_UPPER, cap_bool/None) -> Decimal rate

    for l in lines:
        cap = _bool(
            getattr(l, "is_captive", False)
            or getattr(l, "captive", False)
            or getattr(l, "captive_tick", False)
        )
        code = _canon_code(getattr(l, "rm_code", "") or "")
        name = (getattr(l, "rm_name", "") or "").strip()
        rate = _budget_rate_inr_for_line(l)

        if code:
            price_by_code_cap[(code, cap)] = rate
            price_by_code_cap.setdefault((code, None), rate)

        if name:
            k = (name.upper(), cap)
            price_by_name_cap[k] = rate
            price_by_name_cap.setdefault((name.upper(), None), rate)

    # --------------------------------------------------------------------------
    # Workbook
    # --------------------------------------------------------------------------
    wb = Workbook()

    # =====================================================================================
    # SHEET 1 (kept): RMC_Budget
    # =====================================================================================
    ws = wb.active
    ws.title = "RMC_Budget"

    ws["E1"] = "$/₹"
    ws["F1"] = float(_q(ex, "0.0000"))
    yellow = PatternFill("solid", fgColor="FFF59D")
    ws["E1"].fill = yellow
    ws["F1"].fill = yellow
    ws["E1"].font = Font(bold=True)
    ws["F1"].font = Font(bold=True)
    ws["E1"].alignment = Alignment(horizontal="center")
    ws["F1"].number_format = "0.0000"

    headers = ["Material Name", "Budgeted Rate/Kg", "Import Rate $", "Rate (₹)", "Custom Duty", "Clearance"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1E3A8A")
    header_font = Font(bold=True, color="FFFFFF")
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=2, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    ws.auto_filter.ref = "A2:F2"
    ws.freeze_panes = "A3"

    row_idx = 3
    import_row_fill = PatternFill("solid", fgColor="E9D5FF")

    for l in lines:
        p = _norm_pt(getattr(l, "purchase_type", ""))
        rm_name = (getattr(l, "rm_name", "") or "").strip()

        local_inr = _d(getattr(l, "local_rate_inr", None), Decimal("0"))
        usd = _d(getattr(l, "import_rate_usd", None), Decimal("0"))
        duty_pct = _d(getattr(l, "duty_percent", None), Decimal("0"))
        freight = _d(getattr(l, "freight_inr", None), Decimal("0"))
        clearance = _d(getattr(l, "clearance_inr", None), Decimal("0"))

        clearance_total = freight + clearance

        if p == "IMPORT":
            rate_inr = (usd * ex) if (ex and ex > 0 and usd > 0) else Decimal("0")
            custom_duty = (rate_inr * duty_pct / Decimal("100")) if rate_inr > 0 else Decimal("0")
            budgeted = rate_inr + custom_duty + clearance_total

            row = [
                rm_name,
                float(_q(budgeted, "0.0000")),
                float(_q(usd, "0.0000")),
                float(_q(rate_inr, "0.0000")),
                float(_q(custom_duty, "0.0000")),
                float(_q(clearance_total, "0.0000")),
            ]
        else:
            budgeted = local_inr
            row = [
                rm_name,
                float(_q(budgeted, "0.0000")),
                None,
                float(_q(local_inr, "0.0000")),
                None,
                None,
            ]

        ws.append(row)

        ws.cell(row=row_idx, column=2).number_format = "0.0000"
        ws.cell(row=row_idx, column=3).number_format = '"$"#,##0.0000'
        ws.cell(row=row_idx, column=4).number_format = "0.0000"
        ws.cell(row=row_idx, column=5).number_format = "0.0000"
        ws.cell(row=row_idx, column=6).number_format = "0.0000"

        if p == "IMPORT":
            for c in range(1, 7):
                ws.cell(row=row_idx, column=c).fill = import_row_fill

        row_idx += 1

    widths = [40, 18, 14, 14, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # =====================================================================================
    # SHEET 2 (kept): RMC_Working
    # =====================================================================================
    ws2 = wb.create_sheet("RMC_Working")
    ws2.append([
        "RM Code", "RM Name", "Purchase Type",
        "Local ₹/Kg",
        "Import $/Kg", "USD/INR", "Rate ₹",
        "Duty %", "Custom Duty ₹",
        "Freight ₹/Kg", "Clearance ₹/Kg",
        "Budget ₹/Kg"
    ])

    for col in range(1, 13):
        c = ws2.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    rr = 2
    for l in lines:
        p = _norm_pt(getattr(l, "purchase_type", ""))

        local_inr = _d(getattr(l, "local_rate_inr", None), Decimal("0"))
        usd = _d(getattr(l, "import_rate_usd", None), Decimal("0"))
        duty_pct = _d(getattr(l, "duty_percent", None), Decimal("0"))
        freight = _d(getattr(l, "freight_inr", None), Decimal("0"))
        clearance = _d(getattr(l, "clearance_inr", None), Decimal("0"))

        rate_inr = (usd * ex) if (p == "IMPORT" and ex and ex > 0 and usd > 0) else Decimal("0")
        custom_duty = (rate_inr * duty_pct / Decimal("100")) if (p == "IMPORT" and rate_inr > 0) else Decimal("0")

        if p == "IMPORT":
            budgeted = rate_inr + custom_duty + freight + clearance
        else:
            budgeted = local_inr

        ws2.append([
            getattr(l, "rm_code", "") or "",
            getattr(l, "rm_name", "") or "",
            p,
            float(_q(local_inr, "0.0000")),
            float(_q(usd, "0.0000")),
            float(_q(ex, "0.0000")),
            float(_q(rate_inr, "0.0000")),
            float(_q(duty_pct, "0.0000")),
            float(_q(custom_duty, "0.0000")),
            float(_q(freight, "0.0000")),
            float(_q(clearance, "0.0000")),
            float(_q(budgeted, "0.0000")),
        ])

        for cidx in range(4, 13):
            ws2.cell(row=rr, column=cidx).number_format = "0.0000"
        rr += 1

    for col in range(1, 13):
        ws2.column_dimensions[get_column_letter(col)].width = 18
    ws2.column_dimensions["B"].width = 40

    # =====================================================================================
    # SHEET 3 (kept): RMC_All_Lines
    # =====================================================================================
    ws3 = wb.create_sheet("RMC_All_Lines")

    month_cols = [m.upper() for m in MONTHS]  # APR..MAR
    ws3_headers = [
        "RM Code", "RM Name", "Unit", "Type",
        *month_cols,
        "Annual Qty", "Required Qty",
        "Purchase Type",
        "Local ₹/Kg", "Import $/Kg", "USD/INR",
        "Base ₹", "Duty %", "Duty ₹",
        "Freight ₹/Kg", "Clearance ₹/Kg",
        "Final Budget ₹/Kg",
        "Annual Value ₹"
    ]
    ws3.append(ws3_headers)

    for col in range(1, len(ws3_headers) + 1):
        c = ws3.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = f"A1:{get_column_letter(len(ws3_headers))}1"

    qty_fmt = "0.000000"
    money_fmt = "0.0000"

    start_data_row = 2
    r = start_data_row

    for l in lines:
        p = _norm_pt(getattr(l, "purchase_type", ""))

        code = (getattr(l, "rm_code", "") or "").strip()
        name = (getattr(l, "rm_name", "") or "").strip()
        unit = (getattr(l, "unit", "") or "").strip()

        is_cap = bool(getattr(l, "is_captive", False) or getattr(l, "captive", False) or getattr(l, "captive_tick", False))
        typ = "CAPTIVE" if is_cap else "RAW"

        mvals = []
        for m in MONTHS:
            mv = _d(getattr(l, f"qty_{m}", None), Decimal("0"))
            mvals.append(_q(mv, "0.000000"))

        annual_qty = _q(sum(mvals, Decimal("0.000000")), "0.000000")

        req_qty = getattr(l, "required_qty", None)
        req_qty = _q(req_qty if req_qty not in (None, "") else annual_qty, "0.000000")

        local_inr = _d(getattr(l, "local_rate_inr", None), Decimal("0"))
        usd = _d(getattr(l, "import_rate_usd", None), Decimal("0"))
        duty_pct = _d(getattr(l, "duty_percent", None), Decimal("0"))
        freight = _d(getattr(l, "freight_inr", None), Decimal("0"))
        clearance = _d(getattr(l, "clearance_inr", None), Decimal("0"))

        if p == "IMPORT":
            base_inr = (usd * ex) if (ex and ex > 0 and usd > 0) else Decimal("0")
            duty_amt = (base_inr * duty_pct / Decimal("100")) if base_inr > 0 else Decimal("0")
            final_budget = base_inr + duty_amt + freight + clearance
        else:
            base_inr = local_inr
            duty_amt = Decimal("0")
            final_budget = local_inr

        base_inr = _q(base_inr, "0.0000")
        duty_amt = _q(duty_amt, "0.0000")
        final_budget = _q(final_budget, "0.0000")

        annual_value = _q(req_qty * final_budget, "0.00")

        row = [
            code, name, unit, typ,
            *[float(x) for x in mvals],
            float(annual_qty), float(req_qty),
            p,
            float(_q(local_inr, "0.0000")),
            float(_q(usd, "0.0000")) if p == "IMPORT" else None,
            float(_q(ex, "0.0000")),
            float(base_inr),
            float(_q(duty_pct, "0.0000")) if p == "IMPORT" else None,
            float(duty_amt) if p == "IMPORT" else None,
            float(_q(freight, "0.0000")) if p == "IMPORT" else None,
            float(_q(clearance, "0.0000")) if p == "IMPORT" else None,
            float(final_budget),
            float(annual_value),
        ]
        ws3.append(row)

        months_start_col = 5
        months_end_col = months_start_col + len(MONTHS) - 1

        for cidx in range(months_start_col, months_end_col + 1):
            ws3.cell(row=r, column=cidx).number_format = qty_fmt

        ws3.cell(row=r, column=months_end_col + 1).number_format = qty_fmt
        ws3.cell(row=r, column=months_end_col + 2).number_format = qty_fmt

        def col_of(h): return ws3_headers.index(h) + 1

        for h in [
            "Local ₹/Kg", "Import $/Kg", "USD/INR", "Base ₹", "Duty %", "Duty ₹",
            "Freight ₹/Kg", "Clearance ₹/Kg", "Final Budget ₹/Kg", "Annual Value ₹"
        ]:
            ws3.cell(row=r, column=col_of(h)).number_format = money_fmt

        if p == "IMPORT":
            for cidx in range(1, len(ws3_headers) + 1):
                ws3.cell(row=r, column=cidx).fill = import_row_fill

        r += 1

    ws3.column_dimensions["A"].width = 14
    ws3.column_dimensions["B"].width = 42
    ws3.column_dimensions["C"].width = 10
    ws3.column_dimensions["D"].width = 10
    for i in range(5, 5 + len(MONTHS)):
        ws3.column_dimensions[get_column_letter(i)].width = 11
    ws3.column_dimensions[get_column_letter(5 + len(MONTHS))].width = 12
    ws3.column_dimensions[get_column_letter(6 + len(MONTHS))].width = 12
    ws3.column_dimensions[get_column_letter(7 + len(MONTHS))].width = 14
    for i in range(8 + len(MONTHS), len(ws3_headers) + 1):
        ws3.column_dimensions[get_column_letter(i)].width = 14

    # =====================================================================================
    # ✅ SHEET 4 (FINAL): FG_COGS (FG from Production Budget + Inputs from BOM)
    # =====================================================================================
    ws4 = wb.create_sheet("FG_COGS")

    ws4["D1"] = None
    ws4["F1"] = None
    ws4["H1"] = None

    ws4["D1"].fill = yellow
    ws4["F1"].fill = yellow
    ws4["H1"].fill = PatternFill("solid", fgColor="93C5FD")

    ws4["D1"].font = Font(bold=True)
    ws4["F1"].font = Font(bold=True)
    ws4["H1"].font = Font(bold=True)

    ws4["D1"].number_format = "0.0000"
    ws4["F1"].number_format = "0.00"
    ws4["H1"].number_format = "#,##0"

    ws4_headers = [
        "FG NAME", "QTY (MT)", "MATERIAL NAME", "NORM",
        "Price/kg", "Cost/kg", "MAT QTY (MT)", "RMC (₹)"
    ]
    ws4.append(ws4_headers)

    for col in range(1, len(ws4_headers) + 1):
        c = ws4.cell(row=2, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    ws4.freeze_panes = "A3"
    ws4.auto_filter.ref = f"A2:{get_column_letter(len(ws4_headers))}2"

    # ---- FG list + qty from ProductionBudgetFG (handles plan/budget__plan) ----
    try:
        pb_qs = ProductionBudgetFG.objects.filter(plan=plan)
    except Exception:
        pb_qs = ProductionBudgetFG.objects.filter(budget__plan=plan)

    try:
        sel = list(getattr(rmc_budget, "selected_fgs", None) or [])
    except Exception:
        sel = []
    sel = [(x or "").strip() for x in sel if (x or "").strip()]
    if sel:
        pb_qs = pb_qs.filter(fg_name__in=sel)

    pb_qs = pb_qs.exclude(fg_name__isnull=True).exclude(fg_name__exact="")

    fg_qty_map = {}
    fg_name_map = {}
    for row in pb_qs.values("fg_name", *MONTHS):
        fg = (row.get("fg_name") or "").strip()
        if not fg:
            continue
        total = Decimal("0.000000")
        for m in MONTHS:
            total += _d(row.get(m, 0), Decimal("0"))
        fg_u = fg.upper()
        fg_qty_map[fg_u] = _q(total, "0.000000")
        fg_name_map[fg_u] = fg

    fg_list_upper = sorted(fg_qty_map.keys())

    # ---- Input field adapters (schema tolerant) ----
    CAPTIVE_FIELDS = ("is_captive", "captive_tick", "captive", "is_captive_tick", "captive_flag")
    CODE_FIELDS = ("material_code", "bom_item_code", "rm_code", "item_code", "code")
    NAME_FIELDS = ("material_name", "bom_item_name", "name")
    NORM_FIELDS = ("budget_norm", "norm")

    def _first_attr(obj, names):
        for n in names:
            if hasattr(obj, n):
                return getattr(obj, n)
        return None

    def _rm_code(obj) -> str:
        v = _first_attr(obj, CODE_FIELDS) or ""
        return _canon_code(str(v))

    def _rm_name(obj) -> str:
        v = _first_attr(obj, NAME_FIELDS) or ""
        return (str(v) or "").strip()

    def _norm_value(obj) -> Decimal:
        v = _first_attr(obj, NORM_FIELDS)
        return _q(_d(v, Decimal("0")), "0.000000")

    def _is_cap(obj) -> bool:
        return _bool(_first_attr(obj, CAPTIVE_FIELDS))

    if not fg_list_upper:
        ws4.append(["(No FG found in Production Budget for this plan)", 0, "", 0, 0, 0, 0, 0])
    else:
        ws4_row = 3
        grand_norm = Decimal("0.0000")
        grand_cost_per_kg = Decimal("0.00")
        grand_rmc = Decimal("0")

        for fg_u in fg_list_upper:
            fg_display = fg_name_map.get(fg_u, fg_u)
            fg_qty = fg_qty_map.get(fg_u, Decimal("0.000000"))

            # ✅ Inputs from BOM
            bom = ProductionBOM.objects.filter(fg_name__iexact=fg_display, is_active=True).first()
            if not bom:
                ws4.append([fg_display, float(_q(fg_qty, "0.000000")), "(No BOM found)", 0, 0, 0, 0, 0])
                ws4_row += 1
                continue

            in_qs = ProductionBOMInputLine.objects.filter(bom=bom).order_by("id")

            # filter valid inputs (norm > 0 + has code or name)
            inputs = []
            for ln in in_qs:
                code = _rm_code(ln)
                name = _rm_name(ln)
                norm = _norm_value(ln)
                if not code and not name:
                    continue
                if norm <= 0:
                    continue
                inputs.append(ln)

            if not inputs:
                ws4.append([fg_display, float(_q(fg_qty, "0.000000")), "(No inputs found)", 0, 0, 0, 0, 0])
                ws4_row += 1
                continue

            fg_norm_total = Decimal("0.0000")
            fg_costkg_total = Decimal("0.00")
            fg_rmc_total = Decimal("0")

            for ln in inputs:
                code = _rm_code(ln)
                mat_name = _rm_name(ln)
                norm = _norm_value(ln)
                is_cap = _is_cap(ln)

                # price lookup: code+cap -> code any -> name+cap -> name any -> 0
                price = Decimal("0.0000")
                if code:
                    price = price_by_code_cap.get((code, is_cap))
                    if price is None:
                        price = price_by_code_cap.get((code, None))
                if (price is None) or (price == 0):
                    k = (mat_name.strip().upper(), is_cap)
                    price = price_by_name_cap.get(k)
                    if price is None:
                        price = price_by_name_cap.get((mat_name.strip().upper(), None))
                price = _q(price or 0, "0.0000")

                mat_qty = _q(fg_qty * norm, "0.000000")
                cost_per_kg = _q(norm * price, "0.00")
                rmc_val = _q(mat_qty * price * Decimal("1000"), "0")

                ws4.append([
                    fg_display,
                    float(_q(fg_qty, "0.000000")),
                    mat_name,
                    float(_q(norm, "0.000000")),
                    float(_q(price, "0.00")),
                    float(_q(cost_per_kg, "0.00")),
                    float(_q(mat_qty, "0.000000")),
                    float(_q(rmc_val, "0")),
                ])

                ws4.cell(row=ws4_row, column=2).number_format = "0.000000"
                ws4.cell(row=ws4_row, column=4).number_format = "0.000000"
                ws4.cell(row=ws4_row, column=5).number_format = "0.00"
                ws4.cell(row=ws4_row, column=6).number_format = "0.00"
                ws4.cell(row=ws4_row, column=7).number_format = "0.000000"
                ws4.cell(row=ws4_row, column=8).number_format = "#,##0"

                fg_norm_total += _q(norm, "0.0000")
                fg_costkg_total += _q(cost_per_kg, "0.00")
                fg_rmc_total += _q(rmc_val, "0")

                ws4_row += 1

            grand_norm += _q(fg_norm_total, "0.0000")
            grand_cost_per_kg += _q(fg_costkg_total, "0.00")
            grand_rmc += _q(fg_rmc_total, "0")

        ws4["D1"] = float(_q(grand_norm, "0.0000"))
        ws4["F1"] = float(_q(grand_cost_per_kg, "0.00"))
        ws4["H1"] = float(_q(grand_rmc, "0"))

    ws4.column_dimensions["A"].width = 28
    ws4.column_dimensions["B"].width = 12
    ws4.column_dimensions["C"].width = 44
    ws4.column_dimensions["D"].width = 10
    ws4.column_dimensions["E"].width = 12
    ws4.column_dimensions["F"].width = 12
    ws4.column_dimensions["G"].width = 14
    ws4.column_dimensions["H"].width = 16

    # ------------------------------------------------------------
    # ✅ SUBTOTAL totals (only visible rows when filtering FG NAME)
    # ------------------------------------------------------------
    last_row = ws4.max_row  # last data row in FG_COGS

    # If there is at least 1 data row (row>=3), apply formulas
    if last_row >= 3:
        # D = NORM, F = Cost/kg, H = RMC (₹)
        ws4["D1"] = f"=SUBTOTAL(9,D3:D{last_row})"
        ws4["F1"] = f"=SUBTOTAL(9,F3:F{last_row})"
        ws4["H1"] = f"=SUBTOTAL(9,H3:H{last_row})"
    else:
        ws4["D1"] = 0
        ws4["F1"] = 0
        ws4["H1"] = 0

    ws4["D1"].number_format = "0.0000"
    ws4["F1"].number_format = "0.00"
    ws4["H1"].number_format = "#,##0"

    # =====================================================================================
    # Response
    # =====================================================================================
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"RMC_Budget_{plan.fy}.xlsx"
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp

# =============================================================================
# GL ACCOUNT VIEWS
# =============================================================================

# Helpers

# ACCOUNTS/Budget/views.py

import pandas as pd
import re
from django.db import transaction
from django.http import JsonResponse
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.shortcuts import redirect, render, get_object_or_404
from django.db.models import Q

from .models import DepartmentBudgetHead, GLAccount
from .forms import GLAccountForm, BudgetHeadUploadForm


def _clean_text(v) -> str:
    s = "" if v is None else str(v)
    s = s.replace("\u00A0", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s


@login_required
def ajax_budget_departments(request):
    unit = (request.GET.get("unit") or "").strip()
    qs = DepartmentBudgetHead.objects.filter(is_active=True)
    if unit:
        qs = qs.filter(unit__iexact=unit)
    depts = list(qs.values_list("department", flat=True).distinct().order_by("department"))
    depts = [d for d in depts if d]
    return JsonResponse({"results": depts})


@login_required
def ajax_budget_units(request):
    dept = (request.GET.get("department") or "").strip()
    qs = DepartmentBudgetHead.objects.filter(is_active=True)
    if dept:
        qs = qs.filter(department__iexact=dept)
    units = list(qs.values_list("unit", flat=True).distinct().order_by("unit"))
    units = [u for u in units if u]
    return JsonResponse({"results": units})


@login_required
def ajax_budget_heads(request):
    unit = (request.GET.get("unit") or "").strip()
    dept = (request.GET.get("department") or "").strip()

    qs = DepartmentBudgetHead.objects.filter(is_active=True)
    if unit:
        qs = qs.filter(unit__iexact=unit)
    if dept:
        qs = qs.filter(department__iexact=dept)

    rows = list(qs.values("budget_head", "gl").order_by("budget_head"))
    return JsonResponse({"results": rows})
# -------------------------------------------------------------------------------
# GL Account Views (Excel-driven)
# -------------------------------------------------------------------------------
from openpyxl import load_workbook
from .forms import GLAccountExcelForm, DepartmentBudgetHeadUploadForm

def _norm(v: str) -> str:
    return re.sub(r"\s+", " ", (v or "").strip())

# ----------------------------
# Ledger list (shows uploaded + manual)
# ----------------------------
def _norm(s):
    return (str(s).strip() if s is not None else "")


@login_required
def gl_account_list(request):
    q = (request.GET.get("q") or "").strip()
    unit = (request.GET.get("unit") or "").strip()
    dept = (request.GET.get("department") or "").strip()
    head = (request.GET.get("budget_head") or "").strip()

    qs = GLAccount.objects.all().order_by("department", "unit", "budget_head", "no")

    if unit:
        qs = qs.filter(unit__iexact=unit)
    if dept:
        qs = qs.filter(department__iexact=dept)
    if head:
        qs = qs.filter(budget_head__iexact=head)

    if q:
        qs = qs.filter(Q(no__icontains=q) | Q(name__icontains=q) | Q(budget_head__icontains=q))

    return render(request, "accounts/budget/gl_account_list.html", {
        "rows": qs[:2000],
        "q": q,
        "unit": unit,
        "department": dept,
        "budget_head": head,
    })

# ----------------------------
# Ledger create (Excel-driven)
# ----------------------------
@login_required
def gl_account_create(request):
    if request.method == "POST":
        form = GLAccountForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.created_by = request.user
            obj.save()
            messages.success(request, "Ledger created.")
            return redirect("accounts_budget:gl_account_list")
        messages.error(request, "Please correct the errors.")
    else:
        form = GLAccountForm()

    units = list(DepartmentBudgetHead.objects.filter(is_active=True).values_list("unit", flat=True).distinct())
    depts = list(DepartmentBudgetHead.objects.filter(is_active=True).values_list("department", flat=True).distinct())
    heads = list(DepartmentBudgetHead.objects.filter(is_active=True).values_list("budget_head", flat=True).distinct())

    return render(request, "accounts/budget/gl_account_form.html", {
        "form": form,
        "mode": "create",
        "units": sorted([u for u in units if u]),
        "departments": sorted([d for d in depts if d]),
        "heads": sorted([h for h in heads if h]),
    })

@login_required
def gl_account_edit(request, pk):
    obj = get_object_or_404(GLAccount, pk=pk)
    if request.method == "POST":
        form = GLAccountForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            messages.success(request, "Ledger updated.")
            return redirect("accounts_budget:gl_account_list")
        messages.error(request, "Please correct the errors.")
    else:
        form = GLAccountForm(instance=obj)
    return render(request, "accounts/budget/gl_account_form.html", {"form": form, "mode": "edit", "obj": obj})

@login_required
@require_POST
def gl_account_delete(request, pk):
    obj = get_object_or_404(GLAccount, pk=pk)

    # Optional: soft delete if you have is_active
    if hasattr(obj, "is_active"):
        obj.is_active = False
        obj.save(update_fields=["is_active"])
        messages.success(request, "Ledger disabled (soft deleted).")
    else:
        obj.delete()
        messages.success(request, "Ledger deleted.")

    return redirect("accounts_budget:gl_account_list")

# ----------------------------
# Ledger edit (Excel-driven)
# ----------------------------
@login_required
def gl_account_edit(request, pk):
    obj = get_object_or_404(GLAccount, pk=pk)
    if request.method == "POST":
        form = GLAccountForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            messages.success(request, "Ledger updated.")
            return redirect("accounts_budget:gl_account_list")
        messages.error(request, "Please correct the errors.")
    else:
        form = GLAccountForm(instance=obj)

    units = list(DepartmentBudgetHead.objects.filter(is_active=True).values_list("unit", flat=True).distinct())
    depts = list(DepartmentBudgetHead.objects.filter(is_active=True).values_list("department", flat=True).distinct())
    heads = list(DepartmentBudgetHead.objects.filter(is_active=True).values_list("budget_head", flat=True).distinct())

    return render(request, "accounts/budget/gl_account_form.html", {
        "form": form,
        "mode": "edit",
        "obj": obj,
        "units": sorted([u for u in units if u]),
        "departments": sorted([d for d in depts if d]),
        "heads": sorted([h for h in heads if h]),
    })


# ----------------------------
# Upload Excel -> DepartmentBudgetHead + auto create/update GLAccount
# ----------------------------
@login_required
def gl_excel_upload(request):
    """
    Upload Excel -> fill DepartmentBudgetHead master
    and auto-create/update GLAccount for rows having GL name.
    """
    if request.method == "POST":
        form = GLExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            f = form.cleaned_data["file"]

            wb = openpyxl.load_workbook(f, data_only=True)
            if "Budget-heads" not in wb.sheetnames:
                messages.error(request, "Sheet 'Budget-heads' not found.")
                return redirect("accounts_budget:gl_excel_upload")

            ws = wb["Budget-heads"]

            # header row expected: Unit | Department | Budget Head | GL
            created_master = 0
            updated_master = 0
            created_ledgers = 0
            updated_ledgers = 0

            with transaction.atomic():
                for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    unit = _norm(row[0])
                    dept = _norm(row[1])
                    head = _norm(row[2])
                    gl = _norm(row[3])

                    if not (unit or dept or head or gl):
                        continue

                    unit = unit.strip()
                    dept = dept.strip()
                    head = head.strip()
                    gl = gl.strip()

                    if not (dept and head):
                        # ignore incomplete lines
                        continue

                    # 1) upsert master
                    m, is_new = DepartmentBudgetHead.objects.update_or_create(
                        unit=unit,
                        department=dept,
                        budget_head=head,
                        defaults={
                            "gl_name": gl,
                            "is_active": True,
                            "updated_by": request.user,
                            "created_by": request.user if True else None,
                        }
                    )
                    if is_new:
                        created_master += 1
                    else:
                        updated_master += 1

                    # 2) upsert ledger only if GL name present
                    if gl:
                        led = GLAccount.objects.filter(
                            unit__iexact=unit,
                            department__iexact=dept,
                            budget_head__iexact=head,
                        ).first()

                        if led:
                            # update name if changed
                            if (led.name or "").strip() != gl:
                                led.name = gl
                                led.save()
                                updated_ledgers += 1
                        else:
                            GLAccount.objects.create(
                                no="",  # force auto
                                name=gl,
                                unit=unit,
                                department=dept,
                                budget_head=head,
                                blocked=False,
                                is_active=True,
                                created_by=request.user,
                            )
                            created_ledgers += 1

            messages.success(
                request,
                f"Upload done. Master: +{created_master}/~{updated_master} | Ledgers: +{created_ledgers}/~{updated_ledgers}"
            )
            return redirect("accounts_budget:gl_account_list")
        messages.error(request, "Please choose a valid Excel file.")
    else:
        form = GLExcelUploadForm()

    return render(request, "accounts/budget/gl_excel_upload.html", {"form": form})

def _s(v):
    return (str(v).strip() if v is not None else "")

@login_required
def dept_budget_head_upload(request):
    if request.method == "POST":
        f = request.FILES.get("file")
        if not f:
            messages.error(request, "Please select an Excel file.")
            return redirect("accounts_budget:budget_heads_upload")

        try:
            wb = openpyxl.load_workbook(f, data_only=True)
        except Exception as e:
            messages.error(request, f"Invalid Excel file: {e}")
            return redirect("accounts_budget:budget_heads_upload")

        if "Budget-heads" not in wb.sheetnames:
            messages.error(request, "Sheet 'Budget-heads' not found.")
            return redirect("accounts_budget:budget_heads_upload")

        ws = wb["Budget-heads"]

        # Expect header row: Unit, Department, Budget Head, GL
        headers = [(_s(ws.cell(1, c).value)) for c in range(1, 5)]
        expected = ["Unit", "Department", "Budget Head", "GL"]
        if headers != expected:
            messages.error(request, f"Excel headers must be: {expected}. Found: {headers}")
            return redirect("accounts_budget:budget_heads_upload")

        created_master = updated_master = 0
        created_gl = updated_gl = 0

        with transaction.atomic():
            for r in range(2, ws.max_row + 1):
                unit = _s(ws.cell(r, 1).value)
                dept = _s(ws.cell(r, 2).value)
                head = _s(ws.cell(r, 3).value)
                glname = _s(ws.cell(r, 4).value)

                if not (dept and head):
                    continue

                # 1) Upsert master
                obj, created = DepartmentBudgetHead.objects.update_or_create(
                    unit=unit,
                    department=dept,
                    budget_head=head,
                    defaults={
                        "gl": glname,
                        "is_active": True,
                        "updated_by": request.user,
                        "created_by": request.user if created else None,
                    },
                )
                if created:
                    created_master += 1
                else:
                    updated_master += 1

                # 2) Upsert GLAccount (this is why your "ledger list" was empty)
                gl_obj, gl_created = GLAccount.objects.update_or_create(
                    unit=unit,
                    department=dept,
                    budget_head=head,
                    name=glname or head,   # fallback: if GL blank, use head
                    defaults={
                        "is_active": True,
                        "created_by": request.user if gl_created else None,
                    },
                )
                if gl_created:
                    created_gl += 1
                else:
                    updated_gl += 1

        messages.success(
            request,
            f"Upload complete. Master: +{created_master} / upd {updated_master}. "
            f"Ledgers: +{created_gl} / upd {updated_gl}."
        )
        return redirect("accounts_budget:gl_account_list")

    return render(request, "accounts/budget/budget_heads_upload.html")

@login_required
def budget_heads_json(request):
    """
    Optional helper for keyword search / dependent dropdowns.
    Query params: unit, department, q
    """
    unit = (request.GET.get("unit") or "").strip()
    dept = (request.GET.get("department") or "").strip()
    q = (request.GET.get("q") or "").strip()

    qs = DepartmentBudgetHead.objects.filter(is_active=True)
    if unit:
        qs = qs.filter(unit__iexact=unit)
    if dept:
        qs = qs.filter(department__iexact=dept)
    if q:
        qs = qs.filter(budget_head__icontains=q)

    qs = qs.order_by("budget_head")[:200]
    return JsonResponse({
        "results": [
            {"id": x.id, "budget_head": x.budget_head, "gl_name": (x.gl_name or "")}
            for x in qs
        ]
    })

# ----------------------------
# AJAX: Units (from DepartmentBudgetHead)
# ----------------------------
@login_required
def ajax_units(request):
    units = (
        DepartmentBudgetHead.objects
        .filter(is_active=True)
        .exclude(unit__isnull=True)
        .exclude(unit__exact="")
        .values_list("unit", flat=True)
        .distinct()
        .order_by("unit")
    )
    return JsonResponse({"results": list(units)})


# ----------------------------
# AJAX: Departments (by unit)
# ----------------------------
@login_required
def ajax_departments(request):
    unit = _norm(request.GET.get("unit") or "")

    qs = DepartmentBudgetHead.objects.filter(is_active=True)
    if unit:
        qs = qs.filter(unit=unit)

    depts = (
        qs.exclude(department__isnull=True)
          .exclude(department__exact="")
          .values_list("department", flat=True)
          .distinct()
          .order_by("department")
    )
    return JsonResponse({"results": list(depts)})


# ----------------------------
# AJAX: Budget Heads (by unit + department)
# ----------------------------
# @login_required
# def ajax_budget_heads(request):
#     unit = _norm(request.GET.get("unit") or "")
#     dept = _norm(request.GET.get("department") or "")

#     qs = DepartmentBudgetHead.objects.filter(is_active=True)
#     if unit:
#         qs = qs.filter(unit=unit)
#     if dept:
#         qs = qs.filter(department=dept)

#     heads = (
#         qs.exclude(budget_head__isnull=True)
#           .exclude(budget_head__exact="")
#           .values_list("budget_head", flat=True)
#           .distinct()
#           .order_by("budget_head")
#     )
#     return JsonResponse({"results": list(heads)})


# ----------------------------
# AJAX: GL name (for selected unit+dept+head)
# ----------------------------
@login_required
def ajax_gl_name(request):
    unit = _norm(request.GET.get("unit") or "")
    dept = _norm(request.GET.get("department") or "")
    head = _norm(request.GET.get("budget_head") or "")

    m = DepartmentBudgetHead.objects.filter(
        unit=unit, department=dept, budget_head=head, is_active=True
    ).first()

    return JsonResponse({"gl_name": (m.gl_name if m else "")})


# ----------------------------
# TomSelect search for ledgers (GENERAL)
# Use for other screens needing quick ledger lookup by code/name.
# ----------------------------
def _dept_variants(dept: str) -> list[str]:
    """
    Makes department matching tolerant:
    QA/QC  <->  QA& QC  <->  QA & QC  <->  QA&QC
    """
    d = (dept or "").strip()
    if not d:
        return []

    variants = {d}

    # swap separators
    variants.add(d.replace("/", "&"))
    variants.add(d.replace("&", "/"))

    # add common spacing around &
    variants.add(d.replace("/", "& ").replace("&", "& "))
    variants.add(d.replace("/", " & ").replace("&", " & "))
    variants.add(d.replace("&", "& "))
    variants.add(d.replace("&", " & "))

    # normalize multiple spaces
    out = set()
    for v in variants:
        out.add(re.sub(r"\s+", " ", v).strip())

    return list(out)


@login_required
def gl_search(request):
    # TomSelect sends AJAX GET
    q = (request.GET.get("q") or "").strip()
    dept = (request.GET.get("department") or "").strip()

    qs = GLAccount.objects.all()

    # Department-wise filter
    if dept:
        dvs = _dept_variants(dept)
        if dvs:
            dept_q = reduce(or_, [Q(department__iexact=v) for v in dvs])
            qs = qs.filter(dept_q)

    # Search primarily by Budget Head (you can keep name/no too, but label remains Budget Head)
    if q:
        qs = qs.filter(
            Q(budget_head__icontains=q) |
            Q(name__icontains=q) |
            Q(no__icontains=q)
        )

    qs = qs.order_by("budget_head", "no")[:50]

    results = []
    for obj in qs:
        label = (obj.budget_head or "").strip() or (obj.name or "").strip() or str(obj.no)
        results.append({"id": str(obj.pk), "text": label})

    return JsonResponse({"results": results})

# =============================================================================
# Captive Consumption Budget Views
# =============================================================================

from decimal import Decimal, ROUND_HALF_UP
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.shortcuts import redirect, render

from .models import (
    BudgetPlan,
    CaptiveConsumptionBudget,
    CaptiveConsumptionLine,
    ProductionBudgetFG,   # ✅ for FG dropdown suggestions
)
from .forms import CaptiveConsumptionBudgetForm, CaptiveConsumptionLineFormSet
from .services.maker_checker import mc_get, mc_submit
from .models import MCStatus


def _is_locked(mc) -> bool:
    return mc and (mc.status in (MCStatus.SUBMITTED, MCStatus.APPROVED))


def _is_approver(user) -> bool:
    return user.has_perm("Budget.can_approve_budgets")


def _q(v: Decimal, places: str):
    """
    Quantize safely with explicit rounding.
    """
    q = Decimal(places)
    x = v if isinstance(v, Decimal) else Decimal(str(v or "0"))
    return x.quantize(q, rounding=ROUND_HALF_UP)


def _item_choices_for_plan(plan):
    """
    Captive selection must include:
      - Work in Progress
      - Semi Finished Good
      - WIP FR
      - Finished Goods (from ProductionBudgetFG / SalesBudgetLine)
    Returns sorted distinct names.
    """
    CAPTIVE_TYPES = ("Work in Progress", "Semi Finished Good", "WIP FR")
    FG_TYPES = ("Finished Good", "Finished Goods", "FG")  # safe aliases (if present in ERP rows)

    names = set()

    # 1) Finished goods from your own budgets (most reliable for FG list)
    try:
        names.update(
            ProductionBudgetFG.objects
            .filter(plan=plan)
            .exclude(fg_name__isnull=True)
            .exclude(fg_name__exact="")
            .values_list("fg_name", flat=True)
        )
    except Exception:
        pass

    # 2) Finished goods from Sales budget lines (fallback)
    try:
        from .models import SalesBudgetLine
        names.update(
            SalesBudgetLine.objects
            .filter(budget__plan=plan)
            .exclude(product_name__isnull=True)
            .exclude(product_name__exact="")
            .values_list("product_name", flat=True)
        )
    except Exception:
        pass

    # 3) WIP / SFG / WIP FR products from ERP BOM cache (distinct FG names)
    # NOTE: This assumes your ERPBOMRow.type carries these values for the FG rows you want discoverable.
    # If your ERP stores FG type differently, tell me the column and I will switch the filter accordingly.
    try:
        from .models import ERPBOMRow
        names.update(
            ERPBOMRow.objects
            .filter(type__in=CAPTIVE_TYPES + FG_TYPES)
            .exclude(fg_name__isnull=True)
            .exclude(fg_name__exact="")
            .values_list("fg_name", flat=True)
            .distinct()
        )
    except Exception:
        pass

    # clean + sort
    cleaned = sorted({(n or "").strip() for n in names if (n or "").strip()})
    return cleaned

@login_required
def captive_consumption_home(request):
    plan = (
        BudgetPlan.objects.filter(is_active=True)
        .order_by("-updated_at", "-id")
        .first()
    )
    if not plan:
        messages.error(request, "No active Budget Plan found.")
        return redirect("accounts_budget:budget_home")

    budget, _ = CaptiveConsumptionBudget.objects.get_or_create(
        plan=plan,
        defaults={"created_by": request.user}
    )

    mc = mc_get(budget, "CAPTIVE")
    locked = _is_locked(mc)

    lines = list(budget.lines.all().order_by("item_name"))

    # ✅ compute amount in view if column not present/filled
    total_amt = sum(
        ((_q((l.qty or Decimal("0")) * (l.rate or Decimal("0")), "0.01")) for l in lines),
        Decimal("0.00")
    )

    return render(request, "accounts/budget/captive_consumption_home.html", {
        "plan": plan,
        "budget": budget,
        "lines": lines,
        "total_amt": total_amt,
        "mc": mc,
        "is_locked": locked,
        "is_approver": _is_approver(request.user),
        "mc_scope": "CAPTIVE",
        "mc_model": "accounts_budget.CaptiveConsumptionBudget",
        "mc_pk": budget.id,
    })


@login_required
def captive_consumption_edit(request):
    plan = (
        BudgetPlan.objects.filter(is_active=True)
        .order_by("-updated_at", "-id")
        .first()
    )
    if not plan:
        messages.error(request, "No active Budget Plan found.")
        return redirect("accounts_budget:budget_home")

    budget, _ = CaptiveConsumptionBudget.objects.get_or_create(
        plan=plan,
        defaults={"created_by": request.user}
    )

    scope = "CAPTIVE"
    mc = mc_get(budget, scope)
    locked = _is_locked(mc)

    if request.method == "POST" and locked:
        messages.warning(request, "This entry is locked. It can be edited only if disapproved by approver.")
        return redirect("accounts_budget:captive_consumption_home")

    qs = CaptiveConsumptionLine.objects.filter(budget=budget).order_by("id")

    # ✅ FG suggestions for datalist (keyword search input)
    item_choices = _item_choices_for_plan(plan)

    if request.method == "POST":
        action = (request.POST.get("action") or "save").strip().lower()

        form = CaptiveConsumptionBudgetForm(request.POST, instance=budget)
        formset = CaptiveConsumptionLineFormSet(request.POST, instance=budget, queryset=qs)

        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                obj = form.save(commit=False)
                obj.updated_by = request.user
                obj.save()

                # save valid rows
                lines = formset.save(commit=False)

                for ln in lines:
                    ln.budget = budget

                    # ✅ normalize
                    ln.item_name = (getattr(ln, "item_name", "") or "").strip()

                    # ✅ qty/rate quantize
                    ln.qty = _q(getattr(ln, "qty", None) or Decimal("0"), "0.000")
                    ln.rate = _q(getattr(ln, "rate", None) or Decimal("0"), "0.0000")

                    # ✅ amount: if model has amount field, set it; else ignore
                    if hasattr(ln, "amount"):
                        ln.amount = _q((ln.qty * ln.rate), "0.01")

                    ln.save()

                # delete removed
                for d in formset.deleted_objects:
                    d.delete()

                if action == "submit":
                    mc_submit(budget, scope, request.user)
                    messages.success(request, "Captive Consumption saved and submitted for approval.")
                    return redirect("accounts_budget:captive_consumption_home")

            messages.success(request, "Captive Consumption saved.")
            return redirect("accounts_budget:captive_consumption_edit")

        messages.error(request, "Please correct the errors.")
    else:
        form = CaptiveConsumptionBudgetForm(instance=budget)
        formset = CaptiveConsumptionLineFormSet(instance=budget, queryset=qs)

    if locked:
        for f in formset.forms:
            for name in f.fields:
                f.fields[name].disabled = True
            if "DELETE" in f.fields:
                f.fields["DELETE"].disabled = True

    return render(request, "accounts/budget/captive_consumption_edit.html", {
        "plan": plan,
        "budget": budget,
        "form": form,
        "formset": formset,

        # ✅ required for datalist in template
        "item_choices": item_choices,

        "mc": mc,
        "is_locked": locked,
        "is_approver": _is_approver(request.user),
        "mc_scope": scope,
        "mc_model": "accounts_budget.CaptiveConsumptionBudget",
        "mc_pk": budget.id,
    })





# =============================================================================
# PACKING MATERIAL BUDGET (NEW)
# =============================================================================

PACK_MONTHS = ["apr","may","jun","jul","aug","sep","oct","nov","dec","jan","feb","mar"]

PACK_REQ_FIELD = {
    "apr": "req_apr", "may": "req_may", "jun": "req_jun", "jul": "req_jul",
    "aug": "req_aug", "sep": "req_sep", "oct": "req_oct", "nov": "req_nov",
    "dec": "req_dec", "jan": "req_jan", "feb": "req_feb", "mar": "req_mar",
}

# ✅ NEW: monthwise value fields (₹)
PACK_VAL_FIELD = {
    "apr": "val_apr", "may": "val_may", "jun": "val_jun", "jul": "val_jul",
    "aug": "val_aug", "sep": "val_sep", "oct": "val_oct", "nov": "val_nov",
    "dec": "val_dec", "jan": "val_jan", "feb": "val_feb", "mar": "val_mar",
}

DEC0_MT = Decimal("0.000")
DEC0_REQ = Decimal("0.000000")
DEC0_AMT = Decimal("0.00")


def _d(v, default=Decimal("0")):
    """Robust Decimal parser."""
    if v in (None, "", "None"):
        return default
    if isinstance(v, Decimal):
        return v
    try:
        return Decimal(str(v).strip().replace(",", ""))
    except (InvalidOperation, ValueError, TypeError):
        return default


def _money(v, default=DEC0_AMT):
    """₹ quantize to 2 decimals."""
    x = _d(v, Decimal("0"))
    try:
        return x.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except Exception:
        return default


def _sales_qty_map_for_products(plan: BudgetPlan, products: list[str]):
    """
    Return dict: product_upper -> {apr..mar: Decimal(MT), annual_mt: Decimal}.

    - Sums across all SaleTypes for the product.
    - If SalesBudget doesn't exist: returns zeros.
    """
    out = {}
    prods = [p.strip() for p in (products or []) if (p or "").strip()]
    if not prods:
        return out

    prods_u = {p.upper() for p in prods}
    zeros = {m: DEC0_MT for m in PACK_MONTHS}

    sb = getattr(plan, "sales_budget", None)
    if not sb:
        for p in prods_u:
            out[p] = {**zeros, "annual_mt": DEC0_MT}
        return out

    qs = (
        SalesBudgetLine.objects
        .filter(budget=sb)
        .filter(Q(product_name__in=prods) | Q(product_name__in=list(prods_u)))
        .order_by()
        .values("product_name")
        .annotate(
            apr=Sum("apr"), may=Sum("may"), jun=Sum("jun"), jul=Sum("jul"),
            aug=Sum("aug"), sep=Sum("sep"), oct=Sum("oct"), nov=Sum("nov"),
            dec=Sum("dec"), jan=Sum("jan"), feb=Sum("feb"), mar=Sum("mar"),
            annual=Sum("annual_qty_mt"),
        )
    )

    for p in prods_u:
        out[p] = {**zeros, "annual_mt": DEC0_MT}

    for r in qs:
        pn = (r.get("product_name") or "").strip().upper()
        if not pn:
            continue
        row = {m: (_d(r.get(m), DEC0_MT) if r.get(m) is not None else DEC0_MT) for m in PACK_MONTHS}
        annual = _d(r.get("annual"), sum(row.values(), DEC0_MT))
        row["annual_mt"] = annual
        out[pn] = row

    return out


def _packing_items_from_erp(limit: int = 5000):
    """
    Return distinct packing items from ERPBOMRow (Budget_erpbomrow).

    Uses ERPBOMRow.type containing 'Packing'/'Pack'/'PM'.
    """
    typ_f = "type"
    code_f = "bom_item_code"
    name_f = "bom_item_name"
    unit_f = "unit"

    qs = (
        ERPBOMRow.objects
        .filter(
            Q(**{f"{typ_f}__icontains": "PACK"}) |
            Q(**{f"{typ_f}__icontains": "PACKING"}) |
            Q(**{f"{typ_f}__icontains": "PM"})
        )
        .exclude(**{f"{code_f}__isnull": True})
        .exclude(**{f"{code_f}__exact": ""})
        .exclude(**{f"{name_f}__isnull": True})
        .exclude(**{f"{name_f}__exact": ""})
        .values(code_f)
        .annotate(
            item_name=Min(name_f),
            unit=Min(unit_f),
        )
        .order_by(code_f)[:limit]
    )

    rows = []
    for r in qs:
        rows.append({
            "item_code": (r.get(code_f) or "").strip().upper(),
            "item_name": (r.get("item_name") or "").strip(),
            "unit": (r.get("unit") or "").strip(),
        })
    return rows


def _sync_packing_master_from_erp(user=None) -> dict:
    """Create missing PackingMaterialMaster rows from ERPBOMRow (does not overwrite size/rate)."""
    src = _packing_items_from_erp()
    if not src:
        return {"created": 0, "updated": 0, "seen": 0}

    codes = [r["item_code"] for r in src if r.get("item_code")]
    existing = set(
        PackingMaterialMaster.objects
        .filter(item_code__in=codes)
        .values_list("item_code", flat=True)
    )

    created = 0
    updated = 0
    with transaction.atomic():
        for r in src:
            code = (r.get("item_code") or "").strip().upper()
            if not code:
                continue

            if code in existing:
                # refresh name/unit safely; do NOT touch packing_size/rate
                PackingMaterialMaster.objects.filter(item_code=code).update(
                    item_name=r.get("item_name") or code,
                    unit=r.get("unit") or "",
                    updated_by=user,
                )
                updated += 1
                continue

            PackingMaterialMaster.objects.create(
                item_code=code,
                item_name=r.get("item_name") or code,
                unit=r.get("unit") or "",
                packing_size=None,
                rate=None,          # ✅ NEW: keep empty by default
                is_active=True,
                updated_by=user,
            )
            created += 1

    return {"created": created, "updated": updated, "seen": len(src)}


def _compute_packing_requirements(*, sales_mt_row: dict, pack_size, wastage_pct: Decimal) -> dict:
    """Compute monthwise requirement counts (No. of packs)."""
    out = {m: DEC0_REQ for m in PACK_MONTHS}
    ps = _d(pack_size, Decimal("0"))
    if not ps or ps <= 0:
        return out

    factor = (Decimal("1") + (_d(wastage_pct, Decimal("0")) / Decimal("100")))

    for m in PACK_MONTHS:
        mt = _d(sales_mt_row.get(m), DEC0_MT)
        kg = (mt * Decimal("1000"))
        req = (kg * factor / ps).quantize(Decimal("0.000000"), rounding=ROUND_HALF_UP)
        out[m] = req

    return out


@login_required
def packing_budget_home(request):
    return redirect("accounts_budget:packing_products")


@require_http_methods(["GET", "POST"])
@login_required
def packing_products(request):
    """Page-1: FG list from ProductionNorm + show Sales Qty, select products."""

    plan = BudgetPlan.objects.filter(is_active=True).order_by("-updated_at", "-id").first()
    if not plan:
        messages.error(request, "No active Budget Plan found.")
        return redirect("accounts_budget:budget_home")

    budget, _ = PackingBudget.objects.get_or_create(plan=plan, defaults={"created_by": request.user})

    scope = "PACKING"
    mc = mc_get(budget, scope)
    locked = _is_locked(mc)

    if request.method == "POST" and locked:
        messages.warning(request, "Packing Budget is locked. It can be edited only if disapproved by approver.")
        return redirect("accounts_budget:packing_products")

    norm_products = list(
        ProductionNorm.objects
        .exclude(fg_name__isnull=True)
        .exclude(fg_name__exact="")
        .values_list("fg_name", flat=True)
        .distinct()
        .order_by("fg_name")
    )

    sb = getattr(plan, "sales_budget", None)
    sales_products = []
    if sb:
        sales_products = list(
            SalesBudgetLine.objects
            .filter(budget=sb)
            .exclude(product_name__isnull=True)
            .exclude(product_name__exact="")
            .values_list("product_name", flat=True)
            .distinct()
            .order_by("product_name")
        )

    seen = set()
    products = []
    for p in norm_products + sales_products:
        p2 = (p or "").strip()
        if not p2:
            continue
        u = p2.upper()
        if u in seen:
            continue
        seen.add(u)
        products.append(p2)

    sales_map = _sales_qty_map_for_products(plan, products)
    selected_set = {x.upper() for x in budget.get_selected_products()}

    if request.method == "POST":
        action = (request.POST.get("action") or "save").strip().lower()

        if action == "sync_master":
            res = _sync_packing_master_from_erp(request.user)
            messages.success(request, f"Packing master synced. Added: {res['created']} items (updated: {res['updated']}).")
            return redirect("accounts_budget:packing_products")

        selected = request.POST.getlist("selected_products")
        selected_clean = [s.strip() for s in selected if (s or "").strip()]

        w = _d(request.POST.get("wastage_percent_default"), budget.wastage_percent_default).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )

        with transaction.atomic():
            budget.set_selected_products(selected_clean)
            budget.wastage_percent_default = w
            budget.save()

            if action == "submit":
                mc_submit(budget, scope, request.user)
                messages.success(request, "Packing Budget submitted for approval.")
            else:
                _mc_mark_draft(budget, scope, request.user)
                messages.success(request, "Packing Budget saved as Draft.")

        if action in ("next", "inputs"):
            return redirect("accounts_budget:packing_inputs")
        return redirect("accounts_budget:packing_products")

    rows = []
    for p in products:
        pu = p.upper()
        sm = sales_map.get(pu) or {m: DEC0_MT for m in PACK_MONTHS}
        annual_mt = sm.get("annual_mt") or sum((sm.get(m) or DEC0_MT for m in PACK_MONTHS), DEC0_MT)
        annual_kg = (annual_mt * Decimal("1000")).quantize(Decimal("0.001"), rounding=ROUND_HALF_UP)
        rows.append({
            "product": p,
            "selected": (pu in selected_set),
            "annual_mt": annual_mt,
            "annual_kg": annual_kg,
        })

    return render(request, "accounts/budget/packing_budget_products.html", {
        "plan": plan,
        "packing_budget": budget,
        "rows": rows,
        "mc": mc,
        "is_locked": locked,
        "is_approver": _is_approver(request.user),
        "mc_scope": scope,
        "mc_model": "accounts_budget.PackingBudget",
        "mc_pk": budget.id,
    })


@require_http_methods(["GET", "POST"])
@login_required
def packing_inputs(request):
    """Page-2: assign packing materials per selected product; compute monthwise requirements + values."""

    plan = BudgetPlan.objects.filter(is_active=True).order_by("-updated_at", "-id").first()
    if not plan:
        messages.error(request, "No active Budget Plan found.")
        return redirect("accounts_budget:budget_home")

    budget, _ = PackingBudget.objects.get_or_create(plan=plan, defaults={"created_by": request.user})

    scope = "PACKING"
    mc = mc_get(budget, scope)
    locked = _is_locked(mc)

    selected = budget.get_selected_products()
    if not selected:
        messages.warning(request, "Select products first.")
        return redirect("accounts_budget:packing_products")

    if not PackingMaterialMaster.objects.exists():
        _sync_packing_master_from_erp(request.user)

    pack_choices = list(
        PackingMaterialMaster.objects
        .filter(is_active=True)
        .order_by("item_name", "item_code")
        .only("id", "item_code", "item_name", "unit", "packing_size", "rate")  # ✅ include rate
    )

    sales_map = _sales_qty_map_for_products(plan, selected)

    selected_norm = [p.strip() for p in selected if (p or "").strip()]
    selected_u = [p.upper() for p in selected_norm]

    existing_lines = list(
        PackingBudgetLine.objects
        .filter(budget=budget)
        .filter(Q(product_name__in=selected_norm) | Q(product_name__in=selected_u))
        .select_related("packing_material")
        .order_by("product_name", "packing_name", "packing_code")
    )

    if request.method == "POST" and locked:
        messages.warning(request, "Packing Budget is locked. It can be edited only if disapproved by approver.")
        return redirect("accounts_budget:packing_inputs")

    if request.method == "POST":
        action = (request.POST.get("action") or "save").strip().lower()

        try:
            line_count = int(request.POST.get("line_count") or "0")
        except Exception:
            line_count = 0

        with transaction.atomic():
            for i in range(line_count):
                lid = (request.POST.get(f"line_id__{i}") or "").strip()
                pid = (request.POST.get(f"packing_material__{i}") or "").strip()
                prod = (request.POST.get(f"product__{i}") or "").strip()
                del_flag = (request.POST.get(f"delete__{i}") or "").strip() == "1"

                if not prod:
                    continue

                if del_flag and lid:
                    try:
                        PackingBudgetLine.objects.filter(id=int(lid), budget=budget).delete()
                    except Exception:
                        pass
                    continue

                if not pid:
                    continue

                try:
                    pm = PackingMaterialMaster.objects.get(id=int(pid), is_active=True)
                except Exception:
                    continue

                wastage = _d(request.POST.get(f"wastage__{i}"), budget.wastage_percent_default).quantize(
                    Decimal("0.01"), rounding=ROUND_HALF_UP
                )

                sm = sales_map.get(prod.upper()) or {m: DEC0_MT for m in PACK_MONTHS}

                reqs = _compute_packing_requirements(
                    sales_mt_row=sm,
                    pack_size=pm.packing_size,
                    wastage_pct=wastage,
                )

                # ✅ Save snapshot fields too (packing_size + rate) so values compute correctly
                obj, _ = PackingBudgetLine.objects.update_or_create(
                    budget=budget,
                    product_name=prod,
                    packing_material=pm,
                    defaults={
                        "wastage_percent": wastage,
                        "packing_size": pm.packing_size,
                        "rate": pm.rate,
                        **{PACK_REQ_FIELD[m]: reqs[m] for m in PACK_MONTHS},
                    },
                )
                obj.save()  # model save will compute req_total + val fields

            if action == "submit":
                mc_submit(budget, scope, request.user)
                messages.success(request, "Packing Budget submitted for approval.")
            else:
                _mc_mark_draft(budget, scope, request.user)
                messages.success(request, "Packing Budget saved as Draft.")

        if action in ("summary", "next"):
            return redirect("accounts_budget:packing_summary")
        return redirect("accounts_budget:packing_inputs")

    rows = []
    for ln in existing_lines:
        prod = (ln.product_name or "").strip()
        sm = sales_map.get(prod.upper()) or {m: DEC0_MT for m in PACK_MONTHS}
        annual_mt = sm.get("annual_mt") or sum((sm.get(m) or DEC0_MT for m in PACK_MONTHS), DEC0_MT)
        annual_kg = (annual_mt * Decimal("1000")).quantize(Decimal("0.001"), rounding=ROUND_HALF_UP)

        rows.append({
            "id": ln.id,
            "product": prod,
            "annual_mt": annual_mt,
            "annual_kg": annual_kg,
            "packing_material_id": ln.packing_material_id,
            "packing_size": ln.packing_size,
            "rate": ln.rate,               # ✅ show on form if needed
            "wastage": ln.wastage_percent,
            "req_total": ln.req_total,
            "val_total": getattr(ln, "val_total", DEC0_AMT),  # ✅
        })

    for prod in selected_norm:
        sm = sales_map.get(prod.upper()) or {m: DEC0_MT for m in PACK_MONTHS}
        annual_mt = sm.get("annual_mt") or sum((sm.get(m) or DEC0_MT for m in PACK_MONTHS), DEC0_MT)
        annual_kg = (annual_mt * Decimal("1000")).quantize(Decimal("0.001"), rounding=ROUND_HALF_UP)

        rows.append({
            "id": "",
            "product": prod,
            "annual_mt": annual_mt,
            "annual_kg": annual_kg,
            "packing_material_id": "",
            "packing_size": "",
            "rate": "",
            "wastage": budget.wastage_percent_default,
            "req_total": DEC0_REQ,
            "val_total": DEC0_AMT,
        })

    return render(request, "accounts/budget/packing_budget_inputs.html", {
        "plan": plan,
        "packing_budget": budget,
        "rows": rows,
        "pack_choices": pack_choices,
        "line_count": len(rows),
        "mc": mc,
        "is_locked": locked,
        "is_approver": _is_approver(request.user),
        "mc_scope": scope,
        "mc_model": "accounts_budget.PackingBudget",
        "mc_pk": budget.id,
    })


@login_required
def packing_summary(request):
    """Page-3: summary aggregated by packing material (qty + value)."""

    plan = BudgetPlan.objects.filter(is_active=True).order_by("-updated_at", "-id").first()
    if not plan:
        messages.error(request, "No active Budget Plan found.")
        return redirect("accounts_budget:budget_home")

    budget = getattr(plan, "packing_budget", None)
    if not budget:
        messages.warning(request, "Packing Budget not created yet.")
        return redirect("accounts_budget:packing_products")

    lines = list(
        PackingBudgetLine.objects
        .filter(budget=budget)
        .select_related("packing_material")
        .order_by("packing_name", "packing_code", "product_name")
    )

    agg = {}
    for ln in lines:
        key = (ln.packing_code, ln.packing_name, ln.unit, str(ln.packing_size or ""), str(ln.rate or ""))
        if key not in agg:
            agg[key] = {m: DEC0_REQ for m in PACK_MONTHS}
            agg[key]["total"] = DEC0_REQ
            agg[key]["val_total"] = DEC0_AMT
            agg[key]["packing_size"] = ln.packing_size
            agg[key]["rate"] = ln.rate
            agg[key]["packing_code"] = ln.packing_code
            agg[key]["packing_name"] = ln.packing_name
            agg[key]["unit"] = ln.unit

        for m in PACK_MONTHS:
            agg[key][m] += (_d(getattr(ln, PACK_REQ_FIELD[m]), DEC0_REQ))
        agg[key]["total"] += (_d(ln.req_total, DEC0_REQ))
        agg[key]["val_total"] = _money(_d(agg[key]["val_total"], Decimal("0")) + _d(getattr(ln, "val_total", DEC0_AMT), Decimal("0")))

    summary_rows = list(agg.values())
    summary_rows.sort(key=lambda r: ((r.get("packing_name") or "").lower(), (r.get("packing_code") or "")))

    grand_qty = DEC0_REQ
    grand_val = DEC0_AMT
    for r in summary_rows:
        grand_qty += _d(r.get("total"), DEC0_REQ)
        grand_val = _money(_d(grand_val, Decimal("0")) + _d(r.get("val_total"), Decimal("0")))

    return render(request, "accounts/budget/packing_budget_summary.html", {
        "plan": plan,
        "packing_budget": budget,
        "lines": lines,
        "summary_rows": summary_rows,
        "grand_qty": grand_qty,
        "grand_val": grand_val,
    })


# -------------------------------------------------------------------
# Packing Material Master (Upload / CRUD) - Excel import updated for rate/size if present
# -------------------------------------------------------------------

def _clean_excel_text(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    s = re.sub(r"\.0$", "", s)
    return s


def import_packing_master_from_excel(file_obj, *, user=None) -> dict:
    """
    Create ONLY missing PackingMaterialMaster rows from uploaded .xlsx.
    Does NOT overwrite existing rows.

    Minimum columns:
      - Short Name  -> item_code
      - Name        -> item_name
      - Stock Keeping Unit -> unit (optional)

    Optional columns (if present):
      - Packing Size / Pack Size
      - Rate
    """
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb.active

    header = [(_clean_excel_text(c.value)).lower() for c in ws[1]]

    def find_col(*names):
        for n in names:
            n = n.lower()
            if n in header:
                return header.index(n) + 1
        return None

    col_code = find_col("short name", "item code", "code")
    col_name = find_col("name", "item name", "description")
    col_unit = find_col("stock keeping unit", "unit", "uom")

    col_size = find_col("packing size", "pack size", "size", "packing_size")
    col_rate = find_col("rate", "std rate", "purchase rate")

    if not col_code or not col_name:
        raise ValueError("Excel format invalid. Required columns: Short Name, Name (optional: Stock Keeping Unit).")

    created = 0
    skipped_existing = 0
    skipped_blank = 0

    existing = set(PackingMaterialMaster.objects.values_list("item_code", flat=True))
    to_create = []

    for r in range(2, ws.max_row + 1):
        code = _clean_excel_text(ws.cell(row=r, column=col_code).value).upper()
        name = _clean_excel_text(ws.cell(row=r, column=col_name).value)
        unit = _clean_excel_text(ws.cell(row=r, column=col_unit).value) if col_unit else ""

        if not code or not name:
            skipped_blank += 1
            continue
        if code in existing:
            skipped_existing += 1
            continue

        size = None
        rate = None
        if col_size:
            try:
                size = _d(ws.cell(row=r, column=col_size).value, None)
            except Exception:
                size = None
        if col_rate:
            try:
                rate = _d(ws.cell(row=r, column=col_rate).value, None)
            except Exception:
                rate = None

        to_create.append(PackingMaterialMaster(
            item_code=code,
            item_name=name,
            unit=unit,
            packing_size=size,
            rate=rate,          # ✅ NEW
            is_active=True,
            updated_by=user,
        ))
        existing.add(code)

    with transaction.atomic():
        if to_create:
            PackingMaterialMaster.objects.bulk_create(to_create, ignore_conflicts=True)
            created = len(to_create)

    return {
        "created": created,
        "skipped_existing": skipped_existing,
        "skipped_blank": skipped_blank,
        "total_rows": ws.max_row - 1,
    }

@require_http_methods(["GET", "POST"])
@login_required
def packing_material_master_list(request):
    q = (request.GET.get("q") or "").strip()
    upload_form = PackingMaterialUploadForm()

    if request.method == "POST":
        action = (request.POST.get("action") or "").strip().lower()

        if action == "sync":
            res = _sync_packing_master_from_erp(request.user)
            messages.success(request, f"Synced from ERP. Added: {res['created']}, Updated: {res['updated']}.")
            return redirect("accounts_budget:packing_master_list")

        if action == "upload":
            upload_form = PackingMaterialUploadForm(request.POST, request.FILES)
            if not upload_form.is_valid():
                messages.error(request, "Please select a valid .xlsx file.")
                return redirect("accounts_budget:packing_master_list")

            f = upload_form.cleaned_data["file"]
            try:
                res = import_packing_master_from_excel(f, user=request.user)
                messages.success(
                    request,
                    f"Upload done. Added: {res['created']}, Existing skipped: {res['skipped_existing']}, "
                    f"Blank skipped: {res['skipped_blank']} (Rows checked: {res['total_rows']})."
                )
            except Exception as e:
                messages.error(request, f"Upload failed: {e}")
            return redirect("accounts_budget:packing_master_list")

    qs = PackingMaterialMaster.objects.all().order_by("item_name", "item_code")
    if q:
        qs = qs.filter(Q(item_code__icontains=q) | Q(item_name__icontains=q))

    return render(request, "accounts/budget/packing_material_master_list.html", {
        "q": q,
        "rows": qs[:2000],
        "upload_form": upload_form,
    })

@require_http_methods(["GET", "POST"])
@login_required
def packing_material_master_create(request):
    if request.method == "POST":
        form = PackingMaterialMasterForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.updated_by = request.user
            obj.save()
            messages.success(request, "Packing material created.")
            return redirect("accounts_budget:packing_master_list")
        messages.error(request, "Please correct the errors.")
    else:
        form = PackingMaterialMasterForm(initial={"is_active": True})

    return render(request, "accounts/budget/packing_material_master_form.html", {
        "form": form,
        "is_new": True,
    })

@require_http_methods(["GET", "POST"])
@login_required
def packing_material_master_edit(request, pk: int):
    obj = get_object_or_404(PackingMaterialMaster, pk=pk)

    if request.method == "POST":
        form = PackingMaterialMasterForm(request.POST, instance=obj)
        if form.is_valid():
            o = form.save(commit=False)
            o.updated_by = request.user
            o.save()
            messages.success(request, "Packing material updated.")
            return redirect("accounts_budget:packing_master_list")
        messages.error(request, "Please correct the errors.")
    else:
        form = PackingMaterialMasterForm(instance=obj)

    return render(request, "accounts/budget/packing_material_master_form.html", {
        "form": form,
        "is_new": False,
        "obj": obj,
    })

@login_required
def packing_material_master_json(request):
    """AJAX helper: return pack size + rate by master id."""
    pid = (request.GET.get("id") or "").strip()
    if not pid:
        return JsonResponse({"ok": False, "error": "Missing id"}, status=400)

    try:
        pm = PackingMaterialMaster.objects.get(id=int(pid))
    except Exception:
        return JsonResponse({"ok": False, "error": "Not found"}, status=404)

    return JsonResponse({
        "ok": True,
        "id": pm.id,
        "item_code": pm.item_code,
        "item_name": pm.item_name,
        "unit": pm.unit,
        "packing_size": (str(pm.packing_size) if pm.packing_size is not None else ""),
        "rate": (str(pm.rate) if pm.rate is not None else ""),   # ✅ NEW
    })