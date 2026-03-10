# ----------------------------
# Standard library
# ----------------------------
import json
import re
import threading
import uuid
import base64
from collections import defaultdict
from datetime import date, datetime, timedelta
from decimal import Decimal
from functools import reduce
from itertools import islice
from django.db.models import Sum, Max
from django.views.generic import DeleteView
from django.utils.dateparse import parse_date
from openpyxl.cell import WriteOnlyCell
from typing import List
from typing import Any
from decimal import Decimal, InvalidOperation
from typing import Any, Iterable, Set, Tuple, Optional

# ----------------------------
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ----------------------------
# Django
# ----------------------------
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.cache import cache
from django.core.management import call_command
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import DecimalField, ExpressionWrapper, F, Q, Sum
from django.db.models.functions import TruncMonth
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy
from django.utils import timezone
from django.views.decorators.http import require_GET, require_POST
from django.views.generic import CreateView, ListView, UpdateView
from urllib3 import request
from ACCOUNTS.Receivable.services.receivables_dashboard import BASE_START_DATE
from ACCOUNTS.Receivable.services.company_groups import COMPANY_GROUPS

from django.db.models import Q

# ----------------------------
# Third-party
# ----------------------------
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# ----------------------------
# App imports
# ----------------------------
from ACCOUNTS.Receivable.services.receivables_sync import latest_snapshot_date
from ACCOUNTS.Receivable.services.receivables_dashboard import get_received_rows_for_snapshot
from ACCOUNTS.Receivable.services.receivables_sync_state import get_state, set_state

from .forms import ReceivableForm, PaymentTargetSelectPartyForm, PaymentTargetWeekForm
from .models import Receivable, ReceivableSnapshotRow, PaymentTargetLine, PaymentTargetWeek

from .services import _parse_ui_date
from .services.receivables_dashboard import build_receivable_dashboard_context
from .services.receivables_weekly_fast import (
    build_paid_lookup_for_period,
    get_received_rows_for_period,
    get_received_totals_for_period,
)
from .services.receivables_helpers import get_receivable_entries_for_period
from .models import Receivable, ReceivableSnapshotRow, ReceivableOutstandingRemark


import logging
logger = logging.getLogger("custom_logger")


# NOTE:
# - Do NOT import get_receivable_entries_for_period if Targets "open bills" come from snapshot model.
# - Keep receivables_helpers import only if used elsewhere in this views.py.
# from .services.receivables_helpers import get_receivable_entries_for_period


# -------------------------------------------------------------------
# Receivable Create/Update View helpers
# -------------------------------------------------------------------

def _norm_group(v) -> str:
    s = str(v or "").strip()
    if not s:
        return "ALL"
    sl = s.lower()
    if sl in ("all", "all groups", "all group", "all companies", "all company"):
        return "ALL"
    if sl.startswith("all "):
        return "ALL"
    return s


def _has_field(model, field_name: str) -> bool:
    try:
        model._meta.get_field(field_name)
        return True
    except Exception:
        return False

def _safe_model_fields_dict(model_cls, data: dict, allowed_fields: List[str]) -> dict:
    """
    Return only those keys from `data` that truly exist as fields on `model_cls`.
    Prevents TypeError on create(**kwargs).
    """
    try:
        model_field_names = {f.name for f in model_cls._meta.get_fields()}
    except Exception:
        model_field_names = set()

    out = {}
    for k in allowed_fields:
        if k in data and k in model_field_names:
            out[k] = data.get(k)
    return out

def _to_decimal(val, default=Decimal("0")):
    try:
        s = str(val or "").replace(",", "").strip()
        if s == "":
            return default
        return Decimal(s)
    except Exception:
        return default
    
def _parse_ddmmyyyy(val):
    s = (str(val or "")).strip()
    if not s:
        return None
    for fmt in ("%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def _getlist_any(post, *names):
    """Accept both invoice_number and invoice_number[] patterns."""
    for n in names:
        v = post.getlist(n)
        if v:
            return v
    return []

def extract_invoice_rows(post):
    """
    Supports:
      A) invoice_number (repeated) / invoice_number[]
      B) invoice_number_0, invoice_number_1 ... (indexed)
    Returns list of rows with invoice_number, invoice_date, due_date, invoice_amount, received_amount
    Blank invoice_number rows are skipped.
    """

    # A) Repeated / array-style
    inv_numbers = [x.strip() for x in _getlist_any(post, "invoice_number", "invoice_number[]") if (x or "").strip()]
    if len(inv_numbers) > 0:
        inv_dates   = _getlist_any(post, "invoice_date", "invoice_date[]")
        due_dates   = _getlist_any(post, "due_date", "due_date[]")
        inv_amounts = _getlist_any(post, "invoice_amount", "invoice_amount[]")
        rcv_amounts = _getlist_any(post, "received_amount", "received_amount[]")

        rows = []
        for i, inv_no in enumerate(inv_numbers):
            rows.append({
                "invoice_number": inv_no,
                "invoice_date": _parse_ddmmyyyy(inv_dates[i]) if i < len(inv_dates) else None,
                "due_date": _parse_ddmmyyyy(due_dates[i]) if i < len(due_dates) else None,
                "invoice_amount": _to_decimal(inv_amounts[i], default=Decimal("0")) if i < len(inv_amounts) else Decimal("0"),
                "received_amount": _to_decimal(rcv_amounts[i], default=Decimal("0")) if i < len(rcv_amounts) else Decimal("0"),
            })
        return rows

    # B) Indexed style invoice_number_0, invoice_number_1...
    idxs = set()
    for k in post.keys():
        m = re.match(r"^invoice_number_(\d+)$", k)
        if m:
            idxs.add(int(m.group(1)))

    if not idxs:
        return []

    rows = []
    for i in sorted(idxs):
        inv_no = (post.get(f"invoice_number_{i}") or "").strip()
        if not inv_no:
            continue  # ignore blank rows (your screenshot has blanks)

        rows.append({
            "invoice_number": inv_no,
            "invoice_date": _parse_ddmmyyyy(post.get(f"invoice_date_{i}") or ""),
            "due_date": _parse_ddmmyyyy(post.get(f"due_date_{i}") or ""),
            "invoice_amount": _to_decimal(post.get(f"invoice_amount_{i}") or "", default=Decimal("0")),
            "received_amount": _to_decimal(post.get(f"received_amount_{i}") or "", default=Decimal("0")),
        })
    return rows

def _apply_receivable_filters(qs, params):
    """
    Single source of truth for list + excel filters.
    Keeps existing behavior, but prevents mismatch drift.
    """
    customer = (params.get("customer") or "").strip()
    status = (params.get("status") or "").strip()
    typ = (params.get("type") or "").strip()
    company_group = (params.get("company_group") or "").strip()

    from_ui = (params.get("from") or "").strip()
    to_ui = (params.get("to") or "").strip()

    from_dt = parse_date(from_ui) if from_ui else None
    to_dt = parse_date(to_ui) if to_ui else None

    if customer:
        qs = qs.filter(customer_name__icontains=customer)
    if status:
        qs = qs.filter(status=status)
    if typ:
        qs = qs.filter(type=typ)
    if company_group:
        qs = qs.filter(company_group=company_group)

    # Entry Date filtering (inclusive)
    if from_dt and to_dt:
        qs = qs.filter(entry_date__range=(from_dt, to_dt))
    elif from_dt:
        qs = qs.filter(entry_date__gte=from_dt)
    elif to_dt:
        qs = qs.filter(entry_date__lte=to_dt)

    return qs

class ReceivableCreateView(LoginRequiredMixin, CreateView):
    model = Receivable
    form_class = ReceivableForm
    template_name = "accounts/receivable_form.html"
    success_url = reverse_lazy("accounts:receivable_list")

    def form_valid(self, form):
        rows = extract_invoice_rows(self.request.POST)

        # ✅ Entry Date (different from invoice date) - default today
        entry_date = form.cleaned_data.get("entry_date") or timezone.localdate()

        # Single row -> normal CreateView
        if len(rows) <= 1:
            form.instance.created_by = self.request.user
            form.instance.entry_date = entry_date  # ✅ NEW
            return super().form_valid(form)

        cd = form.cleaned_data

        common = {
            "customer_code": (cd.get("customer_code") or "").strip(),
            "customer_name": (cd.get("customer_name") or "").strip(),
            "entry_date": entry_date,  # ✅ NEW
            "currency": cd.get("currency") or "INR",
            "cheque_no": cd.get("cheque_no") or None,
            "cheque_date": cd.get("cheque_date") or None,
            "remarks": cd.get("remarks") or "",
            "created_by": self.request.user,
        }

        # ✅ Optional common fields (only if model has them)
        common.update(_safe_model_fields_dict(
            Receivable,
            {
                "type": cd.get("type"),
                "company_group": cd.get("company_group"),
                "status": cd.get("status"),
                "narration": cd.get("narration"),
            },
            ["type", "company_group", "status", "narration"]
        ))

        # Prevent duplicates in same submit
        seen = set()
        cleaned_rows = []
        for r in rows:
            inv_no = (r.get("invoice_number") or "").strip()
            if not inv_no:
                continue
            key = (
                common["customer_code"],
                inv_no.upper(),
                common["cheque_no"] or "",
                str(common["cheque_date"] or "")
            )
            if key in seen:
                continue
            seen.add(key)
            cleaned_rows.append(r)

        # Ensure required dates exist; skip incomplete blank rows safely
        cleaned_rows = [r for r in cleaned_rows if r.get("invoice_date") and r.get("due_date")]
        if not cleaned_rows:
            messages.error(self.request, "No valid invoice rows found to save.")
            return redirect(self.success_url)

        with transaction.atomic():
            for r in cleaned_rows:
                Receivable.objects.create(
                    **common,
                    invoice_number=(r.get("invoice_number") or "").strip(),
                    invoice_date=r["invoice_date"],
                    due_date=r["due_date"],
                    invoice_amount=r.get("invoice_amount"),
                    received_amount=r.get("received_amount"),
                )

        messages.success(self.request, f"Receivable saved with {len(cleaned_rows)} invoice line(s).")
        return redirect(self.success_url)
# -------------------------------------------------------------------
# Receivable List View
# -------------------------------------------------------------------
# -----------------------------------------------------------------------------
# Normalizers / Parsers
# -----------------------------------------------------------------------------

def _norm(s: Any) -> str:
    return str(s or "").strip().upper()


def _norm_party(s: Any) -> str:
    return _norm(s)

import re
from typing import Any, Iterable, Set

_INV_CLEAN_RE = re.compile(r"[^A-Z0-9]+")

def _norm_inv(v: Any) -> str:
    s = str(v or "").strip().upper()
    if not s:
        return ""
    return _INV_CLEAN_RE.sub("", s)

def _iter_raw_values(v: Any):
    """Yield all primitive values from nested dict/list raw payloads."""
    if isinstance(v, dict):
        for vv in v.values():
            yield from _iter_raw_values(vv)
    elif isinstance(v, (list, tuple, set)):
        for vv in v:
            yield from _iter_raw_values(vv)
    else:
        yield v

def _tokenize_invoice_text(s: Any) -> Set[str]:
    """
    Pull probable invoice/doc numbers from a text.
    We intentionally ignore short purely-numeric tokens (often cheque/instrument).
    """
    txt = str(s or "").strip()
    if not txt:
        return set()

    for sep in [",", ";", "|", "\n", "\r", "\t"]:
        txt = txt.replace(sep, " ")

    out: Set[str] = set()
    for part in txt.split():
        n = _norm_inv(part)
        if not n:
            continue

        # ignore instrument-like short numbers (e.g., 000070)
        if n.isdigit() and len(n) < 8:
            continue

        # invoices/docs are usually long (CMU...); keep conservative
        if len(n) >= 10:
            out.add(n)

    return out

def _erp_receipt_invoice_norms(raw: Any) -> Set[str]:
    """
    Extract invoice/doc refs from ERP receipt raw payload.
    Checks keys that commonly carry invoice/doc numbers.
    """
    if not isinstance(raw, dict):
        return set()

    key_hints = ("doc", "reference", "ref", "invoice", "bill", "against", "settle", "adjust")
    out: Set[str] = set()

    for k, v in raw.items():
        ks = str(k or "").strip().casefold()
        if not any(h in ks for h in key_hints):
            continue
        for vv in _iter_raw_values(v):
            out |= _tokenize_invoice_text(vv)

    return out

def _norm_inst(s: Any) -> str:
    """
    Normalize instrument/cheque reference for reliable matching.
    - removes leading '#'
    - removes spaces
    - keeps alnum only
    - uppercases
    """
    raw = _norm(s)
    raw = raw.replace("#", "").replace(" ", "")
    raw = "".join(ch for ch in raw if ch.isalnum())
    return raw


def _norm_inst_variants(s: Any) -> Set[str]:
    """
    Return a set of possible normalized values (handle leading zeros).
    """
    n = _norm_inst(s)
    out: Set[str] = set()
    if n:
        out.add(n)
        if n.isdigit():
            out.add(n.lstrip("0") or "0")
    return out


def _inst_db_variants(v: Any) -> Set[str]:
    """
    Build variants that can match what is stored in ReceivableSnapshotRow.instrument_no.
    Includes raw/base/#/clean + leading-zero variants.
    """
    raw = str(v or "").strip()
    if not raw:
        return set()

    base = raw.lstrip("#").strip()
    clean = re.sub(r"[^A-Z0-9]", "", base.upper())

    out = {raw, base, f"#{base}", clean, f"#{clean}"}

    if clean.isdigit():
        nz = clean.lstrip("0") or "0"
        out.add(nz)
        out.add(f"#{nz}")

    return {x for x in out if x}


def _to_dec(v: Any) -> Decimal:
    """
    Robust decimal parser for ERP snapshot values.
    Handles:
      - None / '' / 'NA'
      - commas
      - scientific strings like '0E-11'
    """
    if v in (None, "", "NA", "N/A"):
        return Decimal("0")
    try:
        s = str(v).strip().replace(",", "")
        if not s:
            return Decimal("0")
        return Decimal(s)
    except (InvalidOperation, ValueError, TypeError):
        return Decimal("0")


def _pick_any(d: dict, keys: Iterable[str], default: Any = "") -> Any:
    """
    Case-insensitive key pick for dict where ERP keys vary in case/spaces.
    """
    if not isinstance(d, dict):
        return default

    keymap = {str(k).strip().casefold(): k for k in d.keys()}

    for k in keys:
        if not k:
            continue
        kk = str(k).strip().casefold()
        real_key = keymap.get(kk)
        if real_key is None:
            continue
        val = d.get(real_key)
        if val not in (None, "", "NA", "N/A"):
            return val
    return default


def _erp_paid_from_row(srow) -> Decimal:
    """
    Paid amount as per ERP snapshot row:
    1) use srow.paid_amt if non-zero
    2) else try common raw keys (Paid/Receipt/Credit/Unadjustment etc.)
    """
    paid = _to_dec(getattr(srow, "paid_amt", 0))
    if paid != 0:
        return paid

    raw = getattr(srow, "raw", {}) or {}
    raw_val = _pick_any(
        raw,
        [
            "Paid Amt", "Paid Amount",
            "Receipt Amount", "Receipt Amt",
            "Amount", "Amount (Rs.)",
            "Cr Amt", "Credit Amt", "Credit Amount",
            "Unadjustment Amt", "Unadjusted Amt",
        ],
        default="0",
    )
    return _to_dec(raw_val)


# -----------------------------------------------------------------------------
# Financial Year helpers (India FY: Apr 1 - Mar 31)
# -----------------------------------------------------------------------------

def _fy_bounds(d: Optional[date]) -> Tuple[Optional[date], Optional[date]]:
    if not d:
        return (None, None)

    if d.month >= 4:
        start = date(d.year, 4, 1)
        end = date(d.year + 1, 3, 31)
    else:
        start = date(d.year - 1, 4, 1)
        end = date(d.year, 3, 31)

    return start, end


def _in_fy(d: Optional[date], fy_start: Optional[date], fy_end: Optional[date]) -> bool:
    """
    If FY bounds are known, enforce them.
    If unknown, allow (keeps backward compatibility).
    """
    if not (d and fy_start and fy_end):
        return True
    return fy_start <= d <= fy_end

def _alloc_received_across_lines(total_received, lines, amount_attr="invoice_amount"):
    """
    Returns list[Decimal] allocations aligned with `lines`.

    Allocation is sequential:
      alloc[i] = min(line_amount, remaining_total_received)
      remaining decreases after each line.

    Robustness improvements (no flow change):
    - Accepts total_received as None/str/int/float/Decimal.
    - Accepts line amount as None/str/int/float/Decimal.
    - Treats invalid/negative values as 0.
    - Quantizes to 2 decimals (financial safe) while preserving total integrity.
    """
    def _to_dec(v) -> Decimal:
        if v is None:
            return Decimal("0")
        if isinstance(v, Decimal):
            d = v
        else:
            s = str(v).replace(",", "").strip()
            if s in ("", "NA", "N/A", "None", "null"):
                return Decimal("0")
            try:
                d = Decimal(s)
            except (InvalidOperation, ValueError, TypeError):
                return Decimal("0")
        return d if d > 0 else Decimal("0")

    remaining = _to_dec(total_received)
    allocs = []

    for ln in lines:
        amt = _to_dec(getattr(ln, amount_attr, None))
        if remaining <= 0 or amt <= 0:
            allocs.append(Decimal("0.00"))
            continue

        a = amt if remaining >= amt else remaining

        # keep to 2 decimal places (money)
        a = a.quantize(Decimal("0.01"))
        allocs.append(a)

        remaining = (remaining - a)
        if remaining < 0:
            remaining = Decimal("0.00")

    return allocs

# -----------------------------------------------------------------------------
# Receivable List View
# -----------------------------------------------------------------------------

class ReceivableListView(LoginRequiredMixin, ListView):
    model = Receivable
    template_name = "accounts/receivable_list.html"
    context_object_name = "receivables"
    paginate_by = 50

    def get_queryset(self):
        qs = Receivable.objects.all()

        customer = (self.request.GET.get("customer") or "").strip()
        status = (self.request.GET.get("status") or "").strip()
        typ = (self.request.GET.get("type") or "").strip()
        company_group = (self.request.GET.get("company_group") or "").strip()

        from_ui = (self.request.GET.get("from") or "").strip()
        to_ui = (self.request.GET.get("to") or "").strip()

        from_dt = parse_date(from_ui) if from_ui else None
        to_dt = parse_date(to_ui) if to_ui else None

        if customer:
            qs = qs.filter(customer_name__icontains=customer)
        if status:
            qs = qs.filter(status=status)
        if typ:
            qs = qs.filter(type=typ)
        if company_group:
            qs = qs.filter(company_group=company_group)

        if from_dt and to_dt:
            qs = qs.filter(entry_date__range=(from_dt, to_dt))
        elif from_dt:
            qs = qs.filter(entry_date__gte=from_dt)
        elif to_dt:
            qs = qs.filter(entry_date__lte=to_dt)

        return qs.order_by("-entry_date", "-id")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        page_rows = list(ctx.get(self.context_object_name) or [])
        if not page_rows:
            return ctx

        # ---------------------------------------------------------------------
        # A) Split application received invoice-wise (display only)
        # ---------------------------------------------------------------------
        def _local_inst(r) -> str:
            return (
                (getattr(r, "instrument_no", "") or "").strip()
                or (getattr(r, "cheque_no", "") or "").strip()
            )

        def _dec(v) -> Decimal:
            try:
                return _to_dec(v)
            except Exception:
                try:
                    return Decimal(str(v or "0"))
                except Exception:
                    return Decimal("0")

        groups = defaultdict(list)
        for r in page_rows:
            party = _norm_party((getattr(r, "customer_code", "") or "").strip())
            inst = _local_inst(r)
            inst_key = (inst or "").strip().upper()
            if not party or not inst_key:
                continue
            groups[(party, inst_key)].append(r)

        for (_party, _inst_key), rows in groups.items():
            if len(rows) <= 1:
                # still set display_balance_amount for consistency
                x = rows[0]
                inv_amt = _dec(getattr(x, "invoice_amount", 0))
                rec_amt = _dec(getattr(x, "received_amount", 0))
                x.display_balance_amount = (inv_amt - rec_amt)
                continue

            total_received = max((_dec(getattr(x, "received_amount", 0)) for x in rows), default=Decimal("0"))

            # ✅ If nothing received, just show full invoice as balance (DISPLAY ONLY)
            if total_received <= 0:
                for x in rows:
                    inv_amt = _dec(getattr(x, "invoice_amount", 0))
                    x.display_balance_amount = inv_amt     # ✅ no assignment to balance_amount
                    x.is_received_split = True
                    x.received_amount_original = getattr(x, "received_amount", None)
                    x.received_amount = Decimal("0")
                continue

            remaining = total_received
            for x in rows:
                inv_amt = _dec(getattr(x, "invoice_amount", 0))
                alloc = Decimal("0") if inv_amt <= 0 else (inv_amt if remaining >= inv_amt else remaining)
                remaining -= alloc

                x.received_amount_original = getattr(x, "received_amount", None)
                x.is_received_split = True

                # DISPLAY override only (not saved)
                x.received_amount = alloc
                x.display_balance_amount = (inv_amt - alloc)

                # ---------------------------------------------------------------------
        # B) ERP enrichment logic (FIXED: instrument+invoice+party safe)
        # ---------------------------------------------------------------------

        for r in page_rows:
            r.erp_paid_amount = Decimal("0.00")
            r.erp_last_receipt_date = ""
            r.erp_instrument_no = ""
            r.erp_receipt_no = ""
            r.erp_paid_is_fallback = False

        snap = (
            ReceivableSnapshotRow.objects.order_by("-snapshot_date")
            .values_list("snapshot_date", flat=True)
            .first()
        )
        if not snap:
            return ctx

        wanted_db_values: Set[str] = set()
        wanted_parties: Set[str] = set()

        for r in page_rows:
            inst_local = (getattr(r, "instrument_no", "") or "").strip() or (getattr(r, "cheque_no", "") or "").strip()
            wanted_db_values |= _inst_db_variants(inst_local)

            party = (getattr(r, "customer_code", "") or "").strip()
            if party:
                wanted_parties.add(_norm_party(party))

        if not wanted_db_values:
            return ctx

        q = Q()
        for v in wanted_db_values:
            q |= Q(instrument_no__iexact=v)

        if not q:
            return ctx

        erp_qs = (
            ReceivableSnapshotRow.objects
            .filter(snapshot_date=snap)
            .filter(q)
            .only(
                "party_code", "instrument_no",
                "paid_amt", "trans_date", "trans_date_display",
                "trans_no", "raw",
            )
        )

        # IMPORTANT: __in is case-sensitive; keep as-is for flow,
        # but we will also invoice-check, so wrong-party rows won't attach.
        if wanted_parties:
            erp_qs = erp_qs.filter(party_code__in=list(wanted_parties))

        lookup_party_inst = {}        # (party, inst_key) -> agg
        lookup_inst_only = {}         # inst_key -> list[item]

        lookup_party_inst_inv = {}    # (party, inst_key, inv_norm) -> agg
        lookup_inst_inv_only = {}     # (inst_key, inv_norm) -> list[item]

        def _merge_into(cur: dict, add: dict):
            cur["paid"] = (cur.get("paid") or Decimal("0")) + (add.get("paid") or Decimal("0"))

            add_td = add.get("trans_date")
            cur_td = cur.get("trans_date")
            if add_td and (not cur_td or add_td > cur_td):
                cur["trans_date"] = add_td
                cur["last_date"] = add.get("last_date", "") or cur.get("last_date", "")

            if add.get("receipt_no") and not cur.get("receipt_no"):
                cur["receipt_no"] = add["receipt_no"]

            if add.get("inst_raw") and not cur.get("inst_raw"):
                cur["inst_raw"] = add["inst_raw"]

            # union invoice refs
            cur.setdefault("inv_norms", set())
            cur["inv_norms"] |= (add.get("inv_norms") or set())

        for srow in erp_qs.iterator(chunk_size=2000):
            party = _norm_party(getattr(srow, "party_code", ""))
            inst_raw = (getattr(srow, "instrument_no", "") or "").strip()
            inst_keys = _norm_inst_variants(inst_raw)
            if not inst_keys:
                continue

            raw = getattr(srow, "raw", {}) or {}
            inv_norms = _erp_receipt_invoice_norms(raw)  # ✅ raw is defined before use

            receipt_no = _pick_any(
                raw,
                ["Trans No", "Receipt No", "ReceiptNo", "Voucher No", "VoucherNo", "Vch No", "VchNo", "Number"],
                default=""
            ) or (getattr(srow, "trans_no", "") or "").strip()

            paid = _erp_paid_from_row(srow)

            if paid == 0 and not receipt_no:
                continue

            td = getattr(srow, "trans_date", None)
            td_date = td.date() if hasattr(td, "date") else td

            date_disp = (getattr(srow, "trans_date_display", "") or "").strip()
            if not date_disp and td_date:
                date_disp = td_date.strftime("%d-%b-%Y")

            item = {
                "paid": paid,
                "last_date": date_disp,
                "trans_date": td_date,
                "receipt_no": receipt_no,
                "inst_raw": inst_raw,
                "party": party,
                "inv_norms": inv_norms,  # ✅ store invoice refs
            }

            for inst_key in inst_keys:
                lookup_inst_only.setdefault(inst_key, []).append(item)

                # Strong index: instrument + invoice
                if inv_norms:
                    for invn in inv_norms:
                        lookup_inst_inv_only.setdefault((inst_key, invn), []).append(item)
                        if party:
                            k3 = (party, inst_key, invn)
                            cur3 = lookup_party_inst_inv.get(k3)
                            if not cur3:
                                lookup_party_inst_inv[k3] = dict(item)
                            else:
                                _merge_into(cur3, item)

                # Standard: party + instrument
                if party:
                    k = (party, inst_key)
                    cur = lookup_party_inst.get(k)
                    if not cur:
                        lookup_party_inst[k] = dict(item)
                    else:
                        _merge_into(cur, item)

        # Attach to page objects with FY + party + invoice enforcement
        for r in page_rows:
            party = _norm_party(getattr(r, "customer_code", "") or "")
            inst_local = (getattr(r, "instrument_no", "") or "").strip() or (getattr(r, "cheque_no", "") or "").strip()
            local_keys = _norm_inst_variants(inst_local)
            inv_norm = _norm_inv(getattr(r, "invoice_number", "") or "")

            if not local_keys:
                continue

            base_dt = getattr(r, "invoice_date", None) or getattr(r, "entry_date", None)
            base_date = base_dt if isinstance(base_dt, date) else (base_dt.date() if base_dt and hasattr(base_dt, "date") else None)

            fy_start, fy_end = _fy_bounds(base_date) if base_date else (None, None)

            hit = None

            # 0) ✅ strongest: party + instrument + invoice + FY
            if party and inv_norm:
                for k in local_keys:
                    cand = lookup_party_inst_inv.get((party, k, inv_norm))
                    if cand and _in_fy(cand.get("trans_date"), fy_start, fy_end):
                        hit = cand
                        break

            # 1) party + instrument + FY (ONLY if invoice not contradicted)
            if not hit and party:
                for k in local_keys:
                    cand = lookup_party_inst.get((party, k))
                    if not cand or not _in_fy(cand.get("trans_date"), fy_start, fy_end):
                        continue

                    cand_invs = cand.get("inv_norms") or set()
                    # If ERP receipt has invoice refs, require match
                    if cand_invs and inv_norm and (inv_norm not in cand_invs):
                        continue

                    hit = cand
                    break

            # 1.5) instrument + invoice fallback (ONLY if unique party within FY)
            if not hit and inv_norm:
                for k in local_keys:
                    cand_list = lookup_inst_inv_only.get((k, inv_norm)) or []
                    if not cand_list:
                        continue

                    cand_fy = [c for c in cand_list if _in_fy(c.get("trans_date"), fy_start, fy_end)]
                    if not cand_fy:
                        continue

                    party_set = {c.get("party") for c in cand_fy if c.get("party")}
                    # if current party known, it must be that party; and must be unique
                    if party and (party not in party_set):
                        continue
                    if len(party_set) != 1:
                        continue  # ambiguous instrument+invoice across parties

                    agg = None
                    for it in cand_fy:
                        if agg is None:
                            agg = dict(it)
                        else:
                            _merge_into(agg, it)

                    if agg:
                        hit = agg
                        break

            # 2) instrument-only fallback ONLY if resolves to EXACTLY ONE party within FY
            if not hit:
                for k in local_keys:
                    cand_list = lookup_inst_only.get(k) or []
                    if not cand_list:
                        continue

                    cand_fy = [c for c in cand_list if _in_fy(c.get("trans_date"), fy_start, fy_end)]
                    if not cand_fy:
                        continue

                    party_set = {c.get("party") for c in cand_fy if c.get("party")}
                    if len(party_set) != 1:
                        continue  # ambiguous instrument across parties => DO NOT attach

                    # if current party known, it must match
                    only_party = next(iter(party_set)) if party_set else ""
                    if party and only_party and party != only_party:
                        continue

                    agg = None
                    for it in cand_fy:
                        # If ERP has invoice refs and invoice is known, require match
                        invs = it.get("inv_norms") or set()
                        if invs and inv_norm and (inv_norm not in invs):
                            continue

                        if agg is None:
                            agg = dict(it)
                        else:
                            _merge_into(agg, it)

                    if agg:
                        hit = agg
                        break

            if hit:
                r.erp_paid_amount = hit.get("paid") or Decimal("0.00")
                r.erp_last_receipt_date = hit.get("last_date", "") or ""
                r.erp_receipt_no = hit.get("receipt_no", "") or ""
                r.erp_instrument_no = hit.get("inst_raw", "") or ""

                if (r.erp_paid_amount == 0) and (r.erp_receipt_no or r.erp_last_receipt_date or r.erp_instrument_no):
                    app_received = _to_dec(getattr(r, "received_amount", 0))
                    if app_received != 0:
                        r.erp_paid_amount = app_received
                        r.erp_paid_is_fallback = True

        ctx[self.context_object_name] = page_rows
        ctx["object_list"] = page_rows
        if ctx.get("page_obj") is not None:
            ctx["page_obj"].object_list = page_rows

        return ctx    
# -------------------------------------------------------------------
# List Export Excel View
# -------------------------------------------------------------------
    
# ---------- helpers used by list view (match export logic) ----------

def _apply_invoice_wise_received_split(rows):
    def _local_inst(r):
        return (
            (getattr(r, "instrument_no", "") or "").strip()
            or (getattr(r, "cheque_no", "") or "").strip()
        )

    groups = defaultdict(list)
    for r in rows:
        party = _norm_party(getattr(r, "customer_code", ""))
        inst = _local_inst(r).upper()
        if party and inst:
            groups[(party, inst)].append(r)

    for (_party, _inst), rows in groups.items():
        if len(rows) <= 1:
            continue

        total_received = max((_to_dec(getattr(x, "received_amount", 0)) for x in rows), default=Decimal("0"))
        if total_received <= 0:
            continue

        remaining = total_received
        for x in rows:
            inv_amt = _to_dec(getattr(x, "invoice_amount", 0))
            alloc = min(inv_amt, remaining) if inv_amt > 0 else Decimal("0")
            remaining -= alloc

            x._excel_received_amount = alloc
            x._excel_balance_amount = inv_amt - alloc


def _filtered_receivables_queryset(request):
    """
    Same filters as ReceivableListView.get_queryset().
    Keep this in sync with the list view.
    """
    qs = Receivable.objects.all()

    customer = (request.GET.get("customer") or "").strip()
    status = (request.GET.get("status") or "").strip()
    typ = (request.GET.get("type") or "").strip()
    company_group = (request.GET.get("company_group") or "").strip()

    from_ui = (request.GET.get("from") or "").strip()
    to_ui = (request.GET.get("to") or "").strip()

    from_dt = parse_date(from_ui) if from_ui else None
    to_dt = parse_date(to_ui) if to_ui else None

    if customer:
        qs = qs.filter(customer_name__icontains=customer)
    if status:
        qs = qs.filter(status=status)
    if typ:
        qs = qs.filter(type=typ)
    if company_group:
        qs = qs.filter(company_group=company_group)

    # ✅ Entry Date filtering (inclusive)
    if from_dt and to_dt:
        qs = qs.filter(entry_date__range=(from_dt, to_dt))
    elif from_dt:
        qs = qs.filter(entry_date__gte=from_dt)
    elif to_dt:
        qs = qs.filter(entry_date__lte=to_dt)

    return qs.order_by("-entry_date", "-id")


# ---------- helpers used by export (match list logic) ----------

def _norm_party(s: str) -> str:
    return (str(s or "").strip().upper())


def _to_dec(v) -> Decimal:
    if v in (None, "", "NA", "N/A"):
        return Decimal("0")
    try:
        s = str(v).strip().replace(",", "")
        if not s:
            return Decimal("0")
        return Decimal(s)  # handles "0E-11"
    except (InvalidOperation, ValueError, TypeError):
        return Decimal("0")


def _pick(d, keys, default=""):
    """
    Case-insensitive dict key pick.
    """
    if not isinstance(d, dict):
        return default
    keymap = {str(k).strip().casefold(): k for k in d.keys()}
    for k in keys:
        if not k:
            continue
        kk = str(k).strip().casefold()
        real_key = keymap.get(kk)
        if real_key is None:
            continue
        val = d.get(real_key)
        if val not in (None, "", "NA", "N/A"):
            return val
    return default


def _inst_db_variants(v: str) -> set:
    """
    Variants to match how instrument_no may be stored in ERP snapshot DB:
    - raw
    - base (strip leading '#')
    - clean (A-Z0-9 only)
    - same with '#'
    - numeric: add non-leading-zero variant too (003435 -> 3435)
    """
    raw = (str(v or "")).strip()
    if not raw:
        return set()

    base = raw.lstrip("#").strip()
    clean = re.sub(r"[^A-Z0-9]", "", base.upper())

    out = {raw, base, f"#{base}", clean, f"#{clean}"}
    if clean.isdigit():
        nz = clean.lstrip("0") or "0"
        out.add(nz)
        out.add(f"#{nz}")

    return {x for x in out if x}


def _merge_into(cur: dict, add: dict):
    """
    Merge receipt aggregates safely.
    - Always sums paid
    - Keeps latest trans_date (real date) and aligned last_date display
    - Preserves first non-empty receipt_no / inst_raw
    """
    cur["paid"] = (cur.get("paid") or Decimal("0")) + (add.get("paid") or Decimal("0"))

    # Prefer real date comparison (List view behavior)
    add_td = add.get("trans_date")
    cur_td = cur.get("trans_date")
    if add_td and (not cur_td or add_td > cur_td):
        cur["trans_date"] = add_td
        cur["last_date"] = add.get("last_date", "") or cur.get("last_date", "")

    if add.get("receipt_no") and not cur.get("receipt_no"):
        cur["receipt_no"] = add["receipt_no"]

    if add.get("inst_raw") and not cur.get("inst_raw"):
        cur["inst_raw"] = add["inst_raw"]

    if add.get("party") and not cur.get("party"):
        cur["party"] = add["party"]


@login_required
def receivable_list_excel(request):
    qs = Receivable.objects.all()
    qs = _apply_receivable_filters(qs, request.GET).order_by("-entry_date", "-id")

    # ✅ Latest snapshot date (once)
    snap = (
        ReceivableSnapshotRow.objects.order_by("-snapshot_date")
        .values_list("snapshot_date", flat=True)
        .first()
    )

    # ✅ Fast export: write_only mode
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="Receivables")

    headers = [
        "Entry Date",
        "Customer Code",
        "Customer Name",
        "Invoice Number",
        "Invoice Date",
        "Due Date",
        "Type",
        "Company Group",
        "Cheque/Ref No",
        "Cheque Date",
        "Invoice Amount",
        "Received Amount",

        # ✅ ERP receipt columns
        "ERP Receipt Amount",
        "ERP Receipt No",
        "ERP Receipt Date",
        "ERP Instrument No",

        "Balance Amount",
        "Status",
        "Narration/Remarks",
    ]

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F2937")  # slate-800
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    header_row = []
    for h in headers:
        c = WriteOnlyCell(ws, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        header_row.append(c)
    ws.append(header_row)

    ws.freeze_panes = "A2"

    money_fmt = "₹ #,##0.00"
    date_fmt = "DD-MMM-YYYY"

    def _safe_float(v):
        try:
            return float(v or 0)
        except Exception:
            return 0.0

    # Column widths (safe in write_only)
    widths = {
        1: 14,   # Entry Date
        2: 14,   # Customer Code
        3: 30,   # Customer Name
        4: 18,   # Invoice Number
        5: 14,   # Invoice Date
        6: 14,   # Due Date
        7: 12,   # Type
        8: 14,   # Group
        9: 16,   # Cheque No
        10: 14,  # Cheque Date
        11: 16,  # Invoice Amount
        12: 16,  # Received Amount

        13: 16,  # ERP Receipt Amount
        14: 20,  # ERP Receipt No
        15: 14,  # ERP Receipt Date
        16: 18,  # ERP Instrument No

        17: 16,  # Balance Amount
        18: 12,  # Status
        19: 40,  # Narration
    }
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # -------------------------------------------------------------------------
    # ✅ NEW: invoice normalizer + ERP raw invoice extractor (safe + optional)
    # -------------------------------------------------------------------------
    _INV_CLEAN_RE = re.compile(r"[^A-Z0-9]+")
    _INV_SPLIT_RE = re.compile(r"[,\n;/|]+")

    def _norm_inv(v) -> str:
        s = str(v or "").strip().upper()
        if not s:
            return ""
        return _INV_CLEAN_RE.sub("", s)

    def _erp_receipt_invoice_norms(raw: dict) -> set:
        """
        Extract invoice/bill numbers from ERP receipt raw.
        If ERP raw doesn't contain these keys, returns empty set -> old flow continues.
        """
        if not isinstance(raw, dict):
            return set()

        keys = [
            "Invoice No", "Invoice Number", "Inv No", "InvNo",
            "Bill No", "Bill Number", "BillNo",
            "Against Invoice No", "Against Bill No",
            "Applied Invoice No", "Applied Bill No",
            "Ref No", "Reference No", "RefNo",
            "Document No", "Doc No", "DocNo",
        ]

        v = _pick(raw, keys, default="")  # uses your existing helper

        out = set()

        def _add_token(x):
            n = _norm_inv(x)
            if n:
                out.add(n)

        if isinstance(v, (list, tuple, set)):
            for x in v:
                _add_token(x)
            return out

        if isinstance(v, dict):
            for k in v.keys():
                _add_token(k)
            return out

        s = str(v or "").strip()
        if not s:
            return set()

        parts = [p.strip() for p in _INV_SPLIT_RE.split(s) if p.strip()]
        if not parts:
            parts = [s]

        for p in parts:
            _add_token(p)

        return out

    # -------------------------------------------------------------------------
    # ERP Lookup builder (chunked, fast)
    # -------------------------------------------------------------------------
    def _build_erp_lookup_for_chunk(rows):
        """
        Returns 4 lookups (existing + invoice-aware):
          lookup_party_inst_inv: (party, inst_key, inv_norm) -> agg
          lookup_party_inst:     (party, inst_key)           -> agg
          lookup_inst_inv_only:  (inst_key, inv_norm)        -> list[agg]
          lookup_inst_only:      inst_key                    -> list[agg]
        """
        if not snap or not rows:
            return {}, {}, {}, {}

        wanted_db_values = set()
        wanted_parties = set()

        for r in rows:
            inst_local = (getattr(r, "instrument_no", "") or "").strip() or (getattr(r, "cheque_no", "") or "").strip()
            wanted_db_values |= _inst_db_variants(inst_local)

            pc = (getattr(r, "customer_code", "") or "").strip()
            if pc:
                wanted_parties.add(_norm_party(pc))

        if not wanted_db_values:
            return {}, {}, {}, {}

        q = Q()
        for v in wanted_db_values:
            q |= Q(instrument_no__iexact=v)

        if not q:
            return {}, {}, {}, {}

        erp_qs = (
            ReceivableSnapshotRow.objects
            .filter(snapshot_date=snap)
            .filter(q)
            .only(
                "party_code", "instrument_no",
                "paid_amt", "trans_date", "trans_date_display",
                "trans_no", "raw",
            )
        )

        # ✅ Keep your existing collision prevention
        if wanted_parties:
            erp_qs = erp_qs.filter(party_code__in=list(wanted_parties))

        lookup_party_inst_inv = {}
        lookup_party_inst = {}
        lookup_inst_inv_only = {}
        lookup_inst_only = {}

        for srow in erp_qs.iterator(chunk_size=2000):
            party = _norm_party(getattr(srow, "party_code", ""))
            inst_raw = (getattr(srow, "instrument_no", "") or "").strip()

            inst_keys = _norm_inst_variants(inst_raw)
            if not inst_keys:
                continue

            raw = getattr(srow, "raw", {}) or {}

            receipt_no = _pick(
                raw,
                ["Trans No", "Receipt No", "ReceiptNo", "Voucher No", "VoucherNo", "Vch No", "VchNo"],
                default=""
            ) or (getattr(srow, "trans_no", "") or "").strip()

            paid = _to_dec(getattr(srow, "paid_amt", 0))
            if paid == 0:
                paid = _to_dec(_pick(raw, ["Paid Amt", "Paid Amount", "Receipt Amount", "Amount", "Cr Amt"], default="0"))

            # Keep row if it has receipt info even if paid is zero
            if paid == 0 and not receipt_no:
                continue

            td = getattr(srow, "trans_date", None)
            td_date = td.date() if hasattr(td, "date") else td

            date_disp = (getattr(srow, "trans_date_display", "") or "").strip()
            if not date_disp and td_date:
                date_disp = td_date.strftime("%d-%b-%Y")

            item = {
                "paid": paid,
                "last_date": date_disp,
                "trans_date": td_date,     # ✅ needed for FY enforcement
                "receipt_no": str(receipt_no or "").strip(),
                "inst_raw": inst_raw,
                "party": party,
            }

            # ✅ NEW invoice norms from ERP raw (may be empty; then no change in behavior)
            inv_norms = _erp_receipt_invoice_norms(raw)

            for inst_key in inst_keys:
                lookup_inst_only.setdefault(inst_key, []).append(item)

                # ✅ NEW: instrument + invoice indexing (used as strict preference)
                for invn in inv_norms:
                    lookup_inst_inv_only.setdefault((inst_key, invn), []).append(item)

                    if party:
                        k3 = (party, inst_key, invn)
                        cur3 = lookup_party_inst_inv.get(k3)
                        if not cur3:
                            lookup_party_inst_inv[k3] = dict(item)
                        else:
                            _merge_into(cur3, item)

                # Existing party+inst aggregation (unchanged)
                if party:
                    k = (party, inst_key)
                    cur = lookup_party_inst.get(k)
                    if not cur:
                        lookup_party_inst[k] = dict(item)
                    else:
                        _merge_into(cur, item)

        return lookup_party_inst_inv, lookup_party_inst, lookup_inst_inv_only, lookup_inst_only

    def _find_hit_for_row(r, lookup_party_inst_inv, lookup_party_inst, lookup_inst_inv_only, lookup_inst_only):
        party = _norm_party(getattr(r, "customer_code", "") or "")
        inst_local = (getattr(r, "instrument_no", "") or "").strip() or (getattr(r, "cheque_no", "") or "").strip()
        local_keys = _norm_inst_variants(inst_local)
        if not local_keys:
            return None

        inv_norm = _norm_inv(getattr(r, "invoice_number", "") or "")

        base_dt = getattr(r, "invoice_date", None) or getattr(r, "entry_date", None)
        if hasattr(base_dt, "date"):
            base_date = base_dt.date()
        else:
            base_date = base_dt if isinstance(base_dt, date) else None

        fy_start, fy_end = _fy_bounds(base_date) if base_date else (None, None)

        hit = None

        # 0) ✅ NEW STRICT: party + instrument + invoice + FY
        if party and inv_norm:
            for k in local_keys:
                cand = lookup_party_inst_inv.get((party, k, inv_norm))
                if cand and _in_fy(cand.get("trans_date"), fy_start, fy_end):
                    hit = cand
                    break

        # 1) EXISTING STRICT: party + instrument + FY
        if not hit and party:
            for k in local_keys:
                cand = lookup_party_inst.get((party, k))
                if cand and _in_fy(cand.get("trans_date"), fy_start, fy_end):
                    hit = cand
                    break

        # 1.5) ✅ NEW: instrument + invoice fallback ONLY if unique party within FY
        if not hit and inv_norm:
            for k in local_keys:
                cand_list = lookup_inst_inv_only.get((k, inv_norm)) or []
                if not cand_list:
                    continue

                cand_fy = [c for c in cand_list if _in_fy(c.get("trans_date"), fy_start, fy_end)]
                if not cand_fy:
                    continue

                party_set = {c.get("party") for c in cand_fy if c.get("party")}
                if len(party_set) != 1:
                    continue  # ambiguous => skip

                agg = None
                for it in cand_fy:
                    if agg is None:
                        agg = dict(it)
                    else:
                        _merge_into(agg, it)

                if agg:
                    hit = agg
                    break

        # 2) EXISTING instrument-only fallback ONLY if single party within FY
        if not hit:
            for k in local_keys:
                cand_list = lookup_inst_only.get(k) or []
                if not cand_list:
                    continue

                cand_fy = [c for c in cand_list if _in_fy(c.get("trans_date"), fy_start, fy_end)]
                if not cand_fy:
                    continue

                party_set = {c.get("party") for c in cand_fy if c.get("party")}
                if len(party_set) != 1:
                    continue  # ambiguous => skip

                agg = None
                for it in cand_fy:
                    if agg is None:
                        agg = dict(it)
                    else:
                        _merge_into(agg, it)

                if agg:
                    hit = agg
                    break

        return hit

    # -------------------------------------------------------------------------
    # Data rows (process in chunks so we can build ERP lookups efficiently)
    # -------------------------------------------------------------------------
    EXPORT_CHUNK = 2000
    buf = []

    def _flush_buf(rows):
        if not rows:
            return

        lookup_party_inst_inv, lookup_party_inst, lookup_inst_inv_only, lookup_inst_only = _build_erp_lookup_for_chunk(rows)

        for r in rows:
            inv_amt = _to_dec(getattr(r, "invoice_amount", 0))

            rec_amt = _to_dec(
                getattr(r, "_excel_received_amount", None)
                or getattr(r, "received_amount", 0)
            )

            bal = getattr(r, "_excel_balance_amount", None)
            if bal is None:
                bal = inv_amt - rec_amt

            narration = getattr(r, "narration", "") or getattr(r, "remarks", "") or ""

            erp_paid = Decimal("0.00")
            erp_receipt_no = ""
            erp_receipt_date = ""
            erp_inst = ""

            hit = _find_hit_for_row(r, lookup_party_inst_inv, lookup_party_inst, lookup_inst_inv_only, lookup_inst_only)
            if hit:
                erp_paid = hit.get("paid") or Decimal("0.00")
                erp_receipt_no = hit.get("receipt_no") or ""
                erp_receipt_date = hit.get("last_date") or ""
                erp_inst = hit.get("inst_raw") or ""

                # ✅ keep your existing fallback rule
                if erp_paid == 0 and (erp_receipt_no or erp_receipt_date or erp_inst):
                    if rec_amt != 0:
                        erp_paid = rec_amt

            row_vals = [
                getattr(r, "entry_date", None),
                getattr(r, "customer_code", "") or "",
                getattr(r, "customer_name", "") or "",
                getattr(r, "invoice_number", "") or "",
                getattr(r, "invoice_date", None),
                getattr(r, "due_date", None),
                getattr(r, "type", "") or "",
                getattr(r, "company_group", "") or "",
                getattr(r, "cheque_no", "") or "",
                getattr(r, "cheque_date", None),

                _safe_float(inv_amt),
                _safe_float(rec_amt),

                _safe_float(erp_paid),
                erp_receipt_no,
                erp_receipt_date,
                erp_inst,

                _safe_float(bal),
                getattr(r, "status", "") or "",
                narration,
            ]

            out_row = []
            for idx, v in enumerate(row_vals, start=1):
                cell = WriteOnlyCell(ws, value=v)

                if idx in (1, 5, 6, 10) and v:
                    cell.number_format = date_fmt
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                elif idx in (11, 12, 13, 17):
                    cell.number_format = money_fmt
                    cell.alignment = Alignment(horizontal="right", vertical="center")

                else:
                    if idx in (3, 4, 19):
                        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=(idx == 19))
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")

                out_row.append(cell)

            ws.append(out_row)

    for r in qs.iterator(chunk_size=EXPORT_CHUNK):
        buf.append(r)
        if len(buf) >= EXPORT_CHUNK:
            _flush_buf(buf)
            buf = []

    _flush_buf(buf)

    ts = timezone.localtime(timezone.now()).strftime("%Y%m%d_%H%M")
    filename = f"receivables_{ts}.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


class ReceivableUpdateView(LoginRequiredMixin, UpdateView):
    model = Receivable
    form_class = ReceivableForm
    template_name = "accounts/receivable_form.html"
    success_url = reverse_lazy("accounts:receivable_list")

    def _group_queryset(self, obj):
        """
        Group definition:
        - Always same customer_code
        - If cheque_no exists: same cheque_no + cheque_date
        - Else: only this record (no grouping)
        """
        qs = Receivable.objects.filter(customer_code=obj.customer_code)
        if (obj.cheque_no or "").strip():
            qs = qs.filter(cheque_no=obj.cheque_no, cheque_date=obj.cheque_date)
        else:
            qs = qs.filter(pk=obj.pk)
        return qs.order_by("id")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        obj = self.object

        group_rows = list(self._group_queryset(obj))

        ctx["initial_customer"] = {
            "code": (obj.customer_code or "").strip(),
            "name": (obj.customer_name or "").strip(),
        }

        # For TomSelect: preselect the current invoice number in first row
        ctx["initial_invoice"] = {
            "invoice_number": (obj.invoice_number or "").strip(),
            "invoice_date": obj.invoice_date,
            "due_date": obj.due_date,
            "invoice_amount": obj.invoice_amount,
        }

        # For table rendering (extra rows)
        ctx["existing_lines"] = [
            {
                "invoice_number": (r.invoice_number or "").strip(),
                "invoice_date": r.invoice_date,
                "due_date": r.due_date,
                "invoice_amount": r.invoice_amount,
            }
            for r in group_rows
        ]
        return ctx

    def form_valid(self, form):
        self.object = self.get_object()

        rows = extract_invoice_rows(self.request.POST)

        entry_date = form.cleaned_data.get("entry_date") or self.object.entry_date or timezone.localdate()

        # Single row -> normal UpdateView
        if len(rows) <= 1:
            form.instance.entry_date = entry_date
            return super().form_valid(form)

        cd = form.cleaned_data

        common = {
            "customer_code": (cd.get("customer_code") or "").strip(),
            "customer_name": (cd.get("customer_name") or "").strip(),
            "entry_date": entry_date,
            "currency": cd.get("currency") or "INR",
            "cheque_no": cd.get("cheque_no") or None,
            "cheque_date": cd.get("cheque_date") or None,
            "remarks": cd.get("remarks") or "",
        }

        common.update(_safe_model_fields_dict(
            Receivable,
            {
                "type": cd.get("type"),
                "company_group": cd.get("company_group"),
                "status": cd.get("status"),
                "narration": cd.get("narration"),
            },
            ["type", "company_group", "status", "narration"]
        ))

        created_by = self.object.created_by or self.request.user

        # Remove duplicates
        seen = set()
        cleaned_rows = []
        for r in rows:
            inv_no = (r.get("invoice_number") or "").strip()
            if not inv_no:
                continue
            key = (
                common["customer_code"],
                inv_no.upper(),
                common["cheque_no"] or "",
                str(common["cheque_date"] or "")
            )
            if key in seen:
                continue
            seen.add(key)
            cleaned_rows.append(r)

        cleaned_rows = [r for r in cleaned_rows if r.get("invoice_date") and r.get("due_date")]
        if not cleaned_rows:
            messages.error(self.request, "No valid invoice rows found to save.")
            return redirect(self.success_url)

        first = cleaned_rows[0]

        with transaction.atomic():
            # 1) Update main object with first invoice line
            for k, v in common.items():
                setattr(self.object, k, v)

            self.object.invoice_number = (first.get("invoice_number") or "").strip()
            self.object.invoice_date = first["invoice_date"]
            self.object.due_date = first["due_date"]
            self.object.invoice_amount = first.get("invoice_amount")
            self.object.received_amount = first.get("received_amount")
            if self.object.created_by_id is None:
                self.object.created_by = created_by
            self.object.save()

            # 2) Upsert remaining rows
            for r in cleaned_rows[1:]:
                inv_no = (r.get("invoice_number") or "").strip()
                if not inv_no:
                    continue

                qs = Receivable.objects.filter(
                    customer_code=common["customer_code"],
                    invoice_number=inv_no,
                )

                if common["cheque_no"]:
                    qs = qs.filter(cheque_no=common["cheque_no"], cheque_date=common["cheque_date"])

                existing = qs.order_by("-id").first()

                if existing:
                    for k, v in common.items():
                        setattr(existing, k, v)
                    existing.invoice_date = r["invoice_date"]
                    existing.due_date = r["due_date"]
                    existing.invoice_amount = r.get("invoice_amount")
                    existing.received_amount = r.get("received_amount")
                    if existing.created_by_id is None:
                        existing.created_by = created_by
                    existing.save()
                else:
                    Receivable.objects.create(
                        created_by=created_by,
                        **common,
                        invoice_number=inv_no,
                        invoice_date=r["invoice_date"],
                        due_date=r["due_date"],
                        invoice_amount=r.get("invoice_amount"),
                        received_amount=r.get("received_amount"),
                    )

            # 3) Delete removed rows inside same cheque group (if cheque exists)
            if common["cheque_no"]:
                keep = [(r.get("invoice_number") or "").strip() for r in cleaned_rows if (r.get("invoice_number") or "").strip()]
                Receivable.objects.filter(
                    customer_code=common["customer_code"],
                    cheque_no=common["cheque_no"],
                    cheque_date=common["cheque_date"],
                ).exclude(invoice_number__in=keep).exclude(pk=self.object.pk).delete()

        messages.success(self.request, "Receivable updated with multiple invoice lines.")
        return redirect(self.success_url)
    
class ReceivableDeleteView(LoginRequiredMixin, DeleteView):
    model = Receivable
    template_name = "accounts/receivables/receivable_confirm_delete.html"
    success_url = reverse_lazy("accounts:receivable_list")

    def delete(self, request, *args, **kwargs):
        obj = self.get_object()
        ref = f"{getattr(obj, 'invoice_number', '')}".strip() or f"ID {obj.pk}"
        messages.success(request, f"Receivable entry deleted: {ref}.")
        return super().delete(request, *args, **kwargs)
    
# ----------------------------
# Receivable Dashboard View Helpers
# ----------------------------
#   --------------------------------------------------------
# Dashborad View Helpers
#  --------------------------------------------------------

from decimal import Decimal
from datetime import timedelta
from django.db.models import Max

def _row_raw_trans_type(row) -> str:
    raw = _row_get(row, "raw", default=None) or {}
    if isinstance(raw, dict):
        return str(raw.get("Trans Type") or raw.get("TransType") or "").strip()
    return ""

def _is_opening_balance_row(row) -> bool:
    tn = str(_row_get(row, "trans_no", "invoice_number", default="") or "").upper()
    tt = (_row_raw_trans_type(row) or str(_row_get(row, "trans_type", "transaction_type", default="") or "")).upper()
    return ("OPNBAL" in tn) or ("OPENING BALANCE" in tt)

def _is_non_receipt_adjustment_row(row) -> bool:
    """
    Exclude rows where paid_amt is not a real receipt.
    """
    if _is_opening_balance_row(row):
        return True

    tt = (_row_raw_trans_type(row) or str(_row_get(row, "trans_type", "transaction_type", default="") or "")).upper()

    bad = (
        "CREDIT NOTE",
        "DEBIT NOTE",
        "JOURNAL",
        "ADJUST",
        "CONTRA",
        "REVERSAL",
        "WRITE OFF",
        "ON ACCOUNT",
        "UNADJUST",
    )
    return any(x in tt for x in bad)

def _row_paid_any_decimal(row) -> Decimal:
    """
    Include ALL paid/received values, including JV/DN/CN/adjustments.
    """
    return _row_paid_amount(row)  # your existing paid getter (Decimal)

def _snapshot_date_on_or_before(target_date):
    """
    Latest available snapshot_date <= target_date.
    """
    return (
        ReceivableSnapshotRow.objects
        .filter(snapshot_date__lte=target_date)
        .aggregate(d=Max("snapshot_date"))["d"]
    )

def get_collected_receipts_total_for_snapshot(*, company_group="ALL", snapshot_date=None, party_name=None, party_code=None):
    """
    TOTAL COLLECTED (Paid-only) from beginning as-of snapshot.
    Uses snapshot rows with paid_amt > 0, excluding non-receipt adjustments.
    """
    if not snapshot_date:
        snapshot_date = latest_snapshot_date()
    if not snapshot_date:
        return Decimal("0")

    qs = ReceivableSnapshotRow.objects.filter(snapshot_date=snapshot_date, paid_amt__gt=0)

    cg = (company_group or "ALL").strip().upper()
    if cg and cg != "ALL":
        if cg == "OCSPL":
            qs = qs.filter(company_name__icontains="Special")
        elif cg == "OCCHEM":
            qs = qs.filter(company_name__icontains="Chem")
        else:
            qs = qs.filter(company_name__icontains=cg)

    if party_code:
        qs = qs.filter(party_code__iexact=party_code.strip())
    elif party_name:
        pn = party_name.strip()
        exact = qs.filter(party_name__iexact=pn)
        qs = exact if exact.exists() else qs.filter(party_name__icontains=pn)

    total = Decimal("0")
    qs = qs.only("trans_no", "paid_amt", "raw", "party_name", "party_code", "company_name")

    for r in qs.iterator(chunk_size=2000):
        row = {
            "trans_no": r.trans_no,
            "paid_amt": r.paid_amt,
            "paid_amount": r.paid_amt,
            "raw": r.raw or {},
            "party_name": r.party_name or "",
            "party_code": r.party_code or "",
            "company_name": r.company_name or "",
        }
        total += _row_receipt_paid_decimal(row)

    return total

def get_collected_receipts_last_7d(*, company_group="ALL", party_name=None, party_code=None):
    """
    RECEIVED (7D) (Paid-only) = cumulative(as-of latest snap) - cumulative(as-of snap <= latest-7d).
    """
    snap = latest_snapshot_date()
    if not snap:
        return Decimal("0")

    prev_target = snap - timedelta(days=7)
    prev_snap = _snapshot_date_on_or_before(prev_target)
    if not prev_snap:
        return Decimal("0")

    cur_total = get_collected_receipts_total_for_snapshot(
        company_group=company_group, snapshot_date=snap, party_name=party_name, party_code=party_code
    )
    prev_total = get_collected_receipts_total_for_snapshot(
        company_group=company_group, snapshot_date=prev_snap, party_name=party_name, party_code=party_code
    )

    delta = cur_total - prev_total
    return delta if delta > 0 else Decimal("0")

def _to_decimal(val, default=Decimal("0")):
    try:
        s = str(val or "").replace(",", "").strip()
        if s == "":
            return default
        return Decimal(s)
    except Exception:
        return default


def _row_get(row, *keys, default=""):
    """
    Safe getter for dict/object rows.
    - Supports dict keys in different cases (e.g. COMPANY vs company)
    - Treats None/"" as missing
    """
    if row is None:
        return default

    if isinstance(row, dict):
        for k in keys:
            if not k:
                continue

            if k in row:
                v = row.get(k)
                if v not in (None, ""):
                    return v

            lk = str(k).lower()
            uk = str(k).upper()

            if lk in row:
                v = row.get(lk)
                if v not in (None, ""):
                    return v

            if uk in row:
                v = row.get(uk)
                if v not in (None, ""):
                    return v

        return default

    for k in keys:
        if not k:
            continue

        v = getattr(row, k, None)
        if v not in (None, ""):
            return v

        v = getattr(row, str(k).lower(), None)
        if v not in (None, ""):
            return v

    return default


def _normalize_overdue_flag(overdue_flag: str) -> str:
    s = (overdue_flag or "").strip().lower()
    if s in ("1", "true", "yes", "y", "overdue"):
        return "overdue"
    if s in ("0", "false", "no", "n", "not_overdue", "clean"):
        return "not_overdue"
    return ""


def _norm_text(s: str) -> str:
    return " ".join((str(s or "")).strip().upper().split())


def _row_bill_amount(row):
    return _to_decimal(_row_get(
        row,
        "bill_amount", "bill_amt",
        "invoice_amount", "inv_amount",
        "amount", "BILL_AMT",
        default="0"
    ), default=Decimal("0"))


# ✅ CRITICAL: Paid means ONLY paid values (paid_amt/paid_amount). NOT received_* aliases.
def _row_paid_amount(row):
    return _to_decimal(_row_get(
        row,
        "paid_amt", "paid_amount", "paid",
        "PAID_AMT",
        default="0"
    ), default=Decimal("0"))


def _row_outstanding_decimal(row):
    return _to_decimal(_row_get(
        row,
        "outstanding_amount",
        "outstanding_amt",
        "balance_amount",
        "os_amt",
        "outstanding",
        "BALANCE",
        default="0"
    ), default=Decimal("0"))


# ---------------------------------------
# PO extraction helpers (kept - used for received table)
# ---------------------------------------
def _pick_first_dict(d, keys, default=""):
    for k in keys:
        v = d.get(k)
        if v not in (None, "", "NA", "N/A"):
            return v
    return default


def _row_po_no(row) -> str:
    v = _row_get(row, "customer_po_no", "customer_po_no.", "po_no", "pono", "cust_po_no", default="")
    if v:
        return str(v).strip()

    raw = _row_get(row, "raw", default=None) or {}
    if isinstance(raw, dict):
        po_no = _pick_first_dict(raw, [
            "customer_po_no", "customer_po_no.", "customerPoNo",
            "po_no", "poNo", "pono", "PONO",
            "cust_po_no", "CustPONo", "poNumber", "PONumber",
            "Customer PO No", "Customer PO No.", "CustomerPONo",
        ], default="")
        return str(po_no).strip()
    return ""


def _row_po_date(row) -> str:
    v = _row_get(row, "customer_po_date", "po_date", "podate", "cust_po_date", default="")
    if v:
        return str(v).strip()

    raw = _row_get(row, "raw", default=None) or {}
    if isinstance(raw, dict):
        po_dt = _pick_first_dict(raw, [
            "customer_po_date", "customerPoDate",
            "po_date", "poDate", "podate", "PODate",
            "cust_po_date", "CustPODate",
            "Customer PO Date", "CustomerPODate",
        ], default="")
        return str(po_dt).strip()
    return ""


# ---------------------------------------
# Company / party helpers (kept)
# ---------------------------------------
def _normalize_company_from_name(val: str) -> str:
    s = " ".join((str(val or "")).strip().upper().split())
    if not s:
        return ""

    if "OC SPECIALITIES CHEMICALS" in s:
        return "OCCHEM"
    if "OC SPECIALITIES PRIVATE LIMITED" in s:
        return "OCSPL"
    return ""


def _row_company_group(row) -> str:
    raw = _row_get(
        row,
        "company_group",
        "company_name",
        "company",
        "comp_name",
        "COMPANY",
        default=""
    )
    raw_u = (str(raw or "")).strip().upper()
    if raw_u in ("OCSPL", "OCCHEM"):
        return raw_u
    return _normalize_company_from_name(raw)


def _row_party(row) -> str:
    return (str(_row_get(row, "party_name", "customer", "customer_name", default=""))).strip()


def _parse_date_loose(v):
    if not v:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v

    s = str(v).strip()
    if not s:
        return None

    fmts = (
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%d/%m/%Y",
        "%Y/%m/%d",
        "%d.%m.%Y",
        "%Y.%m.%d",
        "%d %b %Y",
        "%d %B %Y",
    )
    for f in fmts:
        try:
            return datetime.strptime(s, f).date()
        except Exception:
            pass

    try:
        s2 = s.split("T")[0].split(" ")[0]
        return datetime.strptime(s2, "%Y-%m-%d").date()
    except Exception:
        return None


def _row_month_date(row):
    v = _row_get(
        row,
        "trans_date",
        "invoice_date",
        "bill_date",
        "doc_date",
        "inv_date",
        "date",
        "due_date",
        default=""
    )
    return _parse_date_loose(v)


def _row_overdue_days(row) -> int:
    v = _row_get(row, "overdue_days", "days_overdue", "od_days", "OD_DAYS", default=0)
    try:
        return int(float(str(v).replace(",", "").strip() or "0"))
    except Exception:
        return 0


_BUCKET_ORDER = [
    "Not due",
    "0-30 days",
    "31-60 days",
    "61-90 days",
    "91-120 days",
    "121-180 days",
    ">180 days",
]


def _row_aging_bucket(row) -> str:
    b = str(_row_get(row, "aging_bucket", "bucket_label", "aging", "bucket", default="") or "").strip()
    if b:
        return b

    d = _row_overdue_days(row)
    if d <= 0:
        return "Not due"
    if d <= 30:
        return "0-30 days"
    if d <= 60:
        return "31-60 days"
    if d <= 90:
        return "61-90 days"
    if d <= 120:
        return "91-120 days"
    if d <= 180:
        return "121-180 days"
    return ">180 days"


def _build_monthly_data_from_rows(rows):
    buckets = defaultdict(float)
    for r in (rows or []):
        d = _row_month_date(r)
        if not d:
            continue
        month_start = date(d.year, d.month, 1)
        buckets[month_start] += float(_row_outstanding_decimal(r))

    return [{"month": k, "outstanding": v} for k, v in sorted(buckets.items(), key=lambda kv: kv[0])]


def _build_aging_data_from_rows(rows):
    buckets = defaultdict(float)
    for r in (rows or []):
        b = _row_aging_bucket(r)
        buckets[b] += float(_row_outstanding_decimal(r))

    ordered = []
    seen = set()

    for b in _BUCKET_ORDER:
        if b in buckets:
            ordered.append({"aging_bucket": b, "outstanding": buckets[b]})
            seen.add(b)

    for b, v in sorted(buckets.items(), key=lambda kv: kv[0]):
        if b not in seen:
            ordered.append({"aging_bucket": b, "outstanding": v})

    return ordered


def _build_customer_data_from_rows(rows):
    buckets = defaultdict(float)
    for r in (rows or []):
        nm = _row_party(r)
        if not nm:
            continue
        buckets[nm] += float(_row_outstanding_decimal(r))

    data = []
    for k, v in buckets.items():
        data.append({
            "customer_name": k,
            "party_name": k,
            "display_party": k,
            "outstanding": v
        })

    data.sort(key=lambda x: x["outstanding"], reverse=True)
    return data


def _ensure_template_safe_keys(rows):
    """
    Make rows template-safe by ensuring commonly referenced keys exist.
    - Keeps existing values intact (uses setdefault only)
    - Ensures templates never crash on received_amt / received_amount / rcvd_* keys
    - Prefers existing received_* if present, otherwise falls back to paid_*
    """
    if not rows:
        return rows

    for r in rows:
        if not isinstance(r, dict):
            continue

        # -------------------------
        # Party/customer aliases
        # -------------------------
        party_val = (
            r.get("party_name")
            or r.get("display_party")
            or r.get("customer_name")
            or r.get("customer")
            or r.get("party")
            or ""
        )
        r.setdefault("party_name", party_val)
        r.setdefault("display_party", party_val)
        r.setdefault("customer_name", party_val)
        r.setdefault("customer", party_val)
        r.setdefault("party", party_val)

        # -------------------------
        # Identifier aliases
        # -------------------------
        trans_no = r.get("trans_no") or r.get("invoice_number") or ""
        r.setdefault("trans_no", trans_no)
        r.setdefault("invoice_number", trans_no)

        # -------------------------
        # Date aliases
        # -------------------------
        trans_dt = r.get("trans_date") or ""
        inv_dt   = r.get("invoice_date") or ""
        bill_dt  = r.get("bill_date") or ""

        r.setdefault("bill_date", bill_dt or trans_dt or inv_dt or "")
        r.setdefault("invoice_date", inv_dt or trans_dt or bill_dt or "")
        r.setdefault("trans_date", trans_dt or inv_dt or bill_dt or "")

        due_dt = r.get("due_date") or ""
        r.setdefault("due_date_display", r.get("due_date_display") or due_dt or "")
        r.setdefault("overdue_date", r.get("overdue_date") or due_dt or "")

        # -------------------------
        # Bill aliases
        # -------------------------
        bill_val = r.get("bill_amount")
        if bill_val in (None, ""):
            bill_val = r.get("bill_amt")
        if bill_val in (None, ""):
            bill_val = Decimal("0")
        r.setdefault("bill_amount", bill_val)
        r.setdefault("bill_amt", bill_val)

        # -------------------------
        # Paid aliases (source-of-truth if received_* missing)
        # -------------------------
        paid_val = r.get("paid_amount")
        if paid_val in (None, ""):
            paid_val = r.get("paid_amt")
        if paid_val in (None, ""):
            paid_val = Decimal("0")
        r.setdefault("paid_amount", paid_val)
        r.setdefault("paid_amt", paid_val)

        # -------------------------
        # ✅ Received aliases (templates use received_amt in multiple places)
        # Prefer explicit received_* / rcvd_* if present; else fall back to paid_*
        # -------------------------
        received_val = r.get("received_amount")
        if received_val in (None, ""):
            received_val = r.get("received_amt")
        if received_val in (None, ""):
            received_val = r.get("rcvd_amount")
        if received_val in (None, ""):
            received_val = r.get("rcvd_amt")
        if received_val in (None, ""):
            received_val = paid_val
        if received_val in (None, ""):
            received_val = Decimal("0")

        r.setdefault("received_amount", received_val)
        r.setdefault("received_amt", received_val)
        r.setdefault("rcvd_amount", received_val)
        r.setdefault("rcvd_amt", received_val)

        # -------------------------
        # Outstanding aliases
        # -------------------------
        os_val = r.get("outstanding_amount")
        if os_val in (None, ""):
            os_val = r.get("outstanding_amt")
        if os_val in (None, ""):
            os_val = r.get("os_amt")
        if os_val in (None, ""):
            os_val = r.get("balance_amount")
        if os_val in (None, ""):
            os_val = r.get("BALANCE")
        if os_val in (None, ""):
            os_val = Decimal("0")

        r.setdefault("outstanding_amount", os_val)
        r.setdefault("outstanding_amt", os_val)
        r.setdefault("os_amt", os_val)
        r.setdefault("balance_amount", os_val)

    return rows

# ---------------------------
# Summary from filtered rows
# ---------------------------
def _build_summary_from_rows(rows):
    total_bill = Decimal("0")
    total_os = Decimal("0")

    for r in (rows or []):
        total_bill += _to_decimal(_row_bill_amount(r), default=Decimal("0"))
        total_os += _row_outstanding_decimal(r)

    crore = (total_os / Decimal("10000000")) if total_os else Decimal("0")

    return {
        "total_invoiced": total_bill,
        "total_bill_amt": total_bill,
        "total_received": Decimal("0"),
        "total_paid_amt": Decimal("0"),
        "total_outstanding": total_os,
        "total_os_amt": total_os,
        "total_outstanding_crore": float(crore),
        "total_os_crore": float(crore),
    }


# ------------------------------------------------------------
# Receipt classification: exclude non-receipt adjustments
# ------------------------------------------------------------
def _row_raw_trans_type(row) -> str:
    raw = _row_get(row, "raw", default=None) or {}
    if isinstance(raw, dict):
        return str(raw.get("Trans Type") or raw.get("TransType") or "").strip()
    return ""


def _is_opening_balance_row(row) -> bool:
    tn = str(_row_get(row, "trans_no", "invoice_number", default="") or "").upper()
    tt = (_row_raw_trans_type(row) or str(_row_get(row, "trans_type", "transaction_type", default="") or "")).upper()
    return ("OPNBAL" in tn) or ("OPENING BALANCE" in tt)


def _is_non_receipt_adjustment_row(row) -> bool:
    if _is_opening_balance_row(row):
        return True

    tt = (_row_raw_trans_type(row) or str(_row_get(row, "trans_type", "transaction_type", default="") or "")).upper()

    bad = (
        "CREDIT NOTE",
        "DEBIT NOTE",
        "JOURNAL",
        "ADJUST",
        "CONTRA",
        "REVERSAL",
        "WRITE OFF",
        "ON ACCOUNT",
        "UNADJUST",
    )
    return any(x in tt for x in bad)


def _row_receipt_paid_decimal(row) -> Decimal:
    """
    Receipt-only paid amount (ONLY paid keys). Returns Decimal.
    """
    if _is_non_receipt_adjustment_row(row):
        return Decimal("0")
    return _row_paid_amount(row)


# ------------------------------------------------------------
# Snapshot-based providers
# ------------------------------------------------------------
def _apply_company_group_filter(qs, company_group: str):
    g = (company_group or "").strip().upper()
    if not g or g == "ALL":
        return qs

    if g == "OTHER":
        known = set()
        for names in COMPANY_GROUPS.values():
            known.update(names)
        return qs.exclude(company_name__in=list(known))

    names = COMPANY_GROUPS.get(g) or []
    if not names:
        return qs

    return qs.filter(company_name__in=names)

def _apply_company_group_filter_snapshot(qs, company_group: str):
    g = (company_group or "").strip().upper()
    if not g or g == "ALL":
        return qs

    if g == "OTHER":
        known = set()
        for names in COMPANY_GROUPS.values():
            known.update(names)
        return qs.exclude(company_name__in=list(known))

    names = COMPANY_GROUPS.get(g) or []
    if not names:
        return qs
    return qs.filter(company_name__in=names)

def _apply_party_filter(qs, *, party_name=None, party_code=None):
    if party_code:
        return qs.filter(party_code__iexact=str(party_code).strip())
    if party_name:
        return qs.filter(party_name__icontains=str(party_name).strip())
    return qs

def get_received_rows_for_snapshot(
    *,
    company_group="ALL",
    snapshot_date=None,
    party_name=None,
    party_code=None,
    from_date=None,
    to_date=None,
    limit=2000
):
    if not snapshot_date:
        snapshot_date = latest_snapshot_date()
    if not snapshot_date:
        return []

    qs = ReceivableSnapshotRow.objects.filter(snapshot_date=snapshot_date, paid_amt__gt=0)

    qs = _apply_company_group_filter(qs, company_group)
    qs = _apply_party_filter(qs, party_name=party_name, party_code=party_code)

    if from_date:
        qs = qs.filter(trans_date__gte=from_date)
    if to_date:
        qs = qs.filter(trans_date__lte=to_date)

    qs = qs.order_by("-paid_amt", "party_name", "trans_no").only(
        "company_name", "party_code", "party_name",
        "trans_no", "trans_date", "due_date",
        "bill_amt", "paid_amt", "outstanding_amt",
        "item_name", "raw",
    )[:limit]

    out = []
    for r in qs:
        out.append({
            "company_name": r.company_name or "",
            "party_code": r.party_code or "",
            "party_name": r.party_name or "",
            "trans_no": r.trans_no or "",
            "trans_date": r.trans_date,
            "due_date": r.due_date,
            "bill_amount": r.bill_amt or Decimal("0"),
            "paid_amount": r.paid_amt or Decimal("0"),
            "paid_amt": r.paid_amt or Decimal("0"),
            "outstanding_amount": r.outstanding_amt or Decimal("0"),
            "item_name": r.item_name or "",
            "raw": r.raw or {},
        })
    return out

def get_collected_total_for_snapshot_all(*, company_group="ALL", snapshot_date=None, party_name=None, party_code=None):
    """
    Sum of paid_amt for snapshot rows INCLUDING JV/DN/CN etc.
    """
    if not snapshot_date:
        snapshot_date = latest_snapshot_date()
    if not snapshot_date:
        return Decimal("0")

    qs = ReceivableSnapshotRow.objects.filter(snapshot_date=snapshot_date).exclude(paid_amt=0)

    cg = (company_group or "ALL").strip().upper()
    if cg and cg != "ALL":
        if cg == "OCSPL":
            qs = qs.filter(company_name__icontains="Special")
        elif cg == "OCCHEM":
            qs = qs.filter(company_name__icontains="Chem")
        else:
            qs = qs.filter(company_name__icontains=cg)

    if party_code:
        qs = qs.filter(party_code__iexact=party_code.strip())
    elif party_name:
        pn = party_name.strip()
        exact = qs.filter(party_name__iexact=pn)
        qs = exact if exact.exists() else qs.filter(party_name__icontains=pn)

    total = qs.aggregate(s=Sum("paid_amt"))["s"] or Decimal("0")
    return total

def get_paid_total_last_7d_all(*, company_group="ALL", snapshot_date=None, party_name=None, party_code=None):
    """
    Sum paid_amt for last 7 days window INCLUDING JV/DN/CN etc.
    Uses snapshot_date as the end date and trans_date for the window.
    """
    if not snapshot_date:
        snapshot_date = latest_snapshot_date()
    if not snapshot_date:
        return Decimal("0")

    end_dt = snapshot_date
    start_dt = end_dt - timedelta(days=6)

    qs = ReceivableSnapshotRow.objects.filter(
        snapshot_date=snapshot_date,
        trans_date__gte=start_dt,
        trans_date__lte=end_dt,
    ).exclude(paid_amt=0)

    cg = (company_group or "ALL").strip().upper()
    if cg and cg != "ALL":
        if cg == "OCSPL":
            qs = qs.filter(company_name__icontains="Special")
        elif cg == "OCCHEM":
            qs = qs.filter(company_name__icontains="Chem")
        else:
            qs = qs.filter(company_name__icontains=cg)

    if party_code:
        qs = qs.filter(party_code__iexact=party_code.strip())
    elif party_name:
        qs = qs.filter(party_name__icontains=party_name.strip())

    total = qs.aggregate(s=Sum("paid_amt"))["s"] or Decimal("0")
    return total

def _snapshot_date_on_or_before(target_date: date):
    """
    Returns latest snapshot_date <= target_date.
    Falls back to latest_snapshot_date() if none found.
    """
    if not target_date:
        return latest_snapshot_date()

    snap = (
        ReceivableSnapshotRow.objects
        .filter(snapshot_date__lte=target_date)
        .aggregate(d=Max("snapshot_date"))["d"]
    )
    return snap or latest_snapshot_date()


def _invoice_paid_as_of_snapshot(inv_row) -> Decimal:
    """
    Invoice paid (as-of snapshot) - robust:
      1) paid_amt if > 0
      2) else bill_amt - outstanding_amt if > 0
    """
    paid = _to_decimal(getattr(inv_row, "paid_amt", None), default=Decimal("0"))
    if paid > 0:
        return paid

    bill = _to_decimal(getattr(inv_row, "bill_amt", None), default=Decimal("0"))
    os_  = _to_decimal(getattr(inv_row, "outstanding_amt", None), default=Decimal("0"))
    eff = bill - os_
    return eff if eff > 0 else Decimal("0")


def build_paid_lookup_for_period(
    *,
    company_group="ALL",
    start_date=None,
    end_date=None,
    snapshot_date=None,
    wanted_keys=None,
):
    """
    Target 'Received' lookup (Against Target) using SNAPSHOT INVOICE ROWS.

    Returns:
      dict: _bill_key(party_code, invoice_no) -> Decimal(paid_as_of_snapshot)

    Notes:
      - snapshot_date defaults to latest snapshot <= end_date (target week_end)
      - wanted_keys should be list of (party_code, invoice_no) for fast exact fetch
      - If wanted_keys is not provided, it will fall back to old receipt-based scan
        (keeps backward compatibility, but invoice-based is recommended for targets).
    """
    snap = snapshot_date or _snapshot_date_on_or_before(end_date) or latest_snapshot_date()
    if not snap:
        return {}
    

    # ------------------------------------------------------------------
    # FAST PATH (recommended): invoice-row based, only for selected bills
    # ------------------------------------------------------------------
    if wanted_keys:
        party_codes = []
        invoice_nos = []

        for pc, inv in wanted_keys:
            pc = (pc or "").strip()
            inv = (inv or "").strip()
            if pc and inv:
                party_codes.append(pc)
                invoice_nos.append(inv)

        party_codes = list(set(party_codes))
        invoice_nos = list(set(invoice_nos))

        if not party_codes or not invoice_nos:
            return {}

        qs = ReceivableSnapshotRow.objects.filter(
            snapshot_date=snap,
            party_code__in=party_codes,
            trans_no__in=invoice_nos,
        )

        # Apply company group constraint same as elsewhere
        qs = _apply_company_group_filter_snapshot(qs, company_group)

        qs = qs.only("party_code", "trans_no", "paid_amt", "bill_amt", "outstanding_amt")

        out = defaultdict(Decimal)

        # If ERP duplicates invoice row (multiple LIDs), do NOT sum; keep MAX paid.
        for r in qs.iterator(chunk_size=2000):
            pc = (r.party_code or "").strip().upper()
            inv = (r.trans_no or "").strip().upper()
            if not pc or not inv:
                continue

            key = _bill_key(pc, inv)
            paid = _invoice_paid_as_of_snapshot(r)
            if paid > out[key]:
                out[key] = paid

        return dict(out)

    # ------------------------------------------------------------------
    # FALLBACK (compat): previous receipt-based scan (kept if some other
    # call site still relies on it without wanted_keys)
    # ------------------------------------------------------------------
    rows = _fetch_snapshot_receipts_rows_for_bills(
        snapshot_date=snap,
        company_group=company_group,
        start_date=start_date,
        end_date=end_date,
    )

    paid_lookup = defaultdict(Decimal)
    for rr in rows:
        pc = _norm_party(rr.get("party_code"))
        inv_ref = _extract_invoice_ref_for_against_target(rr)
        inv_norm = _norm_inv(inv_ref)

        if not pc or not inv_norm:
            continue

        amt = _erp_receipt_amount_decimal(rr)
        if amt == 0:
            continue

        paid_lookup[_bill_key(pc, inv_norm)] += amt

    return dict(paid_lookup)

# ------------------------------------------------------------

def _invoice_paid_asof_snapshot_row(sr) -> Decimal:
    """
    Paid as-of a snapshot invoice row:
      - if paid_amt > 0 use it
      - else use (bill_amt - outstanding_amt) if positive
    Works even when paid_amt is 0E-11 etc.
    """
    paid = _to_decimal(getattr(sr, "paid_amt", None), default=Decimal("0"))
    if paid > 0:
        return paid

    bill = _to_decimal(getattr(sr, "bill_amt", None), default=Decimal("0"))
    os_  = _to_decimal(getattr(sr, "outstanding_amt", None), default=Decimal("0"))
    eff = bill - os_
    return eff if eff > 0 else Decimal("0")


def _paid_asof_snapshot_for_target_lines(*, snapshot_date, company_group, lines):
    """
    Returns dict: bill_key -> paid_asof_snapshot (MAX per invoice to avoid LID duplicates)
    """
    if not snapshot_date or not lines:
        return {}

    party_codes = set()
    invoice_nos = set()
    for l in lines:
        pc = (getattr(l, "party_code", "") or "").strip()
        inv = (getattr(l, "invoice_no", "") or "").strip()
        if pc and inv:
            party_codes.add(pc)
            invoice_nos.add(inv)

    if not party_codes or not invoice_nos:
        return {}

    qs = ReceivableSnapshotRow.objects.filter(
        snapshot_date=snapshot_date,
        party_code__in=list(party_codes),
        trans_no__in=list(invoice_nos),
    )

    # keep same company group behavior as your other snapshot queries
    qs = _apply_company_group_filter_snapshot(qs, company_group)

    qs = qs.only("party_code", "trans_no", "paid_amt", "bill_amt", "outstanding_amt")

    out = defaultdict(Decimal)

    # IMPORTANT: ERP may duplicate invoice rows across LIDs -> keep MAX paid for that invoice
    for sr in qs.iterator(chunk_size=2000):
        key = _bill_key(sr.party_code, sr.trans_no)
        paid = _invoice_paid_asof_snapshot_row(sr)
        if paid > out[key]:
            out[key] = paid

    return dict(out)


def _weekly_received_lookup_for_target(*, target, lines, company_group):
    """
    Weekly received per invoice = paid_asof(end_snap) - paid_asof(prev_snap).
    If prev snapshot is missing / equals end snapshot, falls back to paid_asof(end_snap)
    so UI does not show 0.
    """
    snap_end = _snapshot_date_on_or_before(target.week_end) or latest_snapshot_date()
    if not snap_end:
        return {}, None, None

    prev_day = target.week_start - timedelta(days=1)
    snap_prev = _snapshot_date_on_or_before(prev_day)

    paid_end = _paid_asof_snapshot_for_target_lines(
        snapshot_date=snap_end,
        company_group=company_group,
        lines=lines,
    )

    # If no previous snapshot (or same as end), use fallback (show paid_asof_end)
    paid_prev = {}
    if snap_prev and snap_prev != snap_end:
        paid_prev = _paid_asof_snapshot_for_target_lines(
            snapshot_date=snap_prev,
            company_group=company_group,
            lines=lines,
        )

    out = {}
    for l in lines:
        k = _bill_key(l.party_code, l.invoice_no)
        end_amt = _to_decimal(paid_end.get(k), default=Decimal("0"))
        prev_amt = _to_decimal(paid_prev.get(k), default=Decimal("0"))
        delta = end_amt - prev_amt
        if delta < 0:
            delta = Decimal("0")
        out[k] = delta

    return out, snap_end, snap_prev

# ------------------------------------------------------------
# VIEW
# ------------------------------------------------------------
@login_required
def receivable_dashboard(request):
    received_rows = []
    received_total = Decimal("0")
    received_count = 0

    customer     = (request.GET.get("customer") or "").strip()
    status       = (request.GET.get("status") or "").strip()
    from_date    = (request.GET.get("from_date") or "").strip()
    to_date      = (request.GET.get("to_date") or "").strip()
    aging        = (request.GET.get("aging") or "").strip()
    overdue_flag = _normalize_overdue_flag(request.GET.get("overdue"))

    company_group = (request.GET.get("company_group") or "").strip().upper()
    if company_group in ("ALL", "ALL COMPANIES"):
        company_group = ""

    month  = (request.GET.get("month") or "").strip()    # YYYY-MM
    bucket = (request.GET.get("bucket") or "").strip()
    party  = (request.GET.get("party") or "").strip()

    if month:
        bucket = ""
        party = ""
        aging = ""
        customer = ""
    elif bucket:
        month = ""
        party = ""
        aging = ""
        customer = ""
    elif party:
        month = ""
        bucket = ""
        aging = ""
        customer = ""

    if party:
        customer = party
    if bucket:
        aging = bucket

    company = (request.GET.get("company") or "").strip()
    company_code = (request.GET.get("company_code") or "").strip()

    filters = {
        "customer":      customer,
        "status":        status,
        "from_date":     from_date,
        "to_date":       to_date,
        "aging":         aging,
        "overdue":       overdue_flag,
        "company_group": company_group,
        "company": company,
        "company_code": company_code,
        "month":         month,
        "bucket":        bucket,
        "party":         party,
    }

    service_filters = dict(filters)
    service_filters["company"] = ""
    service_filters["company_code"] = ""
    service_filters.pop("month", None)
    service_filters.pop("bucket", None)
    service_filters.pop("party", None)
    service_filters["customer"] = ""

    # 1) Main dashboard context from service (unchanged flow)
    ctx = build_receivable_dashboard_context(service_filters)

    receivables_list = (ctx.get("receivables") or [])
    receivables_list = _ensure_template_safe_keys(receivables_list)
    ctx["rows"] = receivables_list

    if company_group and receivables_list:
        receivables_list = [r for r in receivables_list if _row_company_group(r) == company_group]

    if overdue_flag and receivables_list:
        if overdue_flag == "overdue":
            receivables_list = [r for r in receivables_list if _row_overdue_days(r) > 0]
        elif overdue_flag == "not_overdue":
            receivables_list = [r for r in receivables_list if _row_overdue_days(r) <= 0]

    if month and receivables_list:
        try:
            y_str, m_str = month.split("-", 1)
            y, m = int(y_str), int(m_str)
        except Exception:
            y = m = None

        if y and 1 <= m <= 12:
            tmp = []
            for r in receivables_list:
                d = _row_month_date(r)
                if d and d.year == y and d.month == m:
                    tmp.append(r)
            receivables_list = tmp

    if aging and receivables_list:
        receivables_list = [r for r in receivables_list if _row_aging_bucket(r) == aging]

    if customer and receivables_list:
        cust_norm = _norm_text(customer)
        receivables_list = [r for r in receivables_list if _norm_text(_row_party(r)) == cust_norm]

    receivables_list = _ensure_template_safe_keys(receivables_list)

    ctx["summary"] = _build_summary_from_rows(receivables_list)

    base_rows = list(receivables_list)
    ctx["monthly_data"] = _ensure_template_safe_keys(_build_monthly_data_from_rows(base_rows) or [])
    ctx["aging_data"]   = _ensure_template_safe_keys(_build_aging_data_from_rows(base_rows) or [])
    ctx["customer_data"] = _ensure_template_safe_keys(_build_customer_data_from_rows(base_rows) or [])

    ctx["detail_filtered_total_outstanding"] = sum((_row_outstanding_decimal(r) for r in receivables_list), Decimal("0"))
    ctx["detail_filtered_count"] = len(receivables_list)

    paginator = Paginator(receivables_list, 50)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    q = request.GET.copy()
    q.pop("page", None)
    ctx["query_string"] = q.urlencode()

    q2 = request.GET.copy()
    q2.pop("page", None)
    q2.pop("overdue", None)
    ctx["query_string_no_overdue"] = q2.urlencode()

    ctx["page_obj"] = page_obj
    ctx["filters"] = filters

    # ---------------------------
    # Party dropdown options (for Customer select)
    # ---------------------------
    party_filters = dict(service_filters)
    party_filters["customer"] = ""  # always fetch all parties for dropdown
    party_filters["overdue"] = overdue_flag

    party_rows = _ensure_template_safe_keys(ctx.get("receivables") or [])


    # respect company_group / overdue toggle same way
    if company_group and party_rows:
        party_rows = [r for r in party_rows if _row_company_group(r) == company_group]

    if overdue_flag and party_rows:
        if overdue_flag == "overdue":
            party_rows = [r for r in party_rows if _row_overdue_days(r) > 0]
        elif overdue_flag == "not_overdue":
            party_rows = [r for r in party_rows if _row_overdue_days(r) <= 0]

    party_set = set()
    for r in party_rows:
        nm = _row_party(r)
        if nm:
            party_set.add(nm)

    ctx["party_options"] = sorted(party_set, key=lambda x: x.lower())

    # ✅ Ensure currently selected customer is always present in dropdown options
    if customer and customer not in ctx["party_options"]:
        ctx["party_options"] = [customer] + ctx["party_options"]

    # Incoming This Week kept as you had
    incoming_rows = (ctx.get("this_week_incoming") or [])
    incoming_rows = _ensure_template_safe_keys(incoming_rows)

    if incoming_rows:
        if company_group:
            incoming_rows = [r for r in incoming_rows if _row_company_group(r) == company_group]

        if overdue_flag == "overdue":
            incoming_rows = [r for r in incoming_rows if _row_overdue_days(r) > 0]
        elif overdue_flag == "not_overdue":
            incoming_rows = [r for r in incoming_rows if _row_overdue_days(r) <= 0]

        if month:
            try:
                y_str, m_str = month.split("-", 1)
                y, m = int(y_str), int(m_str)
            except Exception:
                y = m = None

            if y and 1 <= m <= 12:
                tmp = []
                for r in incoming_rows:
                    d = _row_month_date(r)
                    if d and d.year == y and d.month == m:
                        tmp.append(r)
                incoming_rows = tmp

        if aging:
            incoming_rows = [r for r in incoming_rows if _row_aging_bucket(r) == aging]

        if customer:
            cust_norm = _norm_text(customer)
            incoming_rows = [r for r in incoming_rows if _norm_text(_row_party(r)) == cust_norm]

    incoming_rows = _ensure_template_safe_keys(incoming_rows)
    ctx["this_week_incoming"] = incoming_rows
    ctx["this_week_incoming_total"] = sum((_row_outstanding_decimal(r) for r in incoming_rows), Decimal("0"))

    # ------------------------------------------------------------
    # ✅ RECEIVED / COLLECTED (Paid only; record-based)
    # ------------------------------------------------------------
    ctx.setdefault("summary", {})
    ctx["received_rows"] = []
    ctx["received_total"] = Decimal("0")
    ctx["received_count"] = 0
    ctx["previous_week_received"] = Decimal("0")

    # Keep these keys for template usage
    ctx["summary"].setdefault("total_received_as_of_snapshot", Decimal("0"))
    ctx["summary"].setdefault("received_snapshot_date", None)

    try:
        snap = latest_snapshot_date()
        svc_cg = company_group or "ALL"

        from_dt = _parse_date_loose(from_date) if from_date else None
        to_dt   = _parse_date_loose(to_date) if to_date else None

        # -------------------------
        # 1) RECEIPT DETAIL ROWS (table) - Paid-only
        # -------------------------
        received_rows = get_received_rows_for_snapshot(
            company_group=svc_cg,
            snapshot_date=snap,
            party_name=customer or None,
            from_date=from_dt,
            to_date=to_dt,
        ) or []

        received_rows = _ensure_template_safe_keys(received_rows)

        # ✅ INCLUDE ALL rows with any paid/received value (including JV/DN/CN)
        received_rows = [r for r in received_rows if _row_paid_any_decimal(r) != 0]

        # PO enrich (safe)
        for r in received_rows:
            if isinstance(r, dict):
                r.setdefault("customer_po_no", _row_po_no(r) or "")
                r.setdefault("customer_po_date", _row_po_date(r) or "")

        # ✅ Total for table (ALL types)
        received_total = sum((_row_paid_any_decimal(r) for r in received_rows), Decimal("0"))

        ctx["received_rows"] = received_rows
        ctx["received_total"] = received_total
        ctx["received_count"] = len(received_rows)

        # ✅ RECEIVED (7D) card: ALL types
        ctx["previous_week_received"] = get_paid_total_last_7d_all(
            company_group=svc_cg,
            snapshot_date=snap,
            party_name=customer or None,
            party_code=None,
        ) or Decimal("0")

        # ✅ TOTAL COLLECTED card: ALL types
        collected_total = get_collected_total_for_snapshot_all(
            company_group=svc_cg,
            snapshot_date=snap,
            party_name=customer or None,
            party_code=None,
        ) or Decimal("0")

        ctx["summary"]["total_received_as_of_snapshot"] = collected_total
        ctx["summary"]["received_snapshot_date"] = snap

        # overwrite keys used by template
        ctx["summary"]["total_received"] = collected_total
        ctx["summary"]["total_paid_amt"] = collected_total
        ctx["summary"]["total_collected"] = collected_total

    except Exception:
        pass

    return render(request, "accounts/receivable_dashboard.html", ctx)

# ---------------------------------------------------------------------------
# Receivables Export to Excel
# ---------------------------------------------------------------------------

def _xlsx_safe_str(v):
    if v is None:
        return ""
    return str(v)


def _xlsx_date(v):
    """
    Excel-friendly date:
    - returns a python date or None
    - accepts date/datetime/str
    - uses your _parse_date_loose for strings
    """
    if not v:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    try:
        return _parse_date_loose(v)  # uses your helper
    except Exception:
        return None


def _xlsx_dec(v) -> Decimal:
    try:
        return _to_decimal(v, default=Decimal("0"))  # uses your helper
    except Exception:
        return Decimal("0")


def _xlsx_autowidth(ws, max_width=60):
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws[letter]:
            val = "" if cell.value is None else str(cell.value)
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[letter].width = min(max_len + 2, max_width)


def _pay_terms_days(invoice_dt, due_dt) -> int:
    """
    Pay Terms = Due Date - Invoice Date (in days)
    """
    if not invoice_dt or not due_dt:
        return 0
    try:
        return int((due_dt - invoice_dt).days)
    except Exception:
        return 0


@login_required
def receivable_dashboard_excel(request):
    """
    Excel export for Receivable Dashboard using SAME filtering logic as receivable_dashboard().
    Adds Pay Terms + Overdue Days.
    """

    # ---------------------------
    # Read filters EXACTLY like dashboard
    # ---------------------------
    customer     = (request.GET.get("customer") or "").strip()
    status       = (request.GET.get("status") or "").strip()
    from_date    = (request.GET.get("from_date") or "").strip()
    to_date      = (request.GET.get("to_date") or "").strip()
    aging        = (request.GET.get("aging") or "").strip()
    overdue_flag = _normalize_overdue_flag(request.GET.get("overdue"))

    company_group = (request.GET.get("company_group") or "").strip().upper()
    if company_group in ("ALL", "ALL COMPANIES"):
        company_group = ""

    month  = (request.GET.get("month") or "").strip()    # YYYY-MM
    bucket = (request.GET.get("bucket") or "").strip()
    party  = (request.GET.get("party") or "").strip()

    # Dashboard mutual exclusivity rules (keep same)
    if month:
        bucket = ""
        party = ""
        aging = ""
        customer = ""
    elif bucket:
        month = ""
        party = ""
        aging = ""
        customer = ""
    elif party:
        month = ""
        bucket = ""
        aging = ""
        customer = ""

    if party:
        customer = party
    if bucket:
        aging = bucket

    company = (request.GET.get("company") or "").strip()
    company_code = (request.GET.get("company_code") or "").strip()

    filters = {
        "customer":      customer,
        "status":        status,
        "from_date":     from_date,
        "to_date":       to_date,
        "aging":         aging,
        "overdue":       overdue_flag,
        "company_group": company_group,
        "company":       company,
        "company_code":  company_code,
        "month":         month,
        "bucket":        bucket,
        "party":         party,
    }

    # service_filters EXACTLY like dashboard
    service_filters = dict(filters)
    service_filters["company"] = ""
    service_filters["company_code"] = ""
    service_filters.pop("month", None)
    service_filters.pop("bucket", None)
    service_filters.pop("party", None)
    service_filters["customer"] = ""  # as in your dashboard

    # ---------------------------
    # Build base context (unchanged flow)
    # ---------------------------
    ctx = build_receivable_dashboard_context(service_filters)

    receivables_list = _ensure_template_safe_keys(list(ctx.get("receivables") or []))
    incoming_rows    = _ensure_template_safe_keys(list(ctx.get("this_week_incoming") or []))

    # apply company_group / overdue / month / aging / customer filters
    if company_group and receivables_list:
        receivables_list = [r for r in receivables_list if _row_company_group(r) == company_group]

    if overdue_flag and receivables_list:
        if overdue_flag == "overdue":
            receivables_list = [r for r in receivables_list if _row_overdue_days(r) > 0]
        elif overdue_flag == "not_overdue":
            receivables_list = [r for r in receivables_list if _row_overdue_days(r) <= 0]

    if month and receivables_list:
        try:
            y_str, m_str = month.split("-", 1)
            y, m = int(y_str), int(m_str)
        except Exception:
            y = m = None
        if y and 1 <= m <= 12:
            tmp = []
            for r in receivables_list:
                d = _row_month_date(r)
                if d and d.year == y and d.month == m:
                    tmp.append(r)
            receivables_list = tmp

    if aging and receivables_list:
        receivables_list = [r for r in receivables_list if _row_aging_bucket(r) == aging]

    if customer and receivables_list:
        cust_norm = _norm_text(customer)
        receivables_list = [r for r in receivables_list if _norm_text(_row_party(r)) == cust_norm]

    receivables_list = _ensure_template_safe_keys(receivables_list)

    # Incoming this week filters (same logic)
    if incoming_rows:
        if company_group:
            incoming_rows = [r for r in incoming_rows if _row_company_group(r) == company_group]

        if overdue_flag == "overdue":
            incoming_rows = [r for r in incoming_rows if _row_overdue_days(r) > 0]
        elif overdue_flag == "not_overdue":
            incoming_rows = [r for r in incoming_rows if _row_overdue_days(r) <= 0]

        if month:
            try:
                y_str, m_str = month.split("-", 1)
                y, m = int(y_str), int(m_str)
            except Exception:
                y = m = None
            if y and 1 <= m <= 12:
                tmp = []
                for r in incoming_rows:
                    d = _row_month_date(r)
                    if d and d.year == y and d.month == m:
                        tmp.append(r)
                incoming_rows = tmp

        if aging:
            incoming_rows = [r for r in incoming_rows if _row_aging_bucket(r) == aging]

        if customer:
            cust_norm = _norm_text(customer)
            incoming_rows = [r for r in incoming_rows if _norm_text(_row_party(r)) == cust_norm]

    incoming_rows = _ensure_template_safe_keys(incoming_rows)

    # ---------------------------
    # Received / collected rows (same logic you do in dashboard)
    # ---------------------------
    received_rows = []
    received_total = Decimal("0")
    snap = None

    try:
        snap = latest_snapshot_date()
        svc_cg = company_group or "ALL"
        from_dt = _parse_date_loose(from_date) if from_date else None
        to_dt   = _parse_date_loose(to_date) if to_date else None

        received_rows = get_received_rows_for_snapshot(
            company_group=svc_cg,
            snapshot_date=snap,
            party_name=customer or None,
            from_date=from_dt,
            to_date=to_dt,
        ) or []

        received_rows = _ensure_template_safe_keys(received_rows)

        # include ALL rows where paid != 0 (as your dashboard)
        received_rows = [r for r in received_rows if _row_paid_any_decimal(r) != 0]

        # PO enrich (as your dashboard)
        for r in received_rows:
            if isinstance(r, dict):
                r.setdefault("customer_po_no", _row_po_no(r) or "")
                r.setdefault("customer_po_date", _row_po_date(r) or "")

        received_total = sum((_row_paid_any_decimal(r) for r in received_rows), Decimal("0"))
    except Exception:
        received_rows = []
        received_total = Decimal("0")

    # ---------------------------
    # Build Excel
    # ---------------------------
    wb = Workbook()

    # Sheet 1: Receivables
    ws1 = wb.active
    ws1.title = "Receivables"

    ws1.append([
        "Company Group",
        "Company Name",
        "Party Name",
        "Party Code",
        "Trans No",
        "Trans Date",
        "Due Date",
        "Pay Terms (Days)",
        "Overdue Days",
        "Aging Bucket",
        "Bill Amount",
        "Outstanding Amount",
        "Paid Amount (Raw)",
        "Item Name",
        "Customer PO No",
        "Customer PO Date",
        "Trans Type (Raw)",
    ])
    ws1.freeze_panes = "A2"

    for r in receivables_list:
        inv_dt = _xlsx_date(_row_get(r, "invoice_date", "trans_date", "bill_date", default=""))
        due_dt = _xlsx_date(_row_get(r, "due_date", "overdue_date", default=""))
        pay_terms = _pay_terms_days(inv_dt, due_dt)

        ws1.append([
            _row_company_group(r),
            _xlsx_safe_str(_row_get(r, "company_name", "company", default="")),
            _xlsx_safe_str(_row_get(r, "party_name", "customer_name", "display_party", default="")),
            _xlsx_safe_str(_row_get(r, "party_code", default="")),
            _xlsx_safe_str(_row_get(r, "trans_no", "invoice_number", default="")),
            inv_dt,
            due_dt,
            pay_terms,
            int(_row_overdue_days(r) or 0),
            _xlsx_safe_str(_row_aging_bucket(r)),
            float(_xlsx_dec(_row_bill_amount(r))),
            float(_xlsx_dec(_row_outstanding_decimal(r))),
            float(_xlsx_dec(_row_paid_any_decimal(r))),  # raw paid (may include adjustments)
            _xlsx_safe_str(_row_get(r, "item_name", default="")),
            _xlsx_safe_str(_row_po_no(r)),
            _xlsx_safe_str(_row_po_date(r)),
            _xlsx_safe_str(_row_raw_trans_type(r)),
        ])

    _xlsx_autowidth(ws1)

    # Sheet 2: Incoming This Week
    ws2 = wb.create_sheet("Incoming This Week")
    ws2.append([
        "Company Group",
        "Company Name",
        "Party Name",
        "Party Code",
        "Trans No",
        "Trans Date",
        "Due Date",
        "Pay Terms (Days)",
        "Overdue Days",
        "Aging Bucket",
        "Bill Amount",
        "Outstanding Amount",
        "Item Name",
        "Customer PO No",
        "Customer PO Date",
    ])
    ws2.freeze_panes = "A2"

    for r in incoming_rows:
        inv_dt = _xlsx_date(_row_get(r, "invoice_date", "trans_date", "bill_date", default=""))
        due_dt = _xlsx_date(_row_get(r, "due_date", "overdue_date", default=""))
        pay_terms = _pay_terms_days(inv_dt, due_dt)

        ws2.append([
            _row_company_group(r),
            _xlsx_safe_str(_row_get(r, "company_name", "company", default="")),
            _xlsx_safe_str(_row_get(r, "party_name", "customer_name", "display_party", default="")),
            _xlsx_safe_str(_row_get(r, "party_code", default="")),
            _xlsx_safe_str(_row_get(r, "trans_no", "invoice_number", default="")),
            inv_dt,
            due_dt,
            pay_terms,
            int(_row_overdue_days(r) or 0),
            _xlsx_safe_str(_row_aging_bucket(r)),
            float(_xlsx_dec(_row_bill_amount(r))),
            float(_xlsx_dec(_row_outstanding_decimal(r))),
            _xlsx_safe_str(_row_get(r, "item_name", default="")),
            _xlsx_safe_str(_row_po_no(r)),
            _xlsx_safe_str(_row_po_date(r)),
        ])

    _xlsx_autowidth(ws2)

    # Sheet 3: Received / Collected
    ws3 = wb.create_sheet("Received")
    ws3.append([
        "Snapshot Date",
        "Company Name",
        "Party Name",
        "Party Code",
        "Trans No",
        "Trans Date",
        "Due Date",
        "Pay Terms (Days)",
        "Paid Amount",
        "Outstanding Amount",
        "Item Name",
        "Customer PO No",
        "Customer PO Date",
        "Trans Type (Raw)",
    ])
    ws3.freeze_panes = "A2"

    for r in received_rows:
        trn_dt = _xlsx_date(_row_get(r, "trans_date", "invoice_date", default=""))
        due_dt = _xlsx_date(_row_get(r, "due_date", "overdue_date", default=""))
        pay_terms = _pay_terms_days(trn_dt, due_dt)

        ws3.append([
            snap,
            _xlsx_safe_str(_row_get(r, "company_name", default="")),
            _xlsx_safe_str(_row_get(r, "party_name", default="")),
            _xlsx_safe_str(_row_get(r, "party_code", default="")),
            _xlsx_safe_str(_row_get(r, "trans_no", default="")),
            trn_dt,
            due_dt,
            pay_terms,
            float(_xlsx_dec(_row_paid_any_decimal(r))),
            float(_xlsx_dec(_row_outstanding_decimal(r))),
            _xlsx_safe_str(_row_get(r, "item_name", default="")),
            _xlsx_safe_str(_row_po_no(r)),
            _xlsx_safe_str(_row_po_date(r)),
            _xlsx_safe_str(_row_raw_trans_type(r)),
        ])

    _xlsx_autowidth(ws3)

    # Sheet 4: Summary
    ws4 = wb.create_sheet("Summary")
    ws4.append(["Key", "Value"])
    ws4.freeze_panes = "A2"

    total_outstanding = sum((_row_outstanding_decimal(r) for r in receivables_list), Decimal("0"))
    total_bill_amt = sum((_xlsx_dec(_row_bill_amount(r)) for r in receivables_list), Decimal("0"))
    incoming_total = sum((_row_outstanding_decimal(r) for r in incoming_rows), Decimal("0"))

    ws4.append(["Snapshot Date", snap or ""])
    ws4.append(["Receivables Rows (Filtered)", len(receivables_list)])
    ws4.append(["Receivables Total Bill Amount", float(total_bill_amt)])
    ws4.append(["Receivables Total Outstanding", float(total_outstanding)])
    ws4.append(["Incoming Rows (Filtered)", len(incoming_rows)])
    ws4.append(["Incoming Total Outstanding", float(incoming_total)])
    ws4.append(["Received Rows (Filtered)", len(received_rows)])
    ws4.append(["Received Total (Paid)", float(received_total)])

    ws4.append(["--- Filters ---", ""])
    for k in [
        "company_group", "customer", "status", "from_date", "to_date",
        "aging", "overdue", "month", "bucket", "party"
    ]:
        ws4.append([k, _xlsx_safe_str(filters.get(k, ""))])

    _xlsx_autowidth(ws4, max_width=80)

    # Response
    today = date.today()
    filename = f"Receivable_Dashboard_{today:%Y%m%d}.xlsx"
    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(resp)

    # ✅ signal frontend to stop spinner
    resp.set_cookie("excel_download", "1", max_age=60, path="/")

    return resp

# ---------------------------------------------------------------------------
# Receivables Snapshot Sync views
# ---------------------------------------------------------------------------


@login_required
@require_POST
def receivables_snapshot_sync(request):
    """
    Trigger receivables snapshot sync from UI (frontend button -> backend endpoint).
    Keeps existing flow unchanged: runs sync then redirects back to dashboard.
    """
    next_url = request.POST.get("next") or request.META.get("HTTP_REFERER") or "/accounts/receivables/dashboard/"
    as_of = request.POST.get("as_of")  # optional, yyyy-mm-dd

    try:
        # If your command supports a date arg, pass it; else call without it.
        if as_of:
            call_command("sync_receivables_snapshot", as_of=as_of)
        else:
            call_command("sync_receivables_snapshot")

        messages.success(request, f"Receivables snapshot sync completed for {as_of or 'latest date'}.")
    except TypeError:
        # If the command doesn't accept as_of, fall back safely
        call_command("sync_receivables_snapshot")
        messages.success(request, "Receivables snapshot sync completed.")
    except Exception as e:
        messages.error(request, f"Receivables snapshot sync failed: {e}")

    return redirect(next_url)

@login_required
@require_POST
def receivables_snapshot_sync_start(request):
    """
    Starts sync in background and returns run_id.
    """
    run_id = str(uuid.uuid4())
    as_of = request.POST.get("as_of")  # optional yyyy-mm-dd

    set_state(run_id, status="queued", percent=1, step="queued", message="Queued...")

    def _runner():
        try:
            set_state(run_id, status="running", percent=2, step="starting", message="Starting...")
            if as_of:
                call_command("sync_receivables_snapshot", as_of=as_of, run_id=run_id)
            else:
                call_command("sync_receivables_snapshot", run_id=run_id)
        except TypeError:
            # if your command doesn't accept run_id/as_of yet
            try:
                call_command("sync_receivables_snapshot")
                set_state(run_id, status="done", percent=100, step="done", message="Snapshot sync complete.")
            except Exception as e:
                set_state(run_id, status="error", percent=100, step="error", message=str(e))
        except Exception as e:
            set_state(run_id, status="error", percent=100, step="error", message=str(e))

    t = threading.Thread(target=_runner, daemon=True)
    t.start()

    return JsonResponse({"ok": True, "run_id": run_id})


@login_required
@require_GET
def receivables_snapshot_sync_status(request):
    run_id = (request.GET.get("run_id") or "").strip()
    if not run_id:
        return JsonResponse({"ok": False, "error": "Missing run_id"}, status=400)
    return JsonResponse({"ok": True, "state": get_state(run_id)})

# ---------------------------------------------------------------------------
# Payment Target Week views
# ---------------------------------------------------------------------------
# -------------------------
# helpers (single source)
# -------------------------

def _as_float(v):
    """
    Safe conversion for openpyxl numeric cells.
    """
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(Decimal(str(v)))
    except Exception:
        return 0.0


def _bill_key(party_code, invoice_no):
    return f"{(party_code or '').strip().upper()}||{(invoice_no or '').strip().upper()}"


def _chunked(iterable, size):
    it = iter(iterable)
    while True:
        chunk = list(islice(it, size))
        if not chunk:
            break
        yield chunk


def _prev_target_for_group(company_group: str, new_week_start: date):
    """
    Previous target for same group strictly before new_week_start.
    """
    cg = (company_group or "ALL").strip()
    cg_db = "" if cg.upper() == "ALL" else cg
    return (
        PaymentTargetWeek.objects
        .filter(company_group=cg_db, week_end__lt=new_week_start)
        .order_by("-week_end", "-id")
        .first()
    )

def _selected_cache_keys_from_request(request) -> set:
    """
    Reads selection map from hidden field 'selected_cache'.
    Expected formats:
      - {}  (dict of selected keys -> metadata)
      - []  (list of selected keys)
    Returns normalized UPPER bill_keys.
    """
    raw = (request.POST.get("selected_cache") or "").strip()
    if not raw:
        return set()

    try:
        data = json.loads(raw)
    except Exception:
        return set()

    out = set()

    if isinstance(data, dict):
        for k in data.keys():
            ks = str(k or "").strip().upper()
            if ks:
                out.add(ks)
        return out

    if isinstance(data, list):
        for k in data:
            ks = str(k or "").strip().upper()
            if ks:
                out.add(ks)
        return out

    return set()


def _resolve_selected_keys(request, valid_keys: set) -> list:
    """
    Ensures ONLY selected bills are saved.

    Strategy:
    - read checkbox keys from POST.getlist("bill_key")
    - read selected_cache keys from POST["selected_cache"]
    - intersect with valid_keys (safety)
    - if checkbox list equals ALL valid keys (hidden-post bug), prefer selected_cache
    """
    post_keys = {
        str(k or "").strip().upper()
        for k in request.POST.getlist("bill_key")
        if str(k or "").strip()
    }
    post_keys &= valid_keys

    cache_keys = _selected_cache_keys_from_request(request) & valid_keys

    # If bill_key is polluted (contains everything), use selected_cache instead
    if cache_keys and post_keys and post_keys == valid_keys and cache_keys != valid_keys:
        return sorted(cache_keys)

    # Prefer checkboxes if they look sane; otherwise fallback to cache
    if post_keys:
        return sorted(post_keys)

    return sorted(cache_keys)


def _carry_forward_open_bills_map(*, prev_target, company_group: str, new_week_end: date, snapshot_date: date):
    """
    Returns:
      carry_map: {bill_key -> bill_dict compatible with preview_rows}
      carry_keys: set(bill_key)

    Only includes invoices that are STILL OPEN as-of snapshot_date (outstanding_amt__gt=0).

    IMPORTANT: avoids SQL Server IN cross-product false positives by re-checking exact keys.
    """
    if not prev_target or not snapshot_date:
        return {}, set()

    prev_pairs = list(
        PaymentTargetLine.objects
        .filter(target=prev_target)
        .values_list("party_code", "invoice_no")
    )
    prev_pairs = [(str(pc).strip(), str(inv).strip()) for pc, inv in prev_pairs if pc and inv]
    if not prev_pairs:
        return {}, set()

    prev_key_set = {_bill_key(pc, inv) for pc, inv in prev_pairs}

    party_codes = list({pc for pc, _ in prev_pairs})
    invoice_nos = list({inv for _, inv in prev_pairs})

    qs = ReceivableSnapshotRow.objects.filter(
        snapshot_date=snapshot_date,
        party_code__in=party_codes,
        trans_no__in=invoice_nos,
        outstanding_amt__gt=0,
    )
    qs = _apply_company_group_filter_snapshot(qs, company_group)

    # keep SAME selection rule as your open bills query (validity by date range)
    qs = qs.filter(
        Q(due_date__lte=new_week_end)
        | Q(due_date__isnull=True, overdue_date__lte=new_week_end)
        | Q(due_date__isnull=True, overdue_date__isnull=True)
    )

    qs = qs.only(
        "party_code", "party_name", "trans_no", "trans_date",
        "due_date", "overdue_date", "bill_amt", "outstanding_amt",
    )

    carry_map = {}
    for r in qs.iterator(chunk_size=2000):
        key = _bill_key(r.party_code, r.trans_no)
        if key not in prev_key_set:
            continue

        carry_map[key] = {
            "bill_key": key,
            "party_code": (r.party_code or "").strip(),
            "party_name": (r.party_name or "").strip(),
            "invoice_no": (r.trans_no or "").strip(),
            "invoice_date": r.trans_date,
            "due_date": r.due_date or r.overdue_date,
            "bill_amount": r.bill_amt or Decimal("0"),
            "outstanding_amount": r.outstanding_amt or Decimal("0"),
            "is_carry_forward": True,  # ✅ used by template
        }

    return carry_map, set(carry_map.keys())


def _attach_bill_keys(rows):
    """
    Ensures each row has bill_key + is_carry_forward default.
    Keeps order stable.
    """
    out = []
    seen = set()
    for r in (rows or []):
        k = _bill_key(r.get("party_code"), r.get("invoice_no"))
        if k in seen:
            continue
        seen.add(k)
        r["bill_key"] = k
        r.setdefault("is_carry_forward", False)
        out.append(r)
    return out


def _merge_carry_forward_rows(rows, carry_map, carry_keys):
    """
    Merge carry-forward rows into existing rows, without breaking ordering.
    - Existing rows keep order.
    - Missing carry rows get prepended (sorted) so they appear prominently.
    """
    rows = _attach_bill_keys(rows)
    seen = {r["bill_key"] for r in rows}

    missing = [carry_map[k] for k in carry_keys if k in carry_map and k not in seen]
    missing.sort(key=lambda r: ((r.get("party_name") or ""), (r.get("party_code") or ""), (r.get("invoice_no") or "")))

    # mark carry flag on existing rows too
    carry_key_set = set(carry_keys or [])
    for r in rows:
        if r.get("bill_key") in carry_key_set:
            r["is_carry_forward"] = True

    return missing + rows


def _exclude_already_selected(preview_rows, batch_size=900):
    """
    Exclude any bill already present in PaymentTargetLine (any target).

    IMPORTANT:
    - PaymentTargetLine does NOT have a 'bill_key' field (per your FieldError),
      so we must match using (party_code, invoice_no).
    - We keep exact key matching via _bill_key() so behavior stays identical.
    - SQL Server-safe: avoids OR chains and large parameter explosions.
    """
    if not preview_rows:
        return preview_rows or []

    pairs = []
    for r in preview_rows:
        pc = (r.get("party_code") or "").strip().upper()
        inv = (r.get("invoice_no") or "").strip().upper()
        if pc and inv:
            pairs.append((pc, inv))

    if not pairs:
        return preview_rows

    existing_set = set()

    # batch pairs to keep SQL params safely below SQL Server limits
    for batch in _chunked(pairs, batch_size):
        party_codes = list({pc for pc, _ in batch})
        invoice_nos = list({inv for _, inv in batch})

        existing = (
            PaymentTargetLine.objects
            .filter(party_code__in=party_codes, invoice_no__in=invoice_nos)
            .values_list("party_code", "invoice_no")
        )

        # exact key match (prevents cross-product false positives)
        for pc, inv in existing:
            existing_set.add(_bill_key(pc, inv))

    return [
        r for r in preview_rows
        if _bill_key(r.get("party_code"), r.get("invoice_no")) not in existing_set
    ]


# default cutoff (01-Jan-2025)
BILL_CUTOFF_DATE = date(2025, 1, 1)


def get_open_bills_for_period_snapshot(
    *,
    start_date,
    end_date,
    company_group="ALL",
    party_code=None,
    party_name=None,
    snapshot_date=None,            # ✅ allow passing snap to avoid recompute
    bill_cutoff_date=None          # ✅ allow passing cutoff (fixes your error)
):
    """
    Open bills from ReceivableSnapshotRow (snapshot table) as-of WEEK END (end_date).

    Keeps logic same, but limits bills to Invoice Date >= cutoff (default 01-Jan-2025)
    to reduce fetch time.
    """
    snap = snapshot_date or latest_snapshot_date()
    if not snap or not end_date:
        return []

    cutoff = bill_cutoff_date or BILL_CUTOFF_DATE

    qs = ReceivableSnapshotRow.objects.filter(
        snapshot_date=snap,
        outstanding_amt__gt=0,
    )
    qs = _apply_company_group_filter_snapshot(qs, company_group)

    cg = (company_group or "ALL").strip().upper()
    if cg and cg != "ALL":
        if cg == "OCSPL":
            qs = qs.filter(company_name__icontains="Special")  # adjust if needed
        elif cg == "OCCHEM":
            qs = qs.filter(company_name__icontains="Chem")     # adjust if needed
        else:
            qs = qs.filter(company_name__icontains=cg)

    if party_code:
        qs = qs.filter(party_code__iexact=party_code.strip())
    elif party_name:
        qs = qs.filter(party_name__icontains=party_name.strip())

    # ✅ enforce invoice date >= 01-Jan-2025 (keep null dates to avoid surprises)
    qs = qs.filter(Q(trans_date__isnull=True) | Q(trans_date__gte=cutoff))

    # ✅ existing due/overdue selection logic (unchanged)
    qs = qs.filter(
        Q(due_date__lte=end_date) |
        Q(due_date__isnull=True, overdue_date__lte=end_date) |
        Q(due_date__isnull=True, overdue_date__isnull=True)
    )

    # optional safety (unchanged)
    qs = qs.filter(Q(trans_date__isnull=True) | Q(trans_date__lte=end_date))

    qs = qs.only(
        "party_code", "party_name", "trans_no",
        "trans_date", "due_date", "overdue_date",
        "bill_amt", "outstanding_amt",
        "erp_lid",
    ).order_by("party_name", "party_code", "trans_no", "erp_lid")

    out = []
    for r in qs:
        out.append({
            "party_code": (r.party_code or "").strip(),
            "party_name": (r.party_name or "").strip(),
            "invoice_no": (r.trans_no or "").strip(),
            "invoice_date": r.trans_date,
            "due_date": r.due_date or r.overdue_date,
            "bill_amount": r.bill_amt or Decimal("0"),
            "outstanding_amount": r.outstanding_amt or Decimal("0"),
        })

    return out


# ---------------------------------------------------------------------------
# LIST
# ---------------------------------------------------------------------------
@login_required
def target_list(request):
    qs = (
        PaymentTargetWeek.objects
        .all()
        .prefetch_related("lines")
        .order_by("-week_start", "-created_at")[:50]
    )
    targets = list(qs)

    snap = latest_snapshot_date()

    # (group, ws, we) -> list of (party_code, invoice_no) needed
    wanted_by_period = defaultdict(list)

    for t in targets:
        group = _norm_group(t.company_group)
        lines = list(t.lines.all())
        key = (group, t.week_start, t.week_end)
        for l in lines:
            if l.party_code and l.invoice_no:
                wanted_by_period[key].append((str(l.party_code).strip(), str(l.invoice_no).strip()))

    # (group, ws, we) -> paid_lookup dict { BILLKEY: Decimal }
    paid_lookup_cache = {}

    for (group, ws, we), wanted_pairs in wanted_by_period.items():

        # 1) Try paid from receipt/bank lookup (may be empty for many ERPs)
        raw_paid_lookup = build_paid_lookup_for_period(
            company_group=group,
            start_date=ws,
            end_date=we,
            snapshot_date=snap,
        ) or {}

        paid_lookup = {}
        for k, v in raw_paid_lookup.items():
            amt = _to_decimal(v, default=Decimal("0"))
            if amt == 0:
                continue

            if isinstance(k, (tuple, list)) and len(k) == 2:
                kk = _bill_key(k[0], k[1])
            else:
                ks = str(k or "").strip()
                if "||" in ks:
                    pc_part, inv_part = ks.split("||", 1)
                    kk = _bill_key(pc_part, inv_part)
                else:
                    kk = str(ks).strip().upper()

            paid_lookup[kk] = paid_lookup.get(kk, Decimal("0")) + amt

        # 2) Backfill missing invoice paid from snapshot invoice rows (reliable)
        wanted_keys = {_bill_key(pc, inv) for pc, inv in wanted_pairs}
        missing_keys = wanted_keys - set(paid_lookup.keys())

        if missing_keys:
            party_set = {pc for pc, _ in wanted_pairs if pc}
            invno_set = {inv for _, inv in wanted_pairs if inv}

            snap_qs = ReceivableSnapshotRow.objects.filter(
                snapshot_date=snap,
                party_code__in=party_set,
                trans_no__in=invno_set,
            ).only(
                "party_code", "trans_no", "trans_type",
                "bill_amt", "paid_amt", "outstanding_amt",
            )

            if group != "ALL":
                if _has_field(ReceivableSnapshotRow, "company_group"):
                    snap_qs = snap_qs.filter(company_group=group)
                elif _has_field(ReceivableSnapshotRow, "company_code"):
                    snap_qs = snap_qs.filter(company_code=group)

            for sr in snap_qs:
                if not _is_invoice_type(sr.trans_type):
                    continue

                kk = _bill_key(sr.party_code, sr.trans_no)
                if kk not in missing_keys:
                    continue

                bill = _to_decimal(getattr(sr, "bill_amt", None), default=Decimal("0"))
                paid = _to_decimal(getattr(sr, "paid_amt", None), default=Decimal("0"))
                os_  = _to_decimal(getattr(sr, "outstanding_amt", None), default=Decimal("0"))

                eff_paid = paid if paid > 0 else (bill - os_)
                if eff_paid > 0:
                    paid_lookup[kk] = max(paid_lookup.get(kk, Decimal("0")), eff_paid)

        paid_lookup_cache[(group, ws, we)] = paid_lookup

    # Attach totals per target
    for t in targets:
        group = _norm_group(t.company_group)
        lines = list(t.lines.all())

        expected_total = sum(((l.expected_amount or Decimal("0")) for l in lines), Decimal("0"))

        lookup = paid_lookup_cache.get((group, t.week_start, t.week_end), {})

        received_total = Decimal("0")
        for l in lines:
            received_total += _to_decimal(lookup.get(_bill_key(l.party_code, l.invoice_no)), default=Decimal("0"))

        t.received_total = received_total
        t.balance_total = expected_total - received_total

    return render(request, "accounts/targets/target_list.html", {"targets": targets})


# ---------------------------------------------------------------------------
# CREATE
# ---------------------------------------------------------------------------
@login_required
def target_create(request):
    preview_rows = []
    preview_totals = None

    today = timezone.localdate()
    default_week_start = today - timedelta(days=today.weekday())
    default_week_end = default_week_start + timedelta(days=6)

    company_group = (request.GET.get("company_group") or "ALL").strip() or "ALL"
    week_start_ui = (request.GET.get("week_start") or "").strip()
    week_end_ui = (request.GET.get("week_end") or "").strip()
    fetch_flag = (request.GET.get("fetch") or "") == "1"

    from_dt = _parse_ui_date(week_start_ui) or default_week_start
    to_dt = _parse_ui_date(week_end_ui) or default_week_end

    form = PaymentTargetWeekForm(initial={
        "company_group": company_group,
        "week_start": from_dt,
        "week_end": to_dt,
    })

    party_form = PaymentTargetSelectPartyForm(request.POST or None)

    selected_cache_json = (request.POST.get("selected_cache") or "").strip() or "{}"
    try:
        json.loads(selected_cache_json)
    except Exception:
        selected_cache_json = "{}"

    # -----------------------
    # POST
    # -----------------------
    if request.method == "POST":
        action = (request.POST.get("action") or "").strip().lower()

        form = PaymentTargetWeekForm(request.POST)
        if not form.is_valid():
            messages.error(request, "Please correct the errors in the form.")
            return render(request, "accounts/targets/target_create.html", {
                "form": form,
                "party_form": party_form,
                "preview_rows": [],
                "preview_totals": None,
                "selected_cache_json": selected_cache_json,
                "filters": {
                    "company_group": company_group,
                    "week_start": from_dt.strftime("%Y-%m-%d"),
                    "week_end": to_dt.strftime("%Y-%m-%d"),
                }
            })

        company_group = (form.cleaned_data.get("company_group") or "ALL").strip() or "ALL"
        from_dt = form.cleaned_data["week_start"]
        to_dt = form.cleaned_data["week_end"]

        # ✅ compute snapshot once
        snap = latest_snapshot_date()
        if not snap:
            messages.error(request, "Snapshot not available. Please try again after snapshot sync.")
            return render(request, "accounts/targets/target_create.html", {
                "form": form,
                "party_form": party_form,
                "preview_rows": [],
                "preview_totals": None,
                "selected_cache_json": selected_cache_json,
                "filters": {
                    "company_group": company_group,
                    "week_start": from_dt.strftime("%Y-%m-%d"),
                    "week_end": to_dt.strftime("%Y-%m-%d"),
                }
            })

        party_code = ""
        party_name = ""
        if party_form.is_valid():
            party_code = (party_form.cleaned_data.get("party_code") or "").strip()
            party_name = (party_form.cleaned_data.get("party_name") or "").strip()

        # Preview rows (filtered)
        preview_rows = get_open_bills_for_period_snapshot(
            start_date=from_dt,
            end_date=to_dt,
            company_group=company_group,
            party_code=party_code or None,
            party_name=party_name or None,
            snapshot_date=snap,
            bill_cutoff_date=BILL_CUTOFF_DATE,
        )

        total_out = sum((_to_decimal(r.get("outstanding_amount")) for r in preview_rows), Decimal("0"))
        preview_totals = {"total_outstanding": total_out}

        if action == "fetch":
            return render(request, "accounts/targets/target_create.html", {
                "form": form,
                "party_form": party_form,
                "preview_rows": preview_rows,
                "preview_totals": preview_totals,
                "selected_cache_json": selected_cache_json,
                "filters": {
                    "company_group": company_group,
                    "week_start": from_dt.strftime("%Y-%m-%d"),
                    "week_end": to_dt.strftime("%Y-%m-%d"),
                }
            })

        if action == "create_save":
            # ✅ Build lookup from ALL open bills (multi-party safe)
            all_rows = get_open_bills_for_period_snapshot(
                start_date=from_dt,
                end_date=to_dt,
                company_group=company_group,
                party_code=None,
                party_name=None,
                snapshot_date=snap,
                bill_cutoff_date=BILL_CUTOFF_DATE,
            )

            bills_by_key = {
                _bill_key(b.get("party_code"), b.get("invoice_no")): b
                for b in all_rows
            }
            valid_keys = set(bills_by_key.keys())

            # ✅ FIX: resolve ONLY selected keys (prevents saving all)
            selected_keys = _resolve_selected_keys(request, valid_keys)

            if not selected_keys:
                messages.error(request, "Please select at least one bill before saving.")
                return render(request, "accounts/targets/target_create.html", {
                    "form": form,
                    "party_form": party_form,
                    "preview_rows": preview_rows,
                    "preview_totals": preview_totals,
                    "selected_cache_json": selected_cache_json,
                    "filters": {
                        "company_group": company_group,
                        "week_start": from_dt.strftime("%Y-%m-%d"),
                        "week_end": to_dt.strftime("%Y-%m-%d"),
                    }
                })

            with transaction.atomic():
                target = PaymentTargetWeek.objects.create(
                    week_start=from_dt,
                    week_end=to_dt,
                    company_group=company_group if company_group != "ALL" else "",
                    notes=form.cleaned_data.get("notes") or "",
                    created_by=request.user,
                )

                created_count = 0
                updated_count = 0
                missing_count = 0

                for key in selected_keys:
                    k = (key or "").strip().upper()
                    b = bills_by_key.get(k)
                    if not b:
                        missing_count += 1
                        continue

                    expected = _to_decimal(
                        request.POST.get(f"expected__{k}", ""),
                        default=_to_decimal(b.get("outstanding_amount"), default=Decimal("0"))
                    )

                    _, created = PaymentTargetLine.objects.update_or_create(
                        target=target,
                        party_code=(b.get("party_code") or ""),
                        invoice_no=(b.get("invoice_no") or ""),
                        defaults={
                            "party_name": (b.get("party_name") or ""),
                            "invoice_date": (b.get("invoice_date") or None),
                            "due_date": (b.get("due_date") or None),
                            "bill_amount": _to_decimal(b.get("bill_amount"), default=Decimal("0")),
                            "outstanding_amount": _to_decimal(b.get("outstanding_amount"), default=Decimal("0")),
                            "expected_amount": expected,
                        }
                    )

                    created_count += 1 if created else 0
                    updated_count += 0 if created else 1

            if missing_count:
                messages.warning(request, f"{missing_count} selected bills could not be matched (snapshot list changed).")

            messages.success(request, f"Target week created. Added: {created_count}, Updated: {updated_count}.")
            return redirect("accounts:payment_target_detail", pk=target.pk)

        messages.error(request, "Invalid action.")
        return redirect("accounts:payment_target_create")

    # -----------------------
    # GET (fetch=1)
    # -----------------------
    if fetch_flag:
        snap = latest_snapshot_date()
        if snap:
            preview_rows = get_open_bills_for_period_snapshot(
                start_date=from_dt,
                end_date=to_dt,
                company_group=company_group,
                party_code=None,
                party_name=None,
                snapshot_date=snap,
                bill_cutoff_date=BILL_CUTOFF_DATE,
            )

            total_out = sum((_to_decimal(r.get("outstanding_amount")) for r in preview_rows), Decimal("0"))
            preview_totals = {"total_outstanding": total_out}
        else:
            preview_rows = []
            preview_totals = None

    return render(request, "accounts/targets/target_create.html", {
        "form": form,
        "party_form": party_form,
        "preview_rows": preview_rows,
        "preview_totals": preview_totals,
        "selected_cache_json": "{}",
        "filters": {
            "company_group": company_group,
            "week_start": from_dt.strftime("%Y-%m-%d"),
            "week_end": to_dt.strftime("%Y-%m-%d"),
        }
    })

# ---------------------------------------------------------------------------
# EDIT
# ---------------------------------------------------------------------------
@login_required
def target_edit(request, pk):
    target = get_object_or_404(PaymentTargetWeek, pk=pk)

    party_form = PaymentTargetSelectPartyForm(request.POST or None)

    preview_rows = []
    preview_totals = None

    def _build_selected_cache_json():
        d = {}
        for l in PaymentTargetLine.objects.filter(target=target):
            k = _bill_key(l.party_code, l.invoice_no)
            d[k] = {"expected": str(l.expected_amount or "")}
        return json.dumps(d)

    selected_cache_json = _build_selected_cache_json()

    if request.method == "POST":
        action = (request.POST.get("action") or "").strip().lower()

        form = PaymentTargetWeekForm(request.POST)
        if not form.is_valid():
            messages.error(request, "Please correct the errors in the form.")
            return render(request, "accounts/targets/target_edit.html", {
                "target": target,
                "form": form,
                "party_form": party_form,
                "preview_rows": preview_rows,
                "preview_totals": preview_totals,
                "selected_cache_json": request.POST.get("selected_cache") or selected_cache_json,
            })

        company_group = (form.cleaned_data.get("company_group") or "ALL").strip() or "ALL"
        from_dt = form.cleaned_data["week_start"]
        to_dt = form.cleaned_data["week_end"]

        snap = latest_snapshot_date()
        if not snap:
            messages.error(request, "Snapshot not available. Please try again after snapshot sync.")
            return render(request, "accounts/targets/target_edit.html", {
                "target": target,
                "form": form,
                "party_form": party_form,
                "preview_rows": [],
                "preview_totals": None,
                "selected_cache_json": request.POST.get("selected_cache") or selected_cache_json,
            })

        preview_rows = get_open_bills_for_period_snapshot(
            start_date=from_dt,
            end_date=to_dt,
            company_group=company_group,
            party_code=None,
            party_name=None,
            snapshot_date=snap,
            bill_cutoff_date=BILL_CUTOFF_DATE,
        )
        total_out = sum((_to_decimal(r.get("outstanding_amount")) for r in preview_rows), Decimal("0"))
        preview_totals = {"total_outstanding": total_out}

        if action in ("update_save", "save", "update"):
            # ✅ rebuild lookup from ALL open bills
            all_rows = get_open_bills_for_period_snapshot(
                start_date=from_dt,
                end_date=to_dt,
                company_group=company_group,
                party_code=None,
                party_name=None,
                snapshot_date=snap,
                bill_cutoff_date=BILL_CUTOFF_DATE,
            )
            bills_by_key = {
                _bill_key(b.get("party_code"), b.get("invoice_no")): b
                for b in all_rows
            }
            valid_keys = set(bills_by_key.keys())

            # ✅ FIX: resolve ONLY selected keys
            selected_keys = _resolve_selected_keys(request, valid_keys)

            with transaction.atomic():
                target.week_start = from_dt
                target.week_end = to_dt
                target.company_group = "" if company_group == "ALL" else company_group
                target.notes = form.cleaned_data.get("notes") or ""
                target.save(update_fields=["week_start", "week_end", "company_group", "notes"])

                PaymentTargetLine.objects.filter(target=target).delete()

                created = 0
                missing = 0

                for key in selected_keys:
                    k = (key or "").strip().upper()
                    b = bills_by_key.get(k)
                    if not b:
                        missing += 1
                        continue

                    expected = _to_decimal(
                        request.POST.get(f"expected__{k}", ""),
                        default=_to_decimal(b.get("outstanding_amount"), default=Decimal("0"))
                    )

                    PaymentTargetLine.objects.create(
                        target=target,
                        party_code=(b.get("party_code") or ""),
                        party_name=(b.get("party_name") or ""),
                        invoice_no=(b.get("invoice_no") or ""),
                        invoice_date=(b.get("invoice_date") or None),
                        due_date=(b.get("due_date") or None),
                        bill_amount=_to_decimal(b.get("bill_amount"), default=Decimal("0")),
                        outstanding_amount=_to_decimal(b.get("outstanding_amount"), default=Decimal("0")),
                        expected_amount=expected,
                    )
                    created += 1

            if missing:
                messages.warning(request, f"{missing} selected bills could not be matched (snapshot list changed).")
            messages.success(request, f"Target updated. Lines saved: {created}.")
            return redirect("accounts:payment_target_detail", pk=target.pk)

    else:
        form = PaymentTargetWeekForm(initial={
            "week_start": target.week_start,
            "week_end": target.week_end,
            "company_group": target.company_group or "ALL",
            "notes": target.notes,
        })

        snap = latest_snapshot_date()
        if snap:
            cg = (target.company_group or "ALL").strip() or "ALL"
            preview_rows = get_open_bills_for_period_snapshot(
                start_date=target.week_start,
                end_date=target.week_end,
                company_group=cg,
                party_code=None,
                party_name=None,
                snapshot_date=snap,
                bill_cutoff_date=BILL_CUTOFF_DATE,
            )
            total_out = sum((_to_decimal(r.get("outstanding_amount")) for r in preview_rows), Decimal("0"))
            preview_totals = {"total_outstanding": total_out}
        else:
            preview_rows = []
            preview_totals = None

    return render(request, "accounts/targets/target_edit.html", {
        "target": target,
        "form": form,
        "party_form": party_form,
        "preview_rows": preview_rows,
        "preview_totals": preview_totals,
        "selected_cache_json": selected_cache_json,
    })

#--------------------------------------------------------------
# DELETE
# -------------------------------------------------------------

@login_required
def target_delete(request, pk):
    target = get_object_or_404(PaymentTargetWeek, pk=pk)

    if request.method == "POST":
        with transaction.atomic():
            PaymentTargetLine.objects.filter(target=target).delete()
            target.delete()
        messages.success(request, "Target week deleted successfully.")
        return redirect("accounts:payment_target_list")

    return render(request, "accounts/targets/target_delete.html", {"target": target})


# =============================================================================
# TARGET DETAIL (Week Target) - FULL UPDATED CODE
# Maintains existing flow:
#   1) Target totals
#   2) Received Against Target (selected bills)  <-- FIXED (invoice ref parsing + stable keys)
#   3) Week receipts (snapshot)
#   4) Cheque/PDC entries (application) enriched with ERP receipt (instrument matching)  <-- unchanged flow
# =============================================================================

import re
from collections import defaultdict
from datetime import date, datetime, timedelta
from decimal import Decimal
from typing import Any

from django.contrib.auth.decorators import login_required
from django.core.cache import cache
from django.db.models import Q
from django.shortcuts import get_object_or_404, render
from django.utils import timezone

# Assumes your project already has these:
# - ReceivableSnapshotRow, Receivable, PaymentTargetWeek, PaymentTargetLine
# - latest_snapshot_date
# - _apply_company_group_filter_snapshot
# - get_received_totals_for_period, get_received_rows_for_period
# - logger
# -----------------------------------------------------------------------------


# ---------------------------------------------------------------------------
# COMMON HELPERS (single definition only – avoids function override bugs)
# ---------------------------------------------------------------------------

def _gv(obj, key, default=None):
    """Get value from dict OR object attribute safely."""
    if obj is None:
        return default
    if isinstance(obj, dict):
        return obj.get(key, default)
    return getattr(obj, key, default)


def _to_decimal(v, default=Decimal("0")) -> Decimal:
    """Safe Decimal conversion (handles 0E-11, commas, blanks)."""
    if v in (None, "", "NA", "N/A"):
        return default
    if isinstance(v, Decimal):
        return v
    try:
        s = str(v).strip().replace(",", "")
        if not s:
            return default
        return Decimal(s)
    except Exception:
        return default


def _safe_dec(v, default=Decimal("0")) -> Decimal:
    return _to_decimal(v, default=default)


def _to_date(v):
    """
    Accepts: date, datetime, 'YYYY-MM-DD', 'DD-MM-YYYY', 'DD/MM/YYYY', '31 Dec 2025', etc.
    Returns: datetime.date or None
    """
    if not v:
        return None

    if isinstance(v, date) and not isinstance(v, datetime):
        return v

    if isinstance(v, datetime):
        return v.date()

    s = str(v).strip()
    if not s:
        return None

    for fmt in (
        "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d",
        "%d %b %Y", "%d %B %Y",
    ):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            continue

    try:
        return datetime.fromisoformat(s[:10]).date()
    except Exception:
        return None


def _days_between(d1: date, d2: date) -> int:
    """d2 - d1 in days, safe."""
    if not d1 or not d2:
        return 0
    try:
        return (d2 - d1).days
    except Exception:
        return 0


# ---------------------------------------------------------------------------
# NORMALIZERS (single source of truth)
# ---------------------------------------------------------------------------

_INV_CLEAN_RE = re.compile(r"[^A-Z0-9]+")
_INST_CLEAN_RE = re.compile(r"[^A-Z0-9]+")

def _norm_inv(v: Any) -> str:
    s = str(v or "").strip().upper()
    if not s:
        return ""
    return _INV_CLEAN_RE.sub("", s)

def _norm_party(v: Any) -> str:
    return str(v or "").strip().upper()

def _bill_key(pc, inv) -> str:
    return f"{_norm_party(pc)}||{_norm_inv(inv)}"

def _norm_inst(v: Any) -> str:
    s = str(v or "").strip().upper()
    if not s:
        return ""
    if s.startswith("#"):
        s = s[1:]
    s = _INST_CLEAN_RE.sub("", s)  # remove spaces, /, -, etc.
    if s.isdigit():
        s = s.lstrip("0") or "0"
    return s


# =============================================================================
# (A) DETAIL: PDC / Instrument-based ERP Receipt Matching (UNCHANGED FLOW)
# =============================================================================

_INST_RE = re.compile(
    r"(?:inst(?:rument)?|chq|cheque|check)\s*[:#-]?\s*([A-Za-z0-9\/-]{4,})",
    re.I
)

def _fetch_snapshot_receipts_rows(*, snapshot_date, company_group="ALL", start_date=None, end_date=None):
    """
    Pull receipt-like rows from ReceivableSnapshotRow for PDC matching.
    IMPORTANT:
      - Do NOT filter by paid_amt__gt=0 (many bank receipt rows show 0E-11 but still contain receipt no/date + instrument).
      - Filter by trans_type contains 'receipt' and instrument_no present to keep it relevant & fast.
    """
    if not snapshot_date:
        return []

    qs = ReceivableSnapshotRow.objects.filter(snapshot_date=snapshot_date)
    qs = _apply_company_group_filter_snapshot(qs, company_group)

    qs = qs.filter(trans_type__icontains="receipt").exclude(instrument_no="")

    if start_date:
        qs = qs.filter(trans_date__gte=start_date)
    if end_date:
        qs = qs.filter(trans_date__lte=end_date)

    qs = qs.only(
        "company_name", "party_code", "party_name",
        "trans_type", "trans_no",
        "instrument_no",
        "trans_date", "trans_date_display",
        "paid_amt", "raw",
    )

    out = []
    for r in qs.iterator(chunk_size=2000):
        out.append({
            "company_name": r.company_name or "",
            "party_code": (r.party_code or "").strip(),
            "party_name": r.party_name or "",
            "trans_type": r.trans_type or "",
            "trans_no": r.trans_no or "",
            "trans_date": r.trans_date,
            "trans_date_display": r.trans_date_display or "",
            "instrument_no": r.instrument_no or "",
            "paid_amount": r.paid_amt or Decimal("0"),
            "paid_amt": r.paid_amt or Decimal("0"),
            "raw": r.raw or {},
        })
    return out


def _extract_invoice_ref_from_receipt_row(r) -> str:
    """
    Generic invoice ref extractor (used in PDC index too).
    For PDC strict key (party+inv+inst) – keep as-is.
    """
    for k in ("ref_no", "Ref No", "RefNo", "Reference No", "ReferenceNo", "invoice_no", "Invoice No", "Bill No", "bill_no"):
        v = _gv(r, k)
        if v:
            return str(v).strip().upper()

    raw = _gv(r, "raw", None)
    if isinstance(raw, dict):
        for k in ("ref_no", "Ref No", "RefNo", "Reference No", "ReferenceNo", "invoice_no", "Invoice No", "Bill No", "bill_no"):
            v = raw.get(k)
            if v:
                return str(v).strip().upper()

    return ""


def _extract_instrument_from_receipt_row(r) -> str:
    """
    Extract ERP instrument/cheque no for matching to Django cheque_no/instrument_no.
    """
    for k in (
        "sExtNo", "sExtNo.", "Ext No", "ExtNo", "External No", "ExternalNo",
        "instrument_no", "Instrument No.", "Instrument No", "InstrumentNo",
        "cheque_no", "Cheque No", "Chq No", "ChequeNo",
    ):
        v = _gv(r, k)
        if v:
            return _norm_inst(str(v))

    raw = _gv(r, "raw", None)
    if isinstance(raw, dict):
        for k in ("sExtNo", "instrument_no", "Instrument No", "Instrument No.", "Cheque No", "Chq No"):
            v = raw.get(k)
            if v:
                return _norm_inst(str(v))

    blob = " ".join([
        str(_gv(r, "remarks", "") or ""),
        str(_gv(r, "narration", "") or ""),
        str(_gv(r, "Narration", "") or ""),
        str(_gv(r, "trans_ref", "") or ""),
        str(_gv(r, "particulars", "") or ""),
    ]).strip()

    if blob:
        m = _INST_RE.search(blob)
        if m:
            return _norm_inst(m.group(1) or "")

        nums = re.findall(r"\b\d{4,}\b", blob)
        nums = list({n for n in nums})
        if len(nums) == 1:
            return _norm_inst(nums[0])

    return ""


def _erp_receipt_amount_decimal(r) -> Decimal:
    """
    Best-effort ERP receipt amount getter.
    (Used by both PDC matching and Against Target matching)
    """
    candidates = [
        _gv(r, "paid_amount"), _gv(r, "paid_amt"),
        _gv(r, "Paid Amt"), _gv(r, "Paid Amount"),
        _gv(r, "Received Amt"), _gv(r, "PDC Received Amount"),
        _gv(r, "Receipt Amount"), _gv(r, "Receipt Amt"),
        _gv(r, "amount"), _gv(r, "Amount"),
        _gv(r, "Cr Amt"), _gv(r, "Credit Amt"),
        _gv(r, "unadjustment_amt"), _gv(r, "Unadjustment Amt"),
        _gv(r, "unadjustment_amount"), _gv(r, "Unadjustment Amount"),
        _gv(r, "bill_amt"), _gv(r, "Bill Amt"),
        _gv(r, "Bill Amt"),
    ]

    raw = _gv(r, "raw", None)
    if isinstance(raw, dict):
        candidates += [
            raw.get("Paid Amt"), raw.get("Paid Amount"),
            raw.get("Received Amt"), raw.get("PDC Received Amount"),
            raw.get("Receipt Amount"), raw.get("Receipt Amt"),
            raw.get("Amount"), raw.get("amount"),
            raw.get("Cr Amt"), raw.get("Credit Amt"),
            raw.get("Unadjustment Amt"), raw.get("Unadjustment Amount"),
            raw.get("Bill Amt"), raw.get("bill_amt"),
        ]

    for v in candidates:
        d = _to_decimal(v, default=Decimal("0"))
        if d != 0:
            return d

    return Decimal("0")


def _build_erp_pdc_lookup_from_rows(rows: list) -> dict:
    strict = defaultdict(list)    # (party_code, inv_norm, inst_norm)
    relaxed = defaultdict(list)   # (party_code, inst_norm)
    inst_only = defaultdict(list) # (inst_norm)

    for r in (rows or []):
        pc = (_gv(r, "party_code", "") or _gv(r, "Party Code", "") or "").strip().upper()

        inst_norm = _extract_instrument_from_receipt_row(r)
        if not inst_norm:
            continue

        invref_raw = _extract_invoice_ref_from_receipt_row(r)
        inv_norm = _norm_inv(invref_raw)

        receipt_no = (
            _gv(r, "receipt_no", "") or _gv(r, "Receipt No", "") or
            _gv(r, "voucher_no", "") or _gv(r, "Voucher No", "") or
            _gv(r, "trans_no", "") or _gv(r, "Trans No", "") or ""
        )

        trans_date = (
            _gv(r, "trans_date", None) or _gv(r, "trans_date_display", "") or
            _gv(r, "Trans Date", "") or _gv(r, "Receipt Date", "") or
            _gv(r, "Voucher Date", "") or ""
        )

        paid_amount = _erp_receipt_amount_decimal(r)

        item = {
            "company_name": _gv(r, "company_name", "") or _gv(r, "Company Name", "") or "",
            "party_code": pc,
            "party_name": _gv(r, "party_name", "") or _gv(r, "Party Name", "") or "",
            "invoice_no": inv_norm,
            "invoice_no_raw": invref_raw,
            "trans_date": trans_date,
            "paid_amount": _to_decimal(paid_amount, default=Decimal("0")),
            "instrument_no": inst_norm,
            "receipt_no": str(receipt_no or "").strip(),
        }

        if pc and inv_norm:
            strict[(pc, inv_norm, inst_norm)].append(item)
        if pc:
            relaxed[(pc, inst_norm)].append(item)

        inst_only[inst_norm].append(item)

    return {"strict": strict, "relaxed": relaxed, "inst_only": inst_only}


def _attach_erp_to_receivable_entry(entry: dict, erp_index: dict) -> None:
    pc = (entry.get("party_code") or "").strip().upper()
    inv_raw = (entry.get("invoice_no") or "").strip().upper()
    inv_norm = _norm_inv(inv_raw)

    inst = _norm_inst(entry.get("instrument_no") or entry.get("cheque_no") or "")

    entry["erp_matched"] = False
    entry["erp_paid_total"] = Decimal("0")
    entry["erp_receipts"] = []
    entry.setdefault("receipt_no", "")
    entry.setdefault("receipt_date", "")

    if not inst:
        return

    strict = erp_index.get("strict") or {}
    relaxed = erp_index.get("relaxed") or {}
    inst_only = erp_index.get("inst_only") or {}

    matches = []

    # 1) Strict
    if pc and inv_norm:
        matches = strict.get((pc, inv_norm, inst)) or []

    # 2) Relaxed: party+instrument
    if not matches and pc:
        cand = relaxed.get((pc, inst)) or []
        cand_nz = [c for c in cand if _to_decimal(c.get("paid_amount"), default=Decimal("0")) != 0]
        if cand_nz:
            cand = cand_nz

        if len(cand) == 1:
            matches = cand
        else:
            inv_set = {c.get("invoice_no") for c in cand if c.get("invoice_no")}
            if len(inv_set) <= 1:
                matches = cand

    # 3) Instrument-only fallback
    if not matches:
        cand = inst_only.get(inst) or []
        cand_nz = [c for c in cand if _to_decimal(c.get("paid_amount"), default=Decimal("0")) != 0]
        if cand_nz:
            cand = cand_nz

        if len(cand) == 1:
            matches = cand
        else:
            inv_set = {c.get("invoice_no") for c in cand if c.get("invoice_no")}
            if len(inv_set) <= 1:
                matches = cand

    if not matches:
        return

    entry["erp_matched"] = True
    entry["erp_receipts"] = matches
    entry["erp_paid_total"] = sum((m.get("paid_amount") or Decimal("0")) for m in matches)

    latest_dt = None
    latest_dt_raw = ""
    for m in matches:
        td_raw = str(m.get("trans_date") or "").strip()
        td = _to_date(td_raw)
        if td and (latest_dt is None or td > latest_dt):
            latest_dt = td
            latest_dt_raw = td_raw
        elif not latest_dt and td_raw and (not latest_dt_raw or td_raw > latest_dt_raw):
            latest_dt_raw = td_raw

    entry["receipt_date"] = latest_dt_raw or (matches[0].get("trans_date") or "")
    entry["receipt_no"] = matches[0].get("receipt_no") or ""


# ---------------------------------------------------------------------------
# PDC entry link instrument helpers
# ---------------------------------------------------------------------------

_CHEQUE_RE = re.compile(r"(?:chq|cheque)\s*[:#-]?\s*([A-Za-z0-9\/-]{4,})", re.I)

def _extract_link_instrument(rv) -> str:
    """
    Link key priority:
      1) rv.instrument_no
      2) rv.cheque_no
      3) parse from remarks
    """
    inst = (getattr(rv, "instrument_no", None) or "").strip()
    if inst:
        return inst

    chq = (getattr(rv, "cheque_no", None) or "").strip()
    if chq:
        return chq

    remarks = (getattr(rv, "remarks", None) or "").strip()
    if remarks:
        m = _CHEQUE_RE.search(remarks)
        if m:
            return (m.group(1) or "").strip()
        m2 = re.search(r"\b(\d{6,})\b", remarks)
        if m2:
            return (m2.group(1) or "").strip()

    return ""


# =============================================================================
# (B) AGAINST TARGET: RECEIVED LOOKUP (FIXED)
#   Problem: Ref No often blank; invoice appears in narration/particulars.
#   Fix: Parse invoice token (e.g. CMU...) and accept more raw keys.
# =============================================================================

# Adjust patterns if your invoice format differs.
_CMU_RE = re.compile(r"\b(CMU\d{6,})\b", re.I)          # example: CMU252610115071
_LONG_TOKEN_RE = re.compile(r"\b([A-Z0-9]{8,})\b", re.I)

def _fetch_snapshot_receipts_rows_for_bills(*, snapshot_date, company_group="ALL", start_date=None, end_date=None):
    """
    Receipt rows for invoice matching (Against Target).
    NOTE:
      - do NOT require instrument_no
      - do NOT require paid_amt > 0
    """
    if not snapshot_date:
        return []

    qs = ReceivableSnapshotRow.objects.filter(snapshot_date=snapshot_date)
    qs = _apply_company_group_filter_snapshot(qs, company_group)
    qs = qs.filter(trans_type__icontains="receipt")

    if start_date:
        qs = qs.filter(trans_date__gte=start_date)
    if end_date:
        qs = qs.filter(trans_date__lte=end_date)

    qs = qs.only(
        "company_name", "party_code", "party_name",
        "trans_type", "trans_no",
        "trans_date", "trans_date_display",
        "paid_amt", "raw",
        "instrument_no",
    )

    out = []
    for r in qs.iterator(chunk_size=2000):
        out.append({
            "company_name": r.company_name or "",
            "party_code": (r.party_code or "").strip(),
            "party_name": r.party_name or "",
            "trans_type": r.trans_type or "",
            "trans_no": r.trans_no or "",
            "trans_date": r.trans_date,
            "trans_date_display": r.trans_date_display or "",
            "instrument_no": r.instrument_no or "",
            "paid_amt": r.paid_amt or Decimal("0"),
            "raw": r.raw or {},
        })
    return out


def _extract_invoice_ref_for_against_target(rr: dict) -> str:
    """
    Stronger invoice ref extractor for "Against Target" matching.
    Tries:
      1) explicit invoice/ref keys
      2) raw keys
      3) parse from narration/remarks/particulars
    """
    # 1) direct keys
    for k in (
        "Agst Bill", "Agst Bill No", "Agst Bill No.", "AgstBillNo",
        "Against Bill No", "Against Bill No.", "AgainstBillNo",
        "Bill Ref", "Bill Ref No", "Bill Ref No.", "BillRefNo",
        "Ref", "Reference", "Bill Reference",
        "Adj Against Bill", "Adjusted Against", "Against Document",
    ):
        v = _gv(rr, k)
        if v:
            return str(v).strip().upper()

    # 2) raw keys
    raw = _gv(rr, "raw", None)
    if isinstance(raw, dict):
        for k in (
            "Agst Bill", "Agst Bill No", "Agst Bill No.", "AgstBillNo",
            "Against Bill No", "Against Bill No.", "AgainstBillNo",
            "Bill Ref", "Bill Ref No", "Bill Ref No.", "BillRefNo",
            "Ref", "Reference", "Bill Reference",
            "Adj Against Bill", "Adjusted Against", "Against Document",
        ):
            v = raw.get(k)
            if v:
                return str(v).strip().upper()

    # 3) parse from text blobs
    blob = " ".join([
        str(_gv(rr, "narration", "") or ""),
        str(_gv(rr, "Narration", "") or ""),
        str(_gv(rr, "remarks", "") or ""),
        str(_gv(rr, "particulars", "") or ""),
        str(_gv(rr, "trans_ref", "") or ""),
        str(_gv(rr, "trans_no", "") or ""),
    ]).strip().upper()

    if blob:
        m = _CMU_RE.search(blob)
        if m:
            return m.group(1).strip().upper()

        # fallback: long token containing digits
        tokens = _LONG_TOKEN_RE.findall(blob)
        for t in tokens:
            tt = (t or "").strip().upper()
            if any(ch.isdigit() for ch in tt):
                return tt
            
        # 4) last resort: sometimes invoice is sitting in trans_no
            tn = str(_gv(rr, "trans_no", "") or "").strip().upper()
            if tn:
                m = _CMU_RE.search(tn)
                if m:
                    return m.group(1).strip().upper()


    return ""


def build_paid_lookup_for_period(*, company_group="ALL", start_date=None, end_date=None, snapshot_date=None):
    """
    Returns dict: key=_bill_key(party_code, invoice_no) -> Decimal(received)
    Used by Week Target (Against Target).
    """
    if not snapshot_date:
        snapshot_date = latest_snapshot_date()
    if not snapshot_date:
        return {}

    rows = _fetch_snapshot_receipts_rows_for_bills(
        snapshot_date=snapshot_date,
        company_group=company_group,
        start_date=start_date,
        end_date=end_date,
    )

    paid_lookup = defaultdict(Decimal)

    for rr in rows:
        pc = _norm_party(rr.get("party_code"))
        inv_ref = _extract_invoice_ref_for_against_target(rr)
        inv_norm = _norm_inv(inv_ref)

        if not pc or not inv_norm:
            continue

        amt = _erp_receipt_amount_decimal(rr)
        if amt == 0:
            continue

        paid_lookup[_bill_key(pc, inv_norm)] += amt

    return dict(paid_lookup)
# =============================================================================
# (C) RECEIVED SECTION: ENRICH RECEIPT ROWS FROM SNAPSHOT
# =============================================================================

import json
from decimal import Decimal
from dateutil import parser as dtparser

RECEIPT_TYPE_HINTS = ("receipt", "pdc", "bank receipt", "bank rcpt", "br")
INVOICE_TYPE_HINTS = ("sales invoice", "invoice", "tax invoice")

def _is_receipt_type(tt: str) -> bool:
    t = (tt or "").strip().lower()
    return any(x in t for x in RECEIPT_TYPE_HINTS)

def _is_invoice_type(trans_type: str) -> bool:
    s = (trans_type or "").strip().lower()
    if not s:
        return False
    # exclude receipt types
    if "receipt" in s or "payment" in s:
        return False
    # include all invoice variants
    return "invoice" in s

def _parse_erp_date(v):
    if v in (None, ""):
        return None
    try:
        # ERP strings like "03 Jan 2026" OR "2026-01-03"
        return dtparser.parse(str(v).strip(), dayfirst=True).date()
    except Exception:
        return None

def _norm_key(k: str) -> str:
    return str(k or "").strip().lower()

def _raw_get(d, *keys, default=None):
    """
    Robust raw getter: matches keys by exact OR normalized (strip+lower).
    """
    if not isinstance(d, dict):
        return default

    # fast path: direct
    for k in keys:
        if not k:
            continue
        if k in d:
            v = d.get(k)
            if v not in (None, "", "NA", "N/A"):
                return v

    # normalized path
    nd = {_norm_key(k): v for k, v in d.items()}
    for k in keys:
        nk = _norm_key(k)
        if not nk:
            continue
        if nk in nd:
            v = nd.get(nk)
            if v not in (None, "", "NA", "N/A"):
                return v

    return default

def _dec(v, default=Decimal("0")) -> Decimal:
    try:
        if v in (None, "", "NA", "N/A"):
            return default
        return Decimal(str(v))
    except Exception:
        return default

def _invoice_effective_paid(inv_row) -> Decimal:
    """
    Paid on invoice can be stored as paid_amt OR computed from bill_amt - outstanding_amt.
    """
    paid = _dec(getattr(inv_row, "paid_amt", None), default=Decimal("0"))
    bill = _dec(getattr(inv_row, "bill_amt", None), default=Decimal("0"))
    os_  = _dec(getattr(inv_row, "outstanding_amt", None), default=Decimal("0"))
    eff = bill - os_
    if paid > 0:
        return paid
    return eff if eff > 0 else Decimal("0")

def enrich_received_rows(snapshot_rows, week_start, week_end):
    """
    snapshot_rows: iterable of ReceivableSnapshotRow (same snapshot_date)
    Returns list of dicts for Received section
    """

    # Build invoice lookup by (party_code, trans_no, trans_date)
    inv_map = {}
    inv_by_party = {}
    for r in snapshot_rows:
        if _is_invoice_type(getattr(r, "trans_type", "")):
            key = (getattr(r, "party_code", ""), getattr(r, "trans_no", ""), getattr(r, "trans_date", None))
            inv_map[key] = r
            pc = getattr(r, "party_code", "") or ""
            inv_by_party.setdefault(pc, []).append(r)

    # Keep a conservative allocation set to avoid double-attributing the same invoice
    allocated_invoices = set()

    received = []

    for r in snapshot_rows:
        if not _is_receipt_type(getattr(r, "trans_type", "")):
            continue

        try:
            raw = json.loads(getattr(r, "raw", None) or "{}")
        except Exception:
            raw = {}

        # 1) payment date: prefer ERP Trans Date from raw
        payment_date = _parse_erp_date(_raw_get(raw, "Trans Date", "TransDate", "Receipt Date", "Doc Date")) \
                       or getattr(r, "trans_date", None)

        if not payment_date or not (week_start <= payment_date <= week_end):
            continue

        # 2) amount: prefer snapshot numeric columns
        amt = _dec(getattr(r, "paid_amt", None), default=Decimal("0"))
        if amt <= 0:
            amt = _dec(getattr(r, "bill_amt", None), default=Decimal("0"))

        # 3) fallback A: map receipt -> invoice via Ref No/Ref Date (when ERP provides it)
        if amt <= 0:
            ref_no = _raw_get(raw, "Ref No", "RefNo", "Against Bill No", "AgainstBillNo")
            ref_dt = _parse_erp_date(_raw_get(raw, "Ref Date", "RefDate", "Against Bill Date", "AgainstBillDate"))
            if ref_no and ref_dt:
                inv = inv_map.get((getattr(r, "party_code", ""), str(ref_no).strip(), ref_dt))
                if inv:
                    inv_eff = _invoice_effective_paid(inv)
                    if inv_eff > 0:
                        amt = inv_eff
                        allocated_invoices.add((getattr(inv, "party_code", ""), getattr(inv, "trans_no", ""), getattr(inv, "trans_date", None)))

        # 4) fallback B (conservative): if no reference is provided, and there is EXACTLY ONE
        #    fully/meaningfully paid invoice candidate for that party in the loaded snapshot_rows window,
        #    attribute it to this receipt.
        inferred_ref = None
        if amt <= 0:
            pc = getattr(r, "party_code", "") or ""
            candidates = []
            for inv in inv_by_party.get(pc, []):
                inv_key = (getattr(inv, "party_code", ""), getattr(inv, "trans_no", ""), getattr(inv, "trans_date", None))
                if inv_key in allocated_invoices:
                    continue
                inv_eff = _invoice_effective_paid(inv)
                if inv_eff <= 0:
                    continue
                # Only consider invoices dated not after the payment date (basic sanity)
                inv_dt = getattr(inv, "trans_date", None)
                if inv_dt and inv_dt > payment_date:
                    continue
                candidates.append((inv, inv_eff))

            if len(candidates) == 1:
                inv, inv_eff = candidates[0]
                amt = inv_eff
                inferred_ref = getattr(inv, "trans_no", "") or ""
                allocated_invoices.add((getattr(inv, "party_code", ""), getattr(inv, "trans_no", ""), getattr(inv, "trans_date", None)))

        if amt <= 0:
            # Nothing reliable to show
            continue

        inst = getattr(r, "instrument_no", None) or _raw_get(raw, "Instrument No", "InstrumentNo", "Cheque No", "ChequeNo")

        received.append({
            "party_code": getattr(r, "party_code", "") or _raw_get(raw, "Party Code", "PartyCode", default="") or "",
            "party_name": getattr(r, "party_name", "") or _raw_get(raw, "Party Name", "PartyName", default="") or "",
            "payment_date": payment_date,
            "instrument_no": inst or "",
            "ref_no": _raw_get(raw, "Ref No", "RefNo") or inferred_ref or getattr(r, "trans_no", ""),
            "amount": amt,
            "trans_type": getattr(r, "trans_type", ""),
        })

    return received

# =============================================================================
# TARGET DETAIL VIEW (FLOW MAINTAINED)
# =============================================================================

@login_required
def target_detail(request, pk):
    target = get_object_or_404(PaymentTargetWeek, pk=pk)

    # ✅ One snapshot lookup only (kept for Sections 1–3)
    snap = latest_snapshot_date()

    group = _norm_group(target.company_group)

    # ---------------------------------------------------------------------
    # Snapshot helpers (safe + fast)
    # ---------------------------------------------------------------------
    def _apply_group_filter_snapshot(qs):
        if group != "ALL":
            if _has_field(ReceivableSnapshotRow, "company_group"):
                return qs.filter(company_group=group)
            if _has_field(ReceivableSnapshotRow, "company_code"):
                return qs.filter(company_code=group)
        return qs

    def _snapshot_date_on_or_before(d):
        if not d:
            return None
        qs = ReceivableSnapshotRow.objects.filter(snapshot_date__lte=d)
        qs = _apply_group_filter_snapshot(qs)
        return qs.order_by("-snapshot_date").values_list("snapshot_date", flat=True).first()

    # For invoice-paid fallback/delta (does NOT change Sections 1–3 snap usage)
    snap_end = _snapshot_date_on_or_before(target.week_end) or snap
    snap_prev = _snapshot_date_on_or_before(target.week_start - timedelta(days=1))

    # ✅ Cache per target+week+group+snapshot (2 minutes)
    # Bump version to avoid old cached results
    cache_key = f"pt_detail::v11::{pk}::{snap_end}::{target.week_start}::{target.week_end}::{group}"
    cached_ctx = cache.get(cache_key)
    if cached_ctx is not None:
        return render(request, "accounts/targets/target_detail.html", cached_ctx)

    lines = list(
        PaymentTargetLine.objects
        .filter(target=target)
        .order_by("party_name", "invoice_no")
    )

    # 1) Target totals
    target_expected = sum(((l.expected_amount or Decimal("0")) for l in lines), Decimal("0"))
    target_os = sum(((l.outstanding_amount or Decimal("0")) for l in lines), Decimal("0"))

    # ---------------------------------------------------------------------
    # 2) Received Against Target (selected bills only)
    #    - Primary: receipts lookup for the week (wanted keys if supported)
    #    - Fallback: invoice snapshot delta between snapshots
    # ---------------------------------------------------------------------
    wanted_keys_pairs = [(l.party_code, l.invoice_no) for l in lines if l.party_code and l.invoice_no]

    # A) Primary receipts-based lookup (weekly)
    try:
        raw_paid_lookup = build_paid_lookup_for_period(
            company_group=(target.company_group or "ALL"),
            start_date=target.week_start,
            end_date=target.week_end,
            snapshot_date=snap,  # keep section 1–3 snapshot usage
            wanted_keys=wanted_keys_pairs,
        ) or {}
    except TypeError:
        # older signature
        raw_paid_lookup = build_paid_lookup_for_period(
            company_group=(target.company_group or "ALL"),
            start_date=target.week_start,
            end_date=target.week_end,
            snapshot_date=snap,
        ) or {}

    # Normalize keys into bill_key
    receipt_paid_lookup = {}
    for k, v in (raw_paid_lookup or {}).items():
        amt = _to_decimal(v, default=Decimal("0"))
        if amt == 0:
            continue

        if isinstance(k, (tuple, list)) and len(k) == 2:
            kk = _bill_key(k[0], k[1])
        else:
            ks = str(k or "").strip()
            if "||" not in ks:
                continue
            pc_part, inv_part = ks.split("||", 1)
            kk = _bill_key(pc_part, inv_part)

        receipt_paid_lookup[kk] = receipt_paid_lookup.get(kk, Decimal("0")) + amt

    # B) Fallback: invoice snapshot "paid during week" (delta between snapshots)
    inv_week_delta_lookup = {}

    party_set = {str(l.party_code).strip() for l in lines if l.party_code}
    invno_set = {str(l.invoice_no).strip() for l in lines if l.invoice_no}

    def _invoice_paid_map(snapshot_date):
        if not snapshot_date or not party_set or not invno_set:
            return {}

        qs = ReceivableSnapshotRow.objects.filter(
            snapshot_date=snapshot_date,
            party_code__in=party_set,
            trans_no__in=invno_set,
        ).only(
            "party_code",
            "trans_no",
            "trans_type",
            "bill_amt",
            "paid_amt",
            "outstanding_amt",
        )
        qs = _apply_group_filter_snapshot(qs)

        out = {}
        for sr in qs.iterator(chunk_size=2000):
            if not _is_invoice_type(sr.trans_type):
                continue

            kk = _bill_key(sr.party_code, sr.trans_no)

            bill = _to_decimal(getattr(sr, "bill_amt", None), default=Decimal("0"))
            paid = _to_decimal(getattr(sr, "paid_amt", None), default=Decimal("0"))
            os_ = _to_decimal(getattr(sr, "outstanding_amt", None), default=Decimal("0"))

            eff_paid = paid if paid > 0 else (bill - os_)
            if eff_paid > out.get(kk, Decimal("0")):
                out[kk] = eff_paid

        return out

    inv_paid_end = _invoice_paid_map(snap_end)
    inv_paid_prev = {}
    if snap_prev and snap_prev != snap_end:
        inv_paid_prev = _invoice_paid_map(snap_prev)

    for l in lines:
        kk = _bill_key(l.party_code, l.invoice_no)
        end_amt = _to_decimal(inv_paid_end.get(kk), default=Decimal("0"))
        prev_amt = _to_decimal(inv_paid_prev.get(kk), default=Decimal("0"))
        delta = end_amt - prev_amt
        if delta < 0:
            delta = Decimal("0")
        inv_week_delta_lookup[kk] = delta if delta > 0 else end_amt

    # C) Compute line-wise received
    received_against_target_total = Decimal("0")
    lines_with_received = []
    today = timezone.localdate()

    for l in lines:
        inv_dt = (
            _to_date(getattr(l, "invoice_date", None)) or
            _to_date(getattr(l, "trans_date", None)) or
            _to_date(getattr(l, "trs_date", None)) or
            _to_date(getattr(l, "bill_date", None))
        )
        due_dt = _to_date(getattr(l, "due_date", None))

        if hasattr(l, "invoice_date") and inv_dt:
            l.invoice_date = inv_dt
        if hasattr(l, "due_date") and due_dt:
            l.due_date = due_dt

        pay_terms_days = 0
        if inv_dt and due_dt:
            pay_terms_days = max(0, _days_between(inv_dt, due_dt))

        overdue_days = 0
        if due_dt:
            overdue_days = max(0, _days_between(due_dt, today))

        key = _bill_key(l.party_code, l.invoice_no)

        receipt_amt = _to_decimal(receipt_paid_lookup.get(key), default=Decimal("0"))
        fallback_amt = _to_decimal(inv_week_delta_lookup.get(key), default=Decimal("0"))
        received_amt = receipt_amt if receipt_amt > 0 else fallback_amt

        received_against_target_total += received_amt

        l.received_amount = received_amt
        l.balance_amount = (l.expected_amount or Decimal("0")) - received_amt

        lines_with_received.append({
            "line": l,
            "received_amount": received_amt,
            "balance_amount": l.balance_amount,
            "pay_terms_days": pay_terms_days,
            "overdue_days": overdue_days,
            "invoice_date": inv_dt,
            "due_date": due_dt,
        })

    target_balance = target_expected - received_against_target_total

    # ---------------------------------------------------------------------
    # 3) All Receipts in Week
    #    Primary: Code-2 fast functions
    #    Fallback: Code-1 enriched receipt scan if list is empty
    # ---------------------------------------------------------------------
    ws = target.week_start
    we = target.week_end

    week_received_total = (
        get_received_totals_for_period(
            company_group=(target.company_group or "ALL"),
            start_date=ws,
            end_date=we,
            snapshot_date=snap,
        ).get("received_total") or Decimal("0")
    )

    # ✅ Primary (Code-2): should show "All receipts in the week"
    week_receipts = get_received_rows_for_period(
        company_group=(target.company_group or "ALL"),
        start_date=ws,
        end_date=we,
        snapshot_date=snap,
        limit=2000,
    ) or []

    # ✅ Fallback (Code-1): if primary returns empty, rebuild from snapshot receipt rows
    if not week_receipts:
        window_start = ws - timedelta(days=365)

        receipt_source_rows = []
        try:
            receipt_source_rows = _fetch_snapshot_receipts_rows(
                snapshot_date=snap,
                company_group=(target.company_group or "ALL"),
                start_date=window_start,
                end_date=we,
            ) or []
        except Exception:
            receipt_source_rows = []

        if not receipt_source_rows:
            qs = ReceivableSnapshotRow.objects.filter(
                snapshot_date=snap,
                trans_date__range=(window_start, we),
            ).only(
                "party_code", "party_name",
                "trans_type", "trans_no", "trans_date",
                "instrument_no",
                "bill_amt", "paid_amt", "outstanding_amt",
                "raw",
            )
            qs = _apply_group_filter_snapshot(qs)
            receipt_source_rows = list(qs[:25000])  # safety cap

        week_receipts = enrich_received_rows(receipt_source_rows, ws, we)[:2000]

        # Normalize fields (date + amount) so template never gets blanks
        normed = []
        for r in (week_receipts or []):
            if not isinstance(r, dict):
                continue

            raw = r.get("raw") or {}
            if isinstance(raw, str):
                try:
                    import json
                    raw = json.loads(raw)
                except Exception:
                    raw = {}

            dt = (
                _to_date(r.get("date")) or
                _to_date(r.get("receipt_date")) or
                _to_date(r.get("rcpt_date")) or
                _to_date(r.get("trans_date")) or
                _to_date(raw.get("receipt_date")) or
                _to_date(raw.get("ReceiptDate")) or
                _to_date(raw.get("trans_date")) or
                _to_date(raw.get("TransDate")) or
                _to_date(raw.get("date")) or
                _to_date(raw.get("Date"))
            )

            amt = (
                _to_decimal(r.get("amount"), default=Decimal("0")) or
                _to_decimal(r.get("received"), default=Decimal("0")) or
                _to_decimal(r.get("received_amount"), default=Decimal("0")) or
                _to_decimal(r.get("rcvd_amt"), default=Decimal("0")) or
                _to_decimal(r.get("paid_amt"), default=Decimal("0")) or
                _to_decimal(r.get("bill_amt"), default=Decimal("0"))
            )

            if (amt or Decimal("0")) <= 0:
                amt = (
                    _to_decimal(raw.get("amount"), default=Decimal("0")) or
                    _to_decimal(raw.get("Amount"), default=Decimal("0")) or
                    _to_decimal(raw.get("paid_amt"), default=Decimal("0")) or
                    _to_decimal(raw.get("PaidAmt"), default=Decimal("0")) or
                    _to_decimal(raw.get("bill_amt"), default=Decimal("0")) or
                    _to_decimal(raw.get("BillAmt"), default=Decimal("0"))
                )

            if dt:
                r["date"] = dt
            if (amt or Decimal("0")) > 0:
                r["amount"] = amt

            if not r.get("company_name") and not r.get("company"):
                r["company_name"] = (target.company_group or "ALL")

            has_any = bool(r.get("party_name") or r.get("party_code") or r.get("invoice_no") or r.get("instrument_no"))
            if not has_any:
                continue

            if (r.get("date") is None) and (_to_decimal(r.get("amount"), default=Decimal("0")) <= 0):
                if not (r.get("invoice_no") or r.get("instrument_no")):
                    continue

            normed.append(r)

        week_receipts = normed[:2000]

    # If SQL total is 0 but list has values, use list sum
    if (week_received_total or Decimal("0")) <= 0 and week_receipts:
        week_received_total = sum(
            ((_to_decimal(x.get("amount"), default=Decimal("0")) for x in week_receipts)),
            Decimal("0")
        )

    # ---------------------------------------------------------------------
    # 4) Cheque/PDC entries — APPLICATION ONLY (unchanged)
    # ---------------------------------------------------------------------
    receivable_qs = (
        Receivable.objects
        .filter(
            Q(cheque_date__range=(ws, we)) |
            Q(due_date__range=(ws, we)) |
            Q(invoice_date__range=(ws, we)) |
            Q(entry_date__range=(ws, we)) |
            Q(created_at__date__range=(ws, we)) |
            Q(updated_at__date__range=(ws, we))
        )
        .only(
            "id",
            "customer_code", "customer_name",
            "invoice_number", "invoice_date", "due_date",
            "invoice_amount",
            "received_amount",
            "cheque_no", "cheque_date",
            "instrument_no",
            "status", "remarks",
        )
        .order_by("customer_name", "invoice_number", "id")
    )

    receivable_entries = []
    total_amt = Decimal("0")
    pdc_total = Decimal("0")
    wanted_inst = set()

    for rv in receivable_qs:
        remarks_l = (rv.remarks or "").lower()

        link_inst = _extract_link_instrument(rv)
        cheque_no = (rv.cheque_no or "").strip()
        cheque_date = getattr(rv, "cheque_date", None)

        is_pdc = ("pdc" in remarks_l) or (cheque_date is not None and cheque_date > today)

        amt = _safe_dec(getattr(rv, "invoice_amount", None), default=Decimal("0"))
        rcvd = _safe_dec(getattr(rv, "received_amount", None), default=Decimal("0"))

        bal_raw = (
            getattr(rv, "balance_amount", None) or
            getattr(rv, "balance_amt", None) or
            getattr(rv, "balance", None)
        )
        bal = _safe_dec(bal_raw, default=(amt - rcvd))

        total_amt += amt
        if is_pdc:
            pdc_total += amt

        inv_dt = _to_date(getattr(rv, "invoice_date", None))
        due_dt = _to_date(getattr(rv, "due_date", None))

        pay_terms_days = 0
        if inv_dt and due_dt:
            pay_terms_days = max(0, _days_between(inv_dt, due_dt))

        overdue_days = 0
        if due_dt:
            overdue_days = max(0, _days_between(due_dt, today))

        inst_norm = _norm_inst(link_inst or cheque_no)
        if inst_norm:
            wanted_inst.add(inst_norm)

        entry = {
            "company_name": (target.company_group or "ALL"),
            "party_code": rv.customer_code or "",
            "party_name": rv.customer_name or "",
            "invoice_no": rv.invoice_number or "",
            "invoice_date": inv_dt,
            "due_date": due_dt,

            "pay_terms_days": pay_terms_days,
            "overdue_days": overdue_days,

            "mode": "Cheque" if (cheque_no or cheque_date) else "N/A",
            "cheque_no": cheque_no,
            "instrument_no": link_inst,
            "cheque_date": cheque_date,

            "is_pdc": is_pdc,
            "pdc_date": cheque_date if is_pdc else None,

            "receipt_no": "",
            "receipt_date": "",
            "erp_matched": False,
            "erp_paid_total": Decimal("0"),
            "erp_receipts": [],
            "erp_receipt_value": Decimal("0"),

            "amount": amt,
            "received_amount": rcvd,
            "balance_amount": bal,

            "status": getattr(rv, "status", "") or "",
            "remarks": rv.remarks or "",
        }

        receivable_entries.append(entry)

    # Prefer all ERP rows if available; fallback to snapshot receipts scan
    erp_rows = []
    try:
        from ACCOUNTS.Receivable.services.receivables_targets import fetch_receivables_raw_all as _fetch_receivables_raw_all
        erp_rows = _fetch_receivables_raw_all() or []
    except Exception as e:
        logger.exception("fetch_receivables_raw_all failed in target_detail: %s", e)
        erp_rows = []

    has_any_inst = False
    for rr in (erp_rows or []):
        if _extract_instrument_from_receipt_row(rr):
            has_any_inst = True
            break

    if not has_any_inst:
        window_start = ws - timedelta(days=180)
        erp_rows = _fetch_snapshot_receipts_rows(
            snapshot_date=snap,
            company_group=(target.company_group or "ALL"),
            start_date=window_start,
            end_date=we,
        )

    filtered_rows = []
    for rr in (erp_rows or []):
        inst = _extract_instrument_from_receipt_row(rr)
        if not inst:
            continue
        if wanted_inst and inst not in wanted_inst:
            continue
        filtered_rows.append(rr)

    erp_pdc_index = _build_erp_pdc_lookup_from_rows(filtered_rows)

    for entry in receivable_entries:
        _attach_erp_to_receivable_entry(entry, erp_pdc_index)

        if entry.get("erp_matched"):
            erp_val = _to_decimal(entry.get("erp_paid_total"), default=Decimal("0"))
            entry["erp_receipt_value"] = erp_val if erp_val > 0 else _to_decimal(entry.get("amount"), default=Decimal("0"))

    ctx = {
        "target": target,
        "lines": lines,
        "lines_with_received": lines_with_received,
        "target_totals": {
            "expected": target_expected,
            "outstanding": target_os,
            "received": received_against_target_total,
            "balance": target_balance,
        },
        "week_receipts": week_receipts,
        "week_receipts_totals": {
            "received_total": week_received_total,
        },
        "receivable_entries": receivable_entries,
        "receivable_entries_totals": {
            "total": total_amt,
            "pdc_total": pdc_total,
        },
        "snapshot_date": snap,
        # optional debug/context
        "against_target_snap_end": snap_end,
        "against_target_snap_prev": snap_prev,
    }

    cache.set(cache_key, ctx, 120)
    return render(request, "accounts/targets/target_detail.html", ctx)


# ---------------------------------------------------------------------------
# Excel Export
# ---------------------------------------------------------------------------
# -------------------------
# helper (put near _as_float)
# -------------------------
def _to_date_safe(v):
    """
    Accepts: date/datetime/'yyyy-mm-dd'/'dd-mm-yyyy'/'dd/mm/yyyy'/'dd-Mon-yyyy'
    Returns: date or None
    """
    if not v:
        return None
    if isinstance(v, date) and not isinstance(v, datetime):
        return v
    if isinstance(v, datetime):
        return v.date()

    s = str(v).strip()
    # 1) ISO yyyy-mm-dd
    d = parse_date(s)
    if d:
        return d

    # 2) common formats
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%d-%b-%Y", "%d %b %Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass

    return None


def _pay_terms_days(invoice_date, due_date):
    inv = _to_date_safe(invoice_date)
    due = _to_date_safe(due_date)
    if not inv or not due:
        return ""
    return (due - inv).days


def _overdue_days(due_date, as_of=None):
    """
    Overdue days = max(0, as_of - due_date)
    Returns: int or "" when due date missing
    """
    due = _to_date_safe(due_date)
    if not due:
        return ""
    if as_of is None:
        as_of = timezone.localdate()
    as_of = _to_date_safe(as_of) or timezone.localdate()
    return max(0, (as_of - due).days)


def _build_paid_lookup_for_target_lines(*, target, lines, snapshot_date=None):
    """
    Excel must match UI totals logic:
      1) receipt/bank lookup (build_paid_lookup_for_period)
      2) snapshot invoice backfill for missing keys
    Returns: dict { BILLKEY: Decimal(received) }
    """
    snap = snapshot_date or latest_snapshot_date()

    # 1) Bank/receipt lookup (may be incomplete depending on ERP)
    raw_paid_lookup = build_paid_lookup_for_period(
        company_group=(target.company_group or "ALL"),
        start_date=target.week_start,
        end_date=target.week_end,
        snapshot_date=snap,   # ✅ keep consistent with other areas
    ) or {}

    paid_lookup = {}
    for k, v in raw_paid_lookup.items():
        amt = _to_decimal(v, default=Decimal("0"))
        if amt == 0:
            continue

        if isinstance(k, (tuple, list)) and len(k) == 2:
            kk = _bill_key(k[0], k[1])
        else:
            ks = str(k or "").strip()
            if "||" in ks:
                pc_part, inv_part = ks.split("||", 1)
                kk = _bill_key(pc_part, inv_part)
            else:
                kk = str(ks).strip().upper()

        paid_lookup[kk] = paid_lookup.get(kk, Decimal("0")) + amt

    # 2) Snapshot invoice backfill for missing keys
    #    This is the reliable "received = bill - outstanding" style fill.
    if snap and lines:
        wanted_pairs = [
            (str(l.party_code).strip(), str(l.invoice_no).strip())
            for l in lines
            if l.party_code and l.invoice_no
        ]
        wanted_keys = {_bill_key(pc, inv) for pc, inv in wanted_pairs}
        missing_keys = wanted_keys - set(paid_lookup.keys())

        if missing_keys:
            party_set = {pc for pc, _ in wanted_pairs if pc}
            invno_set = {inv for _, inv in wanted_pairs if inv}

            snap_qs = ReceivableSnapshotRow.objects.filter(
                snapshot_date=snap,
                party_code__in=party_set,
                trans_no__in=invno_set,
            ).only(
                "party_code", "trans_no", "trans_type",
                "bill_amt", "paid_amt", "outstanding_amt",
            )

            group = (target.company_group or "ALL").strip() or "ALL"
            if group != "ALL":
                if _has_field(ReceivableSnapshotRow, "company_group"):
                    snap_qs = snap_qs.filter(company_group=group)
                elif _has_field(ReceivableSnapshotRow, "company_code"):
                    snap_qs = snap_qs.filter(company_code=group)

            for sr in snap_qs:
                if not _is_invoice_type(getattr(sr, "trans_type", None)):
                    continue

                kk = _bill_key(sr.party_code, sr.trans_no)
                if kk not in missing_keys:
                    continue

                bill = _to_decimal(getattr(sr, "bill_amt", None), default=Decimal("0"))
                paid = _to_decimal(getattr(sr, "paid_amt", None), default=Decimal("0"))
                os_  = _to_decimal(getattr(sr, "outstanding_amt", None), default=Decimal("0"))

                eff_paid = paid if paid > 0 else (bill - os_)
                if eff_paid > 0:
                    # keep max if multiple rows found
                    paid_lookup[kk] = max(paid_lookup.get(kk, Decimal("0")), eff_paid)

    return paid_lookup


# ---------------------------------------------------------------------------
# Excel Export
# ---------------------------------------------------------------------------
@login_required
def target_detail_excel(request, pk):
    target = get_object_or_404(PaymentTargetWeek, pk=pk)
    today = timezone.localdate()

    lines = list(
        PaymentTargetLine.objects
        .filter(target=target)
        .order_by("party_name", "invoice_no")
    )

    # ✅ Build paid lookup consistent with your UI/target_list logic
    snap = latest_snapshot_date()
    paid_lookup = _build_paid_lookup_for_target_lines(target=target, lines=lines, snapshot_date=snap)

    against_rows = []
    for l in lines:
        key = _bill_key(l.party_code, l.invoice_no)
        received_amt = _to_decimal(paid_lookup.get(key), default=Decimal("0"))
        expected_amt = (l.expected_amount or Decimal("0"))

        pt_days = _pay_terms_days(l.invoice_date, l.due_date)
        od_days = _overdue_days(l.due_date, as_of=today)

        against_rows.append({
            "party_code": l.party_code,
            "party_name": l.party_name,
            "invoice_no": l.invoice_no,
            "invoice_date": l.invoice_date,
            "due_date": l.due_date,
            "pay_terms": pt_days,
            "overdue_days": od_days,
            "outstanding": (l.outstanding_amount or Decimal("0")),
            "expected": expected_amt,
            "received": received_amt,
            "balance": expected_amt - received_amt,
        })

    week_receipts = get_received_rows_for_period(
        company_group=(target.company_group or "ALL"),
        start_date=target.week_start,
        end_date=target.week_end,
    ) or []

    receivable_rows = get_receivable_entries_for_period(
        company_group=(target.company_group or "ALL"),
        start_date=target.week_start,
        end_date=target.week_end,
    ) or []

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Against Target"

    # ✅ Pay Terms + Overdue columns
    ws1.append([
        "Party Code", "Party Name", "Invoice No", "Invoice Date", "Due Date",
        "Pay Terms (Days)",
        "Overdue (Days)",
        "Outstanding", "Expected", "Received", "Balance"
    ])

    for r in against_rows:
        ws1.append([
            r["party_code"], r["party_name"], r["invoice_no"], r["invoice_date"], r["due_date"],
            r["pay_terms"],
            r["overdue_days"],
            _as_float(r["outstanding"]), _as_float(r["expected"]), _as_float(r["received"]), _as_float(r["balance"])
        ])

    ws2 = wb.create_sheet("All Receipts")
    ws2.append(["Company", "Party Code", "Party Name", "Invoice No", "Trans Date", "Paid Amount"])
    for r in week_receipts:
        ws2.append([
            r.get("company_name", ""),
            r.get("party_code", ""),
            r.get("party_name", ""),
            r.get("invoice_no", ""),
            r.get("trans_date", ""),
            _as_float(r.get("paid_amount")),
        ])

    ws3 = wb.create_sheet("Receivable Entry")

    # ✅ Pay Terms + Overdue columns
    ws3.append([
        "Party Code", "Party Name", "Invoice No", "Invoice Date", "Due Date",
        "Pay Terms (Days)",
        "Overdue (Days)",
        "Mode", "Cheque No", "Cheque Date", "PDC", "PDC Date",
        "Amount", "Received Amount", "Balance Amount", "Status", "Remarks"
    ])

    for r in receivable_rows:
        pt_days = _pay_terms_days(r.get("invoice_date"), r.get("due_date"))
        od_days = _overdue_days(r.get("due_date"), as_of=today)

        ws3.append([
            r.get("party_code", ""),
            r.get("party_name", ""),
            r.get("invoice_no", ""),
            r.get("invoice_date", ""),
            r.get("due_date", ""),
            pt_days,
            od_days,
            r.get("mode", ""),
            r.get("cheque_no", ""),
            r.get("cheque_date", ""),
            "PDC" if r.get("is_pdc") else "",
            r.get("pdc_date", ""),
            _as_float(r.get("amount")),
            _as_float(r.get("received_amount")),
            _as_float(r.get("balance_amount")),
            r.get("status", ""),
            r.get("remarks", ""),
        ])

    # Auto width
    for sheet in (ws1, ws2, ws3):
        for col in range(1, sheet.max_column + 1):
            max_len = 0
            col_letter = get_column_letter(col)
            for cell in sheet[col_letter]:
                v = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(v))
            sheet.column_dimensions[col_letter].width = min(max_len + 2, 50)

    filename = f"TargetDetail_{target.week_start:%Y%m%d}_{target.week_end:%Y%m%d}_T{target.pk}.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


# ---------------------------------------------------------------------------
# Select Bills page: READ-ONLY
# ---------------------------------------------------------------------------
@login_required
def target_select_bills(request, pk):
    target = get_object_or_404(PaymentTargetWeek, pk=pk)

    party_form = PaymentTargetSelectPartyForm(request.POST or None)
    qs = PaymentTargetLine.objects.filter(target=target).order_by("party_name", "invoice_no")

    if request.method == "POST":
        action = (request.POST.get("action") or "").strip().lower()

        if action == "save":
            messages.info(request, "Bills are selected from the Create Target page. Nothing to save here.")
            return redirect("accounts:payment_target_detail", pk=target.pk)

        if party_form.is_valid():
            party_code = (party_form.cleaned_data.get("party_code") or "").strip()
            party_name = (party_form.cleaned_data.get("party_name") or "").strip()

            if party_code:
                qs = qs.filter(party_code__iexact=party_code)
            elif party_name:
                qs = qs.filter(party_name__icontains=party_name)

    bills = list(qs)

    return render(request, "accounts/targets/target_select_bills.html", {
        "target": target,
        "party_form": party_form,
        "bills": bills,
        "readonly": True,
    })


@login_required
def target_report(request, pk):
    target = get_object_or_404(PaymentTargetWeek, pk=pk)
    lines = PaymentTargetLine.objects.filter(target=target).order_by("party_name", "invoice_no")

    party_rows = {}
    for l in lines:
        pr = party_rows.setdefault(l.party_code, {
            "party_code": l.party_code,
            "party_name": l.party_name,
            "expected": Decimal("0"),
            "bill_count": 0,
            "notes": [],
        })
        pr["expected"] += (l.expected_amount or Decimal("0"))
        pr["bill_count"] += 1
        if getattr(l, "discussion_notes", ""):
            pr["notes"].append(l.discussion_notes)

    totals = {
        "expected": sum(((l.expected_amount or Decimal("0")) for l in lines), Decimal("0")),
    }

    return render(request, "accounts/targets/target_report.html", {
        "target": target,
        "lines": lines,
        "party_rows": list(party_rows.values()),
        "totals": totals,
    })

# ------------------------------------------------------------------------- 
# accounts/views_party.py
# -------------------------------------------------------------------------
from django.contrib import messages

from .models import Party
from .forms import PartyForm, PartyContactFormSet


@login_required
def party_list(request):
    qs = Party.objects.all().order_by("party_name")
    q = (request.GET.get("q") or "").strip()

    if q:
        qs = qs.filter(party_name__icontains=q)

    return render(request, "accounts/masters/party_list.html", {
        "parties": qs,
        "q": q,
    })


@login_required
def party_create(request):
    if request.method == "POST":
        form = PartyForm(request.POST)
        formset = PartyContactFormSet(request.POST)

        if form.is_valid() and formset.is_valid():
            party = form.save()
            formset.instance = party
            formset.save()
            messages.success(request, "Party master created successfully.")
            return redirect("accounts:party_list")
    else:
        form = PartyForm()
        formset = PartyContactFormSet()

    return render(request, "accounts/masters/party_form.html", {
        "form": form,
        "formset": formset,
        "is_edit": False,
    })


@login_required
def party_edit(request, pk):
    party = get_object_or_404(Party, pk=pk)

    if request.method == "POST":
        form = PartyForm(request.POST, instance=party)
        formset = PartyContactFormSet(request.POST, instance=party)

        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            messages.success(request, "Party master updated successfully.")
            return redirect("accounts:party_list")
    else:
        form = PartyForm(instance=party)
        formset = PartyContactFormSet(instance=party)

    return render(request, "accounts/masters/party_form.html", {
        "form": form,
        "formset": formset,
        "is_edit": True,
        "party": party,
    })

# ---------------------------------------------------------------------------
# accounts/views_mail.py
# ---------------------------------------------------------------------------
from .models import OutgoingEmailAccount
from .forms import OutgoingEmailAccountForm

@login_required
def email_account_list(request):
    qs = OutgoingEmailAccount.objects.all().order_by("company_group")

    return render(request, "accounts/masters/email_account_list.html", {
        "accounts": qs,
    })


@login_required
def email_account_create(request):
    if request.method == "POST":
        form = OutgoingEmailAccountForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Sender email configured.")
            return redirect("accounts:email_account_list")
    else:
        form = OutgoingEmailAccountForm()

    return render(request, "accounts/masters/email_account_form.html", {
        "form": form,
        "is_edit": False,
    })

@login_required
def receivable_outstanding_all_pdc(request):
    """
    - Shows ALL outstanding bills as-of a date (snapshot <= as_on)
    - Attaches PDC details from application (Receivable)
    - Marks PDC status as CLEARED/PENDING based on ERP receipt existence (instrument match)
    - Adds: Outstanding min/max filter, overdue days, and line-wise remarks saving
    """

    # -----------------------------
    # Local safe helpers (isolated)
    # -----------------------------
    _INV_CLEAN_RE = re.compile(r"[^A-Z0-9]+")
    _INST_CLEAN_RE = re.compile(r"[^A-Z0-9]+")

    def _norm_party(v) -> str:
        return str(v or "").strip().upper()

    def _norm_inv(v) -> str:
        s = str(v or "").strip().upper()
        return _INV_CLEAN_RE.sub("", s) if s else ""

    def _bill_key(pc, inv) -> str:
        return f"{_norm_party(pc)}||{_norm_inv(inv)}"

    def _norm_inst(v) -> str:
        s = str(v or "").strip().upper()
        if not s:
            return ""
        if s.startswith("#"):
            s = s[1:]
        s = _INST_CLEAN_RE.sub("", s)
        if s.isdigit():
            s = s.lstrip("0") or "0"
        return s

    def _to_dec(v, default=Decimal("0")) -> Decimal:
        try:
            if v in (None, "", "NA", "N/A"):
                return default
            s = str(v).replace(",", "").strip()
            if not s:
                return default
            return Decimal(s)
        except Exception:
            return default

    def _to_date(d):
        if not d:
            return None
        try:
            return d.date() if hasattr(d, "date") else d
        except Exception:
            return None

    def _inst_db_variants(raw: str) -> set:
        raw = (str(raw or "")).strip()
        if not raw:
            return set()

        base = raw.lstrip("#").strip()
        clean = re.sub(r"[^A-Z0-9]", "", base.upper())

        out = {raw, base, f"#{base}", clean, f"#{clean}"}
        if clean.isdigit():
            nz = clean.lstrip("0") or "0"
            out.add(nz)
            out.add(f"#{nz}")
        return {x for x in out if x}

    def _pick_any(d: dict, keys, default=""):
        if not isinstance(d, dict):
            return default
        keymap = {str(k).strip().casefold(): k for k in d.keys()}
        for k in keys:
            kk = str(k).strip().casefold()
            real = keymap.get(kk)
            if real is None:
                continue
            val = d.get(real)
            if val not in (None, "", "NA", "N/A"):
                return val
        return default

    def _iter_raw_values(v):
        if isinstance(v, dict):
            for vv in v.values():
                yield from _iter_raw_values(vv)
        elif isinstance(v, (list, tuple, set)):
            for vv in v:
                yield from _iter_raw_values(vv)
        else:
            yield v

    def _tokenize_invoice_text(s) -> set:
        """
        Extract invoice/doc like tokens from raw text.
        Ignore short numeric tokens (often instrument).
        """
        txt = str(s or "").strip()
        if not txt:
            return set()

        for sep in [",", ";", "|", "\n", "\r", "\t"]:
            txt = txt.replace(sep, " ")

        out = set()
        for part in txt.split():
            n = _norm_inv(part)
            if not n:
                continue

            # ignore short purely numeric tokens (000070 etc)
            if n.isdigit() and len(n) < 8:
                continue

            if len(n) >= 10:
                out.add(n)
        return out

    def _erp_receipt_invoice_norms(raw) -> set:
        """
        Pull invoice/doc numbers that the receipt is settling against.
        Works with typical ERP keys: Doc No, Reference No, Against, Bill, Invoice, etc.
        """
        if not isinstance(raw, dict):
            return set()

        key_hints = ("doc", "reference", "ref", "invoice", "bill", "against", "settle", "adjust")
        out = set()

        for k, v in raw.items():
            ks = str(k or "").strip().casefold()
            if not any(h in ks for h in key_hints):
                continue
            for vv in _iter_raw_values(v):
                out |= _tokenize_invoice_text(vv)

        return out

    def _chunked(iterable, size: int):
        lst = list(iterable)
        for i in range(0, len(lst), size):
            yield lst[i:i + size]

    # -----------------------------
    # Filters (GET)
    # -----------------------------
    as_on_ui = (request.GET.get("as_on") or "").strip()
    as_on = parse_date(as_on_ui) if as_on_ui else None
    if not as_on:
        as_on = timezone.localdate()

    company_group = (request.GET.get("company_group") or "ALL").strip().upper() or "ALL"
    party_code = (request.GET.get("party_code") or "").strip()
    party_name = (request.GET.get("party_name") or "").strip()
    qtext = (request.GET.get("q") or "").strip()

    os_min_ui = (request.GET.get("os_min") or "").strip()
    os_max_ui = (request.GET.get("os_max") or "").strip()
    os_min = _to_dec(os_min_ui, default=Decimal("0")) if os_min_ui else None
    os_max = _to_dec(os_max_ui, default=Decimal("0")) if os_max_ui else None

    # -----------------------------
    # Snapshot date <= as_on  (needed for GET + POST)
    # -----------------------------
    snap = (
        ReceivableSnapshotRow.objects
        .filter(snapshot_date__lte=as_on)
        .aggregate(d=Max("snapshot_date"))["d"]
    )

    # -----------------------------
    # POST: Save remarks (line-wise)
    # -----------------------------
    if request.method == "POST" and request.POST.get("action") == "save_remarks":
        snap_post = parse_date((request.POST.get("snapshot_date") or "").strip()) or None
        snap_for_save = snap_post or snap
        return_qs = request.POST.get("return_qs") or ""

        try:
            row_count = int(request.POST.get("row_count") or "0")
        except Exception:
            row_count = 0

        if not snap_for_save:
            messages.error(request, "Snapshot date missing. Please refresh the page and try again.")
            return redirect(request.path + (("?" + return_qs) if return_qs else ""))

        changed = 0
        cleared = 0

        with transaction.atomic():
            for i in range(row_count):
                pc = (request.POST.get(f"party_code_{i}") or "").strip()
                inv = (request.POST.get(f"invoice_no_{i}") or "").strip()
                rem = (request.POST.get(f"remark_{i}") or "").strip()
                comp = (request.POST.get(f"company_name_{i}") or "").strip()
                pn = (request.POST.get(f"party_name_{i}") or "").strip()

                if not pc or not inv:
                    continue

                # blank => clear remark
                if rem == "":
                    deleted, _ = ReceivableOutstandingRemark.objects.filter(
                        snapshot_date=snap_for_save,
                        party_code=pc,
                        invoice_no=inv,
                    ).delete()
                    if deleted:
                        cleared += 1
                    continue

                obj, created = ReceivableOutstandingRemark.objects.get_or_create(
                    snapshot_date=snap_for_save,
                    party_code=pc,
                    invoice_no=inv,
                    defaults={
                        "company_name": comp,
                        "party_name": pn,
                        "remark": rem,
                        "created_by": request.user,
                        "updated_by": request.user,
                    },
                )

                if created:
                    changed += 1
                else:
                    new_comp = comp or obj.company_name
                    new_pn = pn or obj.party_name
                    if (obj.remark or "") != rem or obj.company_name != new_comp or obj.party_name != new_pn or obj.updated_by_id != request.user.id:
                        obj.remark = rem
                        obj.company_name = new_comp
                        obj.party_name = new_pn
                        obj.updated_by = request.user
                        obj.save(update_fields=["remark", "company_name", "party_name", "updated_by", "updated_at"])
                        changed += 1

        if changed or cleared:
            messages.success(request, f"Remark saved/updated for {changed} row(s), cleared for {cleared} row(s).")
        else:
            messages.info(request, "No changes to save.")
        return redirect(request.path + (("?" + return_qs) if return_qs else ""))

    # -----------------------------
    # If no snapshot available
    # -----------------------------
    if not snap:
        messages.error(request, "Snapshot not available. Please run snapshot sync.")
        return render(request, "accounts/receivables/outstanding_all_pdc.html", {
            "as_on": as_on,
            "snapshot_date": None,
            "page_obj": None,
            "rows": [],
            "summary": {"count": 0, "total_outstanding": Decimal("0")},
            "filters": {
                "as_on": as_on.strftime("%Y-%m-%d"),
                "company_group": company_group,
                "party_code": party_code,
                "party_name": party_name,
                "q": qtext,
                "os_min": os_min_ui,
                "os_max": os_max_ui,
            },
            "company_group_options": ["ALL"],
        })

    # -----------------------------
    # A) Base: ALL outstanding bills from snapshot
    # -----------------------------
    base_qs = (
        ReceivableSnapshotRow.objects
        .filter(snapshot_date=snap, outstanding_amt__gt=0)
        .exclude(trans_type__icontains="receipt")
    )

    try:
        base_qs = _apply_company_group_filter_snapshot(base_qs, company_group)
    except Exception:
        pass

    if party_code:
        base_qs = base_qs.filter(party_code__iexact=party_code)
    if party_name:
        base_qs = base_qs.filter(party_name__icontains=party_name)
    if qtext:
        base_qs = base_qs.filter(
            Q(party_name__icontains=qtext)
            | Q(party_code__icontains=qtext)
            | Q(trans_no__icontains=qtext)
        )

    base_qs = base_qs.only(
        "company_name", "party_code", "party_name",
        "trans_no", "trans_date",
        "due_date", "overdue_date",
        "bill_amt", "outstanding_amt",
        "erp_lid", "trans_type",
    ).order_by("party_name", "party_code", "trans_no", "erp_lid")

    # De-dup: Keep MAX outstanding per invoice key.
    seen = {}
    rows = []
    for r in base_qs.iterator(chunk_size=2000):
        pc = (r.party_code or "").strip()
        inv = (r.trans_no or "").strip()
        if not pc or not inv:
            continue

        k = _bill_key(pc, inv)
        bill_amt = _to_dec(getattr(r, "bill_amt", None), default=Decimal("0"))
        os_amt = _to_dec(getattr(r, "outstanding_amt", None), default=Decimal("0"))
        due = getattr(r, "due_date", None) or getattr(r, "overdue_date", None)

        if k not in seen:
            obj = {
                "company_name": getattr(r, "company_name", "") or "",
                "party_code": pc,
                "party_name": (r.party_name or "").strip(),
                "invoice_no": inv,
                "invoice_date": getattr(r, "trans_date", None),
                "due_date": due,
                "bill_amount": bill_amt,
                "outstanding_amount": os_amt,

                "overdue_days": 0,

                "user_remark": "",
                "remark_updated_at": None,
                "remark_updated_by": "",

                "pdc_ref": "",
                "pdc_date": None,
                "pdc_amount": Decimal("0"),
                "pdc_entry_date": None,
                "pdc_status": "",
                "pdc_erp_receipt_no": "",
                "pdc_erp_receipt_date": "",
                "pdc_erp_inst_no": "",
                "pdc_count": 0,
            }
            seen[k] = obj
            rows.append(obj)
        else:
            obj = seen[k]
            if os_amt > (obj["outstanding_amount"] or Decimal("0")):
                obj["outstanding_amount"] = os_amt
            if bill_amt > (obj["bill_amount"] or Decimal("0")):
                obj["bill_amount"] = bill_amt
            if not obj.get("due_date") and due:
                obj["due_date"] = due

    if not rows:
        return render(request, "accounts/receivables/outstanding_all_pdc.html", {
            "as_on": as_on,
            "snapshot_date": snap,
            "page_obj": None,
            "rows": [],
            "summary": {"count": 0, "total_outstanding": Decimal("0")},
            "filters": {
                "as_on": as_on.strftime("%Y-%m-%d"),
                "company_group": company_group,
                "party_code": party_code,
                "party_name": party_name,
                "q": qtext,
                "os_min": os_min_ui,
                "os_max": os_max_ui,
            },
            "company_group_options": ["ALL"],
        })

    # -----------------------------
    # B) Outstanding min/max filter
    # -----------------------------
    if os_min is not None:
        rows = [r for r in rows if (r.get("outstanding_amount") or Decimal("0")) >= os_min]
    if os_max is not None:
        rows = [r for r in rows if (r.get("outstanding_amount") or Decimal("0")) <= os_max]

    if not rows:
        return render(request, "accounts/receivables/outstanding_all_pdc.html", {
            "as_on": as_on,
            "snapshot_date": snap,
            "page_obj": None,
            "rows": [],
            "summary": {"count": 0, "total_outstanding": Decimal("0")},
            "filters": {
                "as_on": as_on.strftime("%Y-%m-%d"),
                "company_group": company_group,
                "party_code": party_code,
                "party_name": party_name,
                "q": qtext,
                "os_min": os_min_ui,
                "os_max": os_max_ui,
            },
            "company_group_options": ["ALL"],
        })

    # -----------------------------
    # C) Pull PDC entries from application (Receivable)
    # -----------------------------
    open_keys = {_bill_key(r["party_code"], r["invoice_no"]) for r in rows}
    party_set = {r["party_code"] for r in rows if r.get("party_code")}
    inv_set = {r["invoice_no"] for r in rows if r.get("invoice_no")}

    app_qs = (
        Receivable.objects
        .filter(customer_code__in=list(party_set), entry_date__lte=as_on)
        .exclude(Q(cheque_no__isnull=True) | Q(cheque_no=""))
    )

    if len(inv_set) <= 4000:
        app_qs = app_qs.filter(invoice_number__in=list(inv_set))

    app_qs = app_qs.only(
        "id", "customer_code", "customer_name",
        "invoice_number", "invoice_date", "due_date",
        "invoice_amount", "received_amount",
        "cheque_no", "cheque_date",
        "entry_date", "remarks",
    ).order_by("-entry_date", "-id")

    pdc_by_bill = defaultdict(list)
    wanted_inst = set()
    wanted_party_norm = set()

    for p in app_qs.iterator(chunk_size=2000):
        pc = (getattr(p, "customer_code", "") or "").strip()
        inv = (getattr(p, "invoice_number", "") or "").strip()
        if not pc or not inv:
            continue

        k = _bill_key(pc, inv)
        if k not in open_keys:
            continue

        pdc_by_bill[k].append(p)

        inst = (getattr(p, "cheque_no", "") or "").strip()
        if inst:
            wanted_inst.add(inst)
            wanted_party_norm.add(_norm_party(pc))

    # -----------------------------
    # D) ERP receipt lookup ONCE (SAFE + CHUNKED)
    # -----------------------------
    cleared_lookup = {}        # (party, inst) -> latest receipt (also stores inv_norms)
    cleared_lookup_inv = {}    # (party, inst, inv_norm) -> latest receipt

    if wanted_inst:
        wanted_db_values = set()
        for inst in wanted_inst:
            wanted_db_values |= _inst_db_variants(inst)

        # SQL Server safety: chunk OR conditions (avoid 2100 params)
        OR_CHUNK = 250

        for batch in _chunked(sorted(wanted_db_values), OR_CHUNK):
            qq = Q()
            for v in batch:
                qq |= Q(instrument_no__iexact=v)

            erp_receipts_qs = (
                ReceivableSnapshotRow.objects
                .filter(snapshot_date=snap)
                .filter(qq)
                .filter(trans_type__icontains="receipt")
                .exclude(instrument_no="")
                .only("party_code", "instrument_no", "trans_date", "trans_date_display", "trans_no", "paid_amt", "raw")
            )

            if wanted_party_norm:
                erp_receipts_qs = erp_receipts_qs.filter(party_code__in=list(wanted_party_norm))

            for sr in erp_receipts_qs.iterator(chunk_size=2000):
                pc_norm = _norm_party(getattr(sr, "party_code", "") or "")
                inst_raw = (getattr(sr, "instrument_no", "") or "").strip()
                inst_norm = _norm_inst(inst_raw)
                if not pc_norm or not inst_norm:
                    continue

                raw = getattr(sr, "raw", {}) or {}
                inv_norms = _erp_receipt_invoice_norms(raw)

                receipt_no = (
                    _pick_any(raw, ["Trans No", "Receipt No", "Voucher No", "Vch No", "Number"], default="")
                    or (getattr(sr, "trans_no", "") or "").strip()
                )

                dt_disp = (getattr(sr, "trans_date_display", "") or "").strip()
                if not dt_disp and getattr(sr, "trans_date", None):
                    try:
                        td = sr.trans_date.date() if hasattr(sr.trans_date, "date") else sr.trans_date
                        dt_disp = td.strftime("%d-%b-%Y") if td else ""
                    except Exception:
                        dt_disp = ""

                paid = _to_dec(getattr(sr, "paid_amt", None), default=Decimal("0"))
                cur_td = getattr(sr, "trans_date", None)

                item = {
                    "receipt_no": receipt_no,
                    "receipt_date": dt_disp,
                    "inst_raw": inst_raw,
                    "paid": paid,
                    "_td": cur_td,
                    "inv_norms": inv_norms,
                }

                # strict: party + inst + invoice
                if inv_norms:
                    for invn in inv_norms:
                        k3 = (pc_norm, inst_norm, invn)
                        prev = cleared_lookup_inv.get(k3)
                        prev_td = prev.get("_td") if prev else None
                        if (not prev) or (cur_td and (not prev_td or cur_td > prev_td)):
                            cleared_lookup_inv[k3] = dict(item)

                # base: party + inst
                k2 = (pc_norm, inst_norm)
                prev2 = cleared_lookup.get(k2)
                prev2_td = prev2.get("_td") if prev2 else None
                if (not prev2) or (cur_td and (not prev2_td or cur_td > prev2_td)):
                    cleared_lookup[k2] = dict(item)
                else:
                    if prev2 and inv_norms:
                        prev2.setdefault("inv_norms", set())
                        prev2["inv_norms"] |= inv_norms

    # -----------------------------
    # E) Attach PDC + overdue days
    # -----------------------------
    for r in rows:
        due_d = _to_date(r.get("due_date"))
        if due_d:
            dd = (as_on - due_d).days
            r["overdue_days"] = dd if dd > 0 else 0
        else:
            r["overdue_days"] = 0

        k = _bill_key(r["party_code"], r["invoice_no"])
        plist = pdc_by_bill.get(k) or []
        if not plist:
            continue

        p = plist[0]
        r["pdc_count"] = len(plist)
        r["pdc_ref"] = (getattr(p, "cheque_no", "") or "").strip()
        r["pdc_date"] = getattr(p, "cheque_date", None)
        r["pdc_amount"] = _to_dec(getattr(p, "received_amount", None), default=Decimal("0"))
        r["pdc_entry_date"] = getattr(p, "entry_date", None)

        inst_norm = _norm_inst(r["pdc_ref"])
        pc_norm = _norm_party(getattr(p, "customer_code", "") or r["party_code"])
        inv_norm = _norm_inv(r.get("invoice_no"))

        hit = None

        # 1) strict: party + inst + invoice
        if inv_norm:
            hit = cleared_lookup_inv.get((pc_norm, inst_norm, inv_norm))

        # 2) safe fallback: party + inst only if ERP receipt has no invoice refs OR invoice matches
        if not hit:
            cand = cleared_lookup.get((pc_norm, inst_norm))
            if cand:
                cand_invs = cand.get("inv_norms") or set()
                if (not cand_invs) or (inv_norm and inv_norm in cand_invs):
                    hit = cand

        if hit:
            r["pdc_status"] = "CLEARED"
            r["pdc_erp_receipt_no"] = hit.get("receipt_no") or ""
            r["pdc_erp_receipt_date"] = hit.get("receipt_date") or ""
            r["pdc_erp_inst_no"] = hit.get("inst_raw") or ""
        else:
            r["pdc_status"] = "PENDING"

    # -----------------------------
    # Summary + Pagination
    # -----------------------------
    total_outstanding = sum((r.get("outstanding_amount") or Decimal("0") for r in rows), Decimal("0"))
    paginator = Paginator(rows, 50)
    page_obj = paginator.get_page(request.GET.get("page"))
    page_rows = list(page_obj.object_list) if page_obj else []

    # -----------------------------
    # F) Load saved remarks for only current page rows
    # -----------------------------
    if page_rows:
        q = Q()
        for rr in page_rows:
            q |= Q(party_code=rr["party_code"], invoice_no=rr["invoice_no"])

        rem_qs = (
            ReceivableOutstandingRemark.objects
            .filter(snapshot_date=snap)
            .filter(q)
            .select_related("updated_by")
        )

        rem_map = {(x.party_code, x.invoice_no): x for x in rem_qs}

        for rr in page_rows:
            obj = rem_map.get((rr["party_code"], rr["invoice_no"]))
            if obj:
                rr["user_remark"] = obj.remark or ""
                rr["remark_updated_at"] = getattr(obj, "updated_at", None)
                rr["remark_updated_by"] = getattr(obj.updated_by, "username", "") if getattr(obj, "updated_by", None) else ""
            else:
                rr["user_remark"] = ""
                rr["remark_updated_at"] = None
                rr["remark_updated_by"] = ""

    # company group options
    opts = ["ALL"]
    try:
        opts += sorted(list(COMPANY_GROUPS.keys()))
        if "OTHER" not in opts:
            opts.append("OTHER")
    except Exception:
        pass

    return render(request, "accounts/receivables/outstanding_all_pdc.html", {
        "as_on": as_on,
        "snapshot_date": snap,
        "page_obj": page_obj,
        "rows": page_rows,
        "summary": {"count": len(rows), "total_outstanding": total_outstanding},
        "filters": {
            "as_on": as_on.strftime("%Y-%m-%d"),
            "company_group": company_group,
            "party_code": party_code,
            "party_name": party_name,
            "q": qtext,
            "os_min": os_min_ui,
            "os_max": os_max_ui,
        },
        "company_group_options": opts,
    })
