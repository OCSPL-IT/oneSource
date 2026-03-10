# ──────────────────────────────────────────────
# Standard Library
# ──────────────────────────────────────────────
import json
import logging
from calendar import monthrange
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from time import perf_counter
from django.utils import timezone
from xlsxwriter.utility import xl_col_to_name
# ──────────────────────────────────────────────
# Third-Party / Django
# ──────────────────────────────────────────────
import pandas as pd
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db import connections, transaction
from django.db.models import (
    Sum, F, Q, Value, DecimalField, Count, Exists, OuterRef,
)
from django.db.models.functions import Coalesce, Greatest
from django.http import HttpResponse, HttpResponseBadRequest, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import NoReverseMatch, reverse_lazy
from django.utils.safestring import mark_safe
from django.views.decorators.http import require_GET, require_http_methods, require_POST
from django.views.generic import CreateView, DeleteView, DetailView, ListView, UpdateView

# ──────────────────────────────────────────────
# Local Apps
# ──────────────────────────────────────────────
from STORE.forms import AllocationForm, MaterialRequestForm, VehicleForm
from STORE.models import (
    GrnLineCache,
    IssueLineCache,
    MaterialRequest,
    Pallet,
    Rack,
    RackAllocation,
    RackIssue,
    Vehicle,
)
from STORE.services import (
    apply_issue_fifo,
    get_erp_alias,
    sync_grn,
    sync_issues,
    transfer_allocations,
    TransferError,
)

# ---------------------------------------------------------
# Constants
# ---------------------------------------------------------
ALLOWED_ITEM_TYPES = ("Key Raw Material", "Packing Material", "Raw Material")
DEC18_3 = DecimalField(max_digits=18, decimal_places=3)
ZERO_DEC = Value(Decimal("0.000"), output_field=DEC18_3)

# ──────────────────────────────────────────────
# Logging Setup
# ──────────────────────────────────────────────
log = logging.getLogger(__name__)           # module logger
logger = logging.getLogger("custom_logger") # retained if used elsewhere


def search_supplier(request):
    query = request.GET.get('term', '')
    logger.debug(f"Searching suppliers with term: '{query}'")

    sql_query = """
        SELECT DISTINCT SUPP.sName 
        FROM TXNHDR HDR
        LEFT JOIN BUSMST AS SUPP ON HDR.lAccId1 = SUPP.lId
        WHERE HDR.ltypid IN (
            400, 509, 520, 524, 750, 751, 752, 753, 
            754, 755, 756, 757, 758, 759, 760, 761, 
            762, 763, 764, 765, 766, 767, 768, 769, 956,654,547
        ) AND SUPP.sName LIKE %s
    """

    with connections['readonly_db'].cursor() as cursor:
        cursor.execute(sql_query, [f'%{query}%'])
        results = cursor.fetchall()

    # print("Supplier Results:", results)  # Debugging line

    suppliers = [{'id': row[0], 'text': row[0]} for row in results]
    logger.info(f"Found {len(suppliers)} suppliers for search '{query}'")
    return JsonResponse(suppliers, safe=False)


def search_item(request):
    query = request.GET.get('term', '')
    logger.debug(f"Searching items with term: '{query}'")
    sql_query = """
        SELECT DISTINCT ITM.sName 
        FROM TXNHDR HDR
        INNER JOIN TXNDET AS DET ON HDR.lId = DET.lId
        INNER JOIN ITMMST AS ITM ON DET.lItmId = ITM.lId
        WHERE HDR.ltypid IN (
            400, 509, 520, 524, 750, 751, 752, 753, 
            754, 755, 756, 757, 758, 759, 760, 761, 
            762, 763, 764, 765, 766, 767, 768, 769, 956)
			AND DET.lItmTyp IN (57, 66, 77, 60) AND ITM.sName LIKE %s
    """

    with connections['readonly_db'].cursor() as cursor:
        cursor.execute(sql_query, [f'%{query}%'])
        results = cursor.fetchall()

    # print("Item Results:", results)  # Debugging line

    items = [{'id': row[0], 'text': row[0]} for row in results]
    logger.info(f"Found {len(items)} items for search '{query}'")
    return JsonResponse(items, safe=False)



@login_required
def add_vehicle(request):
    user_groups = request.user.groups.values_list('name', flat=True)  # Check if the user is in STORE group
    is_superuser = request.user.is_superuser

    if not request.user.has_perm('STORE.add_vehicle'):
        logger.warning(f"Unauthorized add_vehicle access by user: {request.user.username}")
        messages.error(request, "You do not have permission to add vehicle records.")
        return redirect('indexpage')
        
    if request.method == 'POST':
        form = VehicleForm(request.POST)
        try:
            if form.is_valid():
                form.save()
                logger.info(f"Vehicle added by user: {request.user.username}")
                messages.success(request, "Vehicle entry saved successfully!")
                return redirect('add_vehicle')
            else:
                messages.error(request, "Please fill in the required fields.")
                logger.warning(f"Invalid vehicle form submitted by {request.user.username}")
        except Exception as e:
            logger.exception(f"Error while saving vehicle form by {request.user.username}")
            messages.error(request, "An error occurred while saving the vehicle entry.")

    else:
        form = VehicleForm()
    return render(request, 'store/add_vehicle_form.html', locals())



@login_required
def vehicle_list(request):
    user_groups = request.user.groups.values_list('name', flat=True)
    is_superuser = request.user.is_superuser

    if not request.user.has_perm('STORE.view_vehicle'):
        logger.warning(f"Unauthorized vehicle_list access by user: {request.user.username}")
        messages.error(request, "You do not have permission to view vehicle records.")
        return redirect('indexpage')

    can_edit_vehicle = request.user.has_perm('STORE.change_vehicle')
    can_delete_vehicle = request.user.has_perm('STORE.delete_vehicle')
    can_add_vehicle = request.user.has_perm('STORE.add_vehicle')
    can_view_vehicle = request.user.has_perm('STORE.view_vehicle')

    # --- Filters ---
    from_date = request.GET.get("from_date", "")
    to_date   = request.GET.get("to_date", "")

    filters = {
        'name_of_supplier': request.GET.get('name_of_supplier', ''),
        'material': request.GET.get('material', ''),
        'name_of_transporter': request.GET.get('name_of_transporter', ''),
        'status': request.GET.get('status', ''),
    }

    vehicles = Vehicle.objects.all().order_by('-reporting_date')

    # Apply date range filter
    if from_date and to_date:
        vehicles = vehicles.filter(reporting_date__range=[from_date, to_date])
    elif from_date:
        vehicles = vehicles.filter(reporting_date__gte=from_date)
    elif to_date:
        vehicles = vehicles.filter(reporting_date__lte=to_date)

    # Apply other filters
    for key, value in filters.items():
        if value:
            vehicles = vehicles.filter(**{f"{key}__icontains": value})

    # Pagination
    paginator = Paginator(vehicles, 10)  # 10 records per page
    page = request.GET.get('page')

    try:
        vehicles_paginated = paginator.page(page)
    except PageNotAnInteger:
        vehicles_paginated = paginator.page(1)
    except EmptyPage:
        vehicles_paginated = paginator.page(paginator.num_pages)
    # ---- NEW: compute pending days for each vehicle ----
    today = timezone.localdate()

    for v in vehicles_paginated:
        v.pending_days = ""          # default empty
        if (v.status or "").lower() == "pending" and v.reporting_date:
            # use report_time if present, otherwise midnight
            rep_time = v.report_time or time(0, 0)
            reported_dt = datetime.combine(v.reporting_date, rep_time)
            days = (today - reported_dt.date()).days
            # don't show negative numbers
            v.pending_days = max(days, 0)
            
    context = {
        'vehicles': vehicles_paginated,
        'from_date': from_date,
        'to_date': to_date,
        **filters,
        'user_groups': user_groups,
        'is_superuser': is_superuser,
        'can_edit_vehicle': can_edit_vehicle,
        'can_delete_vehicle': can_delete_vehicle,
        'can_add_vehicle': can_add_vehicle,
        'can_view_vehicle': can_view_vehicle,
    }
    logger.info(
        f"Vehicle list viewed by {request.user.username} "
        f"with date range {from_date or 'NA'} to {to_date or 'NA'} and filters {filters}"
    )
    return render(request, 'store/vehicle_list.html', context)

@login_required
def edit_vehicle(request, vehicle_id):
    user_groups = request.user.groups.values_list('name', flat=True)  # Check if the user is in HR group
    is_superuser = request.user.is_superuser

    """ Edit vehicle details (Permission Required: STORE.change_vehicle) """
    if not request.user.has_perm('STORE.change_vehicle'):
        logger.warning(f"Unauthorized edit attempt by {request.user.username} on vehicle ID {vehicle_id}")
        messages.error(request, "You do not have permission to edit vehicle records.")
        return redirect('indexpage')

    try:
        vehicle = get_object_or_404(Vehicle, id=vehicle_id)
    except Exception as e:
        logger.exception(f"Vehicle ID {vehicle_id} not found for editing by {request.user.username}")
        messages.error(request, "Vehicle not found.")
        return redirect('vehicle_list')
    
    if request.method == "POST":
        try:
            form = VehicleForm(request.POST, instance=vehicle)
            if form.is_valid():
                form.save()
                logger.info(f"Vehicle ID {vehicle_id} edited by {request.user.username}")
                messages.success(request, "Vehicle details updated successfully!")
                return redirect('vehicle_list')
        except Exception as e:
            logger.exception(f"Error while updating vehicle ID {vehicle_id} by {request.user.username}")
            messages.error(request, "An error occurred while updating the vehicle.")
    else:
        form = VehicleForm(instance=vehicle)
    return render(request, 'store/edit_vehicle.html', locals())

@login_required
def delete_vehicle(request, vehicle_id):
    user_groups = request.user.groups.values_list('name', flat=True)  # Check if the user is in HR group
    is_superuser = request.user.is_superuser

    """ Delete a vehicle entry (Permission Required: STORE.delete_vehicle) """
    if not request.user.has_perm('STORE.delete_vehicle'):
        logger.warning(f"Unauthorized delete attempt by {request.user.username} on vehicle ID {vehicle_id}")
        messages.error(request, "You do not have permission to delete vehicle records.")
        return redirect('indexpage')

    if request.method == "POST":  # Ensuring deletion via POST request
        vehicle = get_object_or_404(Vehicle, id=vehicle_id)
        vehicle.delete()
        logger.info(f"Vehicle ID {vehicle_id} deleted by {request.user.username}")
        messages.success(request, "Vehicle deleted successfully!")  # Optional success message
        return redirect('vehicle_list')  # Redirect to the vehicle list page
    
    messages.error(request, "Invalid request method!")  # Optional error message
    return redirect('vehicle_list')  # Redirect even if the method is incorrect


#Edit function
@login_required
def view_vehicle(request, vehicle_id):
    user_groups = request.user.groups.values_list('name', flat=True)  # Check if the user is in HR group
    is_superuser = request.user.is_superuser

    """ View vehicle details (Permission Required: STORE.view_vehicle) """
    if not request.user.has_perm('STORE.view_vehicle'):
        logger.warning(f"Unauthorized vehicle view by {request.user.username} on ID {vehicle_id}")
        messages.error(request, "You do not have permission to view vehicle records.")
        return redirect('indexpage')
    
    logger.info(f"Vehicle ID {vehicle_id} viewed by {request.user.username}")
    vehicle = get_object_or_404(Vehicle, id=vehicle_id)
    return render(request, 'store/view_vehicle.html', locals())



#excel download
@login_required
def vehicle_download_excel(request):
    if not request.user.has_perm('STORE.view_vehicle'):
        logger.warning(f"Unauthorized Excel download by {request.user.username}")
        messages.error(request, "You do not have permission to download vehicle records.")
        return redirect('indexpage')

    # ---- Filters ----
    from_date = request.GET.get("from_date", "")
    to_date   = request.GET.get("to_date", "")

    filters = {
        'name_of_supplier': request.GET.get('name_of_supplier', ''),
        'material': request.GET.get('material', ''),
        'name_of_transporter': request.GET.get('name_of_transporter', ''),
        'status': request.GET.get('status', ''),
    }

    vehicles = Vehicle.objects.all()

    # Apply date range filter
    if from_date and to_date:
        vehicles = vehicles.filter(reporting_date__range=[from_date, to_date])
    elif from_date:
        vehicles = vehicles.filter(reporting_date__gte=from_date)
    elif to_date:
        vehicles = vehicles.filter(reporting_date__lte=to_date)

    # Apply other filters
    for key, value in filters.items():
        if value:
            vehicles = vehicles.filter(**{f"{key}__icontains": value})

    # ---- Data for Excel ----
    base_fields = [
        'reporting_date',
        'invoice',
        'name_of_supplier',
        'material',
        'unit',
        'qty',
        'report_time',
        'unloading_date',
        'unloading_time',
        'unloading_days',
        'vehicle_no',
        'name_of_transporter',
        'status',
        'manufacture',
        'remark',
    ]

    # Convert to list of dicts so we can add a computed column
    rows = list(vehicles.values(*base_fields))

    # ✅ NEW: compute pending_days per row (same logic as list page)
    today = timezone.localdate()
    for row in rows:
        pending = ""
        status = (row.get("status") or "").lower()
        reporting_date = row.get("reporting_date")

        if status == "pending" and reporting_date:
            # we only need the date difference; time is optional
            days = (today - reporting_date).days
            pending = max(days, 0)

        row["pending_days"] = pending

    try:
        df = pd.DataFrame(rows)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=vehicle_records.xlsx'

        with pd.ExcelWriter(response, engine='xlsxwriter') as writer:
            sheet_name = 'Vehicles'
            start_row = 2  # leave space for title
            df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row)

            workbook  = writer.book
            worksheet = writer.sheets[sheet_name]

            # ---- Title formatting ----
            title_fmt = workbook.add_format({
                'bold': True, 'font_size': 14,
                'align': 'center', 'valign': 'vcenter',
                'bg_color': '#D9E1F2'  # light blue background
            })

            # Merge across all columns for title
            ncols = len(df.columns) if not df.empty else 16  # now includes pending_days
            last_col_letter = xl_col_to_name(ncols - 1)
            worksheet.merge_range(f"A1:{last_col_letter}1", "Vehicle Records", title_fmt)

            # ---- Header formatting ----
            header_fmt = workbook.add_format({
                'bold': True, 'align': 'center',
                'valign': 'vcenter', 'bg_color': '#BDD7EE',
                'border': 1
            })

            # Apply header format
            for col_num, value in enumerate(df.columns.values):
                worksheet.write(start_row, col_num, value, header_fmt)

            # Freeze header row
            worksheet.freeze_panes(start_row + 1, 0)

            # Optional: set column widths nicely
            worksheet.set_column(0, ncols - 1, 18)

        logger.info(
            f"{vehicles.count()} vehicle records exported by {request.user.username} "
            f"(from={from_date or 'NA'}, to={to_date or 'NA'}, filters={filters})"
        )
        return response
    except Exception:
        logger.exception(f"Error generating Excel for vehicle records by {request.user.username}")
        messages.error(request, "An error occurred while generating the Excel file.")
        return redirect('vehicle_list')
    

def vehicle_chart_report(request):
    user_groups      = request.user.groups.values_list('name', flat=True)
    is_superuser     = request.user.is_superuser
    today            = date.today()
    year             = int(request.GET.get('year', today.year))
    month            = int(request.GET.get('month', today.month))
    period           = request.GET.get('period', 'monthly')
    material_filter  = request.GET.get('material', '')
    supplier_filter  = request.GET.get('supplier', '')

    # Base queryset
    qs = Vehicle.objects.filter(
        reporting_date__year=year,
        reporting_date__month=month
    )
    if material_filter:
        qs = qs.filter(material=material_filter)
    if supplier_filter:
        qs = qs.filter(name_of_supplier=supplier_filter)

    #
    # 1) Build the Material chart data
    #
    materials = list(
        qs.values_list('material', flat=True)
          .distinct()
          .order_by('material')
    )
    # get counts per material
    mat_counts = (
        qs.values('material')
          .annotate(count=Count('id'))
          .order_by('material')
    )
    mat_map = {row['material']: row['count'] for row in mat_counts}

    labels   = materials
    datasets = [{
        'label': f"Month {month:02d}/{year}",
        'data':  [mat_map.get(m, 0) for m in materials],
        'backgroundColor': 'rgba(54, 162, 235, 0.6)',
        'borderColor':     'rgba(54, 162, 235, 1)',
        'borderWidth': 1,
    }]

    #
    # 2) Build the Supplier chart data (bar-only)
    #
    suppliers = list(
        qs.values_list('name_of_supplier', flat=True)
          .distinct()
          .order_by('name_of_supplier')
    )

    if period == 'fortnightly':
        first_qs  = qs.filter(reporting_date__day__lte=15)
        second_qs = qs.filter(reporting_date__day__gte=16)

        first_counts  = (
            first_qs.values('name_of_supplier')
                    .annotate(count=Count('id'))
                    .order_by('name_of_supplier')
        )
        second_counts = (
            second_qs.values('name_of_supplier')
                     .annotate(count=Count('id'))
                     .order_by('name_of_supplier')
        )

        first_map  = {r['name_of_supplier']: r['count'] for r in first_counts}
        second_map = {r['name_of_supplier']: r['count'] for r in second_counts}

        supplier_labels   = suppliers
        supplier_datasets = [
            {
                'label': f"Days 1–15 ({month:02d}/{year})",
                'data':  [first_map.get(s, 0) for s in suppliers],
            },
            {
                'label': f"Days 16–{monthrange(year,month)[1]} ({month:02d}/{year})",
                'data':  [second_map.get(s, 0) for s in suppliers],
            },
        ]
    else:
        # monthly
        sup_counts = (
            qs.values('name_of_supplier')
              .annotate(count=Count('id'))
              .order_by('name_of_supplier')
        )
        sup_map = {r['name_of_supplier']: r['count'] for r in sup_counts}

        supplier_labels   = suppliers
        supplier_datasets = [
            {
                'label': f"Month {month:02d}/{year}",
                'data':  [sup_map.get(s, 0) for s in suppliers],
            }
        ]

    # --- dropdown choices ---
    month_choices    = [(m, date(2000, m, 1).strftime('%B')) for m in range(1, 13)]
    year_choices     = [today.year - 1, today.year, today.year + 1]
    material_choices = list(
        Vehicle.objects.filter(reporting_date__year=year,
                               reporting_date__month=month)
               .values_list('material', flat=True)
               .distinct().order_by('material')
    )
    supplier_choices = list(
        Vehicle.objects.filter(reporting_date__year=year,
                               reporting_date__month=month)
               .values_list('name_of_supplier', flat=True)
               .distinct().order_by('name_of_supplier')
    )

    return render(request, 'store/vehicle_chart_report.html', {
        # Material chart
        'labels':            labels,
        'datasets':          datasets,
        # Supplier chart
        'supplier_labels':   supplier_labels,
        'supplier_datasets': supplier_datasets,
        # Filters
        'year':              year,
        'month':             month,
        'period':            period,
        'material_choices':  material_choices,
        'selected_material': material_filter,
        'supplier_choices':  supplier_choices,
        'selected_supplier': supplier_filter,
        'month_choices':     month_choices,
        'year_choices':      year_choices,
        # auth
        'user_groups':       user_groups,
        'is_superuser':      is_superuser,
    })



@login_required
def material_list(request):
    """
    List all Dispatch Plan entries.
    """
    qs = MaterialRequest.objects.all().order_by('-created_at')
    logger.info(f"{request.user} viewed material list ({qs.count()} items)")
    return render(request, 'store/material_list.html', {
        'requests': qs,
    })


@login_required
def material_create(request):
    
    """
    Create a new Dispatch Plan.
    """
    if request.method == 'POST':
        form = MaterialRequestForm(request.POST)
        if form.is_valid():
            obj = form.save()
            messages.success(request, "Dispatch Plan saved successfully.")
            logger.info(f"{request.user} created MaterialRequest #{obj.pk}")
            return redirect('material-list')
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = MaterialRequestForm()

    return render(request, 'store/material_form.html', {
        'form': form,
        'view': request,  # so template can do {% if view.object %} safely
    })


@login_required
def material_detail(request, pk):
    """
    Show a single Dispatch Plan.
    """
    obj = get_object_or_404(MaterialRequest, pk=pk)
    logger.info(f"{request.user} viewed MaterialRequest #{pk}")
    return render(request, 'store/material_detail.html', {
        'object': obj,
    })


@login_required
def material_edit(request, pk):
    """
    Edit an existing Dispatch Plan.
    """
    obj = get_object_or_404(MaterialRequest, pk=pk)

    if request.method == 'POST':
        form = MaterialRequestForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            messages.success(request, "Dispatch Plan updated successfully.")
            logger.info(f"{request.user} updated MaterialRequest #{pk}")
            return redirect('material-list')
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = MaterialRequestForm(instance=obj)

    return render(request, 'store/material_form.html', {
        'form': form,
        'view': request,  # template uses view.object to know “edit” vs “new”
    })


@login_required
def material_delete(request, pk):
    """
    Confirm & delete a Dispatch Plan.
    GET  → show confirm page
    POST → delete and redirect
    """
    obj = get_object_or_404(MaterialRequest, pk=pk)

    if request.method == 'POST':
        obj.delete()
        messages.success(request, "Dispatch Plan deleted.")
        logger.info(f"{request.user} deleted MaterialRequest #{pk}")
        return redirect('material-list')

    return render(request, 'store/material_confirm_delete.html', {
        'object': obj,
    })

    
# ---------------------------------------------------------
# Rack_Store RM — Optimised (same behaviour / flow)
# ---------------------------------------------------------
# ---------------------------------------------------------
# Small helpers (DRY + faster)
# ---------------------------------------------------------

def _coalesce_sum(field: str):
    """Coalesced Sum(field) -> Decimal(18,3)."""
    return Coalesce(Sum(field, output_field=DEC18_3), ZERO_DEC, output_field=DEC18_3)

def _annotate_balance(qs, allocated_field="allocated_qty", consumed_field="consumptions__qty"):
    """
    Add allocated/consumed/balance to a queryset of RackAllocation (or values grouped to it).
    Fields are the same ones you used earlier; behaviour unchanged.
    """
    return (
        qs.annotate(
            allocated=_coalesce_sum(allocated_field),
            consumed=_coalesce_sum(consumed_field),
        )
        .annotate(balance=F("allocated") - F("consumed"))
    )

def _blocked_pallet_ids(rack_id: int | None = None):
    """
    Return set of pallet IDs whose net balance > 0.
    Optional rack filter to reduce work when we only need one rack.
    """
    base = RackAllocation.objects
    if rack_id is not None:
        base = base.filter(pallet__rack_id=rack_id)

    qs = (
        base.values("pallet_id")
        .annotate(
            allocated=_coalesce_sum("allocated_qty"),
            consumed=_coalesce_sum("consumptions__qty"),
        )
        .annotate(balance=F("allocated") - F("consumed"))
        .filter(pallet_id__isnull=False, balance__gt=0)
        .values_list("pallet_id", flat=True)
    )
    return set(qs)

def _redir_to_rack_dashboard():
    try:
        return redirect("rack_dashboard")
    except NoReverseMatch:
        return redirect("STORE:rack_dashboard")

def _parse_ymd(s: str | None) -> date | None:
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        return None

def _parse_on_date(s: str | None) -> date | None:
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None

# ---------------------------------------------------------
# Dashboard
# ---------------------------------------------------------

@login_required
def rack_dashboard(request):
    """Live balances by rack & item (unchanged output)."""

    # Existing rows (same fields/order, just DRY balances)
    rows = (
        RackAllocation.objects
        .select_related("rack", "grn", "pallet")
        .filter(grn__pr_item_type__in=ALLOWED_ITEM_TYPES)
        .values(
            "rack__id", "rack__code",
            "pallet__id", "pallet__number",
            "grn__item_code", "grn__item_name", "grn__batch_no", "grn__uom",
        )
    )
    rows = _annotate_balance(rows).order_by("rack__code", "grn__item_code")

    # Racks/pallet list for UI
    racks_qs = Rack.objects.filter(is_active=True).order_by("code").values("id", "code")
    pallets_qs = (
        Pallet.objects.filter(is_active=True)
        .order_by("number")
        .values("id", "number", "rack_id")
    )

    # Blocked pallets (net balance > 0)
    blocked_ids = _blocked_pallet_ids()

    pallets_by_rack: dict[int, list[dict]] = {}
    for p in pallets_qs:
        pallets_by_rack.setdefault(p["rack_id"], []).append({
            "id": p["id"],
            "number": p["number"],
            "blocked": p["id"] in blocked_ids,
        })

    rack_data = [
        {"id": r["id"], "code": r["code"], "pallets": pallets_by_rack.get(r["id"], [])}
        for r in racks_qs
    ]
    rack_data_json = mark_safe(json.dumps(rack_data))

    return render(
        request,
        "store/rack_dashboard.html",
        {
            "rows": rows,
            "rack_data_json": rack_data_json,
        },
    )

# ---------------------------------------------------------
# Allocate single GRN line
# ---------------------------------------------------------

@login_required
@require_http_methods(["GET", "POST"])
def grn_allocate(request, erp_line_id: str):
    """Allocate a single GRN line to a rack/pallet (behaviour unchanged)."""
    grn = get_object_or_404(GrnLineCache, erp_line_id=erp_line_id)

    done = grn.allocations.aggregate(total=_coalesce_sum("allocated_qty"))["total"] or Decimal("0.000")
    remaining = max(grn.qty - done, Decimal("0.000"))

    # ✅ Collect currently allocated rack/pallet combinations for this GRN
    used_allocs = (
        grn.allocations
        .select_related("rack", "pallet")
        .values("rack__id", "rack__code", "pallet__id", "pallet__number")
        .order_by("rack__code")
    )

    # BUGFIX + faster grouping
    used_racks = sorted({a["rack__id"] for a in used_allocs if a["rack__id"]})
    used_pallets: dict[str, list[int]] = {}
    for a in used_allocs:
        rid = a["rack__id"]
        pid = a["pallet__id"]
        if rid and pid:
            used_pallets.setdefault(str(rid), set()).add(pid)
    used_pallets = {k: sorted(v) for k, v in used_pallets.items()}

    if request.method == "POST":
        form = AllocationForm(request.POST)
        if form.is_valid():
            qty = form.cleaned_data["allocated_qty"]
            if qty <= 0 or qty > remaining:
                messages.error(request, f"Invalid quantity. Remaining allocatable: {remaining:.3f}")
            else:
                alloc = form.save(commit=False)
                alloc.grn = grn
                alloc.save()
                messages.success(request, f"Allocated {qty:.3f} to {alloc.rack.code}")

                nxt = (request.GET.get("next") or request.POST.get("next")) or ""
                if nxt:
                    try:
                        return redirect(nxt)
                    except Exception:
                        messages.warning(request, "Could not follow return URL; going to dashboard.")
                return _redir_to_rack_dashboard()
    else:
        form = AllocationForm()

    return render(
        request,
        "store/grn_allocate.html",
        {
            "grn": grn,
            "form": form,
            "remaining": remaining,
            "used_racks": used_racks,
            "used_pallets": used_pallets,
        },
    )

# ---------------------------------------------------------
# ERP Sync page (same UX)
# ---------------------------------------------------------

@login_required
@require_http_methods(["GET", "POST"])
def sync_page(request):
    erp_alias = get_erp_alias()

    if request.method == "POST":
        f_raw = (request.POST.get("from") or "").strip()
        t_raw = (request.POST.get("to") or "").strip()

        from_date = _parse_ymd(f_raw) or date.today().replace(day=1)
        to_date   = _parse_ymd(t_raw) or date.today()

        if from_date > to_date:
            messages.error(request, "Sync failed: From-date cannot be after To-date.")
            return _redir_to_rack_dashboard()

        # Connection test (get DB name for messaging)
        try:
            with connections[erp_alias].cursor() as c:
                c.execute("SELECT DB_NAME()")
                dbname = c.fetchone()[0]
        except Exception as e:
            messages.error(request, f"ERP connection failed ({erp_alias}): {e}")
            return _redir_to_rack_dashboard()

        try:
            t0 = perf_counter()
            n_grn = sync_grn(from_date, to_date, alias=erp_alias) or 0
            n_iss = sync_issues(from_date, to_date, alias=erp_alias) or 0
            dt = perf_counter() - t0
        except Exception as e:
            messages.error(request, f"Sync failed: {e}")
            return _redir_to_rack_dashboard()

        if n_grn == 0 and n_iss == 0:
            messages.warning(
                request,
                f"No ERP rows found in {dbname} for {from_date:%d-%b-%Y} to {to_date:%d-%b-%Y}."
            )
        else:
            messages.success(
                request,
                f"Synced {n_grn} GRN and {n_iss} Issue rows from {dbname} "
                f"({erp_alias}) in {dt:.1f}s."
            )
        return _redir_to_rack_dashboard()

    return render(request, "store/sync.html", {})

# ---------------------------------------------------------
# Apply FIFO issue (unchanged)
# ---------------------------------------------------------

@login_required
def issue_apply(request, erp_line_id: str):
    issue = get_object_or_404(IssueLineCache, erp_line_id=erp_line_id)
    prefer = request.GET.get("rack") or None
    created = apply_issue_fifo(issue, prefer_rack_code=prefer)
    messages.success(request, f"Applied issue {issue.issue_no} – split into {len(created)} rack lines.")
    return _redir_to_rack_dashboard()

# ---------------------------------------------------------
# AJAX pallet options (now uses shared helper)
# ---------------------------------------------------------

@login_required
@require_GET
def pallet_options(request):
    rack_id_raw = (request.GET.get("rack") or "").strip()
    if not rack_id_raw.isdigit():
        return HttpResponseBadRequest("Invalid rack id")
    rack_id = int(rack_id_raw)

    blocked_ids = _blocked_pallet_ids(rack_id=rack_id)

    pallets = (
        Pallet.objects
        .filter(rack_id=rack_id, is_active=True)
        .order_by("number")
        .values_list("id", "number")
    )

    parts = ['<option value="">(No pallet)</option>']
    for pid, num in pallets:
        blocked = pid in blocked_ids
        label = f'{num}{" — Blocked" if blocked else ""}'
        disabled = ' disabled="disabled"' if blocked else ''
        parts.append(f'<option value="{pid}"{disabled}>{label}</option>')

    return HttpResponse("\n".join(parts), content_type="text/html")

# ---------------------------------------------------------
# Allocate-by-date (same UI; DRY internals)
# ---------------------------------------------------------

@login_required
@require_GET
def grn_allocate_date(request):
    """
    List GRN lines for a given date with rack+pallet blocking logic.
    Query params:
      on=YYYY-MM-DD | DD-MM-YYYY | DD/MM/YYYY
      q= free text (GRN no, item code/name, supplier)
      unallocated=1
    """
    on_raw = request.GET.get("on") or ""
    on = _parse_on_date(on_raw)
    qtext = (request.GET.get("q") or "").strip()
    only_unalloc = (request.GET.get("unallocated") == "1")

    grns = []
    if on:
        qs = (
            GrnLineCache.objects
            .filter(doc_date=on, pr_item_type__in=ALLOWED_ITEM_TYPES)
            .select_related()
            .annotate(
                allocated=_coalesce_sum("allocations__allocated_qty"),
            )
            .annotate(
                remaining=Greatest(
                    F("qty") - F("allocated"),
                    ZERO_DEC,
                    output_field=DEC18_3,
                ),
            )
        )

        if qtext:
            qs = qs.filter(
                Q(doc_no__icontains=qtext) |
                Q(item_code__icontains=qtext) |
                Q(item_name__icontains=qtext) |
                Q(supplier_name__icontains=qtext)
            )

        if only_unalloc:
            qs = qs.filter(remaining__gt=0)

        grns = qs.order_by("doc_no", "item_code")

    racks_qs = Rack.objects.filter(is_active=True).order_by("code").values("id", "code")
    pallets_qs = Pallet.objects.filter(is_active=True).order_by("number").values("id", "number", "rack_id")

    blocked_ids = _blocked_pallet_ids()

    pallets_by_rack: dict[int, list[dict]] = {}
    for p in pallets_qs:
        pallets_by_rack.setdefault(p["rack_id"], []).append({
            "id": p["id"],
            "number": p["number"],
            "blocked": p["id"] in blocked_ids,
        })

    rack_data = [
        {"id": r["id"], "code": r["code"], "pallets": pallets_by_rack.get(r["id"], [])}
        for r in racks_qs
    ]
    rack_data_json = mark_safe(json.dumps(rack_data))

    return render(
        request,
        "store/grn_allocate_date.html",
        {
            "grns": grns,
            "rack_list": racks_qs,
            "today": date.today().isoformat(),
            "on_value": (on and on.strftime("%Y-%m-%d")) or (on_raw if on_raw else ""),
            "rack_data_json": rack_data_json,
        },
    )

# ---------------------------------------------------------
# Consume group (FIFO within rack) — trimmed imports, same logic
# ---------------------------------------------------------

@login_required
@require_POST
def rack_consume_group(request):
    rack_code   = (request.POST.get("rack_code") or "").strip()
    pallet_num  = (request.POST.get("pallet_number") or "").strip()
    item_code   = (request.POST.get("item_code") or "").strip()
    batch_no    = (request.POST.get("batch_no") or "").strip()
    qty_raw     = (request.POST.get("qty") or "").strip()

    if not rack_code or not item_code:
        messages.error(request, "Rack and Item Code are required.")
        return _redir_to_rack_dashboard()

    try:
        need = Decimal(qty_raw)
    except (InvalidOperation, TypeError):
        messages.error(request, "Invalid quantity.")
        return _redir_to_rack_dashboard()
    if need <= 0:
        messages.error(request, "Quantity must be greater than zero.")
        return _redir_to_rack_dashboard()

    q = (
        RackAllocation.objects
        .select_related("grn", "rack", "pallet")
        .filter(rack__code=rack_code, grn__item_code=item_code)
        .order_by("grn__doc_date", "created_at", "pk")
    )
    # Optional pallet
    if pallet_num:
        q = q.filter(pallet__number=pallet_num)
    else:
        q = q.filter(pallet__isnull=True)

    # Optional batch ("—" in table means empty string)
    if batch_no and batch_no != "—":
        q = q.filter(grn__batch_no=batch_no)
    else:
        q = q.filter(grn__batch_no="")

    agg = q.aggregate(
        allocated=_coalesce_sum("allocated_qty"),
        consumed=_coalesce_sum("consumptions__qty"),
    )
    available = (agg["allocated"] or Decimal("0.000")) - (agg["consumed"] or Decimal("0.000"))

    if available <= 0:
        messages.error(request, f"No balance available on rack {rack_code}.")
        return _redir_to_rack_dashboard()
    if need > available:
        messages.error(request, f"Not enough balance. Available: {available:.3f}")
        return _redir_to_rack_dashboard()

    today = date.today()
    manual_issue_no = f"MANUAL-{today:%Y%m%d}"
    issue, _ = IssueLineCache.objects.get_or_create(
        issue_no=manual_issue_no,
        issue_date=today,
        item_code=item_code,
        batch_no=(batch_no if batch_no != "—" else ""),
        defaults=dict(
            company_id=0,
            year_id=0,
            cost_center="",
            item_name="",
            uom="",
            qty=Decimal("0.000"),
            warehouse="",
            erp_line_id=f"{manual_issue_no}-{rack_code}-{item_code}-{today:%H%M%S}",
        ),
    )

    created = []
    left = need
    for alloc in q:
        bal = alloc.balance_qty
        if bal <= 0:
            continue
        take = bal if bal < left else left
        created.append(RackIssue(issue=issue, allocation=alloc, qty=take))
        left -= take
        if left <= 0:
            break

    if not created:
        messages.error(request, "Could not create consumption lines.")
        return _redir_to_rack_dashboard()

    with transaction.atomic():
        RackIssue.objects.bulk_create(created)

    messages.success(
        request,
        f"Consumed {need:.3f} from Rack {rack_code}"
        + (f" / Pallet {pallet_num}" if pallet_num else "")
        + (f" / Batch {batch_no}" if batch_no and batch_no != '—' else "")
        + " (FIFO within rack)."
    )
    return _redir_to_rack_dashboard()

# ---------------------------------------------------------
# Rack → Rack transfer (unchanged behaviour)
# ---------------------------------------------------------

@login_required
@require_POST
def rack_transfer(request):
    def _v(name):  # quick fetch/strip
        return (request.POST.get(name) or "").strip()

    from_rack = _v("from_rack_code")
    to_rack   = _v("to_rack_code")
    item_code = _v("item_code")
    batch_no  = _v("batch_no") or None
    src_pal   = _v("from_pallet_number") or None
    dst_pal   = _v("to_pallet_number") or None

    try:
        qty = Decimal(_v("qty"))
    except (InvalidOperation, ValueError):
        messages.error(request, "Invalid quantity.")
        return _redir_to_rack_dashboard()

    if not from_rack or not to_rack or not item_code:
        messages.error(request, "From/To rack and Item Code are required.")
        return _redir_to_rack_dashboard()

    if from_rack == to_rack and (src_pal or "") == (dst_pal or ""):
        messages.error(request, "Source and destination are identical.")
        return _redir_to_rack_dashboard()

    try:
        fragments = transfer_allocations(
            item_code=item_code,
            batch_no=batch_no,
            from_rack_code=from_rack,
            to_rack_code=to_rack,
            qty=qty,
            from_pallet_number=src_pal,
            to_pallet_number=dst_pal,
        )
    except TransferError as e:
        messages.error(request, f"Transfer failed: {e}")
        return _redir_to_rack_dashboard()
    except Exception as e:
        messages.error(request, f"Unexpected error: {e}")
        return _redir_to_rack_dashboard()

    moved = sum((f["moved"] for f in fragments), Decimal("0.000"))
    messages.success(
        request,
        f"Transferred {moved:.3f} of {item_code}"
        + (f" / batch {batch_no}" if batch_no else "")
        + f" from {from_rack}{' '+src_pal if src_pal else ''} to {to_rack}{' '+dst_pal if dst_pal else ''}."
    )
    return _redir_to_rack_dashboard()



# STORE/views.py
from datetime import datetime, date
from io import BytesIO

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone

import xlsxwriter

# import your functions
from .dispatch_report import fetch_alpha_map, build_dispatch_grid


# ───────────────────────────────────────────────────────────────
# Helpers
# ───────────────────────────────────────────────────────────────
def _parse_ymd(s: str | None) -> date | None:
    if not s:
        return None
    try:
        return datetime.strptime(str(s).strip(), "%Y-%m-%d").date()
    except Exception:
        return None


def _n0(x) -> float:
    try:
        return float(x or 0.0)
    except Exception:
        return 0.0


# ───────────────────────────────────────────────────────────────
# Excel Export (same structure as HTML table)
# ───────────────────────────────────────────────────────────────
def _export_dispatch_plan_vs_actual_xlsx(*, disp_grid: dict, period: str, fr: date, to: date) -> HttpResponse:
    output = BytesIO()
    wb = xlsxwriter.Workbook(output, {"in_memory": True})
    ws = wb.add_worksheet("Dispatch")

    is_daily = (period or "").upper() == "DAILY"
    labels = disp_grid.get("months", []) or []
    rows = disp_grid.get("rows", []) or []
    footer_months = (disp_grid.get("footer", {}) or {}).get("months", []) or []
    footer_stock = _n0((disp_grid.get("footer", {}) or {}).get("stock_qty", 0))

    # ---------- Formats ----------
    fmt_title = wb.add_format({"bold": True, "font_size": 14})
    fmt_sub = wb.add_format({"font_size": 10, "font_color": "#64748b"})
    fmt_strip = wb.add_format({"bold": True, "bg_color": "#F1F5F9", "border": 1})

    fmt_hdr_top = wb.add_format({"bold": True, "align": "center", "valign": "vcenter", "bg_color": "#F8FAFC", "border": 1})
    fmt_hdr_mid = wb.add_format({"bold": True, "align": "center", "valign": "vcenter", "bg_color": "#F8FAFC", "border": 1, "font_color": "#64748b"})
    fmt_hdr_left = wb.add_format({"bold": True, "align": "left", "valign": "vcenter", "bg_color": "#F8FAFC", "border": 1})

    fmt_plan_h = wb.add_format({"bold": True, "align": "center", "valign": "vcenter", "bg_color": "#F0F9FF", "font_color": "#0284c7", "border": 1})
    fmt_est_h  = wb.add_format({"bold": True, "align": "center", "valign": "vcenter", "bg_color": "#F5F3FF", "font_color": "#7c3aed", "border": 1})
    fmt_act_h  = wb.add_format({"bold": True, "align": "center", "valign": "vcenter", "bg_color": "#FFF7ED", "font_color": "#c2410c", "border": 1})
    fmt_rate_h = wb.add_format({"bold": True, "align": "center", "valign": "vcenter", "bg_color": "#F0FDF4", "font_color": "#15803d", "border": 1})

    fmt_txt_bold = wb.add_format({"border": 1, "bold": True})
    fmt_qty = wb.add_format({"border": 1, "num_format": "#,##0", "align": "right"})
    fmt_amt = wb.add_format({"border": 1, "num_format": "#,##0", "align": "right"})
    fmt_rate = wb.add_format({"border": 1, "num_format": "#,##0.00", "align": "right"})

    fmt_tot = wb.add_format({"border": 1, "bold": True, "bg_color": "#F1F5F9", "num_format": "#,##0", "align": "right"})
    fmt_tot_rate = wb.add_format({"border": 1, "bold": True, "bg_color": "#F1F5F9", "num_format": "#,##0.00", "align": "right"})
    fmt_tot_lbl = wb.add_format({"border": 1, "bold": True, "bg_color": "#F1F5F9", "align": "right"})

    # ---------- Column widths ----------
    ws.set_column(0, 0, 34)   # Product/alias
    ws.set_column(1, 300, 14)

    # ---------- Title ----------
    r = 0
    ws.write(r, 0, "Dispatch Analysis", fmt_title)
    ws.write(r + 1, 0, f"Plan vs Actual Performance | Period: {period}", fmt_sub)
    ws.write(r + 2, 0, f"Date Range: {fr.strftime('%d %b %Y')} → {to.strftime('%d %b %Y')}", fmt_sub)
    r += 4

    # ---------- Grand totals strip (stays visible) ----------
    grand_plan = sum(_n0(f.get("plan")) for f in footer_months)
    grand_est = sum(_n0(f.get("est")) for f in footer_months)
    grand_act = sum(_n0(f.get("actual")) for f in footer_months)
    grand_val = sum(_n0(f.get("actual_value")) for f in footer_months)
    grand_rate = (grand_val / grand_act) if grand_act else 0.0

    ws.merge_range(r, 0, r, 2, "Grand Totals (overall)", fmt_strip)
    if is_daily:
        ws.write(r, 3, "Plan Qty", fmt_strip);   ws.write_number(r, 4, grand_plan, fmt_tot)
        ws.write(r, 5, "Actual Qty", fmt_strip); ws.write_number(r, 6, grand_act, fmt_tot)
        ws.write(r, 7, "Actual Val", fmt_strip); ws.write_number(r, 8, grand_val, fmt_tot)
        ws.write(r, 9, "Avg Rate", fmt_strip);   ws.write_number(r, 10, grand_rate, fmt_tot_rate)
        ws.write(r, 11, "Stock", fmt_strip);     ws.write_number(r, 12, footer_stock, fmt_tot)
        last_col = 12
    else:
        ws.write(r, 3, "Plan Qty", fmt_strip);   ws.write_number(r, 4, grand_plan, fmt_tot)
        ws.write(r, 5, "Est Qty", fmt_strip);    ws.write_number(r, 6, grand_est, fmt_tot)
        ws.write(r, 7, "Actual Qty", fmt_strip); ws.write_number(r, 8, grand_act, fmt_tot)
        ws.write(r, 9, "Actual Val", fmt_strip); ws.write_number(r, 10, grand_val, fmt_tot)
        ws.write(r, 11, "Avg Rate", fmt_strip);  ws.write_number(r, 12, grand_rate, fmt_tot_rate)
        ws.write(r, 13, "Stock", fmt_strip);     ws.write_number(r, 14, footer_stock, fmt_tot)
        last_col = 14

    r += 2

    # ---------- 3-row header like HTML ----------
    header1 = r
    header2 = r + 1
    header3 = r + 2

    ws.write(header1, 0, "Product / FG", fmt_hdr_left)

    col = 1
    group_w = 4 if is_daily else 7
    for lab in labels:
        ws.merge_range(header1, col, header1, col + group_w - 1, lab, fmt_hdr_top)
        col += group_w

    stock_col = col
    ws.write(header1, stock_col, "Stock", fmt_hdr_top)

    ws.write(header2, 0, "Alias / Alpha", fmt_hdr_left)
    col = 1
    for _ in labels:
        if is_daily:
            ws.write(header2, col + 0, "Plan", fmt_plan_h)
            ws.merge_range(header2, col + 1, header2, col + 2, "Actual", fmt_act_h)
            ws.write(header2, col + 3, "Rate", fmt_rate_h)
            col += 4
        else:
            ws.merge_range(header2, col + 0, header2, col + 1, "Plan", fmt_plan_h)
            ws.merge_range(header2, col + 2, header2, col + 3, "Est (MTD)", fmt_est_h)
            ws.merge_range(header2, col + 4, header2, col + 5, "Actual", fmt_act_h)
            ws.write(header2, col + 6, "Rate", fmt_rate_h)
            col += 7
    ws.write(header2, stock_col, "Qty", fmt_hdr_mid)

    ws.write(header3, 0, "", fmt_hdr_left)
    col = 1
    for _ in labels:
        if is_daily:
            ws.write(header3, col + 0, "Qty", fmt_plan_h)
            ws.write(header3, col + 1, "Qty", fmt_act_h)
            ws.write(header3, col + 2, "Val", fmt_act_h)
            ws.write(header3, col + 3, "Avg", fmt_rate_h)
            col += 4
        else:
            ws.write(header3, col + 0, "Qty", fmt_plan_h)
            ws.write(header3, col + 1, "Val", fmt_plan_h)
            ws.write(header3, col + 2, "Qty", fmt_est_h)
            ws.write(header3, col + 3, "Val", fmt_est_h)
            ws.write(header3, col + 4, "Qty", fmt_act_h)
            ws.write(header3, col + 5, "Val", fmt_act_h)
            ws.write(header3, col + 6, "Avg", fmt_rate_h)
            col += 7
    ws.write(header3, stock_col, "", fmt_hdr_mid)

    # Freeze: keep headers + first column
    ws.freeze_panes(header3 + 1, 1)

    # ---------- Body rows ----------
    row_r = header3 + 1
    for rr in rows:
        ws.write(row_r, 0, rr.get("alias") or "", fmt_txt_bold)

        col = 1
        for c in (rr.get("cells") or []):
            if is_daily:
                ws.write_number(row_r, col + 0, _n0(c.get("plan")), fmt_qty)
                ws.write_number(row_r, col + 1, _n0(c.get("actual")), fmt_qty)
                ws.write_number(row_r, col + 2, _n0(c.get("actual_value")), fmt_amt)
                ws.write_number(row_r, col + 3, _n0(c.get("actual_rate")), fmt_rate)
                col += 4
            else:
                rate = _n0(c.get("actual_rate"))
                pqty = _n0(c.get("plan"))
                eqty = _n0(c.get("est"))
                aqty = _n0(c.get("actual"))
                aval = _n0(c.get("actual_value"))

                # same as your HTML JS: PlanVal / EstVal = Qty * actual_rate
                ws.write_number(row_r, col + 0, pqty, fmt_qty)
                ws.write_number(row_r, col + 1, pqty * rate, fmt_amt)

                ws.write_number(row_r, col + 2, eqty, fmt_qty)
                ws.write_number(row_r, col + 3, eqty * rate, fmt_amt)

                ws.write_number(row_r, col + 4, aqty, fmt_qty)
                ws.write_number(row_r, col + 5, aval, fmt_amt)

                ws.write_number(row_r, col + 6, rate, fmt_rate)
                col += 7

        ws.write_number(row_r, stock_col, _n0(rr.get("stock_qty")), fmt_qty)
        row_r += 1

    # ---------- Totals row (bottom) ----------
    ws.write(row_r, 0, "Totals", fmt_tot_lbl)
    col = 1
    for f in footer_months:
        if is_daily:
            ws.write_number(row_r, col + 0, _n0(f.get("plan")), fmt_tot)
            ws.write_number(row_r, col + 1, _n0(f.get("actual")), fmt_tot)
            ws.write_number(row_r, col + 2, _n0(f.get("actual_value")), fmt_tot)
            ws.write_number(row_r, col + 3, _n0(f.get("actual_rate")), fmt_tot_rate)
            col += 4
        else:
            rate = _n0(f.get("actual_rate"))
            pqty = _n0(f.get("plan"))
            eqty = _n0(f.get("est"))
            aqty = _n0(f.get("actual"))
            aval = _n0(f.get("actual_value"))

            ws.write_number(row_r, col + 0, pqty, fmt_tot)
            ws.write_number(row_r, col + 1, pqty * rate, fmt_tot)

            ws.write_number(row_r, col + 2, eqty, fmt_tot)
            ws.write_number(row_r, col + 3, eqty * rate, fmt_tot)

            ws.write_number(row_r, col + 4, aqty, fmt_tot)
            ws.write_number(row_r, col + 5, aval, fmt_tot)

            ws.write_number(row_r, col + 6, rate, fmt_tot_rate)
            col += 7

    ws.write_number(row_r, stock_col, footer_stock, fmt_tot)

    ws.autofilter(header3, 0, row_r, stock_col)
    ws.set_row(header1, 22)
    ws.set_row(header2, 18)
    ws.set_row(header3, 18)

    wb.close()
    output.seek(0)

    fname = f"dispatch_plan_vs_actual_{period}_{fr.strftime('%Y%m%d')}_{to.strftime('%Y%m%d')}.xlsx"
    resp = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{fname}"'
    return resp


# ───────────────────────────────────────────────────────────────
# Dispatch Plan vs Actual (FIXED FG DROPDOWN SOURCE + Excel)
# ───────────────────────────────────────────────────────────────
@login_required
def dispatch_plan_vs_actual(request):
    """
    STORE → Dispatch Plan vs Actual
    Supports:
      ?from=YYYY-MM-DD&to=YYYY-MM-DD&period=DAILY|MTD|YTD|CUSTOM
      ?fg_name=<alpha> (multi)
      ?export=xlsx  (download excel)
    """
    today = timezone.localdate()

    raw_from = _parse_ymd(request.GET.get("from"))
    raw_to   = _parse_ymd(request.GET.get("to"))

    period = (request.GET.get("period") or "MTD").upper()
    if period not in ("DAILY", "MTD", "YTD", "CUSTOM"):
        period = "MTD"

    # ✅ default: current month 1st day → today
    default_from = today.replace(day=1)
    default_to   = today

    rng_from = raw_from or default_from
    rng_to   = raw_to   or default_to

    # ✅ clamp future end-date to today
    if rng_to > today:
        rng_to = today

    # ✅ Daily mode = single day (rng_to)
    if period == "DAILY":
        rng_from = rng_to

    # ✅ safe swap
    if rng_from > rng_to:
        rng_from, rng_to = rng_to, rng_from

    selected_fgs = request.GET.getlist("fg_name") or []

    alpha_map = fetch_alpha_map()

    # ✅ Dropdown list must come from alpha_map (aliases)
    all_fg_list = sorted(
        {v for v in alpha_map.values() if str(v).strip()},
        key=lambda x: (x or "").upper(),
    )

    disp_grid = build_dispatch_grid(
        period=period,
        fr=rng_from,
        to=rng_to,
        alpha_map=alpha_map,
        selected_fgs=selected_fgs,
    )

    # ✅ Excel download
    exp = (request.GET.get("export") or "").lower()
    if exp in ("xlsx", "excel", "1", "true"):
        return _export_dispatch_plan_vs_actual_xlsx(
            disp_grid=disp_grid,
            period=period,
            fr=rng_from,
            to=rng_to,
        )

    return render(
        request,
        "store/dispatch_plan_vs_actual.html",
        {
            "period": period,
            "from_date": rng_from,
            "to_date": rng_to,
            "selected_fgs": selected_fgs,
            "all_fg_list": all_fg_list,
            "disp_grid": disp_grid,
        },
    )


import logging
from io import BytesIO
from typing import Any
from urllib.parse import urlencode
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.db import connections
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.urls import reverse
from django.utils.http import url_has_allowed_host_and_scheme
from django.views.decorators.http import require_POST

import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger("custom_logger")

SOLAPUR_COMPANY = "OC Specialities Private Limited - Solapur"

# ✅ Use your actual app label here (usually "STORE")
APP_LABEL = "STORE"

# Django permission codenames created automatically by the model:
#   view_invageingpreview  (for report page)
# Custom permission from Meta.permissions:
#   sync_invageingpreview  (for sync button POST)
PERM_VIEW = f"{APP_LABEL}.view_invageingpreview"
PERM_SYNC = f"{APP_LABEL}.sync_invageingpreview"


def _dictfetchall(cursor):
    while cursor.description is None:
        if not cursor.nextset():
            return []
    cols = [c[0] for c in cursor.description]
    return [dict(zip(cols, row)) for row in cursor.fetchall()]


def _safe_int(val, default=0):
    try:
        return int(val)
    except Exception:
        return default


def _xlsx_response(filename: str, rows: list[dict]) -> HttpResponse:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inv Drill"

    headers = [
        "Txn Location",
        "Origin From Location",
        "Item Code",
        "Item Name",
        "Unit",
        "Batch No",
        "Mfg Date",
        "Retest Date",
        "Bin",
        "Closing Qty",
        "Closing Value",
    ]
    ws.append(headers)

    # Header styling
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data
    for r in rows:
        ws.append([
            r.get("txn_location"),
            r.get("from_location"),
            r.get("item_code"),
            r.get("item_name"),
            r.get("unit"),
            r.get("batch_no"),
            r.get("mfg_date"),
            r.get("retest_date"),
            r.get("stock_location"),
            float(r.get("closing_qty") or 0),
            float(r.get("closing_value") or 0),
        ])

    # Number formats
    qty_col = headers.index("Closing Qty") + 1
    val_col = headers.index("Closing Value") + 1
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=qty_col).number_format = "#,##0.000"
        ws.cell(row=row_idx, column=val_col).number_format = "#,##0.00"

    # Column widths
    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(max(len(h) + 2, 14), 45)

    ws.freeze_panes = "A2"

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    resp = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


@require_POST
@login_required
def inv_ageing_sync(request):
    """
    Calls dbo.usp_SyncInvAgeingPreview and redirects back to same page.
    Permission controlled ONLY in backend.
    """
    if not request.user.has_perm(PERM_SYNC):
        raise PermissionDenied

    next_url = request.POST.get("next") or reverse("inv_ageing_drill_report")

    if not url_has_allowed_host_and_scheme(
        url=next_url,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ):
        next_url = reverse("inv_ageing_drill_report")

    try:
        with connections["default"].cursor() as cursor:
            cursor.execute("EXEC dbo.usp_SyncInvAgeingPreview;")
        messages.success(request, "Inventory ageing data synced successfully.")
    except Exception:
        logger.exception("Error while running dbo.usp_SyncInvAgeingPreview")
        messages.error(request, "Sync failed. Please check logs / SQL Server.")

    return redirect(next_url)


@login_required
def inv_ageing_drill_report(request):
    """
    Summary:
      txn_location totals -> origin_from_location totals (expand/collapse)
      ✅ Uses total_closing_per_batch / total_closing_value_per_batch
      ✅ NO doubling: one row per batch
      ✅ company_name fixed = Solapur
      ✅ NO date filter

    Detail:
      One row per batch with origin + batch totals

    Excel:
      ✅ export=1 returns XLSX
      ✅ If from_location present => exports that txn + from only
      ✅ Else exports txn_location only
      ✅ Always respects filters (company, item name/code, txn_sel)
    """
    if not request.user.has_perm(PERM_VIEW):
        raise PermissionDenied

    mode = (request.GET.get("mode") or "summary").lower().strip()
    if mode not in ("summary", "detail"):
        mode = "summary"

    # Filters (NO date)
    item_q = (request.GET.get("item_q") or "").strip()
    item_code_q = (request.GET.get("item_code_q") or "").strip()
    txn_sel = (request.GET.get("txn_sel") or "").strip()

    # paging for detail
    page = max(_safe_int(request.GET.get("page"), 1), 1)
    page_size = _safe_int(request.GET.get("page_size"), 50)
    if page_size not in (25, 50, 100, 200):
        page_size = 50
    offset = (page - 1) * page_size

    # drill keys
    txn_location = (request.GET.get("txn_location") or "").strip()
    from_location = (request.GET.get("from_location") or "").strip()
    export = (request.GET.get("export") or "").strip()

    item_like = f"%{item_q}%" if item_q else "%%"
    item_code_like = f"%{item_code_q}%" if item_code_q else "%%"

    # back url (preserve summary filters) - NO date params
    back_qs = {
        "mode": "summary",
        "item_q": item_q,
        "item_code_q": item_code_q,
        "txn_sel": txn_sel,
    }
    back_url = "?" + urlencode(back_qs, doseq=True)

    # helpful for template
    current_full_path = request.get_full_path()

    try:
        with connections["default"].cursor() as cursor:

            # -----------------------------
            # Last synced time
            # -----------------------------
            cursor.execute(
                """
                SELECT MAX(synced_at) AS last_synced_at
                FROM dbo.InvAgeingPreview
                WHERE company_name = %s AND item_type = 'Engineering Material';
                """,
                [SOLAPUR_COMPANY],
            )
            last_synced = _dictfetchall(cursor)
            last_synced_at = last_synced[0]["last_synced_at"] if last_synced else None

            # -----------------------------
            # Txn Location dropdown list (filtered)
            # ✅ do NOT show locations with 0 stock (qty+value zero)
            # -----------------------------
            txn_list_sql = r"""
;WITH Base AS (
    SELECT
        txn_location,
        from_location,
        item_code,
        item_name,
        batch_no,
        mfg_date,
        stock_location,
        doc_date,
        doc_no,
        closing_qty,
        receipt_qty,
        total_closing_per_batch,
        total_closing_value_per_batch
    FROM dbo.InvAgeingPreview
    WHERE company_name = %s
      AND item_type = 'Engineering Material'
      AND (%s = '' OR item_name LIKE %s)
      AND (%s = '' OR item_code LIKE %s)
      AND ISNULL(LTRIM(RTRIM(txn_location)),'') <> ''
      AND ISNULL(LTRIM(RTRIM(batch_no)),'') <> ''
),
BatchOne AS (
    SELECT
        b.*,
        origin_from_location =
            FIRST_VALUE(b.from_location) OVER (
                PARTITION BY b.txn_location, b.item_code, b.batch_no, b.mfg_date, b.stock_location
                ORDER BY
                    CASE WHEN ISNULL(b.receipt_qty,0) > 0 OR ISNULL(b.closing_qty,0) > 0 THEN 0 ELSE 1 END,
                    b.doc_date ASC,
                    b.doc_no   ASC
            ),
        batch_qty   = MAX(b.total_closing_per_batch) OVER (
                        PARTITION BY b.txn_location, b.item_code, b.batch_no, b.mfg_date, b.stock_location
                     ),
        batch_value = MAX(b.total_closing_value_per_batch) OVER (
                        PARTITION BY b.txn_location, b.item_code, b.batch_no, b.mfg_date, b.stock_location
                     ),
        rn = ROW_NUMBER() OVER (
                PARTITION BY b.txn_location, b.item_code, b.batch_no, b.mfg_date, b.stock_location
                ORDER BY b.doc_date DESC, b.doc_no DESC
             )
    FROM Base b
),
Batches AS (
    SELECT
        txn_location,
        batch_qty,
        batch_value
    FROM BatchOne
    WHERE rn = 1
      AND LTRIM(RTRIM(ISNULL(origin_from_location,''))) <> ''
),
TxnAgg AS (
    SELECT
        LTRIM(RTRIM(txn_location)) AS txn_location,
        SUM(batch_qty)   AS qty_sum,
        SUM(batch_value) AS val_sum
    FROM Batches
    GROUP BY LTRIM(RTRIM(txn_location))
)
SELECT txn_location
FROM TxnAgg
WHERE NOT (ABS(ISNULL(qty_sum,0)) <= 0.001 AND ABS(ISNULL(val_sum,0)) <= 1)
ORDER BY txn_location;
"""
            cursor.execute(
                txn_list_sql,
                [SOLAPUR_COMPANY, item_q, item_like, item_code_q, item_code_like],
            )
            txn_locations = [r["txn_location"] for r in _dictfetchall(cursor) if r.get("txn_location")]
            if txn_sel and txn_sel not in txn_locations:
                txn_locations = [txn_sel] + txn_locations

            # -----------------------------
            # Excel export (batch-level) - NO date filter
            # -----------------------------
            if export == "1":
                export_sql = r"""
;WITH Base AS (
    SELECT
        txn_location,
        from_location,
        item_code,
        item_name,
        unit,
        batch_no,
        mfg_date,
        retest_date,
        stock_location,
        doc_date,
        doc_no,
        receipt_qty,
        closing_qty,
        total_closing_per_batch,
        total_closing_value_per_batch
    FROM dbo.InvAgeingPreview
    WHERE company_name = %s
      AND item_type = 'Engineering Material'
      AND (%s = '' OR item_name LIKE %s)
      AND (%s = '' OR item_code LIKE %s)
      AND (%s = '' OR txn_location = %s)              -- global dropdown filter
      AND (%s = '' OR txn_location = %s)              -- clicked txn_location
      AND ISNULL(LTRIM(RTRIM(batch_no)),'') <> ''
),
BatchOne AS (
    SELECT
        b.*,
        origin_from_location =
            FIRST_VALUE(b.from_location) OVER (
                PARTITION BY b.txn_location, b.item_code, b.batch_no, b.mfg_date, b.stock_location
                ORDER BY
                    CASE WHEN ISNULL(b.receipt_qty,0) > 0 OR ISNULL(b.closing_qty,0) > 0 THEN 0 ELSE 1 END,
                    b.doc_date ASC,
                    b.doc_no   ASC
            ),
        batch_qty   = MAX(b.total_closing_per_batch) OVER (
                        PARTITION BY b.txn_location, b.item_code, b.batch_no, b.mfg_date, b.stock_location
                     ),
        batch_value = MAX(b.total_closing_value_per_batch) OVER (
                        PARTITION BY b.txn_location, b.item_code, b.batch_no, b.mfg_date, b.stock_location
                     ),
        rn = ROW_NUMBER() OVER (
                PARTITION BY b.txn_location, b.item_code, b.batch_no, b.mfg_date, b.stock_location
                ORDER BY b.doc_date DESC, b.doc_no DESC
             )
    FROM Base b
),
Batches AS (
    SELECT
        txn_location,
        origin_from_location,
        item_code,
        item_name,
        unit,
        batch_no,
        mfg_date,
        retest_date,
        stock_location,
        batch_qty,
        batch_value
    FROM BatchOne
    WHERE rn = 1
      AND LTRIM(RTRIM(ISNULL(origin_from_location,''))) <> ''
      AND (%s = '' OR origin_from_location = %s)      -- clicked from_location (optional)
)
SELECT
    txn_location,
    origin_from_location AS from_location,
    item_code,
    item_name,
    unit,
    batch_no,
    mfg_date,
    retest_date,
    stock_location,
    batch_qty   AS closing_qty,
    batch_value AS closing_value
FROM Batches
WHERE NOT (ABS(ISNULL(batch_qty,0)) <= 0.001 AND ABS(ISNULL(batch_value,0)) <= 1)
ORDER BY item_name, batch_no, mfg_date, txn_location;
"""
                params = [
                    SOLAPUR_COMPANY,
                    item_q, item_like,
                    item_code_q, item_code_like,
                    txn_sel, txn_sel,
                    txn_location, txn_location,
                    from_location, from_location,
                ]
                cursor.execute(export_sql, params)
                export_rows = _dictfetchall(cursor)

                safe_txn = (txn_location or "ALL_TXN").replace("/", "-")
                safe_from = (from_location or "").replace("/", "-")
                if from_location:
                    fname = f"InvAgeing_{safe_txn}__{safe_from}.xlsx"
                else:
                    fname = f"InvAgeing_{safe_txn}.xlsx"

                return _xlsx_response(fname, export_rows)

            # -----------------------------
            # SUMMARY - NO date filter
            # -----------------------------
            if mode == "summary":
                summary_sql = r"""
;WITH Base AS (
    SELECT
        txn_location,
        from_location,
        item_code,
        item_name,
        batch_no,
        mfg_date,
        stock_location,
        doc_date,
        doc_no,
        closing_qty,
        receipt_qty,
        total_closing_per_batch,
        total_closing_value_per_batch
    FROM dbo.InvAgeingPreview
    WHERE company_name = %s
      AND item_type = 'Engineering Material'
      AND (%s = '' OR item_name LIKE %s)
      AND (%s = '' OR item_code LIKE %s)
      AND (%s = '' OR txn_location = %s)
      AND ISNULL(LTRIM(RTRIM(txn_location)),'') <> ''
      AND ISNULL(LTRIM(RTRIM(batch_no)),'') <> ''
),
BatchOne AS (
    SELECT
        b.*,
        origin_from_location =
            FIRST_VALUE(b.from_location) OVER (
                PARTITION BY b.txn_location, b.item_code, b.batch_no, b.mfg_date, b.stock_location
                ORDER BY
                    CASE WHEN ISNULL(b.receipt_qty,0) > 0 OR ISNULL(b.closing_qty,0) > 0 THEN 0 ELSE 1 END,
                    b.doc_date ASC,
                    b.doc_no   ASC
            ),
        batch_qty   = MAX(b.total_closing_per_batch) OVER (
                        PARTITION BY b.txn_location, b.item_code, b.batch_no, b.mfg_date, b.stock_location
                     ),
        batch_value = MAX(b.total_closing_value_per_batch) OVER (
                        PARTITION BY b.txn_location, b.item_code, b.batch_no, b.mfg_date, b.stock_location
                     ),
        rn = ROW_NUMBER() OVER (
                PARTITION BY b.txn_location, b.item_code, b.batch_no, b.mfg_date, b.stock_location
                ORDER BY b.doc_date DESC, b.doc_no DESC
             )
    FROM Base b
),
Batches AS (
    SELECT
        txn_location,
        origin_from_location,
        batch_qty,
        batch_value
    FROM BatchOne
    WHERE rn = 1
      AND LTRIM(RTRIM(ISNULL(origin_from_location,''))) <> ''
),
TF AS (
    SELECT
        txn_location,
        origin_from_location AS from_location,
        SUM(batch_qty)   AS from_closing_qty,
        SUM(batch_value) AS from_closing_value
    FROM Batches
    GROUP BY txn_location, origin_from_location
),
T AS (
    SELECT
        txn_location,
        SUM(from_closing_qty)   AS txn_closing_qty,
        SUM(from_closing_value) AS txn_closing_value
    FROM TF
    GROUP BY txn_location
)
SELECT
    t.txn_location,
    t.txn_closing_qty,
    t.txn_closing_value,
    tf.from_location,
    tf.from_closing_qty,
    tf.from_closing_value
FROM T t
LEFT JOIN TF tf
  ON tf.txn_location = t.txn_location
WHERE NOT (ABS(ISNULL(t.txn_closing_qty,0)) <= 0.001 AND ABS(ISNULL(t.txn_closing_value,0)) <= 1)
ORDER BY t.txn_location, tf.from_location;
"""
                params = [
                    SOLAPUR_COMPANY,
                    item_q, item_like,
                    item_code_q, item_code_like,
                    txn_sel, txn_sel,
                ]
                cursor.execute(summary_sql, params)
                rows = _dictfetchall(cursor)

                tree: dict[str, Any] = {}
                for r in rows:
                    tl = (r.get("txn_location") or "").strip() or "(Blank)"
                    node = tree.get(tl)
                    if not node:
                        node = tree[tl] = {
                            "txn_location": tl,
                            "closing_qty": float(r.get("txn_closing_qty") or 0),
                            "closing_value": float(r.get("txn_closing_value") or 0),
                            "children": [],
                        }

                    fl = (r.get("from_location") or "").strip() or "-"
                    node["children"].append({
                        "from_location": fl,
                        "closing_qty": float(r.get("from_closing_qty") or 0),
                        "closing_value": float(r.get("from_closing_value") or 0),
                    })

                return render(
                    request,
                    "store/inv_ageing_drill_report.html",
                    {
                        "mode": "summary",
                        "company_name": SOLAPUR_COMPANY,
                        "item_q": item_q,
                        "item_code_q": item_code_q,
                        "txn_sel": txn_sel,
                        "txn_locations": txn_locations,
                        "txn_list": list(tree.values()),
                        "back_url": back_url,
                        "current_full_path": current_full_path,
                        "last_synced_at": last_synced_at,
                        # If you ever need this in template later (optional):
                        "can_sync": request.user.has_perm(PERM_SYNC),
                    },
                )

            # -----------------------------
            # DETAIL - NO date filter
            # -----------------------------
            detail_sql = r"""
;WITH Base AS (
    SELECT
        txn_location,
        from_location,
        item_code,
        item_name,
        unit,
        batch_no,
        mfg_date,
        retest_date,
        stock_location,
        doc_date,
        doc_no,
        receipt_qty,
        closing_qty,
        total_closing_per_batch,
        total_closing_value_per_batch
    FROM dbo.InvAgeingPreview
    WHERE company_name = %s
      AND item_type = 'Engineering Material'
      AND (%s = '' OR item_name LIKE %s)
      AND (%s = '' OR item_code LIKE %s)
      AND (%s = '' OR txn_location = %s)              -- global dropdown filter
      AND (%s = '' OR txn_location = %s)              -- clicked txn_location
      AND ISNULL(LTRIM(RTRIM(batch_no)),'') <> ''
),
BatchOne AS (
    SELECT
        b.*,
        origin_from_location =
            FIRST_VALUE(b.from_location) OVER (
                PARTITION BY b.txn_location, b.item_code, b.batch_no, b.mfg_date, b.stock_location
                ORDER BY
                    CASE WHEN ISNULL(b.receipt_qty,0) > 0 OR ISNULL(b.closing_qty,0) > 0 THEN 0 ELSE 1 END,
                    b.doc_date ASC,
                    b.doc_no   ASC
            ),
        batch_qty   = MAX(b.total_closing_per_batch) OVER (
                        PARTITION BY b.txn_location, b.item_code, b.batch_no, b.mfg_date, b.stock_location
                     ),
        batch_value = MAX(b.total_closing_value_per_batch) OVER (
                        PARTITION BY b.txn_location, b.item_code, b.batch_no, b.mfg_date, b.stock_location
                     ),
        rn = ROW_NUMBER() OVER (
                PARTITION BY b.txn_location, b.item_code, b.batch_no, b.mfg_date, b.stock_location
                ORDER BY b.doc_date DESC, b.doc_no DESC
             )
    FROM Base b
),
Batches AS (
    SELECT
        txn_location,
        origin_from_location,
        item_code,
        item_name,
        unit,
        batch_no,
        mfg_date,
        retest_date,
        stock_location,
        batch_qty,
        batch_value
    FROM BatchOne
    WHERE rn = 1
      AND LTRIM(RTRIM(ISNULL(origin_from_location,''))) <> ''
      AND (%s = '' OR origin_from_location = %s)
),
C AS (SELECT COUNT(1) AS total_rows FROM Batches
      WHERE NOT (ABS(ISNULL(batch_qty,0)) <= 0.001 AND ABS(ISNULL(batch_value,0)) <= 1))
SELECT
    b.txn_location,
    b.origin_from_location AS from_location,
    b.item_code,
    b.item_name,
    b.unit,
    b.batch_no,
    b.mfg_date,
    b.retest_date,
    b.stock_location,
    b.batch_qty   AS closing_qty,
    b.batch_value AS closing_value,
    c.total_rows
FROM Batches b
CROSS JOIN C c
WHERE NOT (ABS(ISNULL(b.batch_qty,0)) <= 0.001 AND ABS(ISNULL(b.batch_value,0)) <= 1)
ORDER BY b.item_name, b.batch_no, b.mfg_date
OFFSET %s ROWS FETCH NEXT %s ROWS ONLY;
"""
            params = [
                SOLAPUR_COMPANY,
                item_q, item_like,
                item_code_q, item_code_like,
                txn_sel, txn_sel,
                txn_location, txn_location,
                from_location, from_location,
                offset, page_size,
            ]
            cursor.execute(detail_sql, params)
            detail_rows = _dictfetchall(cursor)

            total_rows = int(detail_rows[0]["total_rows"]) if detail_rows else 0
            total_pages = max((total_rows + page_size - 1) // page_size, 1)

            return render(
                request,
                "store/inv_ageing_drill_report.html",
                {
                    "mode": "detail",
                    "company_name": SOLAPUR_COMPANY,
                    "item_q": item_q,
                    "item_code_q": item_code_q,
                    "txn_sel": txn_sel,
                    "txn_locations": txn_locations,
                    "txn_location": txn_location,
                    "from_location": from_location,
                    "rows": detail_rows,
                    "page": page,
                    "page_size": page_size,
                    "total_rows": total_rows,
                    "total_pages": total_pages,
                    "back_url": back_url,
                    "current_full_path": current_full_path,
                    "last_synced_at": last_synced_at,
                    "can_sync": request.user.has_perm(PERM_SYNC),
                },
            )

    except Exception:
        logger.exception("Error in inv_ageing_drill_report")
        messages.error(request, "Something went wrong while loading the report.")
        return redirect("indexpage")

from datetime import datetime
from io import BytesIO
from collections import Counter

from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.shortcuts import render, redirect
from django.db import connections
from django.http import HttpResponse

import logging
logger = logging.getLogger("custom_logger")

import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


def _dictfetchall(cursor):
    # supports multi-statement queries (temp tables etc.)
    while cursor.description is None:
        if not cursor.nextset():
            return []
    cols = [c[0] for c in cursor.description]
    return [dict(zip(cols, row)) for row in cursor.fetchall()]


def _yyyymmdd_int(d: str | None, default_int: int) -> int:
    """Convert 'YYYY-MM-DD' -> int YYYYMMDD (SQL int docdate format)."""
    if not d:
        return default_int
    try:
        dt = datetime.strptime(d.strip(), "%Y-%m-%d")
        return int(dt.strftime("%Y%m%d"))
    except Exception:
        return default_int


def _xlsx_response_pending_grn(filename: str, rows: list[dict]) -> HttpResponse:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pending GRN"

    headers = [
        "DocDate",
        "DocNo",
        "TransactionType",
        "PartyName",
        "ItemType",
        "ItemName",
        "PendingStage",   # ✅ normalized stage (Check By / Approved By)
        "PendingStatus",  # Pending for Check By / Pending for Approved By
    ]
    ws.append(headers)

    # Header styling
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data
    for r in rows:
        ws.append([
            r.get("DocDate"),
            r.get("sDocNo"),
            r.get("TransactionType"),
            r.get("PartyName"),
            r.get("ItemType"),
            r.get("ItemName"),
            r.get("PendingStage"),
            r.get("PendingStatus"),
        ])

    # Column widths
    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(max(len(h) + 2, 14), 55)

    ws.freeze_panes = "A2"

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    resp = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


@login_required
def pending_grn_report(request):
    """
    Pending GRN report (fast) from TXNHDR/TXNDET/TXNSTAT.

    Added:
      ✅ Analysis counters (TransactionType / ItemType / PendingStage)
      ✅ Cross-filter params: txn_type=, item_type=, stage=
      ✅ Excel export exports FULL filtered detail (not only current page)

    Updated:
      ✅ PartyName now comes from BUSMST (instead of ACCMST)

    Normalization:
      ✅ Stage is only 2 values: "Check By" and "Approved By"
      ✅ Fix for django-mssql formatting: LIKE 'check%%' not 'check%'
    """
    # Filters
    from_date_str = (request.GET.get("from_date") or "").strip()   # YYYY-MM-DD
    docno_q       = (request.GET.get("docno") or "").strip()
    party_q       = (request.GET.get("party") or "").strip()
    item_q        = (request.GET.get("item") or "").strip()
    status_q      = (request.GET.get("status") or "").strip()

    # Cross-filter selections (PowerBI-like)
    txn_type_sel  = (request.GET.get("txn_type") or "").strip()
    item_type_sel = (request.GET.get("item_type") or "").strip()
    stage_sel     = (request.GET.get("stage") or "").strip()  # expected: "Check By" / "Approved By"

    # export flag
    export = (request.GET.get("export") or "").strip().lower()

    # Pagination (screen only)
    page = request.GET.get("page", "1")
    page_size = int(request.GET.get("page_size") or 50)
    if page_size not in (25, 50, 100, 200):
        page_size = 50

    # Default from_date = 2025-01-01
    from_date_int = _yyyymmdd_int(from_date_str, 20250101)

    sql = r"""
/* Pending stage report (fast) */

SET NOCOUNT ON;

IF OBJECT_ID('tempdb..#Docs')     IS NOT NULL DROP TABLE #Docs;
IF OBJECT_ID('tempdb..#DetFirst') IS NOT NULL DROP TABLE #DetFirst;

------------------------------------------------------------
-- 1) Docs (working set)
------------------------------------------------------------
SELECT
    h.lTypId,
    h.lId,
    h.sDocNo,
    h.dtDocDate,      -- int yyyymmdd
    h.lDocNo,
    h.lAccId1
INTO #Docs
FROM TXNHDR h WITH (NOLOCK)
WHERE h.lTypId IN (164, 540, 792, 793, 794, 795, 796, 797, 801, 850, 958)
  AND ISNULL(h.bDel, 0) = 0
  AND h.dtDocDate >= %s;

CREATE CLUSTERED INDEX IX_Docs_lId ON #Docs (lId);
CREATE NONCLUSTERED INDEX IX_Docs_TypId_lId ON #Docs (lTypId, lId)
INCLUDE (dtDocDate, lDocNo, sDocNo, lAccId1);

------------------------------------------------------------
-- 2) First item per doc (to get Item + ItemType)
------------------------------------------------------------
;WITH X AS (
    SELECT
        dd.lId,
        dd.lItmId,
        dd.lItmTyp,
        ROW_NUMBER() OVER (PARTITION BY dd.lId ORDER BY dd.lLine) AS rn
    FROM TXNDET dd WITH (NOLOCK)
    INNER JOIN #Docs d
        ON d.lId = dd.lId
    WHERE dd.cFlag = 'I'
      AND ISNULL(dd.bDel,0) = 0
      AND dd.lItmId > 0
)
SELECT lId, lItmId, lItmTyp
INTO #DetFirst
FROM X
WHERE rn = 1;

CREATE CLUSTERED INDEX IX_DetFirst_lId ON #DetFirst (lId);

------------------------------------------------------------
-- 3) Final output
------------------------------------------------------------
SELECT
    CONVERT(date, CONVERT(char(8), d.dtDocDate)) AS DocDate,
    d.sDocNo,
    tt.sName   AS TransactionType,
    b.sName    AS PartyName,   -- ✅ BUSMST
    ityp.sName AS ItemType,
    itm.sName  AS ItemName,

    -- ✅ NORMALIZED to ONLY 2 values
    CASE
        WHEN LOWER(LTRIM(RTRIM(p.sName))) LIKE 'check%%'  THEN 'Check By'
        WHEN LOWER(LTRIM(RTRIM(p.sName))) LIKE 'approv%%' THEN 'Approved By'
        ELSE 'Other'
    END AS PendingStage,

    CONCAT(
        'Pending for ',
        CASE
            WHEN LOWER(LTRIM(RTRIM(p.sName))) LIKE 'check%%'  THEN 'Check By'
            WHEN LOWER(LTRIM(RTRIM(p.sName))) LIKE 'approv%%' THEN 'Approved By'
            ELSE 'Other'
        END
    ) AS PendingStatus

FROM #Docs d
INNER JOIN TXNTYP tt WITH (NOLOCK)
    ON tt.lTypId = d.lTypId

OUTER APPLY (
    SELECT TOP (1) 1 AS IsCancelled
    FROM TXNSTAT s WITH (NOLOCK)
    WHERE s.lTypId = d.lTypId
      AND s.lId    = d.lId
      AND (s.sName LIKE 'Cancel%%' OR s.sName LIKE 'Cancle%%' OR s.lStatusId = 40)
      AND s.dtDate IS NOT NULL
) c

OUTER APPLY (
    SELECT TOP (1) s.lStatusId, s.sName
    FROM TXNSTAT s WITH (NOLOCK)
    WHERE s.lTypId = d.lTypId
      AND s.lId    = d.lId
      AND s.dtDate IS NULL
      AND NOT (s.sName LIKE 'Cancel%%' OR s.sName LIKE 'Cancle%%' OR s.lStatusId = 40)
    ORDER BY s.lStatusId ASC
) p

LEFT JOIN #DetFirst df
    ON df.lId = d.lId
LEFT JOIN ITMMST itm WITH (NOLOCK)
    ON itm.lId = df.lItmId
LEFT JOIN ITMTYP ityp WITH (NOLOCK)
    ON ityp.lTypId = df.lItmTyp
LEFT JOIN BUSMST b WITH (NOLOCK)
    ON b.lId = d.lAccId1

WHERE c.IsCancelled IS NULL
  AND p.sName IS NOT NULL

  -- normal search filters
  AND (%s = '' OR d.sDocNo LIKE %s)
  AND (%s = '' OR b.sName LIKE %s)
  AND (%s = '' OR itm.sName LIKE %s)
  AND (%s = '' OR p.sName  LIKE %s)

  -- Cross-filter exact matches
  AND (%s = '' OR tt.sName   = %s)
  AND (%s = '' OR ityp.sName = %s)

  -- ✅ stage filter uses normalized PendingStage
  AND (
        %s = '' OR
        CASE
            WHEN LOWER(LTRIM(RTRIM(p.sName))) LIKE 'check%%'  THEN 'Check By'
            WHEN LOWER(LTRIM(RTRIM(p.sName))) LIKE 'approv%%' THEN 'Approved By'
            ELSE 'Other'
        END = %s
      )

ORDER BY d.dtDocDate DESC, d.lDocNo DESC
OPTION (RECOMPILE);
"""

    params = [
        from_date_int,

        docno_q, f"%{docno_q}%",
        party_q, f"%{party_q}%",
        item_q,  f"%{item_q}%",
        status_q, f"%{status_q}%",

        txn_type_sel, txn_type_sel,
        item_type_sel, item_type_sel,

        stage_sel, stage_sel,
    ]

    try:
        with connections["readonly_db"].cursor() as cursor:
            cursor.execute(sql, params)
            rows = _dictfetchall(cursor)

        # Analysis counts (based on current filtered rows)
        txn_counter = Counter((r.get("TransactionType") or "Unknown") for r in rows)
        itemtype_counter = Counter((r.get("ItemType") or "Unknown") for r in rows)

        # ✅ ONLY 2 stages in counts
        stage_counter = Counter()
        for r in rows:
            st = (r.get("PendingStage") or "").strip()
            if st in ("Check By", "Approved By"):
                stage_counter[st] += 1

        txn_counts = txn_counter.most_common(15)
        itemtype_counts = itemtype_counter.most_common(15)

        # ✅ fixed order
        stage_counts = [
            (k, stage_counter.get(k, 0))
            for k in ("Check By", "Approved By")
            if stage_counter.get(k, 0) > 0
        ]

        # Excel export (all filtered rows, no pagination)
        if export in ("1", "true", "xlsx", "excel"):
            fname = f"Pending_GRN_{from_date_int}.xlsx"
            return _xlsx_response_pending_grn(fname, rows)

        # Screen pagination
        paginator = Paginator(rows, page_size)
        try:
            page_obj = paginator.page(page)
        except PageNotAnInteger:
            page_obj = paginator.page(1)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages)

        return render(
            request,
            "store/pending_grn_report.html",
            {
                # detail table
                "rows": page_obj,
                "page_size": page_size,

                # filters
                "from_date": from_date_str,
                "docno": docno_q,
                "party": party_q,
                "item": item_q,
                "status": status_q,

                # cross-filter selection
                "txn_type_sel": txn_type_sel,
                "item_type_sel": item_type_sel,
                "stage_sel": stage_sel,

                # analysis panels
                "txn_counts": txn_counts,
                "itemtype_counts": itemtype_counts,
                "stage_counts": stage_counts,
            },
        )

    except Exception:
        logger.exception("Error in pending_grn_report")
        messages.error(request, "Something went wrong while loading Pending GRN report.")
        return redirect("indexpage")
