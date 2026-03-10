# maintenance/views.py
from datetime import timedelta
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import DateField
from django.db.models.functions import Coalesce
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST
from .forms import MaintenanceUpdateForm
from .models import MaintenanceSchedule
import calendar
from datetime import date, datetime, timedelta

@login_required
def dashboard(request):
    today = timezone.localdate()
    tomorrow = today + timedelta(days=1)

    # Use effective date = COALESCE(rescheduled_to, scheduled_date)
    effective = Coalesce("rescheduled_to", "scheduled_date", output_field=DateField())
    base_qs = MaintenanceSchedule.objects.annotate(eff_date=effective)

    today_qs = base_qs.filter(eff_date=today).order_by("equipment_id")
    tomorrow_qs = base_qs.filter(eff_date=tomorrow).order_by("equipment_id")

    return render(request, "maintenance/dashboard.html", {
        "today": today,
        "tomorrow": tomorrow,
        "today_qs": today_qs,
        "tomorrow_qs": tomorrow_qs,
    })


@login_required
def update_schedule(request, pk):
    obj = get_object_or_404(MaintenanceSchedule, pk=pk)
    if request.method == "POST":
        form = MaintenanceUpdateForm(request.POST, instance=obj)
        if form.is_valid():
            schedule = form.save(commit=False)
            if schedule.status == MaintenanceSchedule.STATUS_DONE and schedule.completed_at is None:
                schedule.completed_at = timezone.now()
                schedule.completed_by = request.user
            schedule.save()
            messages.success(request, "Maintenance updated.")
            return redirect("maintenance:dashboard")
    else:
        form = MaintenanceUpdateForm(instance=obj)

    return render(request, "maintenance/update.html", {"form": form, "obj": obj})


# --- Simple Excel uploader (openpyxl) ---
# Accepts columns: "Equipment ID" / "EquipmentId", "Location", "Date" (e.g., 01-04-2025)
import datetime
from django.views.decorators.http import require_http_methods

def _parse_date(cell_value):
    """
    Robust date parser:
    - Excel serials (datetime, date)
    - Strings like '01-04-2025' or '2025-04-01'
    """
    if isinstance(cell_value, (datetime.date, datetime.datetime)):
        return cell_value.date() if isinstance(cell_value, datetime.datetime) else cell_value

    if isinstance(cell_value, (int, float)):
        # Likely Excel serial date; try best-effort (Excel's 1900 system)
        base = datetime.datetime(1899, 12, 30)
        return (base + datetime.timedelta(days=int(cell_value))).date()

    if isinstance(cell_value, str):
        s = cell_value.strip()
        for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%b-%Y"):
            try:
                return datetime.datetime.strptime(s, fmt).date()
            except ValueError:
                pass

    raise ValueError(f"Unrecognized date value: {cell_value!r}")

@login_required
@require_http_methods(["GET", "POST"])
def upload_excel(request):
    """
    Upload an .xlsx file with columns:
      Equipment ID | Location | Date
    """
    if request.method == "POST" and request.FILES.get("file"):
        from openpyxl import load_workbook  # pip install openpyxl

        wb = load_workbook(request.FILES["file"], read_only=True, data_only=True)
        ws = wb.active

        # Find header row
        headers = {}
        for j, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)), start=1):
            if not cell:
                continue
            key = str(cell).strip().lower()
            headers[key] = j

        def col(colnames):
            for name in colnames:
                idx = headers.get(name)
                if idx:
                    return idx
            return None

        c_equip = col(["equipment id", "equipmentid", "equipment"])
        c_loc   = col(["location", "block"])
        c_date  = col(["date", "scheduled date", "schedule date"])

        if not all([c_equip, c_loc, c_date]):
            messages.error(request, "Missing required columns. Need: Equipment ID, Location, Date.")
            return redirect("maintenance:upload_excel")

        created = 0
        skipped = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            equip = (row[c_equip - 1] or "").strip() if row[c_equip - 1] else ""
            loc   = (row[c_loc - 1] or "").strip() if row[c_loc - 1] else ""
            raw_d = row[c_date - 1]

            if not equip or not raw_d:
                skipped += 1
                continue

            try:
                dt = _parse_date(raw_d)
            except Exception:
                skipped += 1
                continue

            obj, is_created = MaintenanceSchedule.objects.get_or_create(
                equipment_id=equip,
                scheduled_date=dt,
                defaults={"location": loc},
            )
            if not is_created:
                # Update location if empty in DB
                if not obj.location and loc:
                    obj.location = loc
                    obj.save(update_fields=["location"])
                skipped += 1
            else:
                created += 1

        messages.success(request, f"Imported: {created}, Skipped: {skipped}")
        return redirect("maintenance:dashboard")

    return render(request, "maintenance/upload.html")


@login_required
@require_POST
def mark_done(request, pk):
    obj = get_object_or_404(MaintenanceSchedule, pk=pk)
    if obj.status != MaintenanceSchedule.STATUS_DONE:
        obj.status = MaintenanceSchedule.STATUS_DONE
        obj.completed_at = timezone.now()
        obj.completed_by = request.user
        obj.save(update_fields=["status", "completed_at", "completed_by", "updated_at"])
        messages.success(request, f"{obj.equipment_id} marked as Done.")
    else:
        messages.info(request, f"{obj.equipment_id} is already Done.")
    return redirect("maintenance:dashboard")




def _first_last_of_month(any_day: date):
    first = any_day.replace(day=1)
    if first.month == 12:
        next_month = first.replace(year=first.year + 1, month=1, day=1)
    else:
        next_month = first.replace(month=first.month + 1, day=1)
    last = next_month - timedelta(days=1)
    return first, last


def calendar_month(request):
    """
    Month calendar showing up to 5 items per day, with effective date:
    eff_date = COALESCE(rescheduled_to, scheduled_date)
    No custom template filters needed — we build weeks_data ready for the template.
    """
    MAX_SHOW = 5
    today = timezone.localdate()

    # Parse ?month=YYYY-MM
    month_param = request.GET.get("month")
    if month_param:
        try:
            y, m = map(int, month_param.split("-"))
            current = date(y, m, 1)
        except Exception:
            current = today.replace(day=1)
    else:
        current = today.replace(day=1)

    month_start, month_end = _first_last_of_month(current)

    # Annotate eff_date and fetch everything for the month
    qs = (
        MaintenanceSchedule.objects
        .annotate(eff_date=Coalesce("rescheduled_to", "scheduled_date"))
        .filter(eff_date__range=(month_start, month_end))
        .order_by("eff_date", "status", "equipment_id")
    )

    # Group by day
    by_day = {}
    for obj in qs:
        d = obj.eff_date
        by_day.setdefault(d, []).append(obj)

    # Build Sunday-first calendar grid (list of weeks; each week = 7 dates)
    cal = calendar.Calendar(firstweekday=6)
    raw_weeks = [list(week) for week in cal.monthdatescalendar(current.year, current.month)]

    # Convert to weeks_data the template can render without custom filters
    weeks_data = []
    for week in raw_weeks:
        day_cells = []
        for d in week:
            items_full = by_day.get(d, [])
            total = len(items_full)
            overflow_count = max(0, total - MAX_SHOW)
            day_cells.append({
                "date": d,
                "in_month": (d.month == current.month),
                "items": items_full[:MAX_SHOW],   # only show first 5
                "total": total,                   # for cap/“view all”
                "overflow_count": overflow_count,
            })
        weeks_data.append(day_cells)

    # For prev/next links
    prev_month = (month_start.replace(day=1) - timedelta(days=1)).replace(day=1)
    next_month = (month_end + timedelta(days=1)).replace(day=1)

    context = {
        "today": today,
        "current": current,
        "month_start": month_start,
        "month_end": month_end,
        "prev_month_str": f"{prev_month:%Y-%m}",
        "next_month_str": f"{next_month:%Y-%m}",
        "weeks_data": weeks_data,   # <— use this in template
        "MAX_SHOW": MAX_SHOW,
    }
    return render(request, "maintenance/calendar.html", context)


def calendar_day(request, datestr: str):
    """
    Day detail list using same 'effective date' logic.
    """
    try:
        target = datetime.datetime.strptime(datestr, "%Y-%m-%d").date()
    except ValueError:
        target = timezone.localdate()

    qs = (
        MaintenanceSchedule.objects
        .annotate(eff_date=Coalesce("rescheduled_to", "scheduled_date"))
        .filter(eff_date=target)
        .order_by("status", "equipment_id")
    )

    return render(request, "maintenance/calendar_day.html", {
        "target": target,
        "items": qs,
        "month_str": f"{target:%Y-%m}",
    })